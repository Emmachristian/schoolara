# finance/views.py

"""
Finance Management Views

Comprehensive view functions for:
- Account Types (CRUD + Print)
- Accounts (CRUD + Print + Reconciliation)
- Expense Categories (CRUD + Print)
- Expenses (CRUD + Print + Approval + Bulk)
- Expense Payments (CRUD + Print + Verification + Reversal)
- Journals (CRUD + Print)
- Journal Entries (CRUD + Print + Posting + Reversal)
- Budgets (CRUD + Print + Approval)
- Reports and Analytics

Changes from original:
- Added PROCESSING status handling throughout expense payment views/exports
- expense_approve, journal_entry_reverse, budget_approve: added GET fallback redirects
- export_expense_payments_excel: added effective_amount, payment_state, account columns
- export_expenses_excel: added payee_type, payee_name, vendor_reference, subtotal, tax columns
- expense_payment_print_view: added processing count, total_bank_charges to stats
- expense_payment_print_receipt: added total_with_fees, effective_amount to context
- Added expense_payment_reversal_print view
- bulk_expense_payment_verification: includes PROCESSING status
- FIX: expense__vendor_name → expense__payee_name in get_filtered_expense_payments
- FIX: removed invalid select_related('academic_session') from expense_detail,
        expense_print_view, expense_category_detail (Expense has no direct FK;
        access via fiscal_period.related_academic_session)
- FIX: expense_list_print_view — vendor_name/is_recurring removed (fields dropped
        from model); replaced with payee_name/payee_type; _get_print_school_context
        corrected to get_print_school_context
- PATTERN: extracted _make_workbook/_xlsx_response/_resolve_columns helpers
        (same pattern as fees/views.py) — all export views now use them
- PATTERN: MAX_PRINT_RECORDS module-level constant (was inline magic number)
- PATTERN: is_htmx checked once at top of POST handler (not duplicated in
        try/except blocks) — matches fees/views.py consistency
- PATTERN: dashboard stats isolated into individual try/except blocks so one
        failing query never blanks the entire dashboard
"""

from django.shortcuts import render, redirect, get_object_or_404
from django.urls import reverse
from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator
from django.contrib import messages
from django.db.models import Q, Count, Sum, Avg, F, Max, Min, DecimalField, Case, When
from django.utils import timezone
from django.http import JsonResponse, HttpResponse
from django.db import transaction
from datetime import timedelta, date, datetime
from decimal import Decimal
import logging

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from .models import (
    AccountType,
    Account,
    ExpenseCategory,
    Expense,
    ExpenseLine,
    ExpensePayment,
    Journal,
    JournalEntry,
    JournalTransaction,
    Budget,
    BudgetLine,
)

from .forms import (
    AccountTypeFilterForm,
    AccountFilterForm,
    ExpenseCategoryFilterForm,
    ExpenseFilterForm,
    ExpensePaymentFilterForm,
    ExpensePaymentReversalForm,
    JournalFilterForm,
    JournalEntryFilterForm,
    BudgetFilterForm,
    AccountTypeForm,
    AccountForm,
    AccountQuickAddForm,
    ExpenseCategoryForm,
    ExpenseForm,
    ExpenseLineForm,
    ExpenseApprovalForm,
    ExpensePaymentForm,
    BulkExpensePaymentForm,
    BulkExpensePaymentVerificationForm,
    JournalForm,
    JournalEntryForm,
    JournalTransactionForm,
    JournalEntryReversalForm,
    BudgetForm,
    BudgetLineForm,
    BudgetApprovalForm,
    AccountReconciliationForm,
)

from .services import (
    ExpenseService,
    ExpensePaymentService,
    JournalEntryService,
    BudgetService,
)

from core.utils import (
    get_school_today,
    get_school_current_time,
    format_money,
    get_school_currency,
)

from core.view_helpers import (
    get_print_school_context,
)

from academics.models import AcademicSession
from core.models import FiscalYear, FiscalPeriod

from . import stats as finance_stats

logger = logging.getLogger(__name__)

MAX_PRINT_RECORDS = 500


# =============================================================================
# EXCEL EXPORT HELPERS  (same pattern as fees/views.py)
# =============================================================================

_HEADER_FILL  = PatternFill(start_color='1E3A5F', end_color='1E3A5F', fill_type='solid')
_HEADER_FONT  = Font(bold=True, color='FFFFFF', size=11, name='Arial')
_HEADER_ALIGN = Alignment(horizontal='center', vertical='center', wrap_text=True)
_DATA_ALIGN   = Alignment(vertical='center', wrap_text=False)


def _make_workbook(sheet_title, columns, rows):
    """
    Build a styled Workbook.

    columns: list of (header, accessor) pairs where accessor is a callable.
    rows:    iterable of model instances.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_title

    ws.append([col[0] for col in columns])
    for cell in ws[1]:
        cell.font      = _HEADER_FONT
        cell.fill      = _HEADER_FILL
        cell.alignment = _HEADER_ALIGN
    ws.row_dimensions[1].height = 28

    for obj in rows:
        ws.append([col[1](obj) for col in columns])

    for col_cells in ws.columns:
        max_len = max(
            (len(str(c.value)) if c.value is not None else 0) for c in col_cells
        )
        ws.column_dimensions[col_cells[0].column_letter].width = min(max_len + 4, 60)

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = _DATA_ALIGN

    return wb


def _xlsx_response(wb, filename_stem):
    """Return an HttpResponse that triggers an xlsx download."""
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = (
        f'attachment; filename="{filename_stem}_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx"'
    )
    wb.save(response)
    return response


def _resolve_columns(all_columns, selected_keys, default_keys):
    """
    Return ordered (header, accessor) pairs matching selected_keys.

    all_columns:   list of (key, header, accessor) triples
    selected_keys: keys requested by caller (e.g. from ?fields= query param)
    default_keys:  fallback when selected_keys is empty or has no matches
    """
    col_map = {c[0]: c for c in all_columns}
    keys    = [k for k in selected_keys if k in col_map] or default_keys
    return [(col_map[k][1], col_map[k][2]) for k in keys if k in col_map]


# =============================================================================
# DASHBOARD
# =============================================================================

@login_required
def finance_dashboard(request):
    """
    Main finance dashboard with overview statistics.

    Each stats block is wrapped in its own try/except so that one
    failing query never blanks the entire dashboard.
    """
    current_period = None
    try:
        current_period = FiscalPeriod.get_current_fiscal_period()
    except Exception as e:
        logger.error(f"Dashboard: error fetching current fiscal period: {e}")

    try:
        account_stats = {
            'total': Account.objects.filter(is_active=True).count(),
            'bank_accounts': Account.objects.filter(is_bank_account=True, is_active=True).count(),
            'cash_accounts': Account.objects.filter(is_cash_account=True, is_active=True).count(),
            'total_cash_balance': Account.objects.filter(
                is_cash_account=True, is_active=True
            ).aggregate(Sum('current_balance'))['current_balance__sum'] or Decimal('0.00'),
            'total_bank_balance': Account.objects.filter(
                is_bank_account=True, is_active=True
            ).aggregate(Sum('current_balance'))['current_balance__sum'] or Decimal('0.00'),
        }
    except Exception as e:
        logger.error(f"Dashboard: error fetching account stats: {e}")
        account_stats = {}

    try:
        today             = get_school_today()
        this_month_start  = today.replace(day=1)
        expense_stats = {
            'total_pending':  Expense.objects.filter(status='PENDING_APPROVAL').count(),
            'total_approved': Expense.objects.filter(status='APPROVED').count(),
            'this_month_count': Expense.objects.filter(
                expense_date__gte=this_month_start
            ).count(),
            'this_month_amount': Expense.objects.filter(
                expense_date__gte=this_month_start
            ).aggregate(Sum('total_amount'))['total_amount__sum'] or Decimal('0.00'),
            'pending_amount': Expense.objects.filter(
                status='PENDING_APPROVAL'
            ).aggregate(Sum('total_amount'))['total_amount__sum'] or Decimal('0.00'),
        }
    except Exception as e:
        logger.error(f"Dashboard: error fetching expense stats: {e}")
        expense_stats = {}

    try:
        active_budgets = Budget.objects.filter(status='ACTIVE')
        budget_stats = {
            'total_active': active_budgets.count(),
            'total_revenue_budget': active_budgets.aggregate(
                Sum('total_revenue_budget'))['total_revenue_budget__sum'] or Decimal('0.00'),
            'total_expense_budget': active_budgets.aggregate(
                Sum('total_expense_budget'))['total_expense_budget__sum'] or Decimal('0.00'),
            'total_actual_revenue': active_budgets.aggregate(
                Sum('actual_revenue_total'))['actual_revenue_total__sum'] or Decimal('0.00'),
            'total_actual_expense': active_budgets.aggregate(
                Sum('actual_expense_total'))['actual_expense_total__sum'] or Decimal('0.00'),
        }
    except Exception as e:
        logger.error(f"Dashboard: error fetching budget stats: {e}")
        budget_stats = {}

    try:
        this_month_start = get_school_today().replace(day=1)
        journal_stats = {
            'total_draft':   JournalEntry.objects.filter(status='DRAFT').count(),
            'total_posted':  JournalEntry.objects.filter(status='POSTED').count(),
            'this_month_count': JournalEntry.objects.filter(
                entry_date__gte=this_month_start
            ).count(),
        }
    except Exception as e:
        logger.error(f"Dashboard: error fetching journal stats: {e}")
        journal_stats = {}

    try:
        payment_stats = {
            'unverified': ExpensePayment.objects.filter(
                is_verified=False,
                status__in=['PROCESSING', 'PROCESSED'],
                reversed=False,
            ).count(),
            'verified': ExpensePayment.objects.filter(is_verified=True).count(),
            'reversed': ExpensePayment.objects.filter(reversed=True).count(),
        }
    except Exception as e:
        logger.error(f"Dashboard: error fetching payment stats: {e}")
        payment_stats = {}

    recent_expenses = Expense.objects.select_related(
        'category', 'fiscal_period'
    ).order_by('-created_at')[:10]

    recent_payments = ExpensePayment.objects.select_related(
        'expense', 'payment_method', 'account'
    ).filter(reversed=False).order_by('-payment_date')[:10]

    recent_entries = JournalEntry.objects.select_related(
        'journal', 'fiscal_period'
    ).order_by('-entry_date')[:10]

    pending_expenses = Expense.objects.filter(
        status='PENDING_APPROVAL'
    ).select_related('category').order_by('expense_date')[:10]

    unverified_payments = ExpensePayment.objects.filter(
        is_verified=False,
        status__in=['PROCESSING', 'PROCESSED'],
        reversed=False,
    ).select_related('expense', 'payment_method').order_by('-payment_date')[:10]

    draft_entries = JournalEntry.objects.filter(
        status='DRAFT'
    ).select_related('journal').order_by('entry_date')[:10]

    over_budget = BudgetLine.objects.filter(
        actual_amount__gt=F('budgeted_amount'),
        budget__status='ACTIVE',
    ).select_related('budget', 'account').order_by('-actual_amount')[:10]

    context = {
        'current_period':    current_period,
        'account_stats':     account_stats,
        'expense_stats':     expense_stats,
        'budget_stats':      budget_stats,
        'journal_stats':     journal_stats,
        'payment_stats':     payment_stats,
        'recent_expenses':   recent_expenses,
        'recent_payments':   recent_payments,
        'recent_entries':    recent_entries,
        'pending_expenses':  pending_expenses,
        'unverified_payments': unverified_payments,
        'draft_entries':     draft_entries,
        'over_budget':       over_budget,
        'currency':          get_school_currency(),
    }

    return render(request, 'finance/dashboard.html', context)


# =============================================================================
# HELPER FUNCTIONS FOR FILTERING
# =============================================================================

def get_filtered_account_types(request):
    account_types = AccountType.objects.annotate(
        account_count=Count('accounts', distinct=True)
    ).order_by('account_type', 'display_order', 'name')

    query        = request.GET.get('q', '').strip()
    account_type = request.GET.get('account_type', '')
    is_active    = request.GET.get('is_active', '')

    if query:
        words = query.strip().split()
        if words:
            combined_q = Q()
            for word in words:
                combined_q &= (
                    Q(name__icontains=word)        |
                    Q(code__icontains=word)        |
                    Q(description__icontains=word)
                )
            account_types = account_types.filter(combined_q)

    if account_type:
        account_types = account_types.filter(account_type=account_type)
    if is_active:
        account_types = account_types.filter(is_active=(is_active.lower() == 'true'))

    return account_types


def get_filtered_accounts(request):
    accounts = Account.objects.select_related(
        'account_type', 'parent_account'
    ).annotate(
        child_count=Count('child_accounts', distinct=True),
        transaction_count=Count('journal_transactions', distinct=True),
    ).order_by('account_type__account_type', 'account_number')

    query            = request.GET.get('q', '').strip()
    account_type     = request.GET.get('account_type', '')
    account_category = request.GET.get('account_category', '')
    is_active        = request.GET.get('is_active', '')
    is_reconcilable  = request.GET.get('is_reconcilable', '')
    min_balance      = request.GET.get('min_balance', '')
    max_balance      = request.GET.get('max_balance', '')

    if query:
        words = query.strip().split()
        if words:
            combined_q = Q()
            for word in words:
                combined_q &= (
                    Q(account_number__icontains=word) |
                    Q(name__icontains=word)           |
                    Q(description__icontains=word)    |
                    Q(bank_name__icontains=word)
                )
            accounts = accounts.filter(combined_q)

    if account_type:
        accounts = accounts.filter(account_type_id=account_type)

    if account_category:
        category_map = {
            'bank':         {'is_bank_account': True},
            'cash':         {'is_cash_account': True},
            'mobile_money': {'is_mobile_money_account': True},
            'receivable':   {'is_receivable_account': True},
            'payable':      {'is_payable_account': True},
            'inventory':    {'is_inventory_account': True},
            'fixed_asset':  {'is_fixed_asset': True},
            'revenue':      {'is_revenue_account': True},
            'expense':      {'is_expense_account': True},
        }
        if account_category in category_map:
            accounts = accounts.filter(**category_map[account_category])

    if is_active:
        accounts = accounts.filter(is_active=(is_active.lower() == 'true'))
    if is_reconcilable:
        accounts = accounts.filter(is_reconcilable=(is_reconcilable.lower() == 'true'))

    if min_balance:
        try:
            accounts = accounts.filter(current_balance__gte=Decimal(min_balance))
        except (ValueError, TypeError):
            pass
    if max_balance:
        try:
            accounts = accounts.filter(current_balance__lte=Decimal(max_balance))
        except (ValueError, TypeError):
            pass

    return accounts


def get_filtered_expense_categories(request):
    categories = ExpenseCategory.objects.select_related(
        'default_expense_account'
    ).annotate(
        expense_count=Count('expenses', distinct=True)
    ).order_by('category_type', 'name')

    query             = request.GET.get('q', '').strip()
    category_type     = request.GET.get('category_type', '')
    is_active         = request.GET.get('is_active', '')
    requires_approval = request.GET.get('requires_approval', '')

    if query:
        words = query.strip().split()
        if words:
            combined_q = Q()
            for word in words:
                combined_q &= (Q(name__icontains=word) | Q(description__icontains=word))
            categories = categories.filter(combined_q)

    if category_type:
        categories = categories.filter(category_type=category_type)
    if is_active:
        categories = categories.filter(is_active=(is_active.lower() == 'true'))
    if requires_approval:
        categories = categories.filter(requires_approval=(requires_approval.lower() == 'true'))

    return categories


def get_filtered_expenses(request):
    expenses = Expense.objects.select_related(
        'category',
        'fiscal_period',
        'fiscal_period__related_academic_session',
        'preferred_payment_method',
        'budget_line',
        'journal_entry',
    ).prefetch_related(
        'lines',
        'payments',
    ).annotate(
        line_count=Count('lines', distinct=True),
        payment_count=Count(
            'payments',
            distinct=True,
            filter=Q(payments__reversed=False),
        ),
    ).order_by('-expense_date', '-created_at')

    query             = request.GET.get('q',                '').strip()
    status            = request.GET.get('status',           '')
    category          = request.GET.get('category',         '')
    payee_type        = request.GET.get('payee_type',       '')
    fiscal_period     = request.GET.get('fiscal_period',    '')
    expense_date_from = request.GET.get('expense_date_from','')
    expense_date_to   = request.GET.get('expense_date_to',  '')
    min_amount        = request.GET.get('min_amount',       '')
    max_amount        = request.GET.get('max_amount',       '')

    if query:
        words = query.split()
        if words:
            combined_q = Q()
            for word in words:
                combined_q &= (
                    Q(expense_number__icontains=word)  |
                    Q(description__icontains=word)     |
                    # FIX: vendor_name replaced by payee_name
                    Q(payee_name__icontains=word)      |
                    Q(vendor_reference__icontains=word)
                )
            expenses = expenses.filter(combined_q)

    if status:
        expenses = expenses.filter(status=status)
    if category:
        expenses = expenses.filter(category_id=category)
    if payee_type:
        expenses = expenses.filter(payee_type=payee_type)
    if fiscal_period:
        expenses = expenses.filter(fiscal_period_id=fiscal_period)
    if expense_date_from:
        expenses = expenses.filter(expense_date__gte=expense_date_from)
    if expense_date_to:
        expenses = expenses.filter(expense_date__lte=expense_date_to)

    if min_amount:
        try:
            expenses = expenses.filter(total_amount__gte=Decimal(min_amount))
        except (ValueError, TypeError):
            pass
    if max_amount:
        try:
            expenses = expenses.filter(total_amount__lte=Decimal(max_amount))
        except (ValueError, TypeError):
            pass

    return expenses


def get_filtered_expense_payments(request):
    payments = ExpensePayment.objects.select_related(
        'expense__category', 'payment_method', 'account', 'fiscal_period'
    ).order_by('-payment_date', '-created_at')

    query             = request.GET.get('q', '').strip()
    status            = request.GET.get('status', '')
    payment_state     = request.GET.get('payment_state', '')
    payment_method    = request.GET.get('payment_method', '')
    account           = request.GET.get('account', '')
    fiscal_period     = request.GET.get('fiscal_period', '')
    is_verified       = request.GET.get('is_verified', '')
    payment_date_from = request.GET.get('payment_date_from', '')
    payment_date_to   = request.GET.get('payment_date_to', '')
    min_amount        = request.GET.get('min_amount', '')
    max_amount        = request.GET.get('max_amount', '')

    if query:
        words = query.strip().split()
        if words:
            combined_q = Q()
            for word in words:
                combined_q &= (
                    Q(reference_number__icontains=word)       |
                    Q(transaction_id__icontains=word)         |
                    Q(expense__expense_number__icontains=word) |
                    # FIX: was expense__vendor_name — field renamed to payee_name
                    Q(expense__payee_name__icontains=word)
                )
            payments = payments.filter(combined_q)

    if status:
        payments = payments.filter(status=status)

    if payment_state:
        if payment_state == 'active':
            payments = payments.filter(reversed=False)
        elif payment_state == 'reversed':
            payments = payments.filter(reversed=True)
        elif payment_state == 'processing':
            payments = payments.filter(status='PROCESSING', reversed=False)

    if payment_method:
        payments = payments.filter(payment_method_id=payment_method)
    if account:
        payments = payments.filter(account_id=account)
    if fiscal_period:
        payments = payments.filter(fiscal_period_id=fiscal_period)
    if is_verified:
        payments = payments.filter(is_verified=(is_verified.lower() == 'true'))
    if payment_date_from:
        payments = payments.filter(payment_date__gte=payment_date_from)
    if payment_date_to:
        payments = payments.filter(payment_date__lte=payment_date_to)

    if min_amount:
        try:
            payments = payments.filter(amount__gte=Decimal(min_amount))
        except (ValueError, TypeError):
            pass
    if max_amount:
        try:
            payments = payments.filter(amount__lte=Decimal(max_amount))
        except (ValueError, TypeError):
            pass

    return payments


def get_filtered_journals(request):
    journals = Journal.objects.annotate(
        entry_count=Count('entries', distinct=True)
    ).order_by('journal_type', 'name')

    query        = request.GET.get('q', '').strip()
    journal_type = request.GET.get('journal_type', '')
    is_active    = request.GET.get('is_active', '')

    if query:
        words = query.strip().split()
        if words:
            combined_q = Q()
            for word in words:
                combined_q &= (Q(name__icontains=word) | Q(description__icontains=word))
            journals = journals.filter(combined_q)

    if journal_type:
        journals = journals.filter(journal_type=journal_type)
    if is_active:
        journals = journals.filter(is_active=(is_active.lower() == 'true'))

    return journals


def get_filtered_journal_entries(request):
    entries = JournalEntry.objects.select_related(
        'journal', 'academic_session', 'fiscal_period'
    ).prefetch_related('transactions__account').annotate(
        transaction_count=Count('transactions', distinct=True),
        total_debit=Sum(
            Case(
                When(transactions__is_debit=True, then=F('transactions__amount')),
                default=0,
                output_field=DecimalField(),
            )
        ),
        total_credit=Sum(
            Case(
                When(transactions__is_debit=False, then=F('transactions__amount')),
                default=0,
                output_field=DecimalField(),
            )
        ),
    ).order_by('-entry_date', '-created_at')

    query            = request.GET.get('q', '').strip()
    status           = request.GET.get('status', '')
    journal          = request.GET.get('journal', '')
    academic_session = request.GET.get('academic_session', '')
    fiscal_period    = request.GET.get('fiscal_period', '')
    entry_date_from  = request.GET.get('entry_date_from', '')
    entry_date_to    = request.GET.get('entry_date_to', '')

    if query:
        words = query.strip().split()
        if words:
            combined_q = Q()
            for word in words:
                combined_q &= (
                    Q(entry_number__icontains=word)    |
                    Q(reference_number__icontains=word) |
                    Q(description__icontains=word)
                )
            entries = entries.filter(combined_q)

    if status:
        entries = entries.filter(status=status)
    if journal:
        entries = entries.filter(journal_id=journal)
    if academic_session:
        entries = entries.filter(academic_session_id=academic_session)
    if fiscal_period:
        entries = entries.filter(fiscal_period_id=fiscal_period)
    if entry_date_from:
        entries = entries.filter(entry_date__gte=entry_date_from)
    if entry_date_to:
        entries = entries.filter(entry_date__lte=entry_date_to)

    return entries


def get_filtered_budgets(request):
    budgets = Budget.objects.select_related(
        'fiscal_year', 'academic_session', 'parent_budget'
    ).prefetch_related('lines').annotate(
        line_count=Count('lines', distinct=True),
        child_count=Count('child_budgets', distinct=True),
    ).order_by('-start_date', 'name')

    query            = request.GET.get('q', '').strip()
    budget_type      = request.GET.get('budget_type', '')
    status           = request.GET.get('status', '')
    fiscal_year      = request.GET.get('fiscal_year', '')
    academic_session = request.GET.get('academic_session', '')

    if query:
        words = query.strip().split()
        if words:
            combined_q = Q()
            for word in words:
                combined_q &= (Q(name__icontains=word) | Q(description__icontains=word))
            budgets = budgets.filter(combined_q)

    if budget_type:
        budgets = budgets.filter(budget_type=budget_type)
    if status:
        budgets = budgets.filter(status=status)
    if fiscal_year:
        budgets = budgets.filter(fiscal_year_id=fiscal_year)
    if academic_session:
        budgets = budgets.filter(academic_session_id=academic_session)

    return budgets


# =============================================================================
# ACCOUNT TYPE VIEWS
# =============================================================================

@login_required
def account_type_list(request):
    filter_form  = AccountTypeFilterForm(request.GET or None)
    account_types = get_filtered_account_types(request)

    stats = {
        'total':          account_types.count(),
        'active':         account_types.filter(is_active=True).count(),
        'asset':          account_types.filter(account_type='ASSET').count(),
        'liability':      account_types.filter(account_type='LIABILITY').count(),
        'equity':         account_types.filter(account_type='EQUITY').count(),
        'revenue':        account_types.filter(account_type='REVENUE').count(),
        'expense':        account_types.filter(account_type='EXPENSE').count(),
        'total_accounts': sum(at.account_count or 0 for at in account_types),
    }

    paginator          = Paginator(account_types, 20)
    account_types_page = paginator.get_page(request.GET.get('page', 1))
    is_htmx            = request.headers.get('HX-Request') == 'true'

    context = {
        'account_types_page': account_types_page,
        'paginator':          paginator,
        'stats':              stats,
        'filter_form':        filter_form,
        'is_htmx':            is_htmx,
    }

    if is_htmx:
        return render(request, 'finance/account_types/partials/_type_results.html', context)
    return render(request, 'finance/account_types/list.html', context)


@login_required
def account_type_create(request):
    if request.method == 'POST':
        form = AccountTypeForm(request.POST)
        if form.is_valid():
            account_type = form.save()
            messages.success(request, f"Account type '{account_type.name}' created successfully", extra_tags='sweetalert')
            return redirect('finance:account_type_detail', pk=account_type.pk)
    else:
        form = AccountTypeForm()

    return render(request, 'finance/account_types/form.html', {'form': form, 'title': 'Create Account Type'})


@login_required
def account_type_edit(request, pk):
    account_type = get_object_or_404(AccountType, pk=pk)

    if request.method == 'POST':
        form = AccountTypeForm(request.POST, instance=account_type)
        if form.is_valid():
            account_type = form.save()
            messages.success(request, f"Account type '{account_type.name}' updated successfully", extra_tags='sweetalert')
            return redirect('finance:account_type_detail', pk=account_type.pk)
    else:
        form = AccountTypeForm(instance=account_type)

    return render(request, 'finance/account_types/form.html', {
        'form': form, 'account_type': account_type, 'title': f'Edit {account_type.name}'
    })


@login_required
def account_type_detail(request, pk):
    account_type = get_object_or_404(AccountType, pk=pk)

    accounts = account_type.accounts.filter(is_active=True).annotate(
        transaction_count=Count('journal_transactions')
    ).order_by('account_number')[:50]

    context = {
        'account_type':  account_type,
        'accounts':      accounts,
        'account_count': account_type.accounts.count(),
        'active_count':  account_type.accounts.filter(is_active=True).count(),
        'total_balance': account_type.accounts.aggregate(
            Sum('current_balance'))['current_balance__sum'] or Decimal('0.00'),
    }
    return render(request, 'finance/account_types/detail.html', context)


@login_required
def account_type_delete(request, pk):
    account_type = get_object_or_404(AccountType, pk=pk)

    if request.method == 'POST':
        is_htmx = request.headers.get('HX-Request') == 'true'

        if account_type.accounts.exists():
            msg = f"Cannot delete '{account_type.name}' because it has associated accounts"
            if is_htmx:
                r = HttpResponse()
                r['HX-Alert-Message'] = msg
                r['HX-Alert-Type']    = 'error'
                r['HX-Alert-Title']   = 'Cannot Delete'
                r['HX-Close-Modal']   = 'true'
                return r
            messages.error(request, msg, extra_tags='sweetalert-error')
            return redirect('finance:account_type_list')

        type_name = account_type.name
        account_type.delete()

        if is_htmx:
            r = HttpResponse()
            r['HX-Alert-Message'] = f"Account type '{type_name}' deleted successfully"
            r['HX-Alert-Type']    = 'success'
            r['HX-Alert-Title']   = 'Deleted!'
            r['HX-Close-Modal']   = 'true'
            r['HX-Redirect']      = reverse('finance:account_type_list')
            return r
        messages.success(request, f"Account type '{type_name}' deleted successfully", extra_tags='sweetalert')
        return redirect('finance:account_type_list')


@login_required
def account_type_list_print_view(request):
    FIELD_NAMES_FULL = {
        'code':           'Code',
        'name':           'Name',
        'account_type':   'Account Category',
        'normal_balance': 'Normal Balance',
        'number_prefix':  'Number Prefix',
        'account_count':  'No. of Accounts',
        'is_active':      'Active',
        'description':    'Description',
    }
    FIELD_NAMES_SHORT = {
        'code':           'Code',
        'name':           'Name',
        'account_type':   'Category',
        'normal_balance': 'Normal Bal.',
        'number_prefix':  'Prefix',
        'account_count':  '# Accounts',
        'is_active':      'Active',
        'description':    'Description',
    }
    DEFAULT_FIELDS  = ['code', 'name', 'account_type', 'normal_balance', 'account_count', 'is_active']
    selected_fields = request.GET.getlist('fields') or DEFAULT_FIELDS
    short_headers   = request.GET.get('short_headers', 'false').lower() == 'true'
    landscape       = request.GET.get('landscape', 'false').lower() == 'true'
    include_stats   = request.GET.get('include_stats', 'true').lower() == 'true'
    field_names     = FIELD_NAMES_SHORT if short_headers else FIELD_NAMES_FULL
    account_types   = get_filtered_account_types(request)

    stats = None
    if include_stats:
        stats = {
            'total':     account_types.count(),
            'active':    account_types.filter(is_active=True).count(),
            'asset':     account_types.filter(account_type='ASSET').count(),
            'liability': account_types.filter(account_type='LIABILITY').count(),
            'equity':    account_types.filter(account_type='EQUITY').count(),
            'revenue':   account_types.filter(account_type='REVENUE').count(),
            'expense':   account_types.filter(account_type='EXPENSE').count(),
        }

    if account_types.count() > MAX_PRINT_RECORDS:
        account_types = account_types[:MAX_PRINT_RECORDS]

    return render(request, 'finance/account_types/print_list.html', {
        **get_print_school_context(request),
        'account_types':        account_types,
        'stats':                stats,
        'selected_fields':      selected_fields,
        'selected_field_names': [field_names.get(f, f.replace('_', ' ').title()) for f in selected_fields],
        'field_names':          field_names,
        'short_headers':        short_headers,
        'landscape':            landscape,
        'now':                  timezone.now(),
        'print_date':           get_school_today(),
        'printed_by':           request.user.get_full_name() or request.user.username,
        'title':                'Account Types',
    })


# =============================================================================
# ACCOUNT VIEWS
# =============================================================================

@login_required
def account_list(request):
    filter_form = AccountFilterForm(request.GET or None)
    accounts    = get_filtered_accounts(request)

    stats = {
        'total':         accounts.count(),
        'active':        accounts.filter(is_active=True).count(),
        'bank_accounts': accounts.filter(is_bank_account=True).count(),
        'cash_accounts': accounts.filter(is_cash_account=True).count(),
        'mobile_money':  accounts.filter(is_mobile_money_account=True).count(),
        'total_balance': accounts.aggregate(Sum('current_balance'))['current_balance__sum'] or Decimal('0.00'),
    }

    paginator     = Paginator(accounts, 20)
    accounts_page = paginator.get_page(request.GET.get('page', 1))
    is_htmx       = request.headers.get('HX-Request') == 'true'

    context = {
        'accounts_page': accounts_page,
        'paginator':     paginator,
        'stats':         stats,
        'filter_form':   filter_form,
        'is_htmx':       is_htmx,
    }

    if is_htmx:
        return render(request, 'finance/accounts/partials/_account_results.html', context)
    return render(request, 'finance/accounts/list.html', context)


@login_required
def account_create(request):
    if request.method == 'POST':
        form = AccountForm(request.POST)
        if form.is_valid():
            account = form.save()
            messages.success(request, f"Account '{account.name}' ({account.account_number}) created successfully", extra_tags='sweetalert')
            return redirect('finance:account_detail', pk=account.pk)
    else:
        form = AccountForm()
    return render(request, 'finance/accounts/form.html', {'form': form, 'title': 'Create Account'})


@login_required
def account_edit(request, pk):
    account = get_object_or_404(Account, pk=pk)

    if request.method == 'POST':
        form = AccountForm(request.POST, instance=account)
        if form.is_valid():
            account = form.save()
            messages.success(request, f"Account '{account.name}' updated successfully", extra_tags='sweetalert')
            return redirect('finance:account_detail', pk=account.pk)
    else:
        form = AccountForm(instance=account)

    return render(request, 'finance/accounts/form.html', {
        'form': form, 'account': account, 'title': f'Edit {account.name}'
    })


@login_required
def account_detail(request, pk):
    account = get_object_or_404(
        Account.objects.select_related('account_type', 'parent_account'), pk=pk
    )

    transactions = account.journal_transactions.select_related(
        'journal_entry__journal', 'journal_entry__fiscal_period'
    ).order_by('-journal_entry__entry_date', '-created_at')[:50]

    child_accounts = account.child_accounts.filter(is_active=True).annotate(
        transaction_count=Count('journal_transactions')
    ).order_by('account_number')

    debit_total = account.journal_transactions.filter(
        is_debit=True, journal_entry__status='POSTED'
    ).aggregate(Sum('amount'))['amount__sum'] or Decimal('0.00')

    credit_total = account.journal_transactions.filter(
        is_debit=False, journal_entry__status='POSTED'
    ).aggregate(Sum('amount'))['amount__sum'] or Decimal('0.00')

    context = {
        'account':           account,
        'transactions':      transactions,
        'child_accounts':    child_accounts,
        'transaction_count': account.journal_transactions.count(),
        'debit_total':       debit_total,
        'credit_total':      credit_total,
    }
    return render(request, 'finance/accounts/detail.html', context)


@login_required
def account_delete(request, pk):
    account = get_object_or_404(Account, pk=pk)

    if request.method == 'POST':
        is_htmx = request.headers.get('HX-Request') == 'true'

        for check, msg in [
            (account.journal_transactions.exists(), f"Cannot delete '{account.name}' because it has transactions"),
            (account.child_accounts.exists(),       f"Cannot delete '{account.name}' because it has child accounts"),
        ]:
            if check:
                if is_htmx:
                    r = HttpResponse()
                    r['HX-Alert-Message'] = msg
                    r['HX-Alert-Type']    = 'error'
                    r['HX-Alert-Title']   = 'Cannot Delete'
                    r['HX-Close-Modal']   = 'true'
                    return r
                messages.error(request, msg, extra_tags='sweetalert-error')
                return redirect('finance:account_list')

        account_name = f"{account.account_number} - {account.name}"
        account.delete()

        if is_htmx:
            r = HttpResponse()
            r['HX-Alert-Message'] = f"Account '{account_name}' deleted successfully"
            r['HX-Alert-Type']    = 'success'
            r['HX-Alert-Title']   = 'Deleted!'
            r['HX-Close-Modal']   = 'true'
            r['HX-Redirect']      = reverse('finance:account_list')
            return r
        messages.success(request, f"Account '{account_name}' deleted successfully", extra_tags='sweetalert')
        return redirect('finance:account_list')


@login_required
def account_toggle_active(request, pk):
    account = get_object_or_404(Account, pk=pk)

    if request.method == 'POST':
        is_htmx     = request.headers.get('HX-Request') == 'true'
        account.is_active = not account.is_active
        account.save(update_fields=['is_active', 'updated_at'])
        status = "activated" if account.is_active else "deactivated"

        if is_htmx:
            r = HttpResponse()
            r['HX-Alert-Message'] = f"Account {account.account_number} {status}"
            r['HX-Alert-Type']    = 'success' if account.is_active else 'warning'
            r['HX-Close-Modal']   = 'true'
            r['HX-Redirect']      = reverse('finance:account_detail', kwargs={'pk': pk})
            return r
        messages.success(request, f"Account {status}", extra_tags='sweetalert')
        return redirect('finance:account_detail', pk=pk)


@login_required
def account_reconcile(request, pk):
    account = get_object_or_404(Account, pk=pk)

    if not account.is_reconcilable:
        messages.warning(request, f"Account '{account.name}' is not marked as reconcilable.")
        return redirect('finance:account_detail', pk=pk)

    if request.method == 'POST':
        form = AccountReconciliationForm(request.POST)
        if form.is_valid():
            try:
                reconciliation_date = form.cleaned_data['reconciliation_date']
                statement_balance   = form.cleaned_data['statement_balance']
                difference          = account.current_balance - statement_balance

                account.last_reconciled_date   = reconciliation_date
                account.reconciliation_balance = statement_balance
                account.save(update_fields=['last_reconciled_date', 'reconciliation_balance', 'updated_at'])

                messages.success(request, f"Account reconciled. Difference: {format_money(abs(difference))}", extra_tags='sweetalert')
                return redirect('finance:account_detail', pk=account.pk)
            except Exception as e:
                logger.error(f"Error reconciling account: {e}")
                messages.error(request, f"Error reconciling account: {str(e)}", extra_tags='sweetalert-error')
    else:
        form = AccountReconciliationForm(initial={'account': account})

    return render(request, 'finance/accounts/reconcile.html', {
        'form': form, 'account': account, 'title': f'Reconcile Account - {account.name}'
    })


@login_required
def account_print_view(request, pk):
    account   = get_object_or_404(Account, pk=pk)
    date_from = request.GET.get('date_from')
    date_to   = request.GET.get('date_to')

    transactions = account.journal_transactions.select_related(
        'journal_entry__journal', 'journal_entry__fiscal_period'
    ).filter(journal_entry__status='POSTED').order_by('journal_entry__entry_date', 'created_at')

    if date_from:
        transactions = transactions.filter(journal_entry__entry_date__gte=date_from)
    if date_to:
        transactions = transactions.filter(journal_entry__entry_date__lte=date_to)

    if date_from:
        opening_debits  = account.journal_transactions.filter(is_debit=True,  journal_entry__status='POSTED', journal_entry__entry_date__lt=date_from).aggregate(Sum('amount'))['amount__sum'] or Decimal('0.00')
        opening_credits = account.journal_transactions.filter(is_debit=False, journal_entry__status='POSTED', journal_entry__entry_date__lt=date_from).aggregate(Sum('amount'))['amount__sum'] or Decimal('0.00')
        opening_balance = opening_debits - opening_credits
    else:
        opening_balance = Decimal('0.00')

    total_debits  = transactions.filter(is_debit=True).aggregate(Sum('amount'))['amount__sum']  or Decimal('0.00')
    total_credits = transactions.filter(is_debit=False).aggregate(Sum('amount'))['amount__sum'] or Decimal('0.00')

    return render(request, 'finance/accounts/print_statement.html', {
        **get_print_school_context(request),
        'account':          account,
        'transactions':     transactions,
        'opening_balance':  opening_balance,
        'total_debits':     total_debits,
        'total_credits':    total_credits,
        'closing_balance':  opening_balance + total_debits - total_credits,
        'date_from':        date_from,
        'date_to':          date_to,
        'now':              timezone.now(),
        'title':            f'Account Statement - {account.name}',
    })


@login_required
def account_list_print_view(request):
    FIELD_NAMES_FULL = {
        'account_number':  'Account Number',
        'name':            'Account Name',
        'account_type':    'Account Type',
        'category':        'Category',
        'parent_account':  'Parent Account',
        'current_balance': 'Current Balance',
        'opening_balance': 'Opening Balance',
        'is_active':       'Active',
        'is_reconcilable': 'Reconcilable',
        'last_reconciled': 'Last Reconciled',
        'description':     'Description',
    }
    FIELD_NAMES_SHORT = {
        'account_number':  'Acc. No.',
        'name':            'Name',
        'account_type':    'Type',
        'category':        'Category',
        'parent_account':  'Parent',
        'current_balance': 'Balance',
        'opening_balance': 'Opening',
        'is_active':       'Active',
        'is_reconcilable': 'Reconcil.',
        'last_reconciled': 'Last Recon.',
        'description':     'Description',
    }
    DEFAULT_FIELDS  = ['account_number', 'name', 'account_type', 'category', 'current_balance', 'is_active']
    selected_fields = request.GET.getlist('fields') or DEFAULT_FIELDS
    short_headers   = request.GET.get('short_headers', 'false').lower() == 'true'
    landscape       = request.GET.get('landscape', 'true').lower() == 'true'
    include_stats   = request.GET.get('include_stats', 'true').lower() == 'true'
    field_names     = FIELD_NAMES_SHORT if short_headers else FIELD_NAMES_FULL
    accounts        = get_filtered_accounts(request)

    stats = None
    if include_stats:
        stats = {
            'total':         accounts.count(),
            'active':        accounts.filter(is_active=True).count(),
            'bank_accounts': accounts.filter(is_bank_account=True).count(),
            'cash_accounts': accounts.filter(is_cash_account=True).count(),
            'mobile_money':  accounts.filter(is_mobile_money_account=True).count(),
            'total_balance': accounts.aggregate(Sum('current_balance'))['current_balance__sum'] or Decimal('0.00'),
        }

    if accounts.count() > MAX_PRINT_RECORDS:
        accounts = accounts[:MAX_PRINT_RECORDS]

    return render(request, 'finance/accounts/print_list.html', {
        **get_print_school_context(request),
        'accounts':             accounts,
        'stats':                stats,
        'selected_fields':      selected_fields,
        'selected_field_names': [field_names.get(f, f.replace('_', ' ').title()) for f in selected_fields],
        'field_names':          field_names,
        'short_headers':        short_headers,
        'landscape':            landscape,
        'now':                  timezone.now(),
        'print_date':           get_school_today(),
        'printed_by':           request.user.get_full_name() or request.user.username,
        'title':                'Chart of Accounts',
    })


@login_required
def export_accounts_excel(request):
    ALL_COLUMNS = [
        ('account_number', 'Account Number', lambda o: o.account_number),
        ('name',           'Account Name',   lambda o: o.name),
        ('account_type',   'Account Type',   lambda o: o.account_type.name),
        ('category',       'Category',       lambda o: o.get_category_display()),
        ('current_balance','Current Balance',lambda o: float(o.current_balance)),
        ('opening_balance','Opening Balance',lambda o: float(o.opening_balance)),
        ('is_active',      'Active',         lambda o: 'Yes' if o.is_active else 'No'),
        ('is_reconcilable','Reconcilable',   lambda o: 'Yes' if o.is_reconcilable else 'No'),
        ('description',    'Description',    lambda o: o.description or ''),
    ]
    DEFAULT_FIELDS = ['account_number', 'name', 'account_type', 'category', 'current_balance', 'is_active']
    accounts = get_filtered_accounts(request)
    columns  = _resolve_columns(ALL_COLUMNS, request.GET.getlist('fields'), DEFAULT_FIELDS)
    return _xlsx_response(_make_workbook('Accounts', columns, accounts), 'accounts')


# =============================================================================
# EXPENSE CATEGORY VIEWS
# =============================================================================

@login_required
def expense_category_list(request):
    filter_form = ExpenseCategoryFilterForm(request.GET or None)
    categories  = get_filtered_expense_categories(request)

    stats = {
        'total':             categories.count(),
        'active':            categories.filter(is_active=True).count(),
        'requires_approval': categories.filter(requires_approval=True).count(),
        'total_expenses':    sum(c.expense_count or 0 for c in categories),
    }

    paginator       = Paginator(categories, 10)
    categories_page = paginator.get_page(request.GET.get('page', 1))
    is_htmx         = request.headers.get('HX-Request') == 'true'

    context = {
        'categories_page': categories_page,
        'paginator':       paginator,
        'stats':           stats,
        'filter_form':     filter_form,
        'is_htmx':         is_htmx,
    }
    if is_htmx:
        return render(request, 'finance/expense_categories/partials/_category_results.html', context)
    return render(request, 'finance/expense_categories/list.html', context)


@login_required
def expense_category_create(request):
    if request.method == 'POST':
        form = ExpenseCategoryForm(request.POST)
        if form.is_valid():
            category = form.save()
            messages.success(request, f"Expense category '{category.name}' created successfully", extra_tags='sweetalert')
            return redirect('finance:expense_category_detail', pk=category.pk)
    else:
        form = ExpenseCategoryForm()
    return render(request, 'finance/expense_categories/form.html', {'form': form, 'title': 'Create Expense Category'})


@login_required
def expense_category_edit(request, pk):
    category = get_object_or_404(ExpenseCategory, pk=pk)
    if request.method == 'POST':
        form = ExpenseCategoryForm(request.POST, instance=category)
        if form.is_valid():
            category = form.save()
            messages.success(request, f"Expense category '{category.name}' updated successfully", extra_tags='sweetalert')
            return redirect('finance:expense_category_detail', pk=category.pk)
    else:
        form = ExpenseCategoryForm(instance=category)
    return render(request, 'finance/expense_categories/form.html', {
        'form': form, 'category': category, 'title': f'Edit {category.name}'
    })


@login_required
def expense_category_detail(request, pk):
    category = get_object_or_404(
        ExpenseCategory.objects.select_related('default_expense_account'), pk=pk
    )

    # FIX: removed select_related('academic_session') — Expense has no direct
    # academic_session FK; session is derived via fiscal_period.related_academic_session
    expenses = category.expenses.select_related(
        'fiscal_period', 'fiscal_period__related_academic_session'
    ).order_by('-expense_date')[:50]

    today = get_school_today()

    context = {
        'category':          category,
        'expenses':          expenses,
        'expense_count':     category.expenses.count(),
        'total_amount':      category.expenses.filter(status__in=['APPROVED', 'PAID']).aggregate(Sum('total_amount'))['total_amount__sum'] or Decimal('0.00'),
        'this_month_amount': category.expenses.filter(expense_date__gte=today.replace(day=1), status__in=['APPROVED', 'PAID']).aggregate(Sum('total_amount'))['total_amount__sum'] or Decimal('0.00'),
    }
    return render(request, 'finance/expense_categories/detail.html', context)


@login_required
def expense_category_delete(request, pk):
    category = get_object_or_404(ExpenseCategory, pk=pk)

    if request.method == 'POST':
        is_htmx = request.headers.get('HX-Request') == 'true'

        if category.expenses.exists():
            msg = f"Cannot delete '{category.name}' because it has associated expenses"
            if is_htmx:
                r = HttpResponse()
                r['HX-Alert-Message'] = msg
                r['HX-Alert-Type']    = 'error'
                r['HX-Alert-Title']   = 'Cannot Delete'
                r['HX-Close-Modal']   = 'true'
                return r
            messages.error(request, msg, extra_tags='sweetalert-error')
            return redirect('finance:expense_category_list')

        category_name = category.name
        category.delete()

        if is_htmx:
            r = HttpResponse()
            r['HX-Alert-Message'] = f"Category '{category_name}' deleted successfully"
            r['HX-Alert-Type']    = 'success'
            r['HX-Alert-Title']   = 'Deleted!'
            r['HX-Close-Modal']   = 'true'
            r['HX-Redirect']      = reverse('finance:expense_category_list')
            return r
        messages.success(request, f"Category '{category_name}' deleted successfully", extra_tags='sweetalert')
        return redirect('finance:expense_category_list')


@login_required
def expense_category_toggle_active(request, pk):
    category = get_object_or_404(ExpenseCategory, pk=pk)

    if request.method == 'POST':
        is_htmx          = request.headers.get('HX-Request') == 'true'
        category.is_active = not category.is_active
        category.save(update_fields=['is_active', 'updated_at'])
        status = "activated" if category.is_active else "deactivated"

        if is_htmx:
            r = HttpResponse()
            r['HX-Alert-Message'] = f"Category '{category.name}' {status}"
            r['HX-Alert-Type']    = 'success' if category.is_active else 'warning'
            r['HX-Close-Modal']   = 'true'
            r['HX-Redirect']      = reverse('finance:expense_category_detail', kwargs={'pk': pk})
            return r
        messages.success(request, f"Category {status}", extra_tags='sweetalert')
        return redirect('finance:expense_category_detail', pk=pk)


@login_required
def expense_category_list_print_view(request):
    FIELD_NAMES_FULL = {
        'name':              'Category Name',
        'category_type':     'Category Type',
        'default_account':   'Default Expense Account',
        'requires_approval': 'Requires Approval',
        'approval_limit':    'Approval Limit',
        'expense_count':     'No. of Expenses',
        'is_active':         'Active',
        'description':       'Description',
    }
    FIELD_NAMES_SHORT = {
        'name':              'Name',
        'category_type':     'Type',
        'default_account':   'Account',
        'requires_approval': 'Approval Req.',
        'approval_limit':    'Limit',
        'expense_count':     '# Expenses',
        'is_active':         'Active',
        'description':       'Description',
    }
    DEFAULT_FIELDS  = ['name', 'category_type', 'requires_approval', 'approval_limit', 'expense_count', 'is_active']
    selected_fields = request.GET.getlist('fields') or DEFAULT_FIELDS
    short_headers   = request.GET.get('short_headers', 'false').lower() == 'true'
    landscape       = request.GET.get('landscape', 'false').lower() == 'true'
    include_stats   = request.GET.get('include_stats', 'true').lower() == 'true'
    field_names     = FIELD_NAMES_SHORT if short_headers else FIELD_NAMES_FULL
    categories      = get_filtered_expense_categories(request)

    stats = None
    if include_stats:
        stats = {
            'total':             categories.count(),
            'active':            categories.filter(is_active=True).count(),
            'requires_approval': categories.filter(requires_approval=True).count(),
            'total_expenses':    sum(c.expense_count or 0 for c in categories),
        }

    if categories.count() > MAX_PRINT_RECORDS:
        categories = categories[:MAX_PRINT_RECORDS]

    return render(request, 'finance/expense_categories/print_list.html', {
        **get_print_school_context(request),
        'categories':           categories,
        'stats':                stats,
        'selected_fields':      selected_fields,
        'selected_field_names': [field_names.get(f, f.replace('_', ' ').title()) for f in selected_fields],
        'field_names':          field_names,
        'short_headers':        short_headers,
        'landscape':            landscape,
        'now':                  timezone.now(),
        'print_date':           get_school_today(),
        'printed_by':           request.user.get_full_name() or request.user.username,
        'title':                'Expense Categories',
    })


# =============================================================================
# EXPENSE VIEWS
# =============================================================================

@login_required
def expense_list(request):
    filter_form = ExpenseFilterForm(request.GET or None)
    expenses    = get_filtered_expenses(request)

    stats = {
        'total':            expenses.count(),
        'draft':            expenses.filter(status='DRAFT').count(),
        'pending_approval': expenses.filter(status='PENDING_APPROVAL').count(),
        'approved':         expenses.filter(status='APPROVED').count(),
        'paid':             expenses.filter(status='PAID').count(),
        'rejected':         expenses.filter(status='REJECTED').count(),
        'total_amount':     expenses.aggregate(
                                Sum('total_amount')
                            )['total_amount__sum'] or Decimal('0.00'),
        'approved_amount':  expenses.filter(
                                status__in=['APPROVED', 'PAID']
                            ).aggregate(
                                Sum('total_amount')
                            )['total_amount__sum'] or Decimal('0.00'),
    }

    paginator     = Paginator(expenses, 10)
    expenses_page = paginator.get_page(request.GET.get('page', 1))
    is_htmx       = request.headers.get('HX-Request') == 'true'

    context = {
        'expenses_page': expenses_page,
        'paginator':     paginator,
        'stats':         stats,
        'filter_form':   filter_form,
        'is_htmx':       is_htmx,
    }

    if is_htmx:
        return render(request, 'finance/expenses/partials/_expense_results.html', context)
    return render(request, 'finance/expenses/list.html', context)


from django.forms import inlineformset_factory

# extra=1 → create view gets one starter blank row
ExpenseLineFormSet = inlineformset_factory(
    Expense, ExpenseLine,
    form=ExpenseLineForm,
    extra=1,
    can_delete=True,
    min_num=0,
)

# extra=0 → edit view shows only the lines that actually exist, no ghost row
ExpenseLineEditFormSet = inlineformset_factory(
    Expense, ExpenseLine,
    form=ExpenseLineForm,
    extra=0,
    can_delete=True,
    min_num=0,
)


@login_required
def expense_create(request):
    if request.method == 'POST':
        form    = ExpenseForm(request.POST, request.FILES)
        formset = ExpenseLineFormSet(request.POST)

        if form.is_valid() and formset.is_valid():
            try:
                expense = form.save(commit=False)
                expense.requested_by_id = str(request.user.pk)
                expense.save()

                formset.instance = expense
                formset.save()

                messages.success(
                    request,
                    f"Expense {expense.expense_number} created successfully.",
                    extra_tags='sweetalert',
                )
                return redirect('finance:expense_detail', pk=expense.pk)

            except Exception as e:
                logger.error(f"Error creating expense: {e}", exc_info=True)
                messages.error(
                    request,
                    f"Error creating expense: {str(e)}",
                    extra_tags='sweetalert-error',
                )
    else:
        form    = ExpenseForm()
        formset = ExpenseLineFormSet()

    return render(request, 'finance/expenses/form.html', {
        'form':    form,
        'formset': formset,
        'title':   'Create Expense',
    })


@login_required
def expense_edit(request, pk):
    expense = get_object_or_404(Expense, pk=pk)

    if expense.status in ['APPROVED', 'PAID']:
        messages.warning(
            request,
            f"Cannot edit expense with status: {expense.get_status_display()}",
            extra_tags='sweetalert',
        )
        return redirect('finance:expense_detail', pk=pk)

    if request.method == 'POST':
        form    = ExpenseForm(request.POST, request.FILES, instance=expense)
        # FIX: use EditFormSet (extra=0) so no ghost blank row appears
        formset = ExpenseLineEditFormSet(request.POST, instance=expense)

        if form.is_valid() and formset.is_valid():
            try:
                expense = form.save()
                formset.save()

                messages.success(
                    request,
                    f"Expense {expense.expense_number} updated successfully.",
                    extra_tags='sweetalert',
                )
                return redirect('finance:expense_detail', pk=expense.pk)

            except Exception as e:
                logger.error(f"Error updating expense: {e}", exc_info=True)
                messages.error(
                    request,
                    f"Error updating expense: {str(e)}",
                    extra_tags='sweetalert-error',
                )
    else:
        form    = ExpenseForm(instance=expense)
        # FIX: extra=0 — only show the lines that already exist
        formset = ExpenseLineEditFormSet(instance=expense)

    return render(request, 'finance/expenses/form.html', {
        'form':    form,
        'formset': formset,
        'expense': expense,
        'title':   f'Edit Expense {expense.expense_number}',
    })


@login_required
def expense_detail(request, pk):
    # FIX: removed select_related('academic_session') — Expense has no direct
    # academic_session FK; session is accessible via
    # fiscal_period.related_academic_session
    expense = get_object_or_404(
        Expense.objects.select_related(
            'category',
            'fiscal_period',
            'fiscal_period__related_academic_session',
            'expense_account',
            'budget_line',
            'journal_entry',
        ).prefetch_related('lines', 'payments'),
        pk=pk,
    )

    # FIX: removed 'expense_account' — ExpenseLine no longer has this FK.
    # Account is resolved once at the Expense level from the category.
    lines    = expense.lines.select_related('unit_of_measure')
    payments = expense.payments.select_related(
        'payment_method', 'account'
    ).order_by('-payment_date')

    total_paid = sum(p.amount for p in payments if p.is_active)
    remaining  = expense.total_amount - total_paid

    paid_percentage = (
        (total_paid / expense.total_amount * 100)
        if expense.total_amount > 0 else 0
    )

    payment_progress = {
        'total_paid':      total_paid,
        'remaining':       remaining,
        'paid_percentage': min(paid_percentage, 100),
    }

    # Resolve the GL account once at the expense level — shown in the line
    # items card header so staff can see which account all lines post to.
    expense_account = expense.get_expense_account()

    return render(request, 'finance/expenses/detail.html', {
        'expense':          expense,
        'lines':            lines,
        'payments':         payments,
        'payment_progress': payment_progress,
        'expense_account':  expense_account,
    })


@login_required
def expense_print_view(request, pk):
    """
    Print-friendly view for a single expense.
    Opened in a new tab — no base.html, standalone print layout.
    Prefetches lines and payments so the template doesn't hit the DB per row.
    """
    expense = get_object_or_404(
        Expense.objects.select_related(
            'category',
            'fiscal_period',
            'fiscal_period__related_academic_session',
            'expense_account',
            'budget_line__budget',
            'journal_entry',
        ).prefetch_related(
            'lines__unit_of_measure',
            'payments__payment_method',
            'payments__account',
        ),
        pk=pk,
    )

    lines    = expense.lines.select_related('unit_of_measure')
    payments = expense.payments.select_related(
        'payment_method', 'account'
    ).order_by('-payment_date')

    total_paid = sum(p.amount for p in payments if p.is_active)

    return render(request, 'finance/expenses/print_expense.html', {
        **get_print_school_context(request),
        'expense':           expense,
        'lines':             lines,
        'payments':          payments,
        'total_paid':        total_paid,
        'remaining':         expense.total_amount - total_paid,
        'expense_account':   expense.get_expense_account(),
        'now':               timezone.now(),
    })


@login_required
def expense_delete(request, pk):
    expense = get_object_or_404(Expense, pk=pk)

    if request.method == 'POST':
        is_htmx = request.headers.get('HX-Request') == 'true'

        for check, msg in [
            (expense.status in ['APPROVED', 'PAID'],
             f"Cannot delete expense with status: {expense.get_status_display()}"),
            (expense.payments.filter(reversed=False).exists(),
             "Cannot delete expense with active payments"),
        ]:
            if check:
                if is_htmx:
                    r = HttpResponse()
                    r['HX-Alert-Message'] = msg
                    r['HX-Alert-Type']    = 'error'
                    r['HX-Alert-Title']   = 'Cannot Delete'
                    r['HX-Close-Modal']   = 'true'
                    return r
                messages.error(request, msg, extra_tags='sweetalert-error')
                return redirect('finance:expense_list')

        expense_number = expense.expense_number
        expense.delete()

        if is_htmx:
            r = HttpResponse()
            r['HX-Alert-Message'] = f"Expense {expense_number} deleted successfully"
            r['HX-Alert-Type']    = 'success'
            r['HX-Alert-Title']   = 'Deleted!'
            r['HX-Close-Modal']   = 'true'
            r['HX-Redirect']      = reverse('finance:expense_list')
            return r
        messages.success(request, f"Expense {expense_number} deleted successfully", extra_tags='sweetalert')
        return redirect('finance:expense_list')


@login_required
def expense_submit(request, pk):
    expense = get_object_or_404(Expense, pk=pk)

    if request.method == 'POST':
        is_htmx = request.headers.get('HX-Request') == 'true'

        if expense.status != 'DRAFT':
            if is_htmx:
                r = HttpResponse()
                r['HX-Alert-Message'] = "Can only submit draft expenses"
                r['HX-Alert-Type']    = 'warning'
                r['HX-Close-Modal']   = 'true'
                return r
            messages.warning(request, "Can only submit draft expenses", extra_tags='sweetalert')
            return redirect('finance:expense_detail', pk=pk)

        try:
            ExpenseService.submit_for_approval(expense, requested_by_id=request.user.pk)
            if is_htmx:
                r = HttpResponse()
                r['HX-Alert-Message'] = f"Expense {expense.expense_number} submitted for approval"
                r['HX-Alert-Type']    = 'success'
                r['HX-Close-Modal']   = 'true'
                r['HX-Redirect']      = reverse('finance:expense_detail', kwargs={'pk': pk})
                return r
            messages.success(request, "Expense submitted for approval", extra_tags='sweetalert')
            return redirect('finance:expense_detail', pk=pk)
        except Exception as e:
            logger.error(f"Error submitting expense: {e}")
            if is_htmx:
                r = HttpResponse()
                r['HX-Alert-Message'] = f"Error submitting expense: {str(e)}"
                r['HX-Alert-Type']    = 'error'
                r['HX-Close-Modal']   = 'true'
                return r
            messages.error(request, f"Error submitting expense: {str(e)}", extra_tags='sweetalert-error')
            return redirect('finance:expense_detail', pk=pk)


@login_required
def expense_approve(request, pk):
    expense = get_object_or_404(Expense, pk=pk)

    if request.method == 'POST':
        # FIX: is_htmx checked once at top of POST handler — not duplicated
        # inside try and again inside except
        is_htmx = request.headers.get('HX-Request') == 'true'
        form    = ExpenseApprovalForm(request.POST)
        if form.is_valid():
            try:
                decision = form.cleaned_data['decision']
                notes    = form.cleaned_data['notes']

                if decision == 'APPROVE':
                    ExpenseService.approve_expense(expense, approved_by_id=request.user.pk, notes=notes)
                    message    = f"Expense {expense.expense_number} approved successfully"
                    alert_type = 'success'
                else:
                    ExpenseService.reject_expense(expense, rejected_by_id=request.user.pk, reason=notes)
                    message    = f"Expense {expense.expense_number} rejected"
                    alert_type = 'warning'

                if is_htmx:
                    r = HttpResponse()
                    r['HX-Alert-Message'] = message
                    r['HX-Alert-Type']    = alert_type
                    r['HX-Close-Modal']   = 'true'
                    r['HX-Redirect']      = reverse('finance:expense_detail', kwargs={'pk': pk})
                    return r
                getattr(messages, 'success' if alert_type == 'success' else 'warning')(
                    request, message, extra_tags='sweetalert'
                )
                return redirect('finance:expense_detail', pk=pk)
            except Exception as e:
                logger.error(f"Error processing expense approval: {e}")
                if is_htmx:
                    r = HttpResponse()
                    r['HX-Alert-Message'] = f"Error: {str(e)}"
                    r['HX-Alert-Type']    = 'error'
                    r['HX-Close-Modal']   = 'true'
                    return r
                messages.error(request, f"Error: {str(e)}", extra_tags='sweetalert-error')
                return redirect('finance:expense_detail', pk=pk)

    return redirect('finance:expense_detail', pk=pk)


@login_required
def expense_cancel(request, pk):
    expense = get_object_or_404(Expense, pk=pk)

    if request.method == 'POST':
        is_htmx = request.headers.get('HX-Request') == 'true'

        if expense.status in ['PAID', 'CANCELLED']:
            msg = f"Cannot cancel expense with status: {expense.get_status_display()}"
            if is_htmx:
                r = HttpResponse()
                r['HX-Alert-Message'] = msg
                r['HX-Alert-Type']    = 'warning'
                r['HX-Close-Modal']   = 'true'
                return r
            messages.warning(request, msg, extra_tags='sweetalert')
            return redirect('finance:expense_detail', pk=pk)

        try:
            reason = request.POST.get('reason', '').strip()
            ExpenseService.cancel_expense(expense=expense, reason=reason)
            msg = f"Expense {expense.expense_number} cancelled"
            if is_htmx:
                r = HttpResponse()
                r['HX-Alert-Message'] = msg
                r['HX-Alert-Type']    = 'success'
                r['HX-Close-Modal']   = 'true'
                r['HX-Redirect']      = reverse('finance:expense_detail', kwargs={'pk': pk})
                return r
            messages.success(request, msg, extra_tags='sweetalert')
            return redirect('finance:expense_detail', pk=pk)
        except Exception as e:
            logger.error(f"Error cancelling expense {pk}: {e}")
            if is_htmx:
                r = HttpResponse()
                r['HX-Alert-Message'] = f"Error cancelling expense: {str(e)}"
                r['HX-Alert-Type']    = 'error'
                r['HX-Close-Modal']   = 'true'
                return r
            messages.error(request, f"Error cancelling expense: {str(e)}", extra_tags='sweetalert-error')
            return redirect('finance:expense_detail', pk=pk)

    return redirect('finance:expense_detail', pk=pk)


@login_required
def expense_reject(request, pk):
    """
    Standalone reject endpoint used by the reject modal (expense_reject_modal).
    The approve modal uses expense_approve with a decision field instead.
    """
    expense = get_object_or_404(Expense, pk=pk)

    if request.method == 'POST':
        is_htmx = request.headers.get('HX-Request') == 'true'

        if expense.status != 'PENDING_APPROVAL':
            msg = f"Cannot reject expense with status: {expense.get_status_display()}"
            if is_htmx:
                r = HttpResponse()
                r['HX-Alert-Message'] = msg
                r['HX-Alert-Type']    = 'warning'
                r['HX-Close-Modal']   = 'true'
                return r
            messages.warning(request, msg, extra_tags='sweetalert')
            return redirect('finance:expense_detail', pk=pk)

        rejection_reason = request.POST.get('rejection_reason', '').strip()
        if not rejection_reason:
            if is_htmx:
                r = HttpResponse()
                r['HX-Alert-Message'] = 'A rejection reason is required.'
                r['HX-Alert-Type']    = 'warning'
                r['HX-Close-Modal']   = 'false'
                return r
            messages.warning(request, 'A rejection reason is required.', extra_tags='sweetalert')
            return redirect('finance:expense_detail', pk=pk)

        try:
            ExpenseService.reject_expense(
                expense=expense, rejected_by_id=request.user.pk, reason=rejection_reason
            )
            msg = f"Expense {expense.expense_number} rejected"
            if is_htmx:
                r = HttpResponse()
                r['HX-Alert-Message'] = msg
                r['HX-Alert-Type']    = 'warning'
                r['HX-Close-Modal']   = 'true'
                r['HX-Redirect']      = reverse('finance:expense_detail', kwargs={'pk': pk})
                return r
            messages.warning(request, msg, extra_tags='sweetalert')
            return redirect('finance:expense_detail', pk=pk)
        except Exception as e:
            logger.error(f"Error rejecting expense {pk}: {e}")
            if is_htmx:
                r = HttpResponse()
                r['HX-Alert-Message'] = f"Error rejecting expense: {str(e)}"
                r['HX-Alert-Type']    = 'error'
                r['HX-Close-Modal']   = 'true'
                return r
            messages.error(request, f"Error: {str(e)}", extra_tags='sweetalert-error')
            return redirect('finance:expense_detail', pk=pk)

    return redirect('finance:expense_detail', pk=pk)

@login_required
def expense_list_print_view(request):
    # FIX: removed 'vendor_name' and 'is_recurring' — both fields were
    # removed from the Expense model. Replaced with 'payee_name'/'payee_type'.
    # FIX: removed 'academic_session' — Expense has no direct FK.
    FIELD_NAMES = {
        'expense_number':   'Expense No.',
        'category':         'Category',
        'description':      'Description',
        'expense_date':     'Date',
        'fiscal_period':    'Period',
        'payee_type':       'Payee Type',
        'payee_name':       'Payee',
        'vendor_reference': 'Vendor Ref.',
        'subtotal_amount':  'Subtotal',
        'tax_amount':       'Tax',
        'total_amount':     'Total Amount',
        'status':           'Status',
        'notes':            'Notes',
    }

    DEFAULT_FIELDS = [
        'expense_number', 'category', 'description',
        'expense_date', 'payee_name', 'total_amount', 'status',
    ]

    selected_fields = request.GET.getlist('fields') or DEFAULT_FIELDS
    include_stats   = request.GET.get('include_stats') == 'true'
    landscape       = request.GET.get('landscape') == 'true'

    expenses = get_filtered_expenses(request)

    # Stats computed BEFORE slice — same pattern as fees/views.py
    stats = {
        'total':            expenses.count(),
        'draft':            expenses.filter(status='DRAFT').count(),
        'pending_approval': expenses.filter(status='PENDING_APPROVAL').count(),
        'approved':         expenses.filter(status='APPROVED').count(),
        'paid':             expenses.filter(status='PAID').count(),
        'rejected':         expenses.filter(status='REJECTED').count(),
        'total_amount':     expenses.aggregate(Sum('total_amount'))['total_amount__sum'] or Decimal('0.00'),
        'approved_amount':  expenses.filter(status__in=['APPROVED', 'PAID']).aggregate(
            Sum('total_amount'))['total_amount__sum'] or Decimal('0.00'),
    }

    if expenses.count() > MAX_PRINT_RECORDS:
        expenses = expenses[:MAX_PRINT_RECORDS]

    return render(request, 'finance/expenses/print_expenses_list.html', {
        # FIX: was _get_print_school_context (undefined); correct name is
        # get_print_school_context (imported from core.view_helpers)
        **get_print_school_context(request),
        'expenses':             expenses,
        'stats':                stats,
        'selected_fields':      selected_fields,
        'selected_field_names': [FIELD_NAMES.get(f, f.replace('_', ' ').title()) for f in selected_fields],
        'field_names':          FIELD_NAMES,
        'include_stats':        include_stats,
        'landscape':            landscape,
        'now':                  timezone.now(),
        'printed_by':           request.user.get_full_name() or request.user.username,
        'title':                'Expenses Report',
    })


@login_required
def export_expenses_excel(request):
    # FIX: replaced vendor_name with payee_name/payee_type (field renamed);
    # removed is_recurring (field removed from model)
    ALL_COLUMNS = [
        ('expense_number',  'Expense Number',   lambda o: o.expense_number),
        ('expense_date',    'Date',             lambda o: o.expense_date.strftime('%Y-%m-%d')),
        ('category',        'Category',         lambda o: o.category.name),
        ('description',     'Description',      lambda o: o.description),
        ('payee_type',      'Payee Type',       lambda o: o.get_payee_type_display()),
        ('payee_name',      'Payee',            lambda o: o.payee_name or ''),
        ('vendor_reference','Vendor Ref.',      lambda o: o.vendor_reference or ''),
        ('subtotal_amount', 'Subtotal',         lambda o: float(o.subtotal_amount)),
        ('tax_amount',      'Tax Amount',       lambda o: float(o.tax_amount)),
        ('total_amount',    'Total Amount',     lambda o: float(o.total_amount)),
        ('status',          'Status',           lambda o: o.get_status_display()),
        ('fiscal_period',   'Fiscal Period',    lambda o: str(o.fiscal_period) if o.fiscal_period else ''),
        ('notes',           'Notes',            lambda o: o.notes or ''),
    ]
    DEFAULT_FIELDS = [
        'expense_number', 'expense_date', 'category', 'description',
        'payee_name', 'vendor_reference', 'subtotal_amount', 'tax_amount',
        'total_amount', 'status',
    ]
    expenses = get_filtered_expenses(request)
    columns  = _resolve_columns(ALL_COLUMNS, request.GET.getlist('fields'), DEFAULT_FIELDS)
    return _xlsx_response(_make_workbook('Expenses', columns, expenses), 'expenses')



# =============================================================================
# EXPENSE PAYMENT ACTIONS  (all payments recorded/viewed via modal from expense_detail)
# =============================================================================

@login_required
def expense_payment_delete(request, pk):
    payment    = get_object_or_404(ExpensePayment, pk=pk)
    expense_pk = payment.expense_id   # save before delete

    if request.method == 'POST':
        is_htmx = request.headers.get('HX-Request') == 'true'

        for check, msg in [
            (payment.is_verified, "Cannot delete verified payment"),
            (payment.reversed,    "Cannot delete reversed payment"),
        ]:
            if check:
                if is_htmx:
                    r = HttpResponse()
                    r['HX-Alert-Message'] = msg
                    r['HX-Alert-Type']    = 'error'
                    r['HX-Alert-Title']   = 'Cannot Delete'
                    r['HX-Close-Modal']   = 'true'
                    return r
                messages.error(request, msg, extra_tags='sweetalert-error')
                return redirect('finance:expense_detail', pk=expense_pk)

        payment.delete()
        if is_htmx:
            r = HttpResponse()
            r['HX-Alert-Message'] = "Payment deleted successfully"
            r['HX-Alert-Type']    = 'success'
            r['HX-Alert-Title']   = 'Deleted!'
            r['HX-Close-Modal']   = 'true'
            r['HX-Redirect']      = reverse('finance:expense_detail', kwargs={'pk': str(expense_pk)})
            return r
        messages.success(request, "Payment deleted successfully", extra_tags='sweetalert')
        return redirect('finance:expense_detail', pk=expense_pk)


@login_required
def expense_payment_verify(request, pk):
    payment    = get_object_or_404(ExpensePayment, pk=pk)
    expense_pk = payment.expense_id

    if request.method == 'POST':
        is_htmx = request.headers.get('HX-Request') == 'true'

        if payment.is_verified:
            if is_htmx:
                r = HttpResponse()
                r['HX-Alert-Message'] = "Payment is already verified"
                r['HX-Alert-Type']    = 'warning'
                r['HX-Close-Modal']   = 'true'
                return r
            messages.warning(request, "Payment is already verified", extra_tags='sweetalert')
            return redirect('finance:expense_detail', pk=expense_pk)

        try:
            # FIX: was user=request.user — service expects verified_by_id (PK)
            ExpensePaymentService.verify_payment(payment, verified_by_id=request.user.pk)
            if is_htmx:
                r = HttpResponse()
                r['HX-Alert-Message'] = "Payment verified successfully"
                r['HX-Alert-Type']    = 'success'
                r['HX-Close-Modal']   = 'true'
                r['HX-Redirect']      = reverse('finance:expense_detail', kwargs={'pk': str(expense_pk)})
                return r
            messages.success(request, "Payment verified successfully", extra_tags='sweetalert')
            return redirect('finance:expense_detail', pk=expense_pk)
        except Exception as e:
            logger.error(f"Error verifying payment: {e}")
            if is_htmx:
                r = HttpResponse()
                r['HX-Alert-Message'] = f"Error: {str(e)}"
                r['HX-Alert-Type']    = 'error'
                r['HX-Close-Modal']   = 'true'
                return r
            messages.error(request, f"Error: {str(e)}", extra_tags='sweetalert-error')
            return redirect('finance:expense_detail', pk=expense_pk)


@login_required
def expense_payment_reverse(request, pk):
    payment    = get_object_or_404(ExpensePayment, pk=pk)
    expense_pk = payment.expense_id

    if request.method == 'POST':
        form = ExpensePaymentReversalForm(payment, request.user, request.POST)
        if form.is_valid():
            # FIX: is_htmx checked once at top of valid POST, not duplicated
            is_htmx = request.headers.get('HX-Request') == 'true'
            try:
                # FIX: was user=request.user — service expects reversed_by_id (PK)
                ExpensePaymentService.reverse_payment(
                    payment=payment,
                    reversed_by_id=request.user.pk,
                    reason=form.cleaned_data['reversal_reason'],
                )
                if is_htmx:
                    r = HttpResponse()
                    r['HX-Alert-Message'] = "Payment reversed successfully"
                    r['HX-Alert-Type']    = 'success'
                    r['HX-Close-Modal']   = 'true'
                    r['HX-Redirect']      = reverse('finance:expense_detail', kwargs={'pk': str(expense_pk)})
                    return r
                messages.success(request, "Payment reversed successfully", extra_tags='sweetalert')
                return redirect('finance:expense_detail', pk=expense_pk)
            except Exception as e:
                logger.error(f"Error reversing payment: {e}")
                if is_htmx:
                    r = HttpResponse()
                    r['HX-Alert-Message'] = f"Error: {str(e)}"
                    r['HX-Alert-Type']    = 'error'
                    r['HX-Close-Modal']   = 'true'
                    return r
                messages.error(request, f"Error: {str(e)}", extra_tags='sweetalert-error')
                return redirect('finance:expense_detail', pk=expense_pk)


@login_required
def bulk_expense_payment_verification(request):
    if request.method == 'POST':
        form = BulkExpensePaymentVerificationForm(request.POST)
        if form.is_valid():
            try:
                payment_ids = form.cleaned_data['payment_ids']
                payments    = ExpensePayment.objects.filter(
                    id__in=payment_ids,
                    is_verified=False,
                    status__in=['PROCESSING', 'PROCESSED'],
                    reversed=False,
                )
                verified_count = 0
                with transaction.atomic():
                    for payment in payments:
                        try:
                            # FIX: was user=request.user — service expects verified_by_id (PK)
                            ExpensePaymentService.verify_payment(
                                payment, verified_by_id=request.user.pk
                            )
                            verified_count += 1
                        except Exception as e:
                            logger.error(f"Error verifying payment {payment.pk}: {e}")

                is_htmx = request.headers.get('HX-Request') == 'true'
                msg = f"Successfully verified {verified_count} payment(s)"
                if is_htmx:
                    r = HttpResponse()
                    r['HX-Alert-Message'] = msg
                    r['HX-Alert-Type']    = 'success'
                    r['HX-Close-Modal']   = 'true'
                    r['HX-Redirect']      = reverse('finance:expense_list')
                    return r
                messages.success(request, msg, extra_tags='sweetalert')
            except Exception as e:
                logger.error(f"Error in bulk verification: {e}")
                messages.error(request, f"Error: {str(e)}", extra_tags='sweetalert-error')

    # FIX: was expense_payment_list (removed) — redirect to expense list instead
    return redirect('finance:expense_list')


@login_required
def expense_payment_print_view(request):
    payments = get_filtered_expense_payments(request)

    # Stats computed BEFORE any slice
    active = payments.filter(reversed=False)
    stats  = {
        'total':              payments.count(),
        'pending':            payments.filter(status='PENDING').count(),
        'processing':         payments.filter(status='PROCESSING').count(),
        'processed':          payments.filter(status='PROCESSED').count(),
        'verified':           payments.filter(is_verified=True).count(),
        'reversed':           payments.filter(reversed=True).count(),
        'total_amount':       active.aggregate(Sum('amount'))['amount__sum']               or Decimal('0.00'),
        'total_fees':         active.aggregate(Sum('processing_fee'))['processing_fee__sum'] or Decimal('0.00'),
        'total_bank_charges': active.aggregate(Sum('bank_charges'))['bank_charges__sum']   or Decimal('0.00'),
    }

    if payments.count() > MAX_PRINT_RECORDS:
        payments = payments[:MAX_PRINT_RECORDS]

    return render(request, 'finance/expense_payments/print.html', {
        **get_print_school_context(request),
        'payments': payments,
        'stats':    stats,
        'now':      timezone.now(),
        'title':    'Expense Payments Report',
    })


@login_required
def export_expense_payments_excel(request):
    ALL_COLUMNS = [
        ('reference_number', 'Reference',        lambda o: o.reference_number or ''),
        ('transaction_id',   'Transaction ID',   lambda o: o.transaction_id or ''),
        ('payment_date',     'Date',             lambda o: o.payment_date.strftime('%Y-%m-%d')),
        ('expense_number',   'Expense Number',   lambda o: o.expense.expense_number),
        # FIX: was expense.vendor_name — field renamed to payee_name
        ('payee_name',       'Payee',            lambda o: o.expense.payee_name or ''),
        ('amount',           'Amount',           lambda o: float(o.amount)),
        ('effective_amount', 'Effective Amount', lambda o: float(o.effective_amount)),
        ('processing_fee',   'Processing Fee',   lambda o: float(o.processing_fee)),
        ('bank_charges',     'Bank Charges',     lambda o: float(o.bank_charges)),
        ('payment_method',   'Payment Method',   lambda o: o.payment_method.name if o.payment_method else ''),
        ('account',          'Account',          lambda o: o.account.name if o.account else ''),
        ('status',           'Status',           lambda o: o.get_status_display()),
        ('payment_state',    'Payment State',    lambda o: o.payment_state),
        ('is_verified',      'Verified',         lambda o: 'Yes' if o.is_verified else 'No'),
        ('reversed',         'Reversed',         lambda o: 'Yes' if o.reversed else 'No'),
        ('batch_number',     'Batch Number',     lambda o: o.batch_number or ''),
        ('fiscal_period',    'Fiscal Period',    lambda o: str(o.fiscal_period) if o.fiscal_period else ''),
    ]
    DEFAULT_FIELDS = [
        'reference_number', 'payment_date', 'expense_number', 'payee_name',
        'amount', 'effective_amount', 'processing_fee', 'bank_charges',
        'payment_method', 'account', 'status', 'payment_state', 'is_verified', 'reversed',
    ]
    payments = get_filtered_expense_payments(request)
    columns  = _resolve_columns(ALL_COLUMNS, request.GET.getlist('fields'), DEFAULT_FIELDS)
    return _xlsx_response(_make_workbook('Expense Payments', columns, payments), 'expense_payments')


@login_required
def expense_payment_print_receipt(request, pk):
    payment = get_object_or_404(
        ExpensePayment.objects.select_related(
            'expense__category',
            'expense__fiscal_period',
            'payment_method',
            'account',
            'fiscal_period',
            'journal_entry',
            'reversal_journal_entry',
        ),
        pk=pk,
    )

    return render(request, 'finance/expense_payments/print_receipt.html', {
        **get_print_school_context(request),
        'payment':           payment,
        'performed_by_user': payment.get_performed_by_user(),
        'verified_by_user':  payment.get_verified_by_user(),
        'reversed_by_user':  payment.get_reversed_by_user(),
        'total_with_fees':   payment.total_amount_including_fees,
        'effective_amount':  payment.effective_amount,
        'now':               timezone.now(),
        'title':             f'Payment Receipt – {payment.reference_number or payment.pk}',
    })


@login_required
def expense_payment_reversal_detail(request, pk):
    payment = get_object_or_404(
        ExpensePayment.objects.select_related(
            'expense__category',
            'expense__fiscal_period',
            'payment_method',
            'account',
            'fiscal_period',
            'journal_entry',
            'reversal_journal_entry',
        ),
        pk=pk,
    )

    if not payment.reversed:
        messages.warning(request, "This payment has not been reversed.", extra_tags='sweetalert')
        return redirect('finance:expense_detail', pk=payment.expense_id)

    return render(request, 'finance/expense_payments/reversal_detail.html', {
        'payment':                   payment,
        'performed_by_user':         payment.get_performed_by_user(),
        'verified_by_user':          payment.get_verified_by_user(),
        'reversed_by_user':          payment.get_reversed_by_user(),
        'reversal_approved_by_user': payment.get_reversal_approved_by_user(),
        'audit_trail':               payment.get_audit_trail(),
        'title':                     f'Reversal Details – {payment.reference_number or payment.pk}',
    })


@login_required
def expense_payment_reversal_print(request, pk):
    """Printable reversal notice for a reversed expense payment."""
    payment = get_object_or_404(
        ExpensePayment.objects.select_related(
            'expense__category',
            'expense__fiscal_period',
            'payment_method',
            'account',
            'fiscal_period',
            'journal_entry',
            'reversal_journal_entry',
        ),
        pk=pk,
    )

    if not payment.reversed:
        return redirect('finance:expense_detail', pk=payment.expense_id)

    return render(request, 'finance/expense_payments/print_reversal.html', {
        **get_print_school_context(request),
        'payment':                   payment,
        'performed_by_user':         payment.get_performed_by_user(),
        'verified_by_user':          payment.get_verified_by_user(),
        'reversed_by_user':          payment.get_reversed_by_user(),
        'reversal_approved_by_user': payment.get_reversal_approved_by_user(),
        'audit_trail':               payment.get_audit_trail(),
        'total_with_fees':           payment.total_amount_including_fees,
        'now':                       timezone.now(),
        'title':                     f'Reversal Notice – {payment.reference_number or payment.pk}',
    })


# =============================================================================
# JOURNAL VIEWS
# =============================================================================

@login_required
def journal_list(request):
    from django.db.models import Count, Sum, Q

    filter_form = JournalFilterForm(request.GET or None)
    is_htmx     = request.headers.get('HX-Request') == 'true'

    # ── Base queryset ─────────────────────────────────────────────────
    journals = get_filtered_journals(request)

    # ── Stats — single aggregate query ────────────────────────────────
    stats_agg = journals.aggregate(
        total         = Count('pk'),
        active        = Count('pk', filter=Q(is_active=True)),
        general       = Count('pk', filter=Q(journal_type='GENERAL')),
        expenses      = Count('pk', filter=Q(journal_type='EXPENSES')),
        total_entries = Sum('entry_count'),
    )
    stats = {k: v or 0 for k, v in stats_agg.items()}

    # ── Paginator — inject fast count ─────────────────────────────────
    paginator        = Paginator(journals, 10)
    paginator._count = stats['total']  # reuse count already computed
    journals_page    = paginator.get_page(request.GET.get('page', 1))

    context = {
        'journals_page': journals_page,
        'paginator':     paginator,
        'stats':         stats,
        'filter_form':   filter_form,
        'is_htmx':       is_htmx,
    }

    if is_htmx:
        return render(request, 'finance/journals/partials/_journal_results.html', context)
    return render(request, 'finance/journals/list.html', context)


@login_required
def journal_create(request):
    if request.method == 'POST':
        form = JournalForm(request.POST)
        if form.is_valid():
            journal = form.save()
            messages.success(request, f"Journal '{journal.name}' created successfully", extra_tags='sweetalert')
            return redirect('finance:journal_detail', pk=journal.pk)
    else:
        form = JournalForm()
    return render(request, 'finance/journals/form.html', {'form': form, 'title': 'Create Journal'})


@login_required
def journal_edit(request, pk):
    journal = get_object_or_404(Journal, pk=pk)
    if request.method == 'POST':
        form = JournalForm(request.POST, instance=journal)
        if form.is_valid():
            journal = form.save()
            messages.success(request, f"Journal '{journal.name}' updated successfully", extra_tags='sweetalert')
            return redirect('finance:journal_detail', pk=journal.pk)
    else:
        form = JournalForm(instance=journal)
    return render(request, 'finance/journals/form.html', {
        'form': form, 'journal': journal, 'title': f'Edit {journal.name}'
    })


@login_required
def journal_detail(request, pk):
    from django.db.models import Count, Sum, Q

    journal = get_object_or_404(Journal, pk=pk)

    # ── Stats — single aggregate on journal's entries ──────────────────
    stats_agg = journal.entries.aggregate(
        total_entries  = Count('pk'),
        posted_count   = Count('pk', filter=Q(status='POSTED')),
        draft_count    = Count('pk', filter=Q(status='DRAFT')),
        reversed_count = Count('pk', filter=Q(status='REVERSED')),
    )
    stats = {k: v or 0 for k, v in stats_agg.items()}

    # ── Paginated entries ──────────────────────────────────────────────
    entries_qs = journal.entries.select_related(
        'fiscal_period', 'academic_session'
    ).annotate(
        transaction_count = Count('transactions', distinct=True),
        total_debit       = Sum(
            'transactions__amount',
            filter=Q(transactions__is_debit=True),
        ),
        total_credit      = Sum(
            'transactions__amount',
            filter=Q(transactions__is_debit=False),
        ),
    ).order_by('-entry_date', '-created_at')

    paginator      = Paginator(entries_qs, 20)
    paginator._count = stats['total_entries']
    entries_page   = paginator.get_page(request.GET.get('page', 1))
    is_htmx        = request.headers.get('HX-Request') == 'true'

    context = {
        'journal':      journal,
        'entries_page': entries_page,
        'paginator':    paginator,
        'stats':        stats,
        'is_htmx':      is_htmx,
    }

    if is_htmx:
        return render(request, 'finance/journals/partials/_journal_entries.html', context)
    return render(request, 'finance/journals/detail.html', context)


@login_required
def journal_delete(request, pk):
    journal = get_object_or_404(Journal, pk=pk)

    if request.method == 'POST':
        is_htmx = request.headers.get('HX-Request') == 'true'

        if journal.entries.exists():
            msg = f"Cannot delete '{journal.name}' because it has entries"
            if is_htmx:
                r = HttpResponse()
                r['HX-Alert-Message'] = msg
                r['HX-Alert-Type']    = 'error'
                r['HX-Alert-Title']   = 'Cannot Delete'
                r['HX-Close-Modal']   = 'true'
                return r
            messages.error(request, msg, extra_tags='sweetalert-error')
            return redirect('finance:journal_list')

        journal_name = journal.name
        journal.delete()
        if is_htmx:
            r = HttpResponse()
            r['HX-Alert-Message'] = f"Journal '{journal_name}' deleted successfully"
            r['HX-Alert-Type']    = 'success'
            r['HX-Alert-Title']   = 'Deleted!'
            r['HX-Close-Modal']   = 'true'
            r['HX-Redirect']      = reverse('finance:journal_list')
            return r
        messages.success(request, f"Journal '{journal_name}' deleted successfully", extra_tags='sweetalert')
        return redirect('finance:journal_list')


@login_required
def journal_toggle_active(request, pk):
    journal = get_object_or_404(Journal, pk=pk)
    if request.method == 'POST':
        is_htmx       = request.headers.get('HX-Request') == 'true'
        journal.is_active = not journal.is_active
        journal.save(update_fields=['is_active', 'updated_at'])
        status = "activated" if journal.is_active else "deactivated"

        if is_htmx:
            r = HttpResponse()
            r['HX-Alert-Message'] = f"Journal '{journal.name}' {status}"
            r['HX-Alert-Type']    = 'success' if journal.is_active else 'warning'
            r['HX-Close-Modal']   = 'true'
            r['HX-Redirect']      = reverse('finance:journal_detail', kwargs={'pk': pk})
            return r
        messages.success(request, f"Journal {status}", extra_tags='sweetalert')
        return redirect('finance:journal_detail', pk=pk)


@login_required
def journal_list_print_view(request):
    FIELD_NAMES_FULL = {
        'name':         'Journal Name',
        'journal_type': 'Journal Type',
        'entry_count':  'No. of Entries',
        'is_active':    'Active',
        'description':  'Description',
    }
    FIELD_NAMES_SHORT = {
        'name':         'Name',
        'journal_type': 'Type',
        'entry_count':  '# Entries',
        'is_active':    'Active',
        'description':  'Description',
    }
    DEFAULT_FIELDS  = ['name', 'journal_type', 'entry_count', 'is_active']
    selected_fields = request.GET.getlist('fields') or DEFAULT_FIELDS
    short_headers   = request.GET.get('short_headers', 'false').lower() == 'true'
    landscape       = request.GET.get('landscape', 'false').lower() == 'true'
    include_stats   = request.GET.get('include_stats', 'true').lower() == 'true'
    field_names     = FIELD_NAMES_SHORT if short_headers else FIELD_NAMES_FULL
    journals        = get_filtered_journals(request)

    stats = None
    if include_stats:
        stats = {
            'total':         journals.count(),
            'active':        journals.filter(is_active=True).count(),
            'general':       journals.filter(journal_type='GENERAL').count(),
            'expenses':      journals.filter(journal_type='EXPENSES').count(),
            'total_entries': sum(j.entry_count or 0 for j in journals),
        }

    if journals.count() > MAX_PRINT_RECORDS:
        journals = journals[:MAX_PRINT_RECORDS]

    return render(request, 'finance/journals/print_list.html', {
        **get_print_school_context(request),
        'journals':             journals,
        'stats':                stats,
        'selected_fields':      selected_fields,
        'selected_field_names': [field_names.get(f, f.replace('_', ' ').title()) for f in selected_fields],
        'field_names':          field_names,
        'short_headers':        short_headers,
        'landscape':            landscape,
        'now':                  timezone.now(),
        'print_date':           get_school_today(),
        'printed_by':           request.user.get_full_name() or request.user.username,
        'title':                'Journals',
    })


# =============================================================================
# JOURNAL ENTRY VIEWS
# =============================================================================

@login_required
def journal_entry_detail(request, pk):
    entry = get_object_or_404(
        JournalEntry.objects.select_related(
            'journal', 'academic_session', 'fiscal_period', 'original_entry'
        ).prefetch_related('transactions__account'),
        pk=pk,
    )

    transactions = entry.transactions.select_related('account__account_type').order_by('id')
    debit_total  = sum(t.amount for t in transactions if t.is_debit)
    credit_total = sum(t.amount for t in transactions if not t.is_debit)

    return render(request, 'finance/journal_entries/detail.html', {
        'entry':        entry,
        'transactions': transactions,
        'debit_total':  debit_total,
        'credit_total': credit_total,
        'is_balanced':  debit_total == credit_total,
    })


@login_required
def journal_entry_delete(request, pk):
    entry = get_object_or_404(JournalEntry, pk=pk)

    if request.method == 'POST':
        is_htmx = request.headers.get('HX-Request') == 'true'

        if entry.status == 'POSTED':
            msg = "Cannot delete posted entries - use reversal instead"
            if is_htmx:
                r = HttpResponse()
                r['HX-Alert-Message'] = msg
                r['HX-Alert-Type']    = 'error'
                r['HX-Alert-Title']   = 'Cannot Delete'
                r['HX-Close-Modal']   = 'true'
                return r
            messages.error(request, msg, extra_tags='sweetalert-error')
            return redirect('finance:journal_entry_list')

        entry_number = entry.entry_number
        entry.delete()
        if is_htmx:
            r = HttpResponse()
            r['HX-Alert-Message'] = f"Entry {entry_number} deleted successfully"
            r['HX-Alert-Type']    = 'success'
            r['HX-Alert-Title']   = 'Deleted!'
            r['HX-Close-Modal']   = 'true'
            r['HX-Redirect']      = reverse('finance:journal_entry_list')
            return r
        messages.success(request, f"Entry {entry_number} deleted successfully", extra_tags='sweetalert')
        return redirect('finance:journal_entry_list')


@login_required
def journal_entry_post(request, pk):
    entry = get_object_or_404(JournalEntry, pk=pk)

    if request.method == 'POST':
        is_htmx = request.headers.get('HX-Request') == 'true'

        if entry.status != 'DRAFT':
            if is_htmx:
                r = HttpResponse()
                r['HX-Alert-Message'] = "Only draft entries can be posted"
                r['HX-Alert-Type']    = 'warning'
                r['HX-Close-Modal']   = 'true'
                return r
            messages.warning(request, "Only draft entries can be posted", extra_tags='sweetalert')
            return redirect('finance:journal_entry_detail', pk=pk)

        try:
            JournalEntryService.post_entry(entry, user=request.user)
            if is_htmx:
                r = HttpResponse()
                r['HX-Alert-Message'] = f"Entry {entry.entry_number} posted successfully"
                r['HX-Alert-Type']    = 'success'
                r['HX-Close-Modal']   = 'true'
                r['HX-Redirect']      = reverse('finance:journal_entry_detail', kwargs={'pk': pk})
                return r
            messages.success(request, "Entry posted successfully", extra_tags='sweetalert')
            return redirect('finance:journal_entry_detail', pk=pk)
        except Exception as e:
            logger.error(f"Error posting entry: {e}")
            if is_htmx:
                r = HttpResponse()
                r['HX-Alert-Message'] = f"Error: {str(e)}"
                r['HX-Alert-Type']    = 'error'
                r['HX-Close-Modal']   = 'true'
                return r
            messages.error(request, f"Error: {str(e)}", extra_tags='sweetalert-error')
            return redirect('finance:journal_entry_detail', pk=pk)


@login_required
def journal_entry_reverse(request, pk):
    entry = get_object_or_404(JournalEntry, pk=pk)

    if request.method == 'POST':
        # FIX: is_htmx checked once at top of POST handler
        is_htmx = request.headers.get('HX-Request') == 'true'
        form    = JournalEntryReversalForm(request.POST)
        if form.is_valid():
            try:
                reversal_entry = JournalEntryService.reverse_entry(
                    entry=entry,
                    user=request.user,
                    reversal_date=form.cleaned_data['reversal_date'],
                    reason=form.cleaned_data['reversal_reason'],
                )
                if is_htmx:
                    r = HttpResponse()
                    r['HX-Alert-Message'] = f"Entry reversed. Reversal entry: {reversal_entry.entry_number}"
                    r['HX-Alert-Type']    = 'success'
                    r['HX-Close-Modal']   = 'true'
                    r['HX-Redirect']      = reverse('finance:journal_entry_detail', kwargs={'pk': reversal_entry.pk})
                    return r
                messages.success(request, f"Entry reversed. Reversal entry: {reversal_entry.entry_number}", extra_tags='sweetalert')
                return redirect('finance:journal_entry_detail', pk=reversal_entry.pk)
            except Exception as e:
                logger.error(f"Error reversing entry: {e}")
                if is_htmx:
                    r = HttpResponse()
                    r['HX-Alert-Message'] = f"Error: {str(e)}"
                    r['HX-Alert-Type']    = 'error'
                    r['HX-Close-Modal']   = 'true'
                    return r
                messages.error(request, f"Error: {str(e)}", extra_tags='sweetalert-error')
                return redirect('finance:journal_entry_detail', pk=pk)

    return redirect('finance:journal_entry_detail', pk=pk)


@login_required
def journal_entry_print_view(request, pk):
    entry = get_object_or_404(
        JournalEntry.objects.select_related(
            'journal', 'fiscal_period', 'academic_session'
        ).prefetch_related('transactions__account'),
        pk=pk,
    )
    return render(request, 'finance/journal_entries/print.html', {
        **get_print_school_context(request),
        'entry': entry,
        'now':   timezone.now(),
        'title': f'Journal Entry {entry.entry_number}',
    })


@login_required
def journal_entry_list_print_view(request):
    FIELD_NAMES_FULL = {
        'entry_number':  'Entry Number',
        'entry_date':    'Entry Date',
        'journal':       'Journal',
        'fiscal_period': 'Fiscal Period',
        'description':   'Description',
        'reference':     'Reference Number',
        'total_debit':   'Total Debit',
        'total_credit':  'Total Credit',
        'status':        'Status',
    }
    FIELD_NAMES_SHORT = {
        'entry_number':  'Entry No.',
        'entry_date':    'Date',
        'journal':       'Journal',
        'fiscal_period': 'Period',
        'description':   'Description',
        'reference':     'Ref.',
        'total_debit':   'Debit',
        'total_credit':  'Credit',
        'status':        'Status',
    }
    DEFAULT_FIELDS  = ['entry_number', 'entry_date', 'journal', 'description', 'total_debit', 'total_credit', 'status']
    selected_fields = request.GET.getlist('fields') or DEFAULT_FIELDS
    short_headers   = request.GET.get('short_headers', 'false').lower() == 'true'
    landscape       = request.GET.get('landscape', 'true').lower() == 'true'
    include_stats   = request.GET.get('include_stats', 'true').lower() == 'true'
    field_names     = FIELD_NAMES_SHORT if short_headers else FIELD_NAMES_FULL
    entries         = get_filtered_journal_entries(request)

    stats = None
    if include_stats:
        posted = entries.filter(status='POSTED')
        stats  = {
            'total':         entries.count(),
            'draft':         entries.filter(status='DRAFT').count(),
            'posted':        entries.filter(status='POSTED').count(),
            'reversed':      entries.filter(status='REVERSED').count(),
            'total_debits':  posted.aggregate(Sum('total_debit'))['total_debit__sum']   or Decimal('0.00'),
            'total_credits': posted.aggregate(Sum('total_credit'))['total_credit__sum'] or Decimal('0.00'),
        }

    if entries.count() > MAX_PRINT_RECORDS:
        entries = entries[:MAX_PRINT_RECORDS]

    return render(request, 'finance/journal_entries/print_list.html', {
        **get_print_school_context(request),
        'entries':              entries,
        'stats':                stats,
        'selected_fields':      selected_fields,
        'selected_field_names': [field_names.get(f, f.replace('_', ' ').title()) for f in selected_fields],
        'field_names':          field_names,
        'short_headers':        short_headers,
        'landscape':            landscape,
        'now':                  timezone.now(),
        'print_date':           get_school_today(),
        'printed_by':           request.user.get_full_name() or request.user.username,
        'title':                'Journal Entries',
    })


@login_required
def export_journal_entries_excel(request):
    ALL_COLUMNS = [
        ('entry_number',  'Entry Number',  lambda o: o.entry_number),
        ('entry_date',    'Date',          lambda o: o.entry_date.strftime('%Y-%m-%d')),
        ('journal',       'Journal',       lambda o: o.journal.name),
        ('description',   'Description',  lambda o: o.description),
        ('total_debit',   'Total Debit',  lambda o: float(o.total_debit or 0)),
        ('total_credit',  'Total Credit', lambda o: float(o.total_credit or 0)),
        ('status',        'Status',       lambda o: o.get_status_display()),
        ('fiscal_period', 'Fiscal Period',lambda o: str(o.fiscal_period) if o.fiscal_period else ''),
        ('reference',     'Reference',    lambda o: o.reference_number or ''),
        ('notes',         'Notes',        lambda o: o.notes or ''),
    ]
    DEFAULT_FIELDS = [
        'entry_number', 'entry_date', 'journal', 'description',
        'total_debit', 'total_credit', 'status',
    ]
    entries = get_filtered_journal_entries(request)
    columns = _resolve_columns(ALL_COLUMNS, request.GET.getlist('fields'), DEFAULT_FIELDS)
    return _xlsx_response(_make_workbook('Journal Entries', columns, entries), 'journal_entries')


# =============================================================================
# BUDGET VIEWS
# =============================================================================

@login_required
def budget_list(request):
    filter_form = BudgetFilterForm(request.GET or None)
    budgets     = get_filtered_budgets(request)

    stats = {
        'total':                budgets.count(),
        'draft':                budgets.filter(status='DRAFT').count(),
        'approved':             budgets.filter(status='APPROVED').count(),
        'active':               budgets.filter(status='ACTIVE').count(),
        'closed':               budgets.filter(status='CLOSED').count(),
        'total_revenue_budget': budgets.aggregate(Sum('total_revenue_budget'))['total_revenue_budget__sum'] or Decimal('0.00'),
        'total_expense_budget': budgets.aggregate(Sum('total_expense_budget'))['total_expense_budget__sum'] or Decimal('0.00'),
    }

    paginator    = Paginator(budgets, 10)
    budgets_page = paginator.get_page(request.GET.get('page', 1))
    is_htmx      = request.headers.get('HX-Request') == 'true'

    context = {
        'budgets_page': budgets_page,
        'paginator':    paginator,
        'stats':        stats,
        'filter_form':  filter_form,
        'is_htmx':      is_htmx,
    }
    if is_htmx:
        return render(request, 'finance/budgets/partials/_budget_results.html', context)
    return render(request, 'finance/budgets/list.html', context)


@login_required
def budget_create(request):
    if request.method == 'POST':
        form = BudgetForm(request.POST)
        if form.is_valid():
            try:
                budget = BudgetService.create_budget(budget_data=form.cleaned_data, user=request.user)
                messages.success(request, f"Budget '{budget.name}' created successfully", extra_tags='sweetalert')
                return redirect('finance:budget_detail', pk=budget.pk)
            except Exception as e:
                logger.error(f"Error creating budget: {e}")
                messages.error(request, f"Error creating budget: {str(e)}", extra_tags='sweetalert-error')
    else:
        form = BudgetForm()
    return render(request, 'finance/budgets/form.html', {'form': form, 'title': 'Create Budget'})


@login_required
def budget_edit(request, pk):
    budget = get_object_or_404(Budget, pk=pk)

    if budget.status in ['APPROVED', 'ACTIVE', 'CLOSED']:
        messages.warning(request, f"Cannot edit budget with status: {budget.get_status_display()}", extra_tags='sweetalert')
        return redirect('finance:budget_detail', pk=pk)

    if request.method == 'POST':
        form = BudgetForm(request.POST, instance=budget)
        if form.is_valid():
            try:
                budget = form.save()
                messages.success(request, f"Budget '{budget.name}' updated successfully", extra_tags='sweetalert')
                return redirect('finance:budget_detail', pk=budget.pk)
            except Exception as e:
                logger.error(f"Error updating budget: {e}")
                messages.error(request, f"Error updating budget: {str(e)}", extra_tags='sweetalert-error')
    else:
        form = BudgetForm(instance=budget)
    return render(request, 'finance/budgets/form.html', {
        'form': form, 'budget': budget, 'title': f'Edit Budget - {budget.name}'
    })


@login_required
def budget_detail(request, pk):
    budget = get_object_or_404(
        Budget.objects.select_related(
            'fiscal_year', 'academic_session', 'parent_budget'
        ).prefetch_related('lines__account'),
        pk=pk,
    )

    revenue_lines    = budget.lines.filter(line_type='REVENUE').select_related('account')
    expense_lines    = budget.lines.filter(line_type='EXPENSE').select_related('account')
    revenue_variance = budget.total_revenue_budget - budget.actual_revenue_total
    expense_variance = budget.total_expense_budget - budget.actual_expense_total

    return render(request, 'finance/budgets/detail.html', {
        'budget':            budget,
        'revenue_lines':     revenue_lines,
        'expense_lines':     expense_lines,
        'revenue_variance':  revenue_variance,
        'expense_variance':  expense_variance,
    })


@login_required
def budget_delete(request, pk):
    budget = get_object_or_404(Budget, pk=pk)

    if request.method == 'POST':
        is_htmx = request.headers.get('HX-Request') == 'true'

        for check, msg in [
            (budget.status in ['APPROVED', 'ACTIVE', 'CLOSED'],
             f"Cannot delete budget with status: {budget.get_status_display()}"),
            (budget.child_budgets.exists(),
             "Cannot delete budget with child budgets"),
        ]:
            if check:
                if is_htmx:
                    r = HttpResponse()
                    r['HX-Alert-Message'] = msg
                    r['HX-Alert-Type']    = 'error'
                    r['HX-Alert-Title']   = 'Cannot Delete'
                    r['HX-Close-Modal']   = 'true'
                    return r
                messages.error(request, msg, extra_tags='sweetalert-error')
                return redirect('finance:budget_list')

        budget_name = budget.name
        budget.delete()
        if is_htmx:
            r = HttpResponse()
            r['HX-Alert-Message'] = f"Budget '{budget_name}' deleted successfully"
            r['HX-Alert-Type']    = 'success'
            r['HX-Alert-Title']   = 'Deleted!'
            r['HX-Close-Modal']   = 'true'
            r['HX-Redirect']      = reverse('finance:budget_list')
            return r
        messages.success(request, f"Budget '{budget_name}' deleted successfully", extra_tags='sweetalert')
        return redirect('finance:budget_list')


@login_required
def budget_approve(request, pk):
    budget = get_object_or_404(Budget, pk=pk)

    if request.method == 'POST':
        # FIX: is_htmx checked once at top of POST handler
        is_htmx = request.headers.get('HX-Request') == 'true'
        form    = BudgetApprovalForm(request.POST)
        if form.is_valid():
            try:
                decision = form.cleaned_data['decision']
                notes    = form.cleaned_data['notes']

                if decision == 'APPROVE':
                    BudgetService.approve_budget(budget=budget, user=request.user, notes=notes)
                    message    = f"Budget '{budget.name}' approved successfully"
                    alert_type = 'success'
                elif decision == 'REQUEST_REVISION':
                    budget.status = 'DRAFT'
                    budget.save(update_fields=['status', 'updated_at'])
                    message    = f"Budget '{budget.name}' sent back for revision"
                    alert_type = 'warning'
                else:
                    budget.status = 'REJECTED'
                    budget.save(update_fields=['status', 'updated_at'])
                    message    = f"Budget '{budget.name}' rejected"
                    alert_type = 'warning'

                if is_htmx:
                    r = HttpResponse()
                    r['HX-Alert-Message'] = message
                    r['HX-Alert-Type']    = alert_type
                    r['HX-Close-Modal']   = 'true'
                    r['HX-Redirect']      = reverse('finance:budget_detail', kwargs={'pk': pk})
                    return r
                getattr(messages, 'success' if alert_type == 'success' else 'warning')(
                    request, message, extra_tags='sweetalert'
                )
                return redirect('finance:budget_detail', pk=pk)
            except Exception as e:
                logger.error(f"Error processing budget approval: {e}")
                if is_htmx:
                    r = HttpResponse()
                    r['HX-Alert-Message'] = f"Error: {str(e)}"
                    r['HX-Alert-Type']    = 'error'
                    r['HX-Close-Modal']   = 'true'
                    return r
                messages.error(request, f"Error: {str(e)}", extra_tags='sweetalert-error')
                return redirect('finance:budget_detail', pk=pk)

    return redirect('finance:budget_detail', pk=pk)


@login_required
def budget_activate(request, pk):
    budget = get_object_or_404(Budget, pk=pk)

    if request.method == 'POST':
        is_htmx = request.headers.get('HX-Request') == 'true'

        if budget.status != 'APPROVED':
            if is_htmx:
                r = HttpResponse()
                r['HX-Alert-Message'] = "Only approved budgets can be activated"
                r['HX-Alert-Type']    = 'warning'
                r['HX-Close-Modal']   = 'true'
                return r
            messages.warning(request, "Only approved budgets can be activated", extra_tags='sweetalert')
            return redirect('finance:budget_detail', pk=pk)

        try:
            BudgetService.activate_budget(budget, user=request.user)
            if is_htmx:
                r = HttpResponse()
                r['HX-Alert-Message'] = f"Budget '{budget.name}' activated successfully"
                r['HX-Alert-Type']    = 'success'
                r['HX-Close-Modal']   = 'true'
                r['HX-Redirect']      = reverse('finance:budget_detail', kwargs={'pk': pk})
                return r
            messages.success(request, "Budget activated successfully", extra_tags='sweetalert')
            return redirect('finance:budget_detail', pk=pk)
        except Exception as e:
            logger.error(f"Error activating budget: {e}")
            if is_htmx:
                r = HttpResponse()
                r['HX-Alert-Message'] = f"Error: {str(e)}"
                r['HX-Alert-Type']    = 'error'
                r['HX-Close-Modal']   = 'true'
                return r
            messages.error(request, f"Error: {str(e)}", extra_tags='sweetalert-error')
            return redirect('finance:budget_detail', pk=pk)


@login_required
def budget_close(request, pk):
    budget = get_object_or_404(Budget, pk=pk)

    if request.method == 'POST':
        is_htmx = request.headers.get('HX-Request') == 'true'

        if budget.status != 'ACTIVE':
            if is_htmx:
                r = HttpResponse()
                r['HX-Alert-Message'] = "Only active budgets can be closed"
                r['HX-Alert-Type']    = 'warning'
                r['HX-Close-Modal']   = 'true'
                return r
            messages.warning(request, "Only active budgets can be closed", extra_tags='sweetalert')
            return redirect('finance:budget_detail', pk=pk)

        try:
            BudgetService.close_budget(budget, user=request.user)
            if is_htmx:
                r = HttpResponse()
                r['HX-Alert-Message'] = f"Budget '{budget.name}' closed successfully"
                r['HX-Alert-Type']    = 'success'
                r['HX-Close-Modal']   = 'true'
                r['HX-Redirect']      = reverse('finance:budget_detail', kwargs={'pk': pk})
                return r
            messages.success(request, "Budget closed successfully", extra_tags='sweetalert')
            return redirect('finance:budget_detail', pk=pk)
        except Exception as e:
            logger.error(f"Error closing budget: {e}")
            if is_htmx:
                r = HttpResponse()
                r['HX-Alert-Message'] = f"Error: {str(e)}"
                r['HX-Alert-Type']    = 'error'
                r['HX-Close-Modal']   = 'true'
                return r
            messages.error(request, f"Error: {str(e)}", extra_tags='sweetalert-error')
            return redirect('finance:budget_detail', pk=pk)


@login_required
def budget_print_view(request, pk):
    budget = get_object_or_404(
        Budget.objects.select_related(
            'fiscal_year', 'academic_session'
        ).prefetch_related('lines__account'),
        pk=pk,
    )
    return render(request, 'finance/budgets/print.html', {
        **get_print_school_context(request),
        'budget': budget,
        'now':    timezone.now(),
        'title':  f'Budget - {budget.name}',
    })


@login_required
def budget_list_print_view(request):
    FIELD_NAMES_FULL = {
        'name':                 'Budget Name',
        'budget_type':          'Budget Type',
        'fiscal_year':          'Fiscal Year',
        'start_date':           'Start Date',
        'end_date':             'End Date',
        'total_revenue_budget': 'Revenue Budget',
        'total_expense_budget': 'Expense Budget',
        'net_budget':           'Net Budget',
        'actual_revenue_total': 'Actual Revenue',
        'actual_expense_total': 'Actual Expenses',
        'status':               'Status',
        'line_count':           'No. of Lines',
    }
    FIELD_NAMES_SHORT = {
        'name':                 'Name',
        'budget_type':          'Type',
        'fiscal_year':          'Year',
        'start_date':           'Start',
        'end_date':             'End',
        'total_revenue_budget': 'Rev. Budget',
        'total_expense_budget': 'Exp. Budget',
        'net_budget':           'Net',
        'actual_revenue_total': 'Act. Revenue',
        'actual_expense_total': 'Act. Expenses',
        'status':               'Status',
        'line_count':           '# Lines',
    }
    DEFAULT_FIELDS  = ['name', 'budget_type', 'fiscal_year', 'start_date', 'end_date', 'total_revenue_budget', 'total_expense_budget', 'net_budget', 'status']
    selected_fields = request.GET.getlist('fields') or DEFAULT_FIELDS
    short_headers   = request.GET.get('short_headers', 'false').lower() == 'true'
    landscape       = request.GET.get('landscape', 'true').lower() == 'true'
    include_stats   = request.GET.get('include_stats', 'true').lower() == 'true'
    field_names     = FIELD_NAMES_SHORT if short_headers else FIELD_NAMES_FULL
    budgets         = get_filtered_budgets(request)

    stats = None
    if include_stats:
        stats = {
            'total':                budgets.count(),
            'draft':                budgets.filter(status='DRAFT').count(),
            'approved':             budgets.filter(status='APPROVED').count(),
            'active':               budgets.filter(status='ACTIVE').count(),
            'closed':               budgets.filter(status='CLOSED').count(),
            'total_revenue_budget': budgets.aggregate(Sum('total_revenue_budget'))['total_revenue_budget__sum'] or Decimal('0.00'),
            'total_expense_budget': budgets.aggregate(Sum('total_expense_budget'))['total_expense_budget__sum'] or Decimal('0.00'),
        }

    if budgets.count() > MAX_PRINT_RECORDS:
        budgets = budgets[:MAX_PRINT_RECORDS]

    return render(request, 'finance/budgets/print_list.html', {
        **get_print_school_context(request),
        'budgets':              budgets,
        'stats':                stats,
        'selected_fields':      selected_fields,
        'selected_field_names': [field_names.get(f, f.replace('_', ' ').title()) for f in selected_fields],
        'field_names':          field_names,
        'short_headers':        short_headers,
        'landscape':            landscape,
        'now':                  timezone.now(),
        'print_date':           get_school_today(),
        'printed_by':           request.user.get_full_name() or request.user.username,
        'title':                'Budgets',
    })


@login_required
def export_budgets_excel(request):
    ALL_COLUMNS = [
        ('name',                 'Budget Name',      lambda o: o.name),
        ('budget_type',          'Type',             lambda o: o.get_budget_type_display()),
        ('start_date',           'Start Date',       lambda o: o.start_date.strftime('%Y-%m-%d')),
        ('end_date',             'End Date',         lambda o: o.end_date.strftime('%Y-%m-%d')),
        ('total_revenue_budget', 'Revenue Budget',   lambda o: float(o.total_revenue_budget)),
        ('total_expense_budget', 'Expense Budget',   lambda o: float(o.total_expense_budget)),
        ('net_budget',           'Net Budget',       lambda o: float(o.net_budget)),
        ('actual_revenue_total', 'Actual Revenue',   lambda o: float(o.actual_revenue_total)),
        ('actual_expense_total', 'Actual Expenses',  lambda o: float(o.actual_expense_total)),
        ('status',               'Status',           lambda o: o.get_status_display()),
        ('fiscal_year',          'Fiscal Year',      lambda o: str(o.fiscal_year) if o.fiscal_year else ''),
        ('description',          'Description',      lambda o: o.description or ''),
    ]
    DEFAULT_FIELDS = [
        'name', 'budget_type', 'start_date', 'end_date',
        'total_revenue_budget', 'total_expense_budget', 'net_budget', 'status',
    ]
    budgets = get_filtered_budgets(request)
    columns = _resolve_columns(ALL_COLUMNS, request.GET.getlist('fields'), DEFAULT_FIELDS)
    return _xlsx_response(_make_workbook('Budgets', columns, budgets), 'budgets')