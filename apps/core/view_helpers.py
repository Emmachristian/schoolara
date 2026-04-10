# core/view_helpers.py

"""
View-layer helper utilities for the core app.

This module contains functions that belong to the view layer — they either
build template context dicts, return HttpResponse objects, or perform
operations that are only meaningful inside a request/response cycle.

WHY SEPARATE FROM core/utils.py
---------------------------------
core/utils.py contains pure business logic with no Django request/response
dependencies. Anything that touches HttpRequest, HttpResponse, or builds
template context lives here instead, keeping utils.py importable from
management commands, signals, and background tasks without side effects.

CONTENTS
--------
- get_print_school_context()   — branding context for print/PDF views
- export_to_csv()              — CSV HttpResponse from a queryset
- export_to_excel()            — XLSX HttpResponse from a queryset
- get_breadcrumb_context()     — breadcrumb list builder for templates
- get_pagination_context()     — pagination metadata for templates
- get_filter_context()         — active filter summary for templates
- get_fiscal_context()         — current fiscal year/period for templates
- get_currency_context()       — currency formatting helpers for templates
- render_to_pdf_response()     — render a template to a PDF HttpResponse
- success_response()           — standardised JSON success response
- error_response()             — standardised JSON error response
- htmx_redirect()              — HTMX-compatible redirect response
"""

import csv
import io
import logging
from decimal import Decimal

from django.http import HttpResponse, JsonResponse
from django.template.loader import render_to_string

logger = logging.getLogger(__name__)


# =============================================================================
# SCHOOL BRANDING CONTEXT
# =============================================================================

def get_print_school_context(request):
    """
    Get school branding context for print and PDF views.

    Tries to derive the school from the logged-in user's profile first,
    then falls back to the first active subscription school.

    Used by all print templates to render the school name, logo, address,
    and contact details in the header.

    Args:
        request: HttpRequest object

    Returns:
        dict with keys:
            school_name     (str)
            school_logo_url (str | None)
            school_address  (str)
            school_contact  (str)

    Example:
        >>> from core.view_helpers import get_print_school_context
        >>>
        >>> def invoice_print(request, pk):
        >>>     invoice = get_object_or_404(FeeInvoice, pk=pk)
        >>>     context = {
        >>>         'invoice': invoice,
        >>>         **get_print_school_context(request),
        >>>     }
        >>>     return render(request, 'fees/invoice_print.html', context)
    """
    school = getattr(getattr(request.user, 'profile', None), 'school', None)

    if not school:
        try:
            from accounts.models import School
            school = School.objects.filter(is_active_subscription=True).first()
        except Exception as e:
            logger.warning(f"Could not fetch school for print context: {e}")
            school = None

    return {
        'school_name':     school.full_name if school else 'School',
        'school_logo_url': (
            school.school_logo.url
            if school and school.school_logo
            else None
        ),
        'school_address':  school.address       if school else '',
        'school_contact':  school.contact_phone if school else '',
    }


# =============================================================================
# CSV EXPORT
# =============================================================================

def export_to_csv(queryset, fields, filename='export.csv', field_labels=None):
    """
    Export a queryset to a CSV HttpResponse.

    Each row is built by calling getattr(obj, field) for every field name.
    Values are coerced to str before writing.

    Args:
        queryset:     Django queryset to export
        fields:       List of field/attribute names to include as columns
        filename:     Output filename (default: 'export.csv')
        field_labels: Optional dict mapping field name → column header label.
                      Falls back to the field name if not provided.
                      Example: {'invoice_number': 'Invoice #', 'total_amount': 'Total'}

    Returns:
        HttpResponse: Content-Type text/csv with Content-Disposition attachment.

    Example:
        >>> from core.view_helpers import export_to_csv
        >>> from fees.models import FeeInvoice
        >>>
        >>> def export_invoices(request):
        >>>     invoices = FeeInvoice.objects.filter(status='PENDING')
        >>>     return export_to_csv(
        >>>         invoices,
        >>>         fields=['invoice_number', 'student', 'total_amount', 'status'],
        >>>         filename='pending_invoices.csv',
        >>>         field_labels={
        >>>             'invoice_number': 'Invoice #',
        >>>             'total_amount':   'Total Amount',
        >>>         },
        >>>     )
    """
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    writer = csv.writer(response)

    # Write header row
    labels = field_labels or {}
    writer.writerow([labels.get(f, f.replace('_', ' ').title()) for f in fields])

    # Write data rows
    for obj in queryset:
        row = []
        for field in fields:
            value = getattr(obj, field, '')
            # Call the value if it is a callable (e.g. a method like get_full_name)
            if callable(value):
                try:
                    value = value()
                except Exception:
                    value = ''
            row.append(str(value) if value is not None else '')
        writer.writerow(row)

    return response


# =============================================================================
# EXCEL EXPORT
# =============================================================================

def export_to_excel(queryset, fields, filename='export.xlsx', field_labels=None,
                    sheet_name='Export'):
    """
    Export a queryset to an XLSX HttpResponse using openpyxl.

    Falls back to CSV export if openpyxl is not installed, logging a warning.

    Args:
        queryset:     Django queryset to export
        fields:       List of field/attribute names to include as columns
        filename:     Output filename (default: 'export.xlsx')
        field_labels: Optional dict mapping field name → column header label
        sheet_name:   Name of the Excel worksheet (default: 'Export')

    Returns:
        HttpResponse: XLSX file or CSV fallback if openpyxl unavailable.

    Example:
        >>> from core.view_helpers import export_to_excel
        >>>
        >>> def export_payments(request):
        >>>     payments = Payment.objects.filter(status='COMPLETED')
        >>>     return export_to_excel(
        >>>         payments,
        >>>         fields=['payment_number', 'student', 'amount', 'payment_date'],
        >>>         filename='payments.xlsx',
        >>>     )
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        logger.warning(
            "openpyxl is not installed. Falling back to CSV export. "
            "Install with: pip install openpyxl"
        )
        csv_filename = filename.replace('.xlsx', '.csv')
        return export_to_csv(queryset, fields, csv_filename, field_labels)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name

    # Header style
    header_font    = Font(bold=True, color='FFFFFF')
    header_fill    = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_align   = Alignment(horizontal='center', vertical='center')

    labels = field_labels or {}
    headers = [labels.get(f, f.replace('_', ' ').title()) for f in fields]

    for col_idx, header in enumerate(headers, start=1):
        cell              = ws.cell(row=1, column=col_idx, value=header)
        cell.font         = header_font
        cell.fill         = header_fill
        cell.alignment    = header_align

    # Data rows
    for row_idx, obj in enumerate(queryset, start=2):
        for col_idx, field in enumerate(fields, start=1):
            value = getattr(obj, field, '')
            if callable(value):
                try:
                    value = value()
                except Exception:
                    value = ''
            # Keep numeric types as numbers for Excel formula compatibility
            if isinstance(value, Decimal):
                value = float(value)
            elif value is None:
                value = ''
            ws.cell(row=row_idx, column=col_idx, value=value)

    # Auto-size columns (approximate)
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                cell_len = len(str(cell.value))
                if cell_len > max_length:
                    max_length = cell_len
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_length + 4, 50)

    # Write to response
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output.getvalue(),
        content_type=(
            'application/vnd.openxmlformats-officedocument'
            '.spreadsheetml.sheet'
        ),
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


# =============================================================================
# BREADCRUMB CONTEXT
# =============================================================================

def get_breadcrumb_context(*crumbs):
    """
    Build a breadcrumb list for template rendering.

    Each crumb is a (label, url) tuple. The last crumb is treated as the
    active page and should have url=None.

    Args:
        *crumbs: Variable number of (label, url) tuples.
                 url should be None for the current active page.

    Returns:
        dict: {'breadcrumbs': [{'label': str, 'url': str|None, 'active': bool}]}

    Example:
        >>> from core.view_helpers import get_breadcrumb_context
        >>> from django.urls import reverse
        >>>
        >>> context = {
        >>>     'invoice': invoice,
        >>>     **get_breadcrumb_context(
        >>>         ('Home',     reverse('dashboard')),
        >>>         ('Invoices', reverse('fees:invoice_list')),
        >>>         ('INV-2025-00042', None),
        >>>     ),
        >>> }
    """
    breadcrumbs = []
    total = len(crumbs)

    for idx, crumb in enumerate(crumbs):
        label, url   = crumb
        is_last      = (idx == total - 1)
        breadcrumbs.append({
            'label':  label,
            'url':    url,
            'active': is_last,
        })

    return {'breadcrumbs': breadcrumbs}


# =============================================================================
# PAGINATION CONTEXT
# =============================================================================

def get_pagination_context(page_obj, paginator, request=None):
    """
    Build pagination metadata for templates.

    Provides page range, neighbour pages, and GET parameter preservation
    so filters are not lost when navigating pages.

    Args:
        page_obj:  Django Page object (from Paginator.page())
        paginator: Django Paginator object
        request:   Optional HttpRequest. When provided, GET params other
                   than 'page' are preserved in pagination links.

    Returns:
        dict: Pagination metadata for use in templates.

    Example:
        >>> from core.utils import paginate_queryset
        >>> from core.view_helpers import get_pagination_context
        >>>
        >>> page_obj, paginator = paginate_queryset(request, invoices)
        >>> context = {
        >>>     'page_obj':  page_obj,
        >>>     'paginator': paginator,
        >>>     **get_pagination_context(page_obj, paginator, request),
        >>> }
    """
    # Build query string without 'page' so filters are preserved
    query_string = ''
    if request:
        params = request.GET.copy()
        params.pop('page', None)
        qs = params.urlencode()
        query_string = f'&{qs}' if qs else ''

    # Visible page range: always show first, last, and neighbours around current
    current   = page_obj.number
    num_pages = paginator.num_pages

    # Build a compact page range with ellipsis markers
    if num_pages <= 7:
        page_range = list(range(1, num_pages + 1))
    else:
        neighbours = {current - 1, current, current + 1}
        edges      = {1, 2, num_pages - 1, num_pages}
        visible    = sorted(neighbours | edges)

        page_range = []
        prev       = None
        for p in visible:
            if p < 1 or p > num_pages:
                continue
            if prev and p - prev > 1:
                page_range.append('...')
            page_range.append(p)
            prev = p

    return {
        'page_range':    page_range,
        'query_string':  query_string,
        'showing_start': page_obj.start_index(),
        'showing_end':   page_obj.end_index(),
        'total_count':   paginator.count,
    }


# =============================================================================
# FILTER SUMMARY CONTEXT
# =============================================================================

def get_filter_context(request, filter_keys, label_map=None):
    """
    Build a list of active filters for display in templates.

    Reads GET parameters and returns only those that have values,
    with human-readable labels. Used to render "Active filters: ..." badges.

    Args:
        request:    HttpRequest object
        filter_keys: List of GET parameter names to check
        label_map:  Optional dict mapping parameter name → display label.
                    Defaults to title-cased parameter name.

    Returns:
        dict: {
            'active_filters': [{'key': str, 'label': str, 'value': str}],
            'has_active_filters': bool,
            'clear_url': str,   # current path with no GET params
        }

    Example:
        >>> from core.view_helpers import get_filter_context
        >>>
        >>> context = {
        >>>     'invoices': page_obj,
        >>>     **get_filter_context(
        >>>         request,
        >>>         filter_keys=['status', 'student', 'date_from', 'date_to'],
        >>>         label_map={
        >>>             'date_from': 'From Date',
        >>>             'date_to':   'To Date',
        >>>         },
        >>>     ),
        >>> }
    """
    labels         = label_map or {}
    active_filters = []

    for key in filter_keys:
        value = request.GET.get(key, '').strip()
        if value:
            active_filters.append({
                'key':   key,
                'label': labels.get(key, key.replace('_', ' ').title()),
                'value': value,
            })

    return {
        'active_filters':     active_filters,
        'has_active_filters': bool(active_filters),
        'clear_url':          request.path,
    }


# =============================================================================
# FISCAL CONTEXT FOR TEMPLATES
# =============================================================================

def get_fiscal_context():
    """
    Build current fiscal year and period context for templates.

    Provides the active fiscal year, active fiscal period, and current
    academic session for use in base templates and dashboards.

    Uses school timezone via core.utils helpers.

    Returns:
        dict with keys:
            active_fiscal_year    (FiscalYear | None)
            active_fiscal_period  (FiscalPeriod | None)
            active_academic_session (AcademicSession | None)
            school_today          (date)

    Example:
        >>> from core.view_helpers import get_fiscal_context
        >>>
        >>> def dashboard(request):
        >>>     context = {
        >>>         'stats': get_dashboard_stats(),
        >>>         **get_fiscal_context(),
        >>>     }
        >>>     return render(request, 'dashboard.html', context)
    """
    from core.utils import (
        get_active_fiscal_year,
        get_active_fiscal_period,
        get_active_academic_session,
        get_school_today,
    )

    return {
        'active_fiscal_year':      get_active_fiscal_year(),
        'active_fiscal_period':    get_active_fiscal_period(),
        'active_academic_session': get_active_academic_session(),
        'school_today':            get_school_today(),
    }


# =============================================================================
# CURRENCY CONTEXT FOR TEMPLATES
# =============================================================================

def get_currency_context():
    """
    Build currency formatting context for templates.

    Provides the school currency code and a format_money callable so
    templates can format amounts without template tags.

    Returns:
        dict with keys:
            school_currency  (str)           ISO 4217 code
            currency_info    (dict)          from FinancialSettings.get_currency_info()
            format_money     (callable)      format_money(amount, include_symbol=True)

    Example:
        >>> from core.view_helpers import get_currency_context
        >>>
        >>> def payment_list(request):
        >>>     context = {
        >>>         'payments': payments,
        >>>         **get_currency_context(),
        >>>     }
        >>>     return render(request, 'payments/list.html', context)
        >>>
        >>> # In the template:
        >>> # {{ format_money(payment.amount) }}
    """
    from core.utils import get_school_currency, format_money
    from core.models import FinancialSettings

    return {
        'school_currency': get_school_currency(),
        'currency_info':   FinancialSettings.get_currency_info(),
        'format_money':    format_money,
    }


# =============================================================================
# PDF RESPONSE
# =============================================================================

def render_to_pdf_response(template_name, context, filename='document.pdf',
                           request=None):
    """
    Render a Django template to a PDF HttpResponse.

    Requires WeasyPrint to be installed:
        pip install weasyprint

    Falls back to an HTML response with a warning header if WeasyPrint
    is not available, so development works without the dependency.

    Args:
        template_name: Django template path (e.g., 'fees/invoice_print.html')
        context:       Template context dict
        filename:      Output PDF filename (default: 'document.pdf')
        request:       Optional HttpRequest (passed to render_to_string
                       for request-aware template rendering)

    Returns:
        HttpResponse: PDF file (Content-Type: application/pdf)
                      or HTML fallback if WeasyPrint unavailable.

    Example:
        >>> from core.view_helpers import render_to_pdf_response, get_print_school_context
        >>>
        >>> def invoice_pdf(request, pk):
        >>>     invoice = get_object_or_404(FeeInvoice, pk=pk)
        >>>     context = {
        >>>         'invoice': invoice,
        >>>         **get_print_school_context(request),
        >>>     }
        >>>     return render_to_pdf_response(
        >>>         'fees/invoice_print.html',
        >>>         context,
        >>>         filename=f'{invoice.invoice_number}.pdf',
        >>>         request=request,
        >>>     )
    """
    html_string = render_to_string(template_name, context, request=request)

    try:
        import weasyprint
        pdf_bytes = weasyprint.HTML(string=html_string).write_pdf()
        response  = HttpResponse(pdf_bytes, content_type='application/pdf')
        response['Content-Disposition'] = (
            f'attachment; filename="{filename}"'
        )
        return response

    except ImportError:
        logger.warning(
            "WeasyPrint is not installed. Returning HTML instead of PDF. "
            "Install with: pip install weasyprint"
        )
        response = HttpResponse(html_string, content_type='text/html')
        response['X-PDF-Fallback'] = (
            'WeasyPrint not installed — HTML rendered instead of PDF'
        )
        return response

    except Exception as e:
        logger.error(f"PDF generation failed for {template_name}: {e}", exc_info=True)
        response = HttpResponse(
            f"<p>PDF generation failed: {e}</p>",
            content_type='text/html',
            status=500,
        )
        return response


# =============================================================================
# STANDARDISED JSON RESPONSES
# =============================================================================

def success_response(message='', data=None, status=200, **kwargs):
    """
    Return a standardised JSON success response.

    Used by AJAX and HTMX endpoints to return consistent payloads.

    Args:
        message: Human-readable success message
        data:    Optional dict of additional data to include
        status:  HTTP status code (default: 200)
        **kwargs: Additional top-level keys to merge into the response

    Returns:
        JsonResponse with structure:
            {'success': True, 'message': str, ...data, ...kwargs}

    Example:
        >>> from core.view_helpers import success_response
        >>>
        >>> def mark_payment_verified(request, pk):
        >>>     payment = get_object_or_404(Payment, pk=pk)
        >>>     payment.is_verified = True
        >>>     payment.save()
        >>>     return success_response(
        >>>         message='Payment verified successfully.',
        >>>         data={'payment_number': payment.payment_number},
        >>>     )
    """
    payload = {'success': True, 'message': message}
    if data:
        payload.update(data)
    if kwargs:
        payload.update(kwargs)
    return JsonResponse(payload, status=status)


def error_response(message='An error occurred.', errors=None,
                   status=400, **kwargs):
    """
    Return a standardised JSON error response.

    Used by AJAX and HTMX endpoints to return consistent error payloads.

    Args:
        message: Human-readable error message
        errors:  Optional dict of field-level errors
                 (mirrors Django form.errors structure)
        status:  HTTP status code (default: 400)
        **kwargs: Additional top-level keys to merge into the response

    Returns:
        JsonResponse with structure:
            {'success': False, 'message': str, 'errors': dict|None, ...kwargs}

    Example:
        >>> from core.view_helpers import error_response
        >>>
        >>> def create_invoice(request):
        >>>     form = FeeInvoiceForm(request.POST)
        >>>     if not form.is_valid():
        >>>         return error_response(
        >>>             message='Please correct the errors below.',
        >>>             errors=form.errors,
        >>>             status=422,
        >>>         )
    """
    payload = {'success': False, 'message': message, 'errors': errors or {}}
    if kwargs:
        payload.update(kwargs)
    return JsonResponse(payload, status=status)


# =============================================================================
# HTMX REDIRECT HELPER
# =============================================================================

def htmx_redirect(url, status=200):
    """
    Return an HTMX-compatible redirect response.

    HTMX cannot follow a standard 302 redirect from a fetch request.
    This helper returns a 200 response with the HX-Redirect header,
    which instructs the HTMX client to perform a full-page navigation.

    Use this instead of Django's redirect() inside HTMX-triggered views
    when you want to redirect after a successful form submission.

    Args:
        url:    URL to redirect to (string)
        status: HTTP status code for the response body (default: 200)

    Returns:
        HttpResponse: Empty body with HX-Redirect header set.

    Example:
        >>> from core.view_helpers import htmx_redirect
        >>> from django.urls import reverse
        >>>
        >>> def create_fiscal_year(request):
        >>>     form = FiscalYearForm(request.POST)
        >>>     if form.is_valid():
        >>>         fiscal_year = form.save()
        >>>         if request.headers.get('HX-Request'):
        >>>             return htmx_redirect(
        >>>                 reverse('core:fiscal_year_detail',
        >>>                         kwargs={'pk': fiscal_year.pk})
        >>>             )
        >>>         return redirect('core:fiscal_year_detail', pk=fiscal_year.pk)
        >>>     return render(request, 'core/fiscal_years/form.html', {'form': form})
    """
    response              = HttpResponse(status=status)
    response['HX-Redirect'] = url
    return response