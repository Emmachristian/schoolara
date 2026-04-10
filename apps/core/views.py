# core/views.py

"""
Core Configuration Views

What lives here
---------------
- Dashboard
- School Configuration (edit)
- Financial Settings (edit + account mappings)
- Fiscal Management combined accordion view
- Fiscal Year  : detail + print + Excel export
- Fiscal Period: detail + print + Excel export
- Payment Method : full CRUD + print + Excel export
- Tax Rate       : full CRUD + print + Excel export
- Unit of Measure: full CRUD + print + Excel export
- JSON quick-stats endpoints

PATTERNS (learned from students/views.py)
-----------------------------------------
Print views:
  - field_names_full / field_names_short dicts
  - short_headers GET param toggles between them
  - include_stats and landscape GET params
  - school context (name, logo, address, contact) from get_print_school_context()

Excel exports:
  - ALL_COLUMNS list of (param_name, header_label, callable) tuples
  - COLUMN_MAP dict for O(1) lookup
  - DEFAULT_FIELDS fallback
  - User-chosen column order preserved
  - openpyxl with header styling, auto-sized columns capped at 60
  - Both print and export reuse the same _get_filtered_* queryset helper
    so active filters are always respected
"""

import io
from datetime import timedelta

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.core.exceptions import ValidationError
from django.db.models import Count, Q, Prefetch
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.views.decorators.http import require_http_methods
import logging

from .models import (
    FinancialSettings,
    FiscalPeriod,
    FiscalYear,
    PaymentMethod,
    SchoolConfiguration,
    TaxRate,
    UnitOfMeasure,
    CoreAccountMappings,
    RevenueAccountMappings,
    PayrollAccountMappings,
    ExpenseAccountMappings,
    SpecialAccountMappings,
)
from .forms import (
    CoreAccountMappingsForm,
    ExpenseAccountMappingsForm,
    FinancialSettingsForm,
    FiscalPeriodFilterForm,
    FiscalPeriodForm,
    FiscalYearFilterForm,
    FiscalYearForm,
    PaymentMethodFilterForm,
    PaymentMethodForm,
    PayrollAccountMappingsForm,
    RevenueAccountMappingsForm,
    SchoolConfigurationForm,
    SpecialAccountMappingsForm,
    TaxRateFilterForm,
    TaxRateForm,
    UnitOfMeasureFilterForm,
    UnitOfMeasureForm,
)
from .utils import (
    get_school_current_time,
    get_school_today,
    paginate_queryset,
)
from .view_helpers import (
    get_print_school_context,
    htmx_redirect,
)
from . import stats as core_stats

logger = logging.getLogger(__name__)


# =============================================================================
# EXCEL HELPER
# =============================================================================

def _build_excel_response(ws, wb, filename):
    """
    Save workbook to an HttpResponse with the correct content-type and
    Content-Disposition header.
    """
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    response = HttpResponse(
        output.getvalue(),
        content_type=(
            'application/vnd.openxmlformats-officedocument'
            '.spreadsheetml.sheet'
        ),
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


def _style_excel_header(ws, header_color='1E3A5F'):
    """
    Apply bold white text on a dark background to row 1.
    Set row height to 30.
    """
    try:
        from openpyxl.styles import Alignment, Font, PatternFill

        header_fill  = PatternFill(
            start_color=header_color, end_color=header_color, fill_type='solid'
        )
        header_font  = Font(bold=True, color='FFFFFF', size=11)
        header_align = Alignment(
            horizontal='center', vertical='center', wrap_text=True
        )

        for cell in ws[1]:
            cell.font      = header_font
            cell.fill      = header_fill
            cell.alignment = header_align

        ws.row_dimensions[1].height = 30
    except ImportError:
        pass  # openpyxl not installed — caller will handle fallback


def _autosize_columns(ws, max_width=60):
    """Auto-size columns based on content, capped at max_width."""
    try:
        from openpyxl.styles import Alignment

        data_align = Alignment(vertical='center', wrap_text=False)

        for col_cells in ws.columns:
            max_len = max(
                (len(str(c.value)) if c.value is not None else 0)
                for c in col_cells
            )
            ws.column_dimensions[col_cells[0].column_letter].width = min(
                max_len + 4, max_width
            )

        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = data_align
    except ImportError:
        pass


# =============================================================================
# DASHBOARD
# =============================================================================

@login_required
def core_dashboard(request):
    """Main core system dashboard."""

    try:
        system_stats  = core_stats.get_core_system_statistics()
        completeness  = core_stats.get_configuration_completeness()
        fiscal_info   = core_stats.get_current_fiscal_info()
        payment_stats = core_stats.get_payment_statistics_summary()
    except Exception as e:
        logger.error(f"Dashboard statistics error: {e}")
        system_stats  = {}
        completeness  = {'percentage': 0, 'is_complete': False}
        fiscal_info   = {}
        payment_stats = {}

    today = get_school_today()

    return render(request, 'core/home.html', {
        'system_stats':  system_stats,
        'completeness':  completeness,
        'fiscal_info':   fiscal_info,
        'payment_stats': payment_stats,
        'fiscal_years_ending': FiscalYear.objects.filter(
            end_date__gte=today,
            end_date__lte=today + timedelta(days=30),
            is_active=True,
        ).order_by('end_date')[:5],
        'periods_needing_closure': FiscalPeriod.objects.filter(
            end_date__lt=today,
            is_closed=False,
            is_active=True,
        ).order_by('end_date')[:5],
        'tax_rates_expiring': TaxRate.objects.filter(
            effective_to__gte=today,
            effective_to__lte=today + timedelta(days=60),
            is_active=True,
        ).order_by('effective_to')[:5],
        'inactive_payment_methods': PaymentMethod.objects.filter(
            is_active=False
        ).count(),
    })


# =============================================================================
# SCHOOL CONFIGURATION
# =============================================================================

@login_required
def school_configuration_edit(request):
    config = SchoolConfiguration.get_cached_instance()

    if request.method == 'POST':
        form = SchoolConfigurationForm(request.POST, instance=config)
        if form.is_valid():
            try:
                config = form.save()
                SchoolConfiguration.clear_cache()
                messages.success(
                    request,
                    'School configuration updated successfully!',
                    extra_tags='sweetalert',   # ← tells the template to use SweetAlert2
                )
                return redirect('core:school_configuration_edit')
            except Exception as e:
                logger.error(f"Configuration update error: {e}")
                messages.error(
                    request,
                    f'Error saving configuration: {str(e)}',
                    extra_tags='sweetalert',
                )
    else:
        form = SchoolConfigurationForm(instance=config)

    return render(request, 'core/configuration/school_config.html', {
        'form':        form,
        'config':      config,
        'title':       'Edit School Configuration',
        'submit_text': 'Save Configuration',
    })


# =============================================================================
# FINANCIAL SETTINGS
# =============================================================================

@login_required
def financial_settings_edit(request):
    settings = FinancialSettings.get_instance()

    if request.method == 'POST':
        form = FinancialSettingsForm(request.POST, instance=settings)
        if form.is_valid():
            try:
                settings = form.save()
                messages.success(
                    request,
                    'Financial settings updated successfully!',
                    extra_tags='sweetalert',
                )
                return redirect('core:financial_settings_edit')
            except Exception as e:
                logger.error(f"Financial settings error: {e}")
                messages.error(
                    request,
                    f'Error saving financial settings: {str(e)}',
                    extra_tags='sweetalert',
                )
    else:
        form = FinancialSettingsForm(instance=settings)

    return render(request, 'core/configuration/financial_settings.html', {
        'form':        form,
        'settings':    settings,
        'title':       'Financial Settings',
        'submit_text': 'Save Financial Settings',
    })

@login_required
def account_mappings_edit(request, mapping_type):
    """Edit one of the five account mapping categories."""
    settings = FinancialSettings.get_instance()

    mapping_config = {
        'core': {
            'form_class': CoreAccountMappingsForm,
            'get_method': 'get_account_mappings',
            'title':      'Core Account Mappings',
        },
        'revenue': {
            'form_class': RevenueAccountMappingsForm,
            'get_method': 'get_revenue_mappings',
            'title':      'Revenue Account Mappings',
        },
        'payroll': {
            'form_class': PayrollAccountMappingsForm,
            'get_method': 'get_payroll_mappings',
            'title':      'Payroll Account Mappings',
        },
        'expense': {
            'form_class': ExpenseAccountMappingsForm,
            'get_method': 'get_expense_mappings',
            'title':      'Expense Account Mappings',
        },
        'special': {
            'form_class': SpecialAccountMappingsForm,
            'get_method': 'get_special_mappings',
            'title':      'Special Account Mappings',
        },
    }

    if mapping_type not in mapping_config:
        messages.error(request, 'Invalid mapping type.', extra_tags='sweetalert')
        return redirect('core:financial_settings_edit')

    cfg              = mapping_config[mapping_type]
    mapping_instance = getattr(settings, cfg['get_method'])()

    if request.method == 'POST':
        form = cfg['form_class'](request.POST, instance=mapping_instance)
        if form.is_valid():
            try:
                form.save()
                messages.success(
                    request,
                    f'{cfg["title"]} updated successfully!',
                    extra_tags='sweetalert',
                )
                return redirect('core:account_mappings_edit', mapping_type=mapping_type)
            except Exception as e:
                logger.error(f"Mappings update error ({mapping_type}): {e}")
                messages.error(
                    request,
                    f'Error saving mappings: {str(e)}',
                    extra_tags='sweetalert',
                )
    else:
        form = cfg['form_class'](instance=mapping_instance)

    return render(request, 'core/configuration/mappings_form.html', {
        'form':         form,
        'mapping_type': mapping_type,
        'title':        f'Edit {cfg["title"]}',
        'submit_text':  'Save Mappings',
    })


# =============================================================================
# FISCAL MANAGEMENT — combined accordion view
# =============================================================================

@login_required
def fiscal_management_view(request):
    """
    Single page showing all fiscal years with their periods nested inside.
    Stats are computed here directly so the template never depends on
    external stats functions whose key names may differ.
    """
    fiscal_years = FiscalYear.objects.annotate(
        period_count=Count('fiscal_periods', distinct=True),
    ).prefetch_related(
        Prefetch(
            'fiscal_periods',
            queryset=FiscalPeriod.objects.select_related(
                'related_academic_session'
            ).order_by('period_number'),
            to_attr='prefetched_periods',
        )
    ).order_by('-start_date')

    fy_agg = FiscalYear.objects.aggregate(
        total=Count('id'),
        active=Count('id', filter=Q(is_active=True)),
        closed=Count('id', filter=Q(is_closed=True)),
        locked=Count('id', filter=Q(is_locked=True)),
        draft=Count('id',  filter=Q(status='DRAFT')),
    )

    period_agg = FiscalPeriod.objects.aggregate(
        total=Count('id'),
        active=Count('id', filter=Q(is_active=True)),
        closed=Count('id', filter=Q(is_closed=True)),
        locked=Count('id', filter=Q(is_locked=True)),
    )

    return render(request, 'core/fiscal_management.html', {
        'fiscal_years': fiscal_years,
        'fy_stats':     fy_agg,
        'period_stats': period_agg,
    })


# =============================================================================
# FISCAL YEAR — queryset helper
# =============================================================================

def _get_filtered_fiscal_years(request):
    """
    Build a filtered FiscalYear queryset from GET parameters.
    Shared by list, print, and export views so filters are always respected.
    """
    qs = FiscalYear.objects.annotate(
        period_count=Count('fiscal_periods', distinct=True)
    ).order_by('-start_date')

    query  = request.GET.get('q',         '').strip()
    status = request.GET.get('status',    '').strip()
    is_active = request.GET.get('is_active', '').strip()
    is_closed = request.GET.get('is_closed', '').strip()
    start_date_from = request.GET.get('start_date_from', '').strip()
    start_date_to   = request.GET.get('start_date_to',   '').strip()

    if query:
        qs = qs.filter(
            Q(name__icontains=query) | Q(code__icontains=query)
        )
    if status:
        qs = qs.filter(status=status)
    if is_active:
        qs = qs.filter(is_active=(is_active.lower() == 'true'))
    if is_closed:
        qs = qs.filter(is_closed=(is_closed.lower() == 'true'))
    if start_date_from:
        try:
            qs = qs.filter(start_date__gte=start_date_from)
        except Exception:
            pass
    if start_date_to:
        try:
            qs = qs.filter(start_date__lte=start_date_to)
        except Exception:
            pass

    return qs


# =============================================================================
# FISCAL YEAR — detail, print, export
# =============================================================================

@login_required
def fiscal_year_detail(request, pk):
    fiscal_year = get_object_or_404(
        FiscalYear.objects.prefetch_related(
            Prefetch(
                'fiscal_periods',
                queryset=FiscalPeriod.objects.select_related(
                    'related_academic_session'
                ).order_by('period_number'),
                to_attr='prefetched_periods',
            )
        ),
        pk=pk,
    )

    return render(request, 'core/fiscal_years/detail.html', {
        'fiscal_year': fiscal_year,
        'periods':     fiscal_year.prefetched_periods,
        'today':       get_school_today(),
    })


# ── Column definitions shared by print and Excel export ─────────────────────

_FISCAL_YEAR_COLUMNS = [
    ('name',                'Fiscal Year',     lambda fy: fy.name),
    ('code',                'Code',            lambda fy: fy.code),
    ('start_date',          'Start Date',      lambda fy: fy.start_date.strftime('%Y-%m-%d') if fy.start_date else ''),
    ('end_date',            'End Date',        lambda fy: fy.end_date.strftime('%Y-%m-%d')   if fy.end_date   else ''),
    ('status',              'Status',          lambda fy: fy.get_status_display()),
    ('is_active',           'Active',          lambda fy: 'Yes' if fy.is_active else 'No'),
    ('is_closed',           'Closed',          lambda fy: 'Yes' if fy.is_closed else 'No'),
    ('is_locked',           'Locked',          lambda fy: 'Yes' if fy.is_locked else 'No'),
    ('period_count',        'Periods',         lambda fy: getattr(fy, 'period_count', fy.get_period_count())),
    ('duration_days',       'Duration (Days)', lambda fy: fy.get_duration_days()),
    ('progress_percentage', 'Progress %',      lambda fy: fy.get_progress_percentage()),
]
_FISCAL_YEAR_COLUMN_MAP    = {c[0]: c for c in _FISCAL_YEAR_COLUMNS}
_FISCAL_YEAR_DEFAULT_FIELDS = [
    'name', 'code', 'start_date', 'end_date', 'status', 'is_active',
]

_FISCAL_YEAR_FIELD_NAMES_FULL = {c[0]: c[1] for c in _FISCAL_YEAR_COLUMNS}
_FISCAL_YEAR_FIELD_NAMES_SHORT = {
    'name':                'Year',
    'code':                'Code',
    'start_date':          'Start',
    'end_date':            'End',
    'status':              'Status',
    'is_active':           'Active',
    'is_closed':           'Closed',
    'is_locked':           'Locked',
    'period_count':        'Periods',
    'duration_days':       'Days',
    'progress_percentage': 'Progress',
}


@login_required
def fiscal_year_print_view(request):
    """
    Printable fiscal year list.
    Respects active filters. Supports short_headers, include_stats, landscape.
    """
    selected_fields = request.GET.getlist('fields') or _FISCAL_YEAR_DEFAULT_FIELDS
    include_stats   = request.GET.get('include_stats') == 'true'
    landscape_mode  = request.GET.get('landscape')    == 'true'
    short_headers   = request.GET.get('short_headers') == 'true'

    fiscal_years = _get_filtered_fiscal_years(request)

    field_names = (
        _FISCAL_YEAR_FIELD_NAMES_SHORT if short_headers
        else _FISCAL_YEAR_FIELD_NAMES_FULL
    )
    selected_field_names = [
        field_names.get(f, f.replace('_', ' ').title())
        for f in selected_fields
    ]

    context = {
        'fiscal_years':         fiscal_years,
        'now':                  get_school_current_time(),
        'selected_fields':      selected_fields,
        'selected_field_names': selected_field_names,
        'field_names':          field_names,
        'landscape':            landscape_mode,
        'short_headers':        short_headers,
        'title':                'Fiscal Years Report',
        **get_print_school_context(request),
    }

    if include_stats:
        context['stats'] = {
            'total':  fiscal_years.count(),
            'active': fiscal_years.filter(is_active=True).count(),
            'closed': fiscal_years.filter(is_closed=True).count(),
            'locked': fiscal_years.filter(is_locked=True).count(),
        }

    return render(request, 'core/fiscal_years/print.html', context)


@login_required
def fiscal_year_export_excel(request):
    """
    Export filtered fiscal years to Excel.
    Respects active filters AND selected columns.
    """
    try:
        import openpyxl
    except ImportError:
        messages.error(
            request,
            'Excel export requires openpyxl. Install with: pip install openpyxl'
        )
        return redirect('core:fiscal_management')

    fiscal_years    = _get_filtered_fiscal_years(request)
    selected        = request.GET.getlist('fields') or _FISCAL_YEAR_DEFAULT_FIELDS
    columns         = [
        _FISCAL_YEAR_COLUMN_MAP[f]
        for f in selected
        if f in _FISCAL_YEAR_COLUMN_MAP
    ] or [_FISCAL_YEAR_COLUMN_MAP[f] for f in _FISCAL_YEAR_DEFAULT_FIELDS]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Fiscal Years'

    ws.append([col[1] for col in columns])
    _style_excel_header(ws)

    for fy in fiscal_years:
        ws.append([col[2](fy) for col in columns])

    _autosize_columns(ws)

    filename = f"fiscal_years_{get_school_today().strftime('%Y%m%d')}.xlsx"
    return _build_excel_response(ws, wb, filename)


# =============================================================================
# FISCAL PERIOD — queryset helper
# =============================================================================

def _get_filtered_fiscal_periods(request):
    """
    Build a filtered FiscalPeriod queryset from GET parameters.
    Shared by list, print, and export views.
    """
    qs = FiscalPeriod.objects.select_related(
        'fiscal_year', 'related_academic_session'
    ).order_by('-fiscal_year__start_date', 'period_number')

    query          = request.GET.get('q',           '').strip()
    fiscal_year_id = request.GET.get('fiscal_year', '').strip()
    period_type    = request.GET.get('period_type', '').strip()
    status         = request.GET.get('status',      '').strip()
    is_active      = request.GET.get('is_active',   '').strip()
    is_closed      = request.GET.get('is_closed',   '').strip()
    start_date_from= request.GET.get('start_date_from', '').strip()
    start_date_to  = request.GET.get('start_date_to',   '').strip()

    if query:
        qs = qs.filter(
            Q(name__icontains=query) | Q(code__icontains=query)
        )
    if fiscal_year_id:
        qs = qs.filter(fiscal_year_id=fiscal_year_id)
    if period_type:
        qs = qs.filter(period_type=period_type)
    if status:
        qs = qs.filter(status=status)
    if is_active:
        qs = qs.filter(is_active=(is_active.lower() == 'true'))
    if is_closed:
        qs = qs.filter(is_closed=(is_closed.lower() == 'true'))
    if start_date_from:
        try:
            qs = qs.filter(start_date__gte=start_date_from)
        except Exception:
            pass
    if start_date_to:
        try:
            qs = qs.filter(start_date__lte=start_date_to)
        except Exception:
            pass

    return qs


# =============================================================================
# FISCAL PERIOD — detail, print, export
# =============================================================================

@login_required
def fiscal_period_detail(request, pk):
    period = get_object_or_404(
        FiscalPeriod.objects.select_related(
            'fiscal_year', 'related_academic_session'
        ),
        pk=pk,
    )

    return render(request, 'core/fiscal_periods/detail.html', {
        'period': period,
        'progress_info': {
            'days_elapsed':            period.get_elapsed_days(),
            'days_remaining':          period.get_remaining_days(),
            'total_days':              period.get_duration_days(),
            'progress_percentage':     period.get_progress_percentage(),
            'is_current':              period.is_current(),
            'is_upcoming':             period.is_upcoming(),
            'is_past':                 period.is_past(),
            'is_in_grace_period':      period.is_in_grace_period(),
            'can_accept_transactions': period.can_accept_transactions(),
        },
        'today': get_school_today(),
    })


_FISCAL_PERIOD_COLUMNS = [
    ('name',                'Period Name',     lambda p: p.name),
    ('code',                'Code',            lambda p: p.code),
    ('fiscal_year',         'Fiscal Year',     lambda p: p.fiscal_year.name),
    ('period_number',       '#',               lambda p: float(p.period_number)),
    ('period_type',         'Type',            lambda p: p.get_period_type_display()),
    ('start_date',          'Start Date',      lambda p: p.start_date.strftime('%Y-%m-%d') if p.start_date else ''),
    ('end_date',            'End Date',        lambda p: p.end_date.strftime('%Y-%m-%d')   if p.end_date   else ''),
    ('status',              'Status',          lambda p: p.get_status_display()),
    ('is_active',           'Active',          lambda p: 'Yes' if p.is_active else 'No'),
    ('is_closed',           'Closed',          lambda p: 'Yes' if p.is_closed else 'No'),
    ('is_locked',           'Locked',          lambda p: 'Yes' if p.is_locked else 'No'),
    ('duration_days',       'Duration (Days)', lambda p: p.get_duration_days()),
    ('duration_weeks',      'Weeks',           lambda p: p.get_duration_weeks()),
    ('progress_percentage', 'Progress %',      lambda p: p.get_progress_percentage()),
    ('grace_period_days',   'Grace Days',      lambda p: p.grace_period_days),
]
_FISCAL_PERIOD_COLUMN_MAP    = {c[0]: c for c in _FISCAL_PERIOD_COLUMNS}
_FISCAL_PERIOD_DEFAULT_FIELDS = [
    'name', 'code', 'period_number', 'period_type',
    'start_date', 'end_date', 'status',
]

_FISCAL_PERIOD_FIELD_NAMES_FULL = {c[0]: c[1] for c in _FISCAL_PERIOD_COLUMNS}
_FISCAL_PERIOD_FIELD_NAMES_SHORT = {
    'name':                'Period',
    'code':                'Code',
    'fiscal_year':         'FY',
    'period_number':       '#',
    'period_type':         'Type',
    'start_date':          'Start',
    'end_date':            'End',
    'status':              'Status',
    'is_active':           'Active',
    'is_closed':           'Closed',
    'is_locked':           'Locked',
    'duration_days':       'Days',
    'duration_weeks':      'Weeks',
    'progress_percentage': 'Progress',
    'grace_period_days':   'Grace',
}


@login_required
def fiscal_period_print_view(request):
    """
    Printable fiscal period list.
    Respects active filters. Supports short_headers, include_stats, landscape.
    """
    selected_fields = request.GET.getlist('fields') or _FISCAL_PERIOD_DEFAULT_FIELDS
    include_stats   = request.GET.get('include_stats')  == 'true'
    landscape_mode  = request.GET.get('landscape')      == 'true'
    short_headers   = request.GET.get('short_headers')  == 'true'

    periods = _get_filtered_fiscal_periods(request)

    field_names = (
        _FISCAL_PERIOD_FIELD_NAMES_SHORT if short_headers
        else _FISCAL_PERIOD_FIELD_NAMES_FULL
    )
    selected_field_names = [
        field_names.get(f, f.replace('_', ' ').title())
        for f in selected_fields
    ]

    context = {
        'periods':              periods,
        'now':                  get_school_current_time(),
        'selected_fields':      selected_fields,
        'selected_field_names': selected_field_names,
        'field_names':          field_names,
        'landscape':            landscape_mode,
        'short_headers':        short_headers,
        'title':                'Fiscal Periods Report',
        **get_print_school_context(request),
    }

    if include_stats:
        context['stats'] = {
            'total':  periods.count(),
            'active': periods.filter(is_active=True).count(),
            'closed': periods.filter(is_closed=True).count(),
            'locked': periods.filter(is_locked=True).count(),
        }

    return render(request, 'core/fiscal_periods/print.html', context)


@login_required
def fiscal_period_export_excel(request):
    """Export filtered fiscal periods to Excel."""
    try:
        import openpyxl
    except ImportError:
        messages.error(request, 'Excel export requires openpyxl.')
        return redirect('core:fiscal_management')

    periods  = _get_filtered_fiscal_periods(request)
    selected = request.GET.getlist('fields') or _FISCAL_PERIOD_DEFAULT_FIELDS
    columns  = [
        _FISCAL_PERIOD_COLUMN_MAP[f]
        for f in selected
        if f in _FISCAL_PERIOD_COLUMN_MAP
    ] or [_FISCAL_PERIOD_COLUMN_MAP[f] for f in _FISCAL_PERIOD_DEFAULT_FIELDS]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Fiscal Periods'

    ws.append([col[1] for col in columns])
    _style_excel_header(ws)

    for p in periods:
        ws.append([col[2](p) for col in columns])

    _autosize_columns(ws)

    filename = f"fiscal_periods_{get_school_today().strftime('%Y%m%d')}.xlsx"
    return _build_excel_response(ws, wb, filename)


# =============================================================================
# PAYMENT METHOD — queryset helper + full CRUD + print + export
# =============================================================================

def _get_filtered_payment_methods(request):
    """
    Build a filtered PaymentMethod queryset from GET parameters.
    Shared by list, print, and export views.
    """
    qs = PaymentMethod.objects.all().order_by('display_order', 'name')

    query               = request.GET.get('q',                    '').strip()
    method_type         = request.GET.get('method_type',          '').strip()
    is_active           = request.GET.get('is_active',            '').strip()
    is_default          = request.GET.get('is_default',           '').strip()
    requires_approval   = request.GET.get('requires_approval',    '').strip()
    has_transaction_fee = request.GET.get('has_transaction_fee',  '').strip()
    mobile_provider     = request.GET.get('mobile_money_provider','').strip()

    if query:
        qs = qs.filter(
            Q(name__icontains=query)
            | Q(code__icontains=query)
            | Q(bank_name__icontains=query)
            | Q(instructions__icontains=query)
        )
    if method_type:
        qs = qs.filter(method_type=method_type)
    if mobile_provider:
        qs = qs.filter(mobile_money_provider=mobile_provider)
    if is_active:
        qs = qs.filter(is_active=(is_active.lower() == 'true'))
    if is_default:
        qs = qs.filter(is_default=(is_default.lower() == 'true'))
    if requires_approval:
        qs = qs.filter(requires_approval=(requires_approval.lower() == 'true'))
    if has_transaction_fee:
        qs = qs.filter(
            has_transaction_fee=(has_transaction_fee.lower() == 'true')
        )

    return qs


_PAYMENT_METHOD_COLUMNS = [
    ('name',                'Method Name',         lambda m: m.name),
    ('code',                'Code',                lambda m: m.code),
    ('method_type',         'Type',                lambda m: m.get_method_type_display()),
    ('is_active',           'Active',              lambda m: 'Yes' if m.is_active else 'No'),
    ('is_default',          'Default',             lambda m: 'Yes' if m.is_default else 'No'),
    ('requires_approval',   'Requires Approval',   lambda m: 'Yes' if m.requires_approval else 'No'),
    ('has_transaction_fee', 'Has Fee',             lambda m: 'Yes' if m.has_transaction_fee else 'No'),
    ('fee_bearer',          'Fee Bearer',          lambda m: m.get_fee_bearer_display() if m.fee_bearer else ''),
    ('minimum_amount',      'Min Amount',          lambda m: float(m.minimum_amount) if m.minimum_amount else ''),
    ('maximum_amount',      'Max Amount',          lambda m: float(m.maximum_amount) if m.maximum_amount else ''),
    ('processing_time',     'Processing Time',     lambda m: m.processing_time or ''),
    ('requires_reference',  'Requires Reference',  lambda m: 'Yes' if m.requires_reference else 'No'),
    ('bank_name',           'Bank Name',           lambda m: m.bank_name or ''),
    ('display_order',       'Display Order',       lambda m: m.display_order),
]
_PAYMENT_METHOD_COLUMN_MAP    = {c[0]: c for c in _PAYMENT_METHOD_COLUMNS}
_PAYMENT_METHOD_DEFAULT_FIELDS = [
    'name', 'code', 'method_type', 'is_active', 'requires_approval',
]

_PAYMENT_METHOD_FIELD_NAMES_FULL  = {c[0]: c[1] for c in _PAYMENT_METHOD_COLUMNS}
_PAYMENT_METHOD_FIELD_NAMES_SHORT = {
    'name':                'Method',
    'code':                'Code',
    'method_type':         'Type',
    'is_active':           'Active',
    'is_default':          'Default',
    'requires_approval':   'Approval',
    'has_transaction_fee': 'Fee',
    'fee_bearer':          'Bearer',
    'minimum_amount':      'Min',
    'maximum_amount':      'Max',
    'processing_time':     'Time',
    'requires_reference':  'Ref.',
    'bank_name':           'Bank',
    'display_order':       'Order',
}


@login_required
def payment_method_list(request):
    is_htmx     = request.headers.get('HX-Request') == 'true'
    filter_form = PaymentMethodFilterForm(request.GET or None)
    methods     = _get_filtered_payment_methods(request)

    stats_qs = PaymentMethod.objects.all()
    q = request.GET.get('q', '').strip()
    if q:
        stats_qs = stats_qs.filter(
            Q(name__icontains=q) | Q(code__icontains=q)
        )

    stats = stats_qs.aggregate(
        total=Count('id'),
        active=Count('id',        filter=Q(is_active=True)),
        cash=Count('id',          filter=Q(method_type='CASH')),
        mobile_money=Count('id',  filter=Q(method_type='MOBILE_MONEY')),
        bank_transfer=Count('id', filter=Q(method_type='BANK_TRANSFER')),
        with_fees=Count('id',     filter=Q(has_transaction_fee=True)),
        requires_approval=Count('id', filter=Q(requires_approval=True)),
    )

    methods_page, paginator = paginate_queryset(request, methods, per_page=20)

    context = {
        'payment_methods_page': methods_page,
        'paginator':            paginator,
        'stats':                stats,
        'filter_form':          filter_form,
        'is_htmx':              is_htmx,
    }

    if is_htmx:
        return render(
            request, 'core/payment_methods/_method_results.html', context
        )
    return render(request, 'core/payment_methods/list.html', context)


@login_required
def payment_method_create(request):
    if request.method == 'POST':
        form = PaymentMethodForm(request.POST)
        if form.is_valid():
            try:
                method = form.save()
                messages.success(
                    request, f'Payment method "{method.name}" created!'
                )
                return redirect('core:payment_method_detail', pk=method.pk)
            except Exception as e:
                logger.error(f"Payment method create error: {e}")
                messages.error(request, f'Error: {str(e)}')
    else:
        form = PaymentMethodForm()

    return render(request, 'core/payment_methods/form.html', {
        'form':        form,
        'title':       'Create Payment Method',
        'submit_text': 'Create Method',
    })


@login_required
def payment_method_detail(request, pk):
    method = get_object_or_404(PaymentMethod, pk=pk)

    try:
        usage_stats  = core_stats.get_payment_method_usage_stats(days=30)
        method_usage = usage_stats.get('by_method', {}).get(method.name, {})
    except Exception as e:
        logger.error(f"Payment method usage stats error: {e}")
        method_usage = {}

    return render(request, 'core/payment_methods/detail.html', {
        'method':       method,
        'method_usage': method_usage,
    })


@login_required
def payment_method_edit(request, pk):
    method = get_object_or_404(PaymentMethod, pk=pk)

    if request.method == 'POST':
        form = PaymentMethodForm(request.POST, instance=method)
        if form.is_valid():
            try:
                method = form.save()
                messages.success(
                    request, f'Payment method "{method.name}" updated!'
                )
                return redirect('core:payment_method_detail', pk=method.pk)
            except Exception as e:
                logger.error(f"Payment method edit error: {e}")
                messages.error(request, f'Error: {str(e)}')
    else:
        form = PaymentMethodForm(instance=method)

    return render(request, 'core/payment_methods/form.html', {
        'form':        form,
        'method':      method,
        'title':       f'Edit {method.name}',
        'submit_text': 'Update Method',
    })


@login_required
def payment_method_print_view(request):
    """
    Printable payment method list.
    Respects active filters. Supports short_headers, include_stats, landscape.
    """
    selected_fields = request.GET.getlist('fields') or _PAYMENT_METHOD_DEFAULT_FIELDS
    include_stats   = request.GET.get('include_stats') == 'true'
    landscape_mode  = request.GET.get('landscape')     == 'true'
    short_headers   = request.GET.get('short_headers') == 'true'

    methods = _get_filtered_payment_methods(request)

    field_names = (
        _PAYMENT_METHOD_FIELD_NAMES_SHORT if short_headers
        else _PAYMENT_METHOD_FIELD_NAMES_FULL
    )
    selected_field_names = [
        field_names.get(f, f.replace('_', ' ').title())
        for f in selected_fields
    ]

    context = {
        'methods':              methods,
        'now':                  get_school_current_time(),
        'selected_fields':      selected_fields,
        'selected_field_names': selected_field_names,
        'field_names':          field_names,
        'landscape':            landscape_mode,
        'short_headers':        short_headers,
        'title':                'Payment Methods Report',
        **get_print_school_context(request),
    }

    if include_stats:
        context['stats'] = {
            'total':  methods.count(),
            'active': methods.filter(is_active=True).count(),
            'cash':   methods.filter(method_type='CASH').count(),
            'mobile': methods.filter(method_type='MOBILE_MONEY').count(),
        }

    return render(request, 'core/payment_methods/print.html', context)


@login_required
def payment_method_export_excel(request):
    """Export filtered payment methods to Excel."""
    try:
        import openpyxl
    except ImportError:
        messages.error(request, 'Excel export requires openpyxl.')
        return redirect('core:payment_method_list')

    methods  = _get_filtered_payment_methods(request)
    selected = request.GET.getlist('fields') or _PAYMENT_METHOD_DEFAULT_FIELDS
    columns  = [
        _PAYMENT_METHOD_COLUMN_MAP[f]
        for f in selected
        if f in _PAYMENT_METHOD_COLUMN_MAP
    ] or [_PAYMENT_METHOD_COLUMN_MAP[f] for f in _PAYMENT_METHOD_DEFAULT_FIELDS]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Payment Methods'

    ws.append([col[1] for col in columns])
    _style_excel_header(ws)

    for method in methods:
        ws.append([col[2](method) for col in columns])

    _autosize_columns(ws)

    filename = f"payment_methods_{get_school_today().strftime('%Y%m%d')}.xlsx"
    return _build_excel_response(ws, wb, filename)


# =============================================================================
# TAX RATE — queryset helper + full CRUD + print + export
# =============================================================================

def _get_filtered_tax_rates(request):
    """
    Build a filtered TaxRate queryset from GET parameters.
    Shared by list, print, and export views.
    """
    qs = TaxRate.objects.all().order_by('-effective_from', 'tax_type')

    query               = request.GET.get('q',                   '').strip()
    tax_type            = request.GET.get('tax_type',            '').strip()
    is_active           = request.GET.get('is_active',           '').strip()
    applies_to_fees     = request.GET.get('applies_to_fees',     '').strip()
    applies_to_services = request.GET.get('applies_to_services', '').strip()
    start_date          = request.GET.get('start_date',          '').strip()
    end_date            = request.GET.get('end_date',            '').strip()

    if query:
        qs = qs.filter(
            Q(name__icontains=query)
            | Q(description__icontains=query)
            | Q(legal_reference__icontains=query)
        )
    if tax_type:
        qs = qs.filter(tax_type=tax_type)
    if is_active:
        qs = qs.filter(is_active=(is_active.lower() == 'true'))
    if applies_to_fees:
        qs = qs.filter(applies_to_fees=(applies_to_fees.lower() == 'true'))
    if applies_to_services:
        qs = qs.filter(
            applies_to_services=(applies_to_services.lower() == 'true')
        )
    if start_date:
        try:
            qs = qs.filter(effective_from__gte=start_date)
        except Exception:
            pass
    if end_date:
        try:
            qs = qs.filter(
                Q(effective_to__lte=end_date) | Q(effective_to__isnull=True)
            )
        except Exception:
            pass

    return qs


_TAX_RATE_COLUMNS = [
    ('name',            'Tax Name',         lambda r: r.name),
    ('tax_type',        'Tax Type',         lambda r: r.get_tax_type_display()),
    ('rate',            'Rate (%)',          lambda r: float(r.rate)),
    ('effective_from',  'Effective From',   lambda r: r.effective_from.strftime('%Y-%m-%d') if r.effective_from else ''),
    ('effective_to',    'Effective To',     lambda r: r.effective_to.strftime('%Y-%m-%d')   if r.effective_to   else 'Open'),
    ('is_active',       'Active',           lambda r: 'Yes' if r.is_active else 'No'),
    ('applies_to_fees', 'Applies to Fees',  lambda r: 'Yes' if r.applies_to_fees else 'No'),
    ('applies_to_services','Applies to Services', lambda r: 'Yes' if r.applies_to_services else 'No'),
    ('legal_reference', 'Legal Reference',  lambda r: r.legal_reference or ''),
]
_TAX_RATE_COLUMN_MAP    = {c[0]: c for c in _TAX_RATE_COLUMNS}
_TAX_RATE_DEFAULT_FIELDS = [
    'name', 'tax_type', 'rate', 'effective_from', 'effective_to', 'is_active',
]

_TAX_RATE_FIELD_NAMES_FULL  = {c[0]: c[1] for c in _TAX_RATE_COLUMNS}
_TAX_RATE_FIELD_NAMES_SHORT = {
    'name':               'Name',
    'tax_type':           'Type',
    'rate':               'Rate',
    'effective_from':     'From',
    'effective_to':       'To',
    'is_active':          'Active',
    'applies_to_fees':    'Fees',
    'applies_to_services':'Services',
    'legal_reference':    'Legal Ref.',
}


@login_required
def tax_rate_list(request):
    is_htmx     = request.headers.get('HX-Request') == 'true'
    filter_form = TaxRateFilterForm(request.GET or None)
    rates       = _get_filtered_tax_rates(request)

    stats_qs = TaxRate.objects.all()
    q = request.GET.get('q', '').strip()
    if q:
        stats_qs = stats_qs.filter(Q(name__icontains=q))

    stats = stats_qs.aggregate(
        total=Count('id'),
        active=Count('id',              filter=Q(is_active=True)),
        vat=Count('id',                 filter=Q(tax_type='VAT')),
        wht=Count('id',                 filter=Q(tax_type__startswith='WHT')),
        applies_to_fees=Count('id',     filter=Q(applies_to_fees=True)),
        applies_to_services=Count('id', filter=Q(applies_to_services=True)),
    )

    rates_page, paginator = paginate_queryset(request, rates, per_page=20)

    context = {
        'tax_rates_page': rates_page,
        'paginator':      paginator,
        'stats':          stats,
        'filter_form':    filter_form,
        'is_htmx':        is_htmx,
    }

    if is_htmx:
        return render(request, 'core/tax_rates/_rate_results.html', context)
    return render(request, 'core/tax_rates/list.html', context)


@login_required
def tax_rate_create(request):
    if request.method == 'POST':
        form = TaxRateForm(request.POST)
        if form.is_valid():
            try:
                rate = form.save()
                messages.success(request, f'Tax rate "{rate.name}" created!')
                return redirect('core:tax_rate_detail', pk=rate.pk)
            except Exception as e:
                logger.error(f"Tax rate create error: {e}")
                messages.error(request, f'Error: {str(e)}')
    else:
        form = TaxRateForm()

    return render(request, 'core/tax_rates/form.html', {
        'form':        form,
        'title':       'Create Tax Rate',
        'submit_text': 'Create Rate',
    })


@login_required
def tax_rate_detail(request, pk):
    rate  = get_object_or_404(TaxRate, pk=pk)
    today = get_school_today()
    return render(request, 'core/tax_rates/detail.html', {
        'rate':         rate,
        'is_effective': rate.is_effective(today),
        'today':        today,
    })


@login_required
def tax_rate_edit(request, pk):
    rate = get_object_or_404(TaxRate, pk=pk)

    if request.method == 'POST':
        form = TaxRateForm(request.POST, instance=rate)
        if form.is_valid():
            try:
                rate = form.save()
                messages.success(request, f'Tax rate "{rate.name}" updated!')
                return redirect('core:tax_rate_detail', pk=rate.pk)
            except Exception as e:
                logger.error(f"Tax rate edit error: {e}")
                messages.error(request, f'Error: {str(e)}')
    else:
        form = TaxRateForm(instance=rate)

    return render(request, 'core/tax_rates/form.html', {
        'form':        form,
        'rate':        rate,
        'title':       f'Edit {rate.name}',
        'submit_text': 'Update Rate',
    })


@login_required
def tax_rate_print_view(request):
    """
    Printable tax rate list.
    Respects active filters. Supports short_headers, include_stats, landscape.
    """
    selected_fields = request.GET.getlist('fields') or _TAX_RATE_DEFAULT_FIELDS
    include_stats   = request.GET.get('include_stats') == 'true'
    landscape_mode  = request.GET.get('landscape')     == 'true'
    short_headers   = request.GET.get('short_headers') == 'true'

    rates = _get_filtered_tax_rates(request)
    today = get_school_today()

    field_names = (
        _TAX_RATE_FIELD_NAMES_SHORT if short_headers
        else _TAX_RATE_FIELD_NAMES_FULL
    )
    selected_field_names = [
        field_names.get(f, f.replace('_', ' ').title())
        for f in selected_fields
    ]

    context = {
        'rates':                rates,
        'today':                today,
        'now':                  get_school_current_time(),
        'selected_fields':      selected_fields,
        'selected_field_names': selected_field_names,
        'field_names':          field_names,
        'landscape':            landscape_mode,
        'short_headers':        short_headers,
        'title':                'Tax Rates Report',
        **get_print_school_context(request),
    }

    if include_stats:
        context['stats'] = {
            'total':     rates.count(),
            'active':    rates.filter(is_active=True).count(),
            'effective': rates.filter(
                is_active=True, effective_from__lte=today
            ).filter(
                Q(effective_to__isnull=True) | Q(effective_to__gte=today)
            ).count(),
        }

    return render(request, 'core/tax_rates/print.html', context)


@login_required
def tax_rate_export_excel(request):
    """Export filtered tax rates to Excel."""
    try:
        import openpyxl
    except ImportError:
        messages.error(request, 'Excel export requires openpyxl.')
        return redirect('core:tax_rate_list')

    rates    = _get_filtered_tax_rates(request)
    selected = request.GET.getlist('fields') or _TAX_RATE_DEFAULT_FIELDS
    columns  = [
        _TAX_RATE_COLUMN_MAP[f]
        for f in selected
        if f in _TAX_RATE_COLUMN_MAP
    ] or [_TAX_RATE_COLUMN_MAP[f] for f in _TAX_RATE_DEFAULT_FIELDS]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Tax Rates'

    ws.append([col[1] for col in columns])
    _style_excel_header(ws)

    for rate in rates:
        ws.append([col[2](rate) for col in columns])

    _autosize_columns(ws)

    filename = f"tax_rates_{get_school_today().strftime('%Y%m%d')}.xlsx"
    return _build_excel_response(ws, wb, filename)


# =============================================================================
# UNIT OF MEASURE — queryset helper + full CRUD + print + export
# =============================================================================

def _get_filtered_units(request):
    """
    Build a filtered UnitOfMeasure queryset from GET parameters.
    Shared by list, print, and export views.
    """
    qs = UnitOfMeasure.objects.select_related('base_unit').annotate(
        derived_units_count=Count('derived_units', distinct=True)
    ).order_by('uom_type', 'name')

    query         = request.GET.get('q',            '').strip()
    uom_type      = request.GET.get('uom_type',     '').strip()
    is_active     = request.GET.get('is_active',    '').strip()
    has_base_unit = request.GET.get('has_base_unit','').strip()

    if query:
        qs = qs.filter(
            Q(name__icontains=query)
            | Q(abbreviation__icontains=query)
            | Q(symbol__icontains=query)
            | Q(description__icontains=query)
        )
    if uom_type:
        qs = qs.filter(uom_type=uom_type)
    if is_active:
        qs = qs.filter(is_active=(is_active.lower() == 'true'))
    if has_base_unit == 'true':
        qs = qs.filter(base_unit__isnull=False)
    elif has_base_unit == 'false':
        qs = qs.filter(base_unit__isnull=True)

    return qs


_UNIT_COLUMNS = [
    ('name',              'Name',               lambda u: u.name),
    ('abbreviation',      'Abbreviation',       lambda u: u.abbreviation),
    ('symbol',            'Symbol',             lambda u: u.symbol or ''),
    ('uom_type',          'Type',               lambda u: u.get_uom_type_display()),
    ('base_unit',         'Base Unit',          lambda u: u.base_unit.name if u.base_unit else '(Base)'),
    ('conversion_factor', 'Conversion Factor',  lambda u: float(u.conversion_factor)),
    ('derived_count',     'Derived Units',      lambda u: getattr(u, 'derived_units_count', u.get_derived_units_count())),
    ('is_active',         'Active',             lambda u: 'Yes' if u.is_active else 'No'),
    ('description',       'Description',        lambda u: u.description or ''),
]
_UNIT_COLUMN_MAP    = {c[0]: c for c in _UNIT_COLUMNS}
_UNIT_DEFAULT_FIELDS = [
    'name', 'abbreviation', 'uom_type', 'base_unit',
    'conversion_factor', 'is_active',
]

_UNIT_FIELD_NAMES_FULL = {c[0]: c[1] for c in _UNIT_COLUMNS}
_UNIT_FIELD_NAMES_SHORT = {
    'name':              'Name',
    'abbreviation':      'Abbr.',
    'symbol':            'Symbol',
    'uom_type':          'Type',
    'base_unit':         'Base',
    'conversion_factor': 'Factor',
    'derived_count':     'Derived',
    'is_active':         'Active',
    'description':       'Desc.',
}


@login_required
def unit_of_measure_list(request):
    is_htmx     = request.headers.get('HX-Request') == 'true'
    filter_form = UnitOfMeasureFilterForm(request.GET or None)
    units       = _get_filtered_units(request)

    stats_qs = UnitOfMeasure.objects.all()
    q = request.GET.get('q', '').strip()
    if q:
        stats_qs = stats_qs.filter(
            Q(name__icontains=q) | Q(abbreviation__icontains=q)
        )

    stats = stats_qs.aggregate(
        total=Count('id'),
        active=Count('id',         filter=Q(is_active=True)),
        base_units=Count('id',     filter=Q(base_unit__isnull=True)),
        derived_units=Count('id',  filter=Q(base_unit__isnull=False)),
        length=Count('id',         filter=Q(uom_type='LENGTH')),
        weight=Count('id',         filter=Q(uom_type='WEIGHT')),
        volume=Count('id',         filter=Q(uom_type='VOLUME')),
        area=Count('id',           filter=Q(uom_type='AREA')),
        quantity=Count('id',       filter=Q(uom_type='QUANTITY')),
    )

    units_page, paginator = paginate_queryset(request, units, per_page=20)

    context = {
        'units_page':  units_page,
        'paginator':   paginator,
        'stats':       stats,
        'filter_form': filter_form,
        'is_htmx':     is_htmx,
    }

    if is_htmx:
        return render(request, 'core/units/_unit_results.html', context)
    return render(request, 'core/units/list.html', context)


@login_required
def unit_of_measure_create(request):
    if request.method == 'POST':
        form = UnitOfMeasureForm(request.POST)
        if form.is_valid():
            try:
                unit = form.save()
                messages.success(request, f'Unit "{unit.name}" created!')
                return redirect('core:unit_of_measure_detail', pk=unit.pk)
            except Exception as e:
                logger.error(f"Unit create error: {e}")
                messages.error(request, f'Error: {str(e)}')
    else:
        form = UnitOfMeasureForm()

    return render(request, 'core/units/form.html', {
        'form':        form,
        'title':       'Create Unit of Measure',
        'submit_text': 'Create Unit',
    })


@login_required
def unit_of_measure_detail(request, pk):
    unit = get_object_or_404(UnitOfMeasure, pk=pk)
    return render(request, 'core/units/detail.html', {
        'unit':                  unit,
        'conversion_examples':   unit.get_conversion_examples(),
        'conversion_table':      unit.get_conversion_table(),
        'derived_units':         unit.get_all_derived_units(),
    })


@login_required
def unit_of_measure_edit(request, pk):
    unit = get_object_or_404(UnitOfMeasure, pk=pk)

    if request.method == 'POST':
        form = UnitOfMeasureForm(request.POST, instance=unit)
        if form.is_valid():
            try:
                unit = form.save()
                messages.success(request, f'Unit "{unit.name}" updated!')
                return redirect('core:unit_of_measure_detail', pk=unit.pk)
            except Exception as e:
                logger.error(f"Unit edit error: {e}")
                messages.error(request, f'Error: {str(e)}')
    else:
        form = UnitOfMeasureForm(instance=unit)

    return render(request, 'core/units/form.html', {
        'form':        form,
        'unit':        unit,
        'title':       f'Edit {unit.name}',
        'submit_text': 'Update Unit',
    })


@login_required
def unit_of_measure_print_view(request):
    """
    Printable unit of measure list.
    Respects active filters. Supports short_headers, include_stats, landscape.
    """
    selected_fields = request.GET.getlist('fields') or _UNIT_DEFAULT_FIELDS
    include_stats   = request.GET.get('include_stats') == 'true'
    landscape_mode  = request.GET.get('landscape')     == 'true'
    short_headers   = request.GET.get('short_headers') == 'true'

    units = _get_filtered_units(request)

    field_names = (
        _UNIT_FIELD_NAMES_SHORT if short_headers
        else _UNIT_FIELD_NAMES_FULL
    )
    selected_field_names = [
        field_names.get(f, f.replace('_', ' ').title())
        for f in selected_fields
    ]

    context = {
        'units':                units,
        'now':                  get_school_current_time(),
        'selected_fields':      selected_fields,
        'selected_field_names': selected_field_names,
        'field_names':          field_names,
        'landscape':            landscape_mode,
        'short_headers':        short_headers,
        'title':                'Units of Measure Report',
        **get_print_school_context(request),
    }

    if include_stats:
        context['stats'] = {
            'total':        units.count(),
            'active':       units.filter(is_active=True).count(),
            'base_units':   units.filter(base_unit__isnull=True,  is_active=True).count(),
            'derived_units':units.filter(base_unit__isnull=False, is_active=True).count(),
        }

    return render(request, 'core/units/print.html', context)


@login_required
def unit_of_measure_export_excel(request):
    """Export filtered units of measure to Excel."""
    try:
        import openpyxl
    except ImportError:
        messages.error(request, 'Excel export requires openpyxl.')
        return redirect('core:unit_of_measure_list')

    units    = _get_filtered_units(request)
    selected = request.GET.getlist('fields') or _UNIT_DEFAULT_FIELDS
    columns  = [
        _UNIT_COLUMN_MAP[f]
        for f in selected
        if f in _UNIT_COLUMN_MAP
    ] or [_UNIT_COLUMN_MAP[f] for f in _UNIT_DEFAULT_FIELDS]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Units of Measure'

    ws.append([col[1] for col in columns])
    _style_excel_header(ws)

    for unit in units:
        ws.append([col[2](unit) for col in columns])

    _autosize_columns(ws)

    filename = f"units_of_measure_{get_school_today().strftime('%Y%m%d')}.xlsx"
    return _build_excel_response(ws, wb, filename)


# =============================================================================
# JSON QUICK-STATS
# =============================================================================

@login_required
@require_http_methods(["GET"])
def fiscal_year_quick_stats(request):
    current = FiscalYear.get_active_fiscal_year()
    return JsonResponse({
        'total':                 FiscalYear.objects.count(),
        'active':                FiscalYear.objects.filter(is_active=True).count(),
        'draft':                 FiscalYear.objects.filter(status='DRAFT').count(),
        'closed':                FiscalYear.objects.filter(is_closed=True).count(),
        'locked':                FiscalYear.objects.filter(is_locked=True).count(),
        'current_year_name':     current.name if current else None,
        'current_year_progress': current.get_progress_percentage() if current else 0,
    })


@login_required
@require_http_methods(["GET"])
def fiscal_period_quick_stats(request):
    current = FiscalPeriod.get_current_fiscal_period()
    return JsonResponse({
        'total':                   FiscalPeriod.objects.count(),
        'active':                  FiscalPeriod.objects.filter(is_active=True).count(),
        'closed':                  FiscalPeriod.objects.filter(is_closed=True).count(),
        'locked':                  FiscalPeriod.objects.filter(is_locked=True).count(),
        'current_period_name':     current.name if current else None,
        'current_period_progress': current.get_progress_percentage() if current else 0,
    })


@login_required
@require_http_methods(["GET"])
def payment_method_quick_stats(request):
    return JsonResponse({
        'total':         PaymentMethod.objects.filter(is_active=True).count(),
        'cash':          PaymentMethod.objects.filter(method_type='CASH',         is_active=True).count(),
        'mobile_money':  PaymentMethod.objects.filter(method_type='MOBILE_MONEY', is_active=True).count(),
        'bank_transfer': PaymentMethod.objects.filter(method_type='BANK_TRANSFER',is_active=True).count(),
        'with_fees':     PaymentMethod.objects.filter(has_transaction_fee=True,   is_active=True).count(),
    })


@login_required
@require_http_methods(["GET"])
def tax_rate_quick_stats(request):
    today = get_school_today()
    return JsonResponse({
        'total':            TaxRate.objects.filter(is_active=True).count(),
        'vat':              TaxRate.objects.filter(tax_type='VAT', is_active=True).count(),
        'current_vat_rate': float(TaxRate.get_vat_rate()),
        'effective_today':  TaxRate.objects.filter(
            is_active=True, effective_from__lte=today,
        ).filter(
            Q(effective_to__isnull=True) | Q(effective_to__gte=today)
        ).count(),
    })


@login_required
@require_http_methods(["GET"])
def unit_of_measure_quick_stats(request):
    return JsonResponse({
        'total':         UnitOfMeasure.objects.filter(is_active=True).count(),
        'base_units':    UnitOfMeasure.objects.filter(base_unit__isnull=True,  is_active=True).count(),
        'derived_units': UnitOfMeasure.objects.filter(base_unit__isnull=False, is_active=True).count(),
        'length':        UnitOfMeasure.objects.filter(uom_type='LENGTH',   is_active=True).count(),
        'weight':        UnitOfMeasure.objects.filter(uom_type='WEIGHT',   is_active=True).count(),
        'volume':        UnitOfMeasure.objects.filter(uom_type='VOLUME',   is_active=True).count(),
        'quantity':      UnitOfMeasure.objects.filter(uom_type='QUANTITY', is_active=True).count(),
    })


@login_required
@require_http_methods(["GET"])
def system_configuration_stats(request):
    school_config      = SchoolConfiguration.get_cached_instance()
    financial_settings = FinancialSettings.get_instance()
    return JsonResponse({
        'term_system':     school_config.get_term_system_display()  if school_config      else 'Not Configured',
        'periods_per_year':school_config.get_period_count()         if school_config      else 0,
        'school_currency': financial_settings.school_currency        if financial_settings else 'UGX',
        'fiscal_years':    FiscalYear.objects.count(),
        'fiscal_periods':  FiscalPeriod.objects.count(),
        'payment_methods': PaymentMethod.objects.filter(is_active=True).count(),
        'tax_rates':       TaxRate.objects.filter(is_active=True).count(),
        'units_of_measure':UnitOfMeasure.objects.filter(is_active=True).count(),
    })