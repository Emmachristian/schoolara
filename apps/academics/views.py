"""
academics/views.py
"""
from django.shortcuts import render, redirect, get_object_or_404
from django.urls import reverse
from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator
from django.contrib import messages
from django.db.models import Q, Count, Sum, Avg, F
from django.utils import timezone
from django.http import JsonResponse, HttpResponse
from django.db import transaction
from django.core.exceptions import ValidationError
from django.views.decorators.http import require_http_methods

from datetime import timedelta, date, datetime
from decimal import Decimal
import logging
import json

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from core.utils import (
    get_school_today,
    get_school_current_time,
    get_school_timezone,
    localize_datetime,
    get_active_academic_session,
    format_money,
    calculate_percentage,
)

from core.view_helpers import (
    get_print_school_context,
)

from .models import (
    AcademicSession, Subject, AcademicLevel, ClassRoom, Class,
    StudentClassEnrollment, ClassSubject, AcademicProgress, Holiday,
)
from .forms import (
    AcademicSessionFilterForm, SubjectFilterForm, AcademicLevelFilterForm,
    ClassRoomFilterForm,  AcademicProgressFilterForm, HolidayFilterForm,
    AcademicSessionForm, SubjectForm, AcademicLevelForm, ClassRoomForm, ClassForm,
    StudentEnrollmentForm, BulkEnrollmentStudentSelectionForm,
    BulkEnrollmentConfirmationForm, ClassSubjectForm,
    AcademicProgressForm, HolidayForm,
)
from students.models import Student
from . import stats as academic_stats

logger = logging.getLogger(__name__)


# =============================================================================
# 2. DASHBOARD
# =============================================================================

@login_required
def academics_dashboard(request):
    try:
        overview         = academic_stats.get_academic_dashboard_statistics()
        current_session  = AcademicSession.get_current_session()
        enrollment_stats = academic_stats.get_enrollment_statistics()
    except Exception as e:
        logger.error(f"Error getting dashboard statistics: {e}")
        overview = enrollment_stats = {}
        current_session = None

    today = get_school_today()
    recent_sessions      = AcademicSession.objects.order_by('-created_at')[:10]
    recent_enrollments   = StudentClassEnrollment.objects.select_related(
        'student', 'class_instance', 'academic_session'
    ).order_by('-created_at')[:10]
    upcoming_holidays    = Holiday.objects.filter(
        start_date__gte=today
    ).order_by('start_date')[:10]
    classes_at_capacity  = Class.objects.annotate(
        enrollment_count=Count(
            'enrollments', filter=Q(enrollments__is_active=True)
        )
    ).filter(
        enrollment_count__gte=F('max_students')
    ).order_by('-enrollment_count')[:10]
    sessions_ending_soon = AcademicSession.objects.filter(
        end_date__gte=today,
        end_date__lte=today + timedelta(days=30),
        is_active=True,
    ).order_by('end_date')[:10]
    pending_progress     = AcademicProgress.objects.filter(
        is_final=False
    ).select_related('student', 'academic_session').order_by('-updated_at')[:10]

    return render(request, 'academics/dashboard.html', {
        'overview':               overview,
        'current_session':        current_session,
        # Access sub-stats from overview rather than re-fetching them
        'session_stats':          overview.get('sessions', {}),
        'class_stats':            overview.get('classes', {}),
        'subject_stats':          overview.get('subjects', {}),
        'enrollment_stats':       enrollment_stats,
        'recent_sessions':        recent_sessions,
        'recent_enrollments':     recent_enrollments,
        'upcoming_holidays':      upcoming_holidays,
        'classes_at_capacity':    classes_at_capacity,
        'sessions_ending_soon':   sessions_ending_soon,
        'pending_progress_records': pending_progress,
    })


# =============================================================================
# 3. ACADEMIC SESSIONS
# =============================================================================

def get_filtered_academic_sessions(request):
    qs = AcademicSession.objects.order_by('-start_date', 'term_number')
    q = request.GET.get('q', '').strip()
    if q:
        qs = qs.filter(Q(year_name__icontains=q) | Q(term_name__icontains=q) | Q(description__icontains=q))
    year_name = request.GET.get('year_name', '')
    if year_name:
        qs = qs.filter(year_name__icontains=year_name)
    period_type = request.GET.get('period_type', '')
    if period_type:
        qs = qs.filter(period_type=period_type)
    for bf in ('is_current', 'is_active', 'is_academically_closed', 'is_special_session', 'allows_promotion'):
        v = request.GET.get(bf, '')
        if v:
            qs = qs.filter(**{bf: v.lower() == 'true'})
    tn = request.GET.get('term_number', '')
    if tn:
        try:
            qs = qs.filter(term_number=int(tn))
        except (ValueError, TypeError):
            pass
    return qs


@login_required
def academic_session_list(request):
    filter_form = AcademicSessionFilterForm(request.GET or None)
    sessions    = get_filtered_academic_sessions(request)
    today       = get_school_today()
    stats = {
        'total': sessions.count(), 'current': sessions.filter(is_current=True).count(),
        'active': sessions.filter(is_active=True).count(), 'closed': sessions.filter(is_academically_closed=True).count(),
        'special': sessions.filter(is_special_session=True).count(), 'regular': sessions.filter(is_special_session=False).count(),
        'upcoming': sessions.filter(start_date__gt=today).count(),
        'ongoing': sessions.filter(start_date__lte=today, end_date__gte=today, is_active=True).count(),
    }
    paginator     = Paginator(sessions, 10)
    sessions_page = paginator.get_page(request.GET.get('page', 1))
    is_htmx       = request.headers.get('HX-Request') == 'true'
    context = {'sessions_page': sessions_page, 'paginator': paginator, 'stats': stats, 'filter_form': filter_form, 'is_htmx': is_htmx}
    if is_htmx:
        return render(request, 'academics/sessions/partials/_session_results.html', context)
    return render(request, 'academics/sessions/list.html', context)


@login_required
def academic_session_detail(request, pk):
    session = get_object_or_404(AcademicSession, pk=pk)
    try:
        session_stats = academic_stats.get_academic_session_statistics({'year_name': session.year_name})
    except Exception as e:
        logger.error(f"Error getting session stats: {e}")
        session_stats = {}
    classes            = session.classes.select_related('academic_level', 'class_teacher__staff').prefetch_related('enrollments')[:10]
    recent_enrollments = session.student_class_enrollments.select_related('student__current_academic_level', 'class_instance__academic_level').order_by('-created_at')[:10]
    holidays           = session.holidays.order_by('start_date')[:10]
    today      = get_school_today()
    total_days = (session.end_date - session.start_date).days + 1
    if today < session.start_date:
        progress_info = {'days_until_start': (session.start_date - today).days, 'days_elapsed': 0, 'days_remaining': total_days, 'total_days': total_days, 'is_current': False, 'is_future': True, 'is_past': False, 'progress_percentage': 0}
    elif today > session.end_date:
        progress_info = {'days_since_end': (today - session.end_date).days, 'days_elapsed': total_days, 'days_remaining': 0, 'total_days': total_days, 'is_current': False, 'is_future': False, 'is_past': True, 'progress_percentage': 100}
    else:
        elapsed = (today - session.start_date).days
        progress_info = {'days_elapsed': elapsed, 'days_remaining': (session.end_date - today).days, 'total_days': total_days, 'is_current': True, 'is_future': False, 'is_past': False, 'progress_percentage': round((elapsed / total_days) * 100, 1) if total_days else 0}
    return render(request, 'academics/sessions/detail.html', {
        'session': session, 'session_stats': session_stats, 'classes': classes,
        'recent_enrollments': recent_enrollments, 'holidays': holidays, 'progress_info': progress_info,
    })


@login_required
def academic_session_create(request):
    if request.method == 'POST':
        form = AcademicSessionForm(request.POST)
        if form.is_valid():
            try:
                with transaction.atomic():
                    session = form.save()
                if request.headers.get('HX-Request') == 'true':
                    r = HttpResponse(); r['HX-Alert-Message'] = f'Session "{session.name}" created successfully'; r['HX-Alert-Type'] = 'success'; r['HX-Close-Modal'] = 'true'; r['HX-Redirect'] = reverse('academics:session_detail', kwargs={'pk': session.pk}); return r
                messages.success(request, f'Session "{session.name}" created successfully')
                return redirect('academics:session_detail', pk=session.pk)
            except Exception as e:
                logger.error(f"Error creating session: {e}")
                if request.headers.get('HX-Request') == 'true':
                    r = HttpResponse(); r['HX-Alert-Message'] = f'Error: {e}'; r['HX-Alert-Type'] = 'error'; return r
                messages.error(request, f'Error: {e}')
        else:
            if request.headers.get('HX-Request') == 'true':
                r = HttpResponse(); r['HX-Alert-Message'] = 'Please correct the errors in the form'; r['HX-Alert-Type'] = 'error'; return r
            messages.error(request, 'Please correct the errors in the form')
    else:
        form = AcademicSessionForm()
    return render(request, 'academics/sessions/form.html', {'form': form, 'title': 'Create academic session'})


@login_required
def academic_session_edit(request, pk):
    session = get_object_or_404(AcademicSession, pk=pk)
    if request.method == 'POST':
        form = AcademicSessionForm(request.POST, instance=session)
        if form.is_valid():
            try:
                with transaction.atomic():
                    session = form.save()
                if request.headers.get('HX-Request') == 'true':
                    r = HttpResponse(); r['HX-Alert-Message'] = f'Session "{session.name}" updated successfully'; r['HX-Alert-Type'] = 'success'; r['HX-Close-Modal'] = 'true'; r['HX-Redirect'] = reverse('academics:session_detail', kwargs={'pk': session.pk}); return r
                messages.success(request, 'Session updated successfully')
                return redirect('academics:session_detail', pk=session.pk)
            except Exception as e:
                logger.error(f"Error updating session: {e}")
                if request.headers.get('HX-Request') == 'true':
                    r = HttpResponse(); r['HX-Alert-Message'] = f'Error: {e}'; r['HX-Alert-Type'] = 'error'; return r
                messages.error(request, f'Error: {e}')
        else:
            if request.headers.get('HX-Request') == 'true':
                r = HttpResponse(); r['HX-Alert-Message'] = 'Please correct the errors in the form'; r['HX-Alert-Type'] = 'error'; return r
            messages.error(request, 'Please correct the errors in the form')
    else:
        form = AcademicSessionForm(instance=session)
    return render(request, 'academics/sessions/form.html', {'form': form, 'session': session, 'title': f'Edit {session.name}'})


@login_required
def academic_session_delete(request, pk):
    session = get_object_or_404(AcademicSession, pk=pk)
    if request.method == 'POST':
        if session.is_current or session.is_active:
            if request.headers.get('HX-Request') == 'true':
                r = HttpResponse(); r['HX-Alert-Message'] = 'Cannot delete active or current sessions'; r['HX-Alert-Type'] = 'error'; r['HX-Close-Modal'] = 'true'; return r
            messages.error(request, 'Cannot delete active or current sessions')
            return redirect('academics:session_detail', pk=pk)
        if session.student_class_enrollments.exists() or session.classes.exists():
            if request.headers.get('HX-Request') == 'true':
                r = HttpResponse(); r['HX-Alert-Message'] = 'Cannot delete session with enrollments or classes'; r['HX-Alert-Type'] = 'error'; r['HX-Close-Modal'] = 'true'; return r
            messages.error(request, 'Cannot delete session with enrollments or classes')
            return redirect('academics:session_detail', pk=pk)
        try:
            name = session.name
            session.delete()
            if request.headers.get('HX-Request') == 'true':
                r = HttpResponse(); r['HX-Alert-Message'] = f'Session "{name}" deleted successfully'; r['HX-Alert-Type'] = 'success'; r['HX-Close-Modal'] = 'true'; r['HX-Redirect'] = reverse('academics:session_list'); return r
            messages.success(request, f'Session "{name}" deleted successfully')
            return redirect('academics:session_list')
        except Exception as e:
            logger.error(f"Error deleting session: {e}")
            if request.headers.get('HX-Request') == 'true':
                r = HttpResponse(); r['HX-Alert-Message'] = f'Error: {e}'; r['HX-Alert-Type'] = 'error'; r['HX-Close-Modal'] = 'true'; return r
            messages.error(request, f'Error: {e}')
            return redirect('academics:session_detail', pk=pk)
    return redirect('academics:session_list')


@login_required
def academic_session_toggle_active(request, pk):
    session = get_object_or_404(AcademicSession, pk=pk)
    if request.method == 'POST':
        try:
            session.is_active = not session.is_active
            session.save(update_fields=['is_active'])
            word = 'activated' if session.is_active else 'deactivated'
            if request.headers.get('HX-Request') == 'true':
                r = HttpResponse(); r['HX-Alert-Message'] = f'Session {word} successfully'; r['HX-Alert-Type'] = 'success'; r['HX-Close-Modal'] = 'true'; r['HX-Redirect'] = reverse('academics:session_detail', kwargs={'pk': pk}); return r
            messages.success(request, f'Session {word}')
        except Exception as e:
            if request.headers.get('HX-Request') == 'true':
                r = HttpResponse(); r['HX-Alert-Message'] = f'Error: {e}'; r['HX-Alert-Type'] = 'error'; r['HX-Close-Modal'] = 'true'; return r
            messages.error(request, f'Error: {e}')
    return redirect('academics:session_detail', pk=pk)


@login_required
def academic_session_set_current(request, pk):
    session = get_object_or_404(AcademicSession, pk=pk)
    if request.method == 'POST':
        try:
            AcademicSession.objects.filter(is_current=True).update(is_current=False)
            session.is_current = True
            session.save()
            if request.headers.get('HX-Request') == 'true':
                r = HttpResponse(); r['HX-Alert-Message'] = f'"{session.name}" is now the current session'; r['HX-Alert-Type'] = 'success'; r['HX-Close-Modal'] = 'true'; r['HX-Redirect'] = reverse('academics:session_detail', kwargs={'pk': pk}); return r
            messages.success(request, f'Current session updated to "{session.name}"')
        except Exception as e:
            if request.headers.get('HX-Request') == 'true':
                r = HttpResponse(); r['HX-Alert-Message'] = f'Error: {e}'; r['HX-Alert-Type'] = 'error'; r['HX-Close-Modal'] = 'true'; return r
            messages.error(request, f'Error: {e}')
    return redirect('academics:session_detail', pk=pk)


@login_required
def academic_session_close(request, pk):
    session = get_object_or_404(AcademicSession, pk=pk)
    if request.method == 'POST':
        if session.is_academically_closed:
            if request.headers.get('HX-Request') == 'true':
                r = HttpResponse(); r['HX-Alert-Message'] = f'Session "{session.name}" is already closed'; r['HX-Alert-Type'] = 'warning'; r['HX-Close-Modal'] = 'true'; return r
            messages.warning(request, f'Session "{session.name}" is already closed')
            return redirect('academics:session_detail', pk=pk)
        if session.is_current:
            if request.headers.get('HX-Request') == 'true':
                r = HttpResponse(); r['HX-Alert-Message'] = 'Cannot close the current session. Set another session as current first.'; r['HX-Alert-Type'] = 'error'; r['HX-Close-Modal'] = 'true'; return r
            messages.error(request, 'Cannot close the current session')
            return redirect('academics:session_detail', pk=pk)
        try:
            session.close_academically(user=request.user)
            if request.headers.get('HX-Request') == 'true':
                r = HttpResponse(); r['HX-Alert-Message'] = f'Session "{session.name}" closed successfully'; r['HX-Alert-Type'] = 'success'; r['HX-Close-Modal'] = 'true'; r['HX-Redirect'] = reverse('academics:session_detail', kwargs={'pk': pk}); return r
            messages.success(request, f'Session "{session.name}" closed successfully')
        except Exception as e:
            logger.error(f"Error closing session: {e}")
            if request.headers.get('HX-Request') == 'true':
                r = HttpResponse(); r['HX-Alert-Message'] = f'Error: {e}'; r['HX-Alert-Type'] = 'error'; r['HX-Close-Modal'] = 'true'; return r
            messages.error(request, f'Error: {e}')
    return redirect('academics:session_detail', pk=pk)


@login_required
def academic_session_reopen(request, pk):
    session = get_object_or_404(AcademicSession, pk=pk)
    if request.method == 'POST':
        if not session.is_academically_closed:
            if request.headers.get('HX-Request') == 'true':
                r = HttpResponse(); r['HX-Alert-Message'] = f'Session "{session.name}" is not closed'; r['HX-Alert-Type'] = 'warning'; r['HX-Close-Modal'] = 'true'; return r
            messages.warning(request, f'Session "{session.name}" is not closed')
            return redirect('academics:session_detail', pk=pk)
        try:
            session.reopen_academically(user=request.user)
            if request.headers.get('HX-Request') == 'true':
                r = HttpResponse(); r['HX-Alert-Message'] = f'Session "{session.name}" reopened successfully'; r['HX-Alert-Type'] = 'success'; r['HX-Close-Modal'] = 'true'; r['HX-Redirect'] = reverse('academics:session_detail', kwargs={'pk': pk}); return r
            messages.success(request, f'Session "{session.name}" reopened')
        except Exception as e:
            logger.error(f"Error reopening session: {e}")
            if request.headers.get('HX-Request') == 'true':
                r = HttpResponse(); r['HX-Alert-Message'] = f'Error: {e}'; r['HX-Alert-Type'] = 'error'; r['HX-Close-Modal'] = 'true'; return r
            messages.error(request, f'Error: {e}')
    return redirect('academics:session_detail', pk=pk)


@login_required
def academic_session_print_view(request):
    sessions        = get_filtered_academic_sessions(request)
    selected_fields = request.GET.getlist('fields') or ['year_name', 'term_name', 'period_type', 'start_date', 'end_date', 'status_display', 'is_current']
    include_stats   = request.GET.get('include_stats') == 'true'
    stats = None
    if include_stats:
        stats = sessions.aggregate(total=Count('id'), active=Count('id', filter=Q(is_active=True)), current=Count('id', filter=Q(is_current=True)), closed=Count('id', filter=Q(is_academically_closed=True)))
    field_labels = {
        'year_name': 'Academic Year', 'term_name': 'Period Name', 'term_number': 'Period #',
        'period_type': 'Period Type', 'start_date': 'Start Date', 'end_date': 'End Date',
        'enrollment_deadline': 'Enrollment Deadline', 'is_current': 'Current',
        'is_active': 'Active', 'is_academically_closed': 'Closed', 'status_display': 'Status',
        'allows_promotion': 'Allows Promotion', 'total_days': 'Total Days',
    }
    return render(request, 'academics/sessions/print.html', {
        **_get_print_school_context(request),
        'sessions': sessions, 'stats': stats, 'now': timezone.now(),
        'selected_fields': selected_fields,
        'selected_field_names': [field_labels.get(f, f.replace('_', ' ').title()) for f in selected_fields],
        'field_labels': field_labels, 'landscape': request.GET.get('landscape') == 'true',
        'title': 'Academic Sessions Report',
    })


@login_required
def export_academic_sessions_excel(request):
    ALL_COLUMNS = [
        ('year_name',        'Academic Year',    lambda s: s.year_name),
        ('term_name',        'Period Name',      lambda s: s.term_name),
        ('term_number',      'Period #',         lambda s: s.term_number),
        ('period_type',      'Period Type',      lambda s: s.get_period_type_display()),
        ('start_date',       'Start Date',       lambda s: s.start_date.strftime('%Y-%m-%d') if s.start_date else ''),
        ('end_date',         'End Date',         lambda s: s.end_date.strftime('%Y-%m-%d') if s.end_date else ''),
        ('total_days',       'Total Days',       lambda s: s.total_days),
        ('is_current',       'Current',          lambda s: 'Yes' if s.is_current else 'No'),
        ('is_active',        'Active',           lambda s: 'Yes' if s.is_active else 'No'),
        ('is_closed',        'Closed',           lambda s: 'Yes' if s.is_academically_closed else 'No'),
        ('is_special',       'Special',          lambda s: 'Yes' if s.is_special_session else 'No'),
        ('status',           'Status',           lambda s: s.status_display),
        ('allows_promotion', 'Allows Promotion', lambda s: 'Yes' if s.allows_promotion else 'No'),
        ('min_attendance',   'Min Attendance %', lambda s: float(s.minimum_attendance_percentage)),
        ('description',      'Description',      lambda s: s.description or ''),
    ]
    DEFAULT  = ['year_name', 'term_name', 'period_type', 'start_date', 'end_date', 'status', 'is_current']
    col_map  = {k: (l, fn) for k, l, fn in ALL_COLUMNS}
    chosen   = request.GET.getlist('fields') or DEFAULT
    columns  = [(col_map[f][0], col_map[f][1]) for f in chosen if f in col_map]
    sessions = get_filtered_academic_sessions(request)
    wb = Workbook(); ws = wb.active; ws.title = 'Academic Sessions'
    ws.append([c[0] for c in columns])
    hf = PatternFill(start_color='1E3A5F', end_color='1E3A5F', fill_type='solid')
    for cell in ws[1]:
        cell.fill = hf; cell.font = Font(bold=True, color='FFFFFF', size=11)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws.row_dimensions[1].height = 28
    af = PatternFill(start_color='F4F6F9', end_color='F4F6F9', fill_type='solid')
    da = Alignment(vertical='center', wrap_text=False)
    for i, obj in enumerate(sessions):
        ws.append([c[1](obj) for c in columns])
        for cell in ws[i + 2]:
            cell.alignment = da
            if i % 2 == 1: cell.fill = af
    for cc in ws.columns:
        ml = max((len(str(c.value)) if c.value else 0) for c in cc)
        ws.column_dimensions[cc[0].column_letter].width = min(ml + 4, 60)
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="academic_sessions_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
    wb.save(response); return response


# =============================================================================
# 4. ACADEMIC LEVELS
# =============================================================================

def get_filtered_academic_levels(request):
    qs = AcademicLevel.objects.select_related('next_level').annotate(class_count=Count('classes', distinct=True)).order_by('order')
    q = request.GET.get('q', '').strip()
    if q:
        qs = qs.filter(Q(name__icontains=q) | Q(code__icontains=q) | Q(description__icontains=q))
    for bf in ('is_active', 'has_sections', 'is_graduation_level'):
        v = request.GET.get(bf, '')
        if v:
            qs = qs.filter(**{bf: v.lower() == 'true'})
    return qs


@login_required
def academic_level_list(request):
    filter_form = AcademicLevelFilterForm(request.GET or None)
    levels      = get_filtered_academic_levels(request)
    stats = {
        'total': levels.count(), 'active': levels.filter(is_active=True).count(),
        'with_sections': levels.filter(has_sections=True).count(),
        'graduation_levels': levels.filter(is_graduation_level=True).count(),
    }
    paginator   = Paginator(levels, 20)
    levels_page = paginator.get_page(request.GET.get('page', 1))
    is_htmx     = request.headers.get('HX-Request') == 'true'
    context = {'levels_page': levels_page, 'paginator': paginator, 'stats': stats, 'filter_form': filter_form, 'is_htmx': is_htmx}
    if is_htmx:
        return render(request, 'academics/levels/partials/_level_results.html', context)
    return render(request, 'academics/levels/list.html', context)


@login_required
def academic_level_detail(request, pk):
    level = get_object_or_404(AcademicLevel, pk=pk)

    previous_levels = AcademicLevel.objects.filter(next_level=level).order_by('order')

    sessions = AcademicSession.objects.filter(
        classes__academic_level=level
    ).distinct().order_by('-start_date')

    stats = {
        'total_classes':   level.classes.filter(is_active=True).count(),
        'total_students':  Student.objects.filter(current_academic_level=level).count(),
        'active_students': Student.objects.filter(
            current_academic_level=level, enrollment_status='ACTIVE'
        ).count(),
    }

    return render(request, 'academics/levels/detail.html', {
        'level':               level,
        'previous_levels':     previous_levels,
        'sessions':            sessions,
        'stats':               stats,
    })


@login_required
def level_classes_partial(request, pk):
    level = get_object_or_404(AcademicLevel, pk=pk)

    qs = level.classes.select_related(
        'academic_session', 'class_teacher__staff', 'classroom'
    ).annotate(
        enrollment_count=Count('enrollments', distinct=True)
    ).order_by('-academic_session__start_date', 'section')

    q       = request.GET.get('q', '').strip()
    session = request.GET.get('academic_session', '')
    active  = request.GET.get('is_active', '')

    if q:
        qs = qs.filter(
            Q(section__icontains=q) |
            Q(class_teacher__staff__first_name__icontains=q) |
            Q(class_teacher__staff__last_name__icontains=q) |
            Q(classroom__room_number__icontains=q)
        )
    if session: qs = qs.filter(academic_session_id=session)
    if active:  qs = qs.filter(is_active=(active.lower() == 'true'))

    paginator = Paginator(qs, 25)
    page      = paginator.get_page(request.GET.get('page', 1))

    return render(request, 'academics/levels/partials/_class_results.html', {
        'classes': page,
        'level':   level,
    })


@login_required
def academic_level_create(request):
    if request.method == 'POST':
        form = AcademicLevelForm(request.POST)
        if form.is_valid():
            try:
                level = form.save()
                if request.headers.get('HX-Request') == 'true':
                    r = HttpResponse(); r['HX-Alert-Message'] = f'Level "{level.name}" created successfully'; r['HX-Alert-Type'] = 'success'; r['HX-Close-Modal'] = 'true'; r['HX-Redirect'] = reverse('academics:level_detail', kwargs={'pk': level.pk}); return r
                messages.success(request, f'Level "{level.name}" created successfully')
                return redirect('academics:level_detail', pk=level.pk)
            except Exception as e:
                logger.error(f"Error creating level: {e}")
                if request.headers.get('HX-Request') == 'true':
                    r = HttpResponse(); r['HX-Alert-Message'] = f'Error: {e}'; r['HX-Alert-Type'] = 'error'; return r
                messages.error(request, f'Error: {e}')
        else:
            if request.headers.get('HX-Request') == 'true':
                r = HttpResponse(); r['HX-Alert-Message'] = 'Please correct the errors in the form'; r['HX-Alert-Type'] = 'error'; return r
            messages.error(request, 'Please correct the errors in the form')
    else:
        form = AcademicLevelForm()
    return render(request, 'academics/levels/form.html', {'form': form, 'title': 'Create academic level'})


@login_required
def academic_level_edit(request, pk):
    level = get_object_or_404(AcademicLevel, pk=pk)
    if request.method == 'POST':
        form = AcademicLevelForm(request.POST, instance=level)
        if form.is_valid():
            try:
                level = form.save()
                if request.headers.get('HX-Request') == 'true':
                    r = HttpResponse(); r['HX-Alert-Message'] = 'Level updated successfully'; r['HX-Alert-Type'] = 'success'; r['HX-Close-Modal'] = 'true'; r['HX-Redirect'] = reverse('academics:level_detail', kwargs={'pk': level.pk}); return r
                messages.success(request, 'Level updated successfully')
                return redirect('academics:level_detail', pk=level.pk)
            except Exception as e:
                logger.error(f"Error updating level: {e}")
                if request.headers.get('HX-Request') == 'true':
                    r = HttpResponse(); r['HX-Alert-Message'] = f'Error: {e}'; r['HX-Alert-Type'] = 'error'; return r
                messages.error(request, f'Error: {e}')
        else:
            if request.headers.get('HX-Request') == 'true':
                r = HttpResponse(); r['HX-Alert-Message'] = 'Please correct the errors in the form'; r['HX-Alert-Type'] = 'error'; return r
            messages.error(request, 'Please correct the errors in the form')
    else:
        form = AcademicLevelForm(instance=level)
    return render(request, 'academics/levels/form.html', {'form': form, 'level': level, 'title': f'Edit {level.name}'})


@login_required
def academic_level_delete(request, pk):
    level = get_object_or_404(AcademicLevel, pk=pk)
    if request.method == 'POST':
        if level.classes.filter(is_active=True).exists():
            if request.headers.get('HX-Request') == 'true':
                r = HttpResponse(); r['HX-Alert-Message'] = 'Cannot delete level with active classes'; r['HX-Alert-Type'] = 'error'; r['HX-Close-Modal'] = 'true'; return r
            messages.error(request, 'Cannot delete level with active classes')
            return redirect('academics:level_detail', pk=pk)
        if Student.objects.filter(current_academic_level=level).exists():
            if request.headers.get('HX-Request') == 'true':
                r = HttpResponse(); r['HX-Alert-Message'] = 'Cannot delete level with assigned students'; r['HX-Alert-Type'] = 'error'; r['HX-Close-Modal'] = 'true'; return r
            messages.error(request, 'Cannot delete level with assigned students')
            return redirect('academics:level_detail', pk=pk)
        try:
            name = level.name; level.delete()
            if request.headers.get('HX-Request') == 'true':
                r = HttpResponse(); r['HX-Alert-Message'] = f'Level "{name}" deleted successfully'; r['HX-Alert-Type'] = 'success'; r['HX-Close-Modal'] = 'true'; r['HX-Redirect'] = reverse('academics:level_list'); return r
            messages.success(request, f'Level "{name}" deleted successfully')
            return redirect('academics:level_list')
        except Exception as e:
            logger.error(f"Error deleting level: {e}")
            if request.headers.get('HX-Request') == 'true':
                r = HttpResponse(); r['HX-Alert-Message'] = f'Error: {e}'; r['HX-Alert-Type'] = 'error'; r['HX-Close-Modal'] = 'true'; return r
            messages.error(request, f'Error: {e}')
            return redirect('academics:level_detail', pk=pk)
    return redirect('academics:level_list')


@login_required
def academic_level_toggle_active(request, pk):
    level = get_object_or_404(AcademicLevel, pk=pk)
    if request.method == 'POST':
        try:
            level.is_active = not level.is_active; level.save(update_fields=['is_active'])
            word = 'activated' if level.is_active else 'deactivated'
            if request.headers.get('HX-Request') == 'true':
                r = HttpResponse(); r['HX-Alert-Message'] = f'Level {word} successfully'; r['HX-Alert-Type'] = 'success'; r['HX-Close-Modal'] = 'true'; r['HX-Redirect'] = reverse('academics:level_detail', kwargs={'pk': pk}); return r
            messages.success(request, f'Level {word}')
        except Exception as e:
            if request.headers.get('HX-Request') == 'true':
                r = HttpResponse(); r['HX-Alert-Message'] = f'Error: {e}'; r['HX-Alert-Type'] = 'error'; r['HX-Close-Modal'] = 'true'; return r
            messages.error(request, f'Error: {e}')
    return redirect('academics:level_detail', pk=pk)


@login_required
def academic_level_print_view(request):
    levels          = get_filtered_academic_levels(request)
    selected_fields = request.GET.getlist('fields') or ['name', 'code', 'order', 'has_sections', 'is_graduation_level', 'is_active']
    include_stats   = request.GET.get('include_stats') == 'true'
    stats = None
    if include_stats:
        stats = levels.aggregate(total=Count('id'), active=Count('id', filter=Q(is_active=True)), with_sections=Count('id', filter=Q(has_sections=True)), graduation_levels=Count('id', filter=Q(is_graduation_level=True)))
    field_labels = {
        'name': 'Level Name', 'code': 'Code', 'order': 'Order',
        'has_sections': 'Has Sections', 'is_graduation_level': 'Graduation Level',
        'is_active': 'Active', 'next_level': 'Next Level', 'class_count': 'Classes', 'description': 'Description',
    }
    return render(request, 'academics/levels/print.html', {
        **_get_print_school_context(request),
        'levels': levels, 'stats': stats, 'now': timezone.now(),
        'selected_fields': selected_fields,
        'selected_field_names': [field_labels.get(f, f.replace('_', ' ').title()) for f in selected_fields],
        'field_labels': field_labels, 'landscape': request.GET.get('landscape') == 'true',
        'title': 'Academic Levels Report',
    })


@login_required
def export_academic_levels_excel(request):
    ALL_COLUMNS = [
        ('name',          'Level Name',       lambda l: l.name),
        ('code',          'Code',             lambda l: l.code),
        ('order',         'Order',            lambda l: l.order),
        ('has_sections',  'Has Sections',     lambda l: 'Yes' if l.has_sections else 'No'),
        ('is_graduation', 'Graduation Level', lambda l: 'Yes' if l.is_graduation_level else 'No'),
        ('next_level',    'Next Level',       lambda l: l.next_level.name if l.next_level else ''),
        ('is_active',     'Active',           lambda l: 'Yes' if l.is_active else 'No'),
        ('class_count',   'Classes',          lambda l: getattr(l, 'class_count', '')),
        ('description',   'Description',      lambda l: l.description or ''),
    ]
    DEFAULT = ['name', 'code', 'order', 'has_sections', 'is_graduation', 'is_active', 'class_count']
    col_map = {k: (lbl, fn) for k, lbl, fn in ALL_COLUMNS}
    chosen  = request.GET.getlist('fields') or DEFAULT
    columns = [(col_map[f][0], col_map[f][1]) for f in chosen if f in col_map]
    levels  = get_filtered_academic_levels(request)
    wb = Workbook(); ws = wb.active; ws.title = 'Academic Levels'
    ws.append([c[0] for c in columns])
    hf = PatternFill(start_color='1E3A5F', end_color='1E3A5F', fill_type='solid')
    for cell in ws[1]:
        cell.fill = hf; cell.font = Font(bold=True, color='FFFFFF', size=11)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws.row_dimensions[1].height = 28
    af = PatternFill(start_color='F4F6F9', end_color='F4F6F9', fill_type='solid')
    da = Alignment(vertical='center', wrap_text=False)
    for i, obj in enumerate(levels):
        ws.append([c[1](obj) for c in columns])
        for cell in ws[i + 2]:
            cell.alignment = da
            if i % 2 == 1: cell.fill = af
    for cc in ws.columns:
        ml = max((len(str(c.value)) if c.value else 0) for c in cc)
        ws.column_dimensions[cc[0].column_letter].width = min(ml + 4, 60)
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="academic_levels_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
    wb.save(response); return response


# =============================================================================
# 5. SUBJECTS
# =============================================================================

def get_filtered_subjects(request):
    qs = Subject.objects.select_related('department').prefetch_related('applicable_levels', 'prerequisites').order_by('subject_type', 'abbreviation')
    q = request.GET.get('q', '').strip()
    if q:
        qs = qs.filter(Q(name__icontains=q) | Q(abbreviation__icontains=q) | Q(code__icontains=q) | Q(description__icontains=q))
    for field in ('subject_type', 'difficulty_level'):
        v = request.GET.get(field, '')
        if v: qs = qs.filter(**{field: v})
    for bf in ('is_active', 'is_compulsory', 'textbook_required'):
        v = request.GET.get(bf, '')
        if v: qs = qs.filter(**{bf: v.lower() == 'true'})
    dept = request.GET.get('department', '')
    if dept: qs = qs.filter(department_id=dept)
    return qs


@login_required
def subject_list(request):
    filter_form = SubjectFilterForm(request.GET or None)
    subjects    = get_filtered_subjects(request)
    stats = {'total': subjects.count(), 'active': subjects.filter(is_active=True).count(), 'compulsory': subjects.filter(is_compulsory=True).count(), 'optional': subjects.filter(is_compulsory=False).count()}
    paginator     = Paginator(subjects, 10)
    subjects_page = paginator.get_page(request.GET.get('page', 1))
    is_htmx       = request.headers.get('HX-Request') == 'true'
    context = {'subjects_page': subjects_page, 'paginator': paginator, 'stats': stats, 'filter_form': filter_form, 'is_htmx': is_htmx}
    if is_htmx:
        return render(request, 'academics/subjects/partials/_subject_results.html', context)
    return render(request, 'academics/subjects/list.html', context)


@login_required
def subject_detail(request, pk):
    subject = get_object_or_404(Subject, pk=pk)
    return render(request, 'academics/subjects/detail.html', {
        'subject': subject,
        'class_assignments': subject.classes.select_related('class_instance', 'teacher').filter(is_active=True)[:10],
        'applicable_levels': subject.applicable_levels.all(),
        'prerequisites':     subject.prerequisites.all(),
    })


@login_required
def subject_create(request):
    if request.method == 'POST':
        form = SubjectForm(request.POST)
        if form.is_valid():
            try:
                subject = form.save()
                if request.headers.get('HX-Request') == 'true':
                    r = HttpResponse(); r['HX-Alert-Message'] = f'Subject "{subject.name}" created successfully'; r['HX-Alert-Type'] = 'success'; r['HX-Close-Modal'] = 'true'; r['HX-Redirect'] = reverse('academics:subject_detail', kwargs={'pk': subject.pk}); return r
                messages.success(request, f'Subject "{subject.name}" created successfully')
                return redirect('academics:subject_detail', pk=subject.pk)
            except Exception as e:
                if request.headers.get('HX-Request') == 'true':
                    r = HttpResponse(); r['HX-Alert-Message'] = f'Error: {e}'; r['HX-Alert-Type'] = 'error'; return r
                messages.error(request, f'Error: {e}')
        else:
            if request.headers.get('HX-Request') == 'true':
                r = HttpResponse(); r['HX-Alert-Message'] = 'Please correct the errors in the form'; r['HX-Alert-Type'] = 'error'; return r
            messages.error(request, 'Please correct the errors in the form')
    else:
        form = SubjectForm()
    return render(request, 'academics/subjects/form.html', {'form': form, 'title': 'Create subject'})


@login_required
def subject_edit(request, pk):
    subject = get_object_or_404(Subject, pk=pk)
    if request.method == 'POST':
        form = SubjectForm(request.POST, instance=subject)
        if form.is_valid():
            try:
                subject = form.save()
                if request.headers.get('HX-Request') == 'true':
                    r = HttpResponse(); r['HX-Alert-Message'] = 'Subject updated successfully'; r['HX-Alert-Type'] = 'success'; r['HX-Close-Modal'] = 'true'; r['HX-Redirect'] = reverse('academics:subject_detail', kwargs={'pk': subject.pk}); return r
                messages.success(request, 'Subject updated successfully')
                return redirect('academics:subject_detail', pk=subject.pk)
            except Exception as e:
                if request.headers.get('HX-Request') == 'true':
                    r = HttpResponse(); r['HX-Alert-Message'] = f'Error: {e}'; r['HX-Alert-Type'] = 'error'; return r
                messages.error(request, f'Error: {e}')
        else:
            if request.headers.get('HX-Request') == 'true':
                r = HttpResponse(); r['HX-Alert-Message'] = 'Please correct the errors in the form'; r['HX-Alert-Type'] = 'error'; return r
            messages.error(request, 'Please correct the errors in the form')
    else:
        form = SubjectForm(instance=subject)
    return render(request, 'academics/subjects/form.html', {'form': form, 'subject': subject, 'title': f'Edit {subject.name}'})


@login_required
def subject_delete(request, pk):
    subject = get_object_or_404(Subject, pk=pk)
    if request.method == 'POST':
        if subject.classes.filter(is_active=True).exists():
            if request.headers.get('HX-Request') == 'true':
                r = HttpResponse(); r['HX-Alert-Message'] = 'Cannot delete subject assigned to active classes'; r['HX-Alert-Type'] = 'error'; r['HX-Close-Modal'] = 'true'; return r
            messages.error(request, 'Cannot delete subject assigned to active classes')
            return redirect('academics:subject_detail', pk=pk)
        try:
            name = subject.name; subject.delete()
            if request.headers.get('HX-Request') == 'true':
                r = HttpResponse(); r['HX-Alert-Message'] = f'Subject "{name}" deleted successfully'; r['HX-Alert-Type'] = 'success'; r['HX-Close-Modal'] = 'true'; r['HX-Redirect'] = reverse('academics:subject_list'); return r
            messages.success(request, f'Subject "{name}" deleted successfully')
            return redirect('academics:subject_list')
        except Exception as e:
            if request.headers.get('HX-Request') == 'true':
                r = HttpResponse(); r['HX-Alert-Message'] = f'Error: {e}'; r['HX-Alert-Type'] = 'error'; r['HX-Close-Modal'] = 'true'; return r
            messages.error(request, f'Error: {e}')
            return redirect('academics:subject_detail', pk=pk)
    return redirect('academics:subject_list')


@login_required
def subject_toggle_active(request, pk):
    subject = get_object_or_404(Subject, pk=pk)
    if request.method == 'POST':
        try:
            subject.is_active = not subject.is_active; subject.save(update_fields=['is_active'])
            word = 'activated' if subject.is_active else 'deactivated'
            if request.headers.get('HX-Request') == 'true':
                r = HttpResponse(); r['HX-Alert-Message'] = f'Subject {word} successfully'; r['HX-Alert-Type'] = 'success'; r['HX-Close-Modal'] = 'true'; r['HX-Redirect'] = reverse('academics:subject_detail', kwargs={'pk': pk}); return r
            messages.success(request, f'Subject {word}')
        except Exception as e:
            if request.headers.get('HX-Request') == 'true':
                r = HttpResponse(); r['HX-Alert-Message'] = f'Error: {e}'; r['HX-Alert-Type'] = 'error'; r['HX-Close-Modal'] = 'true'; return r
            messages.error(request, f'Error: {e}')
    return redirect('academics:subject_detail', pk=pk)


@login_required
def subject_print_view(request):
    subjects        = get_filtered_subjects(request)
    selected_fields = request.GET.getlist('fields') or ['name', 'code', 'abbreviation', 'subject_type', 'credit_hours', 'is_compulsory', 'is_active']
    include_stats   = request.GET.get('include_stats') == 'true'
    stats = None
    if include_stats:
        stats = {'total': subjects.count(), 'active': subjects.filter(is_active=True).count()}
    field_labels = {
        'name': 'Subject Name', 'code': 'Code', 'abbreviation': 'Abbreviation',
        'subject_type': 'Type', 'credit_hours': 'Credit Hours', 'pass_mark': 'Pass Mark',
        'difficulty': 'Difficulty', 'weight_factor': 'Weight', 'is_compulsory': 'Compulsory',
        'is_active': 'Active', 'department': 'Department', 'textbook_req': 'Textbook Req.',
    }
    return render(request, 'academics/subjects/print.html', {
        **_get_print_school_context(request),
        'subjects': subjects, 'stats': stats, 'now': timezone.now(),
        'selected_fields': selected_fields,
        'selected_field_names': [field_labels.get(f, f.replace('_', ' ').title()) for f in selected_fields],
        'field_labels': field_labels, 'landscape': request.GET.get('landscape') == 'true',
        'title': 'Subjects Report',
    })


@login_required
def export_subjects_excel(request):
    ALL_COLUMNS = [
        ('name',          'Subject Name',   lambda s: s.name),
        ('code',          'Code',           lambda s: s.code),
        ('abbreviation',  'Abbreviation',   lambda s: s.abbreviation or ''),
        ('subject_type',  'Type',           lambda s: s.get_subject_type_display()),
        ('credit_hours',  'Credit Hours',   lambda s: float(s.credit_hours)),
        ('pass_mark',     'Pass Mark',      lambda s: float(s.pass_mark)),
        ('difficulty',    'Difficulty',     lambda s: s.get_difficulty_level_display()),
        ('weight_factor', 'Weight Factor',  lambda s: float(s.weight_factor)),
        ('is_compulsory', 'Compulsory',     lambda s: 'Yes' if s.is_compulsory else 'No'),
        ('is_active',     'Active',         lambda s: 'Yes' if s.is_active else 'No'),
        ('department',    'Department',     lambda s: s.department.name if s.department else ''),
        ('textbook_req',  'Textbook Req.',  lambda s: 'Yes' if s.textbook_required else 'No'),
        ('description',   'Description',   lambda s: s.description or ''),
    ]
    DEFAULT = ['name', 'code', 'abbreviation', 'subject_type', 'credit_hours', 'is_compulsory', 'is_active']
    col_map  = {k: (lbl, fn) for k, lbl, fn in ALL_COLUMNS}
    chosen   = request.GET.getlist('fields') or DEFAULT
    columns  = [(col_map[f][0], col_map[f][1]) for f in chosen if f in col_map]
    subjects = get_filtered_subjects(request)
    wb = Workbook(); ws = wb.active; ws.title = 'Subjects'
    ws.append([c[0] for c in columns])
    hf = PatternFill(start_color='1E3A5F', end_color='1E3A5F', fill_type='solid')
    for cell in ws[1]:
        cell.fill = hf; cell.font = Font(bold=True, color='FFFFFF', size=11)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws.row_dimensions[1].height = 28
    af = PatternFill(start_color='F4F6F9', end_color='F4F6F9', fill_type='solid')
    da = Alignment(vertical='center', wrap_text=False)
    for i, obj in enumerate(subjects):
        ws.append([c[1](obj) for c in columns])
        for cell in ws[i + 2]:
            cell.alignment = da
            if i % 2 == 1: cell.fill = af
    for cc in ws.columns:
        ml = max((len(str(c.value)) if c.value else 0) for c in cc)
        ws.column_dimensions[cc[0].column_letter].width = min(ml + 4, 60)
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="subjects_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
    wb.save(response); return response


# =============================================================================
# 6. CLASSROOMS
# =============================================================================

def get_filtered_classrooms(request):
    qs = ClassRoom.objects.annotate(assigned_class_count=Count('assigned_classes', distinct=True)).order_by('building', 'floor', 'room_number')
    q = request.GET.get('q', '').strip()
    if q:
        qs = qs.filter(Q(name__icontains=q) | Q(room_number__icontains=q) | Q(building__icontains=q) | Q(specialized_equipment__icontains=q))
    rt = request.GET.get('room_type', '')
    if rt: qs = qs.filter(room_type=rt)
    for field in ('building', 'floor'):
        v = request.GET.get(field, '')
        if v: qs = qs.filter(**{f'{field}__icontains': v})
    for bf in ('is_active', 'is_bookable', 'has_projector', 'has_computer', 'has_smart_board', 'is_accessible'):
        v = request.GET.get(bf, '')
        if v: qs = qs.filter(**{bf: v.lower() == 'true'})
    mc = request.GET.get('min_capacity', '')
    if mc:
        try: qs = qs.filter(capacity__gte=int(mc))
        except (ValueError, TypeError): pass
    return qs


@login_required
def classroom_list(request):
    filter_form = ClassRoomFilterForm(request.GET or None)
    classrooms  = get_filtered_classrooms(request)
    stats = {
        'total': classrooms.count(), 'active': classrooms.filter(is_active=True).count(),
        'regular': classrooms.filter(room_type='REGULAR').count(),
        'labs': classrooms.filter(room_type__in=['LABORATORY', 'COMPUTER_LAB', 'SCIENCE_LAB']).count(),
        'total_capacity': classrooms.aggregate(Sum('capacity'))['capacity__sum'] or 0,
    }
    paginator       = Paginator(classrooms, 10)
    classrooms_page = paginator.get_page(request.GET.get('page', 1))
    is_htmx         = request.headers.get('HX-Request') == 'true'
    context = {'classrooms_page': classrooms_page, 'paginator': paginator, 'stats': stats, 'filter_form': filter_form, 'is_htmx': is_htmx}
    if is_htmx:
        return render(request, 'academics/classrooms/partials/_classroom_results.html', context)
    return render(request, 'academics/classrooms/list.html', context)


@login_required
def classroom_detail(request, pk):
    classroom       = get_object_or_404(ClassRoom, pk=pk)
    current_classes = classroom.assigned_classes.select_related('academic_level', 'academic_session', 'class_teacher').filter(is_active=True)
    current_students = sum(cls.get_current_enrollment_count() for cls in current_classes)
    utilization      = round((current_students / classroom.capacity * 100), 1) if classroom.capacity else 0
    return render(request, 'academics/classrooms/detail.html', {
        'classroom': classroom, 'current_classes': current_classes,
        'stats': {'current_classes_count': current_classes.count(), 'current_students': current_students, 'utilization_percentage': utilization, 'available_capacity': max(0, classroom.capacity - current_students)},
    })


@login_required
def classroom_create(request):
    if request.method == 'POST':
        form = ClassRoomForm(request.POST)
        if form.is_valid():
            try:
                classroom = form.save()
                if request.headers.get('HX-Request') == 'true':
                    r = HttpResponse(); r['HX-Alert-Message'] = f'Classroom "{classroom.name}" created successfully'; r['HX-Alert-Type'] = 'success'; r['HX-Close-Modal'] = 'true'; r['HX-Redirect'] = reverse('academics:classroom_detail', kwargs={'pk': classroom.pk}); return r
                messages.success(request, f'Classroom "{classroom.name}" created successfully')
                return redirect('academics:classroom_detail', pk=classroom.pk)
            except Exception as e:
                if request.headers.get('HX-Request') == 'true':
                    r = HttpResponse(); r['HX-Alert-Message'] = f'Error: {e}'; r['HX-Alert-Type'] = 'error'; return r
                messages.error(request, f'Error: {e}')
        else:
            if request.headers.get('HX-Request') == 'true':
                r = HttpResponse(); r['HX-Alert-Message'] = 'Please correct the errors in the form'; r['HX-Alert-Type'] = 'error'; return r
            messages.error(request, 'Please correct the errors in the form')
    else:
        form = ClassRoomForm()
    return render(request, 'academics/classrooms/form.html', {'form': form, 'title': 'Create classroom'})


@login_required
def classroom_edit(request, pk):
    classroom = get_object_or_404(ClassRoom, pk=pk)
    if request.method == 'POST':
        form = ClassRoomForm(request.POST, instance=classroom)
        if form.is_valid():
            try:
                classroom = form.save()
                if request.headers.get('HX-Request') == 'true':
                    r = HttpResponse(); r['HX-Alert-Message'] = f'Classroom "{classroom.name}" updated successfully'; r['HX-Alert-Type'] = 'success'; r['HX-Close-Modal'] = 'true'; r['HX-Redirect'] = reverse('academics:classroom_detail', kwargs={'pk': classroom.pk}); return r
                messages.success(request, 'Classroom updated successfully')
                return redirect('academics:classroom_detail', pk=classroom.pk)
            except Exception as e:
                if request.headers.get('HX-Request') == 'true':
                    r = HttpResponse(); r['HX-Alert-Message'] = f'Error: {e}'; r['HX-Alert-Type'] = 'error'; return r
                messages.error(request, f'Error: {e}')
        else:
            if request.headers.get('HX-Request') == 'true':
                r = HttpResponse(); r['HX-Alert-Message'] = 'Please correct the errors in the form'; r['HX-Alert-Type'] = 'error'; return r
            messages.error(request, 'Please correct the errors in the form')
    else:
        form = ClassRoomForm(instance=classroom)
    return render(request, 'academics/classrooms/form.html', {'form': form, 'classroom': classroom, 'title': f'Edit {classroom.name}'})


@login_required
def classroom_delete(request, pk):
    classroom = get_object_or_404(ClassRoom, pk=pk)
    if request.method == 'POST':
        if classroom.assigned_classes.filter(is_active=True).exists():
            if request.headers.get('HX-Request') == 'true':
                r = HttpResponse(); r['HX-Alert-Message'] = 'Cannot delete classroom assigned to active classes'; r['HX-Alert-Type'] = 'error'; r['HX-Close-Modal'] = 'true'; return r
            messages.error(request, 'Cannot delete classroom assigned to active classes')
            return redirect('academics:classroom_detail', pk=pk)
        try:
            name = classroom.name; classroom.delete()
            if request.headers.get('HX-Request') == 'true':
                r = HttpResponse(); r['HX-Alert-Message'] = f'Classroom "{name}" deleted successfully'; r['HX-Alert-Type'] = 'success'; r['HX-Close-Modal'] = 'true'; r['HX-Redirect'] = reverse('academics:classroom_list'); return r
            messages.success(request, f'Classroom "{name}" deleted successfully')
            return redirect('academics:classroom_list')
        except Exception as e:
            if request.headers.get('HX-Request') == 'true':
                r = HttpResponse(); r['HX-Alert-Message'] = f'Error: {e}'; r['HX-Alert-Type'] = 'error'; r['HX-Close-Modal'] = 'true'; return r
            messages.error(request, f'Error: {e}')
            return redirect('academics:classroom_detail', pk=pk)
    return redirect('academics:classroom_list')


@login_required
def classroom_toggle_active(request, pk):
    classroom = get_object_or_404(ClassRoom, pk=pk)
    if request.method == 'POST':
        try:
            classroom.is_active = not classroom.is_active; classroom.save(update_fields=['is_active'])
            word = 'activated' if classroom.is_active else 'deactivated'
            if request.headers.get('HX-Request') == 'true':
                r = HttpResponse(); r['HX-Alert-Message'] = f'Classroom {word} successfully'; r['HX-Alert-Type'] = 'success'; r['HX-Close-Modal'] = 'true'; r['HX-Redirect'] = reverse('academics:classroom_detail', kwargs={'pk': pk}); return r
            messages.success(request, f'Classroom {word}')
        except Exception as e:
            if request.headers.get('HX-Request') == 'true':
                r = HttpResponse(); r['HX-Alert-Message'] = f'Error: {e}'; r['HX-Alert-Type'] = 'error'; r['HX-Close-Modal'] = 'true'; return r
            messages.error(request, f'Error: {e}')
    return redirect('academics:classroom_detail', pk=pk)


@login_required
def classroom_toggle_bookable(request, pk):
    classroom = get_object_or_404(ClassRoom, pk=pk)
    if request.method == 'POST':
        try:
            classroom.is_bookable = not classroom.is_bookable; classroom.save(update_fields=['is_bookable'])
            word = 'bookable' if classroom.is_bookable else 'not bookable'
            if request.headers.get('HX-Request') == 'true':
                r = HttpResponse(); r['HX-Alert-Message'] = f'Classroom is now {word}'; r['HX-Alert-Type'] = 'success'; r['HX-Close-Modal'] = 'true'; r['HX-Redirect'] = reverse('academics:classroom_detail', kwargs={'pk': pk}); return r
            messages.success(request, f'Classroom is now {word}')
        except Exception as e:
            if request.headers.get('HX-Request') == 'true':
                r = HttpResponse(); r['HX-Alert-Message'] = f'Error: {e}'; r['HX-Alert-Type'] = 'error'; r['HX-Close-Modal'] = 'true'; return r
            messages.error(request, f'Error: {e}')
    return redirect('academics:classroom_detail', pk=pk)


@login_required
def classroom_print_view(request):
    classrooms      = get_filtered_classrooms(request)
    selected_fields = request.GET.getlist('fields') or ['room_number', 'name', 'building', 'room_type', 'capacity', 'is_active']
    include_stats   = request.GET.get('include_stats') == 'true'
    stats = None
    if include_stats:
        stats = classrooms.aggregate(total=Count('id'), active=Count('id', filter=Q(is_active=True)), total_capacity=Sum('capacity'), avg_capacity=Avg('capacity'))
    field_labels = {
        'room_number': 'Room Number', 'name': 'Room Name', 'building': 'Building', 'floor': 'Floor',
        'room_type': 'Room Type', 'capacity': 'Capacity', 'is_active': 'Active',
        'has_projector': 'Projector', 'has_computer': 'Computer', 'is_bookable': 'Bookable',
        'assigned_classes': 'Classes Assigned', 'last_maintenance': 'Last Maintenance',
    }
    return render(request, 'academics/classrooms/print.html', {
        **_get_print_school_context(request),
        'classrooms': classrooms, 'stats': stats, 'now': timezone.now(),
        'selected_fields': selected_fields,
        'selected_field_names': [field_labels.get(f, f.replace('_', ' ').title()) for f in selected_fields],
        'field_labels': field_labels, 'landscape': request.GET.get('landscape') == 'true',
        'title': 'Classrooms Report',
    })


@login_required
def export_classrooms_excel(request):
    ALL_COLUMNS = [
        ('room_number',      'Room Number',      lambda r: r.room_number),
        ('name',             'Room Name',        lambda r: r.name),
        ('building',         'Building',         lambda r: r.building or ''),
        ('floor',            'Floor',            lambda r: r.floor or ''),
        ('wing',             'Wing',             lambda r: r.wing or ''),
        ('room_type',        'Room Type',        lambda r: r.get_room_type_display()),
        ('capacity',         'Capacity',         lambda r: r.capacity),
        ('is_active',        'Active',           lambda r: 'Yes' if r.is_active else 'No'),
        ('is_bookable',      'Bookable',         lambda r: 'Yes' if r.is_bookable else 'No'),
        ('has_projector',    'Projector',        lambda r: 'Yes' if r.has_projector else 'No'),
        ('has_computer',     'Computer',         lambda r: 'Yes' if r.has_computer else 'No'),
        ('has_smart_board',  'Smart Board',      lambda r: 'Yes' if r.has_smart_board else 'No'),
        ('has_internet',     'Internet',         lambda r: 'Yes' if r.has_internet else 'No'),
        ('is_accessible',    'Accessible',       lambda r: 'Yes' if r.is_accessible else 'No'),
        ('assigned_classes', 'Classes Assigned', lambda r: getattr(r, 'assigned_class_count', '')),
        ('last_maintenance', 'Last Maintenance', lambda r: r.last_maintenance_date.strftime('%Y-%m-%d') if r.last_maintenance_date else ''),
    ]
    DEFAULT = ['room_number', 'name', 'building', 'room_type', 'capacity', 'is_active']
    col_map    = {k: (lbl, fn) for k, lbl, fn in ALL_COLUMNS}
    chosen     = request.GET.getlist('fields') or DEFAULT
    columns    = [(col_map[f][0], col_map[f][1]) for f in chosen if f in col_map]
    classrooms = get_filtered_classrooms(request)
    wb = Workbook(); ws = wb.active; ws.title = 'Classrooms'
    ws.append([c[0] for c in columns])
    hf = PatternFill(start_color='1E3A5F', end_color='1E3A5F', fill_type='solid')
    for cell in ws[1]:
        cell.fill = hf; cell.font = Font(bold=True, color='FFFFFF', size=11)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws.row_dimensions[1].height = 28
    af = PatternFill(start_color='F4F6F9', end_color='F4F6F9', fill_type='solid')
    da = Alignment(vertical='center', wrap_text=False)
    for i, obj in enumerate(classrooms):
        ws.append([c[1](obj) for c in columns])
        for cell in ws[i + 2]:
            cell.alignment = da
            if i % 2 == 1: cell.fill = af
    for cc in ws.columns:
        ml = max((len(str(c.value)) if c.value else 0) for c in cc)
        ws.column_dimensions[cc[0].column_letter].width = min(ml + 4, 60)
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="classrooms_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
    wb.save(response); return response


# =============================================================================
# 7. CLASSES  (class-subject management embedded here)
# =============================================================================

def get_filtered_classes(request):
    qs = Class.objects.select_related('academic_level', 'academic_session', 'class_teacher', 'classroom').annotate(enrollment_count=Count('enrollments', filter=Q(enrollments__is_active=True))).order_by('-academic_session__start_date', 'academic_level__order', 'section')
    q = request.GET.get('q', '').strip()
    if q:
        qs = qs.filter(Q(academic_level__name__icontains=q) | Q(section__icontains=q) | Q(class_motto__icontains=q))
    for field, lookup in (('academic_level', 'academic_level_id'), ('academic_session', 'academic_session_id'), ('class_teacher', 'class_teacher_id')):
        v = request.GET.get(field, '')
        if v: qs = qs.filter(**{lookup: v})
    section = request.GET.get('section', '')
    if section: qs = qs.filter(section__iexact=section)
    ia = request.GET.get('is_active', '')
    if ia: qs = qs.filter(is_active=(ia.lower() == 'true'))
    hc = request.GET.get('has_capacity', '')
    if hc and hc.lower() == 'true': qs = qs.filter(enrollment_count__lt=F('max_students'))
    return qs

@login_required
def class_detail(request, pk):
    class_instance = get_object_or_404(
        Class.objects.select_related(
            'academic_level', 'academic_session',
            'class_teacher__staff', 'assistant_teacher__staff', 'classroom'
        ),
        pk=pk
    )

    all_subjects    = class_instance.subjects.filter(is_active=True)
    all_enrollments = class_instance.enrollments.all()
    total_students  = all_enrollments.count()

    stats = {
        'total_students':           total_students,
        'male_students':            all_enrollments.filter(student__gender='M').count(),
        'female_students':          all_enrollments.filter(student__gender='F').count(),
        'available_spots':          class_instance.max_students - total_students,
        'total_subjects':           all_subjects.count(),
        'compulsory_subjects':      all_subjects.filter(is_optional=False).count(),
        'optional_subjects':        all_subjects.filter(is_optional=True).count(),
        'subjects_with_teacher':    all_subjects.filter(teacher__isnull=False).count(),
        'subjects_without_teacher': all_subjects.filter(teacher__isnull=True).count(),
    }

    return render(request, 'academics/classes/detail.html', {
        'class': class_instance,
        'stats': stats,
    })


@login_required
def class_subjects_partial(request, pk):
    class_instance = get_object_or_404(Class, pk=pk)

    qs = class_instance.subjects.select_related(
        'subject', 'teacher__staff'
    ).filter(is_active=True)

    q           = request.GET.get('q', '').strip()
    is_optional = request.GET.get('is_optional', '')
    has_teacher = request.GET.get('has_teacher', '')

    if q:
        qs = qs.filter(
            Q(subject__name__icontains=q) | Q(subject__code__icontains=q)
        )
    if is_optional == 'true':    qs = qs.filter(is_optional=True)
    elif is_optional == 'false': qs = qs.filter(is_optional=False)
    if has_teacher == 'true':    qs = qs.filter(teacher__isnull=False)
    elif has_teacher == 'false': qs = qs.filter(teacher__isnull=True)

    paginator = Paginator(qs.order_by('subject__name'), 15)
    page      = paginator.get_page(request.GET.get('page', 1))

    return render(request, 'academics/classes/partials/_subject_results.html', {
        'subjects':       page,
        'class_instance': class_instance,
    })


@login_required
def class_enrollments_partial(request, pk):
    class_instance = get_object_or_404(Class, pk=pk)

    qs = class_instance.enrollments.select_related(
        'student', 'academic_invoice'
    )

    q                 = request.GET.get('q', '').strip()
    gender            = request.GET.get('gender', '')
    completion_status = request.GET.get('completion_status', '')
    has_invoice       = request.GET.get('has_invoice', '')

    if q:
        qs = qs.filter(
            Q(student__first_name__icontains=q) |
            Q(student__last_name__icontains=q) |
            Q(student__admission_number__icontains=q) |
            Q(roll_number__icontains=q)
        )
    if gender:            qs = qs.filter(student__gender=gender)
    if completion_status: qs = qs.filter(completion_status=completion_status)
    if has_invoice == 'true':    qs = qs.filter(academic_invoice__isnull=False)
    elif has_invoice == 'false': qs = qs.filter(academic_invoice__isnull=True)

    paginator = Paginator(qs.order_by('roll_number', 'student__last_name'), 20)
    page      = paginator.get_page(request.GET.get('page', 1))

    # Count enrollments without invoices for the "Generate Missing" button
    missing_invoice_count = class_instance.enrollments.filter(
        academic_invoice__isnull=True,
        is_active=True,
        completion_status='ONGOING'
    ).count()

    return render(request, 'academics/classes/partials/_enrollment_results.html', {
        'enrollments':          page,
        'class_instance':       class_instance,
        'missing_invoice_count': missing_invoice_count,
    })


@login_required
def class_create(request):
    if request.method == 'POST':
        form = ClassForm(request.POST)
        if form.is_valid():
            try:
                with transaction.atomic():
                    ci = form.save()
                if request.headers.get('HX-Request') == 'true':
                    r = HttpResponse(); r['HX-Alert-Message'] = f'Class "{ci}" created successfully'; r['HX-Alert-Type'] = 'success'; r['HX-Close-Modal'] = 'true'; r['HX-Redirect'] = reverse('academics:class_detail', kwargs={'pk': ci.pk}); return r
                messages.success(request, 'Class created successfully')
                return redirect('academics:class_detail', pk=ci.pk)
            except Exception as e:
                logger.error(f"Error creating class: {e}")
                if request.headers.get('HX-Request') == 'true':
                    r = HttpResponse(); r['HX-Alert-Message'] = f'Error: {e}'; r['HX-Alert-Type'] = 'error'; return r
                messages.error(request, f'Error: {e}')
        else:
            if request.headers.get('HX-Request') == 'true':
                r = HttpResponse(); r['HX-Alert-Message'] = 'Please correct the errors in the form'; r['HX-Alert-Type'] = 'error'; return r
            messages.error(request, 'Please correct the errors in the form')
    else:
        form = ClassForm()
    return render(request, 'academics/classes/form.html', {'form': form, 'title': 'Create class'})


@login_required
def class_edit(request, pk):
    ci = get_object_or_404(Class, pk=pk)
    if request.method == 'POST':
        form = ClassForm(request.POST, instance=ci)
        if form.is_valid():
            try:
                with transaction.atomic():
                    ci = form.save()
                if request.headers.get('HX-Request') == 'true':
                    r = HttpResponse(); r['HX-Alert-Message'] = 'Class updated successfully'; r['HX-Alert-Type'] = 'success'; r['HX-Close-Modal'] = 'true'; r['HX-Redirect'] = reverse('academics:class_detail', kwargs={'pk': ci.pk}); return r
                messages.success(request, 'Class updated successfully')
                return redirect('academics:class_detail', pk=ci.pk)
            except Exception as e:
                if request.headers.get('HX-Request') == 'true':
                    r = HttpResponse(); r['HX-Alert-Message'] = f'Error: {e}'; r['HX-Alert-Type'] = 'error'; return r
                messages.error(request, f'Error: {e}')
        else:
            if request.headers.get('HX-Request') == 'true':
                r = HttpResponse(); r['HX-Alert-Message'] = 'Please correct the errors in the form'; r['HX-Alert-Type'] = 'error'; return r
            messages.error(request, 'Please correct the errors in the form')
    else:
        form = ClassForm(instance=ci)
    return render(request, 'academics/classes/form.html', {'form': form, 'class': ci, 'title': f'Edit {ci}'})


@login_required
def class_delete(request, pk):
    ci = get_object_or_404(Class, pk=pk)
    if request.method == 'POST':
        if ci.enrollments.exists():
            if request.headers.get('HX-Request') == 'true':
                r = HttpResponse(); r['HX-Alert-Message'] = 'Cannot delete class with existing enrollments'; r['HX-Alert-Type'] = 'error'; r['HX-Close-Modal'] = 'true'; return r
            messages.error(request, 'Cannot delete class with existing enrollments')
            return redirect('academics:class_detail', pk=pk)
        if ci.subjects.filter(is_active=True).exists():
            if request.headers.get('HX-Request') == 'true':
                r = HttpResponse(); r['HX-Alert-Message'] = 'Cannot delete class with active subject assignments'; r['HX-Alert-Type'] = 'error'; r['HX-Close-Modal'] = 'true'; return r
            messages.error(request, 'Cannot delete class with active subject assignments')
            return redirect('academics:class_detail', pk=pk)
        try:
            name = str(ci); ci.delete()
            if request.headers.get('HX-Request') == 'true':
                r = HttpResponse(); r['HX-Alert-Message'] = f'Class "{name}" deleted successfully'; r['HX-Alert-Type'] = 'success'; r['HX-Close-Modal'] = 'true'; r['HX-Redirect'] = reverse('academics:class_list'); return r
            messages.success(request, f'Class "{name}" deleted successfully')
            return redirect('academics:class_list')
        except Exception as e:
            if request.headers.get('HX-Request') == 'true':
                r = HttpResponse(); r['HX-Alert-Message'] = f'Error: {e}'; r['HX-Alert-Type'] = 'error'; r['HX-Close-Modal'] = 'true'; return r
            messages.error(request, f'Error: {e}')
            return redirect('academics:class_detail', pk=pk)
    return redirect('academics:class_list')


@login_required
def class_toggle_active(request, pk):
    ci = get_object_or_404(Class, pk=pk)
    if request.method == 'POST':
        try:
            ci.is_active = not ci.is_active; ci.save(update_fields=['is_active'])
            word = 'activated' if ci.is_active else 'deactivated'
            if request.headers.get('HX-Request') == 'true':
                r = HttpResponse(); r['HX-Alert-Message'] = f'Class {word} successfully'; r['HX-Alert-Type'] = 'success'; r['HX-Close-Modal'] = 'true'; r['HX-Redirect'] = reverse('academics:class_detail', kwargs={'pk': pk}); return r
            messages.success(request, f'Class {word}')
        except Exception as e:
            if request.headers.get('HX-Request') == 'true':
                r = HttpResponse(); r['HX-Alert-Message'] = f'Error: {e}'; r['HX-Alert-Type'] = 'error'; r['HX-Close-Modal'] = 'true'; return r
            messages.error(request, f'Error: {e}')
    return redirect('academics:class_detail', pk=pk)


@login_required
def class_assign_teacher(request, pk):
    ci = get_object_or_404(Class, pk=pk)
    if request.method == 'POST':
        try:
            from hr.models import Teacher
            tid = request.POST.get('teacher_id')
            if tid:
                teacher = get_object_or_404(Teacher, pk=tid)
                old     = ci.class_teacher
                ci.class_teacher = teacher; ci.save(update_fields=['class_teacher'])
                msg = f'Teacher {teacher.staff.full_name()} assigned'
                if old: msg += f' (replaced {old.staff.full_name()})'
            else:
                ci.class_teacher = None; ci.save(update_fields=['class_teacher']); msg = 'Teacher removed from class'
            if request.headers.get('HX-Request') == 'true':
                r = HttpResponse(); r['HX-Alert-Message'] = msg; r['HX-Alert-Type'] = 'success'; r['HX-Close-Modal'] = 'true'; r['HX-Redirect'] = reverse('academics:class_detail', kwargs={'pk': pk}); return r
            messages.success(request, msg)
        except Exception as e:
            if request.headers.get('HX-Request') == 'true':
                r = HttpResponse(); r['HX-Alert-Message'] = f'Error: {e}'; r['HX-Alert-Type'] = 'error'; r['HX-Close-Modal'] = 'true'; return r
            messages.error(request, f'Error: {e}')
    return redirect('academics:class_detail', pk=pk)


@login_required
def class_assign_classroom(request, pk):
    ci = get_object_or_404(Class, pk=pk)
    if request.method == 'POST':
        try:
            crid = request.POST.get('classroom_id')
            if crid:
                classroom = get_object_or_404(ClassRoom, pk=crid)
                old       = ci.classroom
                ci.classroom = classroom; ci.save(update_fields=['classroom'])
                msg = f'Classroom {classroom.name} assigned'
                if old: msg += f' (replaced {old.name})'
            else:
                ci.classroom = None; ci.save(update_fields=['classroom']); msg = 'Classroom removed from class'
            if request.headers.get('HX-Request') == 'true':
                r = HttpResponse(); r['HX-Alert-Message'] = msg; r['HX-Alert-Type'] = 'success'; r['HX-Close-Modal'] = 'true'; r['HX-Redirect'] = reverse('academics:class_detail', kwargs={'pk': pk}); return r
            messages.success(request, msg)
        except Exception as e:
            if request.headers.get('HX-Request') == 'true':
                r = HttpResponse(); r['HX-Alert-Message'] = f'Error: {e}'; r['HX-Alert-Type'] = 'error'; r['HX-Close-Modal'] = 'true'; return r
            messages.error(request, f'Error: {e}')
    return redirect('academics:class_detail', pk=pk)


@login_required
def class_print_view(request):
    classes         = get_filtered_classes(request)
    selected_fields = request.GET.getlist('fields') or ['level_name', 'section', 'session_name', 'class_teacher', 'max_students', 'enrolled', 'is_active']
    include_stats   = request.GET.get('include_stats') == 'true'
    stats = None
    if include_stats:
        stats = classes.aggregate(total=Count('id'), active=Count('id', filter=Q(is_active=True)), total_capacity=Sum('max_students'), with_teacher=Count('id', filter=Q(class_teacher__isnull=False)))
    field_labels = {
        'level_name': 'Academic Level', 'level_code': 'Code', 'section': 'Section',
        'display_name': 'Class Name', 'session_name': 'Session', 'session_year': 'Year',
        'class_teacher': 'Class Teacher', 'assistant_teacher': 'Asst. Teacher',
        'classroom': 'Classroom', 'max_students': 'Max Students',
        'enrolled': 'Enrolled', 'available': 'Available', 'subject_count': 'Subjects', 'is_active': 'Active',
    }
    return render(request, 'academics/classes/print.html', {
        **_get_print_school_context(request),
        'classes': classes, 'stats': stats, 'now': timezone.now(),
        'selected_fields': selected_fields,
        'selected_field_names': [field_labels.get(f, f.replace('_', ' ').title()) for f in selected_fields],
        'field_labels': field_labels, 'landscape': request.GET.get('landscape') == 'true',
        'title': 'Classes Report',
    })


@login_required
def export_classes_excel(request):
    ALL_COLUMNS = [
        ('level_name',        'Academic Level',   lambda c: c.academic_level.name),
        ('level_code',        'Level Code',       lambda c: c.academic_level.code),
        ('section',           'Section',          lambda c: c.section or ''),
        ('display_name',      'Class Name',       lambda c: c.get_display_name()),
        ('session_name',      'Session',          lambda c: c.academic_session.name),
        ('session_year',      'Year',             lambda c: c.academic_session.year_name),
        ('class_teacher',     'Class Teacher',    lambda c: c.class_teacher.staff.full_name() if c.class_teacher else ''),
        ('assistant_teacher', 'Asst. Teacher',    lambda c: c.assistant_teacher.staff.full_name() if c.assistant_teacher else ''),
        ('classroom',         'Classroom',        lambda c: c.classroom.room_number if c.classroom else ''),
        ('max_students',      'Max Students',     lambda c: c.max_students),
        ('enrolled',          'Enrolled',         lambda c: getattr(c, 'enrollment_count', c.get_current_enrollment_count())),
        ('available',         'Available Spots',  lambda c: max(0, c.max_students - getattr(c, 'enrollment_count', c.get_current_enrollment_count()))),
        ('subject_count',     'Subjects',         lambda c: c.subjects.filter(is_active=True).count()),
        ('is_active',         'Active',           lambda c: 'Yes' if c.is_active else 'No'),
    ]
    DEFAULT = ['level_name', 'section', 'session_name', 'class_teacher', 'max_students', 'enrolled', 'is_active']
    col_map = {k: (lbl, fn) for k, lbl, fn in ALL_COLUMNS}
    chosen  = request.GET.getlist('fields') or DEFAULT
    columns = [(col_map[f][0], col_map[f][1]) for f in chosen if f in col_map]
    classes = get_filtered_classes(request)
    wb = Workbook(); ws = wb.active; ws.title = 'Classes'
    ws.append([c[0] for c in columns])
    hf = PatternFill(start_color='1E3A5F', end_color='1E3A5F', fill_type='solid')
    for cell in ws[1]:
        cell.fill = hf; cell.font = Font(bold=True, color='FFFFFF', size=11)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws.row_dimensions[1].height = 28
    af = PatternFill(start_color='F4F6F9', end_color='F4F6F9', fill_type='solid')
    da = Alignment(vertical='center', wrap_text=False)
    for i, obj in enumerate(classes):
        ws.append([c[1](obj) for c in columns])
        for cell in ws[i + 2]:
            cell.alignment = da
            if i % 2 == 1: cell.fill = af
    for cc in ws.columns:
        ml = max((len(str(c.value)) if c.value else 0) for c in cc)
        ws.column_dimensions[cc[0].column_letter].width = min(ml + 4, 60)
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="classes_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
    wb.save(response); return response


# --- Class invoice generation ------------------------------------------------

@login_required
@require_http_methods(["GET"])
def class_generate_missing_invoices_modal(request, pk):
    """
    Confirmation modal showing how many enrollments are missing invoices.
    GET only — the actual generation is handled by class_generate_missing_invoices (POST).
    """
    class_instance = get_object_or_404(Class, pk=pk)
    count = class_instance.enrollments.filter(
        academic_invoice__isnull=True,
        is_active=True,
        completion_status='ONGOING'
    ).count()
    return render(request, 'academics/classes/modals/generate_missing_invoices.html', {
        'class_instance': class_instance,
        'count': count,
    })


@login_required
@require_http_methods(["POST"])
def class_generate_missing_invoices(request, pk):
    """
    Generate invoices for all active ongoing enrollments in this class
    that don't already have one.

    Each enrollment gets its own transaction.atomic() savepoint so a single
    failure (e.g. missing fee structure) does not roll back successful ones.
    Calls generate_student_enrollment_invoice directly — no reimplementation
    of fee calculation logic here.
    """
    class_instance = get_object_or_404(Class, pk=pk)

    enrollments_without_invoice = class_instance.enrollments.filter(
        academic_invoice__isnull=True,
        is_active=True,
        completion_status='ONGOING'
    ).select_related('student', 'class_instance', 'academic_session')

    if not enrollments_without_invoice.exists():
        if request.headers.get('HX-Request') == 'true':
            r = HttpResponse()
            r['HX-Alert-Message'] = 'All enrollments in this class already have invoices'
            r['HX-Alert-Type']    = 'info'
            r['HX-Close-Modal']   = 'true'
            return r
        messages.info(request, 'All enrollments already have invoices')
        return redirect('academics:class_detail', pk=pk)

    from fees.invoice_generators import (
        generate_student_enrollment_invoice,
        FeeStructureNotFoundError,
    )

    success_count = 0
    failed = []

    for enrollment in enrollments_without_invoice:
        try:
            with transaction.atomic():
                generate_student_enrollment_invoice(enrollment)
            success_count += 1
        except FeeStructureNotFoundError:
            failed.append(f"{enrollment.student.get_full_name()}: no fee structure found")
        except Exception as e:
            failed.append(f"{enrollment.student.get_full_name()}: {e}")

    msg = f"Generated {success_count} invoice(s)."
    alert_type = 'success'

    if failed:
        msg += f" {len(failed)} failed: {'; '.join(failed)}"
        alert_type = 'warning'

    if request.headers.get('HX-Request') == 'true':
        r = HttpResponse()
        r['HX-Alert-Message'] = msg
        r['HX-Alert-Type']    = alert_type
        r['HX-Close-Modal']   = 'true'
        r['HX-Redirect']      = reverse('academics:class_detail', kwargs={'pk': pk})
        return r

    if failed:
        messages.warning(request, msg)
    else:
        messages.success(request, msg)
    return redirect('academics:class_detail', pk=pk)


# --- Class subjects (no standalone list — managed from class detail) ---------

@login_required
def class_subject_detail(request, pk):
    cs = get_object_or_404(
        ClassSubject.objects.select_related(
            'class_instance__academic_level',
            'class_instance__academic_session',
            'class_instance__class_teacher__staff',
            'class_instance__classroom',
            'subject',
            'teacher__staff',
        ),
        pk=pk,
    )
    enrolled_count = cs.class_instance.enrollments.filter(
        is_active=True, completion_status='ONGOING'
    ).count()
    stats = {
        'enrolled_students':    enrolled_count,
        'hours_per_week':       cs.hours_per_week,
        'total_hours':          cs.total_hours or 0,
        'ca_weight':            cs.continuous_assessment_weight,
        'exam_weight':          cs.final_exam_weight,
    }
    return render(request, 'academics/class_subjects/detail.html', {
        'class_subject':  cs,
        'class_instance': cs.class_instance,
        'stats':          stats,
    })


@login_required
def class_subject_create(request):
    class_pk       = request.GET.get('class_pk') or request.POST.get('class_pk')
    class_instance = get_object_or_404(Class, pk=class_pk) if class_pk else None
    if request.method == 'POST':
        form = ClassSubjectForm(request.POST)
        if form.is_valid():
            try:
                with transaction.atomic():
                    cs = form.save(commit=False)
                    if class_instance and not cs.class_instance_id:
                        cs.class_instance = class_instance
                    cs.save()
                if request.headers.get('HX-Request') == 'true':
                    r = HttpResponse()
                    r['HX-Alert-Message'] = f'"{cs.subject.name}" assigned successfully'
                    r['HX-Alert-Type'] = 'success'
                    r['HX-Close-Modal'] = 'true'
                    r['HX-Redirect'] = reverse('academics:class_detail', kwargs={'pk': cs.class_instance.pk})
                    return r
                messages.success(request, f'"{cs.subject.name}" assigned successfully')
                return redirect('academics:class_detail', pk=cs.class_instance.pk)
            except Exception as e:
                logger.error(f"Error assigning subject: {e}")
                if request.headers.get('HX-Request') == 'true':
                    r = HttpResponse()
                    r['HX-Alert-Message'] = f'Error: {e}'
                    r['HX-Alert-Type'] = 'error'
                    return r
                messages.error(request, f'Error: {e}')
        else:
            errors = '; '.join(
                f'{f}: {", ".join(errs)}' if f != '__all__' else ', '.join(errs)
                for f, errs in form.errors.items()
            )
            if request.headers.get('HX-Request') == 'true':
                r = HttpResponse()
                r['HX-Alert-Message'] = errors or 'Please correct the errors in the form'
                r['HX-Alert-Type'] = 'error'
                return r
            messages.error(request, errors or 'Please correct the errors in the form')
    else:
        form = ClassSubjectForm(initial={
            'class_instance':               class_instance,
            'continuous_assessment_weight': 40,
            'final_exam_weight':            60,
            'is_active':                    True,
        })
    return render(request, 'academics/classes/modals/class_subject_form.html', {
        'form': form, 'class_instance': class_instance,
    })


@login_required
def class_subject_edit(request, pk):
    cs = get_object_or_404(
        ClassSubject.objects.select_related('class_instance', 'subject', 'teacher'),
        pk=pk,
    )
    class_instance = cs.class_instance

    if request.method == 'POST':
        form = ClassSubjectForm(request.POST, instance=cs)
        if form.is_valid():
            try:
                with transaction.atomic():
                    cs = form.save()
                if request.headers.get('HX-Request') == 'true':
                    r = HttpResponse()
                    r['HX-Alert-Message'] = f'"{cs.subject.name}" updated successfully'
                    r['HX-Alert-Type']    = 'success'
                    r['HX-Close-Modal']   = 'true'
                    r['HX-Redirect']      = reverse(
                        'academics:class_detail', kwargs={'pk': class_instance.pk}
                    )
                    return r
                messages.success(request, f'"{cs.subject.name}" updated successfully')
                return redirect('academics:class_detail', pk=class_instance.pk)

            except Exception as e:
                logger.error(f"Error updating subject assignment: {e}")
                if request.headers.get('HX-Request') == 'true':
                    r = HttpResponse()
                    r['HX-Alert-Message'] = f'Error: {e}'
                    r['HX-Alert-Type']    = 'error'
                    return r
                messages.error(request, f'Error: {e}')

        else:
            errors = '; '.join(
                f'{f}: {", ".join(errs)}' if f != '__all__' else ', '.join(errs)
                for f, errs in form.errors.items()
            )
            if request.headers.get('HX-Request') == 'true':
                r = HttpResponse()
                r['HX-Alert-Message'] = errors or 'Please correct the errors in the form'
                r['HX-Alert-Type']    = 'error'
                return r
            messages.error(request, errors or 'Please correct the errors in the form')

    else:
        form = ClassSubjectForm(instance=cs)

    return render(request, 'academics/classes/modals/class_subject_form.html', {
        'form':           form,
        'class_subject':  cs,
        'class_instance': class_instance,
    })


@login_required
def class_subject_delete(request, pk):
    cs = get_object_or_404(ClassSubject, pk=pk)
    if request.method == 'POST':
        try:
            subj_name = cs.subject.name; class_pk = cs.class_instance.pk; cs.delete()
            if request.headers.get('HX-Request') == 'true':
                r = HttpResponse(); r['HX-Alert-Message'] = f'Subject "{subj_name}" removed from class'; r['HX-Alert-Type'] = 'success'; r['HX-Close-Modal'] = 'true'; r['HX-Redirect'] = reverse('academics:class_detail', kwargs={'pk': class_pk}); return r
            messages.success(request, f'Subject "{subj_name}" removed from class')
            return redirect('academics:class_detail', pk=class_pk)
        except Exception as e:
            if request.headers.get('HX-Request') == 'true':
                r = HttpResponse(); r['HX-Alert-Message'] = f'Error: {e}'; r['HX-Alert-Type'] = 'error'; r['HX-Close-Modal'] = 'true'; return r
            messages.error(request, f'Error: {e}')
    return redirect('academics:class_list')


@login_required
def class_subject_assign_teacher(request, pk):
    cs = get_object_or_404(ClassSubject, pk=pk)
    if request.method == 'POST':
        try:
            from hr.models import Teacher
            tid = request.POST.get('teacher_id')
            if tid:
                teacher = get_object_or_404(Teacher, pk=tid); old = cs.teacher
                cs.teacher = teacher; cs.save(update_fields=['teacher'])
                msg = f'Teacher {teacher.staff.full_name()} assigned to {cs.subject.name}'
                if old: msg += f' (replaced {old.staff.full_name()})'
            else:
                cs.teacher = None; cs.save(update_fields=['teacher']); msg = f'Teacher removed from {cs.subject.name}'
            if request.headers.get('HX-Request') == 'true':
                r = HttpResponse(); r['HX-Alert-Message'] = msg; r['HX-Alert-Type'] = 'success'; r['HX-Close-Modal'] = 'true'; r['HX-Redirect'] = reverse('academics:class_detail', kwargs={'pk': cs.class_instance.pk}); return r
            messages.success(request, msg)
        except Exception as e:
            if request.headers.get('HX-Request') == 'true':
                r = HttpResponse(); r['HX-Alert-Message'] = f'Error: {e}'; r['HX-Alert-Type'] = 'error'; r['HX-Close-Modal'] = 'true'; return r
            messages.error(request, f'Error: {e}')
    return redirect('academics:class_detail', pk=cs.class_instance.pk)


@login_required
def class_subject_toggle_active(request, pk):
    cs = get_object_or_404(ClassSubject, pk=pk)
    if request.method == 'POST':
        try:
            cs.is_active = not cs.is_active; cs.save(update_fields=['is_active'])
            word = 'activated' if cs.is_active else 'deactivated'
            if request.headers.get('HX-Request') == 'true':
                r = HttpResponse(); r['HX-Alert-Message'] = f'Subject assignment {word}'; r['HX-Alert-Type'] = 'success'; r['HX-Close-Modal'] = 'true'; r['HX-Redirect'] = reverse('academics:class_detail', kwargs={'pk': cs.class_instance.pk}); return r
            messages.success(request, f'Subject assignment {word}')
        except Exception as e:
            if request.headers.get('HX-Request') == 'true':
                r = HttpResponse(); r['HX-Alert-Message'] = f'Error: {e}'; r['HX-Alert-Type'] = 'error'; r['HX-Close-Modal'] = 'true'; return r
            messages.error(request, f'Error: {e}')
    return redirect('academics:class_detail', pk=cs.class_instance.pk)


# =============================================================================
# 8. STUDENT ENROLLMENTS  (bulk wizard embedded here)
# =============================================================================

def get_filtered_enrollments(request=None, filter_params=None):
    qs = StudentClassEnrollment.objects.select_related(
        'student', 'class_instance__academic_level', 'class_instance__academic_session',
        'academic_session', 'academic_invoice',
    ).order_by('-enrollment_date')
    params = {}
    if request:
        params = request.GET
    elif filter_params:
        params = filter_params
    q = (params.get('q') or '').strip()
    if q:
        words   = q.split()
        combined = Q()
        for word in words:
            combined &= (Q(student__first_name__icontains=word) | Q(student__last_name__icontains=word) | Q(student__middle_name__icontains=word) | Q(student__admission_number__icontains=word) | Q(roll_number__icontains=word))
        qs = qs.filter(combined)
    for field, lookup in (('class_instance', 'class_instance_id'), ('academic_session', 'academic_session_id'), ('enrollment_type', 'enrollment_type'), ('completion_status', 'completion_status'), ('progression_type', 'progression_type')):
        v = params.get(field, '')
        if v: qs = qs.filter(**{lookup: v})
    ia = params.get('is_active', '')
    if ia: qs = qs.filter(is_active=(ia.lower() == 'true'))
    hi = params.get('has_invoice', '')
    if hi:
        qs = qs.filter(academic_invoice__isnull=False) if hi.lower() == 'true' else qs.filter(academic_invoice__isnull=True)
    for date_field, lookup in (('enrollment_date_from', 'enrollment_date__gte'), ('enrollment_date_to', 'enrollment_date__lte')):
        v = params.get(date_field, '')
        if v:
            try: qs = qs.filter(**{lookup: datetime.strptime(v, '%Y-%m-%d').date()})
            except (ValueError, TypeError): pass
    return qs


@login_required
def enrollment_detail(request, pk):
    enrollment = get_object_or_404(StudentClassEnrollment.objects.select_related('student', 'class_instance__academic_level', 'academic_session', 'academic_invoice', 'previous_enrollment'), pk=pk)
    history        = StudentClassEnrollment.objects.filter(student=enrollment.student).select_related('class_instance', 'academic_session').order_by('-enrollment_date')
    next_enrollment = StudentClassEnrollment.objects.filter(previous_enrollment=enrollment).select_related('class_instance', 'academic_session').first()
    return render(request, 'academics/enrollments/detail.html', {'enrollment': enrollment, 'enrollment_history': history, 'next_enrollment': next_enrollment})


@login_required
def enrollment_create(request, student_pk=None, class_pk=None):
    """POST-only. Modal form (enrollment_create_modal) handles GET."""
    if request.method != 'POST':
        return HttpResponse(status=405)

    if not class_pk:
        class_pk = request.POST.get('class_pk')

    form = StudentEnrollmentForm(request.POST)
    if form.is_valid():
        try:
            with transaction.atomic():
                e = form.save(commit=False)
                if not e.enrollment_date:
                    e.enrollment_date = get_school_today()
                e.save()

                if e.auto_create_invoice:
                    try:
                        from fees.invoice_generators import (
                            generate_student_enrollment_invoice,
                            FeeStructureNotFoundError,
                        )
                        invoice = generate_student_enrollment_invoice(e)
                        logger.info(
                            f"Created invoice {invoice.invoice_number} "
                            f"for {e.student.get_full_name()}"
                        )
                    except FeeStructureNotFoundError as fee_err:
                        logger.warning(f"Invoice not created: {fee_err}")
                        messages.warning(
                            request,
                            f"Enrolled successfully but no invoice was created: {fee_err}"
                        )
                    except Exception as fee_err:
                        logger.warning(f"Invoice creation failed: {fee_err}")
                        messages.warning(
                            request,
                            f"Enrolled successfully but invoice creation failed: {fee_err}"
                        )

            msg = (
                f'Enrollment created for {e.student.get_full_name()}'
                + (f' (Roll #{e.roll_number})' if e.roll_number else '')
            )
            if request.headers.get('HX-Request') == 'true':
                r = HttpResponse()
                r['HX-Alert-Message'] = msg
                r['HX-Alert-Type']    = 'success'
                r['HX-Close-Modal']   = 'true'
                r['HX-Redirect']      = reverse(
                    'academics:enrollment_detail', kwargs={'pk': e.pk}
                )
                return r
            messages.success(request, msg)
            return redirect('academics:enrollment_detail', pk=e.pk)

        except ValidationError as exc:
            err = '; '.join(exc.messages) if hasattr(exc, 'messages') else str(exc)
            if request.headers.get('HX-Request') == 'true':
                r = HttpResponse()
                r['HX-Alert-Message'] = err
                r['HX-Alert-Type']    = 'error'
                return r
            messages.error(request, err)

        except Exception as exc:
            logger.error(f"Error creating enrollment: {exc}", exc_info=True)
            if request.headers.get('HX-Request') == 'true':
                r = HttpResponse()
                r['HX-Alert-Message'] = f'Error: {exc}'
                r['HX-Alert-Type']    = 'error'
                return r
            messages.error(request, f'Error: {exc}')

    else:
        errors = '; '.join(
            f'{f}: {", ".join(errs)}' if f != '__all__' else ', '.join(errs)
            for f, errs in form.errors.items()
        )
        if request.headers.get('HX-Request') == 'true':
            r = HttpResponse()
            r['HX-Alert-Message'] = errors or 'Please correct the errors in the form'
            r['HX-Alert-Type']    = 'error'
            return r
        messages.error(request, errors or 'Please correct the errors in the form')

    return (
        redirect('academics:class_detail', pk=class_pk)
        if class_pk
        else redirect('academics:dashboard')
    )


@login_required
def enrollment_edit(request, pk):
    """POST-only. Modal form (enrollment_edit_modal) handles GET."""
    enrollment = get_object_or_404(StudentClassEnrollment, pk=pk)
    if request.method != 'POST':
        return HttpResponse(status=405)

    form = StudentEnrollmentForm(request.POST, instance=enrollment)
    if form.is_valid():
        try:
            with transaction.atomic():
                enrollment = form.save()
            if request.headers.get('HX-Request') == 'true':
                r = HttpResponse()
                r['HX-Alert-Message'] = 'Enrollment updated successfully'
                r['HX-Alert-Type']    = 'success'
                r['HX-Close-Modal']   = 'true'
                r['HX-Redirect']      = reverse(
                    'academics:enrollment_detail', kwargs={'pk': enrollment.pk}
                )
                return r
            messages.success(request, 'Enrollment updated successfully')
            return redirect('academics:enrollment_detail', pk=enrollment.pk)

        except Exception as e:
            logger.error(f"Error updating enrollment: {e}", exc_info=True)
            if request.headers.get('HX-Request') == 'true':
                r = HttpResponse()
                r['HX-Alert-Message'] = f'Error: {e}'
                r['HX-Alert-Type']    = 'error'
                return r
            messages.error(request, f'Error: {e}')
    else:
        errors = '; '.join(
            f'{f}: {", ".join(errs)}' if f != '__all__' else ', '.join(errs)
            for f, errs in form.errors.items()
        )
        if request.headers.get('HX-Request') == 'true':
            r = HttpResponse()
            r['HX-Alert-Message'] = errors or 'Please correct the errors in the form'
            r['HX-Alert-Type']    = 'error'
            return r
        messages.error(request, errors or 'Please correct the errors in the form')

    return redirect('academics:enrollment_detail', pk=enrollment.pk)


@login_required
def enrollment_delete(request, pk):
    """
    Delete an enrollment.

    Invoice validity is enforced by the class_enrollment_pre_delete signal
    which raises ValidationError if the invoice cannot be safely removed.
    """
    enrollment = get_object_or_404(StudentClassEnrollment, pk=pk)
    is_htmx    = request.headers.get('HX-Request') == 'true'

    if request.method == 'POST':
        try:
            with transaction.atomic():
                student_name = enrollment.student.get_full_name()
                class_pk     = enrollment.class_instance.pk
                had_invoice  = bool(enrollment.academic_invoice)
                inv_number   = enrollment.academic_invoice.invoice_number if had_invoice else None
                inv_status   = enrollment.academic_invoice.status if had_invoice else None
                enrollment.delete()

            success_msg = f'Enrollment for "{student_name}" deleted.'
            if had_invoice:
                success_msg = (
                    f'Enrollment for "{student_name}" deleted. '
                    f'{inv_status} invoice {inv_number} also removed.'
                )

            if is_htmx:
                r = HttpResponse(status=200)
                r['HX-Trigger']  = json.dumps({
                    'closeModal':  True,
                    'showAlert':   {
                        'message': success_msg,
                        'type':    'success',
                        'title':   'Deleted',
                    }
                })
                r['HX-Redirect'] = reverse(
                    'academics:class_detail', kwargs={'pk': class_pk}
                )
                return r
            messages.success(request, success_msg)
            return redirect('academics:class_detail', pk=class_pk)

        except (ValidationError, Exception) as e:
            logger.error(f"Error deleting enrollment: {e}", exc_info=True)
            err_msg = (
                e.message if hasattr(e, 'message')
                else str(e).replace('\n', ' ')
            )
            if is_htmx:
                r = HttpResponse(status=200)
                r['HX-Trigger'] = json.dumps({
                    'showAlert': {
                        'message': err_msg,
                        'type':    'error',
                        'title':   'Cannot Delete',
                    },
                    'closeModal': True,
                })
                return r
            messages.error(request, err_msg)
            return redirect('academics:enrollment_detail', pk=pk)

    return redirect('academics:enrollment_detail', pk=pk)


@login_required
def enrollment_toggle_active(request, pk):
    enrollment = get_object_or_404(StudentClassEnrollment, pk=pk)
    if request.method == 'POST':
        try:
            enrollment.is_active = not enrollment.is_active
            enrollment.save(update_fields=['is_active'])
            word = 'activated' if enrollment.is_active else 'deactivated'
            if request.headers.get('HX-Request') == 'true':
                r = HttpResponse()
                r['HX-Alert-Message'] = f'Enrollment {word} successfully'
                r['HX-Alert-Type']    = 'success'
                r['HX-Close-Modal']   = 'true'
                r['HX-Redirect']      = reverse(
                    'academics:enrollment_detail', kwargs={'pk': pk}
                )
                return r
            messages.success(request, f'Enrollment {word}')
        except Exception as e:
            if request.headers.get('HX-Request') == 'true':
                r = HttpResponse()
                r['HX-Alert-Message'] = f'Error: {e}'
                r['HX-Alert-Type']    = 'error'
                r['HX-Close-Modal']   = 'true'
                return r
            messages.error(request, f'Error: {e}')
    return redirect('academics:enrollment_detail', pk=pk)


@login_required
def enrollment_create_invoice(request, pk):
    """
    Manually create an invoice for an existing enrollment that does not yet
    have one. Action view — POST only. Modal handled by enrollment_create_invoice_modal.
    """
    enrollment = get_object_or_404(StudentClassEnrollment, pk=pk)

    if request.method == 'POST':
        if enrollment.academic_invoice:
            if request.headers.get('HX-Request') == 'true':
                r = HttpResponse()
                r['HX-Alert-Message'] = 'An invoice already exists for this enrollment'
                r['HX-Alert-Type']    = 'warning'
                r['HX-Close-Modal']   = 'true'
                return r
            messages.warning(request, 'Invoice already exists')
            return redirect('academics:enrollment_detail', pk=pk)

        try:
            from fees.invoice_generators import (
                generate_student_enrollment_invoice,
                FeeStructureNotFoundError,
            )
            with transaction.atomic():
                invoice = generate_student_enrollment_invoice(enrollment)
                if not enrollment.academic_invoice:
                    enrollment.academic_invoice = invoice
                    enrollment.save(update_fields=['academic_invoice'])

            if request.headers.get('HX-Request') == 'true':
                r = HttpResponse()
                r['HX-Alert-Message'] = f'Invoice {invoice.invoice_number} created successfully'
                r['HX-Alert-Type']    = 'success'
                r['HX-Close-Modal']   = 'true'
                r['HX-Redirect']      = reverse(
                    'academics:enrollment_detail', kwargs={'pk': pk}
                )
                return r
            messages.success(request, f'Invoice {invoice.invoice_number} created successfully')

        except FeeStructureNotFoundError as e:
            logger.error(f"Fee structure not found: {e}")
            if request.headers.get('HX-Request') == 'true':
                r = HttpResponse()
                r['HX-Alert-Message'] = f'Cannot create invoice: {e}'
                r['HX-Alert-Type']    = 'error'
                r['HX-Close-Modal']   = 'true'
                return r
            messages.error(request, f'Cannot create invoice: {e}')

        except Exception as e:
            logger.error(f"Error creating invoice: {e}", exc_info=True)
            if request.headers.get('HX-Request') == 'true':
                r = HttpResponse()
                r['HX-Alert-Message'] = f'Error: {e}'
                r['HX-Alert-Type']    = 'error'
                r['HX-Close-Modal']   = 'true'
                return r
            messages.error(request, f'Error: {e}')

    return redirect('academics:enrollment_detail', pk=pk)


# --- Bulk enrollment wizard -------------------------------------------------

@login_required
def bulk_enrollment_create(request):
    class_pk       = request.GET.get('class_id') or request.POST.get('class_id')
    class_instance = get_object_or_404(Class, pk=class_pk) if class_pk else None
    levels         = AcademicLevel.objects.filter(is_active=True).order_by('order')

    return render(request, 'academics/enrollments/bulk_step1.html', {
        'class_instance': class_instance,
        'levels':         levels,
    })


@login_required
def bulk_enrollment_student_search(request):
    class_pk       = request.GET.get('class_id')
    class_instance = get_object_or_404(Class, pk=class_pk) if class_pk else None

    qs = Student.objects.select_related('current_academic_level').order_by(
        'first_name', 'last_name'
    )

    if class_instance:
        qs = qs.filter(current_academic_level=class_instance.academic_level)

    search = request.GET.get('search', '').strip()
    if search:
        qs = qs.filter(
            Q(first_name__icontains=search)        |
            Q(last_name__icontains=search)         |
            Q(admission_number__icontains=search)
        )

    enrollment_status = request.GET.get('enrollment_status', 'ACTIVE')
    if enrollment_status:
        qs = qs.filter(enrollment_status=enrollment_status)

    gender = request.GET.get('gender', '')
    if gender:
        qs = qs.filter(gender=gender)

    enrolled_in_class_ids       = set()
    enrolled_in_other_class_ids = set()

    if class_instance:
        enrolled_in_class_ids = set(
            StudentClassEnrollment.objects.filter(
                class_instance=class_instance,
                is_active=True,
                completion_status='ONGOING',
            ).values_list('student_id', flat=True)
        )
        enrolled_in_other_class_ids = set(
            StudentClassEnrollment.objects.filter(
                academic_session=class_instance.academic_session,
                is_active=True,
                completion_status='ONGOING',
            ).exclude(
                class_instance=class_instance
            ).values_list('student_id', flat=True)
        )

    exclude_enrolled = request.GET.get('exclude_already_enrolled', '')
    if exclude_enrolled and class_instance:
        qs = qs.exclude(
            pk__in=enrolled_in_class_ids | enrolled_in_other_class_ids
        )

    paginator   = Paginator(qs, 25)
    page_obj    = paginator.get_page(request.GET.get('page', 1))
    total_count = paginator.count

    return render(request, 'academics/enrollments/_student_selection_results.html', {
        'page_obj':                    page_obj,
        'total_count':                 total_count,
        'class_instance':              class_instance,
        'enrolled_in_class_ids':       enrolled_in_class_ids,
        'enrolled_in_other_class_ids': enrolled_in_other_class_ids,
    })


@login_required
def bulk_enrollment_step2(request):
    class_pk       = request.GET.get('class_id') or request.POST.get('class_id')
    class_instance = get_object_or_404(Class, pk=class_pk) if class_pk else None
    student_ids    = request.GET.getlist('student_ids') or request.POST.getlist('student_ids')
    students       = Student.objects.filter(pk__in=student_ids).select_related('current_academic_level')

    if request.method == 'POST':
        enrollment_date     = request.POST.get('enrollment_date')
        enrollment_type     = request.POST.get('enrollment_type', 'CONTINUING')
        auto_create_invoice = request.POST.get('auto_create_invoice') == 'on'
        confirm             = request.POST.get('confirm_enrollment') == 'on'

        if not confirm:
            messages.error(request, 'Please tick the confirmation checkbox.')
        elif not student_ids:
            messages.error(request, 'No students selected.')
        else:
            try:
                from datetime import datetime
                parsed_date = (
                    datetime.strptime(enrollment_date, '%Y-%m-%d').date()
                    if enrollment_date
                    else get_school_today()
                )

                from academics.services import BulkEnrollmentService
                service = BulkEnrollmentService()
                result  = service.enroll_students(
                    student_ids=student_ids,
                    academic_session=class_instance.academic_session,
                    class_instance=class_instance,
                    enrollment_date=parsed_date,
                    enrollment_type=enrollment_type,
                    auto_create_invoice=auto_create_invoice,
                    created_by=request.user,
                )

                for warning in result.get('warnings', []):
                    messages.warning(request, warning)

                if result['success']:
                    enrolled = result['enrolled_count']
                    failed   = result['failed_count']

                    if failed:
                        for err in result.get('errors', []):
                            messages.warning(request, err)
                        messages.warning(
                            request,
                            f"Enrolled {enrolled} student(s). "
                            f"{failed} could not be enrolled — see details above."
                        )
                    else:
                        messages.success(
                            request,
                            f"Successfully enrolled {enrolled} student(s) into "
                            f"{class_instance.get_display_name()}."
                        )

                    return redirect('academics:class_detail', pk=class_instance.pk)

                else:
                    for err in result.get('errors', []):
                        messages.error(request, err)

            except Exception as e:
                logger.error(f"Bulk enrollment error: {e}", exc_info=True)
                messages.error(request, f'Enrollment failed: {e}')

        form = BulkEnrollmentConfirmationForm(
            initial={
                'enrollment_date':     enrollment_date or get_school_today(),
                'enrollment_type':     enrollment_type,
                'auto_create_invoice': auto_create_invoice,
            },
            student_count=len(student_ids),
        )
        return render(request, 'academics/enrollments/bulk_step2.html', {
            'class_instance': class_instance,
            'students':       students,
            'form':           form,
            'student_ids':    student_ids,
        })

    form = BulkEnrollmentConfirmationForm(
        initial={
            'enrollment_date':     get_school_today(),
            'enrollment_type':     'CONTINUING',
            'auto_create_invoice': True,
        },
        student_count=len(student_ids),
    )

    return render(request, 'academics/enrollments/bulk_step2.html', {
        'class_instance': class_instance,
        'students':       students,
        'form':           form,
        'student_ids':    student_ids,
    })


# =============================================================================
# 9. HOLIDAYS
# =============================================================================

def get_filtered_holidays(request):
    qs = Holiday.objects.select_related('academic_session').order_by('-start_date')
    q = request.GET.get('q', '').strip()
    if q: qs = qs.filter(Q(name__icontains=q) | Q(description__icontains=q))
    ht = request.GET.get('holiday_type', '')
    if ht: qs = qs.filter(holiday_type=ht)
    sc = request.GET.get('is_school_closed', '')
    if sc: qs = qs.filter(is_school_closed=(sc.lower() == 'true'))
    sess = request.GET.get('academic_session', '')
    if sess: qs = qs.filter(academic_session_id=sess)
    return qs


@login_required
def holiday_list(request):
    filter_form = HolidayFilterForm(request.GET or None)
    holidays    = get_filtered_holidays(request)
    today       = get_school_today()
    stats = {'total': holidays.count(), 'school_closed': holidays.filter(is_school_closed=True).count(), 'upcoming': holidays.filter(start_date__gte=today).count()}
    paginator     = Paginator(holidays, 20)
    holidays_page = paginator.get_page(request.GET.get('page', 1))
    is_htmx       = request.headers.get('HX-Request') == 'true'
    context = {'holidays_page': holidays_page, 'paginator': paginator, 'stats': stats, 'filter_form': filter_form, 'is_htmx': is_htmx}
    if is_htmx:
        return render(request, 'academics/holidays/partials/_holiday_results.html', context)
    return render(request, 'academics/holidays/list.html', context)


@login_required
def holiday_detail(request, pk):
    holiday          = get_object_or_404(Holiday, pk=pk)
    affected_sessions = AcademicSession.objects.filter(start_date__lte=holiday.end_date or holiday.start_date, end_date__gte=holiday.start_date, is_active=True)
    return render(request, 'academics/holidays/detail.html', {'holiday': holiday, 'duration': holiday.duration_days, 'affected_sessions': affected_sessions})


@login_required
def holiday_create(request):
    if request.method == 'POST':
        form = HolidayForm(request.POST)
        if form.is_valid():
            try:
                holiday = form.save()
                if request.headers.get('HX-Request') == 'true':
                    r = HttpResponse(); r['HX-Alert-Message'] = f'Holiday "{holiday.name}" created successfully'; r['HX-Alert-Type'] = 'success'; r['HX-Close-Modal'] = 'true'; r['HX-Redirect'] = reverse('academics:holiday_detail', kwargs={'pk': holiday.pk}); return r
                messages.success(request, f'Holiday "{holiday.name}" created successfully')
                return redirect('academics:holiday_detail', pk=holiday.pk)
            except Exception as e:
                if request.headers.get('HX-Request') == 'true':
                    r = HttpResponse(); r['HX-Alert-Message'] = f'Error: {e}'; r['HX-Alert-Type'] = 'error'; return r
                messages.error(request, f'Error: {e}')
        else:
            if request.headers.get('HX-Request') == 'true':
                r = HttpResponse(); r['HX-Alert-Message'] = 'Please correct the errors in the form'; r['HX-Alert-Type'] = 'error'; return r
            messages.error(request, 'Please correct the errors in the form')
    else:
        form = HolidayForm()
    return render(request, 'academics/holidays/form.html', {'form': form, 'title': 'Create holiday'})


@login_required
def holiday_edit(request, pk):
    holiday = get_object_or_404(Holiday, pk=pk)
    if request.method == 'POST':
        form = HolidayForm(request.POST, instance=holiday)
        if form.is_valid():
            try:
                holiday = form.save()
                if request.headers.get('HX-Request') == 'true':
                    r = HttpResponse(); r['HX-Alert-Message'] = 'Holiday updated successfully'; r['HX-Alert-Type'] = 'success'; r['HX-Close-Modal'] = 'true'; r['HX-Redirect'] = reverse('academics:holiday_detail', kwargs={'pk': holiday.pk}); return r
                messages.success(request, 'Holiday updated successfully')
                return redirect('academics:holiday_detail', pk=holiday.pk)
            except Exception as e:
                if request.headers.get('HX-Request') == 'true':
                    r = HttpResponse(); r['HX-Alert-Message'] = f'Error: {e}'; r['HX-Alert-Type'] = 'error'; return r
                messages.error(request, f'Error: {e}')
        else:
            if request.headers.get('HX-Request') == 'true':
                r = HttpResponse(); r['HX-Alert-Message'] = 'Please correct the errors in the form'; r['HX-Alert-Type'] = 'error'; return r
            messages.error(request, 'Please correct the errors in the form')
    else:
        form = HolidayForm(instance=holiday)
    return render(request, 'academics/holidays/form.html', {'form': form, 'holiday': holiday, 'title': f'Edit {holiday.name}'})


@login_required
def holiday_delete(request, pk):
    holiday = get_object_or_404(Holiday, pk=pk)
    if request.method == 'POST':
        try:
            name = holiday.name; holiday.delete()
            if request.headers.get('HX-Request') == 'true':
                r = HttpResponse(); r['HX-Alert-Message'] = f'Holiday "{name}" deleted successfully'; r['HX-Alert-Type'] = 'success'; r['HX-Close-Modal'] = 'true'; r['HX-Redirect'] = reverse('academics:holiday_list'); return r
            messages.success(request, f'Holiday "{name}" deleted successfully')
            return redirect('academics:holiday_list')
        except Exception as e:
            if request.headers.get('HX-Request') == 'true':
                r = HttpResponse(); r['HX-Alert-Message'] = f'Error: {e}'; r['HX-Alert-Type'] = 'error'; r['HX-Close-Modal'] = 'true'; return r
            messages.error(request, f'Error: {e}')
            return redirect('academics:holiday_detail', pk=pk)
    return redirect('academics:holiday_list')


@login_required
def holiday_print_view(request):
    holidays        = get_filtered_holidays(request)
    selected_fields = request.GET.getlist('fields') or ['name', 'holiday_type', 'start_date', 'end_date', 'duration_days', 'is_school_closed']
    include_stats   = request.GET.get('include_stats') == 'true'
    stats = None
    if include_stats:
        stats = {'total': holidays.count(), 'school_closed': holidays.filter(is_school_closed=True).count()}
    field_labels = {
        'name': 'Holiday Name', 'holiday_type': 'Type',
        'start_date': 'Start Date', 'end_date': 'End Date',
        'duration_days': 'Duration (days)', 'is_school_closed': 'School Closed',
        'is_recurring': 'Recurring', 'academic_session': 'Session',
        'affects_attendance': 'Affects Attendance', 'affects_payroll': 'Affects Payroll',
    }
    return render(request, 'academics/holidays/print.html', {
        **_get_print_school_context(request),
        'holidays': holidays, 'stats': stats, 'now': timezone.now(),
        'selected_fields': selected_fields,
        'selected_field_names': [field_labels.get(f, f.replace('_', ' ').title()) for f in selected_fields],
        'field_labels': field_labels, 'landscape': request.GET.get('landscape') == 'true',
        'title': 'Holidays Report',
    })


@login_required
def export_holidays_excel(request):
    ALL_COLUMNS = [
        ('name',              'Holiday Name',       lambda h: h.name),
        ('holiday_type',      'Type',               lambda h: h.get_holiday_type_display()),
        ('start_date',        'Start Date',         lambda h: h.start_date.strftime('%Y-%m-%d')),
        ('end_date',          'End Date',           lambda h: h.end_date.strftime('%Y-%m-%d') if h.end_date else h.start_date.strftime('%Y-%m-%d')),
        ('duration_days',     'Duration (days)',    lambda h: h.duration_days),
        ('is_school_closed',  'School Closed',      lambda h: 'Yes' if h.is_school_closed else 'No'),
        ('is_partial',        'Partial Closure',    lambda h: 'Yes' if h.is_partial_closure else 'No'),
        ('affects_attendance','Affects Attendance', lambda h: 'Yes' if h.affects_attendance else 'No'),
        ('affects_payroll',   'Affects Payroll',    lambda h: 'Yes' if h.affects_payroll else 'No'),
        ('is_recurring',      'Recurring',          lambda h: 'Yes' if h.is_recurring else 'No'),
        ('academic_session',  'Session',            lambda h: h.academic_session.name if h.academic_session else ''),
        ('description',       'Description',        lambda h: h.description or ''),
    ]
    DEFAULT  = ['name', 'holiday_type', 'start_date', 'end_date', 'duration_days', 'is_school_closed', 'is_recurring']
    col_map  = {k: (lbl, fn) for k, lbl, fn in ALL_COLUMNS}
    chosen   = request.GET.getlist('fields') or DEFAULT
    columns  = [(col_map[f][0], col_map[f][1]) for f in chosen if f in col_map]
    holidays = get_filtered_holidays(request)
    wb = Workbook(); ws = wb.active; ws.title = 'Holidays'
    ws.append([c[0] for c in columns])
    hf = PatternFill(start_color='1E3A5F', end_color='1E3A5F', fill_type='solid')
    for cell in ws[1]:
        cell.fill = hf; cell.font = Font(bold=True, color='FFFFFF', size=11)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws.row_dimensions[1].height = 28
    af = PatternFill(start_color='F4F6F9', end_color='F4F6F9', fill_type='solid')
    da = Alignment(vertical='center', wrap_text=False)
    for i, obj in enumerate(holidays):
        ws.append([c[1](obj) for c in columns])
        for cell in ws[i + 2]:
            cell.alignment = da
            if i % 2 == 1: cell.fill = af
    for cc in ws.columns:
        ml = max((len(str(c.value)) if c.value else 0) for c in cc)
        ws.column_dimensions[cc[0].column_letter].width = min(ml + 4, 60)
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="holidays_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
    wb.save(response); return response


@login_required
def export_holidays_calendar(request):
    try:
        from icalendar import Calendar, Event
    except ImportError:
        messages.error(request, 'Calendar export requires the icalendar package')
        return redirect('academics:holiday_list')
    holidays = Holiday.objects.filter(is_active=True).order_by('start_date')
    cal = Calendar(); cal.add('prodid', '-//Academic Calendar//EN'); cal.add('version', '2.0')
    for h in holidays:
        ev = Event(); ev.add('summary', h.name); ev.add('dtstart', h.start_date); ev.add('dtend', h.end_date or h.start_date)
        if h.description: ev.add('description', h.description)
        cal.add_component(ev)
    response = HttpResponse(cal.to_ical(), content_type='text/calendar')
    response['Content-Disposition'] = f'attachment; filename="holidays_{datetime.now().strftime("%Y%m%d")}.ics"'
    return response


# =============================================================================
# 10. ACADEMIC PROGRESS
# =============================================================================

def get_filtered_progress(request):
    qs = AcademicProgress.objects.select_related('student', 'academic_session', 'class_enrollment').order_by('-academic_session__start_date', 'student__last_name')
    q = request.GET.get('q', '').strip()
    if q: qs = qs.filter(Q(student__first_name__icontains=q) | Q(student__last_name__icontains=q))
    for field, lookup in (('academic_session', 'academic_session_id'), ('progress_status', 'progress_status'), ('promotion_decision', 'promotion_decision')):
        v = request.GET.get(field, '')
        if v: qs = qs.filter(**{lookup: v})
    for bf in ('is_eligible_for_promotion', 'is_final'):
        v = request.GET.get(bf, '')
        if v: qs = qs.filter(**{bf: v.lower() == 'true'})
    for num_field, lookup in (('min_percentage', 'percentage__gte'), ('max_percentage', 'percentage__lte'), ('min_gpa', 'gpa__gte')):
        v = request.GET.get(num_field, '')
        if v:
            try: qs = qs.filter(**{lookup: Decimal(v)})
            except Exception: pass
    return qs


@login_required
def academic_progress_list(request):
    filter_form      = AcademicProgressFilterForm(request.GET or None)
    progress_records = get_filtered_progress(request)
    stats = {'total': progress_records.count(), 'finalized': progress_records.filter(is_final=True).count(), 'eligible_for_promotion': progress_records.filter(is_eligible_for_promotion=True).count()}
    paginator     = Paginator(progress_records, 20)
    progress_page = paginator.get_page(request.GET.get('page', 1))
    is_htmx       = request.headers.get('HX-Request') == 'true'
    context = {'progress_page': progress_page, 'paginator': paginator, 'stats': stats, 'filter_form': filter_form, 'is_htmx': is_htmx}
    if is_htmx:
        return render(request, 'academics/progress/partials/_progress_results.html', context)
    return render(request, 'academics/progress/list.html', context)


@login_required
def academic_progress_detail(request, pk):
    progress = get_object_or_404(AcademicProgress.objects.select_related('student', 'academic_session', 'class_enrollment'), pk=pk)
    return render(request, 'academics/progress/detail.html', {'progress': progress})


@login_required
def academic_progress_create(request, student_pk=None):
    initial = {}
    if student_pk: initial['student'] = get_object_or_404(Student, pk=student_pk)
    if request.method == 'POST':
        form = AcademicProgressForm(request.POST)
        if form.is_valid():
            try:
                progress = form.save()
                if request.headers.get('HX-Request') == 'true':
                    r = HttpResponse(); r['HX-Alert-Message'] = 'Progress record created successfully'; r['HX-Alert-Type'] = 'success'; r['HX-Close-Modal'] = 'true'; r['HX-Redirect'] = reverse('academics:progress_detail', kwargs={'pk': progress.pk}); return r
                messages.success(request, 'Progress record created successfully')
                return redirect('academics:progress_detail', pk=progress.pk)
            except Exception as e:
                if request.headers.get('HX-Request') == 'true':
                    r = HttpResponse(); r['HX-Alert-Message'] = f'Error: {e}'; r['HX-Alert-Type'] = 'error'; return r
                messages.error(request, f'Error: {e}')
        else:
            if request.headers.get('HX-Request') == 'true':
                r = HttpResponse(); r['HX-Alert-Message'] = 'Please correct the errors in the form'; r['HX-Alert-Type'] = 'error'; return r
            messages.error(request, 'Please correct the errors in the form')
    else:
        form = AcademicProgressForm(initial=initial)
    return render(request, 'academics/progress/form.html', {'form': form, 'title': 'Create progress record'})


@login_required
def academic_progress_edit(request, pk):
    progress = get_object_or_404(AcademicProgress, pk=pk)
    if progress.is_final:
        messages.error(request, 'Cannot edit finalized progress records')
        return redirect('academics:progress_detail', pk=pk)
    if request.method == 'POST':
        form = AcademicProgressForm(request.POST, instance=progress)
        if form.is_valid():
            try:
                progress = form.save()
                if request.headers.get('HX-Request') == 'true':
                    r = HttpResponse(); r['HX-Alert-Message'] = 'Progress record updated successfully'; r['HX-Alert-Type'] = 'success'; r['HX-Close-Modal'] = 'true'; r['HX-Redirect'] = reverse('academics:progress_detail', kwargs={'pk': progress.pk}); return r
                messages.success(request, 'Progress record updated successfully')
                return redirect('academics:progress_detail', pk=progress.pk)
            except Exception as e:
                if request.headers.get('HX-Request') == 'true':
                    r = HttpResponse(); r['HX-Alert-Message'] = f'Error: {e}'; r['HX-Alert-Type'] = 'error'; return r
                messages.error(request, f'Error: {e}')
        else:
            if request.headers.get('HX-Request') == 'true':
                r = HttpResponse(); r['HX-Alert-Message'] = 'Please correct the errors in the form'; r['HX-Alert-Type'] = 'error'; return r
            messages.error(request, 'Please correct the errors in the form')
    else:
        form = AcademicProgressForm(instance=progress)
    return render(request, 'academics/progress/form.html', {'form': form, 'progress': progress, 'title': 'Edit progress record'})


@login_required
def academic_progress_delete(request, pk):
    progress = get_object_or_404(AcademicProgress, pk=pk)
    if request.method == 'POST':
        if progress.is_final:
            if request.headers.get('HX-Request') == 'true':
                r = HttpResponse(); r['HX-Alert-Message'] = 'Cannot delete finalized progress records'; r['HX-Alert-Type'] = 'error'; r['HX-Close-Modal'] = 'true'; return r
            messages.error(request, 'Cannot delete finalized progress records')
            return redirect('academics:progress_detail', pk=pk)
        try:
            progress.delete()
            if request.headers.get('HX-Request') == 'true':
                r = HttpResponse(); r['HX-Alert-Message'] = 'Progress record deleted successfully'; r['HX-Alert-Type'] = 'success'; r['HX-Close-Modal'] = 'true'; r['HX-Redirect'] = reverse('academics:progress_list'); return r
            messages.success(request, 'Progress record deleted successfully')
            return redirect('academics:progress_list')
        except Exception as e:
            if request.headers.get('HX-Request') == 'true':
                r = HttpResponse(); r['HX-Alert-Message'] = f'Error: {e}'; r['HX-Alert-Type'] = 'error'; r['HX-Close-Modal'] = 'true'; return r
            messages.error(request, f'Error: {e}')
            return redirect('academics:progress_detail', pk=pk)
    return redirect('academics:progress_list')


@login_required
def academic_progress_finalize(request, pk):
    progress = get_object_or_404(AcademicProgress, pk=pk)
    if request.method == 'POST':
        try:
            finalized  = progress.finalize_record(user=request.user)
            msg        = 'Progress record finalized successfully' if finalized else 'Progress record is already finalized'
            alert_type = 'success' if finalized else 'warning'
            if request.headers.get('HX-Request') == 'true':
                r = HttpResponse(); r['HX-Alert-Message'] = msg; r['HX-Alert-Type'] = alert_type; r['HX-Close-Modal'] = 'true'; r['HX-Redirect'] = reverse('academics:progress_detail', kwargs={'pk': pk}); return r
            getattr(messages, alert_type)(request, msg)
        except Exception as e:
            if request.headers.get('HX-Request') == 'true':
                r = HttpResponse(); r['HX-Alert-Message'] = f'Error: {e}'; r['HX-Alert-Type'] = 'error'; r['HX-Close-Modal'] = 'true'; return r
            messages.error(request, f'Error: {e}')
    return redirect('academics:progress_detail', pk=pk)


@login_required
def academic_progress_update_promotion(request, pk):
    progress = get_object_or_404(AcademicProgress, pk=pk)
    if request.method == 'POST':
        decision = request.POST.get('promotion_decision')
        if not decision:
            if request.headers.get('HX-Request') == 'true':
                r = HttpResponse(); r['HX-Alert-Message'] = 'Please select a promotion decision'; r['HX-Alert-Type'] = 'error'; return r
            messages.error(request, 'Please select a promotion decision')
            return redirect('academics:progress_detail', pk=pk)
        try:
            progress.promotion_decision = decision
            if decision == 'PROMOTED':   progress.is_eligible_for_promotion = True
            elif decision == 'REPEAT':   progress.is_eligible_for_promotion = False
            progress.save()
            if request.headers.get('HX-Request') == 'true':
                r = HttpResponse(); r['HX-Alert-Message'] = 'Promotion decision updated successfully'; r['HX-Alert-Type'] = 'success'; r['HX-Close-Modal'] = 'true'; r['HX-Redirect'] = reverse('academics:progress_detail', kwargs={'pk': pk}); return r
            messages.success(request, 'Promotion decision updated successfully')
        except Exception as e:
            if request.headers.get('HX-Request') == 'true':
                r = HttpResponse(); r['HX-Alert-Message'] = f'Error: {e}'; r['HX-Alert-Type'] = 'error'; r['HX-Close-Modal'] = 'true'; return r
            messages.error(request, f'Error: {e}')
    return redirect('academics:progress_detail', pk=pk)


@login_required
def academic_progress_print_view(request):
    progress_records = get_filtered_progress(request)
    selected_fields  = request.GET.getlist('fields') or ['student_name', 'session_name', 'overall_grade', 'percentage', 'gpa', 'attendance_pct', 'is_eligible', 'promotion_decision', 'is_final']
    include_stats    = request.GET.get('include_stats') == 'true'
    stats = None
    if include_stats:
        stats = {'total': progress_records.count(), 'finalized': progress_records.filter(is_final=True).count(), 'eligible_for_promotion': progress_records.filter(is_eligible_for_promotion=True).count()}
    field_labels = {
        'student_name': 'Student', 'admission_number': 'Admission #',
        'session_name': 'Session', 'overall_grade': 'Grade',
        'gpa': 'GPA', 'percentage': 'Percentage', 'performance': 'Performance',
        'days_attended': 'Days Attended', 'total_days': 'Total Days',
        'attendance_pct': 'Attendance %', 'subjects_total': 'Subjects',
        'subjects_passed': 'Passed', 'subjects_failed': 'Failed',
        'progress_status': 'Progress Status', 'is_eligible': 'Promotion Eligible',
        'promotion_decision': 'Promotion Decision', 'is_final': 'Finalized',
    }
    return render(request, 'academics/progress/print.html', {
        **_get_print_school_context(request),
        'progress_records': progress_records, 'stats': stats, 'now': timezone.now(),
        'selected_fields': selected_fields,
        'selected_field_names': [field_labels.get(f, f.replace('_', ' ').title()) for f in selected_fields],
        'field_labels': field_labels, 'landscape': request.GET.get('landscape') == 'true',
        'title': 'Academic Progress Report',
    })


@login_required
def export_academic_progress_excel(request):
    ALL_COLUMNS = [
        ('student_name',      'Student Name',       lambda p: p.student.get_full_name()),
        ('admission_number',  'Admission #',        lambda p: p.student.admission_number),
        ('session_name',      'Session',            lambda p: p.academic_session.name),
        ('overall_grade',     'Grade',              lambda p: p.overall_grade or ''),
        ('gpa',               'GPA',                lambda p: float(p.gpa) if p.gpa else ''),
        ('percentage',        'Percentage',         lambda p: float(p.percentage) if p.percentage else ''),
        ('performance',       'Performance Level',  lambda p: p.performance_level),
        ('days_attended',     'Days Attended',      lambda p: p.days_attended),
        ('total_days',        'Total School Days',  lambda p: p.total_school_days),
        ('attendance_pct',    'Attendance %',       lambda p: float(p.attendance_percentage) if p.attendance_percentage else ''),
        ('subjects_total',    'Total Subjects',     lambda p: p.total_subjects),
        ('subjects_passed',   'Passed',             lambda p: p.subjects_passed),
        ('subjects_failed',   'Failed',             lambda p: p.subjects_failed),
        ('progress_status',   'Progress Status',    lambda p: p.get_progress_status_display() if p.progress_status else ''),
        ('is_eligible',       'Promotion Eligible', lambda p: 'Yes' if p.is_eligible_for_promotion else 'No'),
        ('promotion_decision','Promotion Decision', lambda p: p.get_promotion_decision_display()),
        ('is_final',          'Finalized',          lambda p: 'Yes' if p.is_final else 'No'),
        ('teacher_comments',  'Teacher Comments',   lambda p: p.teacher_comments or ''),
    ]
    DEFAULT = ['student_name', 'session_name', 'overall_grade', 'percentage', 'gpa', 'attendance_pct', 'is_eligible', 'promotion_decision', 'is_final']
    col_map  = {k: (lbl, fn) for k, lbl, fn in ALL_COLUMNS}
    chosen   = request.GET.getlist('fields') or DEFAULT
    columns  = [(col_map[f][0], col_map[f][1]) for f in chosen if f in col_map]
    progress = get_filtered_progress(request)
    wb = Workbook(); ws = wb.active; ws.title = 'Academic Progress'
    ws.append([c[0] for c in columns])
    hf = PatternFill(start_color='1E3A5F', end_color='1E3A5F', fill_type='solid')
    for cell in ws[1]:
        cell.fill = hf; cell.font = Font(bold=True, color='FFFFFF', size=11)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws.row_dimensions[1].height = 28
    af = PatternFill(start_color='F4F6F9', end_color='F4F6F9', fill_type='solid')
    da = Alignment(vertical='center', wrap_text=False)
    for i, obj in enumerate(progress):
        ws.append([c[1](obj) for c in columns])
        for cell in ws[i + 2]:
            cell.alignment = da
            if i % 2 == 1: cell.fill = af
    for cc in ws.columns:
        ml = max((len(str(c.value)) if c.value else 0) for c in cc)
        ws.column_dimensions[cc[0].column_letter].width = min(ml + 4, 60)
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="academic_progress_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
    wb.save(response); return response


@login_required
def academic_progress_report_card(request, pk):
    progress = get_object_or_404(AcademicProgress, pk=pk)
    return render(request, 'academics/progress/report_card.html', {
        **_get_print_school_context(request),
        'progress': progress, 'student': progress.student,
        'session': progress.academic_session, 'enrollment': progress.class_enrollment, 'now': timezone.now(),
    })


@login_required
def bulk_progress_finalize(request):
    if request.method == 'POST':
        try:
            progress_ids = request.POST.getlist('progress_ids')
            if not progress_ids:
                if request.headers.get('HX-Request') == 'true':
                    r = HttpResponse(); r['HX-Alert-Message'] = 'Please select progress records'; r['HX-Alert-Type'] = 'error'; return r
                messages.error(request, 'Please select progress records')
                return redirect('academics:progress_list')
            count = 0
            with transaction.atomic():
                for pid in progress_ids:
                    rec = get_object_or_404(AcademicProgress, pk=pid)
                    if not rec.is_final and rec.finalize_record(user=request.user): count += 1
            msg = f'Finalized {count} progress record(s)'
            if request.headers.get('HX-Request') == 'true':
                r = HttpResponse(); r['HX-Alert-Message'] = msg; r['HX-Alert-Type'] = 'success'; r['HX-Close-Modal'] = 'true'; r['HX-Redirect'] = reverse('academics:progress_list'); return r
            messages.success(request, msg)
        except Exception as e:
            if request.headers.get('HX-Request') == 'true':
                r = HttpResponse(); r['HX-Alert-Message'] = f'Error: {e}'; r['HX-Alert-Type'] = 'error'; r['HX-Close-Modal'] = 'true'; return r
            messages.error(request, f'Error: {e}')
    return redirect('academics:progress_list')


# =============================================================================
# 11. PROMOTIONS
# =============================================================================

@login_required
def promotion_dashboard(request):
    current_session   = AcademicSession.objects.filter(is_current=True).first()
    eligible_students = []
    if current_session:
        eligible_progress = AcademicProgress.objects.filter(academic_session=current_session, is_eligible_for_promotion=True, promotion_decision='PROMOTED').select_related('student', 'class_enrollment__class_instance__academic_level')
        eligible_students = [{'student': p.student, 'current_level': p.class_enrollment.class_instance.academic_level if p.class_enrollment else None, 'progress': p} for p in eligible_progress]
    level_stats = []
    for level in AcademicLevel.objects.filter(is_active=True).order_by('order'):
        total    = Student.objects.filter(current_academic_level=level, enrollment_status='ACTIVE').count()
        eligible = AcademicProgress.objects.filter(student__current_academic_level=level, academic_session=current_session, is_eligible_for_promotion=True).count() if current_session else 0
        level_stats.append({'level': level, 'total_students': total, 'eligible_for_promotion': eligible})
    return render(request, 'academics/promotions/dashboard.html', {'current_session': current_session, 'eligible_students': eligible_students, 'level_stats': level_stats, 'title': 'Student Promotions'})


@login_required
def promote_student(request):
    if request.method == 'POST':
        student_id = request.POST.get('student_id'); next_level_id = request.POST.get('next_level')
        if not student_id or not next_level_id:
            messages.error(request, 'Please select a student and target level')
            return redirect('academics:promote_student')
        try:
            student = get_object_or_404(Student, pk=student_id); next_level = get_object_or_404(AcademicLevel, pk=next_level_id)
            with transaction.atomic():
                if request.POST.get('complete_current_enrollment') == 'true':
                    completion_date = request.POST.get('completion_date') or get_school_today()
                    StudentClassEnrollment.objects.filter(student=student, is_active=True, completion_status='ONGOING').update(completion_status='COMPLETED', completion_date=completion_date, is_active=False)
                old_level = student.current_academic_level
                student.current_academic_level = next_level; student.save(update_fields=['current_academic_level'])
            messages.success(request, f'{student.get_full_name()} promoted from {old_level.name if old_level else "N/A"} to {next_level.name}')
        except Exception as e:
            logger.error(f"Error promoting student: {e}")
            messages.error(request, f'Promotion failed: {e}')
        return redirect('academics:promote_student')
    search_query  = request.GET.get('search', ''); level_filter  = request.GET.get('level', ''); status_filter = request.GET.get('status', 'ACTIVE')
    students = Student.objects.all()
    if search_query: students = students.filter(Q(first_name__icontains=search_query) | Q(last_name__icontains=search_query) | Q(admission_number__icontains=search_query))
    if level_filter:  students = students.filter(current_academic_level_id=level_filter)
    if status_filter: students = students.filter(enrollment_status=status_filter)
    students = students.select_related('current_academic_level').order_by('first_name', 'last_name')
    selected_student = current_enrollment = next_level_suggestion = None
    student_id = request.GET.get('student_id')
    if student_id:
        try:
            selected_student   = Student.objects.select_related('current_academic_level').get(pk=student_id)
            current_enrollment = StudentClassEnrollment.objects.filter(student=selected_student, is_active=True, completion_status='ONGOING').select_related('class_instance', 'academic_session').first()
            if selected_student.current_academic_level:
                next_level_suggestion = AcademicLevel.objects.filter(order=selected_student.current_academic_level.order + 1, is_active=True).first()
        except Student.DoesNotExist:
            messages.error(request, 'Student not found')
    paginator     = Paginator(students, 25)
    students_page = paginator.get_page(request.GET.get('page', 1))
    return render(request, 'academics/promotions/promote_student.html', {
        'students': students_page, 'selected_student': selected_student,
        'current_enrollment': current_enrollment, 'next_level_suggestion': next_level_suggestion,
        'available_levels': AcademicLevel.objects.filter(is_active=True).order_by('order'),
        'all_levels': AcademicLevel.objects.filter(is_active=True).order_by('order'),
        'search_query': search_query, 'level_filter': level_filter, 'status_filter': status_filter,
        'title': 'Promote student',
    })


@login_required
def bulk_promote_students(request):
    if request.method == 'POST':
        try:
            student_ids          = request.POST.getlist('student_ids')
            next_level_id        = request.POST.get('next_level')
            complete_enrollments = request.POST.get('complete_enrollments') == 'true'
            completion_date      = request.POST.get('completion_date') or get_school_today()
            if not student_ids or not next_level_id:
                messages.error(request, 'Please select students and a target level')
                return redirect('academics:bulk_promote_students')
            next_level = get_object_or_404(AcademicLevel, pk=next_level_id)
            students   = Student.objects.filter(id__in=student_ids)
            promoted   = 0; errors = []
            with transaction.atomic():
                for student in students:
                    try:
                        if complete_enrollments:
                            StudentClassEnrollment.objects.filter(student=student, is_active=True, completion_status='ONGOING').update(completion_status='COMPLETED', completion_date=completion_date, is_active=False)
                        student.current_academic_level = next_level; student.save(update_fields=['current_academic_level']); promoted += 1
                    except Exception as e:
                        errors.append(f'{student.get_full_name()}: {e}')
            msg = f'Successfully promoted {promoted} student(s) to {next_level.name}'
            if errors: messages.warning(request, msg + f'. {len(errors)} error(s) occurred')
            else:       messages.success(request, msg)
        except Exception as e:
            logger.error(f"Bulk promotion error: {e}")
            messages.error(request, f'Bulk promotion failed: {e}')
        return redirect('academics:bulk_promote_students')
    search_query  = request.GET.get('search', ''); level_filter  = request.GET.get('level', '')
    status_filter = request.GET.get('status', 'ACTIVE'); eligible_only = request.GET.get('eligible_only') == 'true'
    students = Student.objects.all()
    if search_query: students = students.filter(Q(first_name__icontains=search_query) | Q(last_name__icontains=search_query) | Q(admission_number__icontains=search_query))
    if level_filter:  students = students.filter(current_academic_level_id=level_filter)
    if status_filter: students = students.filter(enrollment_status=status_filter)
    if eligible_only:
        cs = AcademicSession.objects.filter(is_current=True).first()
        if cs:
            eligible_ids = AcademicProgress.objects.filter(academic_session=cs, is_eligible_for_promotion=True).values_list('student_id', flat=True)
            students = students.filter(id__in=eligible_ids)
    students = students.select_related('current_academic_level').order_by('current_academic_level__order', 'first_name', 'last_name')
    paginator     = Paginator(students, 50)
    students_page = paginator.get_page(request.GET.get('page', 1))
    return render(request, 'academics/promotions/bulk_promote.html', {
        'students': students_page, 'all_levels': AcademicLevel.objects.filter(is_active=True).order_by('order'),
        'all_sessions': AcademicSession.objects.filter(is_active=True).order_by('-start_date'),
        'target_levels': AcademicLevel.objects.filter(is_active=True).order_by('order'),
        'search_query': search_query, 'level_filter': level_filter,
        'status_filter': status_filter, 'eligible_only': eligible_only,
        'total_count': students.count(), 'title': 'Bulk promote students',
    })


# =============================================================================
# 12. AJAX / UTILITY ENDPOINTS
# =============================================================================

@login_required
def academic_calendar(request, year=None, month=None):
    from calendar import monthcalendar, month_name as _month_name
    if not year or not month:
        today = get_school_today(); year = today.year; month = today.month
    holidays = Holiday.objects.filter(start_date__year=year, start_date__month=month)
    return render(request, 'academics/calendar/view.html', {'year': year, 'month': month, 'month_name': _month_name[month], 'calendar': monthcalendar(year, month), 'holidays': holidays})


@login_required
def ajax_get_subjects_for_level(request, level_pk):
    try:
        get_object_or_404(AcademicLevel, pk=level_pk)
        subjects = list(Subject.objects.filter(is_active=True).values('id', 'name', 'code'))
        return JsonResponse({'success': True, 'subjects': subjects})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=400)


@login_required
def ajax_get_classes_for_session(request, session_pk):
    try:
        session = get_object_or_404(AcademicSession, pk=session_pk)
        classes = Class.objects.filter(academic_session=session, is_active=True).select_related('academic_level').values('id', 'academic_level__name', 'section')
        class_list = [{'id': c['id'], 'name': c['academic_level__name'] + (f' {c["section"]}' if c['section'] else '')} for c in classes]
        return JsonResponse({'success': True, 'classes': class_list})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=400)


@login_required
def ajax_get_next_roll_number(request, class_pk):
    try:
        ci   = get_object_or_404(Class, pk=class_pk)
        last = StudentClassEnrollment.objects.filter(class_instance=ci).exclude(roll_number='').order_by('-roll_number').first()
        if last and last.roll_number:
            try:    next_num = int(last.roll_number) + 1
            except ValueError: next_num = 1
        else:
            next_num = 1
        return JsonResponse({'success': True, 'next_roll_number': str(next_num).zfill(3)})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=400)


@login_required
def ajax_check_enrollment_duplicate(request):
    try:
        student_id = request.GET.get('student_id'); class_id = request.GET.get('class_id'); session_id = request.GET.get('session_id')
        if not all([student_id, class_id, session_id]):
            return JsonResponse({'success': False, 'error': 'Missing parameters'}, status=400)
        exists = StudentClassEnrollment.objects.filter(student_id=student_id, class_instance_id=class_id, academic_session_id=session_id).exists()
        return JsonResponse({'success': True, 'is_duplicate': exists})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=400)


@login_required
def ajax_get_class_subjects(request, class_pk):
    try:
        ci = get_object_or_404(Class, pk=class_pk)
        subjects = ClassSubject.objects.filter(class_instance=ci, is_active=True).select_related('subject', 'teacher__staff').values('id', 'subject__name', 'subject__code', 'teacher__staff__first_name', 'teacher__staff__last_name', 'is_optional')
        subject_list = [{'id': s['id'], 'name': s['subject__name'], 'code': s['subject__code'], 'teacher': f"{s['teacher__staff__first_name'] or ''} {s['teacher__staff__last_name'] or ''}".strip(), 'is_optional': s['is_optional']} for s in subjects]
        return JsonResponse({'success': True, 'subjects': subject_list})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=400)


# =============================================================================
# INDIVIDUAL ITEM PRINT DETAIL VIEWS
# =============================================================================

@login_required
def academic_session_print_detail(request, pk):
    session = get_object_or_404(AcademicSession, pk=pk)
    classes = session.classes.select_related('academic_level').filter(is_active=True)
    enrollments = session.student_class_enrollments.select_related('student', 'class_instance')
    stats = {
        'total_classes': classes.count(),
        'total_enrollments': enrollments.count(),
        'active_enrollments': enrollments.filter(is_active=True, completion_status='ONGOING').count(),
        'completed_enrollments': enrollments.filter(completion_status='COMPLETED').count(),
        'male_students': enrollments.filter(student__gender='M').count(),
        'female_students': enrollments.filter(student__gender='F').count(),
    }
    subjects_taught = ClassSubject.objects.filter(
        class_instance__academic_session=session, is_active=True
    ).values_list('subject__name', flat=True).distinct()
    return render(request, 'academics/sessions/print_detail.html', {
        **_get_print_school_context(request),
        'session': session, 'classes': classes, 'enrollments': enrollments,
        'stats': stats, 'subjects_taught': list(subjects_taught),
        'now': timezone.now(), 'title': f'Academic Session: {session.name}',
    })


@login_required
def subject_print_detail(request, pk):
    subject = get_object_or_404(Subject, pk=pk)
    class_assignments = ClassSubject.objects.filter(
        subject=subject, is_active=True
    ).select_related('class_instance', 'teacher__staff')
    stats = {
        'total_classes': class_assignments.count(),
        'with_teacher': class_assignments.filter(teacher__isnull=False).count(),
        'without_teacher': class_assignments.filter(teacher__isnull=True).count(),
        'total_students': sum(cs.class_instance.get_current_enrollment_count() for cs in class_assignments),
    }
    return render(request, 'academics/subjects/print_detail.html', {
        **_get_print_school_context(request),
        'subject': subject, 'class_assignments': class_assignments,
        'stats': stats, 'now': timezone.now(), 'title': f'Subject: {subject.name}',
    })


@login_required
def academic_level_print_detail(request, pk):
    level = get_object_or_404(AcademicLevel, pk=pk)
    classes   = level.classes.select_related('academic_session').filter(is_active=True)
    students  = Student.objects.filter(current_academic_level=level, enrollment_status='ACTIVE')
    stats = {
        'total_classes':   classes.count(),
        'total_students':  students.count(),
        'male_students':   students.filter(gender='M').count(),
        'female_students': students.filter(gender='F').count(),
    }
    return render(request, 'academics/levels/print_detail.html', {
        **_get_print_school_context(request),
        'level': level, 'classes': classes, 'students': students[:50],
        'stats': stats, 'now': timezone.now(), 'title': f'Academic Level: {level.name}',
    })


@login_required
def classroom_print_detail(request, pk):
    classroom       = get_object_or_404(ClassRoom, pk=pk)
    current_classes = classroom.assigned_classes.select_related('academic_level', 'academic_session').filter(is_active=True)
    current_students = sum(cls.get_current_enrollment_count() for cls in current_classes)
    utilization      = round((current_students / classroom.capacity * 100), 1) if classroom.capacity else 0
    stats = {
        'total_capacity':        classroom.capacity,
        'current_students':      current_students,
        'available_capacity':    max(0, classroom.capacity - current_students),
        'utilization_percentage': utilization,
        'assigned_classes':      current_classes.count(),
    }
    return render(request, 'academics/classrooms/print_detail.html', {
        **_get_print_school_context(request),
        'classroom': classroom, 'classes': current_classes,
        'stats': stats, 'now': timezone.now(), 'title': f'Classroom: {classroom.name}',
    })


@login_required
def class_print_detail(request, pk):
    ci = get_object_or_404(Class.objects.select_related('academic_level', 'academic_session', 'class_teacher', 'classroom'), pk=pk)
    enrollments = ci.enrollments.select_related('student').filter(is_active=True, completion_status='ONGOING').order_by('roll_number', 'student__last_name')
    subjects    = ci.subjects.select_related('subject', 'teacher__staff').filter(is_active=True)
    stats = {
        'total_students':     enrollments.count(),
        'male_students':      enrollments.filter(student__gender='M').count(),
        'female_students':    enrollments.filter(student__gender='F').count(),
        'total_subjects':     subjects.count(),
        'compulsory_subjects': subjects.filter(is_optional=False).count(),
        'optional_subjects':  subjects.filter(is_optional=True).count(),
        'capacity_used':      round((enrollments.count() / ci.max_students * 100), 1) if ci.max_students else 0,
    }
    return render(request, 'academics/classes/print_detail.html', {
        **_get_print_school_context(request),
        'class': ci, 'enrollments': enrollments, 'subjects': subjects,
        'stats': stats, 'now': timezone.now(), 'title': f'Class: {ci}',
    })


@login_required
def enrollment_print_detail(request, pk):
    enrollment = get_object_or_404(StudentClassEnrollment.objects.select_related('student', 'class_instance', 'academic_session', 'academic_invoice'), pk=pk)
    history    = StudentClassEnrollment.objects.filter(student=enrollment.student).select_related('class_instance', 'academic_session').order_by('-enrollment_date')
    progress   = AcademicProgress.objects.filter(student=enrollment.student, academic_session=enrollment.academic_session).first()
    return render(request, 'academics/enrollments/print_detail.html', {
        **_get_print_school_context(request),
        'enrollment': enrollment, 'enrollment_history': history,
        'progress': progress, 'now': timezone.now(),
        'title': f'Enrollment: {enrollment.student.get_full_name()}',
    })


@login_required
def class_subject_print_detail(request, pk):
    cs = get_object_or_404(ClassSubject.objects.select_related('class_instance__academic_level', 'subject', 'teacher__staff'), pk=pk)
    enrolled_students = cs.class_instance.enrollments.select_related('student').filter(is_active=True, completion_status='ONGOING')
    stats = {
        'enrolled_students': enrolled_students.count(),
        'hours_per_week':    cs.hours_per_week,
        'total_hours':       cs.total_hours or 0,
    }
    return render(request, 'academics/class_subjects/print_detail.html', {
        **_get_print_school_context(request),
        'class_subject': cs, 'enrolled_students': enrolled_students,
        'stats': stats, 'now': timezone.now(),
        'title': f'{cs.subject.name} — {cs.class_instance}',
    })


@login_required
def academic_progress_print_detail(request, pk):
    progress = get_object_or_404(AcademicProgress.objects.select_related('student', 'academic_session', 'class_enrollment'), pk=pk)
    return render(request, 'academics/progress/print_detail.html', {
        **_get_print_school_context(request),
        'progress': progress, 'student': progress.student,
        'session': progress.academic_session, 'enrollment': progress.class_enrollment,
        'now': timezone.now(), 'title': f'Progress Report: {progress.student.get_full_name()}',
    })


@login_required
def holiday_print_detail(request, pk):
    holiday          = get_object_or_404(Holiday, pk=pk)
    affected_sessions = AcademicSession.objects.filter(
        start_date__lte=holiday.end_date or holiday.start_date,
        end_date__gte=holiday.start_date,
        is_active=True,
    )
    return render(request, 'academics/holidays/print_detail.html', {
        **_get_print_school_context(request),
        'holiday': holiday, 'duration': holiday.duration_days,
        'affected_sessions': affected_sessions,
        'now': timezone.now(), 'title': f'Holiday: {holiday.name}',
    })


# =============================================================================
# ACADEMIC PROGRESS — EXTRA BULK / ALIAS VIEWS
# =============================================================================

@login_required
def academic_progress_list_print_view(request):
    """Alias — same as academic_progress_print_view, retained for URL compatibility."""
    return academic_progress_print_view(request)


@login_required
def bulk_progress_calculate(request):
    if request.method == 'POST':
        try:
            progress_ids = request.POST.getlist('progress_ids')
            if not progress_ids:
                messages.error(request, 'Please select progress records')
                return redirect('academics:progress_list')
            calculated = 0
            with transaction.atomic():
                for pid in progress_ids:
                    rec = get_object_or_404(AcademicProgress, pk=pid)
                    if hasattr(rec, 'calculate_metrics'):
                        rec.calculate_metrics()
                        calculated += 1
            messages.success(request, f'Calculated metrics for {calculated} record(s)')
        except Exception as e:
            logger.error(f"Error in bulk calculation: {e}")
            messages.error(request, f'Error: {e}')
    return redirect('academics:progress_list')


# =============================================================================
# REPORT VIEWS
# =============================================================================

@login_required
def session_summary_report(request):
    session_id = request.GET.get('session_id')
    if not session_id:
        messages.error(request, 'Please select a session')
        return redirect('academics:academics_dashboard')
    session     = get_object_or_404(AcademicSession, pk=session_id)
    classes     = Class.objects.filter(academic_session=session).select_related('academic_level')
    enrollments = StudentClassEnrollment.objects.filter(academic_session=session).select_related('student', 'class_instance')
    stats = {
        'total_classes':         classes.count(),
        'total_enrollments':     enrollments.count(),
        'active_enrollments':    enrollments.filter(is_active=True, completion_status='ONGOING').count(),
        'completed_enrollments': enrollments.filter(completion_status='COMPLETED').count(),
        'male_students':         enrollments.filter(student__gender='M').count(),
        'female_students':       enrollments.filter(student__gender='F').count(),
    }
    return render(request, 'academics/reports/session_summary.html', {
        'session': session, 'classes': classes, 'stats': stats,
    })


@login_required
def class_roster_report(request, class_pk):
    ci = get_object_or_404(Class, pk=class_pk)
    enrollments = ci.enrollments.select_related('student').filter(
        is_active=True, completion_status='ONGOING'
    ).order_by('roll_number', 'student__last_name')
    return render(request, 'academics/reports/class_roster.html', {
        **_get_print_school_context(request),
        'class': ci, 'enrollments': enrollments, 'now': timezone.now(),
    })


@login_required
def teacher_assignment_report(request):
    session_id     = request.GET.get('session_id', '')
    class_subjects = ClassSubject.objects.select_related(
        'class_instance__academic_session', 'subject', 'teacher__staff'
    ).filter(is_active=True)
    if session_id:
        class_subjects = class_subjects.filter(class_instance__academic_session_id=session_id)
    by_teacher = class_subjects.values(
        'teacher__staff__first_name', 'teacher__staff__last_name'
    ).annotate(total_assignments=Count('id'))
    return render(request, 'academics/reports/teacher_assignment.html', {
        'class_subjects': class_subjects[:200], 'by_teacher': by_teacher,
    })