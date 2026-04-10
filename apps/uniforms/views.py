# uniforms/views.py

"""
Uniform Management Views

ARCHITECTURE
------------
1. LIGHTWEIGHT DETAIL VIEWS + HTMX TAB PARTIALS
   uniform_item_detail and uniform_sale_detail are shells that fetch only
   the parent object and lightweight counts. Heavy relational data is loaded
   on demand through tab partial views.

   Tab partials:
     uniform_item_stock_partial     — Stock Records tab
     uniform_item_sales_partial     — Sale History tab
     uniform_item_size_recs_partial — Size Recommendations tab
     uniform_sale_items_partial     — Line Items tab
     uniform_sale_audit_partial     — Audit Trail tab

2. SETTINGS-STYLE SIMPLE VIEWS
   uniform_size_list and measurement_type_list are plain config tables
   (≤20 rows each). No filter bar, no pagination, no stat cards, no HTMX.

3. OOB STAT CARD REFRESH
   List views pass `stats` and `is_htmx` into their partial context.
   Partial templates render `hx-swap-oob="true"` stat blocks when filtering.

4. BANNER WARNINGS
   uniform_item_stock_partial exposes `missing_size_stock` — sizes that
   are in available_sizes but have no stock record yet.
   uniform_sale_list passes `unissued_count` — PAID/PARTIAL sales not issued.

5. PRINT VIEWS — FIELD SELECTION
   All list print views accept a `fields` GET list so the caller (a modal
   with checkboxes) controls which columns appear. Pattern from students/views.py.

6. EXPORT VIEWS — COLUMN SELECTION
   All Excel exports use an ALL_COLUMNS / COLUMN_MAP / DEFAULT_FIELDS
   pattern. Column order follows the user's selection. Pattern from students/views.py.

7. StudentUniformSize
   No list view — recommendations are accessed via the Size Recommendations
   tab on uniform_item_detail or from the student profile. Only
   create / detail / edit / delete are kept.
"""

from django.shortcuts import render, redirect, get_object_or_404
from django.urls import reverse
from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator
from django.contrib import messages
from django.db.models import Q, Count, Sum, Avg, F
from django.utils import timezone
from django.http import JsonResponse, HttpResponse
from django.db import transaction
from django.core.exceptions import ValidationError
from datetime import datetime
from decimal import Decimal
import logging

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from core.utils import (
    get_school_today,
    get_school_current_time,
    get_active_academic_session,
    format_money,
    calculate_percentage,
)
from core.view_helpers import get_print_school_context

from .models import (
    MeasurementType,
    StudentMeasurement,
    UniformSize,
    UniformItem,
    UniformStock,
    UniformPurchaseOrder,
    UniformPurchaseOrderItem,
    UniformSale,
    UniformSaleItem,
    StudentUniformSize,
)
from .forms import (
    MeasurementTypeForm,
    StudentMeasurementForm,
    StudentMeasurementFilterForm,
    BulkMeasurementForm,
    UniformSizeForm,
    UniformItemForm,
    UniformItemFilterForm,
    UniformStockForm,
    UniformStockFilterForm,
    UniformPurchaseOrderForm,
    UniformPurchaseOrderFilterForm,
    UniformSaleForm,
    UniformSaleFilterForm,
    StudentUniformSizeForm,
)
from .services import UniformSaleWorkflowService
from .utils import cancel_uniform_sale, return_uniform_sale
from . import stats as uniform_stats

from students.models import Student
from academics.models import AcademicSession, Class

logger = logging.getLogger(__name__)


# =============================================================================
# DASHBOARD
# =============================================================================

@login_required
def uniforms_dashboard(request):
    """
    Uniforms module dashboard.

    Stat cards come from direct queries (fast, always current).
    Chart data comes from stats.py functions which are designed for
    serialisation and can be cached independently if needed.
    """
    import json
    from datetime import timedelta
    from django.db.models import Sum, F

    today       = get_school_today()
    month_start = today.replace(day=1)

    # ── Stat card queries ─────────────────────────────────────────────────────
    items_qs        = UniformItem.objects.filter(is_active=True)
    total_items     = items_qs.count()
    low_stock_count = items_qs.filter(current_stock__lte=F('reorder_level')).count()

    total_stock_value = (
        UniformStock.objects
        .aggregate(total=Sum('total_cost_value'))['total'] or 0
    )

    active_sales_qs = UniformSale.objects.filter(cancelled=False, returned=False)
    monthly_qs      = active_sales_qs.filter(sale_date__gte=month_start)

    revenue_this_month     = monthly_qs.aggregate(t=Sum('total_amount'))['t'] or 0
    sales_count_this_month = monthly_qs.count()

    outstanding_balance = (
        active_sales_qs
        .filter(status__in=['PENDING', 'PARTIAL', 'ISSUED'])
        .aggregate(t=Sum('balance'))['t'] or 0
    )
    pending_sales_count = active_sales_qs.filter(
        status__in=['DRAFT', 'PENDING', 'PARTIAL']
    ).count()

    pending_pos_count = UniformPurchaseOrder.objects.filter(
        status__in=['SUBMITTED', 'APPROVED', 'ORDERED']
    ).count()

    # ── Recent activity tables ────────────────────────────────────────────────
    recent_sales = (
        UniformSale.objects
        .select_related('student', 'fiscal_period')
        .order_by('-sale_date', '-created_at')[:8]
    )
    recent_pos = (
        UniformPurchaseOrder.objects
        .order_by('-order_date')[:5]
    )
    low_stock_list = (
        UniformItem.objects
        .filter(is_active=True, current_stock__lte=F('reorder_level'))
        .order_by('current_stock')[:10]
    )

    # ── Chart data (from stats.py) ────────────────────────────────────────────

    # 1. 30-day sales trend (area chart)
    trend_start = today - timedelta(days=29)
    trend_data  = uniform_stats.get_sales_trend(trend_start, today, group_by='day')
    chart_trend_labels  = json.dumps([r['period'] for r in trend_data])
    chart_trend_revenue = json.dumps([float(r['revenue']) for r in trend_data])
    chart_trend_counts  = json.dumps([r['sale_count'] for r in trend_data])

    # 2. Revenue by item type (donut chart) — this month
    by_type_data = uniform_stats.get_sales_by_item_type(
        date_from=month_start, date_to=today
    )
    chart_type_labels  = json.dumps([r['item_type_display'] for r in by_type_data])
    chart_type_revenue = json.dumps([float(r['total_revenue']) for r in by_type_data])

    # 3. Top 5 selling items — this month (horizontal bar chart)
    top_sellers = uniform_stats.get_top_selling_items(
        limit=5, date_from=month_start, date_to=today
    )
    chart_seller_names = json.dumps([r['item_name'] for r in top_sellers])
    chart_seller_qty   = json.dumps([r['quantity_sold'] for r in top_sellers])
    chart_seller_rev   = json.dumps([float(r['total_revenue']) for r in top_sellers])

    return render(request, 'uniforms/dashboard.html', {
        # Stat cards
        'total_items':            total_items,
        'low_stock_count':        low_stock_count,
        'total_stock_value':      total_stock_value,
        'revenue_this_month':     revenue_this_month,
        'sales_count_this_month': sales_count_this_month,
        'outstanding_balance':    outstanding_balance,
        'pending_sales_count':    pending_sales_count,
        'pending_pos_count':      pending_pos_count,
        # Tables
        'recent_sales':   recent_sales,
        'recent_pos':     recent_pos,
        'low_stock_list': low_stock_list,
        'today':          today,
        # Chart data (pre-serialised JSON strings for inline <script>)
        'chart_trend_labels':   chart_trend_labels,
        'chart_trend_revenue':  chart_trend_revenue,
        'chart_trend_counts':   chart_trend_counts,
        'chart_type_labels':    chart_type_labels,
        'chart_type_revenue':   chart_type_revenue,
        'chart_seller_names':   chart_seller_names,
        'chart_seller_qty':     chart_seller_qty,
        'chart_seller_rev':     chart_seller_rev,
        'trend_has_data':       len(trend_data) > 0,
        'type_has_data':        len(by_type_data) > 0,
        'sellers_has_data':     len(top_sellers) > 0,
    })


# =============================================================================
# FILTER HELPERS
# =============================================================================

def get_filtered_student_measurements(request):
    qs = StudentMeasurement.objects.select_related(
        'student__current_academic_level',
        'measurement_type__unit',
        'academic_session',
    ).order_by('-measurement_date', 'student__first_name')

    q                   = request.GET.get('q', '').strip()
    student             = request.GET.get('student', '')
    measurement_type    = request.GET.get('measurement_type', '')
    academic_session    = request.GET.get('academic_session', '')
    measurement_context = request.GET.get('measurement_context', '')
    is_verified         = request.GET.get('is_verified', '')
    is_current          = request.GET.get('is_current', '')
    date_from           = request.GET.get('measurement_date_from', '')
    date_to             = request.GET.get('measurement_date_to', '')

    if q:
        qs = qs.filter(
            Q(student__first_name__icontains=q) |
            Q(student__last_name__icontains=q) |
            Q(student__admission_number__icontains=q) |
            Q(notes__icontains=q)
        )
    if student:             qs = qs.filter(student_id=student)
    if measurement_type:    qs = qs.filter(measurement_type_id=measurement_type)
    if academic_session:    qs = qs.filter(academic_session_id=academic_session)
    if measurement_context: qs = qs.filter(measurement_context=measurement_context)
    if is_verified:         qs = qs.filter(is_verified=(is_verified.lower() == 'true'))
    if is_current:          qs = qs.filter(is_current=(is_current.lower() == 'true'))
    if date_from:           qs = qs.filter(measurement_date__gte=date_from)
    if date_to:             qs = qs.filter(measurement_date__lte=date_to)
    return qs


def get_filtered_uniform_items(request):
    """
    NOTE: category filter removed — field no longer exists on UniformItem.
    Use item_type for all category-based filtering.
    """
    qs = UniformItem.objects.select_related(
        'unit_of_measure', 'tax_rate'
    ).prefetch_related('available_sizes').annotate(
        size_count=Count('available_sizes', distinct=True)
    ).order_by('item_type', 'name')

    q            = request.GET.get('q', '').strip()
    item_type    = request.GET.get('item_type', '')
    gender       = request.GET.get('gender', '')
    is_active    = request.GET.get('is_active', '')
    is_mandatory = request.GET.get('is_mandatory', '')
    stock_status = request.GET.get('stock_status', '')

    if q:
        qs = qs.filter(
            Q(name__icontains=q) |
            Q(code__icontains=q) |
            Q(sku__icontains=q) |
            Q(description__icontains=q)
        )
    if item_type:    qs = qs.filter(item_type=item_type)
    if gender:       qs = qs.filter(gender=gender)
    if is_active:    qs = qs.filter(is_active=(is_active.lower() == 'true'))
    if is_mandatory: qs = qs.filter(is_mandatory=(is_mandatory.lower() == 'true'))
    if stock_status == 'low_stock':
        qs = qs.filter(current_stock__lte=F('reorder_level'))
    elif stock_status == 'out_of_stock':
        qs = qs.filter(current_stock=0)
    elif stock_status == 'in_stock':
        qs = qs.filter(current_stock__gt=F('reorder_level'))
    return qs


def get_filtered_uniform_sizes(request):
    """Used by the size chart modal and export — not the list view."""
    qs = UniformSize.objects.annotate(
        item_count=Count('uniform_items', distinct=True)
    ).order_by('display_order', 'name')

    q         = request.GET.get('q', '').strip()
    size_type = request.GET.get('size_type', '')
    is_active = request.GET.get('is_active', '')

    if q:        qs = qs.filter(Q(name__icontains=q) | Q(code__icontains=q))
    if size_type:qs = qs.filter(size_type=size_type)
    if is_active:qs = qs.filter(is_active=(is_active.lower() == 'true'))
    return qs


def get_filtered_purchase_orders(request):
    qs = UniformPurchaseOrder.objects.select_related(
        'fiscal_period', 'journal_entry'
    ).prefetch_related('items').order_by('-order_date')

    q             = request.GET.get('q', '').strip()
    status        = request.GET.get('status', '')
    fiscal_period = request.GET.get('fiscal_period', '')
    date_from     = request.GET.get('order_date_from', '')
    date_to       = request.GET.get('order_date_to', '')

    if q:
        qs = qs.filter(
            Q(po_number__icontains=q) |
            Q(supplier_name__icontains=q) |
            Q(notes__icontains=q)
        )
    if status:        qs = qs.filter(status=status)
    if fiscal_period: qs = qs.filter(fiscal_period_id=fiscal_period)
    if date_from:     qs = qs.filter(order_date__gte=date_from)
    if date_to:       qs = qs.filter(order_date__lte=date_to)
    return qs


def get_filtered_uniform_sales(request):
    """academic_session derives via fiscal_period__related_academic_session."""
    qs = UniformSale.objects.select_related(
        'student__current_academic_level',
        'fiscal_period__related_academic_session',
        'fee_invoice',
        'payment_method',
    ).prefetch_related('items').order_by('-sale_date')

    q                = request.GET.get('q', '').strip()
    status           = request.GET.get('status', '')
    sale_type        = request.GET.get('sale_type', '')
    student          = request.GET.get('student', '')
    academic_session = request.GET.get('academic_session', '')
    fiscal_period    = request.GET.get('fiscal_period', '')
    date_from        = request.GET.get('sale_date_from', '')
    date_to          = request.GET.get('sale_date_to', '')

    if q:
        qs = qs.filter(
            Q(sale_number__icontains=q) |
            Q(student__first_name__icontains=q) |
            Q(student__last_name__icontains=q) |
            Q(student__admission_number__icontains=q)
        )
    if status:           qs = qs.filter(status=status)
    if sale_type:        qs = qs.filter(sale_type=sale_type)
    if student:          qs = qs.filter(student_id=student)
    if academic_session: qs = qs.filter(
        fiscal_period__related_academic_session_id=academic_session
    )
    if fiscal_period: qs = qs.filter(fiscal_period_id=fiscal_period)
    if date_from:     qs = qs.filter(sale_date__gte=date_from)
    if date_to:       qs = qs.filter(sale_date__lte=date_to)
    return qs


# =============================================================================
# HTMX HELPER
# =============================================================================

def _htmx_response(request, alert_message, alert_type='success',
                   redirect_url=None, close_modal=True):
    """
    Return an HttpResponse with HX-* headers for HTMX requests, or None
    for standard requests (caller must handle the redirect in that case).
    """
    if request.headers.get('HX-Request') == 'true':
        response = HttpResponse()
        response['HX-Alert-Message'] = alert_message
        response['HX-Alert-Type']    = alert_type
        if close_modal:
            response['HX-Close-Modal'] = 'true'
        if redirect_url:
            response['HX-Redirect'] = redirect_url
        return response
    if alert_type == 'success':
        messages.success(request, alert_message)
    else:
        messages.error(request, alert_message)
    return None


# =============================================================================
# EXCEL HELPERS
# =============================================================================

def _apply_header_style(ws):
    fill  = PatternFill(start_color='1E3A5F', end_color='1E3A5F', fill_type='solid')
    font  = Font(bold=True, color='FFFFFF', size=11)
    align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    for cell in ws[1]:
        cell.fill = fill
        cell.font = font
        cell.alignment = align
    ws.row_dimensions[1].height = 28


def _auto_fit_columns(ws):
    for col in ws.columns:
        width = max((len(str(c.value)) if c.value else 0) for c in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(width + 4, 60)


def _excel_response(wb, prefix):
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"{prefix}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


def _build_workbook(title, columns, rows):
    """
    Build a workbook from a list of (header, value) column specs and a
    list of row callables.  columns is [(header_str, callable(obj)), ...].
    rows is the iterable of objects.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = title
    ws.append([col[0] for col in columns])
    _apply_header_style(ws)
    for obj in rows:
        ws.append([col[1](obj) for col in columns])
    _auto_fit_columns(ws)
    return wb


# =============================================================================
# MEASUREMENT TYPE VIEWS  (settings-style — no filter bar, no pagination)
# =============================================================================

@login_required
def measurement_type_list(request):
    """
    Settings-style view. Measurement types are a small configuration table
    (~10-20 rows). No filter bar, no pagination, no stat cards needed.
    """
    measurement_types = MeasurementType.objects.select_related('unit').order_by(
        'category', 'display_order', 'name'
    )
    return render(request, 'uniforms/measurement_types/list.html', {
        'measurement_types': measurement_types,
    })


@login_required
def measurement_type_detail(request, pk):
    mt     = get_object_or_404(MeasurementType, pk=pk)
    recent = mt.student_measurements.select_related(
        'student', 'academic_session'
    ).order_by('-measurement_date')[:20]

    return render(request, 'uniforms/measurement_types/detail.html', {
        'measurement_type':    mt,
        'recent_measurements': recent,
        'measurement_count':   mt.student_measurements.count(),
        'verified_count':      mt.student_measurements.filter(is_verified=True).count(),
        'avg_value':           mt.student_measurements.filter(
                                   is_current=True
                               ).aggregate(Avg('value'))['value__avg'],
    })


@login_required
def measurement_type_create(request):
    if request.method == 'POST':
        form = MeasurementTypeForm(request.POST)
        if form.is_valid():
            try:
                mt = form.save()
                messages.success(request, f'Measurement type "{mt.name}" created')
                return redirect('uniforms:measurement_type_detail', pk=mt.pk)
            except Exception as e:
                logger.error(f"Error creating measurement type: {e}")
                messages.error(request, f'Error: {e}')
        else:
            messages.error(request, 'Please correct the errors below')
    else:
        form = MeasurementTypeForm()

    return render(request, 'uniforms/measurement_types/form.html', {
        'form': form, 'title': 'Create Measurement Type', 'submit_text': 'Create',
    })


@login_required
def measurement_type_edit(request, pk):
    mt = get_object_or_404(MeasurementType, pk=pk)

    if request.method == 'POST':
        form = MeasurementTypeForm(request.POST, instance=mt)
        if form.is_valid():
            try:
                mt = form.save()
                messages.success(request, f'Measurement type "{mt.name}" updated')
                return redirect('uniforms:measurement_type_detail', pk=mt.pk)
            except Exception as e:
                logger.error(f"Error updating measurement type: {e}")
                messages.error(request, f'Error: {e}')
        else:
            messages.error(request, 'Please correct the errors below')
    else:
        form = MeasurementTypeForm(instance=mt)

    return render(request, 'uniforms/measurement_types/form.html', {
        'form': form, 'measurement_type': mt,
        'title': f'Edit {mt.name}', 'submit_text': 'Update',
    })


@login_required
def measurement_type_delete(request, pk):
    mt = get_object_or_404(MeasurementType, pk=pk)
    if request.method == 'POST':
        if mt.student_measurements.exists():
            resp = _htmx_response(
                request,
                'Cannot delete — existing measurements reference this type',
                'error',
            )
            return resp or redirect('uniforms:measurement_type_detail', pk=pk)
        try:
            name = mt.name
            mt.delete()
            resp = _htmx_response(
                request, f'Measurement type "{name}" deleted',
                redirect_url=reverse('uniforms:measurement_type_list'),
            )
            return resp or redirect('uniforms:measurement_type_list')
        except Exception as e:
            resp = _htmx_response(request, f'Error: {e}', 'error')
            return resp or redirect('uniforms:measurement_type_detail', pk=pk)


@login_required
def measurement_type_toggle_active(request, pk):
    mt = get_object_or_404(MeasurementType, pk=pk)
    if request.method == 'POST':
        try:
            mt.is_active = not mt.is_active
            mt.save()
            verb = 'activated' if mt.is_active else 'deactivated'
            resp = _htmx_response(
                request, f'"{mt.name}" {verb}',
                redirect_url=reverse('uniforms:measurement_type_detail', kwargs={'pk': pk}),
            )
            return resp or redirect('uniforms:measurement_type_detail', pk=pk)
        except Exception as e:
            resp = _htmx_response(request, f'Error: {e}', 'error')
            return resp or redirect('uniforms:measurement_type_detail', pk=pk)


# =============================================================================
# STUDENT MEASUREMENT VIEWS
# =============================================================================

@login_required
def student_measurement_list(request):
    measurements = get_filtered_student_measurements(request)

    stats = {
        'total':      measurements.count(),
        'current':    measurements.filter(is_current=True).count(),
        'verified':   measurements.filter(is_verified=True).count(),
        'unverified': measurements.filter(is_verified=False).count(),
    }

    paginator = Paginator(measurements, 20)
    page      = paginator.get_page(request.GET.get('page', 1))
    is_htmx   = request.headers.get('HX-Request') == 'true'

    context = {
        'measurements_page': page,
        'paginator':         paginator,
        'stats':             stats,
        'filter_form':       StudentMeasurementFilterForm(request.GET or None),
        'is_htmx':           is_htmx,
    }
    if is_htmx:
        return render(request, 'uniforms/measurements/partials/_measurement_results.html', context)
    return render(request, 'uniforms/measurements/list.html', context)


@login_required
def student_measurement_create(request):
    student_id = request.GET.get('student')
    student    = get_object_or_404(Student, pk=student_id) if student_id else None

    if request.method == 'POST':
        form = StudentMeasurementForm(request.POST, student=student)
        if form.is_valid():
            try:
                m = form.save()
                messages.success(request, f'Measurement recorded for {m.student.get_full_name()}')
                return redirect('uniforms:student_measurement_list')
            except Exception as e:
                logger.error(f"Error creating measurement: {e}")
                messages.error(request, f'Error: {e}')
        else:
            messages.error(request, 'Please correct the errors below')
    else:
        form = StudentMeasurementForm(student=student)

    return render(request, 'uniforms/measurements/form.html', {
        'form': form, 'title': 'Record Measurement', 'submit_text': 'Record',
    })


@login_required
def student_measurement_detail(request, pk):
    m = get_object_or_404(
        StudentMeasurement.objects.select_related(
            'student__current_academic_level',
            'measurement_type__unit',
            'academic_session',
        ),
        pk=pk,
    )
    other = StudentMeasurement.objects.filter(
        student=m.student, measurement_type=m.measurement_type
    ).exclude(pk=pk).order_by('-measurement_date')[:5]

    return render(request, 'uniforms/measurements/detail.html', {
        'measurement':        m,
        'other_measurements': other,
        'verified_by':        m.get_verified_by_user(),
    })


@login_required
def student_measurement_edit(request, pk):
    m = get_object_or_404(StudentMeasurement, pk=pk)

    if request.method == 'POST':
        form = StudentMeasurementForm(request.POST, instance=m)
        if form.is_valid():
            try:
                m = form.save()
                messages.success(request, f'Measurement updated for {m.student.get_full_name()}')
                return redirect('uniforms:student_measurement_detail', pk=m.pk)
            except Exception as e:
                messages.error(request, f'Error: {e}')
        else:
            messages.error(request, 'Please correct the errors below')
    else:
        form = StudentMeasurementForm(instance=m)

    return render(request, 'uniforms/measurements/form.html', {
        'form': form, 'measurement': m,
        'title': 'Edit Measurement', 'submit_text': 'Update',
    })


@login_required
def student_measurement_delete(request, pk):
    m = get_object_or_404(StudentMeasurement, pk=pk)
    if request.method == 'POST':
        try:
            name = m.student.get_full_name()
            m.delete()
            resp = _htmx_response(
                request, f'Measurement deleted for {name}',
                redirect_url=reverse('uniforms:student_measurement_list'),
            )
            return resp or redirect('uniforms:student_measurement_list')
        except Exception as e:
            resp = _htmx_response(request, f'Error: {e}', 'error')
            return resp or redirect('uniforms:student_measurement_detail', pk=pk)


@login_required
def student_measurement_verify(request, pk):
    m = get_object_or_404(StudentMeasurement, pk=pk)
    if request.method == 'POST':
        try:
            with transaction.atomic():
                m.is_verified       = True
                m.verified_by_id    = str(request.user.id)
                m.verification_date = get_school_current_time()
                m.save()
            resp = _htmx_response(
                request, 'Measurement verified',
                redirect_url=reverse(
                    'uniforms:student_measurement_detail', kwargs={'pk': pk}
                ),
            )
            return resp or redirect('uniforms:student_measurement_detail', pk=pk)
        except Exception as e:
            resp = _htmx_response(request, f'Error: {e}', 'error')
            return resp or redirect('uniforms:student_measurement_detail', pk=pk)


@login_required
def student_measurement_bulk_create(request):
    if request.method == 'POST':
        form = BulkMeasurementForm(request.POST)
        if form.is_valid():
            messages.info(request, 'Bulk measurement entry is not yet implemented.')
            return redirect('uniforms:student_measurement_list')
        else:
            messages.error(request, 'Please correct the errors below')
    else:
        form = BulkMeasurementForm()

    return render(request, 'uniforms/measurements/bulk_form.html', {
        'form': form, 'title': 'Bulk Measurement Setup', 'submit_text': 'Continue',
    })


# =============================================================================
# UNIFORM SIZE VIEWS  (settings-style — no filter bar, no pagination)
# =============================================================================

@login_required
def uniform_size_list(request):
    """
    Settings-style view. Sizes are a small config table (10–20 rows).
    No filter bar, no pagination, no stat cards.
    """
    sizes = UniformSize.objects.all().order_by('display_order', 'name')
    return render(request, 'uniforms/sizes/list.html', {'sizes': sizes})


@login_required
def uniform_size_detail(request, pk):
    size          = get_object_or_404(UniformSize, pk=pk)
    items         = size.uniform_items.filter(is_active=True).order_by('name')
    stock_records = size.stock_records.select_related('uniform_item').order_by(
        'uniform_item__name'
    )
    total_stock   = stock_records.aggregate(Sum('quantity'))['quantity__sum'] or 0

    return render(request, 'uniforms/sizes/detail.html', {
        'size':          size,
        'items':         items,
        'stock_records': stock_records,
        'total_stock':   total_stock,
        'total_value':   sum(sr.total_cost_value for sr in stock_records),
    })


@login_required
def uniform_size_create(request):
    if request.method == 'POST':
        form = UniformSizeForm(request.POST)
        if form.is_valid():
            try:
                size = form.save()
                messages.success(request, f'Size "{size.name}" created')
                return redirect('uniforms:uniform_size_list')
            except Exception as e:
                messages.error(request, f'Error: {e}')
        else:
            messages.error(request, 'Please correct the errors below')
    else:
        form = UniformSizeForm()

    return render(request, 'uniforms/sizes/form.html', {
        'form': form, 'title': 'Create Uniform Size', 'submit_text': 'Create',
    })


@login_required
def uniform_size_edit(request, pk):
    size = get_object_or_404(UniformSize, pk=pk)

    if request.method == 'POST':
        form = UniformSizeForm(request.POST, instance=size)
        if form.is_valid():
            try:
                size = form.save()
                messages.success(request, f'Size "{size.name}" updated')
                return redirect('uniforms:uniform_size_detail', pk=size.pk)
            except Exception as e:
                messages.error(request, f'Error: {e}')
        else:
            messages.error(request, 'Please correct the errors below')
    else:
        form = UniformSizeForm(instance=size)

    return render(request, 'uniforms/sizes/form.html', {
        'form': form, 'size': size,
        'title': f'Edit {size.name}', 'submit_text': 'Update',
    })


@login_required
def uniform_size_delete(request, pk):
    size = get_object_or_404(UniformSize, pk=pk)
    if request.method == 'POST':
        if size.uniform_items.exists():
            resp = _htmx_response(
                request, 'Cannot delete — size is used by uniform items', 'error'
            )
            return resp or redirect('uniforms:uniform_size_detail', pk=pk)
        try:
            name = size.name
            size.delete()
            resp = _htmx_response(
                request, f'Size "{name}" deleted',
                redirect_url=reverse('uniforms:uniform_size_list'),
            )
            return resp or redirect('uniforms:uniform_size_list')
        except Exception as e:
            resp = _htmx_response(request, f'Error: {e}', 'error')
            return resp or redirect('uniforms:uniform_size_detail', pk=pk)


@login_required
def uniform_size_toggle_active(request, pk):
    size = get_object_or_404(UniformSize, pk=pk)
    if request.method == 'POST':
        try:
            size.is_active = not size.is_active
            size.save()
            verb = 'activated' if size.is_active else 'deactivated'
            resp = _htmx_response(
                request, f'Size "{size.name}" {verb}',
                redirect_url=reverse('uniforms:uniform_size_detail', kwargs={'pk': pk}),
            )
            return resp or redirect('uniforms:uniform_size_detail', pk=pk)
        except Exception as e:
            resp = _htmx_response(request, f'Error: {e}', 'error')
            return resp or redirect('uniforms:uniform_size_detail', pk=pk)


# =============================================================================
# UNIFORM ITEM VIEWS
# =============================================================================

@login_required
def uniform_item_list(request):
    items = get_filtered_uniform_items(request)

    stats = {
        'total':        items.count(),
        'active':       items.filter(is_active=True).count(),
        'low_stock':    items.filter(current_stock__lte=F('reorder_level')).count(),
        'out_of_stock': items.filter(current_stock=0).count(),
        'total_value':  items.aggregate(
                            total=Sum(F('current_stock') * F('unit_cost'))
                        )['total'] or Decimal('0.00'),
    }

    paginator = Paginator(items, 20)
    page      = paginator.get_page(request.GET.get('page', 1))
    is_htmx   = request.headers.get('HX-Request') == 'true'

    context = {
        'items_page': page, 'paginator': paginator,
        'stats': stats, 'filter_form': UniformItemFilterForm(request.GET or None),
        'is_htmx': is_htmx,
    }
    if is_htmx:
        return render(request, 'uniforms/items/partials/_item_results.html', context)
    return render(request, 'uniforms/items/list.html', context)


@login_required
def uniform_item_detail(request, pk):
    """
    Lightweight shell. Stock, sales, and size recommendations are loaded
    on demand through the three tab partial views below.
    """
    item = get_object_or_404(
        UniformItem.objects.select_related(
            'unit_of_measure', 'tax_rate',
            'inventory_account', 'cogs_account', 'revenue_account',
        ).prefetch_related('available_sizes'),
        pk=pk,
    )
    stats = {
        'current_stock':       item.current_stock,
        'stock_records_count': item.stock_records.count(),
        'sales_count':         item.sale_items.count(),
        'size_recs_count':     item.student_size_recommendations.filter(
                                   is_current=True
                               ).count(),
        'total_sold':          item.sale_items.aggregate(
                                   Sum('quantity')
                               )['quantity__sum'] or 0,
        'total_revenue':       item.sale_items.aggregate(
                                   Sum('total_price')
                               )['total_price__sum'] or Decimal('0.00'),
    }
    return render(request, 'uniforms/items/detail.html', {'item': item, 'stats': stats})


@login_required
def uniform_item_stock_partial(request, pk):
    """
    Stock Records tab.

    Exposes `missing_size_stock` — sizes in available_sizes with no stock
    record — so the template can render a "create missing records" banner.
    """
    item = get_object_or_404(UniformItem, pk=pk)
    qs   = item.stock_records.select_related('size').order_by('size__display_order')

    q       = request.GET.get('q', '').strip()
    has_qty = request.GET.get('has_quantity', '')

    if q:
        qs = qs.filter(
            Q(size__name__icontains=q) |
            Q(location__icontains=q) |
            Q(bin_number__icontains=q)
        )
    if has_qty == 'true':  qs = qs.filter(quantity__gt=0)
    if has_qty == 'false': qs = qs.filter(quantity=0)

    paginator = Paginator(qs, 20)
    page      = paginator.get_page(request.GET.get('page', 1))

    missing_size_stock = []
    if item.requires_sizing:
        sizes_with_records = set(
            item.stock_records.filter(size__isnull=False)
                .values_list('size_id', flat=True)
        )
        missing_size_stock = list(
            item.available_sizes.exclude(pk__in=sizes_with_records)
        )

    return render(request, 'uniforms/items/partials/_stock_tab.html', {
        'item':               item,
        'stock_page':         page,
        'paginator':          paginator,
        'missing_size_stock': missing_size_stock,
    })


@login_required
def uniform_item_sales_partial(request, pk):
    """
    Sale History tab.

    Totals reflect the full filtered queryset so they update when filters change.
    """
    item = get_object_or_404(UniformItem, pk=pk)
    qs   = item.sale_items.select_related(
        'sale__student',
        'sale__fiscal_period__related_academic_session',
        'size',
    ).order_by('-sale__sale_date')

    q      = request.GET.get('q', '').strip()
    size   = request.GET.get('size', '')
    status = request.GET.get('status', '')

    if q:
        qs = qs.filter(
            Q(sale__student__first_name__icontains=q) |
            Q(sale__student__last_name__icontains=q) |
            Q(sale__student__admission_number__icontains=q) |
            Q(sale__sale_number__icontains=q)
        )
    if size:   qs = qs.filter(size_id=size)
    if status: qs = qs.filter(sale__status=status)

    totals = qs.aggregate(
        qty_sold=Sum('quantity'),
        revenue=Sum('total_price'),
        cogs=Sum('total_cost'),
    )
    totals = {
        'qty_sold': totals['qty_sold'] or 0,
        'revenue':  totals['revenue']  or Decimal('0.00'),
        'cogs':     totals['cogs']     or Decimal('0.00'),
    }

    paginator = Paginator(qs, 20)
    page      = paginator.get_page(request.GET.get('page', 1))

    return render(request, 'uniforms/items/partials/_sales_tab.html', {
        'item':            item,
        'sale_items_page': page,
        'paginator':       paginator,
        'totals':          totals,
        'sizes':           item.available_sizes.filter(is_active=True).order_by(
                               'display_order'
                           ),
    })


@login_required
def uniform_item_size_recs_partial(request, pk):
    """Size Recommendations tab."""
    item = get_object_or_404(UniformItem, pk=pk)
    qs   = item.student_size_recommendations.select_related(
        'student', 'recommended_size', 'academic_session',
    ).filter(is_current=True).order_by('student__last_name', 'student__first_name')

    q       = request.GET.get('q', '').strip()
    session = request.GET.get('academic_session', '')

    if q:
        qs = qs.filter(
            Q(student__first_name__icontains=q) |
            Q(student__last_name__icontains=q) |
            Q(student__admission_number__icontains=q)
        )
    if session:
        qs = qs.filter(academic_session_id=session)

    paginator = Paginator(qs, 20)
    page      = paginator.get_page(request.GET.get('page', 1))

    return render(request, 'uniforms/items/partials/_size_recs_tab.html', {
        'item':      item,
        'recs_page': page,
        'paginator': paginator,
        'sessions':  AcademicSession.objects.filter(
                         is_active=True
                     ).order_by('-start_date'),
    })


@login_required
def uniform_item_create(request):
    if request.method == 'POST':
        form = UniformItemForm(request.POST, request.FILES)
        if form.is_valid():
            try:
                item = form.save()
                messages.success(request, f'"{item.name}" created')
                return redirect('uniforms:uniform_item_detail', pk=item.pk)
            except Exception as e:
                logger.error(f"Error creating uniform item: {e}")
                messages.error(request, f'Error: {e}')
        else:
            messages.error(request, 'Please correct the errors below')
    else:
        form = UniformItemForm()

    return render(request, 'uniforms/items/form.html', {
        'form': form, 'title': 'Create Uniform Item', 'submit_text': 'Create',
    })


@login_required
def uniform_item_edit(request, pk):
    item = get_object_or_404(UniformItem, pk=pk)

    if request.method == 'POST':
        form = UniformItemForm(request.POST, request.FILES, instance=item)
        if form.is_valid():
            try:
                item = form.save()
                messages.success(request, f'"{item.name}" updated')
                return redirect('uniforms:uniform_item_detail', pk=item.pk)
            except Exception as e:
                logger.error(f"Error updating uniform item: {e}")
                messages.error(request, f'Error: {e}')
        else:
            messages.error(request, 'Please correct the errors below')
    else:
        form = UniformItemForm(instance=item)

    return render(request, 'uniforms/items/form.html', {
        'form': form, 'item': item,
        'title': f'Edit {item.name}', 'submit_text': 'Update',
    })


@login_required
def uniform_item_delete(request, pk):
    item = get_object_or_404(UniformItem, pk=pk)
    if request.method == 'POST':
        if item.current_stock > 0:
            resp = _htmx_response(
                request, f'Cannot delete — {item.current_stock} units in stock', 'error'
            )
            return resp or redirect('uniforms:uniform_item_detail', pk=pk)
        if item.sale_items.exists():
            resp = _htmx_response(
                request, 'Cannot delete — item has existing sale records', 'error'
            )
            return resp or redirect('uniforms:uniform_item_detail', pk=pk)
        try:
            name = item.name
            item.delete()
            resp = _htmx_response(
                request, f'"{name}" deleted',
                redirect_url=reverse('uniforms:uniform_item_list'),
            )
            return resp or redirect('uniforms:uniform_item_list')
        except Exception as e:
            logger.error(f"Error deleting uniform item: {e}")
            resp = _htmx_response(request, f'Error: {e}', 'error')
            return resp or redirect('uniforms:uniform_item_detail', pk=pk)


@login_required
def uniform_item_toggle_active(request, pk):
    item = get_object_or_404(UniformItem, pk=pk)
    if request.method == 'POST':
        try:
            item.is_active = not item.is_active
            item.save()
            verb = 'activated' if item.is_active else 'deactivated'
            resp = _htmx_response(
                request, f'"{item.name}" {verb}',
                redirect_url=reverse('uniforms:uniform_item_detail', kwargs={'pk': pk}),
            )
            return resp or redirect('uniforms:uniform_item_detail', pk=pk)
        except Exception as e:
            resp = _htmx_response(request, f'Error: {e}', 'error')
            return resp or redirect('uniforms:uniform_item_detail', pk=pk)


@login_required
def uniform_item_adjust_stock(request, pk):
    """Adjust stock for an unsized item via ADD / REMOVE / SET."""
    item = get_object_or_404(UniformItem, pk=pk)
    if request.method == 'POST':
        try:
            adjustment_type = request.POST.get('adjustment_type')
            quantity        = int(request.POST.get('quantity', 0))
            reason          = request.POST.get('reason', '').strip()

            if not reason:
                raise ValidationError('Adjustment reason is required')
            if quantity < 0:
                raise ValidationError('Quantity must be positive')

            with transaction.atomic():
                stock, _ = UniformStock.objects.get_or_create(
                    uniform_item=item, size=None,
                    defaults={'quantity': item.current_stock},
                )
                if adjustment_type == 'ADD':
                    stock.quantity += quantity
                elif adjustment_type == 'REMOVE':
                    if quantity > stock.available_quantity:
                        raise ValidationError(
                            f'Cannot remove {quantity} — only '
                            f'{stock.available_quantity} available'
                        )
                    stock.quantity -= quantity
                elif adjustment_type == 'SET':
                    stock.quantity = quantity
                else:
                    raise ValidationError('Invalid adjustment type')
                stock.save()

            item.refresh_from_db()
            resp = _htmx_response(
                request, f'Stock adjusted. New stock: {item.current_stock}',
                redirect_url=reverse('uniforms:uniform_item_detail', kwargs={'pk': pk}),
            )
            return resp or redirect('uniforms:uniform_item_detail', pk=pk)
        except ValidationError as e:
            resp = _htmx_response(request, str(e), 'error')
            return resp or redirect('uniforms:uniform_item_detail', pk=pk)
        except Exception as e:
            logger.error(f"Error adjusting stock: {e}", exc_info=True)
            resp = _htmx_response(request, f'Error: {e}', 'error')
            return resp or redirect('uniforms:uniform_item_detail', pk=pk)


@login_required
def uniform_item_transfer_stock(request, pk):
    """Transfer stock between size variants of the same item."""
    item = get_object_or_404(UniformItem, pk=pk)
    if request.method == 'POST':
        try:
            source_size_pk = request.POST.get('source_size')
            target_size_pk = request.POST.get('target_size')
            quantity       = int(request.POST.get('quantity', 0))
            reason         = request.POST.get('reason', '').strip()

            if not source_size_pk or not target_size_pk:
                raise ValidationError('Source and target sizes are required')
            if source_size_pk == target_size_pk:
                raise ValidationError('Source and target sizes must be different')
            if quantity <= 0:
                raise ValidationError('Quantity must be greater than zero')
            if not reason:
                raise ValidationError('Transfer reason is required')

            with transaction.atomic():
                source = get_object_or_404(
                    UniformStock, uniform_item=item, size_id=source_size_pk
                )
                if source.available_quantity < quantity:
                    raise ValidationError(
                        f'Only {source.available_quantity} units available '
                        f'(requested {quantity})'
                    )
                target_size  = get_object_or_404(UniformSize, pk=target_size_pk)
                target, _    = UniformStock.objects.get_or_create(
                    uniform_item=item, size=target_size, defaults={'quantity': 0}
                )
                source.quantity -= quantity
                target.quantity += quantity
                source.save()
                target.save()
                logger.info(
                    f"Stock transfer: {item.name} "
                    f"{source.size.name} → {target_size.name} ×{quantity} — {reason}"
                )

            resp = _htmx_response(
                request,
                f'Transferred {quantity} unit(s) from '
                f'{source.size.name} to {target_size.name}',
                redirect_url=reverse('uniforms:uniform_item_detail', kwargs={'pk': pk}),
            )
            return resp or redirect('uniforms:uniform_item_detail', pk=pk)
        except ValidationError as e:
            resp = _htmx_response(request, str(e), 'error')
            return resp or redirect('uniforms:uniform_item_detail', pk=pk)
        except Exception as e:
            logger.error(f"Error transferring stock: {e}", exc_info=True)
            resp = _htmx_response(request, f'Error: {e}', 'error')
            return resp or redirect('uniforms:uniform_item_detail', pk=pk)


# =============================================================================
# UNIFORM STOCK VIEWS
# =============================================================================

@login_required
def uniform_stock_list(request):
    qs = UniformStock.objects.select_related(
        'uniform_item', 'size'
    ).order_by('uniform_item__name', 'size__display_order')

    filter_form  = UniformStockFilterForm(request.GET or None)

    q            = request.GET.get('q', '').strip()
    item_id      = request.GET.get('item', '')
    size_id      = request.GET.get('size', '')
    stock_status = request.GET.get('stock_status', '')

    if q:
        qs = qs.filter(
            Q(uniform_item__name__icontains=q) |
            Q(uniform_item__code__icontains=q) |
            Q(size__name__icontains=q)
        )
    if item_id:
        qs = qs.filter(uniform_item_id=item_id)
    if size_id:
        qs = qs.filter(size_id=size_id)
    if stock_status == 'out_of_stock':
        qs = qs.filter(quantity=0)
    elif stock_status == 'low_stock':
        qs = qs.filter(
            quantity__gt=0,
            quantity__lte=F('uniform_item__reorder_level')
        )
    elif stock_status == 'in_stock':
        qs = qs.filter(quantity__gt=F('uniform_item__reorder_level'))

    stats = {
        'total_records':  qs.count(),
        'total_quantity': qs.aggregate(Sum('quantity'))['quantity__sum'] or 0,
        'total_value':    qs.aggregate(
                              Sum('total_cost_value')
                          )['total_cost_value__sum'] or Decimal('0.00'),
        'low_stock':      qs.filter(
                              quantity__lte=F('uniform_item__reorder_level')
                          ).count(),
    }

    paginator = Paginator(qs, 20)
    page      = paginator.get_page(request.GET.get('page', 1))
    is_htmx   = request.headers.get('HX-Request') == 'true'

    context = {
        'stock_page':   page,
        'paginator':    paginator,
        'stats':        stats,
        'filter_form':  filter_form,
        'is_htmx':      is_htmx,
    }
    if is_htmx:
        return render(request, 'uniforms/stock/partials/_stock_results.html', context)
    return render(request, 'uniforms/stock/list.html', context)


@login_required
def uniform_stock_detail(request, pk):
    stock = get_object_or_404(
        UniformStock.objects.select_related('uniform_item', 'size'), pk=pk
    )
    return render(request, 'uniforms/stock/detail.html', {'stock': stock})


@login_required
def uniform_stock_create(request):
    if request.method == 'POST':
        form = UniformStockForm(request.POST)
        if form.is_valid():
            try:
                stock      = form.save()
                size_label = f' — Size {stock.size.name}' if stock.size else ''
                messages.success(
                    request, f'Stock record created for {stock.uniform_item.name}{size_label}'
                )
                return redirect('uniforms:uniform_stock_detail', pk=stock.pk)
            except Exception as e:
                logger.error(f"Error creating stock record: {e}", exc_info=True)
                messages.error(request, f'Error: {e}')
        else:
            messages.error(request, 'Please correct the errors below')
    else:
        item_id = request.GET.get('item')
        form    = UniformStockForm(initial={'uniform_item': item_id} if item_id else {})

    return render(request, 'uniforms/stock/form.html', {
        'form': form, 'title': 'Create Stock Record', 'submit_text': 'Create',
    })


@login_required
def uniform_stock_edit(request, pk):
    stock = get_object_or_404(UniformStock, pk=pk)

    if request.method == 'POST':
        form = UniformStockForm(request.POST, instance=stock)
        if form.is_valid():
            try:
                stock = form.save()
                messages.success(request, 'Stock record updated')
                return redirect('uniforms:uniform_stock_detail', pk=stock.pk)
            except Exception as e:
                logger.error(f"Error updating stock record: {e}", exc_info=True)
                messages.error(request, f'Error: {e}')
        else:
            messages.error(request, 'Please correct the errors below')
    else:
        form = UniformStockForm(instance=stock)

    return render(request, 'uniforms/stock/form.html', {
        'form': form, 'stock': stock,
        'title': 'Edit Stock Record', 'submit_text': 'Update',
    })


@login_required
def uniform_stock_delete(request, pk):
    stock = get_object_or_404(UniformStock, pk=pk)
    if request.method == 'POST':
        if stock.quantity > 0:
            resp = _htmx_response(
                request, 'Cannot delete — quantity must be zero first', 'error'
            )
            return resp or redirect('uniforms:uniform_stock_detail', pk=pk)
        try:
            label = (
                f'{stock.uniform_item.name} — Size {stock.size.name}'
                if stock.size else stock.uniform_item.name
            )
            stock.delete()
            resp = _htmx_response(
                request, f'Stock record for {label} deleted',
                redirect_url=reverse('uniforms:uniform_stock_list'),
            )
            return resp or redirect('uniforms:uniform_stock_list')
        except Exception as e:
            logger.error(f"Error deleting stock record: {e}", exc_info=True)
            resp = _htmx_response(request, f'Error: {e}', 'error')
            return resp or redirect('uniforms:uniform_stock_detail', pk=pk)


@login_required
def stock_receive(request, stock_pk):
    """Quick-receive against an existing stock record (no PO required)."""
    stock = get_object_or_404(
        UniformStock.objects.select_related('uniform_item', 'size'), pk=stock_pk
    )
    if request.method == 'POST':
        try:
            quantity = int(request.POST.get('quantity_received', 0))
            reason   = request.POST.get('reason', '').strip()
            if quantity <= 0:
                raise ValidationError('Quantity received must be greater than zero')
            if not reason:
                raise ValidationError('Receive reason is required')
            with transaction.atomic():
                stock.quantity += quantity
                stock.save()
            size_label = f' — {stock.size.name}' if stock.size else ''
            resp = _htmx_response(
                request,
                f'Received {quantity} unit(s) for '
                f'{stock.uniform_item.name}{size_label}. Total: {stock.quantity}',
                redirect_url=reverse(
                    'uniforms:uniform_stock_detail', kwargs={'pk': stock_pk}
                ),
            )
            return resp or redirect('uniforms:uniform_stock_detail', pk=stock_pk)
        except (ValidationError, Exception) as e:
            resp = _htmx_response(request, str(e), 'error')
            return resp or redirect('uniforms:uniform_stock_detail', pk=stock_pk)
    return redirect('uniforms:uniform_stock_detail', pk=stock_pk)


@login_required
def stock_transfer(request, stock_pk):
    """Transfer stock between size variants of the same item."""
    source = get_object_or_404(
        UniformStock.objects.select_related('uniform_item', 'size'), pk=stock_pk
    )
    if request.method == 'POST':
        try:
            target_size_pk = request.POST.get('target_size_pk', '').strip()
            quantity       = int(request.POST.get('quantity', 0))
            reason         = request.POST.get('reason', '').strip()

            if not target_size_pk:
                raise ValidationError('Target size is required')
            if not source.size:
                raise ValidationError(
                    'Cannot transfer an unsized stock record — use stock adjustment instead'
                )
            if not reason:
                raise ValidationError('Transfer reason is required')
            if quantity <= 0:
                raise ValidationError('Transfer quantity must be greater than zero')
            if source.available_quantity < quantity:
                raise ValidationError(
                    f'Only {source.available_quantity} unit(s) available'
                )

            target_size = get_object_or_404(UniformSize, pk=target_size_pk)
            if target_size == source.size:
                raise ValidationError('Source and target sizes must be different')
            if not source.uniform_item.available_sizes.filter(pk=target_size.pk).exists():
                raise ValidationError(
                    f'Size "{target_size.name}" is not associated with '
                    f'"{source.uniform_item.name}"'
                )

            with transaction.atomic():
                target, _ = UniformStock.objects.get_or_create(
                    uniform_item=source.uniform_item,
                    size=target_size,
                    defaults={'quantity': 0},
                )
                source.quantity -= quantity
                target.quantity += quantity
                source.save()
                target.save()

            resp = _htmx_response(
                request,
                f'Transferred {quantity} unit(s) from '
                f'{source.size.name} to {target_size.name}',
                redirect_url=reverse('uniforms:uniform_stock_list'),
            )
            return resp or redirect('uniforms:uniform_stock_list')
        except (ValidationError, Exception) as e:
            logger.error(f"Error in stock_transfer: {e}", exc_info=True)
            resp = _htmx_response(request, str(e), 'error')
            return resp or redirect('uniforms:uniform_stock_detail', pk=stock_pk)
    return redirect('uniforms:uniform_stock_detail', pk=stock_pk)


# =============================================================================
# PURCHASE ORDER VIEWS
# =============================================================================

@login_required
def purchase_order_list(request):
    orders = get_filtered_purchase_orders(request)

    stats = {
        'total':        orders.count(),
        'draft':        orders.filter(status='DRAFT').count(),
        'pending':      orders.filter(status__in=['SUBMITTED', 'APPROVED']).count(),
        'received':     orders.filter(status='RECEIVED').count(),
        'total_amount': orders.aggregate(
                            Sum('total_amount')
                        )['total_amount__sum'] or Decimal('0.00'),
    }

    paginator = Paginator(orders, 20)
    page      = paginator.get_page(request.GET.get('page', 1))
    is_htmx   = request.headers.get('HX-Request') == 'true'

    context = {
        'orders_page': page, 'paginator': paginator,
        'stats': stats,
        'filter_form': UniformPurchaseOrderFilterForm(request.GET or None),
        'is_htmx': is_htmx,
    }
    if is_htmx:
        return render(
            request, 'uniforms/purchase_orders/partials/_order_results.html', context
        )
    return render(request, 'uniforms/purchase_orders/list.html', context)


@login_required
def purchase_order_detail(request, pk):
    po = get_object_or_404(
        UniformPurchaseOrder.objects.select_related(
            'fiscal_period', 'journal_entry'
        ).prefetch_related('items__uniform_item', 'items__size'),
        pk=pk,
    )
    return render(request, 'uniforms/purchase_orders/detail.html', {
        'po':          po,
        'approved_by': po.get_approved_by_user() if po.approved_by_id else None,
    })


@login_required
def purchase_order_create(request):
    if request.method == 'POST':
        form = UniformPurchaseOrderForm(request.POST)
        if form.is_valid():
            try:
                with transaction.atomic():
                    po = form.save()
                messages.success(request, f'Purchase order {po.po_number} created')
                return redirect('uniforms:purchase_order_detail', pk=po.pk)
            except Exception as e:
                logger.error(f"Error creating PO: {e}")
                messages.error(request, f'Error: {e}')
        else:
            messages.error(request, 'Please correct the errors below')
    else:
        form = UniformPurchaseOrderForm()

    return render(request, 'uniforms/purchase_orders/form.html', {
        'form': form, 'title': 'Create Purchase Order', 'submit_text': 'Create',
    })


@login_required
def purchase_order_edit(request, pk):
    po = get_object_or_404(UniformPurchaseOrder, pk=pk)
    if po.status != 'DRAFT':
        messages.error(request, 'Only draft purchase orders can be edited')
        return redirect('uniforms:purchase_order_detail', pk=pk)

    if request.method == 'POST':
        form = UniformPurchaseOrderForm(request.POST, instance=po)
        if form.is_valid():
            try:
                po = form.save()
                messages.success(request, f'Purchase order {po.po_number} updated')
                return redirect('uniforms:purchase_order_detail', pk=po.pk)
            except Exception as e:
                messages.error(request, f'Error: {e}')
        else:
            messages.error(request, 'Please correct the errors below')
    else:
        form = UniformPurchaseOrderForm(instance=po)

    return render(request, 'uniforms/purchase_orders/form.html', {
        'form': form, 'po': po,
        'title': f'Edit {po.po_number}', 'submit_text': 'Update',
    })


@login_required
def purchase_order_delete(request, pk):
    po = get_object_or_404(UniformPurchaseOrder, pk=pk)
    if request.method == 'POST':
        if po.status != 'DRAFT':
            resp = _htmx_response(request, 'Only draft POs can be deleted', 'error')
            return resp or redirect('uniforms:purchase_order_detail', pk=pk)
        try:
            number = po.po_number
            po.delete()
            resp = _htmx_response(
                request, f'Purchase order {number} deleted',
                redirect_url=reverse('uniforms:purchase_order_list'),
            )
            return resp or redirect('uniforms:purchase_order_list')
        except Exception as e:
            resp = _htmx_response(request, f'Error: {e}', 'error')
            return resp or redirect('uniforms:purchase_order_detail', pk=pk)


@login_required
def purchase_order_submit(request, pk):
    po = get_object_or_404(UniformPurchaseOrder, pk=pk)
    if request.method == 'POST':
        if po.status != 'DRAFT':
            messages.error(request, 'Only draft POs can be submitted')
        elif not po.items.exists():
            messages.error(request, 'Cannot submit a PO without items')
        else:
            try:
                with transaction.atomic():
                    po.status = 'SUBMITTED'
                    po.save()
                messages.success(request, f'{po.po_number} submitted for approval')
            except Exception as e:
                messages.error(request, f'Error: {e}')
    return redirect('uniforms:purchase_order_detail', pk=pk)


@login_required
def purchase_order_approve(request, pk):
    po = get_object_or_404(UniformPurchaseOrder, pk=pk)
    if request.method == 'POST':
        if po.status != 'SUBMITTED':
            messages.error(request, 'Only submitted POs can be approved')
        else:
            try:
                with transaction.atomic():
                    po.status         = 'APPROVED'
                    po.approved_by_id = str(request.user.id)
                    po.approved_at    = get_school_current_time()
                    po.save()
                messages.success(request, f'{po.po_number} approved')
            except Exception as e:
                messages.error(request, f'Error: {e}')
    return redirect('uniforms:purchase_order_detail', pk=pk)


@login_required
def purchase_order_receive(request, pk):
    """
    Record received quantities. Stock updates are handled by the
    purchase_order_item_post_save signal — this view never touches stock directly.
    """
    po = get_object_or_404(UniformPurchaseOrder, pk=pk)
    if request.method == 'POST':
        if po.status not in ('APPROVED', 'ORDERED', 'PARTIAL'):
            messages.error(request, 'PO must be Approved, Ordered or Partial to receive')
            return redirect('uniforms:purchase_order_detail', pk=pk)
        try:
            with transaction.atomic():
                all_received = True
                for item in po.items.all():
                    received_qty          = int(request.POST.get(f'qty_{item.pk}', 0))
                    item.quantity_received = received_qty
                    item.save()
                    if received_qty < item.quantity_ordered:
                        all_received = False
                po.status = 'RECEIVED' if all_received else 'PARTIAL'
                if all_received:
                    po.actual_delivery_date = get_school_today()
                po.save()
            messages.success(request, f'{po.po_number} received successfully')
        except Exception as e:
            logger.error(f"Error receiving PO: {e}")
            messages.error(request, f'Error: {e}')
    return redirect('uniforms:purchase_order_detail', pk=pk)


@login_required
def purchase_order_cancel(request, pk):
    po = get_object_or_404(UniformPurchaseOrder, pk=pk)
    if request.method == 'POST':
        if po.status in ('RECEIVED', 'CANCELLED'):
            messages.error(request, f'Cannot cancel a {po.get_status_display()} PO')
        else:
            try:
                with transaction.atomic():
                    po.status = 'CANCELLED'
                    po.save()
                messages.success(request, f'{po.po_number} cancelled')
            except Exception as e:
                messages.error(request, f'Error: {e}')
    return redirect('uniforms:purchase_order_detail', pk=pk)


# =============================================================================
# UNIFORM SALE VIEWS
# =============================================================================

@login_required
def uniform_sale_list(request):
    sales = get_filtered_uniform_sales(request)

    stats = {
        'total':         sales.count(),
        'pending':       sales.filter(status='PENDING',  cancelled=False, returned=False).count(),
        'paid':          sales.filter(status='PAID',     cancelled=False, returned=False).count(),
        'issued':        sales.filter(status='ISSUED',   cancelled=False, returned=False).count(),
        'cancelled':     sales.filter(cancelled=True).count(),
        'returned':      sales.filter(returned=True).count(),
        'total_revenue': sales.filter(cancelled=False, returned=False).aggregate(
                             Sum('total_amount')
                         )['total_amount__sum'] or Decimal('0.00'),
    }

    # System-wide count of PAID/PARTIAL sales not yet issued — banner warning
    unissued_count = UniformSale.objects.filter(
        status__in=['PAID', 'PARTIAL'], cancelled=False, returned=False,
    ).count()

    paginator = Paginator(sales, 20)
    page      = paginator.get_page(request.GET.get('page', 1))
    is_htmx   = request.headers.get('HX-Request') == 'true'

    context = {
        'sales_page':     page,
        'paginator':      paginator,
        'stats':          stats,
        'filter_form':    UniformSaleFilterForm(request.GET or None),
        'is_htmx':        is_htmx,
        'unissued_count': unissued_count,
    }
    if is_htmx:
        return render(request, 'uniforms/sales/partials/_sale_results.html', context)
    return render(request, 'uniforms/sales/list.html', context)


@login_required
def uniform_sale_detail(request, pk):
    """
    Lightweight shell. Line items and audit trail are loaded on demand
    through the two tab partial views below.
    """
    sale = get_object_or_404(
        UniformSale.objects.select_related(
            'student',
            'fiscal_period__related_academic_session',
            'fee_invoice',
            'journal_entry',
            'cancellation_journal_entry',
            'return_journal_entry',
            'payment_method',
        ),
        pk=pk,
    )
    stats = {
        'item_count':    sale.items.count(),
        'total_qty':     sale.items.aggregate(Sum('quantity'))['quantity__sum'] or 0,
        'audit_entries': len(sale.get_audit_trail()),
    }
    return render(request, 'uniforms/sales/detail.html', {'sale': sale, 'stats': stats})


@login_required
def uniform_sale_items_partial(request, pk):
    """Line Items tab. No filtering needed — a sale's items are a fixed small set."""
    sale  = get_object_or_404(UniformSale, pk=pk)
    items = sale.items.select_related(
        'uniform_item', 'size', 'tax_rate'
    ).order_by('uniform_item__name')
    return render(request, 'uniforms/sales/partials/_items_tab.html', {
        'sale': sale, 'items': items,
    })


@login_required
def uniform_sale_audit_partial(request, pk):
    """Audit Trail tab."""
    sale = get_object_or_404(
        UniformSale.objects.select_related(
            'student',
            'fiscal_period__related_academic_session',
            'fee_invoice',
            'journal_entry',
        ),
        pk=pk,
    )
    return render(request, 'uniforms/sales/partials/_audit_tab.html', {
        'sale':        sale,
        'audit_trail': sale.get_audit_trail(),
    })


@login_required
def uniform_sale_create(request):
    student_id = request.GET.get('student')
    student    = get_object_or_404(Student, pk=student_id) if student_id else None

    if request.method == 'POST':
        form = UniformSaleForm(request.POST, student=student)
        if form.is_valid():
            try:
                with transaction.atomic():
                    sale = form.save(commit=False)
                    sale.save()
                    form.save_m2m()
                    items_data = _parse_sale_items(request.POST)
                    if items_data:
                        _save_sale_items(sale, items_data)
                    if request.POST.get('action') == 'finalize':
                        success, msg, sale = UniformSaleWorkflowService.finalise_sale(
                            sale, user=request.user
                        )
                        if not success:
                            raise ValueError(msg)
                        messages.success(request, f'Sale {sale.sale_number} created and finalised.')
                    else:
                        messages.success(request, f'Sale {sale.sale_number} created.')
                return redirect('uniforms:uniform_sale_detail', pk=sale.pk)
            except ValueError as e:
                messages.error(request, str(e))
            except Exception as e:
                logger.error(f"Error creating sale: {e}", exc_info=True)
                messages.error(request, f'Error: {e}')
        else:
            messages.error(request, 'Please correct the errors below')
    else:
        form = UniformSaleForm(student=student)

    return render(request, 'uniforms/sales/form.html', {
        'form':            form,
        'available_items': UniformItem.objects.filter(
                               is_active=True
                           ).select_related('tax_rate').order_by('name'),
        'title':           'Create Uniform Sale',
        'submit_text':     'Save Sale',
    })


@login_required
def uniform_sale_edit(request, pk):
    sale = get_object_or_404(
        UniformSale.objects.prefetch_related('items__uniform_item', 'items__size'),
        pk=pk,
    )
    if sale.status != 'DRAFT' or sale.cancelled or sale.returned:
        messages.error(request, 'Only active draft sales can be edited')
        return redirect('uniforms:uniform_sale_detail', pk=pk)

    if request.method == 'POST':
        form = UniformSaleForm(request.POST, instance=sale)
        if form.is_valid():
            try:
                with transaction.atomic():
                    form.save()
                    _save_sale_items(sale, _parse_sale_items(request.POST))
                    if request.POST.get('action') == 'finalize':
                        success, msg, sale = UniformSaleWorkflowService.finalise_sale(
                            sale, user=request.user
                        )
                        if not success:
                            raise ValueError(msg)
                        messages.success(request, f'Sale {sale.sale_number} finalised.')
                    else:
                        messages.success(request, f'Sale {sale.sale_number} updated.')
                return redirect('uniforms:uniform_sale_detail', pk=sale.pk)
            except ValueError as e:
                messages.error(request, str(e))
            except Exception as e:
                logger.exception("Error updating sale")
                messages.error(request, f'Error: {e}')
        else:
            messages.error(request, 'Please correct the errors below')
    else:
        form = UniformSaleForm(instance=sale)

    return render(request, 'uniforms/sales/form.html', {
        'form':            form,
        'sale':            sale,
        'available_items': UniformItem.objects.filter(
                               is_active=True
                           ).select_related('tax_rate').order_by('name'),
        'title':           f'Edit Sale — {sale.sale_number}',
        'submit_text':     'Save Sale',
    })


@login_required
def uniform_sale_finalize(request, pk):
    sale = get_object_or_404(UniformSale, pk=pk)
    if request.method == 'POST':
        success, msg, sale = UniformSaleWorkflowService.finalise_sale(
            sale, user=request.user
        )
        if success:
            messages.success(request, msg)
        else:
            messages.error(request, msg)
    return redirect('uniforms:uniform_sale_detail', pk=pk)


@login_required
def uniform_sale_issue(request, pk):
    sale = get_object_or_404(UniformSale, pk=pk)
    if request.method == 'POST':
        success, msg = UniformSaleWorkflowService.issue_sale(sale, request.user)
        if success:
            messages.success(request, msg)
        else:
            messages.error(request, msg)
    return redirect('uniforms:uniform_sale_detail', pk=pk)


@login_required
def uniform_sale_cancel(request, pk):
    sale = get_object_or_404(UniformSale, pk=pk)
    if request.method == 'POST':
        reason = request.POST.get('cancellation_reason', '').strip()
        if not reason:
            resp = _htmx_response(request, 'Cancellation reason is required', 'error')
            return resp or redirect('uniforms:uniform_sale_detail', pk=pk)
        success, msg = cancel_uniform_sale(sale, request.user, reason)
        resp = _htmx_response(
            request, msg,
            alert_type='success' if success else 'error',
            redirect_url=reverse(
                'uniforms:uniform_sale_detail', kwargs={'pk': pk}
            ) if success else None,
        )
        if resp:
            return resp
        (messages.success if success else messages.error)(request, msg)
    return redirect('uniforms:uniform_sale_detail', pk=pk)


@login_required
def uniform_sale_return(request, pk):
    sale = get_object_or_404(UniformSale, pk=pk)
    if request.method == 'POST':
        reason    = request.POST.get('return_reason', '').strip()
        condition = request.POST.get('return_condition', '').strip()
        if not reason or not condition:
            resp = _htmx_response(
                request, 'Return reason and condition are required', 'error'
            )
            return resp or redirect('uniforms:uniform_sale_detail', pk=pk)
        success, msg, _ = return_uniform_sale(sale, request.user, reason, condition)
        resp = _htmx_response(
            request, msg,
            alert_type='success' if success else 'error',
            redirect_url=reverse(
                'uniforms:uniform_sale_detail', kwargs={'pk': pk}
            ) if success else None,
        )
        if resp:
            return resp
        (messages.success if success else messages.error)(request, msg)
    return redirect('uniforms:uniform_sale_detail', pk=pk)


@login_required
def uniform_sale_delete(request, pk):
    sale = get_object_or_404(UniformSale, pk=pk)
    if request.method == 'POST':
        if sale.status != 'DRAFT':
            resp = _htmx_response(request, 'Only draft sales can be deleted', 'error')
            return resp or redirect('uniforms:uniform_sale_detail', pk=pk)
        try:
            number = sale.sale_number
            sale.delete()
            resp = _htmx_response(
                request, f'Sale {number} deleted',
                redirect_url=reverse('uniforms:uniform_sale_list'),
            )
            return resp or redirect('uniforms:uniform_sale_list')
        except Exception as e:
            logger.error(f"Error deleting sale: {e}")
            resp = _htmx_response(request, f'Error: {e}', 'error')
            return resp or redirect('uniforms:uniform_sale_detail', pk=pk)


# =============================================================================
# STUDENT UNIFORM SIZE VIEWS  (no list — accessed via tabs and student profile)
# =============================================================================

@login_required
def student_uniform_size_create(request):
    student_id = request.GET.get('student')
    student    = get_object_or_404(Student, pk=student_id) if student_id else None

    if request.method == 'POST':
        form = StudentUniformSizeForm(request.POST, student=student)
        if form.is_valid():
            try:
                rec = form.save()
                messages.success(
                    request,
                    f'Size recommendation created for {rec.student.get_full_name()}'
                )
                return redirect('uniforms:student_uniform_size_detail', pk=rec.pk)
            except Exception as e:
                logger.error(f"Error creating size recommendation: {e}")
                messages.error(request, f'Error: {e}')
        else:
            messages.error(request, 'Please correct the errors below')
    else:
        form = StudentUniformSizeForm(student=student)

    return render(request, 'uniforms/student_sizes/form.html', {
        'form': form, 'title': 'Create Size Recommendation', 'submit_text': 'Create',
    })


@login_required
def student_uniform_size_detail(request, pk):
    rec = get_object_or_404(
        StudentUniformSize.objects.select_related(
            'student', 'uniform_item', 'recommended_size', 'academic_session'
        ),
        pk=pk,
    )
    others = StudentUniformSize.objects.filter(
        student=rec.student
    ).exclude(pk=pk).select_related(
        'uniform_item', 'recommended_size'
    ).order_by('-recommendation_date')[:10]

    return render(request, 'uniforms/student_sizes/detail.html', {
        'size_rec':              rec,
        'other_recommendations': others,
    })


@login_required
def student_uniform_size_edit(request, pk):
    rec = get_object_or_404(StudentUniformSize, pk=pk)

    if request.method == 'POST':
        form = StudentUniformSizeForm(request.POST, instance=rec)
        if form.is_valid():
            try:
                rec = form.save()
                messages.success(request, 'Size recommendation updated')
                return redirect('uniforms:student_uniform_size_detail', pk=rec.pk)
            except Exception as e:
                messages.error(request, f'Error: {e}')
        else:
            messages.error(request, 'Please correct the errors below')
    else:
        form = StudentUniformSizeForm(instance=rec)

    return render(request, 'uniforms/student_sizes/form.html', {
        'form': form, 'size_rec': rec,
        'title': 'Edit Size Recommendation', 'submit_text': 'Update',
    })


@login_required
def student_uniform_size_delete(request, pk):
    rec = get_object_or_404(StudentUniformSize, pk=pk)
    if request.method == 'POST':
        try:
            name = rec.student.get_full_name()
            rec.delete()
            resp = _htmx_response(request, f'Size recommendation deleted for {name}')
            # No redirect_url — caller decides where to return (item tab or profile)
            return resp or redirect('uniforms:uniform_item_detail',
                                    pk=rec.uniform_item_id)
        except Exception as e:
            resp = _htmx_response(request, f'Error: {e}', 'error')
            return resp or redirect('uniforms:student_uniform_size_detail', pk=pk)


# =============================================================================
# REPORT VIEWS
# =============================================================================

@login_required
def inventory_report(request):
    items = UniformItem.objects.filter(is_active=True).select_related(
        'unit_of_measure'
    ).order_by('item_type', 'name')
    return render(request, 'uniforms/reports/inventory.html', {
        'items':     items,
        'stats':     uniform_stats.get_inventory_stats(),
        'valuation': uniform_stats.get_stock_valuation(),
        'title':     'Inventory Report',
    })


@login_required
def sales_report(request):
    date_from           = request.GET.get('date_from')
    date_to             = request.GET.get('date_to')
    academic_session_id = request.GET.get('academic_session')

    sales = UniformSale.objects.filter(
        cancelled=False, returned=False
    ).select_related('student', 'fiscal_period__related_academic_session')

    if date_from:           sales = sales.filter(sale_date__gte=date_from)
    if date_to:             sales = sales.filter(sale_date__lte=date_to)
    if academic_session_id: sales = sales.filter(
        fiscal_period__related_academic_session_id=academic_session_id
    )

    return render(request, 'uniforms/reports/sales.html', {
        'sales':        sales[:100],
        'stats':        uniform_stats.get_sales_stats(
                            date_from=date_from, date_to=date_to,
                            academic_session_id=academic_session_id or None,
                        ),
        'by_item_type': uniform_stats.get_sales_by_item_type(
                            date_from=date_from, date_to=date_to,
                        ),
        'cogs_margin':  uniform_stats.get_cogs_and_margin_stats(
                            date_from=date_from, date_to=date_to,
                            academic_session_id=academic_session_id or None,
                        ),
        'date_from': date_from, 'date_to': date_to, 'title': 'Sales Report',
    })


@login_required
def low_stock_report(request):
    items = UniformItem.objects.filter(
        is_active=True, current_stock__lte=F('reorder_level')
    ).select_related('unit_of_measure').order_by('current_stock')
    return render(request, 'uniforms/reports/low_stock.html', {
        'items': items, 'title': 'Low Stock Report',
    })


@login_required
def measurement_summary_report(request):
    academic_session_id = request.GET.get('academic_session')
    class_id            = request.GET.get('class')

    by_type = uniform_stats.get_measurement_coverage_by_type(
        academic_session_id=academic_session_id or None
    )
    summary = uniform_stats.get_measurement_stats(
        academic_session_id=academic_session_id or None
    )
    class_coverage = []
    if academic_session_id:
        class_coverage = uniform_stats.get_measurement_coverage_by_class(
            academic_session_id=academic_session_id
        )
        if class_id:
            class_coverage = [c for c in class_coverage if str(c['class_id']) == class_id]

    return render(request, 'uniforms/reports/measurement_summary.html', {
        'stats':           by_type,
        'summary':         summary,
        'class_coverage':  class_coverage,
        'total_students':  summary.get('students_measured', 0),
        'filters_applied': bool(request.GET),
    })


@login_required
def student_orders_report(request):
    from collections import defaultdict

    academic_session_id = request.GET.get('academic_session')
    class_id            = request.GET.get('class')

    sales = UniformSale.objects.select_related(
        'student', 'fiscal_period__related_academic_session',
    ).filter(cancelled=False, returned=False)

    if academic_session_id:
        sales = sales.filter(
            fiscal_period__related_academic_session_id=academic_session_id
        )
    if class_id:
        sales = sales.filter(
            student__class_enrollments__class_instance_id=class_id,
            student__class_enrollments__is_active=True,
            student__class_enrollments__completion_status='ONGOING',
        ).distinct()

    by_student = defaultdict(list)
    for sale in sales:
        by_student[sale.student].append(sale)

    student_totals = sorted(
        [
            {
                'student':      stu,
                'total_sales':  len(stu_sales),
                'total_amount': sum(s.total_amount for s in stu_sales),
                'paid_amount':  sum(s.paid_amount  for s in stu_sales),
                'balance':      sum(s.balance       for s in stu_sales),
            }
            for stu, stu_sales in by_student.items()
        ],
        key=lambda x: x['total_amount'],
        reverse=True,
    )

    return render(request, 'uniforms/reports/student_orders.html', {
        'student_totals':  student_totals[:100],
        'filters_applied': bool(request.GET),
    })


# =============================================================================
# PRINT VIEWS — SINGLE ITEM  (no field selection needed)
# =============================================================================

@login_required
def measurement_type_print_detail(request, pk):
    mt = get_object_or_404(MeasurementType, pk=pk)
    return render(request, 'uniforms/measurement_types/print_detail.html', {
        'measurement_type':  mt,
        'measurement_count': mt.student_measurements.count(),
        'verified_count':    mt.student_measurements.filter(is_verified=True).count(),
        'avg_value':         mt.student_measurements.filter(
                                 is_current=True
                             ).aggregate(Avg('value'))['value__avg'],
        'print_date':        get_school_current_time(),
    })


@login_required
def student_measurement_print_detail(request, pk):
    m = get_object_or_404(
        StudentMeasurement.objects.select_related(
            'student', 'measurement_type', 'academic_session'
        ),
        pk=pk,
    )
    return render(request, 'uniforms/measurements/print_detail.html', {
        'measurement':        m,
        'other_measurements': StudentMeasurement.objects.filter(
                                  student=m.student,
                                  measurement_type=m.measurement_type,
                              ).exclude(pk=pk).order_by('-measurement_date')[:10],
        'verified_by':        m.get_verified_by_user(),
        'print_date':         get_school_current_time(),
    })


@login_required
def uniform_item_print_detail(request, pk):
    item = get_object_or_404(UniformItem, pk=pk)
    return render(request, 'uniforms/items/print_detail.html', {
        'item':         item,
        'stock_records':item.stock_records.select_related('size').order_by(
                            'size__display_order'
                        ),
        'recent_sales': item.sale_items.select_related(
                            'sale__student', 'size'
                        ).order_by('-sale__sale_date')[:20],
        'print_date':   get_school_current_time(),
    })


@login_required
def uniform_stock_print_detail(request, pk):
    stock = get_object_or_404(
        UniformStock.objects.select_related('uniform_item', 'size'), pk=pk
    )
    return render(request, 'uniforms/stock/print_detail.html', {
        'stock': stock, 'print_date': get_school_current_time(),
    })


@login_required
def purchase_order_print_detail(request, pk):
    po = get_object_or_404(
        UniformPurchaseOrder.objects.select_related('fiscal_period'), pk=pk
    )
    return render(request, 'uniforms/purchase_orders/print_detail.html', {
        'po':          po,
        'items':       po.items.select_related('uniform_item', 'size'),
        'approved_by': po.get_approved_by_user(),
        'print_date':  get_school_current_time(),
    })


@login_required
def uniform_sale_print_detail(request, pk):
    sale = get_object_or_404(
        UniformSale.objects.select_related(
            'student', 'fiscal_period__related_academic_session'
        ),
        pk=pk,
    )
    return render(request, 'uniforms/sales/print_detail.html', {
        'sale':       sale,
        'items':      sale.items.select_related('uniform_item', 'size'),
        'print_date': get_school_current_time(),
    })


@login_required
def uniform_sale_print_invoice(request, pk):
    sale = get_object_or_404(
        UniformSale.objects.select_related(
            'student', 'fiscal_period__related_academic_session'
        ),
        pk=pk,
    )
    return render(request, 'uniforms/sales/print_invoice.html', {
        'sale':       sale,
        'items':      sale.items.select_related('uniform_item', 'size'),
        'print_date': get_school_current_time(),
    })


# =============================================================================
# PRINT VIEWS — LISTS  (field selection via modal checkboxes)
# =============================================================================

@login_required
def student_measurements_print_view(request):
    """
    Printable measurement list. Accepts a `fields` GET list so the caller
    controls which columns appear. Respects active filters.
    """
    selected_fields = request.GET.getlist('fields') or [
        'student_name', 'admission_number', 'measurement_type',
        'value', 'unit', 'measurement_date', 'is_verified',
    ]
    include_stats  = request.GET.get('include_stats') == 'true'
    landscape_mode = request.GET.get('landscape') == 'true'

    measurements = get_filtered_student_measurements(request)

    stats = None
    if include_stats:
        stats = {
            'total':    measurements.count(),
            'verified': measurements.filter(is_verified=True).count(),
            'current':  measurements.filter(is_current=True).count(),
        }

    field_names = {
        'student_name':       'Student',
        'admission_number':   'Adm. No.',
        'measurement_type':   'Type',
        'value':              'Value',
        'unit':               'Unit',
        'measurement_date':   'Date',
        'measurement_context':'Context',
        'measurement_method': 'Method',
        'is_verified':        'Verified',
        'is_current':         'Current',
        'academic_session':   'Session',
        'notes':              'Notes',
    }

    return render(request, 'uniforms/measurements/print_list.html', {
        'measurements':         measurements,
        'stats':                stats,
        'selected_fields':      selected_fields,
        'selected_field_names': [
            field_names.get(f, f.replace('_', ' ').title()) for f in selected_fields
        ],
        'field_names':          field_names,
        'landscape':            landscape_mode,
        'print_date':           get_school_current_time(),
        'filters_applied':      bool(request.GET),
    })


@login_required
def uniform_items_print_view(request):
    selected_fields = request.GET.getlist('fields') or [
        'code', 'name', 'item_type', 'gender',
        'selling_price', 'current_stock', 'is_active',
    ]
    include_stats  = request.GET.get('include_stats') == 'true'
    landscape_mode = request.GET.get('landscape') == 'true'

    items = get_filtered_uniform_items(request)

    stats = None
    if include_stats:
        stats = {
            'total':        items.count(),
            'active':       items.filter(is_active=True).count(),
            'low_stock':    items.filter(current_stock__lte=F('reorder_level')).count(),
            'total_value':  items.aggregate(
                                total=Sum(F('current_stock') * F('unit_cost'))
                            )['total'] or Decimal('0.00'),
        }

    field_names = {
        'code':              'Code',
        'name':              'Item Name',
        'item_type':         'Type',
        'gender':            'Gender',
        'sku':               'SKU',
        'unit_cost':         'Unit Cost',
        'selling_price':     'Selling Price',
        'markup_percentage': 'Markup %',
        'current_stock':     'Stock',
        'reorder_level':     'Reorder Level',
        'is_active':         'Status',
        'is_mandatory':      'Mandatory',
        'color':             'Color',
        'description':       'Description',
    }

    return render(request, 'uniforms/items/print_list.html', {
        'items':                items,
        'stats':                stats,
        'selected_fields':      selected_fields,
        'selected_field_names': [
            field_names.get(f, f.replace('_', ' ').title()) for f in selected_fields
        ],
        'field_names':          field_names,
        'landscape':            landscape_mode,
        'print_date':           get_school_current_time(),
        'filters_applied':      bool(request.GET),
    })


@login_required
def uniform_stock_print_view(request):
    records = UniformStock.objects.select_related(
        'uniform_item', 'size'
    ).order_by('uniform_item__name', 'size__display_order')

    selected_fields = request.GET.getlist('fields') or [
        'item_code', 'item_name', 'size', 'quantity', 'available_quantity',
        'total_cost_value',
    ]
    landscape_mode = request.GET.get('landscape') == 'true'

    field_names = {
        'item_code':          'Code',
        'item_name':          'Item',
        'size':               'Size',
        'quantity':           'Qty',
        'reserved_quantity':  'Reserved',
        'available_quantity': 'Available',
        'location':           'Location',
        'bin_number':         'Bin',
        'total_cost_value':   'Cost Value',
        'total_selling_value':'Sell Value',
    }

    return render(request, 'uniforms/stock/print_list.html', {
        'stock_records':        records[:200],
        'selected_fields':      selected_fields,
        'selected_field_names': [
            field_names.get(f, f.replace('_', ' ').title()) for f in selected_fields
        ],
        'field_names':          field_names,
        'landscape':            landscape_mode,
        'print_date':           get_school_current_time(),
    })


@login_required
def purchase_orders_print_view(request):
    orders = get_filtered_purchase_orders(request)

    selected_fields = request.GET.getlist('fields') or [
        'po_number', 'supplier_name', 'order_date',
        'status', 'total_amount', 'balance_due',
    ]
    include_stats  = request.GET.get('include_stats') == 'true'
    landscape_mode = request.GET.get('landscape') == 'true'

    stats = None
    if include_stats:
        stats = {
            'total':        orders.count(),
            'total_amount': orders.aggregate(
                                Sum('total_amount')
                            )['total_amount__sum'] or Decimal('0.00'),
        }

    field_names = {
        'po_number':             'PO Number',
        'supplier_name':         'Supplier',
        'order_date':            'Order Date',
        'expected_delivery_date':'Expected Delivery',
        'actual_delivery_date':  'Actual Delivery',
        'status':                'Status',
        'subtotal':              'Subtotal',
        'tax_amount':            'Tax',
        'shipping_cost':         'Shipping',
        'total_amount':          'Total',
        'paid_amount':           'Paid',
        'balance_due':           'Balance',
    }

    return render(request, 'uniforms/purchase_orders/print_list.html', {
        'orders':               orders[:100],
        'stats':                stats,
        'selected_fields':      selected_fields,
        'selected_field_names': [
            field_names.get(f, f.replace('_', ' ').title()) for f in selected_fields
        ],
        'field_names':          field_names,
        'landscape':            landscape_mode,
        'print_date':           get_school_current_time(),
        'filters_applied':      bool(request.GET),
    })


@login_required
def uniform_sales_print_view(request):
    sales = get_filtered_uniform_sales(request)

    selected_fields = request.GET.getlist('fields') or [
        'sale_number', 'student_name', 'sale_date',
        'total_amount', 'paid_amount', 'balance', 'status',
    ]
    include_stats  = request.GET.get('include_stats') == 'true'
    landscape_mode = request.GET.get('landscape') == 'true'

    stats = None
    if include_stats:
        active = sales.filter(cancelled=False, returned=False)
        stats = {
            'total':         sales.count(),
            'total_revenue': active.aggregate(
                                 Sum('total_amount')
                             )['total_amount__sum'] or Decimal('0.00'),
        }

    field_names = {
        'sale_number':    'Sale No.',
        'student_name':   'Student',
        'admission_number':'Adm. No.',
        'sale_date':      'Date',
        'sale_type':      'Type',
        'subtotal':       'Subtotal',
        'discount_amount':'Discount',
        'tax_amount':     'Tax',
        'total_amount':   'Total',
        'paid_amount':    'Paid',
        'balance':        'Balance',
        'status':         'Status',
    }

    return render(request, 'uniforms/sales/print_list.html', {
        'sales':                sales[:100],
        'stats':                stats,
        'selected_fields':      selected_fields,
        'selected_field_names': [
            field_names.get(f, f.replace('_', ' ').title()) for f in selected_fields
        ],
        'field_names':          field_names,
        'landscape':            landscape_mode,
        'print_date':           get_school_current_time(),
        'filters_applied':      bool(request.GET),
    })


# =============================================================================
# EXCEL EXPORT VIEWS  (column selection via modal checkboxes)
# =============================================================================

@login_required
def export_measurements_excel(request):
    measurements = get_filtered_student_measurements(request)

    ALL_COLUMNS = [
        ('student_name',      'Student',        lambda m: m.student.get_full_name()),
        ('admission_number',  'Adm. No.',       lambda m: m.student.admission_number),
        ('measurement_type',  'Type',           lambda m: m.measurement_type.name),
        ('value',             'Value',          lambda m: float(m.value)),
        ('unit',              'Unit',           lambda m: m.measurement_type.unit.abbreviation),
        ('measurement_date',  'Date',           lambda m: m.measurement_date.strftime('%Y-%m-%d')),
        ('measurement_context','Context',       lambda m: m.get_measurement_context_display()),
        ('measurement_method','Method',         lambda m: m.get_measurement_method_display()),
        ('is_verified',       'Verified',       lambda m: 'Yes' if m.is_verified else 'No'),
        ('is_current',        'Current',        lambda m: 'Yes' if m.is_current else 'No'),
        ('academic_session',  'Session',        lambda m: str(m.academic_session) if m.academic_session else ''),
        ('notes',             'Notes',          lambda m: m.notes or ''),
    ]

    COLUMN_MAP   = {col[0]: col for col in ALL_COLUMNS}
    DEFAULT      = ['student_name', 'admission_number', 'measurement_type',
                    'value', 'unit', 'measurement_date', 'is_verified']
    selected     = request.GET.getlist('fields') or DEFAULT
    columns      = [COLUMN_MAP[f] for f in selected if f in COLUMN_MAP] or \
                   [COLUMN_MAP[f] for f in DEFAULT]

    wb = _build_workbook(
        'Measurements',
        [(col[1], col[2]) for col in columns],
        measurements,
    )
    return _excel_response(wb, 'measurements')


@login_required
def export_uniform_items_excel(request):
    items = get_filtered_uniform_items(request)

    ALL_COLUMNS = [
        ('code',              'Code',           lambda i: i.code),
        ('name',              'Item Name',      lambda i: i.name),
        ('item_type',         'Type',           lambda i: i.get_item_type_display()),
        ('gender',            'Gender',         lambda i: i.get_gender_display()),
        ('sku',               'SKU',            lambda i: i.sku or ''),
        ('unit_cost',         'Unit Cost',      lambda i: float(i.unit_cost)),
        ('selling_price',     'Selling Price',  lambda i: float(i.selling_price)),
        ('markup_percentage', 'Markup %',       lambda i: float(i.markup_percentage)),
        ('current_stock',     'Stock',          lambda i: i.current_stock),
        ('reorder_level',     'Reorder Level',  lambda i: i.reorder_level),
        ('is_active',         'Status',         lambda i: 'Active' if i.is_active else 'Inactive'),
        ('is_mandatory',      'Mandatory',      lambda i: 'Yes' if i.is_mandatory else 'No'),
        ('color',             'Color',          lambda i: i.color or ''),
        ('description',       'Description',    lambda i: i.description or ''),
    ]

    COLUMN_MAP = {col[0]: col for col in ALL_COLUMNS}
    DEFAULT    = ['code', 'name', 'item_type', 'gender',
                  'selling_price', 'current_stock', 'reorder_level', 'is_active']
    selected   = request.GET.getlist('fields') or DEFAULT
    columns    = [COLUMN_MAP[f] for f in selected if f in COLUMN_MAP] or \
                 [COLUMN_MAP[f] for f in DEFAULT]

    wb = _build_workbook(
        'Uniform Items',
        [(col[1], col[2]) for col in columns],
        items,
    )
    return _excel_response(wb, 'uniform_items')


@login_required
def export_uniform_sizes_excel(request):
    sizes = get_filtered_uniform_sizes(request)

    ALL_COLUMNS = [
        ('code',       'Code',       lambda s: s.code),
        ('name',       'Name',       lambda s: s.name),
        ('size_type',  'Type',       lambda s: s.get_size_type_display()),
        ('min_height', 'Min Height', lambda s: float(s.min_height) if s.min_height else ''),
        ('max_height', 'Max Height', lambda s: float(s.max_height) if s.max_height else ''),
        ('min_chest',  'Min Chest',  lambda s: float(s.min_chest)  if s.min_chest  else ''),
        ('max_chest',  'Max Chest',  lambda s: float(s.max_chest)  if s.max_chest  else ''),
        ('min_waist',  'Min Waist',  lambda s: float(s.min_waist)  if s.min_waist  else ''),
        ('max_waist',  'Max Waist',  lambda s: float(s.max_waist)  if s.max_waist  else ''),
        ('is_active',  'Active',     lambda s: 'Yes' if s.is_active else 'No'),
    ]

    COLUMN_MAP = {col[0]: col for col in ALL_COLUMNS}
    DEFAULT    = ['code', 'name', 'size_type',
                  'min_height', 'max_height', 'min_chest', 'max_chest', 'is_active']
    selected   = request.GET.getlist('fields') or DEFAULT
    columns    = [COLUMN_MAP[f] for f in selected if f in COLUMN_MAP] or \
                 [COLUMN_MAP[f] for f in DEFAULT]

    wb = _build_workbook(
        'Uniform Sizes',
        [(col[1], col[2]) for col in columns],
        sizes,
    )
    return _excel_response(wb, 'uniform_sizes')


@login_required
def export_uniform_stock_excel(request):
    records = UniformStock.objects.select_related(
        'uniform_item', 'size'
    ).order_by('uniform_item__name', 'size__display_order')

    ALL_COLUMNS = [
        ('item_code',          'Item Code',    lambda r: r.uniform_item.code),
        ('item_name',          'Item Name',    lambda r: r.uniform_item.name),
        ('size',               'Size',         lambda r: r.size.name if r.size else 'N/A'),
        ('quantity',           'Quantity',     lambda r: r.quantity),
        ('reserved_quantity',  'Reserved',     lambda r: r.reserved_quantity),
        ('available_quantity', 'Available',    lambda r: r.available_quantity),
        ('location',           'Location',     lambda r: r.location or ''),
        ('bin_number',         'Bin',          lambda r: r.bin_number or ''),
        ('total_cost_value',   'Cost Value',   lambda r: float(r.total_cost_value)),
        ('total_selling_value','Selling Value',lambda r: float(r.total_selling_value)),
    ]

    COLUMN_MAP = {col[0]: col for col in ALL_COLUMNS}
    DEFAULT    = ['item_code', 'item_name', 'size', 'quantity',
                  'reserved_quantity', 'available_quantity', 'total_cost_value']
    selected   = request.GET.getlist('fields') or DEFAULT
    columns    = [COLUMN_MAP[f] for f in selected if f in COLUMN_MAP] or \
                 [COLUMN_MAP[f] for f in DEFAULT]

    wb = _build_workbook(
        'Uniform Stock',
        [(col[1], col[2]) for col in columns],
        records,
    )
    return _excel_response(wb, 'uniform_stock')


@login_required
def export_purchase_orders_excel(request):
    orders = get_filtered_purchase_orders(request)

    ALL_COLUMNS = [
        ('po_number',             'PO Number',        lambda o: o.po_number),
        ('supplier_name',         'Supplier',         lambda o: o.supplier_name),
        ('order_date',            'Order Date',       lambda o: o.order_date.strftime('%Y-%m-%d')),
        ('expected_delivery_date','Expected Delivery',
            lambda o: o.expected_delivery_date.strftime('%Y-%m-%d') if o.expected_delivery_date else ''),
        ('actual_delivery_date',  'Actual Delivery',
            lambda o: o.actual_delivery_date.strftime('%Y-%m-%d') if o.actual_delivery_date else ''),
        ('status',                'Status',           lambda o: o.get_status_display()),
        ('subtotal',              'Subtotal',         lambda o: float(o.subtotal)),
        ('tax_amount',            'Tax',              lambda o: float(o.tax_amount)),
        ('shipping_cost',         'Shipping',         lambda o: float(o.shipping_cost)),
        ('total_amount',          'Total',            lambda o: float(o.total_amount)),
        ('paid_amount',           'Paid',             lambda o: float(o.paid_amount)),
        ('balance_due',           'Balance',          lambda o: float(o.balance_due)),
        ('payment_terms',         'Payment Terms',    lambda o: o.payment_terms or ''),
    ]

    COLUMN_MAP = {col[0]: col for col in ALL_COLUMNS}
    DEFAULT    = ['po_number', 'supplier_name', 'order_date',
                  'status', 'total_amount', 'paid_amount', 'balance_due']
    selected   = request.GET.getlist('fields') or DEFAULT
    columns    = [COLUMN_MAP[f] for f in selected if f in COLUMN_MAP] or \
                 [COLUMN_MAP[f] for f in DEFAULT]

    wb = _build_workbook(
        'Purchase Orders',
        [(col[1], col[2]) for col in columns],
        orders,
    )
    return _excel_response(wb, 'purchase_orders')


@login_required
def export_uniform_sales_excel(request):
    sales = get_filtered_uniform_sales(request)

    ALL_COLUMNS = [
        ('sale_number',    'Sale No.',    lambda s: s.sale_number),
        ('student_name',   'Student',     lambda s: s.student.get_full_name()),
        ('admission_number','Adm. No.',   lambda s: s.student.admission_number),
        ('sale_date',      'Date',        lambda s: s.sale_date.strftime('%Y-%m-%d')),
        ('sale_type',      'Type',        lambda s: s.get_sale_type_display()),
        ('subtotal',       'Subtotal',    lambda s: float(s.subtotal)),
        ('discount_amount','Discount',    lambda s: float(s.discount_amount)),
        ('tax_amount',     'Tax',         lambda s: float(s.tax_amount)),
        ('total_amount',   'Total',       lambda s: float(s.total_amount)),
        ('paid_amount',    'Paid',        lambda s: float(s.paid_amount)),
        ('balance',        'Balance',     lambda s: float(s.balance)),
        ('status',         'Status',      lambda s: s.get_status_display()),
        ('fiscal_period',  'Fiscal Period',lambda s: str(s.fiscal_period) if s.fiscal_period else ''),
    ]

    COLUMN_MAP = {col[0]: col for col in ALL_COLUMNS}
    DEFAULT    = ['sale_number', 'student_name', 'sale_date',
                  'total_amount', 'paid_amount', 'balance', 'status']
    selected   = request.GET.getlist('fields') or DEFAULT
    columns    = [COLUMN_MAP[f] for f in selected if f in COLUMN_MAP] or \
                 [COLUMN_MAP[f] for f in DEFAULT]

    wb = _build_workbook(
        'Uniform Sales',
        [(col[1], col[2]) for col in columns],
        sales,
    )
    return _excel_response(wb, 'uniform_sales')


# =============================================================================
# SALE FORM HELPERS
# =============================================================================

def _parse_sale_items(post_data):
    """Parse items[N][field] keys from POST into a list of dicts."""
    import re
    items   = {}
    pattern = re.compile(r'^items\[(\d+)\]\[(\w+)\]$')
    for key, value in post_data.items():
        m = pattern.match(key)
        if m:
            idx, field = int(m.group(1)), m.group(2)
            items.setdefault(idx, {})[field] = value
    return [items[k] for k in sorted(items.keys())]


def _save_sale_items(sale, items_data):
    """Replace all line items on a sale from parsed POST data."""
    from core.models import TaxRate

    sale.items.all().delete()

    for data in items_data:
        item_pk = data.get('uniform_item')
        if not item_pk:
            continue

        uniform_item = UniformItem.objects.get(pk=item_pk, is_active=True)
        size         = None
        size_pk      = data.get('size')
        if size_pk:
            size = UniformSize.objects.get(pk=size_pk)

        quantity   = max(1, int(data.get('quantity', 1)))
        unit_price = Decimal(str(data.get('unit_price') or uniform_item.selling_price))
        unit_cost  = Decimal(str(data.get('unit_cost')  or uniform_item.unit_cost))
        tax_pct    = Decimal(str(data.get('tax_percentage', '0.00')))

        tax_rate    = None
        tax_rate_pk = data.get('tax_rate')
        if tax_rate_pk:
            try:
                tax_rate = TaxRate.objects.get(pk=tax_rate_pk)
            except TaxRate.DoesNotExist:
                pass

        UniformSaleItem.objects.create(
            sale=sale,
            uniform_item=uniform_item,
            size=size,
            quantity=quantity,
            unit_price=unit_price,
            unit_cost=unit_cost,
            tax_rate=tax_rate,
            tax_percentage=tax_pct,
        )

    sale.calculate_totals()


# =============================================================================
# AJAX ENDPOINTS
# =============================================================================

@login_required
def ajax_get_item_sizes(request, item_pk):
    try:
        item  = UniformItem.objects.get(pk=item_pk)
        sizes = item.available_sizes.filter(is_active=True).order_by('display_order')
        return JsonResponse({'success': True, 'sizes': [
            {'id': s.id, 'name': s.name, 'code': s.code} for s in sizes
        ]})
    except UniformItem.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Item not found'}, status=404)
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@login_required
def ajax_get_item_price(request, item_pk):
    try:
        item = UniformItem.objects.get(pk=item_pk)
        return JsonResponse({'success': True, 'data': {
            'unit_cost':     float(item.unit_cost),
            'selling_price': float(item.selling_price),
            'is_taxable':    item.is_taxable,
            'tax_rate_id':   item.tax_rate_id if item.tax_rate else None,
        }})
    except UniformItem.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Item not found'}, status=404)
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@login_required
def ajax_get_stock_quantity(request, item_pk, size_pk):
    try:
        stock = UniformStock.objects.get(uniform_item_id=item_pk, size_id=size_pk)
        return JsonResponse({'success': True, 'data': {
            'quantity':  stock.quantity,
            'reserved':  stock.reserved_quantity,
            'available': stock.available_quantity,
        }})
    except UniformStock.DoesNotExist:
        return JsonResponse({'success': True, 'data': {
            'quantity': 0, 'reserved': 0, 'available': 0,
        }})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@login_required
def ajax_get_item_stock(request, item_pk):
    """Stock data for an unsized item (size=None record)."""
    try:
        item = get_object_or_404(UniformItem, pk=item_pk)
        try:
            stock = UniformStock.objects.get(uniform_item=item, size__isnull=True)
            data  = {
                'quantity':  stock.quantity,
                'reserved':  stock.reserved_quantity,
                'available': stock.available_quantity,
                'stock_pk':  str(stock.pk),
            }
        except UniformStock.DoesNotExist:
            data = {'quantity': 0, 'reserved': 0, 'available': 0, 'stock_pk': None}
        return JsonResponse({'success': True, 'data': data})
    except Exception as e:
        logger.error(f"Error in ajax_get_item_stock: {e}", exc_info=True)
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@login_required
def ajax_get_student_measurements(request, student_pk):
    try:
        measurements = StudentMeasurement.objects.filter(
            student_id=student_pk, is_current=True
        ).select_related('measurement_type__unit')
        return JsonResponse({'success': True, 'measurements': [
            {
                'type_id':   m.measurement_type_id,
                'type_name': m.measurement_type.name,
                'type_code': m.measurement_type.code,
                'value':     float(m.value),
                'unit':      m.measurement_type.unit.abbreviation,
            }
            for m in measurements
        ]})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@login_required
def ajax_get_size_recommendation(request, student_pk, item_pk):
    try:
        rec = StudentUniformSize.objects.filter(
            student_id=student_pk, uniform_item_id=item_pk, is_current=True
        ).select_related('recommended_size').first()
        if rec:
            return JsonResponse({'success': True, 'data': {
                'size_id':    rec.recommended_size_id,
                'size_name':  rec.recommended_size.name,
                'confidence': rec.confidence_level,
            }})
        return JsonResponse({'success': False, 'error': 'No recommendation found'})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@login_required
def ajax_check_po_number(request):
    po_number  = request.GET.get('po_number', '').strip()
    exclude_id = request.GET.get('exclude_id')
    if not po_number:
        return JsonResponse({'exists': False})
    qs = UniformPurchaseOrder.objects.filter(po_number=po_number)
    if exclude_id:
        qs = qs.exclude(pk=exclude_id)
    return JsonResponse({'exists': qs.exists()})


@login_required
def ajax_calculate_sale_total(request):
    """Placeholder — total is calculated server-side in calculate_totals()."""
    return JsonResponse({'success': True, 'total': 0})