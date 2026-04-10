"""
hr/views.py

Human Resources Management Views

Organisation
────────────
Each section covers one entity end-to-end in this order:
  get_filtered_<entity>  →  list  →  detail/profile  →  create  →  edit
  →  delete / toggles / actions  →  print  →  export (where applicable)

All HTMX modal GET triggers live in modal_views.py.
Business-logic helpers that are not views use a leading underscore.
POST-only action views return HTTP 405 on GET and include a comment
pointing to the modal view that handles the GET.
"""

from django.shortcuts import render, redirect, get_object_or_404
from django.urls import reverse
from django.http import HttpResponse, HttpResponseNotAllowed, JsonResponse
from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator
from django.contrib import messages
from django.db.models import (
    Q, Count, Sum, Avg, Prefetch, F, IntegerField, Case, When,
)
from django.utils import timezone
from django.http import HttpResponse
from django.db import transaction
from django.core.files.storage import FileSystemStorage
from formtools.wizard.views import SessionWizardView

from datetime import timedelta
from decimal import Decimal
import os
import logging

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from core.utils import get_school_today

from core.view_helpers import (
    get_print_school_context,
)

from .models import (
    Department, Designation, Contract,
    Staff, StaffDesignation, Teacher,
    Attendance, Payroll, PayrollAllowance, PayrollDeduction, PayrollBonus,
    SalaryHistory, ContractBenefit,
)
from .forms import (
    STAFF_WIZARD_FORMS, STAFF_WIZARD_STEP_NAMES,
    StaffForm, DepartmentForm, DesignationForm, ContractForm,
    TeacherForm, AttendanceForm, PayrollForm,
    PayrollAllowanceFormSet, PayrollBonusFormSet, PayrollDeductionFormSet,
    StaffDesignationForm,
    DepartmentFilterForm, DesignationFilterForm, StaffFilterForm,
    ContractFilterForm, TeacherFilterForm, AttendanceFilterForm,
    PayrollFilterForm, PayrollReversalForm,
)
from . import stats as hr_stats

logger = logging.getLogger(__name__)


# =============================================================================
# DASHBOARD
# =============================================================================

@login_required
def hr_dashboard(request):
    """Main HR dashboard with overview statistics."""
    try:
        staff_stats       = hr_stats.get_staff_statistics()
        department_stats  = hr_stats.get_department_statistics()
        designation_stats = hr_stats.get_designation_statistics()
        contract_stats    = hr_stats.get_contract_statistics()
        teacher_stats     = hr_stats.get_teacher_statistics()
    except Exception as e:
        logger.error(f"Error getting dashboard statistics: {e}")
        staff_stats = department_stats = designation_stats = contract_stats = teacher_stats = {}

    today = get_school_today()

    expiring_contracts = Contract.objects.filter(
        status='ACTIVE',
        end_date__gte=today,
        end_date__lte=today + timedelta(days=30),
    ).select_related('staff').order_by('end_date')[:10]

    staff_without_contracts = Staff.objects.filter(is_active=True).annotate(
        contract_count=Count('contracts', filter=Q(contracts__status='ACTIVE'))
    ).filter(contract_count=0).order_by('date_of_joining')[:10]

    upcoming_birthdays = Staff.objects.filter(
        is_active=True,
        date_of_birth__month=today.month,
        date_of_birth__day__gte=today.day,
        date_of_birth__day__lte=(today + timedelta(days=7)).day,
    ).order_by('date_of_birth__day')[:10]

    probation_ending = []
    for staff in Staff.objects.filter(employment_status='PR', is_active=True):
        contract = Contract.objects.filter(
            staff=staff, status='ACTIVE', probation_period_months__gt=0
        ).first()
        if contract:
            probation_end = contract.start_date + timedelta(
                days=contract.probation_period_months * 30
            )
            if today <= probation_end <= today + timedelta(days=30):
                probation_ending.append({
                    'staff': staff,
                    'contract': contract,
                    'probation_end': probation_end,
                    'days_remaining': (probation_end - today).days,
                })

    return render(request, 'hr/dashboard.html', {
        'staff_statistics':        staff_stats,
        'department_statistics':   department_stats,
        'designation_statistics':  designation_stats,
        'contract_statistics':     contract_stats,
        'teacher_statistics':      teacher_stats,
        'recent_staff':            Staff.objects.select_related(
                                       'primary_department'
                                   ).order_by('-created_at')[:10],
        'recent_contracts':        Contract.objects.select_related(
                                       'staff'
                                   ).order_by('-created_at')[:10],
        'expiring_contracts':      expiring_contracts,
        'staff_without_contracts': staff_without_contracts,
        'upcoming_birthdays':      upcoming_birthdays,
        'probation_ending':        probation_ending,
        'recent_salary_changes':   SalaryHistory.objects.select_related(
                                       'staff', 'contract', 'effective_period'
                                   ).order_by('-effective_date')[:10],
    })


# =============================================================================
# DEPARTMENTS
# =============================================================================

def get_filtered_departments(request):
    departments = Department.objects.select_related('parent_department').annotate(
        staff_count=Count(
            'primary_staff', filter=Q(primary_staff__is_active=True), distinct=True
        ),
        designation_count=Count(
            'designations', filter=Q(designations__is_active=True), distinct=True
        ),
        sub_department_count=Count('sub_departments', distinct=True),
    ).order_by('department_type', 'name')

    query = request.GET.get('q', '').strip()
    if query:
        combined_q = Q()
        for word in query.split():
            combined_q &= (
                Q(name__icontains=word) |
                Q(code__icontains=word) |
                Q(description__icontains=word)
            )
        departments = departments.filter(combined_q)

    department_type = request.GET.get('department_type', '')
    if department_type:
        departments = departments.filter(department_type=department_type)

    academic_subtype = request.GET.get('academic_subtype', '')
    if academic_subtype:
        departments = departments.filter(academic_subtype=academic_subtype)

    is_academic = request.GET.get('is_academic', '')
    if is_academic:
        departments = departments.filter(is_academic=(is_academic.lower() == 'true'))

    is_active = request.GET.get('is_active', '')
    if is_active:
        departments = departments.filter(is_active=(is_active.lower() == 'true'))

    parent_department = request.GET.get('parent_department', '')
    if parent_department == 'null':
        departments = departments.filter(parent_department__isnull=True)
    elif parent_department == 'has_parent':
        departments = departments.filter(parent_department__isnull=False)
    elif parent_department:
        departments = departments.filter(parent_department_id=parent_department)

    return departments


@login_required
def department_list(request):
    filter_form = DepartmentFilterForm(request.GET or None)
    departments = get_filtered_departments(request)

    stats = {
        'total':              departments.count(),
        'active':             departments.filter(is_active=True).count(),
        'academic':           departments.filter(is_academic=True).count(),
        'administrative':     departments.filter(department_type='ADMINISTRATIVE').count(),
        'support':            departments.filter(department_type='SUPPORT').count(),
        'parent_departments': departments.filter(parent_department__isnull=True).count(),
        'sub_departments':    departments.filter(parent_department__isnull=False).count(),
        'total_staff':        sum(d.staff_count for d in departments),
        'total_budget':       departments.aggregate(
                                  Sum('annual_budget')
                              )['annual_budget__sum'] or 0,
    }

    paginator        = Paginator(departments, 10)
    departments_page = paginator.get_page(request.GET.get('page', 1))
    is_htmx          = request.headers.get('HX-Request') == 'true'

    context = {
        'departments_page': departments_page,
        'paginator':        paginator,
        'stats':            stats,
        'filter_form':      filter_form,
        'is_htmx':          is_htmx,
    }
    if is_htmx:
        return render(request, 'hr/departments/partials/_department_results.html', context)
    return render(request, 'hr/departments/list.html', context)


@login_required
def department_create(request):
    if request.method == 'POST':
        form = DepartmentForm(request.POST)
        if form.is_valid():
            department = form.save()
            if request.headers.get('HX-Request') == 'true':
                r = HttpResponse()
                r['HX-Alert-Message'] = f'Department "{department.name}" created successfully'
                r['HX-Alert-Type']    = 'success'
                r['HX-Close-Modal']   = 'true'
                r['HX-Redirect']      = reverse('hr:department_list')
                return r
            messages.success(request, f'Department "{department.name}" created successfully', extra_tags='sweetalert')
            return redirect('hr:department_list')
        else:
            if request.headers.get('HX-Request') == 'true':
                r = HttpResponse()
                r['HX-Alert-Message'] = 'Please correct the errors in the form'
                r['HX-Alert-Type']    = 'error'
                return r
            messages.error(request, 'Please correct the errors in the form', extra_tags='sweetalert-error')
    else:
        form = DepartmentForm()
    return render(request, 'hr/departments/form.html', {'form': form, 'title': 'Create Department'})


@login_required
def department_edit(request, pk):
    department = get_object_or_404(Department, pk=pk)
    if request.method == 'POST':
        form = DepartmentForm(request.POST, instance=department)
        if form.is_valid():
            department = form.save()
            if request.headers.get('HX-Request') == 'true':
                r = HttpResponse()
                r['HX-Alert-Message'] = f'Department "{department.name}" updated successfully'
                r['HX-Alert-Type']    = 'success'
                r['HX-Close-Modal']   = 'true'
                r['HX-Redirect']      = reverse('hr:department_list')
                return r
            messages.success(request, f'Department "{department.name}" updated successfully', extra_tags='sweetalert')
            return redirect('hr:department_list')
        else:
            if request.headers.get('HX-Request') == 'true':
                r = HttpResponse()
                r['HX-Alert-Message'] = 'Please correct the errors in the form'
                r['HX-Alert-Type']    = 'error'
                return r
            messages.error(request, 'Please correct the errors in the form', extra_tags='sweetalert-error')
    else:
        form = DepartmentForm(instance=department)
    return render(request, 'hr/departments/form.html', {
        'form': form, 'department': department, 'title': 'Update Department',
    })


@login_required
def department_delete(request, pk):
    """POST-only. modal_views.department_delete_modal handles GET."""
    if request.method != 'POST':
        return HttpResponse(status=405)

    department = get_object_or_404(Department, pk=pk)

    if department.primary_staff.exists():
        if request.headers.get('HX-Request') == 'true':
            r = HttpResponse()
            r['HX-Alert-Message'] = f'Cannot delete "{department.name}" — it has staff members'
            r['HX-Alert-Type']    = 'error'
            r['HX-Close-Modal']   = 'true'
            return r
        messages.error(request, f'Cannot delete "{department.name}" — it has staff members', extra_tags='sweetalert-error')
        return redirect('hr:department_list')

    if department.sub_departments.exists():
        if request.headers.get('HX-Request') == 'true':
            r = HttpResponse()
            r['HX-Alert-Message'] = f'Cannot delete "{department.name}" — it has sub-departments'
            r['HX-Alert-Type']    = 'error'
            r['HX-Close-Modal']   = 'true'
            return r
        messages.error(request, f'Cannot delete "{department.name}" — it has sub-departments', extra_tags='sweetalert-error')
        return redirect('hr:department_list')

    name = department.name
    department.delete()

    if request.headers.get('HX-Request') == 'true':
        r = HttpResponse()
        r['HX-Alert-Message'] = f'Department "{name}" deleted successfully'
        r['HX-Alert-Type']    = 'success'
        r['HX-Close-Modal']   = 'true'
        r['HX-Redirect']      = reverse('hr:department_list')
        return r
    messages.success(request, f'Department "{name}" deleted successfully', extra_tags='sweetalert')
    return redirect('hr:department_list')


# =============================================================================
# DESIGNATIONS
# =============================================================================

def get_filtered_designations(request):
    designations = Designation.objects.select_related(
        'department', 'reports_to'
    ).annotate(
        staff_count=Count(
            'staffdesignation', filter=Q(staffdesignation__is_active=True), distinct=True
        ),
        subordinate_count=Count('subordinate_designations', distinct=True),
    ).order_by('rank_order', 'name')

    query = request.GET.get('q', '').strip()
    if query:
        combined_q = Q()
        for word in query.split():
            combined_q &= (
                Q(name__icontains=word) |
                Q(code__icontains=word) |
                Q(description__icontains=word)
            )
        designations = designations.filter(combined_q)

    department = request.GET.get('department', '')
    if department:
        designations = designations.filter(department_id=department)

    is_teaching = request.GET.get('is_teaching', '')
    if is_teaching:
        designations = designations.filter(is_teaching=(is_teaching.lower() == 'true'))

    is_management = request.GET.get('is_management', '')
    if is_management:
        designations = designations.filter(is_management=(is_management.lower() == 'true'))

    is_active = request.GET.get('is_active', '')
    if is_active:
        designations = designations.filter(is_active=(is_active.lower() == 'true'))

    min_salary = request.GET.get('min_salary', '')
    if min_salary:
        try:
            designations = designations.filter(min_salary__gte=Decimal(min_salary))
        except (ValueError, TypeError):
            pass

    max_salary = request.GET.get('max_salary', '')
    if max_salary:
        try:
            designations = designations.filter(max_salary__lte=Decimal(max_salary))
        except (ValueError, TypeError):
            pass

    return designations


@login_required
def designation_list(request):
    filter_form  = DesignationFilterForm(request.GET or None)
    designations = get_filtered_designations(request)

    stats = {
        'total':           designations.count(),
        'active':          designations.filter(is_active=True).count(),
        'teaching':        designations.filter(is_teaching=True).count(),
        'management':      designations.filter(is_management=True).count(),
        'with_reports_to': designations.filter(reports_to__isnull=False).count(),
        'total_staff':     sum(d.staff_count for d in designations),
        'avg_min_salary':  designations.filter(
                               min_salary__isnull=False
                           ).aggregate(Avg('min_salary'))['min_salary__avg'] or 0,
        'avg_max_salary':  designations.filter(
                               max_salary__isnull=False
                           ).aggregate(Avg('max_salary'))['max_salary__avg'] or 0,
    }

    paginator         = Paginator(designations, 10)
    designations_page = paginator.get_page(request.GET.get('page', 1))
    is_htmx           = request.headers.get('HX-Request') == 'true'

    context = {
        'designations_page': designations_page,
        'paginator':         paginator,
        'stats':             stats,
        'filter_form':       filter_form,
        'is_htmx':           is_htmx,
    }
    if is_htmx:
        return render(request, 'hr/designations/partials/_designation_results.html', context)
    return render(request, 'hr/designations/list.html', context)


@login_required
def designation_detail(request, pk):
    designation = get_object_or_404(
        Designation.objects.select_related('department', 'reports_to'), pk=pk
    )
    staff_assignments = StaffDesignation.objects.filter(
        designation=designation, is_active=True
    ).select_related('staff__primary_department').order_by('-is_primary', 'staff__first_name')

    return render(request, 'hr/designations/detail.html', {
        'designation':       designation,
        'staff_assignments': staff_assignments,
        'subordinates':      designation.subordinate_designations.filter(is_active=True),
        'staff_count':       staff_assignments.count(),
    })


@login_required
def designation_create(request):
    if request.method == 'POST':
        form = DesignationForm(request.POST)
        if form.is_valid():
            designation = form.save()
            if request.headers.get('HX-Request') == 'true':
                r = HttpResponse()
                r['HX-Alert-Message'] = f'Designation "{designation.name}" created successfully'
                r['HX-Alert-Type']    = 'success'
                r['HX-Close-Modal']   = 'true'
                r['HX-Redirect']      = reverse('hr:designation_list')
                return r
            messages.success(request, f'Designation "{designation.name}" created successfully', extra_tags='sweetalert')
            return redirect('hr:designation_list')
        else:
            if request.headers.get('HX-Request') == 'true':
                r = HttpResponse()
                r['HX-Alert-Message'] = 'Please correct the errors in the form'
                r['HX-Alert-Type']    = 'error'
                return r
            messages.error(request, 'Please correct the errors in the form', extra_tags='sweetalert-error')
    else:
        form = DesignationForm()
    return render(request, 'hr/designations/form.html', {'form': form, 'title': 'Create Designation'})


@login_required
def designation_edit(request, pk):
    designation = get_object_or_404(Designation, pk=pk)
    if request.method == 'POST':
        form = DesignationForm(request.POST, instance=designation)
        if form.is_valid():
            designation = form.save()
            if request.headers.get('HX-Request') == 'true':
                r = HttpResponse()
                r['HX-Alert-Message'] = f'Designation "{designation.name}" updated successfully'
                r['HX-Alert-Type']    = 'success'
                r['HX-Close-Modal']   = 'true'
                r['HX-Redirect']      = reverse('hr:designation_list')
                return r
            messages.success(request, f'Designation "{designation.name}" updated successfully', extra_tags='sweetalert')
            return redirect('hr:designation_list')
        else:
            if request.headers.get('HX-Request') == 'true':
                r = HttpResponse()
                r['HX-Alert-Message'] = 'Please correct the errors in the form'
                r['HX-Alert-Type']    = 'error'
                return r
            messages.error(request, 'Please correct the errors in the form', extra_tags='sweetalert-error')
    else:
        form = DesignationForm(instance=designation)
    return render(request, 'hr/designations/form.html', {
        'form': form, 'designation': designation, 'title': 'Update Designation',
    })


@login_required
def designation_delete(request, pk):
    """POST-only. modal_views.designation_delete_modal handles GET."""
    if request.method != 'POST':
        return HttpResponse(status=405)

    designation = get_object_or_404(Designation, pk=pk)

    if designation.staff_members.exists():
        if request.headers.get('HX-Request') == 'true':
            r = HttpResponse()
            r['HX-Alert-Message'] = f'Cannot delete "{designation.name}" — it has staff assignments'
            r['HX-Alert-Type']    = 'error'
            r['HX-Close-Modal']   = 'true'
            return r
        messages.error(request, f'Cannot delete "{designation.name}" — it has staff assignments', extra_tags='sweetalert-error')
        return redirect('hr:designation_list')

    name = designation.name
    designation.delete()

    if request.headers.get('HX-Request') == 'true':
        r = HttpResponse()
        r['HX-Alert-Message'] = f'Designation "{name}" deleted successfully'
        r['HX-Alert-Type']    = 'success'
        r['HX-Close-Modal']   = 'true'
        r['HX-Redirect']      = reverse('hr:designation_list')
        return r
    messages.success(request, f'Designation "{name}" deleted successfully', extra_tags='sweetalert')
    return redirect('hr:designation_list')


# =============================================================================
# STAFF
# =============================================================================

def get_filtered_staff(request):
    staff = Staff.objects.select_related('primary_department').prefetch_related(
        Prefetch(
            'staffdesignation_set',
            queryset=StaffDesignation.objects.filter(
                is_primary=True, is_active=True
            ).select_related('designation'),
            to_attr='primary_staff_designation',
        ),
        'contracts',
    ).annotate(
        active_contract_count=Count(
            'contracts', filter=Q(contracts__status='ACTIVE'), distinct=True
        ),
        designation_count=Count(
            'staffdesignation', filter=Q(staffdesignation__is_active=True), distinct=True
        ),
    ).order_by('-is_active', 'first_name', 'last_name')

    query = request.GET.get('q', '').strip()
    if query:
        combined_q = Q()
        for word in query.split():
            combined_q &= (
                Q(first_name__icontains=word)     |
                Q(middle_name__icontains=word)    |
                Q(last_name__icontains=word)      |
                Q(staff_id__icontains=word)       |
                Q(phone_number__icontains=word)   |
                Q(personal_email__icontains=word) |
                Q(national_id__icontains=word)
            )
        staff = staff.filter(combined_q)

    employment_status = request.GET.get('employment_status', '')
    if employment_status:
        staff = staff.filter(employment_status=employment_status)

    gender = request.GET.get('gender', '')
    if gender:
        staff = staff.filter(gender=gender)

    primary_department = request.GET.get('primary_department', '')
    if primary_department:
        staff = staff.filter(primary_department_id=primary_department)

    marital_status = request.GET.get('marital_status', '')
    if marital_status:
        staff = staff.filter(marital_status=marital_status)

    nationality = request.GET.get('nationality', '')
    if nationality:
        staff = staff.filter(nationality=nationality)

    is_active = request.GET.get('is_active', '')
    if is_active:
        staff = staff.filter(is_active=(is_active.lower() == 'true'))

    return staff


@login_required
def staff_list(request):
    filter_form = StaffFilterForm(request.GET or None)
    staff       = get_filtered_staff(request)

    stats = {
        'total_staff':          staff.count(),
        'active_staff':         staff.filter(is_active=True).count(),
        'full_time_staff':      staff.filter(employment_status='FT').count(),
        'part_time':            staff.filter(employment_status='PT').count(),
        'contract':             staff.filter(employment_status='CT').count(),
        'male':                 staff.filter(gender='M').count(),
        'female':               staff.filter(gender='F').count(),
        'with_active_contract': staff.filter(active_contract_count__gt=0).count(),
        'teachers':             staff.filter(teacher__isnull=False).count(),
    }

    paginator  = Paginator(staff, 10)
    staff_page = paginator.get_page(request.GET.get('page', 1))
    is_htmx    = request.headers.get('HX-Request') == 'true'

    context = {
        'staff_page':  staff_page,
        'paginator':   paginator,
        'stats':       stats,
        'filter_form': filter_form,
        'is_htmx':     is_htmx,
    }
    if is_htmx:
        return render(request, 'hr/staff/partials/_staff_results.html', context)
    return render(request, 'hr/staff/list.html', context)


@login_required
def staff_profile(request, pk):
    staff = get_object_or_404(
        Staff.objects.prefetch_related(
            Prefetch(
                'staffdesignation_set',
                queryset=StaffDesignation.objects.filter(
                    is_active=True
                ).select_related('designation'),
            ),
            Prefetch(
                'contracts',
                queryset=Contract.objects.order_by('-start_date'),
            ),
        ),
        pk=pk,
    )

    from .utils import (
        get_staff_age, get_years_of_service,
        get_days_until_birthday, is_birthday_today, is_staff_due_for_retirement,
    )
    try:
        summary = {
            'staff_id':            staff.staff_id,
            'full_name':           staff.full_name(),
            'employment_status':   staff.get_employment_status_display(),
            'age':                 get_staff_age(staff),
            'years_of_service':    get_years_of_service(staff),
            'days_until_birthday': get_days_until_birthday(staff),
            'is_birthday_today':   is_birthday_today(staff),
            'retirement_info':     is_staff_due_for_retirement(staff),
            'designation_count':   staff.designations.count(),
            'contract_count':      staff.contracts.count(),
            'payroll_count':       Payroll.objects.filter(
                                       staff=staff, reversed=False
                                   ).count(),
            'attendance_count':    Attendance.objects.filter(staff=staff).count(),
        }
    except Exception as e:
        logger.error(f"Error getting staff summary: {e}")
        summary = {}

    designations        = staff.staffdesignation_set.filter(is_active=True)
    primary_designation = designations.filter(is_primary=True).first()
    contracts           = staff.contracts.order_by('-start_date')
    active_contract     = contracts.filter(status='ACTIVE').first()
    teacher_profile     = getattr(staff, 'teacher', None)
    salary_history      = staff.salary_history.select_related(
        'contract', 'effective_period'
    ).order_by('-effective_date')[:5]

    return render(request, 'hr/staff/profile.html', {
        'staff':               staff,
        'summary':             summary,
        'designations':        designations,
        'primary_designation': primary_designation,
        'contracts':           contracts,
        'active_contract':     active_contract,
        'teacher_profile':     teacher_profile,
        'salary_history':      salary_history,
    })

@login_required
def staff_payrolls_partial(request, pk):
    staff = get_object_or_404(Staff, pk=pk)

    qs = Payroll.objects.filter(staff=staff).select_related(
        'fiscal_period', 'payment_method'
    ).order_by('-payment_date')

    q             = request.GET.get('q', '').strip()
    status        = request.GET.get('status', '')
    only_reversed = request.GET.get('only_reversed', '')
    date_from     = request.GET.get('date_from', '')
    date_to       = request.GET.get('date_to', '')

    if q:
        qs = qs.filter(
            Q(pay_period_label__icontains=q) |
            Q(payment_reference__icontains=q)
        )
    if status:
        qs = qs.filter(status=status)
    if only_reversed == 'true':
        qs = qs.filter(reversed=True)
    elif only_reversed == 'false':
        qs = qs.filter(reversed=False)
    if date_from:
        qs = qs.filter(payment_date__gte=date_from)
    if date_to:
        qs = qs.filter(payment_date__lte=date_to)

    # Stats from filtered, unsliced queryset
    from django.db.models import Sum, Count, Avg
    payroll_stats = qs.aggregate(
        total_gross     = Sum('gross_pay'),
        total_net       = Sum('net_pay'),
        total_deductions= Sum('total_deductions'),
        total_allowances= Sum('total_allowances'),
        total_paye      = Sum('paye_amount'),
        total_nssf      = Sum('nssf_employee'),
        count           = Count('id'),
        avg_net         = Avg('net_pay'),
    )

    paginator = Paginator(qs, 10)
    page      = paginator.get_page(request.GET.get('page', 1))

    return render(request, 'hr/staff/partials/_payroll_results.html', {
        'payrolls':      page,
        'payroll_stats': payroll_stats,
        'staff':         staff,
    })


@login_required
def staff_attendance_partial(request, pk):
    staff = get_object_or_404(Staff, pk=pk)

    qs = Attendance.objects.filter(staff=staff).order_by('-date')

    q         = request.GET.get('q', '').strip()
    status    = request.GET.get('status', '')
    work_mode = request.GET.get('work_mode', '')
    date_from = request.GET.get('date_from', '')
    date_to   = request.GET.get('date_to', '')

    if q:
        qs = qs.filter(date__icontains=q)
    if status:
        qs = qs.filter(status=status)
    if work_mode:
        qs = qs.filter(work_mode=work_mode)
    if date_from:
        qs = qs.filter(date__gte=date_from)
    if date_to:
        qs = qs.filter(date__lte=date_to)

    # Summary counts from unsliced queryset
    attendance_summary = {
        'present':  qs.filter(status='PRESENT').count(),
        'absent':   qs.filter(status='ABSENT').count(),
        'late':     qs.filter(status='LATE').count(),
        'on_leave': qs.filter(status='LEAVE').count(),
    }

    paginator = Paginator(qs, 10)
    page      = paginator.get_page(request.GET.get('page', 1))

    return render(request, 'hr/staff/partials/_attendance_results.html', {
        'attendance':          page,
        'attendance_summary':  attendance_summary,
        'staff':               staff,
    })

@login_required
def staff_print_view(request):
    selected_fields = request.GET.getlist('fields') or [
        'staff_id', 'full_name', 'gender',
        'primary_department', 'employment_status', 'phone_number',
    ]
    include_stats  = request.GET.get('include_stats') == 'true'
    landscape_mode = request.GET.get('landscape') == 'true'
    short_headers  = request.GET.get('short_headers') == 'true'
    gender_display = request.GET.get('gender_display', 'full')

    staff = get_filtered_staff(request)

    stats = None
    if include_stats:
        stats = {
            'total':     staff.count(),
            'active':    staff.filter(is_active=True).count(),
            'male':      staff.filter(gender='M').count(),
            'female':    staff.filter(gender='F').count(),
            'full_time': staff.filter(employment_status='FT').count(),
        }

    field_names_full = {
        'staff_id': 'Staff ID', 'full_name': 'Full Name', 'first_name': 'First Name',
        'last_name': 'Last Name', 'date_of_birth': 'Date of Birth', 'gender': 'Gender',
        'nationality': 'Nationality', 'national_id': 'National ID',
        'marital_status': 'Marital Status', 'religious_affiliation': 'Religious Affiliation',
        'phone_number': 'Phone Number', 'personal_email': 'Email',
        'primary_department': 'Department', 'employment_status': 'Employment Status',
        'date_of_joining': 'Date of Joining', 'date_of_leaving': 'Date of Leaving',
        'is_active': 'Active Status',
    }
    field_names_short = {
        'staff_id': 'Staff ID', 'full_name': 'Name', 'first_name': 'First',
        'last_name': 'Last', 'date_of_birth': 'DOB', 'gender': 'Gender',
        'nationality': 'Nationality', 'national_id': 'Nat. ID',
        'marital_status': 'Marital', 'religious_affiliation': 'Religion',
        'phone_number': 'Phone', 'personal_email': 'Email',
        'primary_department': 'Dept.', 'employment_status': 'Emp. Status',
        'date_of_joining': 'Joined', 'date_of_leaving': 'Left', 'is_active': 'Active',
    }
    field_names = field_names_short if short_headers else field_names_full

    return render(request, 'hr/staff/print.html', {
        **_get_print_school_context(request),
        'staff':                staff,
        'stats':                stats,
        'now':                  timezone.now(),
        'selected_fields':      selected_fields,
        'selected_field_names': [
            field_names.get(f, f.replace('_', ' ').title()) for f in selected_fields
        ],
        'field_names':          field_names,
        'landscape':            landscape_mode,
        'short_headers':        short_headers,
        'gender_display':       gender_display,
        'title':                'Staff Report',
    })


@login_required
def export_staff_excel(request):
    staff          = get_filtered_staff(request)
    gender_display = request.GET.get('gender_display', 'full')

    def get_gender(s):
        return s.gender if gender_display == 'short' else s.get_gender_display()

    ALL_COLUMNS = [
        ('staff_id',             'Staff ID',           lambda s: s.staff_id),
        ('full_name',            'Full Name',           lambda s: s.full_name()),
        ('first_name',           'First Name',          lambda s: s.first_name),
        ('last_name',            'Last Name',           lambda s: s.last_name),
        ('date_of_birth',        'Date of Birth',       lambda s: s.date_of_birth.strftime('%Y-%m-%d') if s.date_of_birth else ''),
        ('gender',               'Gender',              lambda s: get_gender(s)),
        ('nationality',          'Nationality',         lambda s: str(s.nationality) if s.nationality else ''),
        ('national_id',          'National ID',         lambda s: s.national_id or ''),
        ('marital_status',       'Marital Status',      lambda s: s.get_marital_status_display() if s.marital_status else ''),
        ('religious_affiliation','Religion',            lambda s: s.religious_affiliation or ''),
        ('phone_number',         'Phone Number',        lambda s: s.phone_number or ''),
        ('personal_email',       'Email',               lambda s: s.personal_email or ''),
        ('primary_department',   'Department',          lambda s: s.primary_department.name if s.primary_department else ''),
        ('employment_status',    'Employment Status',   lambda s: s.get_employment_status_display()),
        ('date_of_joining',      'Date of Joining',     lambda s: s.date_of_joining.strftime('%Y-%m-%d') if s.date_of_joining else ''),
        ('date_of_leaving',      'Date of Leaving',     lambda s: s.date_of_leaving.strftime('%Y-%m-%d') if s.date_of_leaving else ''),
        ('qualification',        'Qualification',       lambda s: s.qualification or ''),
        ('is_active',            'Active',              lambda s: 'Yes' if s.is_active else 'No'),
    ]
    DEFAULT = ['staff_id', 'full_name', 'gender', 'primary_department', 'employment_status', 'phone_number']
    col_map = {k: (lbl, fn) for k, lbl, fn in ALL_COLUMNS}
    chosen  = request.GET.getlist('fields') or DEFAULT
    columns = [col_map[f] for f in chosen if f in col_map] or [col_map[f] for f in DEFAULT]

    wb = Workbook(); ws = wb.active; ws.title = 'Staff'
    ws.append([c[0] for c in columns])
    hf = PatternFill(start_color='1E3A5F', end_color='1E3A5F', fill_type='solid')
    for cell in ws[1]:
        cell.fill = hf
        cell.font = Font(bold=True, color='FFFFFF', size=11)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws.row_dimensions[1].height = 28
    af = PatternFill(start_color='F4F6F9', end_color='F4F6F9', fill_type='solid')
    da = Alignment(vertical='center')
    for i, s in enumerate(staff):
        ws.append([c[1](s) for c in columns])
        for cell in ws[i + 2]:
            cell.alignment = da
            if i % 2 == 1:
                cell.fill = af
    for col_cells in ws.columns:
        ml = max((len(str(c.value)) if c.value else 0) for c in col_cells)
        ws.column_dimensions[col_cells[0].column_letter].width = min(ml + 4, 60)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = (
        f'attachment; filename="staff_{timezone.now().strftime("%Y%m%d_%H%M")}.xlsx"'
    )
    wb.save(response)
    return response


# --- Staff wizard (creation) -------------------------------------------------

class StaffWizardFileStorage(FileSystemStorage):
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.location = os.path.join(self.location, 'wizard_temp')


class StaffCreateWizard(SessionWizardView):
    """Multi-step wizard for creating a new staff member."""

    form_list     = STAFF_WIZARD_FORMS
    template_name = 'hr/staff/wizard.html'
    file_storage  = StaffWizardFileStorage()

    def get_context_data(self, form, **kwargs):
        context       = super().get_context_data(form=form, **kwargs)
        total_steps   = len(self.form_list)
        current_index = list(self.form_list).index(self.steps.current)
        context.update({
            'step_names':          STAFF_WIZARD_STEP_NAMES,
            'current_step_name':   STAFF_WIZARD_STEP_NAMES.get(self.steps.current, 'Step'),
            'progress_percentage': (current_index / (total_steps - 1) * 100)
                                   if total_steps > 1 else 100,
        })
        if self.steps.current == 'confirmation':
            for step in ('basic_info', 'contact_info', 'employment_info',
                         'qualifications', 'banking_info', 'designation_contract'):
                context[f'{step}_data'] = self.get_cleaned_data_for_step(step)
        return context

    @transaction.atomic
    def done(self, form_list, **kwargs):
        try:
            form_data = {}
            for form in form_list:
                form_data.update(form.cleaned_data)

            staff = Staff(
                salutation=form_data.get('salutation', ''),
                first_name=form_data.get('first_name'),
                middle_name=form_data.get('middle_name', ''),
                last_name=form_data.get('last_name'),
                date_of_birth=form_data.get('date_of_birth'),
                gender=form_data.get('gender'),
                marital_status=form_data.get('marital_status', ''),
                nationality=form_data.get('nationality', ''),
                ethnicity=form_data.get('ethnicity', ''),
                religious_affiliation=form_data.get('religious_affiliation', ''),
                national_id=form_data.get('national_id', ''),
                passport_number=form_data.get('passport_number', ''),
                phone_number=form_data.get('phone_number'),
                alternative_phone=form_data.get('alternative_phone', ''),
                personal_email=form_data.get('personal_email', ''),
                emergency_contact_name=form_data.get('emergency_contact_name'),
                emergency_contact_relationship=form_data.get('emergency_contact_relationship'),
                emergency_contact_phone=form_data.get('emergency_contact_phone'),
                emergency_contact_address=form_data.get('emergency_contact_address', ''),
                primary_department=form_data.get('primary_department'),
                employment_status=form_data.get('employment_status', 'FT'),
                date_of_joining=form_data.get('date_of_joining'),
                date_of_leaving=form_data.get('date_of_leaving'),
                qualification=form_data.get('qualification', ''),
                experience=form_data.get('experience', ''),
                skills=form_data.get('skills', ''),
                languages_spoken=form_data.get('languages_spoken', ''),
                professional_memberships=form_data.get('professional_memberships', ''),
                certifications=form_data.get('certifications', ''),
                bank_account_name=form_data.get('bank_account_name', ''),
                bank_account_number=form_data.get('bank_account_number', ''),
                bank_name=form_data.get('bank_name', ''),
                bank_branch=form_data.get('bank_branch', ''),
                tax_identification_number=form_data.get('tax_identification_number', ''),
                social_security_number=form_data.get('social_security_number', ''),
                is_active=True,
            )
            staff.save()

            if form_data.get('create_designation'):
                designation = form_data.get('designation')
                if designation:
                    StaffDesignation.objects.create(
                        staff=staff,
                        designation=designation,
                        is_primary=form_data.get('is_primary_designation', True),
                        start_date=get_school_today(),
                        is_active=True,
                        role_allowance=form_data.get('role_allowance', Decimal('0.00')),
                        assignment_type='PERMANENT',
                    )

            if form_data.get('create_contract'):
                contract_start = form_data.get('contract_start_date')
                duration       = form_data.get('contract_duration_months', 12)
                if all([form_data.get('contract_type'), contract_start,
                        form_data.get('basic_salary'), form_data.get('job_title')]):
                    Contract.objects.create(
                        staff=staff,
                        contract_type=form_data['contract_type'],
                        start_date=contract_start,
                        end_date=contract_start + timedelta(days=duration * 30),
                        basic_salary=form_data['basic_salary'],
                        job_title=form_data['job_title'],
                        status='DRAFT',
                        salary_frequency='MONTHLY',
                        working_hours_per_week=40,
                        annual_leave_days=21,
                    )

            messages.success(
                self.request,
                f'Staff member {staff.full_name()} ({staff.staff_id}) created successfully!',
                extra_tags='sweetalert',
            )
            return redirect('hr:staff_profile', pk=staff.pk)

        except Exception as e:
            logger.exception('Error in StaffCreateWizard.done:')
            messages.error(self.request, f'Error creating staff: {e}', extra_tags='sweetalert-error')
            return redirect('hr:staff_list')


staff_create = StaffCreateWizard.as_view()


@login_required
def staff_edit(request, pk):
    staff = get_object_or_404(Staff, pk=pk)
    if request.method == 'POST':
        form = StaffForm(request.POST, request.FILES, instance=staff)
        if form.is_valid():
            staff = form.save()
            messages.success(request, f'{staff.full_name()} updated successfully', extra_tags='sweetalert')
            return redirect('hr:staff_profile', pk=staff.pk)
        else:
            messages.error(request, 'Please correct the errors in the form', extra_tags='sweetalert-error')
    else:
        form = StaffForm(instance=staff)
    return render(request, 'hr/staff/form.html', {
        'form': form, 'staff': staff, 'title': 'Update Staff',
    })


@login_required
def staff_delete(request, pk):
    """POST-only. modal_views.staff_delete_modal handles GET."""
    if request.method != 'POST':
        return HttpResponse(status=405)

    staff = get_object_or_404(Staff, pk=pk)

    if staff.is_active:
        if request.headers.get('HX-Request') == 'true':
            r = HttpResponse()
            r['HX-Alert-Message'] = 'Cannot delete active staff — deactivate first'
            r['HX-Alert-Type']    = 'error'
            r['HX-Close-Modal']   = 'true'
            return r
        messages.error(request, 'Cannot delete active staff — deactivate first', extra_tags='sweetalert-error')
        return redirect('hr:staff_list')

    if staff.contracts.filter(status='ACTIVE').exists():
        if request.headers.get('HX-Request') == 'true':
            r = HttpResponse()
            r['HX-Alert-Message'] = 'Cannot delete staff with active contracts'
            r['HX-Alert-Type']    = 'error'
            r['HX-Close-Modal']   = 'true'
            return r
        messages.error(request, 'Cannot delete staff with active contracts', extra_tags='sweetalert-error')
        return redirect('hr:staff_list')

    if hasattr(staff, 'payrolls') and staff.payrolls.exists():
        if request.headers.get('HX-Request') == 'true':
            r = HttpResponse()
            r['HX-Alert-Message'] = 'Cannot delete staff with payroll records'
            r['HX-Alert-Type']    = 'error'
            r['HX-Close-Modal']   = 'true'
            return r
        messages.error(request, 'Cannot delete staff with payroll records', extra_tags='sweetalert-error')
        return redirect('hr:staff_list')

    name = staff.full_name()
    staff.delete()

    if request.headers.get('HX-Request') == 'true':
        r = HttpResponse()
        r['HX-Alert-Message'] = f'Staff member "{name}" deleted successfully'
        r['HX-Alert-Type']    = 'success'
        r['HX-Close-Modal']   = 'true'
        r['HX-Redirect']      = reverse('hr:staff_list')
        return r
    messages.success(request, f'Staff member "{name}" deleted successfully', extra_tags='sweetalert')
    return redirect('hr:staff_list')


@login_required
def staff_activate(request, pk):
    """POST-only. modal_views.staff_activate_modal handles GET."""
    if request.method != 'POST':
        return HttpResponse(status=405)

    staff = get_object_or_404(Staff, pk=pk)

    if staff.is_active:
        if request.headers.get('HX-Request') == 'true':
            r = HttpResponse()
            r['HX-Alert-Message'] = f'{staff.full_name()} is already active'
            r['HX-Alert-Type']    = 'warning'
            r['HX-Close-Modal']   = 'true'
            return r
        messages.warning(request, f'{staff.full_name()} is already active', extra_tags='sweetalert')
        return redirect('hr:staff_profile', pk=pk)

    staff.is_active = True
    staff.save(update_fields=['is_active', 'updated_at'])

    if request.headers.get('HX-Request') == 'true':
        r = HttpResponse()
        r['HX-Alert-Message'] = f'{staff.full_name()} activated successfully'
        r['HX-Alert-Type']    = 'success'
        r['HX-Close-Modal']   = 'true'
        r['HX-Redirect']      = reverse('hr:staff_profile', kwargs={'pk': pk})
        return r
    messages.success(request, f'{staff.full_name()} activated successfully', extra_tags='sweetalert')
    return redirect('hr:staff_profile', pk=pk)


@login_required
def staff_deactivate(request, pk):
    """POST-only. modal_views.staff_deactivate_modal handles GET."""
    if request.method != 'POST':
        return HttpResponse(status=405)

    staff = get_object_or_404(Staff, pk=pk)

    if not staff.is_active:
        if request.headers.get('HX-Request') == 'true':
            r = HttpResponse()
            r['HX-Alert-Message'] = f'{staff.full_name()} is already inactive'
            r['HX-Alert-Type']    = 'warning'
            r['HX-Close-Modal']   = 'true'
            return r
        messages.warning(request, f'{staff.full_name()} is already inactive', extra_tags='sweetalert')
        return redirect('hr:staff_profile', pk=pk)

    staff.is_active = False
    staff.save(update_fields=['is_active', 'updated_at'])

    if request.headers.get('HX-Request') == 'true':
        r = HttpResponse()
        r['HX-Alert-Message'] = f'{staff.full_name()} deactivated'
        r['HX-Alert-Type']    = 'warning'
        r['HX-Close-Modal']   = 'true'
        r['HX-Redirect']      = reverse('hr:staff_profile', kwargs={'pk': pk})
        return r
    messages.warning(request, f'{staff.full_name()} deactivated', extra_tags='sweetalert')
    return redirect('hr:staff_profile', pk=pk)


# --- Staff designation management (lives in staff context) ------------------

@login_required
def staff_assign_designation(request, staff_pk):
    """POST-only. modal_views.staff_assign_designation_modal handles GET."""
    if request.method != 'POST':
        return HttpResponse(status=405)

    staff = get_object_or_404(Staff, pk=staff_pk)
    form  = StaffDesignationForm(request.POST)

    if form.is_valid():
        form.save()
        if request.headers.get('HX-Request') == 'true':
            r = HttpResponse()
            r['HX-Alert-Message'] = f'Designation assigned to {staff.full_name()}'
            r['HX-Alert-Type']    = 'success'
            r['HX-Close-Modal']   = 'true'
            r['HX-Redirect']      = reverse('hr:staff_profile', kwargs={'pk': staff_pk}) + '#employment'
            return r
        messages.success(request, 'Designation assigned', extra_tags='sweetalert')
        return redirect('hr:staff_profile', pk=staff_pk)
    else:
        # Re-render modal with errors
        return render(request, 'hr/staff/modals/assign_designation.html', {
            'staff': staff,
            'form':  form,
        })


@login_required
def staff_designation_edit(request, pk):
    """POST-only. modal_views.staff_designation_edit_modal handles GET."""
    if request.method != 'POST':
        return HttpResponse(status=405)

    sd   = get_object_or_404(StaffDesignation, pk=pk)
    form = StaffDesignationForm(request.POST, instance=sd)

    if form.is_valid():
        form.save()
        if request.headers.get('HX-Request') == 'true':
            r = HttpResponse()
            r['HX-Alert-Message'] = f'Designation updated for {sd.staff.full_name()}'
            r['HX-Alert-Type']    = 'success'
            r['HX-Close-Modal']   = 'true'
            r['HX-Redirect']      = reverse('hr:staff_profile', kwargs={'pk': sd.staff.pk}) + '#employment'
            return r
        messages.success(request, 'Designation updated', extra_tags='sweetalert')
        return redirect('hr:staff_profile', pk=sd.staff.pk)
    else:
        return render(request, 'hr/staff/modals/edit_designation.html', {
            'sd':   sd,
            'form': form,
        })


@login_required
def staff_designation_activate(request, pk):
    """POST-only. modal_views.staff_designation_activate_modal handles GET."""
    if request.method != 'POST':
        return HttpResponse(status=405)

    sd = get_object_or_404(StaffDesignation, pk=pk)
    sd.end_date  = None
    sd.is_active = True
    sd.save()

    if request.headers.get('HX-Request') == 'true':
        r = HttpResponse()
        r['HX-Alert-Message'] = f'Designation {sd.designation.name} reactivated'
        r['HX-Alert-Type']    = 'success'
        r['HX-Close-Modal']   = 'true'
        r['HX-Redirect']      = reverse('hr:staff_profile', kwargs={'pk': sd.staff.pk})
        return r
    messages.success(request, f'Designation {sd.designation.name} reactivated', extra_tags='sweetalert')
    return redirect('hr:staff_profile', pk=sd.staff.pk)


@login_required
def staff_designation_deactivate(request, pk):
    """POST-only. modal_views.staff_designation_deactivate_modal handles GET."""
    if request.method != 'POST':
        return HttpResponse(status=405)

    sd = get_object_or_404(StaffDesignation, pk=pk)

    if sd.is_primary:
        other_active = StaffDesignation.objects.filter(
            staff=sd.staff, is_active=True
        ).exclude(pk=sd.pk).exists()
        if not other_active:
            if request.headers.get('HX-Request') == 'true':
                r = HttpResponse()
                r['HX-Alert-Message'] = 'Cannot deactivate the only active designation — assign another first'
                r['HX-Alert-Type']    = 'error'
                r['HX-Close-Modal']   = 'true'
                return r
            messages.error(
                request,
                'Cannot deactivate the only active designation — assign another first',
                extra_tags='sweetalert-error',
            )
            return redirect('hr:staff_profile', pk=sd.staff.pk)

    sd.end_date  = get_school_today()
    sd.is_active = False
    sd.save()

    if request.headers.get('HX-Request') == 'true':
        r = HttpResponse()
        r['HX-Alert-Message'] = f'Designation {sd.designation.name} deactivated'
        r['HX-Alert-Type']    = 'warning'
        r['HX-Close-Modal']   = 'true'
        r['HX-Redirect']      = reverse('hr:staff_profile', kwargs={'pk': sd.staff.pk})
        return r
    messages.warning(request, f'Designation {sd.designation.name} deactivated', extra_tags='sweetalert')
    return redirect('hr:staff_profile', pk=sd.staff.pk)


@login_required
def staff_designation_set_primary(request, pk):
    """POST-only. modal_views.staff_designation_set_primary_modal handles GET."""
    if request.method != 'POST':
        return HttpResponse(status=405)

    sd = get_object_or_404(StaffDesignation, pk=pk)

    if not sd.is_active:
        if request.headers.get('HX-Request') == 'true':
            r = HttpResponse()
            r['HX-Alert-Message'] = 'Cannot set inactive designation as primary — activate it first'
            r['HX-Alert-Type']    = 'error'
            r['HX-Close-Modal']   = 'true'
            return r
        messages.error(
            request,
            'Cannot set inactive designation as primary — activate it first',
            extra_tags='sweetalert-error',
        )
        return redirect('hr:staff_profile', pk=sd.staff.pk)

    StaffDesignation.objects.filter(
        staff=sd.staff, is_primary=True
    ).update(is_primary=False)
    sd.is_primary = True
    sd.save()

    if request.headers.get('HX-Request') == 'true':
        r = HttpResponse()
        r['HX-Alert-Message'] = f'{sd.designation.name} set as primary designation'
        r['HX-Alert-Type']    = 'success'
        r['HX-Close-Modal']   = 'true'
        r['HX-Redirect']      = reverse('hr:staff_profile', kwargs={'pk': sd.staff.pk})
        return r
    messages.success(request, f'{sd.designation.name} set as primary designation', extra_tags='sweetalert')
    return redirect('hr:staff_profile', pk=sd.staff.pk)


@login_required
def staff_designation_delete(request, pk):
    """POST-only. modal_views.staff_designation_delete_modal handles GET."""
    if request.method != 'POST':
        return HttpResponse(status=405)

    sd         = get_object_or_404(StaffDesignation, pk=pk)
    staff      = sd.staff
    desig_name = sd.designation.name
    sd.delete()

    if request.headers.get('HX-Request') == 'true':
        r = HttpResponse()
        r['HX-Alert-Message'] = f'Designation {desig_name} removed from {staff.full_name()}'
        r['HX-Alert-Type']    = 'success'
        r['HX-Close-Modal']   = 'true'
        r['HX-Redirect']      = reverse('hr:staff_profile', kwargs={'pk': staff.pk})
        return r
    messages.success(
        request,
        f'Designation {desig_name} removed from {staff.full_name()}',
        extra_tags='sweetalert',
    )
    return redirect('hr:staff_profile', pk=staff.pk)


# =============================================================================
# CONTRACTS
# =============================================================================

def get_filtered_contracts(request):
    contracts = Contract.objects.select_related('staff').annotate(
        status_order=Case(
            When(status='ACTIVE', then=0),
            default=1,
            output_field=IntegerField(),
        )
    ).order_by('status_order', '-start_date', 'staff__first_name')

    query = request.GET.get('q', '').strip()
    if query:
        combined_q = Q()
        for word in query.split():
            combined_q &= (
                Q(contract_number__icontains=word)   |
                Q(staff__first_name__icontains=word) |
                Q(staff__last_name__icontains=word)  |
                Q(staff__staff_id__icontains=word)   |
                Q(job_title__icontains=word)
            )
        contracts = contracts.filter(combined_q)

    contract_type = request.GET.get('contract_type', '')
    if contract_type:
        contracts = contracts.filter(contract_type=contract_type)

    status = request.GET.get('status', '')
    if status:
        contracts = contracts.filter(status=status)

    staff = request.GET.get('staff', '')
    if staff:
        contracts = contracts.filter(staff_id=staff)

    salary_frequency = request.GET.get('salary_frequency', '')
    if salary_frequency:
        contracts = contracts.filter(salary_frequency=salary_frequency)

    start_date = request.GET.get('start_date', '')
    if start_date:
        contracts = contracts.filter(start_date__gte=start_date)

    end_date = request.GET.get('end_date', '')
    if end_date:
        contracts = contracts.filter(end_date__lte=end_date)

    if request.GET.get('is_permanent', '').lower() == 'true':
        contracts = contracts.filter(contract_type='PERMANENT')

    if request.GET.get('expiring_soon', '').lower() == 'true':
        threshold = timezone.now().date() + timedelta(days=30)
        contracts = contracts.filter(
            status='ACTIVE',
            end_date__lte=threshold,
            end_date__gte=timezone.now().date(),
        )

    return contracts


@login_required
def contract_list(request):
    filter_form = ContractFilterForm(request.GET or None)
    contracts   = get_filtered_contracts(request)
    today       = timezone.now().date()

    stats = {
        'total':                   contracts.count(),
        'active':                  contracts.filter(status='ACTIVE').count(),
        'draft':                   contracts.filter(status='DRAFT').count(),
        'expired':                 contracts.filter(status='EXPIRED').count(),
        'terminated':              contracts.filter(status='TERMINATED').count(),
        'permanent':               contracts.filter(contract_type='PERMANENT').count(),
        'fixed_term':              contracts.filter(contract_type='FIXED_TERM').count(),
        'expiring_soon':           contracts.filter(
                                       status='ACTIVE',
                                       end_date__lte=today + timedelta(days=30),
                                       end_date__gte=today,
                                   ).count(),
        'avg_salary':              contracts.filter(status='ACTIVE').aggregate(
                                       Avg('basic_salary')
                                   )['basic_salary__avg'] or 0,
        'total_salary_obligation': contracts.filter(status='ACTIVE').aggregate(
                                       Sum('basic_salary')
                                   )['basic_salary__sum'] or 0,
    }

    paginator      = Paginator(contracts, 20)
    contracts_page = paginator.get_page(request.GET.get('page', 1))
    is_htmx        = request.headers.get('HX-Request') == 'true'

    context = {
        'contracts_page': contracts_page,
        'paginator':      paginator,
        'stats':          stats,
        'filter_form':    filter_form,
        'is_htmx':        is_htmx,
    }
    if is_htmx:
        return render(request, 'hr/contracts/partials/_contract_results.html', context)
    return render(request, 'hr/contracts/list.html', context)


@login_required
def contract_detail(request, pk):
    contract = get_object_or_404(
        Contract.objects.select_related('staff', 'staff__primary_department'), pk=pk
    )
    from .utils import (
        calculate_monthly_salary, calculate_annual_salary,
        calculate_daily_rate, calculate_hourly_rate,
    )
    return render(request, 'hr/contracts/detail.html', {
        'contract': contract,
        'metrics': {
            'monthly_salary':    calculate_monthly_salary(contract),
            'annual_salary':     calculate_annual_salary(contract),
            'daily_rate':        calculate_daily_rate(contract),
            'hourly_rate':       calculate_hourly_rate(contract),
            'days_until_expiry': contract.days_until_expiry,
            'duration_in_months': contract.duration_in_months,
            'is_permanent':      contract.is_permanent,
            'is_probationary':   contract.is_probationary,
        },
        'benefits': contract.benefits.filter(is_active=True),
    })


@login_required
def contract_create(request):
    if request.method == 'POST':
        form = ContractForm(request.POST, request.FILES)
        if form.is_valid():
            contract = form.save()
            messages.success(request, f'Contract {contract.contract_number} created successfully', extra_tags='sweetalert')
            return redirect('hr:contract_detail', pk=contract.pk)
        else:
            messages.error(request, 'Please correct the errors in the form', extra_tags='sweetalert-error')
    else:
        form = ContractForm()
    return render(request, 'hr/contracts/form.html', {'form': form, 'title': 'Create Contract'})


@login_required
def contract_edit(request, pk):
    contract = get_object_or_404(Contract, pk=pk)
    if request.method == 'POST':
        form = ContractForm(request.POST, request.FILES, instance=contract)
        if form.is_valid():
            contract = form.save()
            messages.success(request, f'Contract {contract.contract_number} updated successfully', extra_tags='sweetalert')
            return redirect('hr:contract_detail', pk=contract.pk)
        else:
            messages.error(request, 'Please correct the errors in the form', extra_tags='sweetalert-error')
    else:
        form = ContractForm(instance=contract)
    return render(request, 'hr/contracts/form.html', {
        'form': form, 'contract': contract, 'title': 'Update Contract',
    })


@login_required
def contract_delete(request, pk):
    """POST-only. modal_views.contract_delete_modal handles GET."""
    if request.method != 'POST':
        return HttpResponse(status=405)

    contract = get_object_or_404(Contract, pk=pk)

    if contract.status == 'ACTIVE':
        if request.headers.get('HX-Request') == 'true':
            r = HttpResponse()
            r['HX-Alert-Message'] = 'Cannot delete active contracts'
            r['HX-Alert-Type']    = 'error'
            r['HX-Close-Modal']   = 'true'
            return r
        messages.error(request, 'Cannot delete active contracts', extra_tags='sweetalert-error')
        return redirect('hr:contract_list')

    number = contract.contract_number
    contract.delete()

    if request.headers.get('HX-Request') == 'true':
        r = HttpResponse()
        r['HX-Alert-Message'] = f'Contract "{number}" deleted successfully'
        r['HX-Alert-Type']    = 'success'
        r['HX-Close-Modal']   = 'true'
        r['HX-Redirect']      = reverse('hr:contract_list')
        return r
    messages.success(request, f'Contract "{number}" deleted successfully', extra_tags='sweetalert')
    return redirect('hr:contract_list')


@login_required
def contract_activate(request, pk):
    """POST-only. modal_views.contract_activate_modal handles GET."""
    if request.method != 'POST':
        return HttpResponse(status=405)

    contract = get_object_or_404(Contract, pk=pk)
    contract.activate(user=request.user)

    if request.headers.get('HX-Request') == 'true':
        r = HttpResponse()
        r['HX-Alert-Message'] = f'Contract {contract.contract_number} activated'
        r['HX-Alert-Type']    = 'success'
        r['HX-Close-Modal']   = 'true'
        r['HX-Redirect']      = reverse('hr:contract_detail', kwargs={'pk': pk})
        return r
    messages.success(request, f'Contract {contract.contract_number} activated', extra_tags='sweetalert')
    return redirect('hr:contract_detail', pk=pk)


@login_required
def contract_terminate(request, pk):
    """POST-only. modal_views.contract_terminate_modal handles GET."""
    if request.method != 'POST':
        return HttpResponse(status=405)

    contract      = get_object_or_404(Contract, pk=pk)
    reason        = request.POST.get('termination_reason', 'OTHER')
    notes         = request.POST.get('termination_notes', '')
    term_date_raw = request.POST.get('termination_date')

    from datetime import datetime as dt
    term_date = (
        dt.strptime(term_date_raw, '%Y-%m-%d').date()
        if term_date_raw else get_school_today()
    )
    contract.terminate(reason=reason, user=request.user, termination_date=term_date, notes=notes)

    if request.headers.get('HX-Request') == 'true':
        r = HttpResponse()
        r['HX-Alert-Message'] = f'Contract {contract.contract_number} terminated'
        r['HX-Alert-Type']    = 'warning'
        r['HX-Close-Modal']   = 'true'
        r['HX-Redirect']      = reverse('hr:contract_detail', kwargs={'pk': pk})
        return r
    messages.warning(request, f'Contract {contract.contract_number} terminated', extra_tags='sweetalert')
    return redirect('hr:contract_detail', pk=pk)


@login_required
def contract_renew(request, pk):
    """POST-only. modal_views.contract_renew_modal handles GET."""
    if request.method != 'POST':
        return HttpResponse(status=405)

    contract = get_object_or_404(Contract, pk=pk)

    if contract.status not in ['ACTIVE', 'EXPIRED']:
        if request.headers.get('HX-Request') == 'true':
            r = HttpResponse()
            r['HX-Alert-Message'] = f'Cannot renew contract with status: {contract.get_status_display()}'
            r['HX-Alert-Type']    = 'error'
            r['HX-Close-Modal']   = 'true'
            return r
        messages.error(
            request,
            f'Cannot renew contract with status: {contract.get_status_display()}',
            extra_tags='sweetalert-error',
        )
        return redirect('hr:contract_detail', pk=pk)

    new_end_date = None
    raw = request.POST.get('new_end_date')
    if raw:
        from datetime import datetime as dt
        try:
            new_end_date = dt.strptime(raw, '%Y-%m-%d').date()
        except ValueError:
            pass

    try:
        contract.renew(new_end_date=new_end_date, user=request.user)
        if request.headers.get('HX-Request') == 'true':
            r = HttpResponse()
            r['HX-Alert-Message'] = f'Contract {contract.contract_number} renewed successfully'
            r['HX-Alert-Type']    = 'success'
            r['HX-Close-Modal']   = 'true'
            r['HX-Redirect']      = reverse('hr:contract_detail', kwargs={'pk': pk})
            return r
        messages.success(request, f'Contract {contract.contract_number} renewed successfully', extra_tags='sweetalert')
    except Exception as e:
        logger.error(f'Error renewing contract: {e}')
        if request.headers.get('HX-Request') == 'true':
            r = HttpResponse()
            r['HX-Alert-Message'] = f'Error renewing contract: {e}'
            r['HX-Alert-Type']    = 'error'
            r['HX-Close-Modal']   = 'true'
            return r
        messages.error(request, f'Error renewing contract: {e}', extra_tags='sweetalert-error')

    return redirect('hr:contract_detail', pk=pk)


# =============================================================================
# TEACHERS
# =============================================================================

def get_filtered_teachers(request):
    teachers = Teacher.objects.select_related(
        'staff', 'staff__primary_department',
    ).prefetch_related(
        'assigned_classes', 'assigned_classes__academic_level',
        'qualified_subjects', 'preferred_academic_levels',
        Prefetch(
            'staff__staffdesignation_set',
            queryset=StaffDesignation.objects.filter(
                is_primary=True, is_active=True
            ).select_related('designation'),
            to_attr='primary_staff_designation',
        ),
    ).annotate(
        assigned_classes_count=Count('assigned_classes', distinct=True),
        qualified_subjects_count=Count('qualified_subjects', distinct=True),
    )

    is_active = request.GET.get('is_active', '')
    if is_active:
        teachers = teachers.filter(is_active=(is_active.lower() == 'true'))
    else:
        teachers = teachers.filter(is_active=True)

    query = request.GET.get('q', '').strip()
    if query:
        combined_q = Q()
        for word in query.split():
            combined_q &= (
                Q(staff__first_name__icontains=word)    |
                Q(staff__middle_name__icontains=word)   |
                Q(staff__last_name__icontains=word)     |
                Q(staff__staff_id__icontains=word)      |
                Q(specialization__icontains=word)       |
                Q(staff__phone_number__icontains=word)  |
                Q(staff__personal_email__icontains=word)
            )
        teachers = teachers.filter(combined_q)

    department = request.GET.get('department', '')
    if department:
        teachers = teachers.filter(staff__primary_department_id=department)

    specialization = request.GET.get('specialization', '')
    if specialization:
        teachers = teachers.filter(specialization__icontains=specialization)

    is_class_teacher = request.GET.get('is_class_teacher', '')
    if is_class_teacher:
        teachers = teachers.filter(is_class_teacher=(is_class_teacher.lower() == 'true'))

    can_teach_online = request.GET.get('can_teach_online', '')
    if can_teach_online:
        teachers = teachers.filter(can_teach_online=(can_teach_online.lower() == 'true'))

    digital_literacy_level = request.GET.get('digital_literacy_level', '')
    if digital_literacy_level:
        teachers = teachers.filter(digital_literacy_level=digital_literacy_level)

    return teachers.order_by('-is_active', 'staff__first_name', 'staff__last_name')


@login_required
def teacher_list(request):
    filter_form = TeacherFilterForm(request.GET or None)
    teachers    = get_filtered_teachers(request)

    stats = {
        'total_teachers':            teachers.count(),
        'active_teachers':           teachers.filter(is_active=True).count(),
        'inactive_teachers':         teachers.filter(is_active=False).count(),
        'class_teachers':            teachers.filter(is_class_teacher=True, is_active=True).count(),
        'teachers_can_teach_online': teachers.filter(can_teach_online=True, is_active=True).count(),
        'overloaded':                teachers.filter(
                                         is_active=True,
                                         current_teaching_load__gt=F('max_hours_per_week'),
                                     ).count(),
        'avg_teaching_load':         round(
                                         teachers.filter(is_active=True).aggregate(
                                             Avg('current_teaching_load')
                                         )['current_teaching_load__avg'] or 0, 1
                                     ),
    }

    paginator     = Paginator(teachers, 20)
    teachers_page = paginator.get_page(request.GET.get('page', 1))
    is_htmx       = request.headers.get('HX-Request') == 'true'

    context = {
        'teachers_page': teachers_page,
        'paginator':     paginator,
        'stats':         stats,
        'filter_form':   filter_form,
        'is_htmx':       is_htmx,
    }
    if is_htmx:
        return render(request, 'hr/teachers/partials/_teacher_results.html', context)
    return render(request, 'hr/teachers/list.html', context)

@login_required
def teacher_edit(request, pk):
    """POST-only. modal_views.teacher_edit_modal handles GET."""
    if request.method != 'POST':
        return HttpResponse(status=405)

    teacher = get_object_or_404(Teacher, pk=pk)
    form    = TeacherForm(request.POST, instance=teacher)

    if form.is_valid():
        form.save()
        if request.headers.get('HX-Request') == 'true':
            r = HttpResponse()
            r['HX-Alert-Message'] = f'Teacher profile updated for {teacher.staff.full_name()}'
            r['HX-Alert-Type']    = 'success'
            r['HX-Close-Modal']   = 'true'
            r['HX-Redirect']      = reverse('hr:staff_profile', kwargs={'pk': teacher.staff.pk}) + '#teaching'
            return r
        messages.success(request, 'Teacher profile updated', extra_tags='sweetalert')
        return redirect('hr:staff_profile', pk=teacher.staff.pk)
    else:
        # Re-render modal with errors
        return render(request, 'hr/teachers/modals/edit.html', {
            'teacher': teacher,
            'form':    form,
        })


@login_required
def teacher_delete(request, pk):
    """POST-only. modal_views.teacher_delete_modal handles GET."""
    if request.method != 'POST':
        return HttpResponse(status=405)

    teacher = get_object_or_404(Teacher, pk=pk)

    if teacher.assigned_classes.exists():
        if request.headers.get('HX-Request') == 'true':
            r = HttpResponse()
            r['HX-Alert-Message'] = 'Cannot delete teacher with assigned classes'
            r['HX-Alert-Type']    = 'error'
            r['HX-Close-Modal']   = 'true'
            return r
        messages.error(request, 'Cannot delete teacher with assigned classes', extra_tags='sweetalert-error')
        return redirect('hr:teacher_list')

    name     = teacher.staff.full_name()
    staff_pk = teacher.staff.pk
    teacher.delete()

    if request.headers.get('HX-Request') == 'true':
        r = HttpResponse()
        r['HX-Alert-Message'] = f'Teacher profile for "{name}" deleted'
        r['HX-Alert-Type']    = 'success'
        r['HX-Close-Modal']   = 'true'
        r['HX-Redirect']      = reverse('hr:staff_profile', kwargs={'pk': staff_pk})
        return r
    messages.success(request, f'Teacher profile for "{name}" deleted', extra_tags='sweetalert')
    return redirect('hr:staff_profile', pk=staff_pk)


@login_required
def teacher_activate(request, pk):
    """POST-only. modal_views.teacher_reactivate_modal handles GET."""
    if request.method != 'POST':
        return HttpResponse(status=405)

    teacher = get_object_or_404(Teacher, pk=pk)
    teacher.is_active = True
    teacher.save(update_fields=['is_active', 'updated_at'])

    if request.headers.get('HX-Request') == 'true':
        r = HttpResponse()
        r['HX-Alert-Message'] = f'Teacher profile for {teacher.staff.full_name()} reactivated'
        r['HX-Alert-Type']    = 'success'
        r['HX-Close-Modal']   = 'true'
        r['HX-Redirect']      = reverse('hr:teacher_profile', kwargs={'pk': pk})
        return r
    messages.success(
        request,
        f'Teacher profile for {teacher.staff.full_name()} reactivated',
        extra_tags='sweetalert',
    )
    return redirect('hr:teacher_profile', pk=pk)


@login_required
def teacher_deactivate(request, pk):
    """POST-only. modal_views.teacher_deactivate_modal handles GET."""
    if request.method != 'POST':
        return HttpResponse(status=405)

    teacher = get_object_or_404(Teacher, pk=pk)

    has_teaching_designation = StaffDesignation.objects.filter(
        staff=teacher.staff, designation__is_teaching=True, is_active=True
    ).exists()

    if has_teaching_designation:
        if request.headers.get('HX-Request') == 'true':
            r = HttpResponse()
            r['HX-Alert-Message'] = (
                f'Cannot deactivate {teacher.staff.full_name()} '
                '— remove teaching designations first'
            )
            r['HX-Alert-Type']  = 'warning'
            r['HX-Close-Modal'] = 'true'
            return r
        messages.warning(
            request,
            f'Cannot deactivate {teacher.staff.full_name()} — remove teaching designations first',
            extra_tags='sweetalert',
        )
        return redirect('hr:teacher_profile', pk=pk)

    teacher.is_active = False
    teacher.save(update_fields=['is_active', 'updated_at'])

    if request.headers.get('HX-Request') == 'true':
        r = HttpResponse()
        r['HX-Alert-Message'] = f'Teacher profile for {teacher.staff.full_name()} deactivated'
        r['HX-Alert-Type']    = 'warning'
        r['HX-Close-Modal']   = 'true'
        r['HX-Redirect']      = reverse('hr:teacher_list')
        return r
    messages.success(
        request,
        f'Teacher profile for {teacher.staff.full_name()} deactivated',
        extra_tags='sweetalert',
    )
    return redirect('hr:teacher_list')

# =============================================================================
# ATTENDANCE
# =============================================================================

def get_filtered_attendance(request):
    records = Attendance.objects.select_related(
        'staff__primary_department'
    ).order_by('-date', 'staff__first_name')

    query = request.GET.get('q', '').strip()
    if query:
        combined_q = Q()
        for word in query.split():
            combined_q &= (
                Q(staff__first_name__icontains=word) |
                Q(staff__last_name__icontains=word)  |
                Q(staff__staff_id__icontains=word)
            )
        records = records.filter(combined_q)

    staff = request.GET.get('staff', '')
    if staff:
        records = records.filter(staff_id=staff)

    status = request.GET.get('status', '')
    if status:
        records = records.filter(status=status)

    work_mode = request.GET.get('work_mode', '')
    if work_mode:
        records = records.filter(work_mode=work_mode)

    date_from = request.GET.get('date_from', '')
    if date_from:
        records = records.filter(date__gte=date_from)

    date_to = request.GET.get('date_to', '')
    if date_to:
        records = records.filter(date__lte=date_to)

    return records


@login_required
def attendance_list(request):
    filter_form = AttendanceFilterForm(request.GET or None)
    records     = get_filtered_attendance(request)

    stats = {
        'total':          records.count(),
        'present':        records.filter(status='PRESENT').count(),
        'absent':         records.filter(status='ABSENT').count(),
        'late':           records.filter(status='LATE').count(),
        'on_leave':       records.filter(status='LEAVE').count(),
        'office':         records.filter(work_mode='OFFICE').count(),
        'remote':         records.filter(work_mode='REMOTE').count(),
        'avg_work_hours': records.filter(work_hours__isnull=False).aggregate(
                              Avg('work_hours')
                          )['work_hours__avg'] or 0,
        'total_overtime': records.aggregate(
                              Sum('overtime_hours')
                          )['overtime_hours__sum'] or 0,
    }

    paginator       = Paginator(records, 20)
    attendance_page = paginator.get_page(request.GET.get('page', 1))
    is_htmx         = request.headers.get('HX-Request') == 'true'

    context = {
        'attendance_page': attendance_page,
        'paginator':       paginator,
        'stats':           stats,
        'filter_form':     filter_form,
        'is_htmx':         is_htmx,
    }
    if is_htmx:
        return render(request, 'hr/attendance/partials/_attendance_results.html', context)
    return render(request, 'hr/attendance/list.html', context)


@login_required
def attendance_detail(request, pk):
    attendance = get_object_or_404(Attendance, pk=pk)
    return render(request, 'hr/attendance/detail.html', {'attendance': attendance})


@login_required
def attendance_create(request):
    if request.method == 'POST':
        form = AttendanceForm(request.POST)
        if form.is_valid():
            attendance = form.save()
            messages.success(
                request,
                f'Attendance recorded for {attendance.staff.full_name()}',
                extra_tags='sweetalert',
            )
            return redirect('hr:attendance_list')
        else:
            messages.error(request, 'Please correct the errors in the form', extra_tags='sweetalert-error')
    else:
        form = AttendanceForm()
    return render(request, 'hr/attendance/form.html', {'form': form, 'title': 'Record Attendance'})


@login_required
def attendance_edit(request, pk):
    attendance = get_object_or_404(Attendance, pk=pk)
    if request.method == 'POST':
        form = AttendanceForm(request.POST, instance=attendance)
        if form.is_valid():
            attendance = form.save()
            messages.success(
                request,
                f'Attendance updated for {attendance.staff.full_name()}',
                extra_tags='sweetalert',
            )
            return redirect('hr:attendance_list')
        else:
            messages.error(request, 'Please correct the errors in the form', extra_tags='sweetalert-error')
    else:
        form = AttendanceForm(instance=attendance)
    return render(request, 'hr/attendance/form.html', {
        'form': form, 'attendance': attendance, 'title': 'Update Attendance',
    })


@login_required
def attendance_delete(request, pk):
    """POST-only. modal_views.attendance_delete_modal handles GET."""
    if request.method != 'POST':
        return HttpResponse(status=405)

    attendance      = get_object_or_404(Attendance, pk=pk)
    staff_name      = attendance.staff.full_name()
    attendance_date = attendance.date
    attendance.delete()

    if request.headers.get('HX-Request') == 'true':
        r = HttpResponse()
        r['HX-Alert-Message'] = f'Attendance for {staff_name} on {attendance_date} deleted'
        r['HX-Alert-Type']    = 'success'
        r['HX-Close-Modal']   = 'true'
        r['HX-Redirect']      = reverse('hr:attendance_list')
        return r
    messages.success(
        request,
        f'Attendance for {staff_name} on {attendance_date} deleted',
        extra_tags='sweetalert',
    )
    return redirect('hr:attendance_list')


@login_required
@transaction.atomic
def bulk_attendance_record(request):
    """POST-only. modal_views.bulk_attendance_modal handles GET."""
    if request.method != 'POST':
        return HttpResponse(status=405)

    attendance_date_raw = request.POST.get('date')
    status    = request.POST.get('status', 'PRESENT')
    work_mode = request.POST.get('work_mode', 'OFFICE')

    from datetime import datetime as dt
    attendance_date = (
        dt.strptime(attendance_date_raw, '%Y-%m-%d').date()
        if attendance_date_raw else get_school_today()
    )

    existing = Attendance.objects.filter(date=attendance_date)
    if existing.exists():
        if request.headers.get('HX-Request') == 'true':
            r = HttpResponse()
            r['HX-Alert-Message'] = (
                f'Attendance already recorded for {existing.count()} '
                f'staff on {attendance_date}'
            )
            r['HX-Alert-Type']  = 'warning'
            r['HX-Close-Modal'] = 'true'
            return r
        messages.warning(
            request,
            f'Attendance already recorded for {existing.count()} staff on {attendance_date}',
            extra_tags='sweetalert',
        )
        return redirect('hr:attendance_list')

    active_staff = Staff.objects.filter(is_active=True)
    Attendance.objects.bulk_create([
        Attendance(staff=s, date=attendance_date, status=status, work_mode=work_mode)
        for s in active_staff
    ])
    count = active_staff.count()

    if request.headers.get('HX-Request') == 'true':
        r = HttpResponse()
        r['HX-Alert-Message'] = f'Attendance recorded for {count} staff on {attendance_date}'
        r['HX-Alert-Type']    = 'success'
        r['HX-Close-Modal']   = 'true'
        r['HX-Redirect']      = reverse('hr:attendance_list')
        return r
    messages.success(
        request,
        f'Attendance recorded for {count} staff on {attendance_date}',
        extra_tags='sweetalert',
    )
    return redirect('hr:attendance_list')


# =============================================================================
# PAYROLL
# =============================================================================

def get_filtered_payrolls(request):
    from calendar import monthrange
    from datetime import date

    payrolls = Payroll.objects.select_related(
        'staff', 'staff__primary_department',
        'fiscal_period', 'fiscal_period__fiscal_year',
        'payment_method',
    ).prefetch_related(
        'allowances', 'deductions', 'bonuses',
    ).order_by('-payment_date', 'staff__first_name')

    query = request.GET.get('q', '').strip()
    if query:
        combined_q = Q()
        for word in query.split():
            combined_q &= (
                Q(staff__first_name__icontains=word)   |
                Q(staff__last_name__icontains=word)    |
                Q(staff__staff_id__icontains=word)     |
                Q(payment_reference__icontains=word)   |
                Q(pay_period_label__icontains=word)
            )
        payrolls = payrolls.filter(combined_q)

    staff = request.GET.get('staff', '')
    if staff:
        payrolls = payrolls.filter(staff_id=staff)

    fiscal_period = request.GET.get('fiscal_period', '')
    if fiscal_period:
        payrolls = payrolls.filter(fiscal_period_id=fiscal_period)

    fiscal_year = request.GET.get('fiscal_year', '')
    if fiscal_year:
        payrolls = payrolls.filter(fiscal_period__fiscal_year_id=fiscal_year)

    pay_frequency = request.GET.get('pay_frequency', '')
    if pay_frequency:
        payrolls = payrolls.filter(pay_frequency=pay_frequency)

    status = request.GET.get('status', '')
    if status:
        payrolls = payrolls.filter(status=status)

    payment_method = request.GET.get('payment_method', '')
    if payment_method:
        payrolls = payrolls.filter(payment_method_id=payment_method)

    currency = request.GET.get('currency', '')
    if currency:
        payrolls = payrolls.filter(currency=currency.upper())

    if request.GET.get('only_reversed') == 'on':
        payrolls = payrolls.filter(reversed=True)

    if request.GET.get('only_prorated') == 'on':
        payrolls = payrolls.filter(is_prorated=True)

    quick_filter = request.GET.get('quick_filter', '')
    if quick_filter:
        today = date.today()
        if quick_filter == 'current_month':
            payrolls = payrolls.filter(
                pay_period_start__gte=today.replace(day=1),
                pay_period_end__lte=today.replace(day=monthrange(today.year, today.month)[1]),
            )
        elif quick_filter == 'last_month':
            first_of_current = today.replace(day=1)
            last_of_previous = first_of_current - timedelta(days=1)
            payrolls = payrolls.filter(
                pay_period_start__gte=last_of_previous.replace(day=1),
                pay_period_end__lte=last_of_previous,
            )
        elif quick_filter == 'current_year':
            payrolls = payrolls.filter(
                pay_period_start__gte=today.replace(month=1, day=1),
                pay_period_end__lte=today.replace(month=12, day=31),
            )
        elif quick_filter == 'last_year':
            ly = today.year - 1
            payrolls = payrolls.filter(
                pay_period_start__gte=date(ly, 1, 1),
                pay_period_end__lte=date(ly, 12, 31),
            )

    payment_date_from = request.GET.get('payment_date_from', '')
    if payment_date_from:
        payrolls = payrolls.filter(payment_date__gte=payment_date_from)

    payment_date_to = request.GET.get('payment_date_to', '')
    if payment_date_to:
        payrolls = payrolls.filter(payment_date__lte=payment_date_to)

    pay_period_from = request.GET.get('pay_period_from', '')
    if pay_period_from:
        payrolls = payrolls.filter(pay_period_end__gte=pay_period_from)

    pay_period_to = request.GET.get('pay_period_to', '')
    if pay_period_to:
        payrolls = payrolls.filter(pay_period_start__lte=pay_period_to)

    return payrolls


@login_required
def payroll_list(request):
    filter_form = PayrollFilterForm(request.GET or None)
    payrolls    = get_filtered_payrolls(request)
    active      = payrolls.filter(reversed=False)

    stats = {
        'total':            payrolls.count(),
        'active':           active.count(),
        'reversed':         payrolls.filter(reversed=True).count(),
        'draft':            active.filter(status='DRAFT').count(),
        'approved':         active.filter(status='APPROVED').count(),
        'processing':       active.filter(status='PROCESSING').count(),
        'paid':             active.filter(status='PAID').count(),
        'cancelled':        active.filter(status='CANCELLED').count(),
        'total_gross_pay':  active.filter(status='PAID').aggregate(
                                Sum('gross_pay'))['gross_pay__sum'] or 0,
        'total_net_pay':    active.filter(status='PAID').aggregate(
                                Sum('net_pay'))['net_pay__sum'] or 0,
        'total_deductions': active.filter(status='PAID').aggregate(
                                Sum('total_deductions'))['total_deductions__sum'] or 0,
        'avg_net_pay':      active.filter(status='PAID').aggregate(
                                Avg('net_pay'))['net_pay__avg'] or 0,
        'prorated_count':   active.filter(is_prorated=True).count(),
        'monthly_count':    active.filter(pay_frequency='MONTHLY').count(),
        'weekly_count':     active.filter(pay_frequency='WEEKLY').count(),
    }

    paginator     = Paginator(payrolls, 20)
    payrolls_page = paginator.get_page(request.GET.get('page', 1))
    is_htmx       = request.headers.get('HX-Request') == 'true'

    context = {
        'payrolls_page': payrolls_page,
        'paginator':     paginator,
        'stats':         stats,
        'filter_form':   filter_form,
        'is_htmx':       is_htmx,
    }
    if is_htmx:
        return render(request, 'hr/payroll/partials/_payroll_results.html', context)
    return render(request, 'hr/payroll/list.html', context)


@login_required
def payroll_detail(request, pk):
    payroll = get_object_or_404(
        Payroll.objects.select_related(
            'staff__primary_department', 'fiscal_period', 'payment_method',
            'journal_entry', 'payment_journal_entry', 'reversal_journal_entry',
        ).prefetch_related(
            'allowances', 'deductions', 'bonuses',
            'salary_payments__payment_method',
            'salary_payments__journal_entry',
        ),
        pk=pk,
    )
    can_reverse, reversal_message = payroll.can_be_reversed()

    return render(request, 'hr/payroll/detail.html', {
        'payroll':                    payroll,
        'total_allowances':           payroll.total_allowances,
        'total_bonuses':              payroll.total_bonuses,
        'gross_pay':                  payroll.gross_pay,
        'taxable_income':             payroll.taxable_income,
        'paye_amount':                payroll.paye_amount,
        'nssf_employee':              payroll.nssf_employee,
        'local_service_tax':          payroll.local_service_tax,
        'total_statutory_deductions': payroll.total_statutory_deductions,
        'total_voluntary_deductions': payroll.total_voluntary_deductions,
        'total_deductions':           payroll.total_deductions,
        'net_pay':                    payroll.net_pay,
        'nssf_employer':              payroll.nssf_employer,
        'employer_total_cost':        payroll.employer_total_cost,
        'effective_net_pay':          payroll.effective_net_pay,
        'effective_gross_pay':        payroll.effective_gross_pay,
        'effective_employer_cost':    payroll.effective_employer_cost,
        # Payment tracking
        'salary_payments':            payroll.salary_payments.all(),
        'total_paid':                 payroll.total_paid,
        'balance_due':                payroll.balance_due,
        'payment_completion_status':  payroll.payment_completion_status,
        'payment_completion_display': payroll.payment_completion_display,
        'is_fully_paid':              payroll.is_fully_paid,
        # Auth
        'can_reverse':                can_reverse,
        'reversal_message':           reversal_message,
        'requires_statutory':         payroll.requires_statutory_adjustments(),
        'approved_by':                payroll.get_approved_by_user(),
        'paid_by':                    payroll.get_paid_by_user(),
        'reversed_by':                payroll.get_reversed_by_user(),
        'reversal_approved_by':       payroll.get_reversal_approved_by_user(),
    })

@login_required
def payroll_staff_defaults(request):
    """
    AJAX: return active contract defaults for a given staff member.
    Called by the payroll form when the user selects a staff member
    from the Select2 dropdown.

    GET /hr/payroll/staff-defaults/?staff_id=<uuid>

    Response JSON:
    {
        "has_contract":    true | false,
        "basic_salary":    "1500000.00" | "",
        "contract_number": "CONT/2025/FT/0001" | "",
        "contract_type":   "Fixed Term Contract" | "",
        "salary_frequency": "MONTHLY" | "",
        "bank_account":    "1234567890" | "",
    }
    """
    staff_pk = request.GET.get('staff_id', '').strip()
    if not staff_pk:
        return JsonResponse({'error': 'staff_id is required'}, status=400)

    try:
        staff = Staff.objects.get(pk=staff_pk, is_active=True)
    except (Staff.DoesNotExist, ValueError):
        return JsonResponse({'error': 'Staff not found'}, status=404)

    contract = Contract.get_staff_active_contract(staff)

    return JsonResponse({
        'has_contract':     contract is not None,
        'basic_salary':     str(contract.basic_salary) if contract else '',
        'contract_number':  contract.contract_number if contract else '',
        'contract_type':    contract.get_contract_type_display() if contract else '',
        'salary_frequency': contract.get_salary_frequency_display() if contract else '',
        'bank_account':     staff.bank_account_number or '',
    })

@login_required
@transaction.atomic
def payroll_create(request):
    from core.utils import get_school_today
    from calendar import monthrange

    today    = get_school_today()
    last_day = monthrange(today.year, today.month)[1]

    date_defaults = {
        'pay_period_start': today.replace(day=1),
        'pay_period_end':   today.replace(day=last_day),
        'payment_date':     today,
    }

    if request.method == 'POST':
        form              = PayrollForm(request.POST)
        allowance_formset = PayrollAllowanceFormSet(request.POST, prefix='allowances')
        deduction_formset = PayrollDeductionFormSet(request.POST, prefix='deductions')
        bonus_formset     = PayrollBonusFormSet(request.POST, prefix='bonuses')

        if all([form.is_valid(), allowance_formset.is_valid(),
                deduction_formset.is_valid(), bonus_formset.is_valid()]):
            payroll = form.save()

            allowance_formset.instance = payroll
            deduction_formset.instance = payroll
            bonus_formset.instance     = payroll
            allowance_formset.save()
            deduction_formset.save()
            bonus_formset.save()

            # Recalculate all totals now that the payroll and its line items
            # are saved. The pre_save signal only catches *changes* to existing
            # instances, so a fresh create needs an explicit recalculation.
            payroll.recalculate_all()
            payroll.save(update_fields=[
                'basic_salary', 'total_allowances', 'total_bonuses', 'gross_pay',
                'taxable_income',
                'paye_amount', 'nssf_employee', 'local_service_tax',
                'total_statutory_deductions', 'total_voluntary_deductions',
                'total_deductions', 'net_pay',
                'nssf_employer', 'employer_total_cost',
                'updated_at',
            ])

            payroll.refresh_from_db()

            messages.success(
                request,
                f'Payroll created for {payroll.staff.full_name()} '
                f'({payroll.pay_period_label}) — '
                f'Gross: {payroll.gross_pay:,.0f}, Net: {payroll.net_pay:,.0f}',
                extra_tags='sweetalert',
            )
            return redirect('hr:payroll_detail', pk=payroll.pk)
        else:
            messages.error(
                request,
                'Please correct the errors below.',
                extra_tags='sweetalert-error',
            )
    else:
        form              = PayrollForm(initial=date_defaults)
        allowance_formset = PayrollAllowanceFormSet(prefix='allowances')
        deduction_formset = PayrollDeductionFormSet(prefix='deductions')
        bonus_formset     = PayrollBonusFormSet(prefix='bonuses')

    return render(request, 'hr/payroll/form.html', {
        'form':              form,
        'allowance_formset': allowance_formset,
        'deduction_formset': deduction_formset,
        'bonus_formset':     bonus_formset,
        'title':             'Create Payroll',
        'is_create':         True,
        'default_payment_date':     today.strftime('%Y-%m-%d'),
        'default_pay_period_start': today.replace(day=1).strftime('%Y-%m-%d'),
        'default_pay_period_end':   today.replace(day=last_day).strftime('%Y-%m-%d'),
    })


@login_required
@transaction.atomic
def payroll_edit(request, pk):
    payroll = get_object_or_404(Payroll, pk=pk)

    if payroll.reversed:
        messages.error(
            request,
            f'Cannot edit reversed payroll for {payroll.staff.full_name()} ({payroll.pay_period_label}).',
            extra_tags='sweetalert-error',
        )
        return redirect('hr:payroll_detail', pk=pk)

    # Use fiscal_period_id (raw FK) to avoid RelatedObjectDoesNotExist
    if payroll.fiscal_period_id and getattr(payroll.fiscal_period, 'is_closed', False):
        messages.error(
            request,
            f'Cannot edit payroll from closed fiscal period ({payroll.fiscal_period.name}).',
            extra_tags='sweetalert-error',
        )
        return redirect('hr:payroll_detail', pk=pk)

    if payroll.status == 'PAID':
        messages.warning(
            request,
            'Payroll is paid — only notes and payment reference can be updated.',
            extra_tags='sweetalert',
        )

    if request.method == 'POST':
        form              = PayrollForm(request.POST, instance=payroll)
        allowance_formset = PayrollAllowanceFormSet(request.POST, instance=payroll, prefix='allowances')
        deduction_formset = PayrollDeductionFormSet(request.POST, instance=payroll, prefix='deductions')
        bonus_formset     = PayrollBonusFormSet(request.POST, instance=payroll, prefix='bonuses')

        if all([form.is_valid(), allowance_formset.is_valid(),
                deduction_formset.is_valid(), bonus_formset.is_valid()]):
            payroll = form.save()
            allowance_formset.save()
            deduction_formset.save()
            bonus_formset.save()

            # Recalculate after all line items are saved so totals reflect
            # any changes to basic salary, allowances, deductions, or bonuses.
            payroll.recalculate_all()
            payroll.save(update_fields=[
                'basic_salary', 'total_allowances', 'total_bonuses', 'gross_pay',
                'taxable_income',
                'paye_amount', 'nssf_employee', 'local_service_tax',
                'total_statutory_deductions', 'total_voluntary_deductions',
                'total_deductions', 'net_pay',
                'nssf_employer', 'employer_total_cost',
                'updated_at',
            ])

            payroll.refresh_from_db()

            messages.success(
                request,
                f'Payroll updated for {payroll.staff.full_name()} '
                f'({payroll.pay_period_label}) — '
                f'Gross: {payroll.gross_pay:,.0f}, Net: {payroll.net_pay:,.0f}',
                extra_tags='sweetalert',
            )
            return redirect('hr:payroll_detail', pk=pk)
        else:
            messages.error(
                request,
                'Please correct the errors below.',
                extra_tags='sweetalert-error',
            )
    else:
        form              = PayrollForm(instance=payroll)
        allowance_formset = PayrollAllowanceFormSet(instance=payroll, prefix='allowances')
        deduction_formset = PayrollDeductionFormSet(instance=payroll, prefix='deductions')
        bonus_formset     = PayrollBonusFormSet(instance=payroll, prefix='bonuses')

    return render(request, 'hr/payroll/form.html', {
        'form':              form,
        'allowance_formset': allowance_formset,
        'deduction_formset': deduction_formset,
        'bonus_formset':     bonus_formset,
        'payroll':           payroll,
        'title':             f'Edit Payroll — {payroll.pay_period_label}',
        'is_create':         False,
        'is_editable':       payroll.status in ('DRAFT', 'APPROVED', 'PARTIAL'),
        'is_paid':           payroll.status == 'PAID',
        'is_reversed':       payroll.reversed,
    })


@login_required
def payroll_delete(request, pk):
    """POST-only. modal_views.payroll_delete_modal handles GET."""
    if request.method != 'POST':
        return HttpResponse(status=405)

    payroll = get_object_or_404(Payroll, pk=pk)

    if payroll.reversed:
        if request.headers.get('HX-Request') == 'true':
            r = HttpResponse()
            r['HX-Alert-Message'] = 'Cannot delete reversed payrolls — they must be kept for audit trail'
            r['HX-Alert-Type']    = 'error'
            r['HX-Close-Modal']   = 'true'
            return r
        messages.error(
            request,
            'Cannot delete reversed payrolls — they must be kept for audit trail',
            extra_tags='sweetalert-error',
        )
        return redirect('hr:payroll_list')

    if payroll.status not in ['DRAFT', 'CANCELLED']:
        if request.headers.get('HX-Request') == 'true':
            r = HttpResponse()
            r['HX-Alert-Message'] = (
                f'Cannot delete payroll with status: {payroll.get_status_display()}. '
                'Only DRAFT or CANCELLED payrolls can be deleted.'
            )
            r['HX-Alert-Type']  = 'error'
            r['HX-Close-Modal'] = 'true'
            return r
        messages.error(
            request,
            f'Cannot delete payroll with status: {payroll.get_status_display()}',
            extra_tags='sweetalert-error',
        )
        return redirect('hr:payroll_list')

    staff_name       = payroll.staff.full_name()
    pay_period_label = payroll.pay_period_label
    payroll.delete()

    if request.headers.get('HX-Request') == 'true':
        r = HttpResponse()
        r['HX-Alert-Message'] = f'Payroll for {staff_name} ({pay_period_label}) deleted'
        r['HX-Alert-Type']    = 'success'
        r['HX-Close-Modal']   = 'true'
        r['HX-Redirect']      = reverse('hr:payroll_list')
        return r
    messages.success(
        request,
        f'Payroll for {staff_name} ({pay_period_label}) deleted',
        extra_tags='sweetalert',
    )
    return redirect('hr:payroll_list')


@login_required
def payroll_approve(request, pk):
    """POST-only. modal_views.payroll_approve_modal handles GET."""
    if request.method != 'POST':
        return HttpResponse(status=405)

    payroll = get_object_or_404(Payroll, pk=pk)

    if payroll.reversed:
        if request.headers.get('HX-Request') == 'true':
            r = HttpResponse()
            r['HX-Alert-Message'] = 'Cannot approve a reversed payroll'
            r['HX-Alert-Type']    = 'error'
            r['HX-Close-Modal']   = 'true'
            return r
        messages.error(request, 'Cannot approve a reversed payroll', extra_tags='sweetalert-error')
        return redirect('hr:payroll_detail', pk=pk)

    if payroll.status != 'DRAFT':
        if request.headers.get('HX-Request') == 'true':
            r = HttpResponse()
            r['HX-Alert-Message'] = f'Payroll is already {payroll.get_status_display()}'
            r['HX-Alert-Type']    = 'warning'
            r['HX-Close-Modal']   = 'true'
            return r
        messages.warning(request, f'Payroll is already {payroll.get_status_display()}', extra_tags='sweetalert')
        return redirect('hr:payroll_detail', pk=pk)

    payroll.status         = 'APPROVED'
    payroll.approved_at    = timezone.now()
    payroll.approved_by_id = str(request.user.id)
    payroll.save(update_fields=['status', 'approved_at', 'approved_by_id', 'updated_at'])

    if request.headers.get('HX-Request') == 'true':
        r = HttpResponse()
        r['HX-Alert-Message'] = (
            f'Payroll for {payroll.staff.full_name()} ({payroll.pay_period_label}) approved'
        )
        r['HX-Alert-Type']  = 'success'
        r['HX-Close-Modal'] = 'true'
        r['HX-Redirect']    = reverse('hr:payroll_detail', kwargs={'pk': pk})
        return r
    messages.success(
        request,
        f'Payroll for {payroll.staff.full_name()} ({payroll.pay_period_label}) approved',
        extra_tags='sweetalert',
    )
    return redirect('hr:payroll_detail', pk=pk)


@login_required
def payroll_record_payment(request, pk):
    from .forms import PayrollPaymentForm

    payroll = get_object_or_404(
        Payroll.objects.select_related('staff', 'payment_method')
                       .prefetch_related('salary_payments'),
        pk=pk,
    )

    if request.method != 'POST':
        return HttpResponseNotAllowed(['POST'])

    # Guard — only APPROVED or PROCESSING payrolls can receive payments
    if payroll.reversed or payroll.status not in ['APPROVED', 'PROCESSING']:
        response = HttpResponse(status=204)
        response['HX-Alert-Type']    = 'error'
        response['HX-Alert-Message'] = (
            f'Cannot record payment for a {payroll.get_status_display()} payroll.'
        )
        return response

    form = PayrollPaymentForm(payroll=payroll, data=request.POST)

    if form.is_valid():
        payment = form.save(commit=False)
        payment.payroll      = payroll
        payment.recorded_by_id = str(request.user.id)
        payment.save()

        # Signal handles payroll status update (PROCESSING → PAID when fully paid)
        # Re-fetch to get updated status
        payroll.refresh_from_db()

        response = HttpResponse(status=204)
        response['HX-Close-Modal']   = 'true'
        response['HX-Alert-Type']    = 'success'
        response['HX-Alert-Title']   = 'Payment Recorded'
        response['HX-Alert-Message'] = (
            f'Payment of {payment.amount:,.0f} {payroll.currency} recorded '
            f'for {payroll.staff.full_name()}. '
            f'Ref: {payment.payment_reference}. '
            + (
                f'Payroll is now fully paid.'
                if payroll.status == 'PAID'
                else f'Balance remaining: {payroll.balance_due:,.0f} {payroll.currency}.'
            )
        )
        response['HX-Redirect'] = request.META.get(
            'HTTP_REFERER',
            f'/hr/payrolls/{payroll.pk}/'
        )
        return response

    # Validation failed — re-render modal with errors
    return render(request, 'hr/payroll/modals/record_payment.html', {
        'payroll': payroll,
        'form':    form,
    })

@login_required
def payroll_reverse(request, pk):
    """POST-only. modal_views.payroll_reverse_modal handles GET."""
    if request.method != 'POST':
        return HttpResponse(status=405)

    payroll = get_object_or_404(Payroll, pk=pk)
    can_reverse, reason = payroll.can_be_reversed()

    if not can_reverse:
        if request.headers.get('HX-Request') == 'true':
            r = HttpResponse()
            r['HX-Alert-Message'] = f'Cannot reverse payroll: {reason}'
            r['HX-Alert-Type']    = 'error'
            r['HX-Close-Modal']   = 'true'
            return r
        messages.error(request, f'Cannot reverse payroll: {reason}', extra_tags='sweetalert-error')
        return redirect('hr:payroll_detail', pk=pk)

    form = PayrollReversalForm(payroll, request.user, request.POST)
    if form.is_valid():
        payroll.reversed        = True
        payroll.reversed_on     = timezone.now()
        payroll.reversed_by_id  = str(request.user.id)
        payroll.reversal_reason = form.cleaned_data['reversal_reason']
        payroll.status          = 'REVERSED'
        if form.cleaned_data.get('statutory_adjustments_notes'):
            payroll.statutory_adjustments_notes = form.cleaned_data['statutory_adjustments_notes']
        payroll.save()

        if request.headers.get('HX-Request') == 'true':
            r = HttpResponse()
            r['HX-Alert-Message'] = (
                f'Payroll for {payroll.staff.full_name()} ({payroll.pay_period_label}) reversed'
            )
            r['HX-Alert-Type']  = 'warning'
            r['HX-Close-Modal'] = 'true'
            r['HX-Redirect']    = reverse('hr:payroll_detail', kwargs={'pk': pk})
            return r
        messages.warning(
            request,
            f'Payroll for {payroll.staff.full_name()} ({payroll.pay_period_label}) reversed',
            extra_tags='sweetalert',
        )
    else:
        messages.error(request, 'Please correct the errors in the form', extra_tags='sweetalert-error')

    return redirect('hr:payroll_detail', pk=pk)


@login_required
def payroll_recalculate(request, pk):
    """Force-recalculate all payroll summary fields. Useful after manual corrections."""
    payroll = get_object_or_404(Payroll, pk=pk)

    if payroll.reversed:
        messages.error(request, 'Cannot recalculate a reversed payroll.', extra_tags='sweetalert-error')
        return redirect('hr:payroll_detail', pk=pk)

    old_gross = payroll.gross_pay
    old_net   = payroll.net_pay

    payroll.recalculate_all()
    payroll.save(update_fields=[
        'total_allowances', 'total_bonuses', 'gross_pay', 'taxable_income',
        'paye_amount', 'nssf_employee', 'local_service_tax',
        'total_statutory_deductions', 'total_voluntary_deductions',
        'total_deductions', 'net_pay', 'employer_total_cost', 'updated_at',
    ])

    if old_gross != payroll.gross_pay or old_net != payroll.net_pay:
        messages.success(
            request,
            f'Recalculated. Gross: {old_gross:,.0f} → {payroll.gross_pay:,.0f}, '
            f'Net: {old_net:,.0f} → {payroll.net_pay:,.0f}',
            extra_tags='sweetalert',
        )
    else:
        messages.info(request, 'Recalculated — no changes detected.', extra_tags='sweetalert')

    return redirect('hr:payroll_detail', pk=pk)


# =============================================================================
# SALARY HISTORY
# =============================================================================

@login_required
def salary_history_list(request):
    return render(request, 'hr/salary_history/list.html', {})


@login_required
def salary_history_create(request):
    from .forms import SalaryHistoryForm
    if request.method == 'POST':
        form = SalaryHistoryForm(request.POST, request.FILES)
        if form.is_valid():
            record = form.save()
            messages.success(
                request,
                f'Salary change recorded for {record.staff.full_name()}',
                extra_tags='sweetalert',
            )
            return redirect('hr:salary_history_list')
        else:
            messages.error(request, 'Please correct the errors in the form', extra_tags='sweetalert-error')
    else:
        form = SalaryHistoryForm()
    return render(request, 'hr/salary_history/form.html', {
        'form': form, 'title': 'Record Salary Change',
    })


# =============================================================================
# BULK STAFF OPERATIONS
# =============================================================================

@login_required
@transaction.atomic
def bulk_staff_action(request):
    """POST-only. modal_views.bulk_staff_action_modal handles GET."""
    if request.method != 'POST':
        return HttpResponse(status=405)

    action       = request.POST.get('action')
    selected_ids = [i for i in request.POST.get('selected_ids', '').split(',') if i]

    if not action or not selected_ids:
        messages.error(request, 'No action or staff members selected', extra_tags='sweetalert-error')
        return redirect('hr:staff_list')

    staff_members = Staff.objects.filter(id__in=selected_ids)
    count         = staff_members.count()

    if action == 'activate':
        staff_members.update(is_active=True)
        messages.success(request, f'{count} staff member(s) activated', extra_tags='sweetalert')

    elif action == 'deactivate':
        with_active_contracts = staff_members.filter(contracts__status='ACTIVE').distinct()
        if with_active_contracts.exists():
            messages.error(
                request,
                f'{with_active_contracts.count()} staff have active contracts and cannot be deactivated',
                extra_tags='sweetalert-error',
            )
            return redirect('hr:staff_list')
        staff_members.update(is_active=False)
        messages.warning(request, f'{count} staff member(s) deactivated', extra_tags='sweetalert')

    else:
        messages.error(request, f'Unknown action: {action}', extra_tags='sweetalert-error')

    return redirect('hr:staff_list')


# =============================================================================
# REPORTS
# =============================================================================

@login_required
def hr_reports(request):
    try:
        dashboard_stats = hr_stats.get_hr_dashboard_statistics()
    except Exception as e:
        logger.error(f'Error getting HR dashboard statistics: {e}')
        dashboard_stats = {}
    return render(request, 'hr/reports/index.html', {'dashboard_stats': dashboard_stats})

@login_required
def staff_report(request):
    return render(request, 'hr/reports/staff_report.html', {})

@login_required
def contract_report(request):
    return render(request, 'hr/reports/contract_report.html', {})

@login_required
def teacher_report(request):
    return render(request, 'hr/reports/teacher_report.html', {})