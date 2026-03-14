# students/views.py

"""
Student Management Views

Comprehensive view functions for:
- Student Registration and Profile Management (using Wizard)
- Guardian Management
- Student-Guardian Relationships
- Sibling Relationships
- Enrollment Status Tracking
- Reports and Analytics

All views delegate business logic to services.py where appropriate
Uses stats.py for comprehensive statistics and analytics
Uses SweetAlert2 for all notifications via Django messages
Preserves SessionWizardView for student registration
"""

from django.shortcuts import render, redirect, get_object_or_404
from django.urls import reverse
from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator
from django.contrib import messages
from django.db.models import Q, Count, Sum, Avg, Prefetch, F, Case, When, IntegerField
from django.utils import timezone
from django.http import JsonResponse, HttpResponse
from django.db import transaction
from django.core.files.storage import FileSystemStorage
from formtools.wizard.views import SessionWizardView
from datetime import timedelta, date, datetime
from decimal import Decimal
import os
import logging

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from io import BytesIO

from core.utils import get_school_today, format_money

from .models import (
    Student,
    Guardian,
    StudentGuardian,
    SiblingRelationship,
    EnrollmentStatusHistory,
)

from .forms import (
    STUDENT_WIZARD_FORMS,
    STUDENT_WIZARD_STEP_NAMES,
    StudentForm,
    GuardianForm,
    StudentGuardianForm,
    StudentFilterForm,
    GuardianFilterForm,
)

# Import stats functions
from . import stats as student_stats

logger = logging.getLogger(__name__)


# =============================================================================
# DASHBOARD
# =============================================================================

@login_required
def students_dashboard(request):
    """Main students dashboard with overview statistics - USES stats.py"""
    
    try:
        # Get comprehensive statistics using SCHOOL timezone
        today = get_school_today()
        thirty_days_ago = today - timedelta(days=30)
        
        student_statistics = student_stats.get_student_statistics()
        guardian_statistics = student_stats.get_guardian_statistics()
        sibling_statistics = student_stats.get_sibling_statistics()
        family_statistics = student_stats.get_family_statistics()
        
    except Exception as e:
        logger.error(f"Error getting dashboard statistics: {e}")
        student_statistics = {}
        guardian_statistics = {}
        sibling_statistics = {}
        family_statistics = {}
    
    # Get recent activities (limited queries for display)
    recent_students = Student.objects.select_related(
        'current_academic_level'
    ).order_by('-created_at')[:10]
    
    pending_approval = Student.objects.filter(
        enrollment_status='PENDING_APPROVAL'
    ).order_by('admission_date')[:10]
    
    # Get students needing attention
    students_without_guardians = Student.objects.filter(
        enrollment_status='ACTIVE'
    ).annotate(
        guardian_count=Count('guardians')
    ).filter(guardian_count=0).order_by('admission_date')[:10]
    
    medical_alerts = Student.objects.filter(
        enrollment_status='ACTIVE'
    ).filter(
        Q(medical_conditions__isnull=False) & ~Q(medical_conditions='') |
        Q(allergies__isnull=False) & ~Q(allergies='') |
        Q(has_special_needs=True)
    ).order_by('-updated_at')[:10]
    
    # Get birthdays this week using SCHOOL timezone
    today = get_school_today()
    week_from_now = today + timedelta(days=7)
    
    upcoming_birthdays = Student.objects.filter(
        enrollment_status='ACTIVE',
        date_of_birth__month=today.month,
        date_of_birth__day__gte=today.day,
        date_of_birth__day__lte=week_from_now.day
    ).order_by('date_of_birth__day')[:10]
    
    # Recent status changes
    recent_status_changes = EnrollmentStatusHistory.objects.select_related(
        'student', 'academic_session'
    ).order_by('-effective_date')[:10]
    
    context = {
        'student_statistics': student_statistics,
        'guardian_statistics': guardian_statistics,
        'sibling_statistics': sibling_statistics,
        'family_statistics': family_statistics,
        'recent_students': recent_students,
        'pending_approval': pending_approval,
        'students_without_guardians': students_without_guardians,
        'medical_alerts': medical_alerts,
        'upcoming_birthdays': upcoming_birthdays,
        'recent_status_changes': recent_status_changes,
    }
    
    return render(request, 'students/dashboard.html', context)


# =============================================================================
# HELPER FUNCTIONS FOR FILTERING
# =============================================================================

def get_filtered_students(request):
    """
    Helper function to get filtered student queryset.
    Reusable across student_list, exports, and print views.

    Ordering: active students always first, then by most recent admission date,
    then by admission number for stable tie-breaking.
    """
    students = Student.objects.select_related(
        'current_academic_level',
        'admission_academic_level'
    ).prefetch_related(
        'guardians',
        'guardian_relationships'
    ).annotate(
        guardian_count=Count('guardians', distinct=True),
        sibling_count=Count('sibling_relationships', distinct=True),
        # 0 = ACTIVE floats to top; 1 = everything else follows
        status_order=Case(
            When(enrollment_status='ACTIVE', then=0),
            default=1,
            output_field=IntegerField()
        )
    ).order_by('status_order', '-admission_date', 'admission_number')

    # -------------------------------------------------------------------------
    # FILTER PARAMETERS
    # -------------------------------------------------------------------------

    query                  = request.GET.get('q', '').strip()
    enrollment_status      = request.GET.get('enrollment_status', '')
    gender                 = request.GET.get('gender', '')
    current_academic_level = request.GET.get('current_academic_level', '')
    admission_academic_level = request.GET.get('admission_academic_level', '')
    nationality            = request.GET.get('nationality', '')
    religious_affiliation  = request.GET.get('religious_affiliation', '')
    health_condition       = request.GET.get('health_condition', '')
    blood_type             = request.GET.get('blood_type', '')
    has_special_needs      = request.GET.get('has_special_needs', '')
    transportation_required = request.GET.get('transportation_required', '')
    admission_date_from    = request.GET.get('admission_date_from', '')
    admission_date_to      = request.GET.get('admission_date_to', '')
    min_age                = request.GET.get('age_min', '')
    max_age                = request.GET.get('age_max', '')

    # -------------------------------------------------------------------------
    # TEXT SEARCH — multi-word AND logic across key fields
    # -------------------------------------------------------------------------

    if query:
        words = query.split()
        if words:
            combined_q = Q()
            for word in words:
                combined_q &= (
                    Q(admission_number__icontains=word) |
                    Q(national_student_number__icontains=word) |
                    Q(first_name__icontains=word) |
                    Q(middle_name__icontains=word) |
                    Q(last_name__icontains=word) |
                    Q(phone_number__icontains=word) |
                    Q(personal_email__icontains=word) |
                    Q(birth_certificate_number__icontains=word)
                )
            students = students.filter(combined_q)

    # -------------------------------------------------------------------------
    # CHOICE FILTERS
    # -------------------------------------------------------------------------

    if enrollment_status:
        students = students.filter(enrollment_status=enrollment_status)
    if gender:
        students = students.filter(gender=gender)
    if current_academic_level:
        students = students.filter(current_academic_level_id=current_academic_level)
    if admission_academic_level:
        students = students.filter(admission_academic_level_id=admission_academic_level)
    if nationality:
        students = students.filter(nationality=nationality)
    if religious_affiliation:
        students = students.filter(religious_affiliation=religious_affiliation)
    if health_condition:
        students = students.filter(health_condition=health_condition)
    if blood_type:
        students = students.filter(blood_type=blood_type)

    # -------------------------------------------------------------------------
    # BOOLEAN FILTERS
    # -------------------------------------------------------------------------

    if has_special_needs:
        students = students.filter(
            has_special_needs=(has_special_needs.lower() == 'true')
        )
    if transportation_required:
        students = students.filter(
            transportation_required=(transportation_required.lower() == 'true')
        )

    # -------------------------------------------------------------------------
    # DATE RANGE FILTERS
    # -------------------------------------------------------------------------

    if admission_date_from:
        students = students.filter(admission_date__gte=admission_date_from)
    if admission_date_to:
        students = students.filter(admission_date__lte=admission_date_to)

    # -------------------------------------------------------------------------
    # AGE RANGE FILTERS — computed against school timezone today
    # -------------------------------------------------------------------------

    if min_age or max_age:
        today = get_school_today()

        if max_age:
            try:
                # Youngest allowed birth date for this max age
                min_birth_date = date(
                    today.year - int(max_age) - 1,
                    today.month,
                    today.day
                )
                students = students.filter(date_of_birth__gte=min_birth_date)
            except (ValueError, TypeError):
                pass

        if min_age:
            try:
                # Oldest allowed birth date for this min age
                max_birth_date = date(
                    today.year - int(min_age),
                    today.month,
                    today.day
                )
                students = students.filter(date_of_birth__lte=max_birth_date)
            except (ValueError, TypeError):
                pass

    return students


def get_filtered_guardians(request):
    """Helper function to get filtered guardian queryset"""
    guardians = Guardian.objects.annotate(
        student_count=Count('students', distinct=True),
        primary_student_count=Count(
            'student_relationships',
            filter=Q(student_relationships__is_primary=True),
            distinct=True
        ),
        financial_responsibility_count=Count(
            'student_relationships',
            filter=Q(student_relationships__is_financial_responsible=True),
            distinct=True
        )
    ).order_by('last_name', 'first_name')
    
    query = request.GET.get('q', '').strip()
    guardian_type = request.GET.get('guardian_type', '')
    gender = request.GET.get('gender', '')
    is_active = request.GET.get('is_active', '')
    country = request.GET.get('country', '')
    occupation = request.GET.get('occupation', '')
    
    # Apply text search
    if query:
        words = query.strip().split()
        if words:
            combined_q = Q()
            for word in words:
                word_q = (
                    Q(first_name__icontains=word) |
                    Q(middle_name__icontains=word) |
                    Q(last_name__icontains=word) |
                    Q(primary_phone__icontains=word) |
                    Q(secondary_phone__icontains=word) |
                    Q(email__icontains=word) |
                    Q(national_id__icontains=word) |
                    Q(employer__icontains=word)
                )
                combined_q &= word_q
            guardians = guardians.filter(combined_q)
    
    # Apply filters
    if guardian_type:
        guardians = guardians.filter(guardian_type=guardian_type)
    if gender:
        guardians = guardians.filter(gender=gender)
    if occupation:
        guardians = guardians.filter(occupation__icontains=occupation)
    if country:
        guardians = guardians.filter(country=country)
    if is_active is not None and is_active:
        guardians = guardians.filter(is_active=(is_active.lower() == 'true'))
    
    return guardians


def get_filtered_student_guardians(request):
    """Helper function to get filtered student-guardian relationship queryset"""
    relationships = StudentGuardian.objects.select_related(
        'student__current_academic_level',
        'guardian'
    ).order_by('student__admission_number', 'emergency_contact_priority')
    
    query = request.GET.get('q', '').strip()
    student = request.GET.get('student', '')
    guardian = request.GET.get('guardian', '')
    relationship = request.GET.get('relationship', '')
    is_primary = request.GET.get('is_primary', '')
    is_financial_responsible = request.GET.get('is_financial_responsible', '')
    is_active = request.GET.get('is_active', '')
    
    # Apply text search
    if query:
        words = query.strip().split()
        if words:
            combined_q = Q()
            for word in words:
                word_q = (
                    Q(student__first_name__icontains=word) |
                    Q(student__last_name__icontains=word) |
                    Q(student__admission_number__icontains=word) |
                    Q(guardian__first_name__icontains=word) |
                    Q(guardian__last_name__icontains=word) |
                    Q(guardian__primary_phone__icontains=word)
                )
                combined_q &= word_q
            relationships = relationships.filter(combined_q)
    
    # Apply filters
    if student:
        relationships = relationships.filter(student_id=student)
    if guardian:
        relationships = relationships.filter(guardian_id=guardian)
    if relationship:
        relationships = relationships.filter(relationship=relationship)
    if is_primary:
        relationships = relationships.filter(is_primary=(is_primary.lower() == 'true'))
    if is_financial_responsible:
        relationships = relationships.filter(
            is_financial_responsible=(is_financial_responsible.lower() == 'true')
        )
    if is_active:
        relationships = relationships.filter(is_active=(is_active.lower() == 'true'))
    
    return relationships


def get_filtered_siblings(request):
    """Helper function to get filtered sibling relationship queryset"""
    siblings = SiblingRelationship.objects.select_related(
        'from_student__current_academic_level',
        'to_student__current_academic_level'
    ).order_by('from_student__admission_number')
    
    query = request.GET.get('q', '').strip()
    from_student = request.GET.get('from_student', '')
    to_student = request.GET.get('to_student', '')
    relationship_type = request.GET.get('relationship_type', '')
    is_verified = request.GET.get('is_verified', '')
    
    # Apply text search
    if query:
        words = query.strip().split()
        if words:
            combined_q = Q()
            for word in words:
                word_q = (
                    Q(from_student__first_name__icontains=word) |
                    Q(from_student__last_name__icontains=word) |
                    Q(from_student__admission_number__icontains=word) |
                    Q(to_student__first_name__icontains=word) |
                    Q(to_student__last_name__icontains=word) |
                    Q(to_student__admission_number__icontains=word)
                )
                combined_q &= word_q
            siblings = siblings.filter(combined_q)
    
    # Apply filters
    if from_student:
        siblings = siblings.filter(from_student_id=from_student)
    if to_student:
        siblings = siblings.filter(to_student_id=to_student)
    if relationship_type:
        siblings = siblings.filter(relationship_type=relationship_type)
    if is_verified:
        siblings = siblings.filter(is_verified=(is_verified.lower() == 'true'))
    
    return siblings


# =============================================================================
# STUDENT VIEWS
# =============================================================================

@login_required
def student_list(request):
    """
    Handle BOTH full page loads AND HTMX search/filter requests
    This is the ONLY view needed for student listing and searching
    """
    filter_form = StudentFilterForm(request.GET or None)
    students = get_filtered_students(request)
    
    # Calculate statistics (always in sync with current filters!)
    stats = students.aggregate(
        total=Count('id'),
        active=Count('id', filter=Q(enrollment_status='ACTIVE')),
        suspended=Count('id', filter=Q(enrollment_status='SUSPENDED')),
        graduated=Count('id', filter=Q(enrollment_status='GRADUATED')),
        transferred=Count('id', filter=Q(enrollment_status='TRANSFERRED')),
        withdrawn=Count('id', filter=Q(enrollment_status='WITHDRAWN')),
        male=Count('id', filter=Q(gender='M')),
        female=Count('id', filter=Q(gender='F')),
        special_needs=Count('id', filter=Q(has_special_needs=True)),
        transportation=Count('id', filter=Q(transportation_required=True)),
        medical_alerts=Count('id', filter=(
            Q(medical_conditions__isnull=False) & ~Q(medical_conditions='') |
            Q(allergies__isnull=False) & ~Q(allergies='') |
            Q(medications__isnull=False) & ~Q(medications='')
        )),
    )
    
    # Calculate average age
    today = get_school_today()
    students_with_dob = students.exclude(date_of_birth__isnull=True)
    if students_with_dob.exists():
        ages = [(today - s.date_of_birth).days // 365 for s in students_with_dob]
        stats['avg_age'] = round(sum(ages) / len(ages), 1) if ages else 0
    else:
        stats['avg_age'] = 0
    
    # Pagination
    paginator = Paginator(students, 20)
    page_number = request.GET.get('page', 1)
    students_page = paginator.get_page(page_number)
    
    # Detect HTMX request
    is_htmx = request.headers.get('HX-Request') == 'true'
    
    context = {
        'students_page': students_page,
        'paginator': paginator,
        'stats': stats,
        'filter_form': filter_form,
        'is_htmx': is_htmx,
    }
    
    # Return appropriate template
    if is_htmx:
        return render(request, 'students/partials/_student_results.html', context)
    else:
        return render(request, 'students/list.html', context)

@login_required
def student_print_view(request):
    """Generate printable student list with selected fields"""
    
    # Get selected fields from the modal
    selected_fields = request.GET.getlist('fields')
    if not selected_fields:
        # Default fields if none selected
        selected_fields = [
            'admission_number', 'full_name', 'date_of_birth', 'gender',
            'current_academic_level', 'enrollment_status', 'phone_number'
        ]
    
    # Get additional options
    include_stats = request.GET.get('include_stats') == 'true'
    landscape_mode = request.GET.get('landscape') == 'true'
    
    # Use the same helper function to get filtered students
    students = get_filtered_students(request)
    
    # Calculate stats only if requested
    stats = None
    if include_stats:
        stats = students.aggregate(
            total=Count('id'),
            active=Count('id', filter=Q(enrollment_status='ACTIVE')),
            male=Count('id', filter=Q(gender='M')),
            female=Count('id', filter=Q(gender='F')),
            special_needs=Count('id', filter=Q(has_special_needs=True)),
        )
    
    # Field display names mapping
    field_names = {
        'admission_number': 'Admission Number',
        'full_name': 'Full Name',
        'first_name': 'First Name',
        'last_name': 'Last Name',
        'national_student_number': 'National Student Number',
        'date_of_birth': 'Date of Birth',
        'age': 'Age',
        'gender': 'Gender',
        'nationality': 'Nationality',
        'phone_number': 'Phone',
        'personal_email': 'Email',
        'home_address': 'Home Address',
        'current_academic_level': 'Current Grade/Class',
        'admission_academic_level': 'Admission Grade/Class',
        'enrollment_status': 'Status',
        'admission_date': 'Admission Date',
        'health_condition': 'Health',
        'has_special_needs': 'Special Needs',
        'transportation_required': 'Transport',
        'religious_affiliation': 'Religion',
    }
    
    # Create ordered list of field display names for template
    selected_field_names = [
        field_names.get(field, field.replace('_', ' ').title()) 
        for field in selected_fields
    ]
    
    # Get school context explicitly for print view
    school = None
    if hasattr(request.user, 'profile') and hasattr(request.user.profile, 'school'):
        school = request.user.profile.school
    
    context = {
        'students': students,
        'stats': stats,
        'now': timezone.now(),
        'selected_fields': selected_fields,
        'selected_field_names': selected_field_names,
        'field_names': field_names,
        'landscape': landscape_mode,
        
        # Add school context explicitly
        'school_name': school.name if school else 'School',
        'school_logo': school.logo.url if (school and school.logo) else None,
        'school_address': school.address if school else '',
        'school_contact': school.phone if school else '',
    }
    
    return render(request, 'students/print.html', context)


# =============================================================================
# STUDENT WIZARD FOR CREATION
# =============================================================================

class StudentWizardFileStorage(FileSystemStorage):
    """Custom storage for handling file uploads in wizard"""
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.location = os.path.join(self.location, 'wizard_temp')


class StudentCreateWizard(SessionWizardView):
    """
    Multi-step wizard for creating a student.
    
    Steps:
    1. Basic Information - personal details and identification
    2. Contact Information - address and contact details
    3. Academic Information - academic level and previous education
    4. Health Information - medical and health details
    5. Guardian Information - primary guardian (optional)
    6. Confirmation - review and confirm
    
    Note: Admission number is automatically generated by pre_save signal in signals.py
    """

    form_list = STUDENT_WIZARD_FORMS
    template_name = 'students/wizard.html'
    file_storage = StudentWizardFileStorage()

    def get_template_names(self):
        """Return the template for all steps"""
        return [self.template_name]

    def get_context_data(self, form, **kwargs):
        """Add step names and progress tracking"""
        context = super().get_context_data(form=form, **kwargs)

        total_steps = len(self.form_list)
        current_step_index = list(self.form_list).index(self.steps.current)

        context.update({
            'step_names': STUDENT_WIZARD_STEP_NAMES,
            'current_step_name': STUDENT_WIZARD_STEP_NAMES.get(
                self.steps.current, 'Step'
            ),
            'progress_percentage': ((current_step_index) / (total_steps - 1)) * 100 if total_steps > 1 else 100,
        })

        # Add review data for confirmation step
        if self.steps.current == 'confirmation':
            context['basic_data'] = self.get_cleaned_data_for_step('basic_info')
            context['contact_data'] = self.get_cleaned_data_for_step('contact_info')
            context['academic_data'] = self.get_cleaned_data_for_step('academic_info')
            context['health_data'] = self.get_cleaned_data_for_step('health_info')
            context['guardian_data'] = self.get_cleaned_data_for_step('guardian_info')

        return context

    @transaction.atomic
    def done(self, form_list, **kwargs):
        """
        Persist all wizard data and create student.
        Admission number generation is handled automatically by the pre_save signal
        in signals.py (generate_admission_number function).
        """
        
        logger.info("=" * 80)
        logger.info("WIZARD DONE - Creating Student")
        logger.info("=" * 80)

        try:
            # Merge cleaned data from all steps
            form_data = {}
            
            for step, form in zip(self.form_list.keys(), form_list):
                form_data.update(form.cleaned_data)

            # ------------------------------------------------------------------
            # Create Student
            # ------------------------------------------------------------------
            student = Student(
                # Basic info
                first_name=form_data.get('first_name'),
                middle_name=form_data.get('middle_name', ''),
                last_name=form_data.get('last_name'),
                date_of_birth=form_data.get('date_of_birth'),
                gender=form_data.get('gender'),
                admission_date=form_data.get('admission_date'),
                national_student_number=form_data.get('national_student_number', ''),
                birth_certificate_number=form_data.get('birth_certificate_number', ''),
                nationality=form_data.get('nationality', ''),
                ethnicity=form_data.get('ethnicity', ''),
                birth_place=form_data.get('birth_place', ''),
                birth_country=form_data.get('birth_country', ''),
                religious_affiliation=form_data.get('religious_affiliation', ''),
                
                # Contact info
                personal_email=form_data.get('personal_email', ''),
                phone_number=form_data.get('phone_number', ''),
                home_address=form_data.get('home_address'),
                mailing_address=form_data.get('mailing_address', ''),
                district=form_data.get('district', ''),
                region=form_data.get('region', ''),
                country_of_residence=form_data.get('country_of_residence', ''),
                transportation_required=form_data.get('transportation_required', False),
                transport_route=form_data.get('transport_route', ''),
                pickup_point=form_data.get('pickup_point', ''),
                pickup_time=form_data.get('pickup_time'),
                
                # Academic info
                current_academic_level=form_data.get('current_academic_level'),
                admission_academic_level=form_data.get('admission_academic_level'),
                enrollment_status=form_data.get('enrollment_status', 'ACTIVE'),
                previous_school=form_data.get('previous_school', ''),
                previous_school_address=form_data.get('previous_school_address', ''),
                previous_academic_level=form_data.get('previous_academic_level'),
                transfer_reason=form_data.get('transfer_reason', ''),
                transfer_certificate_number=form_data.get('transfer_certificate_number', ''),
                previous_school_completion_date=form_data.get('previous_school_completion_date'),
                
                # Health info
                health_condition=form_data.get('health_condition', 'GOOD'),
                blood_type=form_data.get('blood_type', 'UNKNOWN'),
                medical_conditions=form_data.get('medical_conditions', ''),
                allergies=form_data.get('allergies', ''),
                medications=form_data.get('medications', ''),
                special_medical_needs=form_data.get('special_medical_needs', ''),
                emergency_medical_contact=form_data.get('emergency_medical_contact', ''),
                preferred_hospital=form_data.get('preferred_hospital', ''),
                medical_insurance=form_data.get('medical_insurance', ''),
                insurance_policy_number=form_data.get('insurance_policy_number', ''),
                has_special_needs=form_data.get('has_special_needs', False),
                special_needs_description=form_data.get('special_needs_description', ''),
                learning_disabilities=form_data.get('learning_disabilities', ''),
                learning_accommodations=form_data.get('learning_accommodations', ''),
                requires_special_diet=form_data.get('requires_special_diet', False),
                special_diet_details=form_data.get('special_diet_details', ''),
            )
            
            # Note: admission_number is NOT set here - it will be auto-generated
            # by the pre_save signal in signals.py
            
            student.save()

            # ------------------------------------------------------------------
            # Handle Guardian
            # ------------------------------------------------------------------
            guardian_option = form_data.get('guardian_option', 'skip')
            
            if guardian_option == 'new':
                # Create new guardian
                guardian = Guardian.objects.create(
                    first_name=form_data.get('guardian_first_name'),
                    last_name=form_data.get('guardian_last_name'),
                    primary_phone=form_data.get('guardian_phone'),
                    email=form_data.get('guardian_email', ''),
                    home_address=form_data.get('guardian_address', ''),
                    occupation=form_data.get('guardian_occupation', ''),
                    guardian_type='PRIMARY',
                )
                
                # Create relationship
                StudentGuardian.objects.create(
                    student=student,
                    guardian=guardian,
                    relationship=form_data.get('relationship'),
                    is_primary=True,
                    is_financial_responsible=True,
                    emergency_contact_priority=1,
                )
                
                logger.info(f"Created new guardian: {guardian.get_full_name()}")
            
            elif guardian_option == 'existing':
                guardian = form_data.get('existing_guardian')
                if guardian:
                    StudentGuardian.objects.create(
                        student=student,
                        guardian=guardian,
                        relationship=form_data.get('relationship'),
                        is_primary=True,
                        is_financial_responsible=True,
                        emergency_contact_priority=1,
                    )
                    logger.info(f"Linked existing guardian: {guardian.get_full_name()}")

            # ------------------------------------------------------------------
            # Success - Admission number was auto-generated by signal
            # ------------------------------------------------------------------
            messages.success(
                self.request,
                f"Student {student.get_full_name()} "
                f"(#{student.admission_number}) was created successfully!",
                extra_tags='sweetalert'
            )

            return redirect('students:student_profile', pk=student.pk)

        except Exception as exc:
            logger.exception("Error in wizard done method:")
            logger.exception(exc)
            
            messages.error(
                self.request,
                f"Error creating student: {exc}",
                extra_tags='sweetalert-error'
            )
            return redirect('students:student_list')


# View entry point
student_create = StudentCreateWizard.as_view()


@login_required
def student_edit(request, pk):
    """Edit existing student"""
    student = get_object_or_404(Student, pk=pk)

    if request.method == "POST":
        form = StudentForm(request.POST, request.FILES, instance=student)
        if form.is_valid():
            student = form.save()
            
            messages.success(
                request,
                f"Student {student.get_full_name()} was updated successfully",
                extra_tags='sweetalert'
            )
            return redirect("students:student_profile", pk=student.pk)
        else:
            messages.error(
                request,
                "Please correct the errors in the form",
                extra_tags='sweetalert-error'
            )
    else:
        form = StudentForm(instance=student)

    context = {
        'form': form,
        'student': student,
        'title': 'Update Student',
    }

    return render(request, 'students/form.html', context)


@login_required
def student_profile(request, pk):
    """View student profile with all related information - USES stats.py"""
    student = get_object_or_404(
        Student.objects.prefetch_related(
            Prefetch(
                'guardian_relationships',
                queryset=StudentGuardian.objects.filter(is_active=True).select_related('guardian')
            ),
            'sibling_relationships',
            'reverse_sibling_relationships',
        ),
        pk=pk
    )

    # Get student summary
    try:
        summary = {
            'admission_number': student.admission_number,
            'full_name': student.get_full_name(),
            'status': student.get_enrollment_status_display(),
            'age': student.age,
            'years_in_school': student_stats.get_years_in_school(student),
            'days_until_birthday': student_stats.get_days_until_birthday(student),
            'is_birthday_today': student_stats.is_birthday_today(student),
            
            # Related counts
            'guardian_count': student.guardians.count(),
            'sibling_count': student_stats.get_sibling_count_for_student(student),
            'has_medical_alerts': student.has_medical_alert(),
        }
    except Exception as e:
        logger.error(f"Error getting student summary: {e}")
        summary = {}

    # Get related data
    guardians = student.guardian_relationships.filter(is_active=True)
    primary_guardian = guardians.filter(is_primary=True).first()
    emergency_contacts = guardians.filter(emergency_contact_priority__lte=5).order_by('emergency_contact_priority')
    
    # Sibling relationships
    siblings_forward = student.sibling_relationships.select_related('to_student__current_academic_level')
    siblings_reverse = student.reverse_sibling_relationships.select_related('from_student__current_academic_level')
    
    # Recent status changes
    status_history = student.status_history.select_related('academic_session').order_by('-effective_date')[:5]

    context = {
        'student': student,
        'summary': summary,
        'guardians': guardians,
        'primary_guardian': primary_guardian,
        'emergency_contacts': emergency_contacts,
        'siblings_forward': siblings_forward,
        'siblings_reverse': siblings_reverse,
        'status_history': status_history,
    }
    
    return render(request, "students/profile.html", context)


@login_required
def student_delete(request, pk):
    """Delete student with HTMX support"""
    student = get_object_or_404(Student, pk=pk)
    
    if request.method == 'POST':
        # Check if deletion is allowed
        can_delete = True
        errors = []
        
        if student.enrollment_status == 'ACTIVE':
            can_delete = False
            errors.append("Cannot delete active students")
        
        if hasattr(student, 'class_enrollments') and student.class_enrollments.exists():
            can_delete = False
            errors.append("Student has class enrollments")
        
        if hasattr(student, 'invoices') and student.invoices.exists():
            can_delete = False
            errors.append("Student has financial records")
        
        if not can_delete:
            is_htmx = request.headers.get('HX-Request') == 'true'
            
            if is_htmx:
                response = HttpResponse()
                response['HX-Alert-Message'] = '; '.join(errors)
                response['HX-Alert-Type'] = 'error'
                response['HX-Alert-Title'] = 'Cannot Delete'
                response['HX-Close-Modal'] = 'true'
                return response
            else:
                messages.error(request, '; '.join(errors), extra_tags='sweetalert-error')
                return redirect('students:student_list')
        
        student_name = student.get_full_name()
        student.delete()
        
        is_htmx = request.headers.get('HX-Request') == 'true'
        
        if is_htmx:
            # Return updated student list
            filter_form = StudentFilterForm(request.GET or None)
            students = get_filtered_students(request)
            
            stats = students.aggregate(
                total=Count('id'),
                active=Count('id', filter=Q(enrollment_status='ACTIVE')),
            )
            
            paginator = Paginator(students, 20)
            page_number = request.GET.get('page', 1)
            students_page = paginator.get_page(page_number)
            
            response = render(request, 'students/_student_results.html', {
                'students_page': students_page,
                'paginator': paginator,
                'stats': stats,
                'filter_form': filter_form,
                'is_htmx': True,
            })
            
            response['HX-Alert-Message'] = f'Student "{student_name}" deleted successfully'
            response['HX-Alert-Type'] = 'success'
            response['HX-Alert-Title'] = 'Deleted!'
            response['HX-Close-Modal'] = 'true'
            
            return response
        else:
            messages.success(request, f"Student {student_name} deleted successfully", extra_tags='sweetalert')
            return redirect('students:student_list')


@login_required
def student_activate(request, pk):
    """Activate a student"""
    student = get_object_or_404(Student, pk=pk)
    
    if request.method == 'POST':
        student.enrollment_status = 'ACTIVE'
        student.save()
        
        messages.success(
            request,
            f"Student {student.get_full_name()} has been activated successfully",
            extra_tags='sweetalert'
        )
        
        return redirect('students:student_profile', pk=student.pk)
    
    return redirect('students:student_profile', pk=student.pk)


@login_required
def student_suspend(request, pk):
    """Suspend a student"""
    student = get_object_or_404(Student, pk=pk)
    
    if request.method == 'POST':
        reason = request.POST.get('reason', 'No reason provided')
        
        student.enrollment_status = 'SUSPENDED'
        student.save()
        
        messages.warning(
            request,
            f"Student {student.get_full_name()} has been suspended",
            extra_tags='sweetalert'
        )
        
        return redirect('students:student_profile', pk=student.pk)
    
    return redirect('students:student_profile', pk=student.pk)


# =============================================================================
# GUARDIAN VIEWS
# =============================================================================

@login_required
def guardian_list(request):
    """
    Handle BOTH full page loads AND HTMX search/filter requests
    """
    filter_form = GuardianFilterForm(request.GET or None)
    guardians = get_filtered_guardians(request)
    
    # Calculate statistics
    stats = guardians.aggregate(
        total=Count('id'),
        active=Count('id', filter=Q(is_active=True)),
        primary=Count('id', filter=Q(guardian_type='PRIMARY')),
        secondary=Count('id', filter=Q(guardian_type='SECONDARY')),
        emergency=Count('id', filter=Q(guardian_type='EMERGENCY')),
        financial=Count('id', filter=Q(guardian_type='FINANCIAL')),
        male=Count('id', filter=Q(gender='M')),
        female=Count('id', filter=Q(gender='F')),
        with_email=Count('id', filter=~Q(email='') & ~Q(email__isnull=True)),
    )
    
    # Pagination
    paginator = Paginator(guardians, 20)
    page_number = request.GET.get('page', 1)
    guardians_page = paginator.get_page(page_number)
    
    # Detect HTMX request
    is_htmx = request.headers.get('HX-Request') == 'true'
    
    context = {
        'guardians_page': guardians_page,
        'paginator': paginator,
        'stats': stats,
        'filter_form': filter_form,
        'is_htmx': is_htmx,
    }
    
    # Return appropriate template
    if is_htmx:
        return render(request, 'guardians/partials/_guardian_results.html', context)
    else:
        return render(request, 'guardians/list.html', context)

@login_required
def guardian_create(request):
    """Create new guardian"""
    if request.method == 'POST':
        form = GuardianForm(request.POST, request.FILES)
        if form.is_valid():
            guardian = form.save()
            
            messages.success(
                request,
                f"Guardian {guardian.get_full_name()} was created successfully",
                extra_tags='sweetalert'
            )
            return redirect('students:guardian_profile', pk=guardian.pk)
        else:
            messages.error(
                request,
                "Please correct the errors in the form",
                extra_tags='sweetalert-error'
            )
    else:
        form = GuardianForm()
    
    context = {
        'form': form,
        'title': 'Create Guardian',
    }
    
    return render(request, 'guardians/form.html', context)


@login_required
def guardian_edit(request, pk):
    """Edit existing guardian"""
    guardian = get_object_or_404(Guardian, pk=pk)
    
    if request.method == 'POST':
        form = GuardianForm(request.POST, request.FILES, instance=guardian)
        if form.is_valid():
            guardian = form.save()
            
            messages.success(
                request,
                f"Guardian {guardian.get_full_name()} was updated successfully",
                extra_tags='sweetalert'
            )
            return redirect('students:guardian_profile', pk=guardian.pk)
        else:
            messages.error(
                request,
                "Please correct the errors in the form",
                extra_tags='sweetalert-error'
            )
    else:
        form = GuardianForm(instance=guardian)
    
    context = {
        'form': form,
        'guardian': guardian,
        'title': 'Update Guardian',
    }
    
    return render(request, 'guardians/form.html', context)


@login_required
def guardian_profile(request, pk):
    """View guardian profile"""
    guardian = get_object_or_404(
        Guardian.objects.prefetch_related(
            Prefetch(
                'student_relationships',
                queryset=StudentGuardian.objects.filter(is_active=True).select_related(
                    'student__current_academic_level'
                )
            )
        ),
        pk=pk
    )
    
    # Get related students
    students = guardian.student_relationships.filter(is_active=True)
    primary_students = students.filter(is_primary=True)
    financial_students = students.filter(is_financial_responsible=True)
    
    context = {
        'guardian': guardian,
        'students': students,
        'primary_students': primary_students,
        'financial_students': financial_students,
    }
    
    return render(request, 'guardians/profile.html', context)


@login_required
def guardian_delete(request, pk):
    """Delete guardian with HTMX support"""
    guardian = get_object_or_404(Guardian, pk=pk)
    
    if request.method == 'POST':
        # Check if deletion is allowed
        if guardian.students.exists():
            is_htmx = request.headers.get('HX-Request') == 'true'
            
            if is_htmx:
                response = HttpResponse()
                response['HX-Alert-Message'] = "Cannot delete guardian with active student relationships"
                response['HX-Alert-Type'] = 'error'
                response['HX-Alert-Title'] = 'Cannot Delete'
                response['HX-Close-Modal'] = 'true'
                return response
            else:
                messages.error(request, "Cannot delete guardian with active student relationships", extra_tags='sweetalert-error')
                return redirect('students:guardian_list')
        
        guardian_name = guardian.get_full_name()
        guardian.delete()
        
        is_htmx = request.headers.get('HX-Request') == 'true'
        
        if is_htmx:
            filter_form = GuardianFilterForm(request.GET or None)
            guardians = get_filtered_guardians(request)
            
            stats = guardians.aggregate(
                total=Count('id'),
                active=Count('id', filter=Q(is_active=True)),
            )
            
            paginator = Paginator(guardians, 20)
            page_number = request.GET.get('page', 1)
            guardians_page = paginator.get_page(page_number)
            
            response = render(request, 'students/guardians/_guardian_results.html', {
                'guardians_page': guardians_page,
                'paginator': paginator,
                'stats': stats,
                'filter_form': filter_form,
                'is_htmx': True,
            })
            
            response['HX-Alert-Message'] = f'Guardian "{guardian_name}" deleted successfully'
            response['HX-Alert-Type'] = 'success'
            response['HX-Alert-Title'] = 'Deleted!'
            response['HX-Close-Modal'] = 'true'
            
            return response
        else:
            messages.success(request, f"Guardian {guardian_name} deleted successfully", extra_tags='sweetalert')
            return redirect('students:guardian_list')


@login_required
def guardian_print_view(request):
    """Generate printable guardian list"""
    
    selected_fields = request.GET.getlist('fields')
    if not selected_fields:
        selected_fields = [
            'full_name', 'guardian_type', 'primary_phone', 
            'email', 'occupation', 'is_active'
        ]
    
    include_stats = request.GET.get('include_stats') == 'true'
    landscape_mode = request.GET.get('landscape') == 'true'
    
    guardians = get_filtered_guardians(request)
    
    stats = None
    if include_stats:
        stats = guardians.aggregate(
            total=Count('id'),
            active=Count('id', filter=Q(is_active=True)),
            primary=Count('id', filter=Q(guardian_type='PRIMARY')),
            with_email=Count('id', filter=~Q(email='') & ~Q(email__isnull=True)),
        )
    
    field_names = {
        'full_name': 'Full Name',
        'guardian_type': 'Guardian Type',
        'primary_phone': 'Primary Phone',
        'secondary_phone': 'Secondary Phone',
        'email': 'Email',
        'occupation': 'Occupation',
        'employer': 'Employer',
        'home_address': 'Home Address',
        'is_active': 'Active Status',
        'student_count': 'Number of Students',
    }
    
    selected_field_names = [
        field_names.get(field, field.replace('_', ' ').title()) 
        for field in selected_fields
    ]
    
    context = {
        'guardians': guardians,
        'stats': stats,
        'now': timezone.now(),
        'selected_fields': selected_fields,
        'selected_field_names': selected_field_names,
        'field_names': field_names,
        'landscape': landscape_mode,
    }
    
    return render(request, 'students/guardians/print.html', context)


# =============================================================================
# STUDENT-GUARDIAN RELATIONSHIP VIEWS
# =============================================================================

@login_required
def student_guardian_save(request, student_pk, relationship_pk=None):
    """
    Unified save handler for creating or updating student-guardian relationship
    """
    student = get_object_or_404(Student, pk=student_pk)
    relationship = get_object_or_404(StudentGuardian, pk=relationship_pk) if relationship_pk else None
    
    if request.method == 'POST':
        guardian_id = request.POST.get('guardian_id', '').strip()
        relation = request.POST.get('relationship', '').strip()
        
        if not all([guardian_id, relation]):
            return render(request, 'students/partials/modals/_student_guardian_form.html', {
                'student': student,
                'relationship': relationship,
                'relationship_choices': StudentGuardian.RELATIONSHIP_CHOICES,
                'available_guardians': Guardian.objects.filter(is_active=True).order_by('last_name', 'first_name'),
                'error_message': 'Guardian and relationship type are required',
            })
        
        try:
            guardian = Guardian.objects.get(pk=guardian_id)
            
            # Common fields
            common_data = {
                'relationship': relation,
                'is_primary': request.POST.get('is_primary') == 'on',
                'is_financial_responsible': request.POST.get('is_financial_responsible') == 'on',
                'can_pickup': request.POST.get('can_pickup') == 'on',
                'can_authorize_medical': request.POST.get('can_authorize_medical') == 'on',
                'emergency_contact_priority': int(request.POST.get('emergency_contact_priority', 999)),
                'has_custody': request.POST.get('has_custody') == 'on',
                'receives_academic_reports': request.POST.get('receives_academic_reports') == 'on',
                'receives_financial_statements': request.POST.get('receives_financial_statements') == 'on',
                'receives_emergency_notifications': request.POST.get('receives_emergency_notifications') == 'on',
                'is_active': request.POST.get('is_active') == 'on',
                'notes': request.POST.get('notes', '').strip() or None,
            }
            
            if relationship:
                # UPDATE
                for key, value in common_data.items():
                    setattr(relationship, key, value)
                relationship.save()
                message = 'Relationship updated successfully'
            else:
                # CREATE
                relationship = StudentGuardian.objects.create(
                    student=student,
                    guardian=guardian,
                    **common_data
                )
                message = 'Relationship created successfully'
            
            # If setting as primary, unset others
            if relationship.is_primary:
                StudentGuardian.objects.filter(student=student).exclude(pk=relationship.pk).update(is_primary=False)
            
            is_htmx = request.headers.get('HX-Request') == 'true'
            
            if is_htmx:
                response = HttpResponse('')
                response['HX-Alert-Message'] = message
                response['HX-Alert-Type'] = 'success'
                response['HX-Close-Modal'] = 'true'
                response['HX-Redirect'] = reverse('students:student_profile', kwargs={'pk': student.pk})
                return response
            else:
                messages.success(request, message)
                return redirect('students:student_profile', pk=student.pk)
                
        except Guardian.DoesNotExist:
            return render(request, 'students/partials/modals/_student_guardian_form.html', {
                'student': student,
                'relationship': relationship,
                'error_message': 'Guardian not found',
            })
    
    return redirect('students:student_profile', pk=student.pk)


@login_required
def student_guardian_delete(request, pk):
    """Delete student-guardian relationship with HTMX support"""
    relationship = get_object_or_404(StudentGuardian, pk=pk)
    student = relationship.student
    
    if request.method == 'POST':
        relationship.delete()
        
        is_htmx = request.headers.get('HX-Request') == 'true'
        
        if is_htmx:
            response = HttpResponse('')
            response['HX-Alert-Message'] = 'Guardian relationship removed successfully'
            response['HX-Alert-Type'] = 'success'
            response['HX-Close-Modal'] = 'true'
            response['HX-Redirect'] = reverse('students:student_profile', kwargs={'pk': student.pk})
            return response
        else:
            messages.success(request, "Guardian relationship removed successfully", extra_tags='sweetalert')
            return redirect('students:student_profile', pk=student.pk)
    
    return redirect('students:student_profile', pk=student.pk)


@login_required
def student_guardian_set_primary(request, pk):
    """Set student-guardian relationship as primary with HTMX support"""
    relationship = get_object_or_404(StudentGuardian, pk=pk)
    student = relationship.student
    
    if request.method == 'POST':
        # Make this relationship primary
        StudentGuardian.objects.filter(student=student).update(is_primary=False)
        relationship.is_primary = True
        relationship.save()
        
        is_htmx = request.headers.get('HX-Request') == 'true'
        
        if is_htmx:
            response = HttpResponse('')
            response['HX-Alert-Message'] = 'Primary guardian updated successfully'
            response['HX-Alert-Type'] = 'success'
            response['HX-Close-Modal'] = 'true'
            response['HX-Redirect'] = reverse('students:student_profile', kwargs={'pk': student.pk})
            return response
        else:
            messages.success(request, "Primary guardian updated successfully", extra_tags='sweetalert')
            return redirect('students:student_profile', pk=student.pk)
    
    return redirect('students:student_profile', pk=student.pk)


# =============================================================================
# SIBLING RELATIONSHIP VIEWS
# =============================================================================

@login_required
def sibling_save(request, student_pk, sibling_pk=None):
    """
    Unified save handler for creating or updating sibling relationship
    """
    student = get_object_or_404(Student, pk=student_pk)
    sibling_rel = get_object_or_404(SiblingRelationship, pk=sibling_pk) if sibling_pk else None
    
    if request.method == 'POST':
        to_student_id = request.POST.get('to_student_id', '').strip()
        relationship_type = request.POST.get('relationship_type', 'FULL')
        
        if not to_student_id:
            return render(request, 'students/partials/modals/_sibling_form.html', {
                'student': student,
                'sibling_rel': sibling_rel,
                'relationship_types': SiblingRelationship.RELATIONSHIP_TYPES,
                'error_message': 'Sibling student is required',
            })
        
        try:
            to_student = Student.objects.get(pk=to_student_id)
            
            if sibling_rel:
                # UPDATE
                sibling_rel.relationship_type = relationship_type
                sibling_rel.is_verified = request.POST.get('is_verified') == 'on'
                sibling_rel.notes = request.POST.get('notes', '').strip() or None
                sibling_rel.save()
                message = 'Sibling relationship updated successfully'
            else:
                # CREATE
                SiblingRelationship.objects.create(
                    from_student=student,
                    to_student=to_student,
                    relationship_type=relationship_type,
                )
                message = 'Sibling relationship created successfully'
            
            is_htmx = request.headers.get('HX-Request') == 'true'
            
            if is_htmx:
                response = HttpResponse('')
                response['HX-Alert-Message'] = message
                response['HX-Alert-Type'] = 'success'
                response['HX-Close-Modal'] = 'true'
                response['HX-Redirect'] = reverse('students:student_profile', kwargs={'pk': student.pk})
                return response
            else:
                messages.success(request, message)
                return redirect('students:student_profile', pk=student.pk)
                
        except Student.DoesNotExist:
            return render(request, 'students/partials/modals/_sibling_form.html', {
                'student': student,
                'sibling_rel': sibling_rel,
                'error_message': 'Student not found',
            })
    
    return redirect('students:student_profile', pk=student.pk)


@login_required
def sibling_delete(request, pk):
    """Delete sibling relationship with HTMX support"""
    sibling_rel = get_object_or_404(SiblingRelationship, pk=pk)
    student = sibling_rel.from_student
    
    if request.method == 'POST':
        sibling_rel.delete()
        
        is_htmx = request.headers.get('HX-Request') == 'true'
        
        if is_htmx:
            response = HttpResponse('')
            response['HX-Alert-Message'] = 'Sibling relationship removed successfully'
            response['HX-Alert-Type'] = 'success'
            response['HX-Close-Modal'] = 'true'
            response['HX-Redirect'] = reverse('students:student_profile', kwargs={'pk': student.pk})
            return response
        else:
            messages.success(request, "Sibling relationship removed successfully", extra_tags='sweetalert')
            return redirect('students:student_profile', pk=student.pk)
    
    return redirect('students:student_profile', pk=student.pk)


# =============================================================================
# ENROLLMENT STATUS HISTORY VIEWS
# =============================================================================

@login_required
def enrollment_history_list(request):
    """List enrollment status history"""
    
    # Get all status history with related data
    history = EnrollmentStatusHistory.objects.select_related(
        'student__current_academic_level',
        'academic_session'
    ).order_by('-effective_date', 'student__admission_number')
    
    # Apply filters if any
    student_id = request.GET.get('student')
    status = request.GET.get('status')
    from_date = request.GET.get('from_date')
    to_date = request.GET.get('to_date')
    
    if student_id:
        history = history.filter(student_id=student_id)
    if status:
        history = history.filter(new_status=status)
    if from_date:
        history = history.filter(effective_date__gte=from_date)
    if to_date:
        history = history.filter(effective_date__lte=to_date)
    
    # Pagination
    paginator = Paginator(history, 50)
    page_number = request.GET.get('page', 1)
    history_page = paginator.get_page(page_number)
    
    context = {
        'history_page': history_page,
        'paginator': paginator,
        'title': 'Enrollment Status History',
    }
    
    return render(request, 'students/enrollment_history/list.html', context)


@login_required
def enrollment_history_detail(request, pk):
    """View enrollment status history entry details"""
    
    history = get_object_or_404(
        EnrollmentStatusHistory.objects.select_related(
            'student__current_academic_level',
            'academic_session'
        ),
        pk=pk
    )
    
    context = {
        'history': history,
    }
    
    return render(request, 'students/enrollment_history/detail.html', context)


# =============================================================================
# BULK ACTION VIEWS
# =============================================================================

@login_required
def bulk_status_change(request):
    """Bulk change student enrollment status"""
    
    if request.method != 'POST':
        return redirect('students:student_list')
    
    student_ids = request.POST.getlist('student_ids')
    new_status = request.POST.get('new_status')
    reason = request.POST.get('reason', 'Bulk status change')
    
    if not student_ids or not new_status:
        messages.error(
            request,
            'Invalid request: missing student IDs or status',
            extra_tags='sweetalert-error'
        )
        return redirect('students:student_list')
    
    try:
        with transaction.atomic():
            students = Student.objects.filter(pk__in=student_ids)
            count = 0
            
            for student in students:
                old_status = student.enrollment_status
                
                # Skip if already at this status
                if old_status == new_status:
                    continue
                
                student.enrollment_status = new_status
                student.save()
                
                # Create status history record
                EnrollmentStatusHistory.objects.create(
                    student=student,
                    previous_status=old_status,
                    new_status=new_status,
                    effective_date=get_school_today(),
                    reason=reason
                )
                
                count += 1
        
        is_htmx = request.headers.get('HX-Request') == 'true'
        
        if is_htmx:
            # Return updated student list
            filter_form = StudentFilterForm(request.GET or None)
            students = get_filtered_students(request)
            
            stats = students.aggregate(
                total=Count('id'),
                active=Count('id', filter=Q(enrollment_status='ACTIVE')),
            )
            
            paginator = Paginator(students, 20)
            page_number = request.GET.get('page', 1)
            students_page = paginator.get_page(page_number)
            
            response = render(request, 'students/_student_results.html', {
                'students_page': students_page,
                'paginator': paginator,
                'stats': stats,
                'filter_form': filter_form,
                'is_htmx': True,
            })
            
            response['HX-Alert-Message'] = f'Successfully changed status for {count} student(s)'
            response['HX-Alert-Type'] = 'success'
            response['HX-Alert-Title'] = 'Bulk Update Complete'
            response['HX-Close-Modal'] = 'true'
            
            return response
        else:
            messages.success(
                request,
                f"Successfully changed status for {count} student(s)",
                extra_tags='sweetalert'
            )
            return redirect('students:student_list')
        
    except Exception as e:
        logger.error(f"Error in bulk status change: {e}")
        
        is_htmx = request.headers.get('HX-Request') == 'true'
        
        if is_htmx:
            response = HttpResponse()
            response['HX-Alert-Message'] = f'Error: {str(e)}'
            response['HX-Alert-Type'] = 'error'
            response['HX-Alert-Title'] = 'Bulk Update Failed'
            response['HX-Close-Modal'] = 'true'
            return response
        else:
            messages.error(request, f'Error: {str(e)}', extra_tags='sweetalert-error')
            return redirect('students:student_list')


@login_required
def bulk_assign_guardian(request):
    """Bulk assign guardian to multiple students"""
    
    if request.method != 'POST':
        return redirect('students:student_list')
    
    student_ids = request.POST.getlist('student_ids')
    guardian_id = request.POST.get('guardian_id')
    relationship = request.POST.get('relationship')
    is_primary = request.POST.get('is_primary') == 'on'
    is_financial = request.POST.get('is_financial_responsible') == 'on'
    
    if not student_ids or not guardian_id or not relationship:
        messages.error(
            request,
            'Invalid request: missing required fields',
            extra_tags='sweetalert-error'
        )
        return redirect('students:student_list')
    
    try:
        guardian = Guardian.objects.get(pk=guardian_id)
        
        with transaction.atomic():
            students = Student.objects.filter(pk__in=student_ids)
            count = 0
            skipped = 0
            
            for student in students:
                # Check if relationship already exists
                if StudentGuardian.objects.filter(
                    student=student,
                    guardian=guardian
                ).exists():
                    skipped += 1
                    continue
                
                # Create relationship
                StudentGuardian.objects.create(
                    student=student,
                    guardian=guardian,
                    relationship=relationship,
                    is_primary=is_primary,
                    is_financial_responsible=is_financial,
                    emergency_contact_priority=1 if is_primary else 999,
                )
                count += 1
        
        is_htmx = request.headers.get('HX-Request') == 'true'
        
        message = f'Successfully linked {count} student(s) to {guardian.get_full_name()}'
        if skipped > 0:
            message += f' ({skipped} skipped - already linked)'
        
        if is_htmx:
            # Return updated student list
            filter_form = StudentFilterForm(request.GET or None)
            students = get_filtered_students(request)
            
            stats = students.aggregate(
                total=Count('id'),
                active=Count('id', filter=Q(enrollment_status='ACTIVE')),
            )
            
            paginator = Paginator(students, 20)
            page_number = request.GET.get('page', 1)
            students_page = paginator.get_page(page_number)
            
            response = render(request, 'students/_student_results.html', {
                'students_page': students_page,
                'paginator': paginator,
                'stats': stats,
                'filter_form': filter_form,
                'is_htmx': True,
            })
            
            response['HX-Alert-Message'] = message
            response['HX-Alert-Type'] = 'success'
            response['HX-Alert-Title'] = 'Bulk Assignment Complete'
            response['HX-Close-Modal'] = 'true'
            
            return response
        else:
            messages.success(request, message, extra_tags='sweetalert')
            return redirect('students:student_list')
        
    except Guardian.DoesNotExist:
        messages.error(request, 'Guardian not found', extra_tags='sweetalert-error')
        return redirect('students:student_list')
    except Exception as e:
        logger.error(f"Error in bulk assign guardian: {e}")
        
        is_htmx = request.headers.get('HX-Request') == 'true'
        
        if is_htmx:
            response = HttpResponse()
            response['HX-Alert-Message'] = f'Error: {str(e)}'
            response['HX-Alert-Type'] = 'error'
            response['HX-Alert-Title'] = 'Bulk Assignment Failed'
            response['HX-Close-Modal'] = 'true'
            return response
        else:
            messages.error(request, f'Error: {str(e)}', extra_tags='sweetalert-error')
            return redirect('students:student_list')


# =============================================================================
# REPORTS & ANALYTICS
# =============================================================================

@login_required
def student_reports_dashboard(request):
    """Dashboard for student reports and analytics"""
    
    try:
        stats = student_stats.get_comprehensive_statistics()
    except Exception as e:
        logger.error(f"Error getting comprehensive statistics: {e}")
        stats = {}
    
    context = {
        'stats': stats,
        'title': 'Student Reports & Analytics',
    }
    
    return render(request, 'students/reports/dashboard.html', context)


@login_required
def demographics_report(request):
    """Demographics report"""
    
    try:
        student_stats_data = student_stats.get_student_statistics()
        family_stats_data = student_stats.get_family_statistics()
    except Exception as e:
        logger.error(f"Error getting demographics: {e}")
        student_stats_data = {}
        family_stats_data = {}
    
    context = {
        'student_stats': student_stats_data,
        'family_stats': family_stats_data,
        'title': 'Demographics Report',
    }
    
    return render(request, 'students/reports/demographics.html', context)


@login_required
def health_report(request):
    """Health and medical report"""
    
    # Get students with medical alerts
    medical_alerts = Student.objects.filter(
        enrollment_status='ACTIVE'
    ).filter(
        Q(medical_conditions__isnull=False) & ~Q(medical_conditions='') |
        Q(allergies__isnull=False) & ~Q(allergies='') |
        Q(medications__isnull=False) & ~Q(medications='') |
        Q(has_special_needs=True)
    ).select_related('current_academic_level').order_by('admission_number')
    
    # Statistics
    total_active = Student.objects.filter(enrollment_status='ACTIVE').count()
    stats = {
        'total_with_alerts': medical_alerts.count(),
        'with_conditions': medical_alerts.exclude(Q(medical_conditions='') | Q(medical_conditions__isnull=True)).count(),
        'with_allergies': medical_alerts.exclude(Q(allergies='') | Q(allergies__isnull=True)).count(),
        'on_medications': medical_alerts.exclude(Q(medications='') | Q(medications__isnull=True)).count(),
        'special_needs': medical_alerts.filter(has_special_needs=True).count(),
        'special_diet': Student.objects.filter(enrollment_status='ACTIVE', requires_special_diet=True).count(),
        'percentage': round((medical_alerts.count() / total_active * 100), 1) if total_active > 0 else 0,
    }
    
    context = {
        'medical_alerts': medical_alerts,
        'stats': stats,
        'title': 'Health & Medical Report',
    }
    
    return render(request, 'students/reports/health.html', context)


@login_required
def guardian_report(request):
    """Guardian report"""
    
    try:
        guardian_stats_data = student_stats.get_guardian_statistics()
        occupation_stats = student_stats.get_guardian_occupation_stats()
    except Exception as e:
        logger.error(f"Error getting guardian stats: {e}")
        guardian_stats_data = {}
        occupation_stats = {}
    
    # Students without guardians
    students_without_guardians = Student.objects.filter(
        enrollment_status='ACTIVE'
    ).annotate(
        guardian_count=Count('guardians')
    ).filter(guardian_count=0).select_related('current_academic_level')
    
    context = {
        'guardian_stats': guardian_stats_data,
        'occupation_stats': occupation_stats,
        'students_without_guardians': students_without_guardians,
        'title': 'Guardian Report',
    }
    
    return render(request, 'students/reports/guardians.html', context)


@login_required
def sibling_report(request):
    """Sibling relationships report"""
    
    try:
        sibling_stats_data = student_stats.get_sibling_statistics()
        largest_groups = student_stats.get_largest_sibling_groups()
    except Exception as e:
        logger.error(f"Error getting sibling stats: {e}")
        sibling_stats_data = {}
        largest_groups = []
    
    context = {
        'sibling_stats': sibling_stats_data,
        'largest_groups': largest_groups,
        'title': 'Sibling Relationships Report',
    }
    
    return render(request, 'students/reports/siblings.html', context)


@login_required
def birthday_report(request):
    """Birthday report (upcoming birthdays)"""
    
    today = get_school_today()
    
    # Get birthdays for next 30 days
    upcoming_birthdays = []
    for i in range(30):
        check_date = today + timedelta(days=i)
        students = Student.objects.filter(
            enrollment_status='ACTIVE',
            date_of_birth__month=check_date.month,
            date_of_birth__day=check_date.day
        ).select_related('current_academic_level').order_by('date_of_birth')
        
        if students:
            upcoming_birthdays.append({
                'date': check_date,
                'students': students,
                'days_away': i,
            })
    
    context = {
        'upcoming_birthdays': upcoming_birthdays,
        'title': 'Birthday Report',
    }
    
    return render(request, 'students/reports/birthdays.html', context)


# =============================================================================
# EXPORT FUNCTIONS
# =============================================================================

@login_required
def export_students_excel(request):
    """Export students to Excel"""
    
    students = get_filtered_students(request)
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Students"
    
    # Headers
    headers = [
        'Admission Number', 'Full Name', 'Gender', 'Date of Birth', 'Age',
        'Current Grade', 'Status', 'Phone', 'Email', 'Admission Date'
    ]
    ws.append(headers)
    
    # Style headers
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        cell.font = Font(bold=True, color='FFFFFF')
    
    # Data rows
    for student in students:
        ws.append([
            student.admission_number,
            student.get_full_name(),
            student.get_gender_display(),
            student.date_of_birth.strftime('%Y-%m-%d') if student.date_of_birth else '',
            student.age,
            str(student.current_academic_level) if student.current_academic_level else '',
            student.get_enrollment_status_display(),
            student.phone_number or '',
            student.personal_email or '',
            student.admission_date.strftime('%Y-%m-%d') if student.admission_date else '',
        ])
    
    # Prepare response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="students_{timezone.now().strftime("%Y%m%d")}.xlsx"'
    
    wb.save(response)
    return response


@login_required
def export_students_pdf(request):
    """Export students to PDF"""
    
    students = get_filtered_students(request)
    
    # Create PDF
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4))
    elements = []
    
    # Styles
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=24,
        textColor=colors.HexColor('#1a1a1a'),
        spaceAfter=30,
        alignment=TA_CENTER
    )
    
    # Title
    elements.append(Paragraph('Student List', title_style))
    elements.append(Spacer(1, 20))
    
    # Table data
    data = [['Admission #', 'Name', 'Gender', 'Age', 'Grade', 'Status']]
    
    for student in students:
        data.append([
            student.admission_number,
            student.get_full_name()[:30],
            student.get_gender_display(),
            str(student.age),
            str(student.current_academic_level)[:20] if student.current_academic_level else '',
            student.get_enrollment_status_display()
        ])
    
    # Create table
    table = Table(data)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    
    elements.append(table)
    doc.build(elements)
    
    # Prepare response
    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="students_{timezone.now().strftime("%Y%m%d")}.pdf"'
    
    return response


@login_required
def export_guardians_excel(request):
    """Export guardians to Excel"""
    
    guardians = get_filtered_guardians(request)
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Guardians"
    
    # Headers
    headers = [
        'Full Name', 'Type', 'Primary Phone', 'Email', 'Occupation',
        'Employer', 'Home Address', '# Students', 'Status'
    ]
    ws.append(headers)
    
    # Style headers
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        cell.font = Font(bold=True, color='FFFFFF')
    
    # Data rows
    for guardian in guardians:
        ws.append([
            guardian.get_full_name(),
            guardian.get_guardian_type_display(),
            guardian.primary_phone,
            guardian.email or '',
            guardian.occupation or '',
            guardian.employer or '',
            guardian.home_address or '',
            guardian.student_count,
            'Active' if guardian.is_active else 'Inactive',
        ])
    
    # Prepare response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="guardians_{timezone.now().strftime("%Y%m%d")}.xlsx"'
    
    wb.save(response)
    return response


@login_required
def export_guardians_pdf(request):
    """Export guardians to PDF"""
    
    guardians = get_filtered_guardians(request)
    
    # Create PDF
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4))
    elements = []
    
    # Styles
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=24,
        textColor=colors.HexColor('#1a1a1a'),
        spaceAfter=30,
        alignment=TA_CENTER
    )
    
    # Title
    elements.append(Paragraph('Guardian List', title_style))
    elements.append(Spacer(1, 20))
    
    # Table data
    data = [['Name', 'Type', 'Phone', 'Email', 'Occupation', '# Students']]
    
    for guardian in guardians:
        data.append([
            guardian.get_full_name()[:30],
            guardian.get_guardian_type_display(),
            guardian.primary_phone,
            guardian.email[:25] if guardian.email else '',
            guardian.occupation[:20] if guardian.occupation else '',
            str(guardian.student_count),
        ])
    
    # Create table
    table = Table(data)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    
    elements.append(table)
    doc.build(elements)
    
    # Prepare response
    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="guardians_{timezone.now().strftime("%Y%m%d")}.pdf"'
    
    return response


# =============================================================================
# QUICK STATS API ENDPOINTS (for dashboard widgets/AJAX)
# =============================================================================

@login_required
def student_quick_stats(request):
    """Get quick statistics for students (JSON API)"""
    
    today = get_school_today()
    
    # Age calculations
    age_ranges = {
        'under_5': Student.objects.filter(
            enrollment_status='ACTIVE',
            date_of_birth__gte=today.replace(year=today.year - 5)
        ).count(),
        'age_5_10': Student.objects.filter(
            enrollment_status='ACTIVE',
            date_of_birth__lt=today.replace(year=today.year - 5),
            date_of_birth__gte=today.replace(year=today.year - 11)
        ).count(),
        'age_11_15': Student.objects.filter(
            enrollment_status='ACTIVE',
            date_of_birth__lt=today.replace(year=today.year - 11),
            date_of_birth__gte=today.replace(year=today.year - 16)
        ).count(),
        'age_16_plus': Student.objects.filter(
            enrollment_status='ACTIVE',
            date_of_birth__lt=today.replace(year=today.year - 16)
        ).count(),
    }
    
    stats = {
        'total': Student.objects.count(),
        'active': Student.objects.filter(enrollment_status='ACTIVE').count(),
        'suspended': Student.objects.filter(enrollment_status='SUSPENDED').count(),
        'graduated': Student.objects.filter(enrollment_status='GRADUATED').count(),
        'male': Student.objects.filter(enrollment_status='ACTIVE', gender='M').count(),
        'female': Student.objects.filter(enrollment_status='ACTIVE', gender='F').count(),
        'special_needs': Student.objects.filter(
            enrollment_status='ACTIVE',
            has_special_needs=True
        ).count(),
        'transportation': Student.objects.filter(
            enrollment_status='ACTIVE',
            transportation_required=True
        ).count(),
        **age_ranges
    }
    
    return JsonResponse(stats)


@login_required
def guardian_quick_stats(request):
    """Get quick statistics for guardians (JSON API)"""
    
    stats = {
        'total': Guardian.objects.filter(is_active=True).count(),
        'primary': Guardian.objects.filter(
            guardian_type='PRIMARY',
            is_active=True
        ).count(),
        'secondary': Guardian.objects.filter(
            guardian_type='SECONDARY',
            is_active=True
        ).count(),
        'financial': Guardian.objects.filter(
            guardian_type='FINANCIAL',
            is_active=True
        ).count(),
        'with_email': Guardian.objects.filter(
            is_active=True
        ).exclude(Q(email='') | Q(email__isnull=True)).count(),
        'total_students': StudentGuardian.objects.filter(is_active=True).count(),
    }
    
    return JsonResponse(stats)


@login_required
def medical_alerts_quick_stats(request):
    """Get quick statistics for students with medical alerts (JSON API)"""
    
    stats = {
        'total_medical_alerts': Student.objects.filter(
            enrollment_status='ACTIVE'
        ).filter(
            Q(medical_conditions__isnull=False) & ~Q(medical_conditions='') |
            Q(allergies__isnull=False) & ~Q(allergies='') |
            Q(medications__isnull=False) & ~Q(medications='')
        ).count(),
        'with_conditions': Student.objects.filter(
            enrollment_status='ACTIVE'
        ).exclude(Q(medical_conditions='') | Q(medical_conditions__isnull=True)).count(),
        'with_allergies': Student.objects.filter(
            enrollment_status='ACTIVE'
        ).exclude(Q(allergies='') | Q(allergies__isnull=True)).count(),
        'on_medications': Student.objects.filter(
            enrollment_status='ACTIVE'
        ).exclude(Q(medications='') | Q(medications__isnull=True)).count(),
        'special_needs': Student.objects.filter(
            enrollment_status='ACTIVE',
            has_special_needs=True
        ).count(),
        'special_diet': Student.objects.filter(
            enrollment_status='ACTIVE',
            requires_special_diet=True
        ).count(),
    }
    
    return JsonResponse(stats)