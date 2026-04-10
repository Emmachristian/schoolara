# fees/views.py

"""
Fees Management Views

Structure:
  Dashboard
  Display Groups       (filter + CRUD + print + export)
  Fee Categories       (filter + CRUD + print + export)
  Fee Structures       (filter + CRUD + print + export)
  Student Accounts     (filter + CRUD + print + export)
  Account Transactions (filter + list + print + export)
  Fee Invoices         (filter + CRUD + print + export)
  Payments             (filter + CRUD + print + export + API)
  Scholarship Programs (filter + CRUD + print + export)
  Scholarship Applications (filter + CRUD + print + export)
  Student Scholarships (filter + CRUD + print + export)
  Discounts            (filter + CRUD + print + export)
  Reports
"""

from django.shortcuts import render, redirect, get_object_or_404
from django.urls import reverse
from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator
from django.contrib import messages
from django.db.models import Q, Count, Sum, Avg, F, Max, Min, Prefetch, Case, When, Value, IntegerField
from django.utils import timezone
from django.http import JsonResponse, HttpResponse
from django.db import transaction
from django.views.decorators.http import require_http_methods
from datetime import timedelta, date, datetime
from decimal import Decimal, InvalidOperation
import logging

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from django.core.mail import send_mail
from django.template.loader import render_to_string
from django.conf import settings

from .models import (
    DisplayGroup,
    FeesCategory,
    FeesStructure,
    FeesStructureItem,
    FeesStructureBillingSplit,
    FeeInvoice,
    FeeInvoiceItem,
    Payment,
    ScholarshipProgram,
    StudentScholarship,
    StudentScholarshipApplication,
    ScholarshipApplicationLog,
    DiscountPolicy,
    StudentDiscount,
    DiscountApplication,
    StudentAccount,
    AccountTransaction,
)

from academics.models import AcademicSession

from .forms import (
    DisplayGroupForm,
    DisplayGroupFilterForm,
    FeesCategoryForm,
    FeesCategoryFilterForm,
    FeesStructureForm,
    FeesStructureItemInlineFormSet,
    FeesStructureBillingSplitInlineFormSet,
    FeesStructureItemForm,
    FeesStructureFilterForm,
    FeeInvoiceFilterForm,
    PaymentForm,
    MultipleInvoicePaymentForm,
    PaymentReversalForm,
    PaymentRefundForm,
    PaymentFilterForm,
    BulkPaymentVerificationForm,
    ScholarshipProgramForm,
    ScholarshipProgramFilterForm,
    StudentScholarshipForm,
    StudentScholarshipFilterForm,
    StudentScholarshipApplicationForm,
    ScholarshipApplicationFilterForm,
    ScholarshipApplicationApprovalForm,
    DiscountTierFormSet,
    DiscountPolicyForm,
    DiscountPolicyFilterForm,
    StudentDiscountForm,
    StudentDiscountFilterForm,
    StudentAccountForm,
    StudentAccountFilterForm,
    StudentAccountAdjustmentForm,
    AccountTransactionFilterForm,
)

from core.utils import (
    get_school_today,
    get_school_current_time,
    get_school_timezone,
    localize_datetime,
    get_active_academic_session,
    format_money,
    calculate_percentage,
    paginate_queryset,
    parse_filters,

)

from core.view_helpers import (
    get_print_school_context,
)

from core.models import FiscalPeriod
from students.models import Student

from fees.invoice_generators import UnifiedStudentInvoiceGenerator
from fees.utils import (
    get_invoice_items_organized,
    get_invoice_status_color,
    validate_payment_data,
    calculate_line_item_totals,
)

from . import stats as fees_stats


from formtools.wizard.views import SessionWizardView

MAX_PRINT_RECORDS = 500

logger = logging.getLogger(__name__)


# =============================================================================
# EXCEL EXPORT HELPERS
# =============================================================================

_HEADER_FILL  = PatternFill(start_color='1E3A5F', end_color='1E3A5F', fill_type='solid')
_HEADER_FONT  = Font(bold=True, color='FFFFFF', size=11, name='Arial')
_HEADER_ALIGN = Alignment(horizontal='center', vertical='center', wrap_text=True)
_DATA_ALIGN   = Alignment(vertical='center', wrap_text=False)


def _make_workbook(sheet_title, columns, rows):
    """Build a styled Workbook. columns: list of (header, accessor). rows: iterable."""
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_title

    ws.append([col[0] for col in columns])
    for cell in ws[1]:
        cell.font      = _HEADER_FONT
        cell.fill      = _HEADER_FILL
        cell.alignment = _HEADER_ALIGN
    ws.row_dimensions[1].height = 28

    for obj in rows:
        ws.append([col[1](obj) for col in columns])

    for col_cells in ws.columns:
        max_len = max(
            (len(str(c.value)) if c.value is not None else 0) for c in col_cells
        )
        ws.column_dimensions[col_cells[0].column_letter].width = min(max_len + 4, 60)

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = _DATA_ALIGN

    return wb


def _xlsx_response(wb, filename_stem):
    """Return an HttpResponse that triggers an xlsx download."""
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = (
        f'attachment; filename="{filename_stem}_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx"'
    )
    wb.save(response)
    return response


def _resolve_columns(all_columns, selected_keys, default_keys):
    """Return ordered (header, accessor) pairs matching selected_keys."""
    col_map = {c[0]: c for c in all_columns}
    keys    = [k for k in selected_keys if k in col_map] or default_keys
    return [(col_map[k][1], col_map[k][2]) for k in keys if k in col_map]


# =============================================================================
# DASHBOARD
# =============================================================================

@login_required
def fees_dashboard(request):
    """Main fees dashboard with overview statistics."""
    from django.db.models import Value
    from django.db.models.functions import Coalesce

    # ── Session is fetched independently — a stats error must never wipe it ──
    current_session = get_active_academic_session()

    # ── Stats — each fetched individually so one failure doesn't blank the rest ──
    def _safe(fn, *args, **kwargs):
        try:
            return fn(*args, **kwargs)
        except Exception as e:
            logger.error(f"Dashboard stats error ({fn.__name__}): {e}")
            return {}

    session_id    = current_session.id if current_session else None
    dashboard_data    = _safe(fees_stats.get_financial_dashboard, academic_session_id=session_id)
    account_stats     = _safe(fees_stats.get_student_account_statistics)
    invoice_stats     = _safe(fees_stats.get_invoice_statistics)
    payment_stats     = _safe(fees_stats.get_payment_statistics)
    scholarship_stats = _safe(fees_stats.get_scholarship_statistics)
    discount_stats    = _safe(fees_stats.get_discount_statistics)

    today = get_school_today()

    recent_invoices = FeeInvoice.objects.select_related(
        'student', 'academic_session'
    ).order_by('-created_at')[:10]

    recent_payments = Payment.objects.select_related(
        'student', 'invoice', 'payment_method'
    ).order_by('-created_at')[:10]

    overdue_invoices = FeeInvoice.objects.filter(
        due_date__lt=today,
        status__in=['PENDING', 'PARTIALLY_PAID', 'OVERDUE'],
    ).select_related('student', 'academic_session').order_by('due_date')[:10]

    unverified_payments = Payment.objects.filter(
        is_verified=False, status='COMPLETED',
    ).select_related('student', 'payment_method').order_by('-created_at')[:10]

    pending_scholarship_applications = StudentScholarshipApplication.objects.filter(
        status='SUBMITTED'
    ).select_related('student', 'scholarship_program').order_by('-application_date')[:10]

    accounts_in_debt = (
        StudentAccount.objects
        .annotate(
            computed_balance=Coalesce(Sum('transactions__amount'), Value(Decimal('0.00')))
        )
        .filter(computed_balance__lt=0)
        .select_related('student')
        .order_by('computed_balance')[:10]
    )

    return render(request, 'fees/dashboard.html', {
        'dashboard_data':                    dashboard_data,
        'account_stats':                     account_stats,
        'invoice_stats':                     invoice_stats,
        'payment_stats':                     payment_stats,
        'scholarship_stats':                 scholarship_stats,
        'discount_stats':                    discount_stats,
        'current_session':                   current_session,
        'recent_invoices':                   recent_invoices,
        'recent_payments':                   recent_payments,
        'overdue_invoices':                  overdue_invoices,
        'unverified_payments':               unverified_payments,
        'pending_scholarship_applications':  pending_scholarship_applications,
        'accounts_in_debt':                  accounts_in_debt,
    })


# =============================================================================
# DISPLAY GROUPS
# =============================================================================

def get_filtered_display_groups(request):
    groups = DisplayGroup.objects.annotate(
        category_count=Count('feescategory')
    ).order_by('display_order', 'name')

    query        = request.GET.get('q', '').strip()
    is_active    = request.GET.get('is_active', '')
    show_as_group = request.GET.get('show_as_group', '')

    if query:
        words = query.split()
        q = Q()
        for w in words:
            q &= Q(name__icontains=w) | Q(description__icontains=w)
        groups = groups.filter(q)

    if is_active:
        groups = groups.filter(is_active=(is_active.lower() == 'true'))
    if show_as_group:
        groups = groups.filter(show_as_group=(show_as_group.lower() == 'true'))

    return groups


@login_required
def display_group_list(request):
    filter_form = DisplayGroupFilterForm(request.GET or None)
    groups      = get_filtered_display_groups(request)

    try:
        raw_stats = fees_stats.get_display_group_statistics()
    except Exception as e:
        logger.error(f"Error getting display group statistics: {e}")
        raw_stats = {}

    stats = {
        'total':             groups.count(),
        'active':            groups.filter(is_active=True).count(),
        'grouped':           groups.filter(show_as_group=True).count(),
        'ungrouped':         groups.filter(show_as_group=False).count(),
        'total_categories':  raw_stats.get('category_distribution', {}).get('total_categories_assigned', 0),
    }

    paginator    = Paginator(groups, 10)
    groups_page  = paginator.get_page(request.GET.get('page', 1))
    is_htmx      = request.headers.get('HX-Request') == 'true'

    context = {
        'groups_page': groups_page,
        'paginator':   paginator,
        'stats':       stats,
        'filter_form': filter_form,
        'is_htmx':     is_htmx,
    }

    if is_htmx:
        return render(request, 'fees/display_groups/partials/_group_results.html', context)
    return render(request, 'fees/display_groups/list.html', context)


@login_required
def display_group_create(request):
    if request.method == 'POST':
        form = DisplayGroupForm(request.POST)
        if form.is_valid():
            try:
                group   = form.save()
                is_htmx = request.headers.get('HX-Request') == 'true'
                if is_htmx:
                    r = HttpResponse()
                    r['HX-Alert-Message'] = f"Display group '{group.name}' created successfully!"
                    r['HX-Alert-Type']    = 'success'
                    r['HX-Alert-Title']   = 'Created!'
                    r['HX-Redirect']      = reverse('fees:display_group_detail', kwargs={'pk': group.pk})
                    return r
                messages.success(request, f"Display group '{group.name}' created successfully!", extra_tags='sweetalert')
                return redirect('fees:display_group_detail', pk=group.pk)
            except Exception as e:
                logger.error(f"Error creating display group: {e}")
                is_htmx = request.headers.get('HX-Request') == 'true'
                if is_htmx:
                    r = HttpResponse()
                    r['HX-Alert-Message'] = f'Error creating group: {str(e)}'
                    r['HX-Alert-Type']    = 'error'
                    r['HX-Alert-Title']   = 'Error!'
                    return r
                messages.error(request, f'Error creating group: {str(e)}')
    else:
        form = DisplayGroupForm()

    return render(request, 'fees/display_groups/form.html', {
        'form': form, 'title': 'Create Display Group', 'submit_text': 'Create Group',
    })


@login_required
def display_group_detail(request, pk):
    group      = get_object_or_404(DisplayGroup, pk=pk)
    categories = group.feescategory_set.order_by('display_order')
    return render(request, 'fees/display_groups/detail.html', {
        'group': group, 'categories': categories,
    })


@login_required
def display_group_edit(request, pk):
    group = get_object_or_404(DisplayGroup, pk=pk)
    if request.method == 'POST':
        form = DisplayGroupForm(request.POST, instance=group)
        if form.is_valid():
            try:
                group   = form.save()
                is_htmx = request.headers.get('HX-Request') == 'true'
                if is_htmx:
                    r = HttpResponse()
                    r['HX-Alert-Message'] = f"Display group '{group.name}' updated successfully!"
                    r['HX-Alert-Type']    = 'success'
                    r['HX-Alert-Title']   = 'Updated!'
                    r['HX-Redirect']      = reverse('fees:display_group_detail', kwargs={'pk': group.pk})
                    return r
                messages.success(request, f"Display group '{group.name}' updated successfully!", extra_tags='sweetalert')
                return redirect('fees:display_group_detail', pk=group.pk)
            except Exception as e:
                logger.error(f"Error updating display group: {e}")
                is_htmx = request.headers.get('HX-Request') == 'true'
                if is_htmx:
                    r = HttpResponse()
                    r['HX-Alert-Message'] = f'Error updating group: {str(e)}'
                    r['HX-Alert-Type']    = 'error'
                    r['HX-Alert-Title']   = 'Error!'
                    return r
                messages.error(request, f'Error updating group: {str(e)}')
    else:
        form = DisplayGroupForm(instance=group)

    return render(request, 'fees/display_groups/form.html', {
        'form': form, 'group': group,
        'title': f'Edit Display Group - {group.name}', 'submit_text': 'Update Group',
    })


@login_required
@require_http_methods(["POST"])
def display_group_delete(request, pk):
    group = get_object_or_404(DisplayGroup, pk=pk)
    if group.feescategory_set.exists():
        is_htmx = request.headers.get('HX-Request') == 'true'
        if is_htmx:
            r = HttpResponse()
            r['HX-Alert-Message'] = f"Cannot delete '{group.name}' — it has associated fee categories"
            r['HX-Alert-Type']    = 'error'
            r['HX-Alert-Title']   = 'Cannot Delete'
            r['HX-Close-Modal']   = 'true'
            return r
        messages.error(request, f"Cannot delete '{group.name}' — it has associated fee categories", extra_tags='sweetalert-error')
        return redirect('fees:display_group_list')

    name = group.name
    group.delete()
    is_htmx = request.headers.get('HX-Request') == 'true'
    if is_htmx:
        r = HttpResponse()
        r['HX-Alert-Message'] = f"Display group '{name}' deleted successfully"
        r['HX-Alert-Type']    = 'success'
        r['HX-Alert-Title']   = 'Deleted!'
        r['HX-Close-Modal']   = 'true'
        r['HX-Redirect']      = reverse('fees:display_group_list')
        return r
    messages.success(request, f"Display group '{name}' deleted successfully", extra_tags='sweetalert')
    return redirect('fees:display_group_list')


@login_required
@require_http_methods(["POST"])
def display_group_toggle_active(request, pk):
    group = get_object_or_404(DisplayGroup, pk=pk)
    try:
        group.is_active = not group.is_active
        group.save()
        status  = "activated" if group.is_active else "deactivated"
        is_htmx = request.headers.get('HX-Request') == 'true'
        if is_htmx:
            r = HttpResponse()
            r['HX-Alert-Message'] = f"Display group '{group.name}' {status}!"
            r['HX-Alert-Type']    = 'success'
            r['HX-Alert-Title']   = 'Status Updated'
            r['HX-Close-Modal']   = 'true'
            r['HX-Redirect']      = reverse('fees:display_group_detail', kwargs={'pk': group.pk})
            return r
        messages.success(request, f"Display group '{group.name}' {status}!")
        return redirect('fees:display_group_detail', pk=group.pk)
    except Exception as e:
        logger.error(f"Error toggling display group status: {e}")
        is_htmx = request.headers.get('HX-Request') == 'true'
        if is_htmx:
            r = HttpResponse()
            r['HX-Alert-Message'] = f'Error updating status: {str(e)}'
            r['HX-Alert-Type']    = 'error'
            r['HX-Alert-Title']   = 'Error'
            return r
        messages.error(request, f'Error updating status: {str(e)}')
        return redirect('fees:display_group_detail', pk=pk)


@login_required
def display_group_list_print_view(request):
    FIELD_NAMES_FULL = {
        'name':           'Name',
        'description':    'Description',
        'display_order':  'Display Order',
        'color_code':     'Color Code',
        'show_as_group':  'Show as Group',
        'show_subtotal':  'Show Subtotal',
        'category_count': 'No. of Categories',
        'is_active':      'Active',
    }
    FIELD_NAMES_SHORT = {
        'name':           'Name',
        'description':    'Description',
        'display_order':  'Order',
        'color_code':     'Color',
        'show_as_group':  'Grouped',
        'show_subtotal':  'Subtotal',
        'category_count': '# Cats',
        'is_active':      'Active',
    }
    DEFAULT_FIELDS   = ['name', 'display_order', 'show_as_group', 'category_count', 'is_active']
    selected_fields  = request.GET.getlist('fields') or DEFAULT_FIELDS
    short_headers    = request.GET.get('short_headers', 'false').lower() == 'true'
    landscape        = request.GET.get('landscape', 'false').lower() == 'true'
    field_names      = FIELD_NAMES_SHORT if short_headers else FIELD_NAMES_FULL
    groups           = get_filtered_display_groups(request)

    return render(request, 'fees/display_groups/print_display_group_list.html', {
        **get_print_school_context(request),
        'groups':               groups,
        'selected_fields':      selected_fields,
        'selected_field_names': [field_names.get(f, f.replace('_', ' ').title()) for f in selected_fields],
        'field_names':          field_names,
        'short_headers':        short_headers,
        'landscape':            landscape,
        'now':                  timezone.now(),
        'print_date':           get_school_today(),
        'printed_by':           request.user.get_full_name() or request.user.username,
        'title':                'Display Groups',
    })


@login_required
def export_display_groups_excel(request):
    ALL_COLUMNS = [
        ('name',          'Name',           lambda o: o.name),
        ('description',   'Description',    lambda o: o.description or ''),
        ('display_order', 'Display Order',  lambda o: o.display_order),
        ('color_code',    'Color',          lambda o: o.color_code or ''),
        ('show_as_group', 'Show as Group',  lambda o: 'Yes' if o.show_as_group else 'No'),
        ('show_subtotal', 'Show Subtotal',  lambda o: 'Yes' if o.show_group_subtotal else 'No'),
        ('category_count','# Categories',   lambda o: o.category_count or 0),
        ('is_active',     'Active',         lambda o: 'Yes' if o.is_active else 'No'),
    ]
    DEFAULT_FIELDS = ['name', 'display_order', 'show_as_group', 'category_count', 'is_active']
    groups  = get_filtered_display_groups(request)
    columns = _resolve_columns(ALL_COLUMNS, request.GET.getlist('fields'), DEFAULT_FIELDS)
    return _xlsx_response(_make_workbook('Display Groups', columns, groups), 'display_groups')


# =============================================================================
# FEE CATEGORIES
# =============================================================================

def get_filtered_fee_categories(request):
    categories = FeesCategory.objects.select_related('display_group').annotate(
        structure_count=Count('structure_items', distinct=True)
    ).order_by('-is_active', 'display_order', 'name')

    query         = request.GET.get('q', '').strip()
    category_type = request.GET.get('category_type', '')
    is_active     = request.GET.get('is_active', '')
    is_mandatory  = request.GET.get('is_mandatory', '')
    is_refundable = request.GET.get('is_refundable', '')
    is_taxable    = request.GET.get('is_taxable', '')
    applicability = request.GET.get('applicability', '')
    display_group = request.GET.get('display_group', '')
    frequency     = request.GET.get('frequency', '')

    if query:
        words = query.split()
        q = Q()
        for w in words:
            q &= Q(name__icontains=w) | Q(code__icontains=w) | Q(description__icontains=w)
        categories = categories.filter(q)

    if category_type:
        categories = categories.filter(category_type=category_type)
    if display_group:
        categories = categories.filter(display_group_id=display_group)
    if frequency:
        categories = categories.filter(frequency=frequency)
    if applicability:
        categories = categories.filter(applicability=applicability)
    if is_active:
        categories = categories.filter(is_active=(is_active.lower() == 'true'))
    if is_mandatory:
        categories = categories.filter(is_mandatory=(is_mandatory.lower() == 'true'))
    if is_refundable:
        categories = categories.filter(is_refundable=(is_refundable.lower() == 'true'))
    if is_taxable:
        categories = categories.filter(is_taxable=(is_taxable.lower() == 'true'))

    return categories


@login_required
def fee_category_list(request):
    filter_form = FeesCategoryFilterForm(request.GET or None)
    categories  = get_filtered_fee_categories(request)

    try:
        raw_stats = fees_stats.get_fee_category_statistics()
    except Exception as e:
        logger.error(f"Error getting fee category statistics: {e}")
        raw_stats = {}

    config = raw_stats.get('configuration', {})
    stats  = {
        'total':      categories.count(),
        'active':     categories.filter(is_active=True).count(),
        'mandatory':  config.get('mandatory', categories.filter(is_mandatory=True).count()),
        'optional':   config.get('optional',  categories.filter(is_mandatory=False).count()),
        'refundable': config.get('refundable', categories.filter(is_refundable=True).count()),
        'taxable':    config.get('taxable',    categories.filter(is_taxable=True).count()),
        'tuition':    categories.filter(category_type='TUITION').count(),
        'boarding':   categories.filter(category_type='BOARDING').count(),
    }

    paginator       = Paginator(categories, 10)
    categories_page = paginator.get_page(request.GET.get('page', 1))
    is_htmx         = request.headers.get('HX-Request') == 'true'

    context = {
        'categories_page': categories_page,
        'paginator':       paginator,
        'stats':           stats,
        'filter_form':     filter_form,
        'is_htmx':         is_htmx,
    }
    if is_htmx:
        return render(request, 'fees/categories/partials/_category_results.html', context)
    return render(request, 'fees/categories/list.html', context)


@login_required
def fee_category_create(request):
    if request.method == 'POST':
        form = FeesCategoryForm(request.POST)
        if form.is_valid():
            try:
                category = form.save()
                is_htmx  = request.headers.get('HX-Request') == 'true'
                if is_htmx:
                    r = HttpResponse()
                    r['HX-Alert-Message'] = f"Fee category '{category.name}' created successfully!"
                    r['HX-Alert-Type']    = 'success'
                    r['HX-Alert-Title']   = 'Created!'
                    r['HX-Redirect']      = reverse('fees:category_detail', kwargs={'pk': category.pk})
                    return r
                messages.success(request, f"Fee category '{category.name}' created successfully!", extra_tags='sweetalert')
                return redirect('fees:category_detail', pk=category.pk)
            except Exception as e:
                logger.error(f"Error creating category: {e}")
                is_htmx = request.headers.get('HX-Request') == 'true'
                if is_htmx:
                    r = HttpResponse()
                    r['HX-Alert-Message'] = f'Error creating category: {str(e)}'
                    r['HX-Alert-Type']    = 'error'
                    r['HX-Alert-Title']   = 'Error!'
                    return r
                messages.error(request, f'Error creating category: {str(e)}')
    else:
        form = FeesCategoryForm()

    return render(request, 'fees/categories/form.html', {
        'form': form, 'title': 'Create Fee Category', 'submit_text': 'Create Category',
    })


@login_required
def fee_category_detail(request, pk):
    category = get_object_or_404(FeesCategory, pk=pk)

    structure_items = category.structure_items.filter(
        fee_structure__isnull=False
    ).select_related(
        'fee_structure', 'fee_structure__academic_year', 'fee_category',
    ).order_by('-fee_structure__is_active', 'fee_structure__name')[:15]

    structure_count    = category.structure_items.filter(fee_structure__isnull=False).values('fee_structure').distinct().count()
    invoice_item_count = category.feeinvoiceitem_set.count()

    return render(request, 'fees/categories/detail.html', {
        'category':          category,
        'structure_items':   structure_items,
        'structure_count':   structure_count,
        'invoice_item_count': invoice_item_count,
    })


@login_required
def fee_category_edit(request, pk):
    category = get_object_or_404(FeesCategory, pk=pk)
    if request.method == 'POST':
        form = FeesCategoryForm(request.POST, instance=category)
        if form.is_valid():
            try:
                category = form.save()
                is_htmx  = request.headers.get('HX-Request') == 'true'
                if is_htmx:
                    r = HttpResponse()
                    r['HX-Alert-Message'] = f"Fee category '{category.name}' updated successfully!"
                    r['HX-Alert-Type']    = 'success'
                    r['HX-Alert-Title']   = 'Updated!'
                    r['HX-Redirect']      = reverse('fees:category_detail', kwargs={'pk': category.pk})
                    return r
                messages.success(request, f"Fee category '{category.name}' updated successfully!", extra_tags='sweetalert')
                return redirect('fees:category_detail', pk=category.pk)
            except Exception as e:
                logger.error(f"Error updating category: {e}")
                is_htmx = request.headers.get('HX-Request') == 'true'
                if is_htmx:
                    r = HttpResponse()
                    r['HX-Alert-Message'] = f'Error updating category: {str(e)}'
                    r['HX-Alert-Type']    = 'error'
                    r['HX-Alert-Title']   = 'Error!'
                    return r
                messages.error(request, f'Error updating category: {str(e)}')
    else:
        form = FeesCategoryForm(instance=category)

    return render(request, 'fees/categories/form.html', {
        'form': form, 'category': category,
        'title': f'Edit Category - {category.name}', 'submit_text': 'Update Category',
    })


@login_required
@require_http_methods(["POST"])
def fee_category_delete(request, pk):
    category = get_object_or_404(FeesCategory, pk=pk)
    if category.structure_items.exists() or category.feeinvoiceitem_set.exists():
        is_htmx = request.headers.get('HX-Request') == 'true'
        if is_htmx:
            r = HttpResponse()
            r['HX-Alert-Message'] = f"Cannot delete '{category.name}' — it is used in fee structures or invoices"
            r['HX-Alert-Type']    = 'error'
            r['HX-Alert-Title']   = 'Cannot Delete'
            r['HX-Close-Modal']   = 'true'
            return r
        messages.error(request, f"Cannot delete '{category.name}' — it is used in fee structures or invoices", extra_tags='sweetalert-error')
        return redirect('fees:category_list')

    name = category.name
    category.delete()
    is_htmx = request.headers.get('HX-Request') == 'true'
    if is_htmx:
        r = HttpResponse()
        r['HX-Alert-Message'] = f"Fee category '{name}' deleted successfully"
        r['HX-Alert-Type']    = 'success'
        r['HX-Alert-Title']   = 'Deleted!'
        r['HX-Close-Modal']   = 'true'
        r['HX-Redirect']      = reverse('fees:category_list')
        return r
    messages.success(request, f"Fee category '{name}' deleted successfully", extra_tags='sweetalert')
    return redirect('fees:category_list')


@login_required
@require_http_methods(["POST"])
def fee_category_toggle_active(request, pk):
    category = get_object_or_404(FeesCategory, pk=pk)
    try:
        category.is_active = not category.is_active
        category.save()
        status  = "activated" if category.is_active else "deactivated"
        is_htmx = request.headers.get('HX-Request') == 'true'
        if is_htmx:
            r = HttpResponse()
            r['HX-Alert-Message'] = f"Fee category '{category.name}' {status}!"
            r['HX-Alert-Type']    = 'success'
            r['HX-Alert-Title']   = 'Status Updated'
            r['HX-Close-Modal']   = 'true'
            r['HX-Redirect']      = reverse('fees:category_detail', kwargs={'pk': category.pk})
            return r
        messages.success(request, f"Fee category '{category.name}' {status}!")
        return redirect('fees:category_detail', pk=category.pk)
    except Exception as e:
        logger.error(f"Error toggling category status: {e}")
        is_htmx = request.headers.get('HX-Request') == 'true'
        if is_htmx:
            r = HttpResponse()
            r['HX-Alert-Message'] = f'Error updating status: {str(e)}'
            r['HX-Alert-Type']    = 'error'
            r['HX-Alert-Title']   = 'Error'
            return r
        messages.error(request, f'Error updating status: {str(e)}')
        return redirect('fees:category_detail', pk=pk)


@login_required
def fee_category_list_print_view(request):
    FIELD_NAMES_FULL = {
        'code':             'Code',
        'name':             'Name',
        'category_type':    'Category Type',
        'display_group':    'Display Group',
        'display_order':    'Display Order',
        'applicability':    'Applicable To',
        'frequency':        'Frequency',
        'is_recurring':     'Recurring',
        'is_mandatory':     'Mandatory',
        'is_refundable':    'Refundable',
        'allows_partial':   'Partial Payment',
        'is_taxable':       'Taxable',
        'default_tax_rate': 'Tax Rate %',
        'structure_count':  'No. of Structures',
        'is_active':        'Active',
        'description':      'Description',
    }
    FIELD_NAMES_SHORT = {
        'code':             'Code',
        'name':             'Name',
        'category_type':    'Type',
        'display_group':    'Group',
        'display_order':    'Order',
        'applicability':    'Applies To',
        'frequency':        'Freq.',
        'is_recurring':     'Recur.',
        'is_mandatory':     'Mand.',
        'is_refundable':    'Refund.',
        'allows_partial':   'Partial',
        'is_taxable':       'Tax',
        'default_tax_rate': 'Tax %',
        'structure_count':  '# Structs',
        'is_active':        'Active',
        'description':      'Description',
    }
    DEFAULT_FIELDS  = ['code', 'name', 'category_type', 'display_group', 'applicability', 'frequency', 'is_mandatory', 'is_refundable', 'is_taxable', 'is_active']
    selected_fields = request.GET.getlist('fields') or DEFAULT_FIELDS
    short_headers   = request.GET.get('short_headers', 'false').lower() == 'true'
    landscape       = request.GET.get('landscape', 'false').lower() == 'true'
    include_stats   = request.GET.get('include_stats', 'true').lower() == 'true'
    field_names     = FIELD_NAMES_SHORT if short_headers else FIELD_NAMES_FULL
    categories      = get_filtered_fee_categories(request)

    stats = None
    if include_stats:
        stats = {
            'total':      categories.count(),
            'active':     categories.filter(is_active=True).count(),
            'mandatory':  categories.filter(is_mandatory=True).count(),
            'optional':   categories.filter(is_mandatory=False).count(),
            'refundable': categories.filter(is_refundable=True).count(),
            'taxable':    categories.filter(is_taxable=True).count(),
        }

    return render(request, 'fees/categories/print_fee_categories_list.html', {
        **get_print_school_context(request),
        'categories':           categories,
        'stats':                stats,
        'selected_fields':      selected_fields,
        'selected_field_names': [field_names.get(f, f.replace('_', ' ').title()) for f in selected_fields],
        'field_names':          field_names,
        'short_headers':        short_headers,
        'landscape':            landscape,
        'now':                  timezone.now(),
        'print_date':           get_school_today(),
        'printed_by':           request.user.get_full_name() or request.user.username,
        'title':                'Fee Categories',
    })


@login_required
def export_fee_categories_excel(request):
    ALL_COLUMNS = [
        ('code',            'Code',            lambda o: o.code),
        ('name',            'Name',            lambda o: o.name),
        ('category_type',   'Type',            lambda o: o.get_category_type_display()),
        ('display_group',   'Display Group',   lambda o: o.display_group.name if o.display_group else ''),
        ('display_order',   'Display Order',   lambda o: o.display_order),
        ('applicability',   'Applicable To',   lambda o: o.get_applicability_display()),
        ('frequency',       'Frequency',       lambda o: o.get_frequency_display() if o.frequency else ''),
        ('is_recurring',    'Recurring',       lambda o: 'Yes' if o.is_recurring else 'No'),
        ('is_mandatory',    'Mandatory',       lambda o: 'Yes' if o.is_mandatory else 'No'),
        ('is_refundable',   'Refundable',      lambda o: 'Yes' if o.is_refundable else 'No'),
        ('allows_partial',  'Partial Payment', lambda o: 'Yes' if o.allows_partial_payment else 'No'),
        ('is_taxable',      'Taxable',         lambda o: 'Yes' if o.is_taxable else 'No'),
        ('default_tax_rate','Tax Rate %',      lambda o: float(o.default_tax_rate)),
        ('structure_count', '# Structures',    lambda o: o.structure_count or 0),
        ('is_active',       'Active',          lambda o: 'Yes' if o.is_active else 'No'),
        ('description',     'Description',     lambda o: o.description or ''),
    ]
    DEFAULT_FIELDS = ['code', 'name', 'category_type', 'display_group', 'applicability', 'frequency', 'is_mandatory', 'is_refundable', 'is_taxable', 'is_active']
    categories = get_filtered_fee_categories(request)
    columns    = _resolve_columns(ALL_COLUMNS, request.GET.getlist('fields'), DEFAULT_FIELDS)
    return _xlsx_response(_make_workbook('Fee Categories', columns, categories), 'fee_categories')


# =============================================================================
# FEE STRUCTURES
# =============================================================================

def get_filtered_fee_structures(request):
    structures = FeesStructure.objects.select_related('academic_year').prefetch_related(
        'academic_levels', 'applicable_sessions', 'applicable_classes', 'items'
    ).annotate(total_amount=Sum('items__amount')).order_by(
        '-is_active',       # active (True=1) before inactive (False=0)
        'priority',         # lower priority number = higher importance
        '-academic_year__start_date',  # newest academic year first
        'name',             # alphabetical tiebreak
    )

    query                = request.GET.get('q', '').strip()
    structure_type       = request.GET.get('structure_type', '')
    academic_year_id     = request.GET.get('academic_year', '')
    billing_frequency    = request.GET.get('billing_frequency', '')
    boarding_type_filter = request.GET.get('boarding_type_filter', '')
    academic_session_id  = request.GET.get('academic_session', '')
    academic_level_id    = request.GET.get('academic_level', '')
    is_active            = request.GET.get('is_active', '')
    effective_from       = request.GET.get('effective_from', '')
    effective_to         = request.GET.get('effective_to', '')

    if query:
        words = query.split()
        q = Q()
        for w in words:
            q &= Q(name__icontains=w) | Q(description__icontains=w) | Q(academic_year__name__icontains=w)
        structures = structures.filter(q)

    if structure_type:
        structures = structures.filter(structure_type=structure_type)
    if academic_year_id:
        structures = structures.filter(academic_year_id=academic_year_id)
    if billing_frequency:
        structures = structures.filter(billing_frequency=billing_frequency)
    if boarding_type_filter:
        structures = structures.filter(boarding_type_filter=boarding_type_filter)
    if academic_session_id:
        structures = structures.filter(applicable_sessions__id=academic_session_id)
    if academic_level_id:
        structures = structures.filter(academic_levels__id=academic_level_id)
    if is_active:
        structures = structures.filter(is_active=(is_active.lower() == 'true'))
    if effective_from:
        structures = structures.filter(effective_date__gte=effective_from)
    if effective_to:
        structures = structures.filter(effective_date__lte=effective_to)

    return structures.distinct()


@login_required
def fee_structure_list(request):
    filter_form = FeesStructureFilterForm(request.GET or None)
    structures  = get_filtered_fee_structures(request)

    stats = {
        'total':           structures.count(),
        'active':          structures.filter(is_active=True).count(),
        'standard':        structures.filter(structure_type='STANDARD').count(),
        'with_late_fees':  structures.filter(charges_late_fee=True).count(),
    }

    paginator       = Paginator(structures, 25)
    structures_page = paginator.get_page(request.GET.get('page', 1))
    is_htmx         = request.headers.get('HX-Request') == 'true'

    context = {
        'structures_page': structures_page,
        'paginator':       paginator,
        'stats':           stats,
        'filter_form':     filter_form,
        'is_htmx':         is_htmx,
    }
    if is_htmx:
        return render(request, 'fees/structures/partials/_structure_results.html', context)
    return render(request, 'fees/structures/list.html', context)

@login_required
def fee_structure_detail(request, pk):
    structure = get_object_or_404(
        FeesStructure.objects.prefetch_related(
            'academic_levels', 'applicable_sessions',
            'applicable_classes__academic_level',
            'items__fee_category__display_group',
        ),
        pk=pk,
    )
    items        = structure.items.select_related('fee_category__display_group').order_by(
        'fee_category__display_group__display_order', 'fee_category__display_order'
    )
    total_amount = sum(item.amount for item in items)
    total_tax    = sum((item.amount * item.tax_percentage / 100) for item in items)
    invoices     = structure.invoices.select_related('student', 'academic_session').prefetch_related('items__fee_category').order_by('-issue_date')[:10]

    return render(request, 'fees/structures/detail.html', {
        'structure':            structure,
        'items':                items,
        'total_amount':         total_amount,
        'total_tax':            total_tax,
        'total_with_tax':       total_amount + total_tax,
        'invoices':             invoices,
        'invoice_count':        structure.invoices.count(),
        'active_invoice_count': structure.invoices.filter(status__in=['PENDING', 'PARTIALLY_PAID']).count(),
    })

# fees/views.py — fee structure create / edit / helpers
#
# Both public views delegate to _fee_structure_form_view.
# Do not add POST logic directly to fee_structure_edit — that was the bug
# in the original which would have caused create and edit to diverge silently.

@login_required
def fee_structure_create(request):
    return _fee_structure_form_view(request, instance=None)


@login_required
def fee_structure_edit(request, pk):
    structure = get_object_or_404(
        FeesStructure.objects.prefetch_related(
            'academic_levels', 'applicable_sessions', 'applicable_classes',
            'billing_splits__fiscal_period',
            'items__fee_category__display_group',
        ),
        pk=pk,
    )
    return _fee_structure_form_view(request, instance=structure)


def _fee_structure_form_view(request, instance):
    """
    Shared create / edit handler for FeesStructure.

    POST validation notes
    ---------------------
    All three formsets are validated unconditionally so that every
    validation error surfaces in a single round-trip.  The short-circuit
    `and` that was in the original caused items/splits errors to be
    silently discarded when the main form was invalid.

    Billing split formset is only *saved* when billing_frequency is
    SPLIT_CUSTOM; for all other frequencies _auto_generate_splits()
    rebuilds the splits automatically.
    """
    is_edit = instance is not None
    title   = f"Edit Structure — {instance.name}" if is_edit else "Create Fee Structure"

    if request.method == 'POST':
        form           = FeesStructureForm(request.POST, instance=instance)
        items_formset  = FeesStructureItemInlineFormSet(
            request.POST, instance=instance, prefix='items',
        )
        splits_formset = FeesStructureBillingSplitInlineFormSet(
            request.POST, instance=instance, prefix='splits',
        )

        billing_frequency = request.POST.get('billing_frequency', '')
        is_split_custom   = billing_frequency == 'SPLIT_CUSTOM'

        # Validate all three unconditionally — never short-circuit.
        form_valid    = form.is_valid()
        items_valid   = items_formset.is_valid()
        # Only enforce split validation when the user actually chose SPLIT_CUSTOM.
        splits_valid  = splits_formset.is_valid() if is_split_custom else True

        if form_valid and items_valid and splits_valid:
            # Extra check: custom splits must sum to exactly 100 %.
            if is_split_custom:
                split_error = _validate_splits_total(splits_formset)
                if split_error:
                    splits_formset._non_form_errors = split_error   # surface in template
                    return render(request, 'fees/structures/form.html', {
                        'form':            form,
                        'items_formset':   items_formset,
                        'splits_formset':  splits_formset,
                        'structure':       instance,
                        'title':           title,
                        'is_edit':         is_edit,
                    })

            try:
                with transaction.atomic():
                    structure = form.save()

                    items_formset.instance = structure
                    items_formset.save()

                    if is_split_custom:
                        splits_formset.instance = structure
                        splits_formset.save()
                    else:
                        _auto_generate_splits(structure)

                messages.success(
                    request,
                    f'Fee structure "{structure.name}" saved successfully.',
                    extra_tags='sweetalert',
                )
                return redirect('fees:structure_detail', pk=structure.pk)

            except Exception as e:
                logger.error(f"Error saving fee structure: {e}", exc_info=True)
                messages.error(request, f'Error saving structure: {e}')

    else:
        form           = FeesStructureForm(instance=instance)
        items_formset  = FeesStructureItemInlineFormSet(
            instance=instance, prefix='items',
        )
        splits_formset = FeesStructureBillingSplitInlineFormSet(
            instance=instance, prefix='splits',
        )

    return render(request, 'fees/structures/form.html', {
        'form':                   form,
        'items_formset':          items_formset,
        'splits_formset':         splits_formset,
        'structure':              instance,
        'title':                  title,
        'is_edit':                is_edit,
        # Invoice counts — computed in Python from the model's STATUS_CHOICES.
        # Active = invoices still requiring payment action.
        # Terminal = closed invoices (no further action expected).
        'invoice_count':           instance.invoices.count()                                                                   if is_edit else 0,
        'active_invoice_count':    instance.invoices.filter(status__in=['PENDING', 'PARTIALLY_PAID', 'OVERDUE']).count()       if is_edit else 0,
        'paid_invoice_count':      instance.invoices.filter(status='PAID').count()                                             if is_edit else 0,
        'cancelled_invoice_count': instance.invoices.filter(status__in=['CANCELLED', 'VOID', 'BAD_DEBT', 'WRITTEN_OFF', 'UNCOLLECTIBLE']).count() if is_edit else 0,
    })


def _validate_splits_total(splits_formset):
    """
    Return a non-empty error list if the non-deleted split rows do not
    sum to exactly 100.00 %, otherwise return None.

    This is a view-level guard.  The model's clean() still enforces the
    per-row constraint (no single row > 100 %) but cannot see the running
    total across rows, so we check it here.
    """
    total = Decimal('0.00')
    for form in splits_formset.forms:
        if not form.cleaned_data.get('DELETE', False):
            pct = form.cleaned_data.get('percentage') or Decimal('0.00')
            total += pct

    if total != Decimal('100.00'):
        from django.core.exceptions import ValidationError
        return ValidationError(
            f'Billing split percentages must total exactly 100.00 % '
            f'(current total: {total:.2f} %).'
        )
    return None


def _auto_generate_splits(structure):
    """
    Rebuild billing splits automatically for non-SPLIT_CUSTOM frequencies.

    Scoping
    -------
    Fiscal periods are scoped to structure.academic_year when set,
    otherwise falls back to all active, open periods ordered by number.
    The original implementation queried all active periods globally,
    which would mix periods from different academic years.

    Frequencies handled
    -------------------
    ONCE           → one split at 100 % on the first available period
    PER_PERIOD     → equal split across all periods in the academic year
    ON_ENROLLMENT  → one split at 100 % on the first available period
                     (fee is billed at the moment of enrolment; the
                     period just anchors the fiscal posting date)
    SPLIT_CUSTOM   → handled by the formset; never reaches this function
    """
    structure.billing_splits.all().delete()
    freq = structure.billing_frequency

    # Build the base queryset scoped to the structure's academic year.
    if structure.academic_year_id:
        period_qs = FiscalPeriod.objects.filter(
            fiscal_year=structure.academic_year,
            is_active=True,
            is_closed=False,
        ).order_by('period_number')
    else:
        period_qs = FiscalPeriod.objects.filter(
            is_active=True,
            is_closed=False,
        ).order_by('period_number')

    if freq in ('ONCE', 'ON_ENROLLMENT'):
        first_period = period_qs.first()
        if first_period:
            label = 'Full payment' if freq == 'ONCE' else 'Billed on enrolment'
            FeesStructureBillingSplit.objects.create(
                fee_structure=structure,
                fiscal_period=first_period,
                percentage=Decimal('100.00'),
                sequence=1,
                description=label,
            )
        else:
            logger.warning(
                f"Fee structure '{structure.name}' (freq={freq}): "
                f"no active fiscal period found — billing split not created."
            )

    elif freq == 'PER_PERIOD':
        periods = list(period_qs)
        count   = len(periods)
        if not count:
            logger.warning(
                f"Fee structure '{structure.name}' (freq=PER_PERIOD): "
                f"no active fiscal periods found — billing splits not created."
            )
            return

        # Distribute evenly; assign any rounding remainder to the last period.
        base_pct  = (Decimal('100.00') / count).quantize(Decimal('0.01'))
        total_pct = base_pct * count
        remainder = Decimal('100.00') - total_pct   # may be ±0.01 due to rounding

        splits = []
        for idx, period in enumerate(periods, 1):
            pct = base_pct + (remainder if idx == count else Decimal('0.00'))
            splits.append(FeesStructureBillingSplit(
                fee_structure=structure,
                fiscal_period=period,
                percentage=pct,
                sequence=idx,
                description=f'Instalment {idx} of {count}',
            ))
        FeesStructureBillingSplit.objects.bulk_create(splits)

    # SPLIT_CUSTOM is never passed to this function — guard just in case.
    elif freq == 'SPLIT_CUSTOM':
        logger.error(
            f"_auto_generate_splits called with SPLIT_CUSTOM for "
            f"structure '{structure.name}' — this is a bug in the caller."
        )

@login_required
def fee_structure_clone(request, pk):
    original = get_object_or_404(
        FeesStructure.objects.prefetch_related('academic_levels', 'applicable_classes', 'applicable_sessions', 'items__fee_category'),
        pk=pk,
    )
    if request.method == 'POST':
        try:
            with transaction.atomic():
                new_structure = FeesStructure.objects.get(pk=original.pk)
                new_structure.pk             = None
                new_structure._state.adding  = True
                new_structure.name           = f"{original.name} (Copy)"
                new_structure.is_active      = False
                new_structure.save()
                new_structure.academic_levels.set(original.academic_levels.all())
                new_structure.applicable_classes.set(original.applicable_classes.all())
                new_structure.applicable_sessions.set(original.applicable_sessions.all())
                for item in original.items.all():
                    FeesStructureItem.objects.create(
                        fee_structure=new_structure,
                        fee_category=item.fee_category,
                        amount=item.amount,
                        tax_percentage=item.tax_percentage,
                        default_discount_percentage=item.default_discount_percentage,
                        scholarship_eligible=item.scholarship_eligible,
                        max_scholarship_discount=item.max_scholarship_discount,
                        is_conditional=item.is_conditional,
                        condition_description=item.condition_description,
                        condition_criteria=item.condition_criteria,
                        is_payable_in_installments=item.is_payable_in_installments,
                        number_of_installments=item.number_of_installments,
                    )
            messages.success(request, f'Fee structure cloned as "{new_structure.name}"! Inactive by default — edit and activate when ready.')
            return redirect(f"{reverse('fees:structure_edit', kwargs={'pk': new_structure.pk})}?cloned=1")
        except Exception as e:
            logger.error(f"Error cloning structure: {e}", exc_info=True)
            messages.error(request, f'Error cloning structure: {str(e)}')
            return redirect('fees:structure_detail', pk=pk)

    items = original.items.select_related('fee_category__display_group').order_by(
        'fee_category__display_group__display_order', 'fee_category__display_order'
    )
    return render(request, 'fees/structures/clone.html', {
        'original_structure': original, 'items': items,
        'title': f'Clone Structure - {original.name}',
    })


@login_required
@require_http_methods(["POST"])
def fee_structure_activate(request, pk):
    structure = get_object_or_404(FeesStructure, pk=pk)
    if not structure.items.exists():
        messages.error(request, 'Cannot activate structure with no fee items.')
        return redirect('fees:structure_detail', pk=pk)
    if not structure.academic_levels.exists():
        messages.error(request, 'Cannot activate structure with no academic levels assigned.')
        return redirect('fees:structure_detail', pk=pk)
    if not structure.applicable_sessions.exists():
        messages.error(request, 'Cannot activate structure with no academic sessions assigned.')
        return redirect('fees:structure_detail', pk=pk)
    try:
        structure.is_active = True
        structure.save()
        is_htmx = request.headers.get('HX-Request') == 'true'
        if is_htmx:
            r = HttpResponse()
            r['HX-Alert-Message'] = f"Fee structure '{structure.name}' activated!"
            r['HX-Alert-Type']    = 'success'
            r['HX-Alert-Title']   = 'Structure Activated'
            r['HX-Close-Modal']   = 'true'
            r['HX-Redirect']      = reverse('fees:structure_detail', kwargs={'pk': structure.pk})
            return r
        messages.success(request, f"Fee structure '{structure.name}' activated!")
        return redirect('fees:structure_detail', pk=structure.pk)
    except Exception as e:
        logger.error(f"Error activating fee structure: {e}")
        is_htmx = request.headers.get('HX-Request') == 'true'
        if is_htmx:
            r = HttpResponse()
            r['HX-Alert-Message'] = f'Error activating structure: {str(e)}'
            r['HX-Alert-Type']    = 'error'
            r['HX-Alert-Title']   = 'Error'
            return r
        messages.error(request, f'Error activating structure: {str(e)}')
        return redirect('fees:structure_detail', pk=pk)


@login_required
@require_http_methods(["POST"])
def fee_structure_deactivate(request, pk):
    structure = get_object_or_404(FeesStructure, pk=pk)
    try:
        reason = request.POST.get('deactivation_reason', '')
        structure.is_active = False
        if reason:
            structure.description = f"{structure.description}\n\nDeactivated: {reason}".strip()
        structure.save()
        is_htmx = request.headers.get('HX-Request') == 'true'
        if is_htmx:
            r = HttpResponse()
            r['HX-Alert-Message'] = f"Fee structure '{structure.name}' deactivated!"
            r['HX-Alert-Type']    = 'success'
            r['HX-Alert-Title']   = 'Structure Deactivated'
            r['HX-Close-Modal']   = 'true'
            r['HX-Redirect']      = reverse('fees:structure_detail', kwargs={'pk': structure.pk})
            return r
        messages.success(request, f"Fee structure '{structure.name}' deactivated!")
        return redirect('fees:structure_detail', pk=structure.pk)
    except Exception as e:
        logger.error(f"Error deactivating fee structure: {e}")
        is_htmx = request.headers.get('HX-Request') == 'true'
        if is_htmx:
            r = HttpResponse()
            r['HX-Alert-Message'] = f'Error deactivating structure: {str(e)}'
            r['HX-Alert-Type']    = 'error'
            r['HX-Alert-Title']   = 'Error'
            return r
        messages.error(request, f'Error deactivating structure: {str(e)}')
        return redirect('fees:structure_detail', pk=pk)


@login_required
def fee_structure_delete(request, pk):
    structure = get_object_or_404(FeesStructure, pk=pk)
    if request.method == 'POST':
        invoice_count = structure.invoices.count()
        if invoice_count > 0:
            messages.warning(request, f'Cannot delete "{structure.name}" — it has {invoice_count} invoice(s). Consider deactivating instead.')
            return redirect('fees:structure_detail', pk=pk)
        try:
            name = structure.name
            structure.delete()
            messages.success(request, f'Fee structure "{name}" deleted successfully!')
            return redirect('fees:structure_list')
        except Exception as e:
            logger.error(f"Error deleting structure: {e}", exc_info=True)
            messages.error(request, f'Error deleting structure: {str(e)}')
            return redirect('fees:structure_detail', pk=pk)

    invoice_count = structure.invoices.count()
    return render(request, 'fees/structures/delete_confirm.html', {
        'structure':    structure,
        'invoice_count': invoice_count,
        'can_delete':   invoice_count == 0,
    })


@login_required
def fee_structure_list_print_view(request):
    FIELD_NAMES_FULL = {
        'name':               'Name',
        'structure_type':     'Structure Type',
        'academic_year':      'Academic Year',
        'billing_frequency':  'Billing Frequency',
        'boarding_filter':    'Boarding Filter',
        'student_filter':     'Student Filter',
        'total_amount':       'Total Amount',
        'item_count':         'No. of Items',
        'payment_terms_days': 'Payment Terms (Days)',
        'priority':           'Priority',
        'charges_late_fee':   'Charges Late Fees',
        'effective_date':     'Effective Date',
        'expiry_date':        'Expiry Date',
        'is_active':          'Active',
        'applicable_levels':  'Academic Levels',
        'applicable_sessions': 'Sessions',
    }
    FIELD_NAMES_SHORT = {
        'name':               'Name',
        'structure_type':     'Type',
        'academic_year':      'Year',
        'billing_frequency':  'Billing',
        'boarding_filter':    'Boarding',
        'student_filter':     'Student',
        'total_amount':       'Total',
        'item_count':         '# Items',
        'payment_terms_days': 'Terms',
        'priority':           'Pri.',
        'charges_late_fee':   'Late Fees',
        'effective_date':     'Eff. Date',
        'expiry_date':        'Exp. Date',
        'is_active':          'Active',
        'applicable_levels':  'Levels',
        'applicable_sessions': 'Sessions',
    }
    DEFAULT_FIELDS  = ['name', 'structure_type', 'academic_year', 'billing_frequency', 'boarding_filter', 'total_amount', 'item_count', 'priority', 'effective_date', 'is_active', 'applicable_levels']
    selected_fields = request.GET.getlist('fields') or DEFAULT_FIELDS
    short_headers   = request.GET.get('short_headers', 'false').lower() == 'true'
    landscape       = request.GET.get('landscape', 'true').lower() == 'true'
    include_stats   = request.GET.get('include_stats', 'true').lower() == 'true'
    field_names     = FIELD_NAMES_SHORT if short_headers else FIELD_NAMES_FULL
    structures      = get_filtered_fee_structures(request)

    stats = None
    if include_stats:
        stats = {
            'total':          structures.count(),
            'active':         structures.filter(is_active=True).count(),
            'inactive':       structures.filter(is_active=False).count(),
            'with_late_fees': structures.filter(charges_late_fee=True).count(),
        }

    return render(request, 'fees/structures/print_fee_structures_list.html', {
        **get_print_school_context(request),
        'structures':           structures,
        'stats':                stats,
        'selected_fields':      selected_fields,
        'selected_field_names': [field_names.get(f, f.replace('_', ' ').title()) for f in selected_fields],
        'field_names':          field_names,
        'short_headers':        short_headers,
        'landscape':            landscape,
        'now':                  timezone.now(),
        'print_date':           get_school_today(),
        'printed_by':           request.user.get_full_name() or request.user.username,
        'title':                'Fee Structures',
    })


@login_required
def fee_structure_print_view(request, pk):
    structure = get_object_or_404(
        FeesStructure.objects.prefetch_related(
            'academic_levels', 'applicable_sessions',
            'applicable_classes__academic_level',
            'items__fee_category__display_group',
        ),
        pk=pk,
    )
    items        = structure.items.select_related('fee_category__display_group').order_by(
        'fee_category__display_group__display_order', 'fee_category__display_order'
    )
    total_amount = sum(item.amount for item in items)
    return render(request, 'fees/structures/print_detail.html', {
        **get_print_school_context(request),
        'structure':    structure,
        'items':        items,
        'total_amount': total_amount,
        'now':          timezone.now(),
    })


@login_required
def export_fee_structures_excel(request):
    ALL_COLUMNS = [
        ('name',               'Name',              lambda o: o.name),
        ('structure_type',     'Structure Type',    lambda o: o.get_structure_type_display()),
        ('academic_year',      'Academic Year',     lambda o: str(o.academic_year) if o.academic_year else ''),
        ('billing_frequency',  'Billing Frequency', lambda o: o.get_billing_frequency_display()),
        ('boarding_filter',    'Boarding Filter',   lambda o: o.get_boarding_type_filter_display()),
        ('student_filter',     'Student Filter',    lambda o: o.get_student_type_filter_display()),
        ('total_amount',       'Total Amount',      lambda o: float(o.get_total_amount())),
        ('item_count',         '# Items',           lambda o: o.items.count()),
        ('payment_terms_days', 'Payment Terms',     lambda o: f"{o.payment_terms_days} days"),
        ('priority',           'Priority',          lambda o: o.priority),
        ('charges_late_fee',   'Late Fees',         lambda o: 'Yes' if o.charges_late_fee else 'No'),
        ('late_fee_amount',    'Late Fee Amount',   lambda o: float(o.late_fee_amount)),
        ('late_fee_pct',       'Late Fee %',        lambda o: float(o.late_fee_percentage)),
        ('grace_period',       'Grace Period',      lambda o: f"{o.grace_period_days} days"),
        ('effective_date',     'Effective Date',    lambda o: o.effective_date.strftime('%Y-%m-%d') if o.effective_date else ''),
        ('expiry_date',        'Expiry Date',       lambda o: o.expiry_date.strftime('%Y-%m-%d') if o.expiry_date else 'No expiry'),
        ('is_active',          'Active',            lambda o: 'Yes' if o.is_active else 'No'),
        ('applicable_levels',  'Academic Levels',   lambda o: ', '.join(str(l) for l in o.academic_levels.all())),
        ('applicable_sessions','Sessions',          lambda o: ', '.join(str(s) for s in o.applicable_sessions.all())),
        ('description',        'Description',       lambda o: o.description or ''),
    ]
    DEFAULT_FIELDS = ['name', 'structure_type', 'academic_year', 'billing_frequency', 'boarding_filter', 'total_amount', 'item_count', 'priority', 'effective_date', 'is_active', 'applicable_levels']
    structures = get_filtered_fee_structures(request)
    columns    = _resolve_columns(ALL_COLUMNS, request.GET.getlist('fields'), DEFAULT_FIELDS)
    return _xlsx_response(_make_workbook('Fee Structures', columns, structures), 'fee_structures')


# =============================================================================
# STUDENT ACCOUNTS
# =============================================================================

def get_filtered_student_accounts(request):
    from django.db.models import Sum, Count, Q, Case, When, Value, DecimalField

    accounts = (
        StudentAccount.objects
        .select_related(
            'student',
            'student__current_academic_level',
        )
        .annotate(
            calculated_balance = Sum('transactions__amount'),
            transaction_count  = Count('transactions', distinct=True),
            total_charges      = Sum(
                'transactions__amount',
                filter=Q(transactions__transaction_type__in=['INVOICE', 'DEBIT']),
            ),
            total_payments     = Sum(
                'transactions__amount',
                filter=Q(transactions__transaction_type='PAYMENT'),
            ),
            total_discounts    = Sum(
                'transactions__amount',
                filter=Q(transactions__transaction_type='DISCOUNT'),
            ),
        )
        .order_by('calculated_balance', 'student__first_name')
    )

    query          = request.GET.get('q', '').strip()
    status         = request.GET.get('status', '')
    balance_status = request.GET.get('balance_status', '')
    min_balance    = request.GET.get('min_balance', '')
    max_balance    = request.GET.get('max_balance', '')

    if query:
        words = query.split()
        q = Q()
        for w in words:
            q &= (
                Q(student__first_name__icontains=w) |
                Q(student__last_name__icontains=w) |
                Q(student__admission_number__icontains=w)
            )
        accounts = accounts.filter(q)

    if status:
        accounts = accounts.filter(status=status)

    if min_balance:
        try:
            accounts = accounts.filter(
                calculated_balance__gte=Decimal(min_balance)
            )
        except (ValueError, InvalidOperation):
            pass

    if max_balance:
        try:
            accounts = accounts.filter(
                calculated_balance__lte=Decimal(max_balance)
            )
        except (ValueError, InvalidOperation):
            pass

    if balance_status == 'positive':
        accounts = accounts.filter(calculated_balance__gt=0)
    elif balance_status == 'zero':
        accounts = accounts.filter(
            Q(calculated_balance=0) | Q(calculated_balance__isnull=True)
        )
    elif balance_status == 'negative':
        accounts = accounts.filter(calculated_balance__lt=0)

    return accounts


@login_required
def student_account_list(request):
    from django.db.models import Count, Sum, Q

    filter_form = StudentAccountFilterForm(request.GET or None)
    is_htmx     = request.headers.get('HX-Request') == 'true'

    # ── Base queryset ─────────────────────────────────────────────────
    base_qs = get_filtered_student_accounts(request)

    # ── Stats — single aggregate query ────────────────────────────────
    stats_agg = base_qs.aggregate(
        total_accounts        = Count('pk'),
        active                = Count('pk', filter=Q(status='ACTIVE')),
        suspended             = Count('pk', filter=Q(status='SUSPENDED')),
        frozen                = Count('pk', filter=Q(status='FROZEN')),
        closed                = Count('pk', filter=Q(status='CLOSED')),
        total_debtors         = Count('pk', filter=Q(calculated_balance__lt=0)),
        total_outstanding     = Sum('calculated_balance', filter=Q(calculated_balance__lt=0)),
        accounts_with_credit  = Count('pk', filter=Q(calculated_balance__gt=0)),
        total_credit          = Sum('calculated_balance', filter=Q(calculated_balance__gt=0)),
        zero_balance_accounts = Count('pk', filter=Q(calculated_balance=0)),
    )

    total   = stats_agg['total_accounts'] or 0
    debtors = stats_agg['total_debtors']  or 0

    stats = {
        'total_accounts': total,
        'by_status': {
            'active':    stats_agg['active']    or 0,
            'suspended': stats_agg['suspended'] or 0,
            'frozen':    stats_agg['frozen']    or 0,
            'closed':    stats_agg['closed']    or 0,
        },
        'debt_analysis': {
            'total_debtors':         debtors,
            'total_outstanding':     abs(stats_agg['total_outstanding'] or Decimal('0.00')),
            'accounts_with_credit':  stats_agg['accounts_with_credit']  or 0,
            'total_credit':          stats_agg['total_credit']           or Decimal('0.00'),
            'zero_balance_accounts': stats_agg['zero_balance_accounts']  or 0,
        },
        'collection_rate': (
            (total - debtors) / total * 100 if total > 0 else 0
        ),
    }

    # ── Paginator — inject fast count ─────────────────────────────────
    paginator        = Paginator(base_qs, 20)
    paginator._count = total  # already computed above — no extra query
    accounts_page    = paginator.get_page(request.GET.get('page', 1))

    context = {
        'accounts_page': accounts_page,
        'paginator':     paginator,
        'stats':         stats,
        'filter_form':   filter_form,
        'is_htmx':       is_htmx,
    }

    if is_htmx:
        return render(request, 'fees/accounts/partials/_account_results.html', context)
    return render(request, 'fees/accounts/list.html', context)


@login_required
def student_account_detail(request, pk):
    account      = get_object_or_404(StudentAccount.objects.select_related('student'), pk=pk)
    transactions = account.transactions.select_related('invoice', 'payment', 'academic_session', 'fiscal_period').order_by('-created_at')[:50]
    invoices     = FeeInvoice.objects.filter(student=account.student).select_related('academic_session', 'fiscal_period').order_by('-issue_date')[:10]
    payments     = Payment.objects.filter(student=account.student).select_related('payment_method', 'invoice').order_by('-payment_date')[:10]
    scholarships = StudentScholarship.objects.filter(student=account.student, status='ACTIVE').select_related('scholarship_program')

    return render(request, 'fees/accounts/detail.html', {
        'account':      account,
        'transactions': transactions,
        'invoices':     invoices,
        'payments':     payments,
        'scholarships': scholarships,
    })


@login_required
def student_account_edit(request, pk):
    account = get_object_or_404(StudentAccount, pk=pk)
    if request.method == 'POST':
        form = StudentAccountForm(request.POST, instance=account)
        if form.is_valid():
            try:
                account = form.save()
                is_htmx = request.headers.get('HX-Request') == 'true'
                if is_htmx:
                    r = HttpResponse()
                    r['HX-Alert-Message'] = f"Account for {account.student.get_full_name()} updated successfully!"
                    r['HX-Alert-Type']    = 'success'
                    r['HX-Alert-Title']   = 'Updated!'
                    r['HX-Redirect']      = reverse('fees:account_detail', kwargs={'pk': account.pk})
                    return r
                messages.success(request, f'Account for {account.student.get_full_name()} updated successfully!', extra_tags='sweetalert')
                return redirect('fees:account_detail', pk=account.pk)
            except Exception as e:
                logger.error(f"Error updating account: {e}")
                is_htmx = request.headers.get('HX-Request') == 'true'
                if is_htmx:
                    r = HttpResponse()
                    r['HX-Alert-Message'] = f'Error updating account: {str(e)}'
                    r['HX-Alert-Type']    = 'error'
                    r['HX-Alert-Title']   = 'Error!'
                    return r
                messages.error(request, f'Error updating account: {str(e)}')
    else:
        form = StudentAccountForm(instance=account)

    return render(request, 'fees/accounts/form.html', {
        'form': form, 'account': account,
        'title': f'Edit Account - {account.student.get_full_name()}',
    })


@login_required
@require_http_methods(["POST"])
def student_account_adjust(request, pk):
    account = get_object_or_404(StudentAccount, pk=pk)
    try:
        adjustment_type = request.POST.get('adjustment_type')
        amount          = Decimal(request.POST.get('amount', '0.00'))
        reason          = request.POST.get('reason', '')
        reference       = request.POST.get('reference', '')

        if amount <= 0:
            raise ValueError("Adjustment amount must be positive")
        if not reason:
            raise ValueError("Adjustment reason is required")

        with transaction.atomic():
            if adjustment_type == 'CREDIT':
                transaction_amount = amount
                description        = f"Manual Credit: {reason}"
            else:
                transaction_amount = -amount
                description        = f"Manual Debit: {reason}"

            AccountTransaction.objects.create(
                student_account=account,
                transaction_type='ADJUSTMENT',
                amount=transaction_amount,
                description=description,
                reference_number=reference or f"ADJ-{get_school_today().strftime('%Y%m%d-%H%M%S')}",
                balance_after=account.get_current_balance() + transaction_amount,
                academic_session=get_active_academic_session(),
                fiscal_period=FiscalPeriod.get_current_fiscal_period(),
                processed_by_id=str(request.user.id),
            )

            try:
                from finance.models import Journal, JournalEntry, JournalTransaction
                from core.models import FinancialSettings

                settings_obj = FinancialSettings.get_instance()
                if settings_obj:
                    mappings           = settings_obj.get_account_mappings()
                    receivable_account = mappings.student_receivables_account
                    adjustment_account = getattr(mappings, 'adjustment_account', None)

                    if receivable_account and adjustment_account:
                        from finance.utils import generate_journal_entry_number
                        fees_journal, _ = Journal.objects.get_or_create(
                            journal_type='FEES', defaults={'name': 'Fee Collection Journal'}
                        )
                        journal_entry = JournalEntry.objects.create(
                            journal=fees_journal,
                            entry_number=generate_journal_entry_number(fees_journal),
                            entry_date=get_school_today(),
                            fiscal_period=FiscalPeriod.get_current_fiscal_period(),
                            description=description,
                            reference_number=reference,
                            status='POSTED',
                        )
                        if adjustment_type == 'CREDIT':
                            JournalTransaction.objects.create(journal_entry=journal_entry, account=adjustment_account, amount=amount, is_debit=True, description=description)
                            JournalTransaction.objects.create(journal_entry=journal_entry, account=receivable_account, amount=amount, is_debit=False, description=f"Adjustment for {account.student.get_full_name()}")
                        else:
                            JournalTransaction.objects.create(journal_entry=journal_entry, account=receivable_account, amount=amount, is_debit=True, description=f"Adjustment for {account.student.get_full_name()}")
                            JournalTransaction.objects.create(journal_entry=journal_entry, account=adjustment_account, amount=amount, is_debit=False, description=description)
            except Exception as e:
                logger.error(f"Error creating adjustment journal entry: {e}", exc_info=True)

        is_htmx = request.headers.get('HX-Request') == 'true'
        if is_htmx:
            r = HttpResponse()
            r['HX-Alert-Message'] = f"Account adjustment of {amount:,.2f} processed successfully!"
            r['HX-Alert-Type']    = 'success'
            r['HX-Alert-Title']   = 'Adjustment Applied'
            r['HX-Redirect']      = reverse('fees:account_detail', kwargs={'pk': account.pk})
            return r
        messages.success(request, f"Account adjustment of {amount:,.2f} processed successfully!")
        return redirect('fees:account_detail', pk=account.pk)

    except Exception as e:
        logger.error(f"Error processing account adjustment: {e}", exc_info=True)
        is_htmx = request.headers.get('HX-Request') == 'true'
        if is_htmx:
            r = HttpResponse()
            r['HX-Alert-Message'] = f'Error processing adjustment: {str(e)}'
            r['HX-Alert-Type']    = 'error'
            r['HX-Alert-Title']   = 'Error'
            return r
        messages.error(request, f'Error processing adjustment: {str(e)}')
        return redirect('fees:account_detail', pk=account.pk)


@login_required
def student_account_list_print_view(request):
    FIELD_NAMES_FULL = {
        'admission_number': 'Admission Number',
        'student_name':     'Student Name',
        'current_class':    'Current Class',
        'contact':          'Contact Information',
        'current_balance':  'Current Balance',
        'total_charged':    'Total Fees Charged',
        'total_paid':       'Total Payments',
        'credit_limit':     'Credit Limit',
        'status':           'Account Status',
        'last_transaction': 'Last Transaction Date',
        'last_payment':     'Last Payment Date',
    }
    FIELD_NAMES_SHORT = {
        'admission_number': 'Adm. No.',
        'student_name':     'Student',
        'current_class':    'Class',
        'contact':          'Contact',
        'current_balance':  'Balance',
        'total_charged':    'Charged',
        'total_paid':       'Paid',
        'credit_limit':     'Limit',
        'status':           'Status',
        'last_transaction': 'Last Trans.',
        'last_payment':     'Last Pmt.',
    }
    DEFAULT_FIELDS  = ['admission_number', 'student_name', 'current_balance', 'total_charged', 'total_paid', 'status']
    selected_fields = request.GET.getlist('fields') or DEFAULT_FIELDS
    short_headers   = request.GET.get('short_headers', 'false').lower() == 'true'
    landscape       = request.GET.get('landscape', 'false').lower() == 'true'
    include_stats   = request.GET.get('include_stats', 'true').lower() == 'true'
    show_transactions = request.GET.get('show_transactions', 'false').lower() == 'true'
    field_names     = FIELD_NAMES_SHORT if short_headers else FIELD_NAMES_FULL
    accounts        = get_filtered_student_accounts(request)

    if accounts.count() > MAX_PRINT_RECORDS:
        messages.warning(request, f'Only the first {MAX_PRINT_RECORDS} accounts will be printed.')
        accounts = accounts[:MAX_PRINT_RECORDS]

    stats = None
    if include_stats:
        accounts_with_debt   = accounts.filter(calculated_balance__lt=0)
        accounts_with_credit = accounts.filter(calculated_balance__gt=0)
        total_accounts       = accounts.count()
        stats = {
            'total_accounts': total_accounts,
            'by_status': {
                'active':    accounts.filter(status='ACTIVE').count(),
                'suspended': accounts.filter(status='SUSPENDED').count(),
                'frozen':    accounts.filter(status='FROZEN').count(),
                'closed':    accounts.filter(status='CLOSED').count(),
            },
            'debt_analysis': {
                'total_debtors':       accounts_with_debt.count(),
                'total_outstanding':   abs(accounts_with_debt.aggregate(total=Sum('calculated_balance'))['total'] or Decimal('0.00')),
                'accounts_with_credit': accounts_with_credit.count(),
                'total_credit':        accounts_with_credit.aggregate(total=Sum('calculated_balance'))['total'] or Decimal('0.00'),
                'zero_balance_accounts': accounts.filter(calculated_balance=0).count(),
            },
            'collection_rate': (total_accounts - accounts_with_debt.count()) / total_accounts * 100 if total_accounts > 0 else 0,
        }

    transactions_by_account = {}
    if show_transactions:
        for account in accounts:
            transactions_by_account[account.pk] = account.transactions.select_related('invoice', 'payment', 'academic_session').order_by('-created_at')[:5]

    return render(request, 'fees/accounts/print_list.html', {
        **get_print_school_context(request),
        'accounts':               accounts,
        'stats':                  stats,
        'transactions_by_account': transactions_by_account,
        'selected_fields':        selected_fields,
        'selected_field_names':   [field_names.get(f, f.replace('_', ' ').title()) for f in selected_fields],
        'field_names':            field_names,
        'short_headers':          short_headers,
        'landscape':              landscape,
        'show_transactions':      show_transactions,
        'now':                    timezone.now(),
        'print_date':             get_school_today(),
        'printed_by':             request.user.get_full_name() or request.user.username,
        'title':                  'Student Accounts Report',
    })


@login_required
def student_account_print_view(request):
    selected_fields  = request.GET.getlist('fields') or ['created_at', 'description', 'reference_number', 'amount', 'balance_after']
    short_headers    = request.GET.get('short_headers', 'false').lower() == 'true'
    account_id       = request.GET.get('account_id')
    date_from        = request.GET.get('date_from')
    date_to          = request.GET.get('date_to')
    include_summary  = request.GET.get('include_summary') == 'true'

    if not account_id:
        messages.error(request, 'No account specified')
        return redirect('fees:account_list')

    account      = get_object_or_404(StudentAccount, pk=account_id)
    transactions = account.transactions.select_related('invoice', 'payment', 'academic_session').order_by('-created_at')

    if date_from:
        transactions = transactions.filter(created_at__gte=date_from)
    if date_to:
        transactions = transactions.filter(created_at__lte=date_to)

    if date_from:
        opening_balance = account.transactions.filter(created_at__lt=date_from).aggregate(Sum('amount'))['amount__sum'] or Decimal('0.00')
    else:
        opening_balance = Decimal('0.00')

    summary = None
    if include_summary:
        summary = {
            'opening_balance': opening_balance,
            'total_charges':   transactions.filter(amount__lt=0).aggregate(Sum('amount'))['amount__sum'] or Decimal('0.00'),
            'total_payments':  transactions.filter(amount__gt=0).aggregate(Sum('amount'))['amount__sum'] or Decimal('0.00'),
            'closing_balance': account.current_balance,
        }

    FIELD_NAMES_FULL  = {'created_at': 'Date', 'transaction_type': 'Type', 'description': 'Description', 'reference_number': 'Reference', 'amount': 'Amount', 'balance_after': 'Balance', 'invoice': 'Invoice', 'payment': 'Payment', 'academic_session': 'Session'}
    FIELD_NAMES_SHORT = {'created_at': 'Date', 'transaction_type': 'Type', 'description': 'Description', 'reference_number': 'Ref.', 'amount': 'Amount', 'balance_after': 'Balance', 'invoice': 'Invoice', 'payment': 'Pmt.', 'academic_session': 'Session'}
    field_names       = FIELD_NAMES_SHORT if short_headers else FIELD_NAMES_FULL

    return render(request, 'fees/accounts/print_statement.html', {
        **get_print_school_context(request),
        'account':              account,
        'transactions':         transactions,
        'opening_balance':      opening_balance,
        'summary':              summary,
        'date_from':            date_from,
        'date_to':              date_to,
        'now':                  timezone.now(),
        'selected_fields':      selected_fields,
        'selected_field_names': [field_names.get(f, f.replace('_', ' ').title()) for f in selected_fields],
        'field_names':          field_names,
        'short_headers':        short_headers,
        'title':                f'Account Statement - {account.student.get_full_name()}',
    })


@login_required
def export_student_accounts_excel(request):
    ALL_COLUMNS = [
        ('admission_number', 'Admission No.',    lambda o: o.student.admission_number),
        ('student_name',     'Student Name',     lambda o: o.student.get_full_name()),
        ('academic_level',   'Level',            lambda o: str(o.student.current_academic_level) if o.student.current_academic_level else ''),
        ('current_balance',  'Current Balance',  lambda o: float(o.get_current_balance())),
        ('outstanding',      'Outstanding',      lambda o: float(o.get_outstanding_amount())),
        ('credit_balance',   'Credit Balance',   lambda o: float(o.get_credit_amount())),
        ('total_charges',    'Total Charged',    lambda o: float(o.get_total_charges())),
        ('total_payments',   'Total Paid',       lambda o: float(o.get_total_payments())),
        ('total_discounts',  'Total Discounts',  lambda o: float(o.get_total_discounts())),
        ('total_refunds',    'Total Refunds',    lambda o: float(o.get_total_refunds())),
        ('credit_limit',     'Credit Limit',     lambda o: float(o.credit_limit)),
        ('over_limit',       'Over Limit',       lambda o: 'Yes' if o.is_over_credit_limit() else 'No'),
        ('status',           'Account Status',   lambda o: o.get_status_display()),
        ('is_settled',       'Settled',          lambda o: 'Yes' if o.is_account_settled() else 'No'),
        ('last_transaction', 'Last Transaction', lambda o: o.last_transaction_date.strftime('%Y-%m-%d') if o.last_transaction_date else ''),
        ('last_payment',     'Last Payment',     lambda o: o.last_payment_date.strftime('%Y-%m-%d') if o.last_payment_date else ''),
    ]
    DEFAULT_FIELDS = ['admission_number', 'student_name', 'academic_level', 'current_balance', 'total_charges', 'total_payments', 'status']
    accounts = get_filtered_student_accounts(request)
    columns  = _resolve_columns(ALL_COLUMNS, request.GET.getlist('fields'), DEFAULT_FIELDS)
    return _xlsx_response(_make_workbook('Student Accounts', columns, accounts), 'student_accounts')


# =============================================================================
# ACCOUNT TRANSACTIONS
# =============================================================================

def get_filtered_account_transactions(request):
    transactions = AccountTransaction.objects.select_related(
        'student_account__student', 'invoice', 'payment', 'academic_session', 'fiscal_period'
    ).order_by('-created_at')

    query            = request.GET.get('q', '').strip()
    transaction_type = request.GET.get('transaction_type', '')
    student_account  = request.GET.get('student_account', '')
    academic_session = request.GET.get('academic_session', '')
    fiscal_period    = request.GET.get('fiscal_period', '')
    start_date       = request.GET.get('start_date', '')
    end_date         = request.GET.get('end_date', '')
    min_amount       = request.GET.get('min_amount', '')
    max_amount       = request.GET.get('max_amount', '')

    if query:
        words = query.split()
        q = Q()
        for w in words:
            q &= (Q(description__icontains=w) | Q(reference_number__icontains=w) |
                  Q(student_account__student__first_name__icontains=w) |
                  Q(student_account__student__last_name__icontains=w))
        transactions = transactions.filter(q)

    if transaction_type:
        transactions = transactions.filter(transaction_type=transaction_type)
    if student_account:
        transactions = transactions.filter(student_account_id=student_account)
    if academic_session:
        transactions = transactions.filter(academic_session_id=academic_session)
    if fiscal_period:
        transactions = transactions.filter(fiscal_period_id=fiscal_period)
    if start_date:
        transactions = transactions.filter(created_at__gte=start_date)
    if end_date:
        transactions = transactions.filter(created_at__lte=end_date)
    if min_amount:
        try:
            transactions = transactions.filter(amount__gte=Decimal(min_amount))
        except (ValueError, InvalidOperation):
            pass
    if max_amount:
        try:
            transactions = transactions.filter(amount__lte=Decimal(max_amount))
        except (ValueError, InvalidOperation):
            pass

    return transactions

@login_required
def transaction_detail(request, pk):
    transaction_obj = get_object_or_404(
        AccountTransaction.objects.select_related('student_account__student', 'invoice', 'payment', 'academic_session'),
        pk=pk,
    )
    return render(request, 'fees/transactions/detail.html', {'transaction': transaction_obj})


@login_required
def transaction_list_print_view(request):
    FIELD_NAMES_FULL = {
        'date':             'Date',
        'student_id':       'Admission Number',
        'student_name':     'Student Name',
        'transaction_type': 'Transaction Type',
        'amount':           'Amount',
        'balance_after':    'Balance After',
        'description':      'Description',
        'reference_number': 'Reference Number',
        'invoice':          'Invoice',
        'payment':          'Payment',
        'academic_session': 'Academic Session',
        'fiscal_period':    'Fiscal Period',
    }
    FIELD_NAMES_SHORT = {
        'date':             'Date',
        'student_id':       'Adm. No.',
        'student_name':     'Student',
        'transaction_type': 'Type',
        'amount':           'Amount',
        'balance_after':    'Balance',
        'description':      'Description',
        'reference_number': 'Ref.',
        'invoice':          'Invoice',
        'payment':          'Pmt.',
        'academic_session': 'Session',
        'fiscal_period':    'Period',
    }
    DEFAULT_FIELDS  = ['date', 'student_id', 'student_name', 'transaction_type', 'amount', 'balance_after', 'description', 'reference_number']
    selected_fields = request.GET.getlist('fields') or DEFAULT_FIELDS
    short_headers   = request.GET.get('short_headers', 'false').lower() == 'true'
    landscape       = request.GET.get('landscape', 'true').lower() == 'true'
    include_stats   = request.GET.get('include_stats', 'true').lower() == 'true'
    field_names     = FIELD_NAMES_SHORT if short_headers else FIELD_NAMES_FULL
    transactions    = get_filtered_account_transactions(request)

    stats = None
    if include_stats:
        stats = {
            'total':            transactions.count(),
            'total_credits':    transactions.filter(amount__gt=0).aggregate(v=Sum('amount'))['v'] or Decimal('0.00'),
            'total_debits':     abs(transactions.filter(amount__lt=0).aggregate(v=Sum('amount'))['v'] or Decimal('0.00')),
            'payment_count':    transactions.filter(transaction_type='PAYMENT').count(),
            'invoice_count':    transactions.filter(transaction_type='INVOICE').count(),
            'adjustment_count': transactions.filter(transaction_type='ADJUSTMENT').count(),
        }

    if transactions.count() > MAX_PRINT_RECORDS:
        messages.warning(request, f'Only the first {MAX_PRINT_RECORDS} transactions will be printed.')
        transactions = transactions[:MAX_PRINT_RECORDS]

    return render(request, 'fees/transactions/print_list.html', {
        **get_print_school_context(request),
        'transactions':         transactions,
        'stats':                stats,
        'selected_fields':      selected_fields,
        'selected_field_names': [field_names.get(f, f.replace('_', ' ').title()) for f in selected_fields],
        'field_names':          field_names,
        'short_headers':        short_headers,
        'landscape':            landscape,
        'now':                  timezone.now(),
        'print_date':           get_school_today(),
        'printed_by':           request.user.get_full_name() or request.user.username,
        'title':                'Account Transactions',
    })


@login_required
def export_account_transactions_excel(request):
    ALL_COLUMNS = [
        ('date',             'Date',             lambda o: o.created_at.strftime('%Y-%m-%d %H:%M') if o.created_at else ''),
        ('admission_number', 'Admission No.',    lambda o: o.student_account.student.admission_number),
        ('student_name',     'Student Name',     lambda o: o.student_account.student.get_full_name()),
        ('transaction_type', 'Type',             lambda o: o.get_transaction_type_display()),
        ('amount',           'Amount',           lambda o: float(o.amount)),
        ('balance_after',    'Balance After',    lambda o: float(o.balance_after)),
        ('description',      'Description',      lambda o: o.description or ''),
        ('reference_number', 'Reference',        lambda o: o.reference_number or ''),
        ('invoice',          'Invoice',          lambda o: o.invoice.invoice_number if o.invoice else ''),
        ('payment',          'Payment',          lambda o: o.payment.payment_number if o.payment else ''),
        ('academic_session', 'Academic Session', lambda o: o.academic_session.name if o.academic_session else ''),
        ('fiscal_period',    'Fiscal Period',    lambda o: o.fiscal_period.name if o.fiscal_period else ''),
    ]
    DEFAULT_FIELDS = ['date', 'admission_number', 'student_name', 'transaction_type', 'amount', 'balance_after', 'description', 'reference_number']
    transactions = get_filtered_account_transactions(request)
    columns      = _resolve_columns(ALL_COLUMNS, request.GET.getlist('fields'), DEFAULT_FIELDS)
    return _xlsx_response(_make_workbook('Transactions', columns, transactions), 'account_transactions')


# =============================================================================
# FEE INVOICES
# =============================================================================

@login_required
def invoice_list(request):
    from django.db.models import Count, Sum, Q, Prefetch
    from academics.models import StudentClassEnrollment

    filter_form = FeeInvoiceFilterForm(request.GET or None)
    today       = get_school_today()
    is_htmx     = request.headers.get('HX-Request') == 'true'

    # ── Lean base queryset — filters only, no joins or annotations ────
    base_qs = get_filtered_fee_invoices(request)

    # ── Stats — one query on the lean queryset ────────────────────────
    stats_agg = base_qs.aggregate(
        total          = Count('pk'),
        draft          = Count('pk', filter=Q(status='DRAFT')),
        pending        = Count('pk', filter=Q(status='PENDING')),
        partially_paid = Count('pk', filter=Q(status='PARTIALLY_PAID')),
        paid           = Count('pk', filter=Q(status='PAID')),
        overdue        = Count('pk', filter=Q(
                             due_date__lt=today,
                             status__in=['PENDING', 'PARTIALLY_PAID', 'OVERDUE'],
                         )),
        total_amount   = Sum('total_amount'),
        total_paid     = Sum('paid_amount'),
        total_balance  = Sum('balance'),
    )
    stats = {k: v or 0 for k, v in stats_agg.items()}

    # ── Page queryset — heavy select_related + prefetch, no annotations ─
    page_qs = (
        base_qs
        .select_related(
            'student',
            'academic_session',
            'fiscal_period',
            'fee_structure',
        )
        .prefetch_related(
            Prefetch(
                'student__class_enrollments',
                queryset=StudentClassEnrollment.objects.select_related(
                    'class_instance__academic_level',
                ).filter(is_active=True),
                to_attr='prefetched_class_enrollments',
            ),
        )
        .order_by('-issue_date', '-created_at')
    )

    # ── Paginator — inject fast count from lean queryset ──────────────
    paginator         = Paginator(page_qs, 20)
    paginator._count  = base_qs.count()
    invoices_page     = paginator.get_page(request.GET.get('page', 1))

    context = {
        'invoices_page': invoices_page,
        'paginator':     paginator,
        'stats':         stats,
        'filter_form':   filter_form,
        'is_htmx':       is_htmx,
        'today':         today,
    }

    if is_htmx:
        return render(request, 'fees/invoices/partials/_invoice_results.html', context)
    return render(request, 'fees/invoices/list.html', context)


def get_filtered_fee_invoices(request):
    """
    Lean filtered queryset — WHERE clauses only.
    No select_related, no prefetch_related, no annotations.
    Used for stats aggregation and as base for the page queryset.
    """
    from decimal import Decimal

    invoices = FeeInvoice.objects.all()

    query            = request.GET.get('q', '').strip()
    status           = request.GET.get('status', '')
    academic_session = request.GET.get('academic_session', '')
    fiscal_period    = request.GET.get('fiscal_period', '')
    student          = request.GET.get('student', '')
    fee_structure    = request.GET.get('fee_structure', '')
    has_scholarships = request.GET.get('has_scholarships', '')
    has_discounts    = request.GET.get('has_discounts', '')
    has_any_reduction= request.GET.get('has_any_reduction', '')
    is_overdue       = request.GET.get('is_overdue', '')
    issue_date_from  = request.GET.get('issue_date_from', '')
    issue_date_to    = request.GET.get('issue_date_to', '')
    due_date_from    = request.GET.get('due_date_from', '')
    due_date_to      = request.GET.get('due_date_to', '')
    min_amount       = request.GET.get('min_amount', '')
    max_amount       = request.GET.get('max_amount', '')

    if query:
        words = query.split()
        q = Q()
        for w in words:
            q &= (
                Q(invoice_number__icontains=w) |
                Q(student__first_name__icontains=w) |
                Q(student__middle_name__icontains=w) |
                Q(student__last_name__icontains=w) |
                Q(student__admission_number__icontains=w) |
                Q(notes__icontains=w)
            )
        invoices = invoices.filter(q)

    if status:
        invoices = invoices.filter(status=status)

    if academic_session:
        invoices = invoices.filter(academic_session_id=academic_session)

    if fiscal_period:
        invoices = invoices.filter(fiscal_period_id=fiscal_period)

    if student:
        invoices = invoices.filter(student_id=student)

    if fee_structure:
        invoices = invoices.filter(fee_structure_id=fee_structure)

    if has_scholarships and has_scholarships.lower() == 'true':
        invoices = invoices.filter(has_scholarships_applied=True)

    if has_discounts and has_discounts.lower() == 'true':
        invoices = invoices.filter(has_discounts_applied=True)

    if has_any_reduction and has_any_reduction.lower() == 'true':
        invoices = invoices.filter(
            Q(has_scholarships_applied=True) | Q(has_discounts_applied=True)
        )

    if is_overdue and is_overdue.lower() == 'true':
        today    = get_school_today()
        invoices = invoices.filter(
            due_date__lt=today,
            status__in=['PENDING', 'PARTIALLY_PAID', 'OVERDUE'],
        )

    if issue_date_from:
        invoices = invoices.filter(issue_date__gte=issue_date_from)

    if issue_date_to:
        invoices = invoices.filter(issue_date__lte=issue_date_to)

    if due_date_from:
        invoices = invoices.filter(due_date__gte=due_date_from)

    if due_date_to:
        invoices = invoices.filter(due_date__lte=due_date_to)

    if min_amount:
        try:
            invoices = invoices.filter(total_amount__gte=Decimal(min_amount))
        except (ValueError, TypeError):
            pass

    if max_amount:
        try:
            invoices = invoices.filter(total_amount__lte=Decimal(max_amount))
        except (ValueError, TypeError):
            pass

    return invoices

@login_required
def invoice_list(request):
    from django.db.models import Count, Sum, Q, Prefetch
    from academics.models import StudentClassEnrollment

    filter_form = FeeInvoiceFilterForm(request.GET or None)
    today       = get_school_today()
    is_htmx     = request.headers.get('HX-Request') == 'true'

    # Lean base queryset — filters only, no joins or annotations
    base_qs = get_filtered_fee_invoices(request)

    # Stats — one aggregate query on the lean queryset
    stats_agg = base_qs.aggregate(
        total          = Count('pk'),
        draft          = Count('pk', filter=Q(status='DRAFT')),
        pending        = Count('pk', filter=Q(status='PENDING')),
        partially_paid = Count('pk', filter=Q(status='PARTIALLY_PAID')),
        paid           = Count('pk', filter=Q(status='PAID')),
        overdue        = Count('pk', filter=Q(
                             due_date__lt=today,
                             status__in=['PENDING', 'PARTIALLY_PAID', 'OVERDUE'],
                         )),
        total_amount   = Sum('total_amount'),
        total_paid     = Sum('paid_amount'),
        total_balance  = Sum('balance'),
    )
    stats = {k: v or 0 for k, v in stats_agg.items()}

    # Page queryset — joins and prefetches applied here only
    # No annotations (item_count/payment_count removed — they caused 1.9s page fetch)
    page_qs = (
        base_qs
        .select_related(
            'student',
            'academic_session',
            'fiscal_period',
            'fee_structure',
        )
        .prefetch_related(
            Prefetch(
                'student__class_enrollments',
                queryset=StudentClassEnrollment.objects.select_related(
                    'class_instance__academic_level',
                ).filter(is_active=True),
                to_attr='prefetched_class_enrollments',
            ),
        )
        .order_by('-issue_date', '-created_at')
    )

    # Inject fast count from lean queryset — avoids COUNT on joined queryset
    paginator        = Paginator(page_qs, 20)
    paginator._count = base_qs.count()
    invoices_page    = paginator.get_page(request.GET.get('page', 1))

    context = {
        'invoices_page': invoices_page,
        'paginator':     paginator,
        'stats':         stats,
        'filter_form':   filter_form,
        'is_htmx':       is_htmx,
        'today':         today,
    }

    if is_htmx:
        return render(request, 'fees/invoices/partials/_invoice_results.html', context)
    return render(request, 'fees/invoices/list.html', context)


@login_required
def invoice_detail(request, pk):
    invoice  = get_object_or_404(
        FeeInvoice.objects.select_related('student', 'academic_session', 'fiscal_period', 'fee_structure').prefetch_related('items__fee_category'),
        pk=pk,
    )
    payments          = invoice.payments.select_related('payment_method').order_by('-payment_date')
    today             = get_school_today()
    is_overdue        = invoice.due_date < today and invoice.status in ['PENDING', 'PARTIALLY_PAID', 'OVERDUE']
    organized_items   = get_invoice_items_organized(invoice)
    status_color      = get_invoice_status_color(invoice.status)

    payment_progress = {
        'paid_percentage':  round((invoice.paid_amount / invoice.total_amount * 100), 1) if invoice.total_amount > 0 else 0,
        'is_overdue':       is_overdue,
        'days_overdue':     (today - invoice.due_date).days if is_overdue else 0,
        'days_until_due':   (invoice.due_date - today).days if not is_overdue else 0,
    }

    from uniforms.models import UniformSale
    uniform_sale = UniformSale.objects.filter(fee_invoice=invoice).prefetch_related('items__uniform_item', 'items__size').first()

    return render(request, 'fees/invoices/detail.html', {
        'invoice':          invoice,
        'payments':         payments,
        'payment_progress': payment_progress,
        'organized_items':  organized_items,
        'status_color':     status_color,
        'uniform_sale':     uniform_sale,
    })


@login_required
@require_http_methods(["POST"])
def invoice_delete(request, pk):
    invoice = get_object_or_404(FeeInvoice, pk=pk)
    if invoice.status in ['PAID', 'PARTIALLY_PAID'] or invoice.payments.exists():
        is_htmx = request.headers.get('HX-Request') == 'true'
        if is_htmx:
            r = HttpResponse()
            r['HX-Alert-Message'] = f"Cannot delete invoice '{invoice.invoice_number}' — it has payments"
            r['HX-Alert-Type']    = 'error'
            r['HX-Alert-Title']   = 'Cannot Delete'
            r['HX-Close-Modal']   = 'true'
            return r
        messages.error(request, f"Cannot delete invoice '{invoice.invoice_number}' — it has payments", extra_tags='sweetalert-error')
        return redirect('fees:invoice_detail', pk=pk)

    number = invoice.invoice_number
    invoice.delete()
    is_htmx = request.headers.get('HX-Request') == 'true'
    if is_htmx:
        r = HttpResponse()
        r['HX-Alert-Message'] = f"Invoice '{number}' deleted successfully"
        r['HX-Alert-Type']    = 'success'
        r['HX-Alert-Title']   = 'Deleted!'
        r['HX-Close-Modal']   = 'true'
        r['HX-Redirect']      = reverse('fees:invoice_list')
        return r
    messages.success(request, f"Invoice '{number}' deleted successfully", extra_tags='sweetalert')
    return redirect('fees:invoice_list')


@login_required
@require_http_methods(["POST"])
def invoice_void(request, pk):
    invoice = get_object_or_404(FeeInvoice, pk=pk)
    if invoice.status in ['VOID', 'CANCELLED']:
        messages.warning(request, 'Invoice is already voided/cancelled.')
        return redirect('fees:invoice_detail', pk=pk)
    if invoice.status == 'PAID':
        messages.error(request, 'Cannot void paid invoices. Use refund process instead.')
        return redirect('fees:invoice_detail', pk=pk)
    try:
        void_reason = request.POST.get('void_reason', '')
        with transaction.atomic():
            invoice.status = 'VOID'
            invoice.notes  = f"{invoice.notes}\n\nVOIDED on {get_school_today()}: {void_reason}".strip()
            invoice.save()
            for payment in invoice.payments.all():
                payment.reversed       = True
                payment.reversed_on    = get_school_current_time()
                payment.reversed_by_id = str(request.user.id)
                payment.reversal_reason = f"Invoice voided: {void_reason}"
                payment.status          = 'REVERSED'
                payment.save()
            student_account, _ = StudentAccount.objects.get_or_create(student=invoice.student)
            AccountTransaction.objects.create(
                student_account=student_account,
                transaction_type='ADJUSTMENT',
                amount=invoice.total_amount,
                description=f"Void invoice {invoice.invoice_number}: {void_reason}",
                reference_number=f"VOID-{invoice.invoice_number}",
                balance_after=student_account.get_current_balance() + invoice.total_amount,
                invoice=invoice,
                academic_session=invoice.academic_session,
                fiscal_period=invoice.fiscal_period,
                processed_by_id=str(request.user.id),
            )
        is_htmx = request.headers.get('HX-Request') == 'true'
        if is_htmx:
            r = HttpResponse()
            r['HX-Alert-Message'] = f"Invoice {invoice.invoice_number} voided successfully!"
            r['HX-Alert-Type']    = 'success'
            r['HX-Alert-Title']   = 'Invoice Voided'
            r['HX-Close-Modal']   = 'true'
            r['HX-Redirect']      = reverse('fees:invoice_detail', kwargs={'pk': invoice.pk})
            return r
        messages.success(request, f"Invoice {invoice.invoice_number} voided successfully!")
        return redirect('fees:invoice_detail', pk=invoice.pk)
    except Exception as e:
        logger.error(f"Error voiding invoice: {e}", exc_info=True)
        is_htmx = request.headers.get('HX-Request') == 'true'
        if is_htmx:
            r = HttpResponse()
            r['HX-Alert-Message'] = f'Error voiding invoice: {str(e)}'
            r['HX-Alert-Type']    = 'error'
            r['HX-Alert-Title']   = 'Error'
            return r
        messages.error(request, f'Error voiding invoice: {str(e)}')
        return redirect('fees:invoice_detail', pk=pk)


@login_required
@require_http_methods(["POST"])
def invoice_send_email(request, pk):
    invoice = get_object_or_404(FeeInvoice, pk=pk)
    try:
        recipient_emails = request.POST.getlist('recipients')
        custom_message   = request.POST.get('custom_message', '')
        if not recipient_emails:
            raise ValueError("At least one recipient email is required")
        email_body = render_to_string('fees/emails/invoice_email.html', {
            'invoice': invoice, 'custom_message': custom_message,
            'school_name': getattr(settings, 'SCHOOL_NAME', 'School'),
        })
        send_mail(
            subject=f"Fee Invoice {invoice.invoice_number} - {invoice.student.get_full_name()}",
            message='', from_email=settings.DEFAULT_FROM_EMAIL,
            recipient_list=recipient_emails, html_message=email_body, fail_silently=False,
        )
        invoice.notes = f"{invoice.notes}\n\nInvoice emailed to {', '.join(recipient_emails)} on {get_school_current_time()}".strip()
        invoice.save()
        is_htmx = request.headers.get('HX-Request') == 'true'
        if is_htmx:
            r = HttpResponse()
            r['HX-Alert-Message'] = f"Invoice sent to {len(recipient_emails)} recipient(s) successfully!"
            r['HX-Alert-Type']    = 'success'
            r['HX-Alert-Title']   = 'Email Sent'
            r['HX-Close-Modal']   = 'true'
            r['HX-Redirect']      = reverse('fees:invoice_detail', kwargs={'pk': invoice.pk})
            return r
        messages.success(request, f"Invoice sent to {len(recipient_emails)} recipient(s) successfully!")
        return redirect('fees:invoice_detail', pk=invoice.pk)
    except Exception as e:
        logger.error(f"Error sending invoice email: {e}", exc_info=True)
        is_htmx = request.headers.get('HX-Request') == 'true'
        if is_htmx:
            r = HttpResponse()
            r['HX-Alert-Message'] = f'Error sending email: {str(e)}'
            r['HX-Alert-Type']    = 'error'
            r['HX-Alert-Title']   = 'Error'
            return r
        messages.error(request, f'Error sending email: {str(e)}')
        return redirect('fees:invoice_detail', pk=pk)


@login_required
@require_http_methods(["POST"])
def send_payment_reminder(request, pk):
    invoice = get_object_or_404(FeeInvoice, pk=pk)
    today   = get_school_today()
    if invoice.due_date >= today:
        messages.warning(request, 'Invoice is not yet overdue.')
        return redirect('fees:invoice_detail', pk=pk)
    try:
        recipient_emails = request.POST.getlist('recipients')
        reminder_type    = request.POST.get('reminder_type', 'FRIENDLY')
        if not recipient_emails:
            raise ValueError("At least one recipient email is required")
        days_overdue = (today - invoice.due_date).days
        email_body   = render_to_string('fees/emails/payment_reminder.html', {
            'invoice': invoice, 'days_overdue': days_overdue, 'reminder_type': reminder_type,
            'school_name': getattr(settings, 'SCHOOL_NAME', 'School'),
        })
        subject_map = {
            'FRIENDLY': f"Payment Reminder: Invoice {invoice.invoice_number}",
            'URGENT':   f"URGENT: Overdue Payment - Invoice {invoice.invoice_number}",
            'FINAL':    f"FINAL NOTICE: Payment Required - Invoice {invoice.invoice_number}",
        }
        send_mail(subject=subject_map.get(reminder_type, subject_map['FRIENDLY']), message='',
                  from_email=settings.DEFAULT_FROM_EMAIL, recipient_list=recipient_emails,
                  html_message=email_body, fail_silently=False)
        invoice.notes = f"{invoice.notes}\n\n{reminder_type} payment reminder sent to {', '.join(recipient_emails)} on {get_school_current_time()}".strip()
        invoice.save()
        is_htmx = request.headers.get('HX-Request') == 'true'
        if is_htmx:
            r = HttpResponse()
            r['HX-Alert-Message'] = f"Payment reminder sent to {len(recipient_emails)} recipient(s)!"
            r['HX-Alert-Type']    = 'success'
            r['HX-Alert-Title']   = 'Reminder Sent'
            r['HX-Close-Modal']   = 'true'
            r['HX-Redirect']      = reverse('fees:invoice_detail', kwargs={'pk': invoice.pk})
            return r
        messages.success(request, f"Payment reminder sent to {len(recipient_emails)} recipient(s)!")
        return redirect('fees:invoice_detail', pk=invoice.pk)
    except Exception as e:
        logger.error(f"Error sending payment reminder: {e}", exc_info=True)
        is_htmx = request.headers.get('HX-Request') == 'true'
        if is_htmx:
            r = HttpResponse()
            r['HX-Alert-Message'] = f'Error sending reminder: {str(e)}'
            r['HX-Alert-Type']    = 'error'
            r['HX-Alert-Title']   = 'Error'
            return r
        messages.error(request, f'Error sending reminder: {str(e)}')
        return redirect('fees:invoice_detail', pk=pk)


@login_required
@require_http_methods(["POST"])
def invoice_finalize(request, pk):
    invoice = get_object_or_404(FeeInvoice, pk=pk)

    is_htmx = request.headers.get('HX-Request') == 'true'

    def _error(title, msg, level='error'):
        if is_htmx:
            r = HttpResponse()
            r['HX-Alert-Message'] = msg
            r['HX-Alert-Type']    = level
            r['HX-Alert-Title']   = title
            return r
        getattr(messages, level)(request, msg)
        return redirect('fees:invoice_detail', pk=pk)

    if invoice.status != 'DRAFT':
        return _error('Cannot Finalize', f'Invoice is already {invoice.get_status_display()}', 'warning')

    try:
        with transaction.atomic():
            from finance.models import Journal, JournalEntry, JournalTransaction
            from core.models import FinancialSettings
            from finance.utils import generate_journal_entry_number
            from django.db.models import Sum

            # ── Zero-amount (100% scholarship) ────────────────────────────
            if invoice.total_amount <= Decimal('0.00'):
                invoice.status = 'PENDING'
                invoice.save(update_fields=['status'])
                msg = f"Invoice {invoice.invoice_number} finalized (zero amount — no journal entry)."
                if is_htmx:
                    r = HttpResponse()
                    r['HX-Alert-Message'] = msg
                    r['HX-Alert-Type']    = 'success'
                    r['HX-Alert-Title']   = 'Invoice Finalized'
                    r['HX-Close-Modal']   = 'true'
                    r['HX-Redirect']      = reverse('fees:invoice_detail', kwargs={'pk': invoice.pk})
                    return r
                messages.success(request, msg)
                return redirect('fees:invoice_detail', pk=invoice.pk)

            fin_settings = FinancialSettings.get_instance()
            if not fin_settings or not hasattr(fin_settings, 'account_mappings'):
                raise ValueError("Account mappings not configured.")

            mappings           = fin_settings.account_mappings
            receivable_account = mappings.student_receivables_account
            BOARDING_TYPES     = {'BOARDING', 'LAUNDRY'}

            fees_journal, _ = Journal.objects.get_or_create(
                journal_type='FEES',
                defaults={
                    'name':        'Fee Collection Journal',
                    'description': 'Student fee invoices and collections',
                    'is_active':   True,
                },
            )

            # ── Build / update journal entry ───────────────────────────────
            journal_entry_created = False
            if invoice.journal_entry:
                je = invoice.journal_entry
                je.entry_date       = invoice.issue_date
                je.fiscal_period    = invoice.fiscal_period
                je.academic_session = invoice.academic_session
                je.description      = f"Student Fee Invoice - {invoice.student.get_full_name()} (Updated)"
                je.save(update_fields=[
                    'entry_date', 'fiscal_period', 'academic_session', 'description',
                ])
                je.transactions.all().delete()
            else:
                je = JournalEntry.objects.create(
                    entry_number     = generate_journal_entry_number(fees_journal),
                    journal          = fees_journal,
                    entry_date       = invoice.issue_date,
                    fiscal_period    = invoice.fiscal_period,
                    academic_session = invoice.academic_session,
                    reference_number = invoice.invoice_number,
                    description      = f"Student Fee Invoice - {invoice.student.get_full_name()}",
                    status           = 'DRAFT',
                )
                invoice.journal_entry = je
                journal_entry_created = True

            # ── Debit: student receivable ──────────────────────────────────
            JournalTransaction.objects.create(
                journal_entry = je,
                account       = receivable_account,
                amount        = invoice.total_amount,
                is_debit      = True,
                description   = f"Student fees - {invoice.student.get_full_name()}",
            )

            # ── Credits: revenue by category type ─────────────────────────
            revenue_breakdown = (
                invoice.items
                .values('fee_category__category_type', 'fee_category__code')
                .annotate(total_amount=Sum('final_amount'))
                .order_by('fee_category__category_type')
            )
            positive_rows  = [
                row for row in revenue_breakdown
                if (row['total_amount'] or Decimal('0.00')) > Decimal('0.00')
            ]
            total_credited = Decimal('0.00')

            if not positive_rows:
                JournalTransaction.objects.create(
                    journal_entry = je,
                    account       = mappings.default_revenue_account,
                    amount        = invoice.total_amount,
                    is_debit      = False,
                    description   = (
                        f"Fee revenue - {invoice.academic_session.name}"
                        if invoice.academic_session else "Fee revenue"
                    ),
                )
                total_credited = invoice.total_amount
            else:
                for item in positive_rows:
                    cat_type = item['fee_category__category_type'] or ''
                    amount   = item['total_amount']
                    if cat_type in BOARDING_TYPES:
                        account     = mappings.boarding_revenue_account or mappings.default_revenue_account
                        description = "Boarding services revenue"
                    elif cat_type == 'UNIFORM':
                        account     = mappings.uniform_and_book_sales_account or mappings.default_revenue_account
                        description = "Uniform sales revenue"
                    else:
                        account     = mappings.default_revenue_account
                        description = (
                            f"{cat_type.replace('_', ' ').title()} revenue"
                            if cat_type else "Other revenue"
                        )
                    JournalTransaction.objects.create(
                        journal_entry = je,
                        account       = account,
                        amount        = amount,
                        is_debit      = False,
                        description   = description,
                    )
                    total_credited += amount

                remainder = invoice.total_amount - total_credited
                if remainder > Decimal('0.00'):
                    JournalTransaction.objects.create(
                        journal_entry = je,
                        account       = mappings.default_revenue_account,
                        amount        = remainder,
                        is_debit      = False,
                        description   = "Fee revenue (rounding adjustment)",
                    )

            # ── Save invoice + post JE ─────────────────────────────────────
            # Signal handle_invoice_status_transition fires on this save and
            # creates the INVOICE AccountTransaction automatically.
            invoice.status = 'PENDING'
            invoice.save(update_fields=['status', 'journal_entry'])

            je.status       = 'POSTED'
            je.posted_at    = get_school_current_time()
            je.posted_by_id = str(request.user.id)
            je.save(update_fields=['status', 'posted_at', 'posted_by_id'])

            action            = "created and posted" if journal_entry_created else "updated and posted"
            transaction_count = je.transactions.count()
            success_msg       = (
                f"Invoice {invoice.invoice_number} finalized. "
                f"Journal entry {je.entry_number} {action} "
                f"({transaction_count} transactions)."
            )

            if is_htmx:
                r = HttpResponse()
                r['HX-Alert-Message'] = success_msg
                r['HX-Alert-Type']    = 'success'
                r['HX-Alert-Title']   = 'Invoice Finalized'
                r['HX-Close-Modal']   = 'true'
                r['HX-Redirect']      = reverse('fees:invoice_detail', kwargs={'pk': invoice.pk})
                return r
            messages.success(request, success_msg)
            return redirect('fees:invoice_detail', pk=invoice.pk)

    except Exception as e:
        logger.error(f"Error finalizing invoice: {e}", exc_info=True)
        return _error('Error', f'Error finalizing invoice: {str(e)}')


@login_required
@require_http_methods(["POST"])
def invoice_bulk_finalize(request):
    from finance.models import Journal, JournalEntry, JournalTransaction
    from core.models import FinancialSettings
    from finance.utils import generate_journal_entry_number
    from django.db.models import Sum

    is_htmx = request.headers.get('HX-Request') == 'true'

    def _htmx_error(title, msg):
        r = HttpResponse()
        r['HX-Alert-Message'] = msg
        r['HX-Alert-Type']    = 'error'
        r['HX-Alert-Title']   = title
        return r

    # ── Resolve PKs ───────────────────────────────────────────────────────
    if request.POST.get('all_draft_pages'):
        pks = list(
            FeeInvoice.objects.filter(status='DRAFT').values_list('pk', flat=True)
        )
    else:
        pks = request.POST.getlist('pk')

    if not pks:
        msg = "No invoices provided."
        if is_htmx:
            return _htmx_error('Bulk Finalize', msg)
        messages.error(request, msg)
        return redirect('fees:invoice_list')

    fin_settings = FinancialSettings.get_instance()
    if not fin_settings or not hasattr(fin_settings, 'account_mappings'):
        msg = "Account mappings not configured — cannot finalize invoices."
        if is_htmx:
            return _htmx_error('Configuration Error', msg)
        messages.error(request, msg)
        return redirect('fees:invoice_list')

    mappings           = fin_settings.account_mappings
    receivable_account = mappings.student_receivables_account
    BOARDING_TYPES     = {'BOARDING', 'LAUNDRY'}
    user_id_str        = str(request.user.id)

    fees_journal, _ = Journal.objects.get_or_create(
        journal_type='FEES',
        defaults={
            'name':        'Fee Collection Journal',
            'description': 'Student fee invoices and collections',
            'is_active':   True,
        },
    )

    invoices = (
        FeeInvoice.objects
        .filter(pk__in=pks, status='DRAFT')
        .select_related(
            'student', 'academic_session', 'fiscal_period', 'journal_entry',
        )
        .prefetch_related('items__fee_category')
    )

    succeeded = []
    failed    = []

    for invoice in invoices:
        try:
            with transaction.atomic():

                # ── Zero-amount ───────────────────────────────────────────
                if invoice.total_amount <= Decimal('0.00'):
                    invoice.status = 'PENDING'
                    invoice.save(update_fields=['status'])
                    succeeded.append({'invoice': invoice, 'je': None})
                    continue

                # ── Build / update journal entry ──────────────────────────
                journal_entry_created = False
                if invoice.journal_entry:
                    je = invoice.journal_entry
                    je.entry_date       = invoice.issue_date
                    je.fiscal_period    = invoice.fiscal_period
                    je.academic_session = invoice.academic_session
                    je.description      = f"Student Fee Invoice - {invoice.student.get_full_name()} (Updated)"
                    je.save(update_fields=[
                        'entry_date', 'fiscal_period', 'academic_session', 'description',
                    ])
                    je.transactions.all().delete()
                else:
                    je = JournalEntry.objects.create(
                        entry_number     = generate_journal_entry_number(fees_journal),
                        journal          = fees_journal,
                        entry_date       = invoice.issue_date,
                        fiscal_period    = invoice.fiscal_period,
                        academic_session = invoice.academic_session,
                        reference_number = invoice.invoice_number,
                        description      = f"Student Fee Invoice - {invoice.student.get_full_name()}",
                        status           = 'DRAFT',
                    )
                    invoice.journal_entry = je
                    journal_entry_created = True

                # ── Debit: student receivable ─────────────────────────────
                JournalTransaction.objects.create(
                    journal_entry = je,
                    account       = receivable_account,
                    amount        = invoice.total_amount,
                    is_debit      = True,
                    description   = f"Student fees - {invoice.student.get_full_name()}",
                )

                # ── Credits: revenue by category type ─────────────────────
                revenue_breakdown = (
                    invoice.items
                    .values('fee_category__category_type', 'fee_category__code')
                    .annotate(total_amount=Sum('final_amount'))
                    .order_by('fee_category__category_type')
                )
                total_credited = Decimal('0.00')
                positive_rows  = [
                    row for row in revenue_breakdown
                    if (row['total_amount'] or Decimal('0.00')) > Decimal('0.00')
                ]

                if not positive_rows:
                    JournalTransaction.objects.create(
                        journal_entry = je,
                        account       = mappings.default_revenue_account,
                        amount        = invoice.total_amount,
                        is_debit      = False,
                        description   = (
                            f"Fee revenue - {invoice.academic_session.name}"
                            if invoice.academic_session else "Fee revenue"
                        ),
                    )
                    total_credited = invoice.total_amount
                else:
                    for row in positive_rows:
                        cat_type = row['fee_category__category_type'] or ''
                        amount   = row['total_amount']
                        if cat_type in BOARDING_TYPES:
                            account     = mappings.boarding_revenue_account or mappings.default_revenue_account
                            description = "Boarding services revenue"
                        elif cat_type == 'UNIFORM':
                            account     = mappings.uniform_and_book_sales_account or mappings.default_revenue_account
                            description = "Uniform sales revenue"
                        else:
                            account     = mappings.default_revenue_account
                            description = (
                                f"{cat_type.replace('_', ' ').title()} revenue"
                                if cat_type else "Other revenue"
                            )
                        JournalTransaction.objects.create(
                            journal_entry = je,
                            account       = account,
                            amount        = amount,
                            is_debit      = False,
                            description   = description,
                        )
                        total_credited += amount

                    remainder = invoice.total_amount - total_credited
                    if remainder > Decimal('0.00'):
                        JournalTransaction.objects.create(
                            journal_entry = je,
                            account       = mappings.default_revenue_account,
                            amount        = remainder,
                            is_debit      = False,
                            description   = "Fee revenue (rounding adjustment)",
                        )

                # ── Save invoice + post JE ────────────────────────────────
                # Signal handle_invoice_status_transition fires on this save
                # and creates the INVOICE AccountTransaction automatically.
                invoice.status = 'PENDING'
                invoice.save(update_fields=['status', 'journal_entry'])

                je.status       = 'POSTED'
                je.posted_at    = get_school_current_time()
                je.posted_by_id = user_id_str
                je.save(update_fields=['status', 'posted_at', 'posted_by_id'])

                succeeded.append({
                    'invoice': invoice,
                    'je':      je,
                    'created': journal_entry_created,
                })

        except Exception as e:
            logger.error(
                f"Bulk finalize: error on invoice {invoice.invoice_number}: {e}",
                exc_info=True,
            )
            failed.append({'invoice': invoice, 'error': str(e)})

    # ── Build response ────────────────────────────────────────────────────
    total     = len(succeeded) + len(failed)
    ok_count  = len(succeeded)
    err_count = len(failed)

    if err_count == 0:
        alert_type = 'success'
        alert_msg  = f"Finalized {ok_count} of {total} invoice(s) successfully."
    elif ok_count == 0:
        alert_type = 'error'
        alert_msg  = f"All {total} invoice(s) failed to finalize."
    else:
        alert_type = 'warning'
        alert_msg  = f"Finalized {ok_count} invoice(s); {err_count} failed."

    if is_htmx:
        r = HttpResponse()
        r['HX-Alert-Message'] = alert_msg
        r['HX-Alert-Type']    = alert_type
        r['HX-Alert-Title']   = 'Bulk Finalize'
        r['HX-Close-Modal']   = 'true'
        r['HX-Redirect']      = reverse('fees:invoice_list')
        return r

    if err_count == 0:
        messages.success(request, alert_msg)
    elif ok_count == 0:
        messages.error(request, alert_msg)
    else:
        messages.warning(request, alert_msg)
    return redirect('fees:invoice_list')


@login_required
@require_http_methods(["POST"])
def invoice_revert_to_draft(request, pk):
    invoice = get_object_or_404(FeeInvoice, pk=pk)

    is_htmx = request.headers.get('HX-Request') == 'true'

    def _error(title, msg, level='error'):
        if is_htmx:
            r = HttpResponse()
            r['HX-Alert-Message'] = msg
            r['HX-Alert-Type']    = level
            r['HX-Alert-Title']   = title
            return r
        getattr(messages, level)(request, msg)
        return redirect('fees:invoice_detail', pk=pk)

    if invoice.status != 'PENDING':
        return _error('Cannot Revert', f'Invoice is {invoice.get_status_display()}, not PENDING', 'warning')

    if invoice.paid_amount > 0:
        return _error('Cannot Revert', f'Cannot revert — invoice has payments ({invoice.paid_amount:,.2f})')

    if invoice.fiscal_period and invoice.fiscal_period.is_closed:
        return _error('Cannot Revert', f'Cannot revert — fiscal period {invoice.fiscal_period.name} is closed')

    try:
        with transaction.atomic():
            # ── Revert invoice ─────────────────────────────────────────────
            # Signal handle_invoice_status_transition fires on this save and
            # handles: un-posting the JE + creating ADJUSTMENT AccountTransaction.
            invoice.status = 'DRAFT'
            reverted_by    = request.user.get_full_name() or str(request.user)
            invoice.notes  = f"{invoice.notes}\n\nReverted to DRAFT by {reverted_by} on {get_school_today()}".strip()
            invoice.save()

        success_msg = f'Invoice {invoice.invoice_number} reverted to DRAFT. Journal entry un-posted.'
        if is_htmx:
            r = HttpResponse()
            r['HX-Alert-Message'] = success_msg
            r['HX-Alert-Type']    = 'success'
            r['HX-Alert-Title']   = 'Invoice Reverted'
            r['HX-Close-Modal']   = 'true'
            r['HX-Redirect']      = reverse('fees:invoice_detail', kwargs={'pk': invoice.pk})
            return r
        messages.success(request, success_msg)
        return redirect('fees:invoice_detail', pk=invoice.pk)

    except Exception as e:
        logger.error(f"Error reverting invoice: {e}", exc_info=True)
        return _error('Error', f'Error reverting invoice: {str(e)}')


@login_required
def invoice_print_view(request, pk):
    invoice = get_object_or_404(
        FeeInvoice.objects.select_related('student', 'academic_session', 'fiscal_period', 'fee_structure').prefetch_related('items__fee_category'),
        pk=pk,
    )
    return render(request, 'fees/invoices/print_invoice.html', {
        **get_print_school_context(request),
        'invoice': invoice, 'now': timezone.now(),
        'title':   f'Invoice {invoice.invoice_number}',
    })


@login_required
def invoice_list_print_view(request):
    FIELD_NAMES_FULL = {
        'invoice_number':     'Invoice Number',
        'student_id':         'Admission Number',
        'student_name':       'Student Name',
        'student_class':      'Class',
        'academic_session':   'Academic Session',
        'fiscal_period':      'Fiscal Period',
        'issue_date':         'Issue Date',
        'due_date':           'Due Date',
        'subtotal_amount':    'Subtotal',
        'discount_amount':    'Discounts',
        'scholarship_amount': 'Scholarship Discount',
        'tax_amount':         'Tax',
        'total_amount':       'Total Amount',
        'paid_amount':        'Paid Amount',
        'balance':            'Balance',
        'late_fee_amount':    'Late Fees',
        'status':             'Status',
        'has_scholarships':   'Has Scholarships',
        'has_discounts':      'Has Discounts',
        'notes':              'Notes',
    }
    FIELD_NAMES_SHORT = {
        'invoice_number':     'Inv. No.',
        'student_id':         'Adm. No.',
        'student_name':       'Student',
        'student_class':      'Class',
        'academic_session':   'Session',
        'fiscal_period':      'Period',
        'issue_date':         'Issued',
        'due_date':           'Due',
        'subtotal_amount':    'Subtotal',
        'discount_amount':    'Disc.',
        'scholarship_amount': 'Scholar.',
        'tax_amount':         'Tax',
        'total_amount':       'Total',
        'paid_amount':        'Paid',
        'balance':            'Balance',
        'late_fee_amount':    'Late',
        'status':             'Status',
        'has_scholarships':   'Scholar.',
        'has_discounts':      'Disc.',
        'notes':              'Notes',
    }
    DEFAULT_FIELDS  = ['invoice_number', 'student_id', 'student_name', 'student_class', 'academic_session', 'issue_date', 'due_date', 'total_amount', 'paid_amount', 'balance', 'status']
    selected_fields = request.GET.getlist('fields') or DEFAULT_FIELDS
    short_headers   = request.GET.get('short_headers', 'false').lower() == 'true'
    landscape       = request.GET.get('landscape', 'true').lower() == 'true'
    include_stats   = request.GET.get('include_stats', 'true').lower() == 'true'
    field_names     = FIELD_NAMES_SHORT if short_headers else FIELD_NAMES_FULL
    invoices        = get_filtered_fee_invoices(request)
    today           = get_school_today()

    stats = None
    if include_stats:
        stats = {
            'total':          invoices.count(),
            'pending':        invoices.filter(status='PENDING').count(),
            'paid':           invoices.filter(status='PAID').count(),
            'partially_paid': invoices.filter(status='PARTIALLY_PAID').count(),
            'overdue':        invoices.filter(due_date__lt=today, status__in=['PENDING', 'PARTIALLY_PAID', 'OVERDUE']).count(),
            'total_amount':   invoices.aggregate(v=Sum('total_amount'))['v'] or Decimal('0.00'),
            'total_paid':     invoices.aggregate(v=Sum('paid_amount'))['v'] or Decimal('0.00'),
            'total_balance':  invoices.aggregate(v=Sum('balance'))['v'] or Decimal('0.00'),
        }

    if invoices.count() > MAX_PRINT_RECORDS:
        messages.warning(request, f'Only the first {MAX_PRINT_RECORDS} invoices will be printed.')
        invoices = invoices[:MAX_PRINT_RECORDS]

    return render(request, 'fees/invoices/print_invoice_list.html', {
        **get_print_school_context(request),
        'invoices':             invoices,
        'stats':                stats,
        'selected_fields':      selected_fields,
        'selected_field_names': [field_names.get(f, f.replace('_', ' ').title()) for f in selected_fields],
        'field_names':          field_names,
        'short_headers':        short_headers,
        'landscape':            landscape,
        'now':                  timezone.now(),
        'print_date':           today,
        'printed_by':           request.user.get_full_name() or request.user.username,
        'title':                'Invoice List',
    })


@login_required
def export_invoices_excel(request):
    ALL_COLUMNS = [
        ('invoice_number',     'Invoice Number',    lambda o: o.invoice_number),
        ('student_id',         'Admission No.',     lambda o: o.student.admission_number),
        ('student_name',       'Student Name',      lambda o: o.student.get_full_name()),
        ('student_class',      'Class',             lambda o: (o.get_student_class().name if o.get_student_class() else '')),
        ('academic_session',   'Academic Session',  lambda o: o.academic_session.name if o.academic_session else ''),
        ('fiscal_period',      'Fiscal Period',     lambda o: o.fiscal_period.name if o.fiscal_period else ''),
        ('fee_structure',      'Fee Structure',     lambda o: o.fee_structure.name if o.fee_structure else ''),
        ('issue_date',         'Issue Date',        lambda o: o.issue_date.strftime('%Y-%m-%d') if o.issue_date else ''),
        ('due_date',           'Due Date',          lambda o: o.due_date.strftime('%Y-%m-%d') if o.due_date else ''),
        ('subtotal_amount',    'Subtotal',          lambda o: float(o.subtotal_amount)),
        ('discount_amount',    'Discounts',         lambda o: float(o.discount_amount)),
        ('scholarship_amount', 'Scholarship Disc.', lambda o: float(o.scholarship_discount_amount)),
        ('tax_amount',         'Tax',               lambda o: float(o.tax_amount)),
        ('total_amount',       'Total Amount',      lambda o: float(o.total_amount)),
        ('paid_amount',        'Paid Amount',       lambda o: float(o.paid_amount)),
        ('balance',            'Balance',           lambda o: float(o.balance)),
        ('late_fee_amount',    'Late Fees',         lambda o: float(o.late_fee_amount)),
        ('status',             'Status',            lambda o: o.get_status_display()),
        ('has_scholarships',   'Has Scholarships',  lambda o: 'Yes' if o.has_scholarships_applied else 'No'),
        ('has_discounts',      'Has Discounts',     lambda o: 'Yes' if o.has_discounts_applied else 'No'),
        ('payment_terms',      'Payment Terms',     lambda o: o.payment_terms or ''),
        ('notes',              'Notes',             lambda o: o.notes or ''),
    ]
    DEFAULT_FIELDS = ['invoice_number', 'student_id', 'student_name', 'student_class', 'academic_session', 'issue_date', 'due_date', 'total_amount', 'paid_amount', 'balance', 'status']
    invoices = get_filtered_fee_invoices(request)
    columns  = _resolve_columns(ALL_COLUMNS, request.GET.getlist('fields'), DEFAULT_FIELDS)
    return _xlsx_response(_make_workbook('Invoices', columns, invoices), 'fee_invoices')


# =============================================================================
# PAYMENTS
# =============================================================================

def get_filtered_payments(request):
    payments = Payment.objects.select_related(
        'student', 'invoice', 'payment_method', 'academic_session', 'fiscal_period'
    ).order_by('-payment_date', '-created_at')

    query            = request.GET.get('q', '').strip()
    status           = request.GET.get('status', '')
    payment_method   = request.GET.get('payment_method', '')
    academic_session = request.GET.get('academic_session', '')
    fiscal_period    = request.GET.get('fiscal_period', '')
    student          = request.GET.get('student', '')
    invoice          = request.GET.get('invoice', '')
    is_verified      = request.GET.get('is_verified', '')
    start_date       = request.GET.get('start_date', '')
    end_date         = request.GET.get('end_date', '')
    min_amount       = request.GET.get('min_amount', '')
    max_amount       = request.GET.get('max_amount', '')
    payment_state    = request.GET.get('payment_state', '')

    if query:
        words = query.split()
        q = Q()
        for w in words:
            q &= (Q(payment_number__icontains=w) | Q(receipt_number__icontains=w) |
                  Q(reference_number__icontains=w) | Q(transaction_id__icontains=w) |
                  Q(student__first_name__icontains=w) | Q(student__last_name__icontains=w) |
                  Q(paid_by_name__icontains=w))
        payments = payments.filter(q)

    if status:
        payments = payments.filter(status=status)
    if payment_method:
        payments = payments.filter(payment_method_id=payment_method)
    if academic_session:
        payments = payments.filter(academic_session_id=academic_session)
    if fiscal_period:
        payments = payments.filter(fiscal_period_id=fiscal_period)
    if student:
        payments = payments.filter(student_id=student)
    if invoice:
        payments = payments.filter(invoice_id=invoice)
    if is_verified:
        payments = payments.filter(is_verified=(is_verified.lower() == 'true'))
    if start_date:
        payments = payments.filter(payment_date__gte=start_date)
    if end_date:
        payments = payments.filter(payment_date__lte=end_date)
    if min_amount:
        try:
            payments = payments.filter(amount__gte=Decimal(min_amount))
        except (ValueError, TypeError):
            pass
    if max_amount:
        try:
            payments = payments.filter(amount__lte=Decimal(max_amount))
        except (ValueError, TypeError):
            pass
    if payment_state == 'active':
        payments = payments.filter(reversed=False, refunded=False)
    elif payment_state == 'reversed':
        payments = payments.filter(reversed=True)
    elif payment_state == 'refunded':
        payments = payments.filter(refunded=True)
    elif payment_state == 'inactive':
        payments = payments.filter(Q(reversed=True) | Q(refunded=True))

    return payments


@login_required
def payment_list(request):
    filter_form = PaymentFilterForm(request.GET or None)
    payments    = get_filtered_payments(request)

    stats = {
        'total':        payments.count(),
        'completed':    payments.filter(status='COMPLETED').count(),
        'pending':      payments.filter(status='PENDING').count(),
        'failed':       payments.filter(status='FAILED').count(),
        'verified':     payments.filter(is_verified=True).count(),
        'unverified':   payments.filter(is_verified=False).count(),
        'reversed':     payments.filter(reversed=True).count(),
        'refunded':     payments.filter(refunded=True).count(),
        'total_amount': payments.filter(status='COMPLETED').aggregate(Sum('amount'))['amount__sum'] or 0,
    }

    paginator     = Paginator(payments, 20)
    payments_page = paginator.get_page(request.GET.get('page', 1))
    is_htmx       = request.headers.get('HX-Request') == 'true'

    context = {
        'payments_page': payments_page,
        'paginator':     paginator,
        'stats':         stats,
        'filter_form':   filter_form,
        'is_htmx':       is_htmx,
    }
    if is_htmx:
        return render(request, 'fees/payments/partials/_payment_results.html', context)
    return render(request, 'fees/payments/list.html', context)


@login_required
def payment_create(request):
    invoice_id = request.GET.get('invoice')
    invoice    = get_object_or_404(FeeInvoice, pk=invoice_id) if invoice_id else None

    if request.method == 'POST':
        form = PaymentForm(request.POST, invoice=invoice)
        if not form.is_valid():
            is_htmx = request.headers.get('HX-Request') == 'true'
            if is_htmx:
                r = HttpResponse()
                r['HX-Alert-Message'] = 'Please correct the errors below.'
                r['HX-Alert-Type']    = 'error'
                r['HX-Alert-Title']   = 'Validation Error'
                return r
            messages.error(request, 'Please correct the errors below.')
        else:
            try:
                with transaction.atomic():
                    payment                 = form.save(commit=False)
                    payment.received_by_id  = str(request.user.id)
                    payment.processed_by_id = str(request.user.id)
                    payment.receipt_issued  = True
                    # payment_date and fiscal_period are set by payment_pre_save signal
                    payment.save()

                    is_htmx = request.headers.get('HX-Request') == 'true'
                    if is_htmx:
                        r = HttpResponse()
                        r['HX-Alert-Message'] = (
                            f"Payment '{payment.payment_number}' recorded! "
                            f"Receipt: {payment.receipt_number}"
                        )
                        r['HX-Alert-Type']  = 'success'
                        r['HX-Alert-Title'] = 'Payment Recorded!'
                        r['HX-Redirect']    = reverse('fees:payment_list')
                        return r

                    messages.success(
                        request,
                        f'Payment {payment.payment_number} recorded! '
                        f'Receipt: {payment.receipt_number}'
                    )
                    return redirect('fees:payment_list')

            except Exception as e:
                logger.error(f"Error creating payment: {e}", exc_info=True)
                is_htmx = request.headers.get('HX-Request') == 'true'
                if is_htmx:
                    r = HttpResponse()
                    r['HX-Alert-Message'] = f'Error creating payment: {str(e)}'
                    r['HX-Alert-Type']    = 'error'
                    r['HX-Alert-Title']   = 'Error!'
                    return r
                messages.error(request, f'Error creating payment: {str(e)}')
    else:
        initial = {}
        if invoice:
            initial['amount'] = invoice.balance
        form = PaymentForm(initial=initial, invoice=invoice)

    return render(request, 'fees/payments/single_payment_form.html', {
        'form':        form,
        'invoice':     invoice,
        'title':       'Record Payment',
        'submit_text': 'Record Payment',
    })


@login_required
def outstanding_invoices_for_student(request):
    """
    HTMX endpoint — returns outstanding invoice rows for a selected student.
    Triggered by the student/session filter selects on the multiple payment form.
    Supports multiple session IDs since the session filter is a Select2 multi-select.
    """
    student_id  = request.GET.get('student_filter')
    session_ids = request.GET.getlist('session_filter')
    invoice_rows = []

    if student_id:
        qs = FeeInvoice.objects.filter(
            student_id=student_id,
        ).exclude(
            status__in=[
                'CANCELLED', 'VOID', 'WRITTEN_OFF',
                'PAID', 'BAD_DEBT', 'UNCOLLECTIBLE',
            ]
        ).filter(
            balance__gt=0
        ).select_related(
            'student', 'academic_session'
        ).order_by('due_date')

        # Filter by one or more sessions when provided.
        # Empty strings can appear in getlist if the field submits a blank
        # option — strip those before applying the queryset filter.
        clean_session_ids = [s for s in session_ids if s]
        if clean_session_ids:
            qs = qs.filter(academic_session_id__in=clean_session_ids)

        invoice_rows = [
            {
                'invoice':  inv,
                'amount':   inv.balance,
                'selected': True,
                'errors':   [],
                'index':    i,
            }
            for i, inv in enumerate(qs)
        ]

    return render(
        request,
        'fees/payments/partials/_outstanding_invoice_rows.html',
        {'invoice_rows': invoice_rows},
    )


@login_required
def multiple_invoice_payment_create(request):

    # ── Shared context for both GET and failed POST ───────────────
    def get_base_context(form, invoice_rows=None):
        return {
            'form':              form,
            'title':             'Multiple Invoice Payment',
            'submit_text':       'Record Payments',
            'students':          Student.objects.filter(
                                     enrollment_status='ACTIVE'
                                 ).order_by('first_name', 'last_name'),
            'academic_sessions': AcademicSession.objects.order_by('-start_date')[:20],
            'invoice_rows':      invoice_rows or [],
        }

    if request.method == 'POST':
        form = MultipleInvoicePaymentForm(request.POST)
        if form.is_valid():
            try:
                with transaction.atomic():

                    # ── Read form fields ──────────────────────────
                    payment_method       = form.cleaned_data['payment_method']
                    reference_number     = form.cleaned_data.get('reference_number', '')
                    transaction_id_val   = form.cleaned_data.get('transaction_id', '')
                    bank_name            = form.cleaned_data.get('bank_name', '')
                    account_number       = form.cleaned_data.get('account_number', '')
                    cheque_number        = form.cleaned_data.get('cheque_number', '')
                    cheque_date          = form.cleaned_data.get('cheque_date')
                    mobile_money_provider= form.cleaned_data.get('mobile_money_provider', '')
                    mobile_number        = form.cleaned_data.get('mobile_number', '')
                    currency             = form.cleaned_data.get('currency', '')
                    exchange_rate        = form.cleaned_data.get('exchange_rate') or Decimal('1.000000')
                    paid_by_name         = form.cleaned_data.get('paid_by_name', '')
                    paid_by_phone        = form.cleaned_data.get('paid_by_phone', '')
                    paid_by_relationship = form.cleaned_data.get('paid_by_relationship', '')
                    remarks              = form.cleaned_data.get('remarks', '')

                    # ── School currency fallback ───────────────────
                    try:
                        from core.models import FinancialSettings
                        school_currency = FinancialSettings.get_school_currency() or 'UGX'
                    except Exception:
                        school_currency = 'UGX'

                    if not currency:
                        currency = school_currency

                    # ── Read per-row invoice PKs and amounts from POST ──
                    # The form has no selected_invoices field — the template
                    # posts invoice_N / amount_N / selected_N per row.
                    invoice_rows_post = []
                    i = 0
                    while True:
                        inv_id = request.POST.get(f'invoice_{i}')
                        if inv_id is None:
                            break
                        # Only include rows that were checked
                        selected = request.POST.get(f'selected_{i}') == inv_id
                        raw_amount = request.POST.get(f'amount_{i}', '').strip()
                        if selected and raw_amount:
                            try:
                                amount = Decimal(raw_amount)
                                if amount > 0:
                                    inv = FeeInvoice.objects.select_related(
                                        'student', 'academic_session'
                                    ).get(pk=inv_id)
                                    invoice_rows_post.append((inv, amount))
                            except (FeeInvoice.DoesNotExist, Exception):
                                pass
                        i += 1

                    if not invoice_rows_post:
                        raise ValueError(
                            "No invoices selected or no amounts entered. "
                            "Please select invoices and enter amounts before saving."
                        )

                    # ── fiscal_period set by signal — verify it exists ─
                    # Validate upfront so we get a clean error before any
                    # payments are created.
                    fiscal_period = FiscalPeriod.get_current_fiscal_period()
                    if not fiscal_period:
                        raise ValueError(
                            "No active fiscal period found. "
                            "Please activate a fiscal period before recording payments."
                        )

                    # ── Create one Payment per selected invoice row ──
                    created_payments = []

                    for inv, amount_applied in invoice_rows_post:

                        overpayment = max(
                            Decimal('0.00'),
                            amount_applied - inv.balance,
                        )

                        amount_in_school_currency = (
                            amount_applied * exchange_rate
                        ).quantize(Decimal('0.01'))

                        payment = Payment.objects.create(
                            invoice                   = inv,
                            student                   = inv.student,
                            amount                    = amount_applied,
                            amount_applied_to_invoice = amount_applied,
                            overpayment_amount        = overpayment,
                            # payment_date and fiscal_period set by
                            # payment_pre_save signal automatically
                            payment_method            = payment_method,
                            currency                  = currency,
                            exchange_rate             = exchange_rate,
                            amount_in_school_currency = amount_in_school_currency,
                            reference_number          = reference_number,
                            transaction_id            = transaction_id_val,
                            bank_name                 = bank_name,
                            account_number            = account_number,
                            cheque_number             = cheque_number,
                            cheque_date               = cheque_date,
                            mobile_money_provider     = mobile_money_provider,
                            mobile_number             = mobile_number,
                            paid_by_name              = paid_by_name,
                            paid_by_phone             = paid_by_phone,
                            paid_by_relationship      = paid_by_relationship,
                            remarks                   = (
                                f"{remarks}\n[Multiple invoice payment]".strip()
                                if remarks else "Multiple invoice payment"
                            ),
                            received_by_id            = str(request.user.id),
                            processed_by_id           = str(request.user.id),
                            academic_session          = inv.academic_session,
                            status                    = 'COMPLETED',
                            receipt_issued            = True,
                        )

                        created_payments.append(payment)

                    total_recorded = sum(p.amount for p in created_payments)

                    success_msg = (
                        f"Created {len(created_payments)} payment(s) totalling "
                        f"{school_currency} {total_recorded:,.2f}"
                    )

                    is_htmx = request.headers.get('HX-Request') == 'true'
                    if is_htmx:
                        r = HttpResponse()
                        r['HX-Alert-Message'] = success_msg
                        r['HX-Alert-Type']    = 'success'
                        r['HX-Alert-Title']   = 'Payments Created!'
                        r['HX-Redirect']      = (
                            reverse(
                                'fees:payment_detail',
                                kwargs={'pk': created_payments[0].pk},
                            )
                            if created_payments
                            else reverse('fees:payment_list')
                        )
                        return r

                    messages.success(request, success_msg)
                    return (
                        redirect('fees:payment_detail', pk=created_payments[0].pk)
                        if created_payments
                        else redirect('fees:payment_list')
                    )

            except Exception as e:
                logger.error(
                    f"Error creating multiple invoice payment: {e}",
                    exc_info=True,
                )
                is_htmx = request.headers.get('HX-Request') == 'true'
                if is_htmx:
                    r = HttpResponse()
                    r['HX-Alert-Message'] = f'Error creating payment: {str(e)}'
                    r['HX-Alert-Type']    = 'error'
                    r['HX-Alert-Title']   = 'Error!'
                    return r
                messages.error(request, f'Error creating payment: {str(e)}')

        # ── POST with validation errors — rebuild invoice rows from POST data ─
        # Re-load the invoices the cashier had selected so the table
        # re-renders with their entered amounts intact.
        invoice_rows = []
        i = 0
        while True:
            inv_id = request.POST.get(f'invoice_{i}')
            if inv_id is None:
                break
            try:
                inv    = FeeInvoice.objects.select_related(
                    'student', 'academic_session'
                ).get(pk=inv_id)
                invoice_rows.append({
                    'invoice':  inv,
                    'amount':   request.POST.get(f'amount_{i}', ''),
                    'selected': request.POST.get(f'selected_{i}') == inv_id,
                    'errors':   [],
                    'index':    i,
                })
            except FeeInvoice.DoesNotExist:
                pass
            i += 1

        return render(
            request,
            'fees/payments/multiple_payment_form.html',
            get_base_context(form, invoice_rows),
        )

    else:
        # ── GET ───────────────────────────────────────────────────
        form = MultipleInvoicePaymentForm()

        # If arriving from the invoice list with ?student=UUID,
        # pre-load that student's invoices immediately.
        student_id          = request.GET.get('student')
        invoice_rows        = []
        selected_student_id = None

        if student_id:
            selected_student_id = student_id
            qs = FeeInvoice.objects.filter(
                student_id=student_id,
            ).exclude(
                status__in=[
                    'CANCELLED', 'VOID', 'WRITTEN_OFF',
                    'PAID', 'BAD_DEBT', 'UNCOLLECTIBLE',
                ]
            ).filter(
                balance__gt=0
            ).select_related(
                'student', 'academic_session'
            ).order_by('due_date')

            invoice_rows = [
                {
                    'invoice':  inv,
                    'amount':   inv.balance,
                    'selected': True,
                    'errors':   [],
                    'index':    i,
                }
                for i, inv in enumerate(qs)
            ]

        ctx = get_base_context(form, invoice_rows)
        ctx['selected_student_id'] = selected_student_id
        return render(
            request,
            'fees/payments/multiple_payment_form.html',
            ctx,
        )


@login_required
def payment_detail(request, pk):
    payment = get_object_or_404(
        Payment.objects.select_related('student', 'invoice', 'payment_method', 'academic_session', 'fiscal_period', 'journal_entry', 'reversal_journal_entry', 'refund_journal_entry'),
        pk=pk,
    )
    try:
        audit_trail = payment.get_audit_trail()
    except Exception as e:
        logger.error(f"Error getting audit trail: {e}")
        audit_trail = []
    try:
        account_summary = payment.student.financial_account.get_account_summary()
    except Exception as e:
        logger.error(f"Error getting account summary: {e}")
        account_summary = None

    return render(request, 'fees/payments/detail.html', {
        'payment':         payment,
        'audit_trail':     audit_trail,
        'account_summary': account_summary,
    })

@login_required
@require_http_methods(["POST"])
def payment_delete(request, pk):
    payment = get_object_or_404(Payment, pk=pk)
    if not payment.reversed and not payment.refunded:
        if payment.is_verified or payment.status == 'COMPLETED':
            is_htmx = request.headers.get('HX-Request') == 'true'
            if is_htmx:
                r = HttpResponse()
                r['HX-Alert-Message'] = f"Cannot delete verified payment '{payment.payment_number}'. Use 'Reverse Payment' instead."
                r['HX-Alert-Type']    = 'error'
                r['HX-Alert-Title']   = 'Cannot Delete'
                r['HX-Close-Modal']   = 'true'
                return r
            messages.error(request, f"Cannot delete verified payment '{payment.payment_number}'", extra_tags='sweetalert-error')
            return redirect('fees:payment_detail', pk=pk)

    number = payment.payment_number
    try:
        payment.delete()
        is_htmx = request.headers.get('HX-Request') == 'true'
        if is_htmx:
            r = HttpResponse()
            r['HX-Alert-Message'] = f"Payment '{number}' deleted successfully"
            r['HX-Alert-Type']    = 'success'
            r['HX-Alert-Title']   = 'Deleted!'
            r['HX-Close-Modal']   = 'true'
            r['HX-Trigger']       = 'refreshPaymentList'
            return r
        messages.success(request, f"Payment '{number}' deleted successfully", extra_tags='sweetalert')
        return redirect('fees:payment_list')
    except Exception as e:
        logger.error(f"Error deleting payment {number}: {e}", exc_info=True)
        is_htmx = request.headers.get('HX-Request') == 'true'
        if is_htmx:
            r = HttpResponse()
            r['HX-Alert-Message'] = f"Error deleting payment: {str(e)}"
            r['HX-Alert-Type']    = 'error'
            r['HX-Alert-Title']   = 'Delete Failed'
            r['HX-Close-Modal']   = 'true'
            return r
        messages.error(request, f"Error deleting payment: {str(e)}", extra_tags='sweetalert-error')
        return redirect('fees:payment_list')


@login_required
def payment_reverse(request, pk):
    payment = get_object_or_404(Payment, pk=pk)
    can_reverse, reason = payment.can_be_reversed()
    if not can_reverse:
        messages.error(request, f'Cannot reverse: {reason}')
        return redirect('fees:payment_detail', pk=pk)
    if request.method == 'POST':
        form = PaymentReversalForm(payment, request.user, request.POST)
        if form.is_valid():
            try:
                with transaction.atomic():
                    payment.reversed        = True
                    payment.reversed_on     = get_school_current_time()
                    payment.reversed_by_id  = str(request.user.id)
                    payment.reversal_reason = form.cleaned_data['reversal_reason']
                    payment.status          = 'REVERSED'
                    payment.save()
                messages.success(request, 'Payment reversed successfully')
                return redirect('fees:payment_detail', pk=pk)
            except Exception as e:
                logger.error(f"Error reversing payment: {e}", exc_info=True)
                messages.error(request, f'Error: {str(e)}')
    else:
        form = PaymentReversalForm(payment, request.user)
    return render(request, 'fees/payments/reverse_form.html', {'form': form, 'payment': payment})


@login_required
def payment_refund(request, pk):
    payment = get_object_or_404(Payment, pk=pk)
    can_refund, reason = payment.can_be_refunded()
    if not can_refund:
        messages.error(request, f'Cannot refund this payment: {reason}')
        return redirect('fees:payment_detail', pk=pk)

    if request.method == 'POST':
        form = PaymentRefundForm(payment, request.user, request.POST)
        if form.is_valid():
            try:
                with transaction.atomic():
                    refund_amount    = form.cleaned_data['refund_amount']
                    refund_method    = form.cleaned_data['refund_method']
                    refund_reference = form.cleaned_data['refund_reference']
                    refund_reason    = form.cleaned_data['refund_reason']
                    refund_notes     = form.cleaned_data.get('refund_notes', '')

                    if refund_amount < payment.amount:
                        messages.error(request, 'Partial refunds are not supported. Please refund the full amount.')
                        return render(request, 'fees/payments/refund_form.html', {'form': form, 'payment': payment})

                    payment.refunded       = True
                    payment.refunded_on    = get_school_current_time()
                    payment.refunded_by_id = str(request.user.id)
                    payment.refund_method  = refund_method
                    payment.refund_reference = refund_reference
                    payment.refund_notes   = f"{refund_reason}\n\n{refund_notes}".strip()
                    payment.status         = 'REFUNDED'
                    payment.save()

                messages.success(request, f'Payment {payment.payment_number} refunded. Amount: UGX {refund_amount:,.2f} via {refund_method}')
                return redirect('fees:payment_detail', pk=payment.pk)
            except Exception as e:
                logger.error(f"Error refunding payment: {e}", exc_info=True)
                messages.error(request, f'Error refunding payment: {str(e)}')
    else:
        form = PaymentRefundForm(payment, request.user)

    return render(request, 'fees/payments/refund_form.html', {
        'form': form, 'payment': payment,
        'title': f'Refund Payment {payment.payment_number}', 'submit_text': 'Issue Refund',
    })


@login_required
@require_http_methods(["POST"])
def payment_verify(request, pk):
    payment = get_object_or_404(Payment, pk=pk)
    if payment.is_verified:
        is_htmx = request.headers.get('HX-Request') == 'true'
        if is_htmx:
            r = HttpResponse()
            r['HX-Alert-Message'] = 'Payment is already verified.'
            r['HX-Alert-Type']    = 'info'
            r['HX-Alert-Title']   = 'Already Verified'
            r['HX-Redirect']      = reverse('fees:payment_detail', kwargs={'pk': pk})
            return r
        messages.info(request, 'Payment is already verified.')
        return redirect('fees:payment_detail', pk=pk)
    if payment.reversed or payment.refunded:
        is_htmx = request.headers.get('HX-Request') == 'true'
        if is_htmx:
            r = HttpResponse()
            r['HX-Alert-Message'] = 'Cannot verify reversed or refunded payments.'
            r['HX-Alert-Type']    = 'error'
            r['HX-Alert-Title']   = 'Cannot Verify'
            r['HX-Redirect']      = reverse('fees:payment_detail', kwargs={'pk': pk})
            return r
        messages.error(request, 'Cannot verify reversed or refunded payments.')
        return redirect('fees:payment_detail', pk=pk)
    try:
        with transaction.atomic():
            verification_notes      = request.POST.get('verification_notes', '').strip()
            payment.is_verified     = True
            payment.verified_by_id  = str(request.user.id)
            payment.verification_date = get_school_current_time()
            payment.status          = 'COMPLETED'
            if verification_notes:
                payment.internal_notes = f"{payment.internal_notes}\n\nVerification: {verification_notes}".strip() if payment.internal_notes else f"Verification: {verification_notes}"
            payment.save()
        is_htmx = request.headers.get('HX-Request') == 'true'
        if is_htmx:
            r = HttpResponse()
            r['HX-Alert-Message'] = f'Payment {payment.payment_number} verified successfully!'
            r['HX-Alert-Type']    = 'success'
            r['HX-Alert-Title']   = 'Verified!'
            r['HX-Redirect']      = reverse('fees:payment_detail', kwargs={'pk': payment.pk})
            return r
        messages.success(request, f'Payment {payment.payment_number} verified successfully!', extra_tags='sweetalert')
        return redirect('fees:payment_detail', pk=payment.pk)
    except Exception as e:
        logger.error(f"Error verifying payment: {e}", exc_info=True)
        is_htmx = request.headers.get('HX-Request') == 'true'
        if is_htmx:
            r = HttpResponse()
            r['HX-Alert-Message'] = f'Error verifying payment: {str(e)}'
            r['HX-Alert-Type']    = 'error'
            r['HX-Alert-Title']   = 'Error'
            return r
        messages.error(request, f'Error verifying payment: {str(e)}')
        return redirect('fees:payment_detail', pk=pk)


@login_required
@require_http_methods(["POST"])
def payment_bulk_verify(request):
    try:
        with transaction.atomic():
            payment_ids        = request.POST.getlist('payment_ids')
            if not payment_ids:
                raise ValueError("No payments selected for verification")
            verification_notes = request.POST.get('verification_notes', '').strip()
            payments           = Payment.objects.filter(id__in=payment_ids, is_verified=False)
            if not payments.exists():
                raise ValueError("No unverified payments found")
            verified_count     = 0
            verification_time  = get_school_current_time()
            for payment in payments:
                if payment.is_verified or payment.reversed or payment.refunded:
                    continue
                payment.is_verified      = True
                payment.verified_by_id   = str(request.user.id)
                payment.verification_date = verification_time
                payment.status           = 'COMPLETED'
                if verification_notes:
                    payment.internal_notes = f"{payment.internal_notes}\n\nBulk Verification: {verification_notes}".strip() if payment.internal_notes else f"Bulk Verification: {verification_notes}"
                payment.save()
                verified_count += 1
        is_htmx = request.headers.get('HX-Request') == 'true'
        if is_htmx:
            r = HttpResponse()
            r['HX-Alert-Message'] = f'Successfully verified {verified_count} payment(s)!'
            r['HX-Alert-Type']    = 'success'
            r['HX-Alert-Title']   = 'Bulk Verification Complete'
            r['HX-Redirect']      = reverse('fees:payment_list')
            return r
        messages.success(request, f'Successfully verified {verified_count} payment(s)!', extra_tags='sweetalert')
        return redirect('fees:payment_list')
    except ValueError as e:
        is_htmx = request.headers.get('HX-Request') == 'true'
        if is_htmx:
            r = HttpResponse()
            r['HX-Alert-Message'] = str(e)
            r['HX-Alert-Type']    = 'warning'
            r['HX-Alert-Title']   = 'Validation Error'
            return r
        messages.warning(request, str(e))
        return redirect('fees:payment_list')
    except Exception as e:
        logger.error(f"Error bulk verifying payments: {e}", exc_info=True)
        is_htmx = request.headers.get('HX-Request') == 'true'
        if is_htmx:
            r = HttpResponse()
            r['HX-Alert-Message'] = f'Error verifying payments: {str(e)}'
            r['HX-Alert-Type']    = 'error'
            r['HX-Alert-Title']   = 'Error'
            return r
        messages.error(request, f'Error verifying payments: {str(e)}')
        return redirect('fees:payment_list')


@login_required
@require_http_methods(["POST"])
def payment_send_receipt(request, pk):
    payment = get_object_or_404(Payment, pk=pk)
    try:
        recipient_emails = request.POST.getlist('recipients')
        if not recipient_emails:
            raise ValueError("At least one recipient email is required")
        email_body = render_to_string('fees/emails/payment_receipt.html', {
            'payment': payment, 'school_name': getattr(settings, 'SCHOOL_NAME', 'School'),
        })
        send_mail(
            subject=f"Payment Receipt {payment.receipt_number or payment.payment_number}",
            message='', from_email=settings.DEFAULT_FROM_EMAIL,
            recipient_list=recipient_emails, html_message=email_body, fail_silently=False,
        )
        payment.internal_notes = f"{payment.internal_notes}\n\nReceipt emailed to {', '.join(recipient_emails)} on {get_school_current_time()}".strip()
        payment.save()
        is_htmx = request.headers.get('HX-Request') == 'true'
        if is_htmx:
            r = HttpResponse()
            r['HX-Alert-Message'] = f"Receipt sent to {len(recipient_emails)} recipient(s)!"
            r['HX-Alert-Type']    = 'success'
            r['HX-Alert-Title']   = 'Receipt Sent'
            r['HX-Close-Modal']   = 'true'
            r['HX-Redirect']      = reverse('fees:payment_detail', kwargs={'pk': payment.pk})
            return r
        messages.success(request, f"Receipt sent to {len(recipient_emails)} recipient(s)!")
        return redirect('fees:payment_detail', pk=payment.pk)
    except Exception as e:
        logger.error(f"Error sending receipt: {e}", exc_info=True)
        is_htmx = request.headers.get('HX-Request') == 'true'
        if is_htmx:
            r = HttpResponse()
            r['HX-Alert-Message'] = f'Error sending receipt: {str(e)}'
            r['HX-Alert-Type']    = 'error'
            r['HX-Alert-Title']   = 'Error'
            return r
        messages.error(request, f'Error sending receipt: {str(e)}')
        return redirect('fees:payment_detail', pk=pk)


@login_required
def payment_print_receipt(request, pk):
    payment = get_object_or_404(Payment.objects.select_related('student', 'invoice', 'payment_method'), pk=pk)
    return render(request, 'fees/payments/print_receipt.html', {
        **get_print_school_context(request),
        'payment': payment, 'now': timezone.now(),
        'title':   f'Receipt {payment.receipt_number or payment.payment_number}',
    })


@login_required
def payment_list_print_view(request):
    FIELD_NAMES_FULL = {
        'payment_number':       'Payment Number',
        'receipt_number':       'Receipt Number',
        'payment_date':         'Payment Date',
        'student_id':           'Admission Number',
        'student_name':         'Student Name',
        'invoice_number':       'Invoice Number',
        'academic_session':     'Academic Session',
        'fiscal_period':        'Fiscal Period',
        'amount':               'Amount',
        'amount_applied':       'Amount Applied',
        'overpayment':          'Overpayment',
        'payment_method':       'Payment Method',
        'reference_number':     'Reference Number',
        'transaction_id':       'Transaction ID',
        'bank_name':            'Bank Name',
        'cheque_number':        'Cheque Number',
        'mobile_provider':      'Mobile Provider',
        'mobile_number':        'Mobile Number',
        'paid_by_name':         'Paid By',
        'paid_by_phone':        'Payer Phone',
        'paid_by_relationship': 'Relationship to Student',
        'status':               'Status',
        'is_verified':          'Verified',
        'reversed':             'Reversed',
        'refunded':             'Refunded',
        'remarks':              'Remarks',
    }
    FIELD_NAMES_SHORT = {
        'payment_number':       'Pmt. No.',
        'receipt_number':       'Rcpt. No.',
        'payment_date':         'Date',
        'student_id':           'Adm. No.',
        'student_name':         'Student',
        'invoice_number':       'Inv. No.',
        'academic_session':     'Session',
        'fiscal_period':        'Period',
        'amount':               'Amount',
        'amount_applied':       'Applied',
        'overpayment':          'Overpmt.',
        'payment_method':       'Method',
        'reference_number':     'Ref.',
        'transaction_id':       'Trans. ID',
        'bank_name':            'Bank',
        'cheque_number':        'Chq. No.',
        'mobile_provider':      'Provider',
        'mobile_number':        'Mobile',
        'paid_by_name':         'Paid By',
        'paid_by_phone':        'Phone',
        'paid_by_relationship': 'Relation',
        'status':               'Status',
        'is_verified':          'Verified',
        'reversed':             'Rev.',
        'refunded':             'Ref.',
        'remarks':              'Remarks',
    }
    DEFAULT_FIELDS  = ['payment_number', 'receipt_number', 'payment_date', 'student_id', 'student_name', 'invoice_number', 'amount', 'payment_method', 'reference_number', 'status', 'is_verified']
    selected_fields = request.GET.getlist('fields') or DEFAULT_FIELDS
    short_headers   = request.GET.get('short_headers', 'false').lower() == 'true'
    landscape       = request.GET.get('landscape', 'true').lower() == 'true'
    include_stats   = request.GET.get('include_stats', 'true').lower() == 'true'
    field_names     = FIELD_NAMES_SHORT if short_headers else FIELD_NAMES_FULL
    payments        = get_filtered_payments(request)

    stats = None
    if include_stats:
        completed = payments.filter(status='COMPLETED')
        stats = {
            'total':        payments.count(),
            'completed':    completed.count(),
            'verified':     payments.filter(is_verified=True).count(),
            'reversed':     payments.filter(reversed=True).count(),
            'refunded':     payments.filter(refunded=True).count(),
            'total_amount': completed.aggregate(v=Sum('amount'))['v'] or Decimal('0.00'),
        }

    if payments.count() > MAX_PRINT_RECORDS:
        messages.warning(request, f'Only the first {MAX_PRINT_RECORDS} payments will be printed.')
        payments = payments[:MAX_PRINT_RECORDS]

    return render(request, 'fees/payments/print_payment_list.html', {
        **get_print_school_context(request),
        'payments':             payments,
        'stats':                stats,
        'selected_fields':      selected_fields,
        'selected_field_names': [field_names.get(f, f.replace('_', ' ').title()) for f in selected_fields],
        'field_names':          field_names,
        'short_headers':        short_headers,
        'landscape':            landscape,
        'now':                  timezone.now(),
        'print_date':           get_school_today(),
        'printed_by':           request.user.get_full_name() or request.user.username,
        'title':                'Payment List',
    })


@login_required
def export_payments_excel(request):
    ALL_COLUMNS = [
        ('payment_number',       'Payment Number',    lambda o: o.payment_number),
        ('receipt_number',       'Receipt Number',    lambda o: o.receipt_number or ''),
        ('payment_date',         'Payment Date',      lambda o: o.payment_date.strftime('%Y-%m-%d') if o.payment_date else ''),
        ('student_id',           'Admission No.',     lambda o: o.student.admission_number),
        ('student_name',         'Student Name',      lambda o: o.student.get_full_name()),
        ('invoice_number',       'Invoice Number',    lambda o: o.invoice.invoice_number if o.invoice else ''),
        ('academic_session',     'Academic Session',  lambda o: o.academic_session.name if o.academic_session else ''),
        ('fiscal_period',        'Fiscal Period',     lambda o: o.fiscal_period.name if o.fiscal_period else ''),
        ('amount',               'Amount',            lambda o: float(o.amount)),
        ('amount_applied',       'Amount Applied',    lambda o: float(o.amount_applied_to_invoice)),
        ('overpayment',          'Overpayment',       lambda o: float(o.overpayment_amount)),
        ('payment_method',       'Payment Method',    lambda o: o.payment_method.name if o.payment_method else ''),
        ('reference_number',     'Reference No.',     lambda o: o.reference_number or ''),
        ('transaction_id',       'Transaction ID',    lambda o: o.transaction_id or ''),
        ('bank_name',            'Bank',              lambda o: o.bank_name or ''),
        ('cheque_number',        'Cheque No.',        lambda o: o.cheque_number or ''),
        ('mobile_provider',      'Mobile Provider',   lambda o: o.mobile_money_provider or ''),
        ('mobile_number',        'Mobile Number',     lambda o: o.mobile_number or ''),
        ('paid_by_name',         'Paid By',           lambda o: o.paid_by_name or ''),
        ('paid_by_phone',        'Payer Phone',       lambda o: o.paid_by_phone or ''),
        ('paid_by_relationship', 'Relationship',      lambda o: o.get_paid_by_relationship_display() if o.paid_by_relationship else ''),
        ('status',               'Status',            lambda o: o.get_status_display()),
        ('is_verified',          'Verified',          lambda o: 'Yes' if o.is_verified else 'No'),
        ('reversed',             'Reversed',          lambda o: 'Yes' if o.reversed else 'No'),
        ('refunded',             'Refunded',          lambda o: 'Yes' if o.refunded else 'No'),
        ('remarks',              'Remarks',           lambda o: o.remarks or ''),
    ]
    DEFAULT_FIELDS = ['payment_number', 'receipt_number', 'payment_date', 'student_id', 'student_name', 'invoice_number', 'amount', 'payment_method', 'reference_number', 'status', 'is_verified']
    payments = get_filtered_payments(request)
    columns  = _resolve_columns(ALL_COLUMNS, request.GET.getlist('fields'), DEFAULT_FIELDS)
    return _xlsx_response(_make_workbook('Payments', columns, payments), 'payments')


# ─── Payment API endpoints ────────────────────────────────────────────────────

@login_required
@require_http_methods(["GET"])
def api_get_student_invoices(request):
    student_ids = request.GET.getlist('students[]')
    if not student_ids:
        return JsonResponse({'error': 'No students selected'}, status=400)
    try:
        invoices = FeeInvoice.objects.filter(
            student_id__in=student_ids,
            status__in=['PENDING', 'PARTIALLY_PAID', 'OVERDUE'],
        ).select_related('student').order_by('issue_date')
        data = [{
            'id': str(inv.id), 'number': inv.invoice_number,
            'student': inv.student.get_full_name(), 'student_id': str(inv.student.id),
            'balance': float(inv.balance), 'total': float(inv.total_amount),
            'paid': float(inv.paid_amount), 'date': inv.issue_date.strftime('%Y-%m-%d'),
            'session': inv.academic_session.name if inv.academic_session else 'N/A',
        } for inv in invoices]
        return JsonResponse({'success': True, 'invoices': data, 'total_balance': sum(d['balance'] for d in data), 'count': len(data)})
    except Exception as e:
        logger.error(f"Error fetching student invoices: {e}", exc_info=True)
        return JsonResponse({'error': str(e)}, status=500)


@login_required
@require_http_methods(["POST"])
def api_validate_invoice_numbers(request):
    import json
    try:
        data            = json.loads(request.body)
        numbers_str     = data.get('invoice_numbers', '').strip()
        if not numbers_str:
            return JsonResponse({'error': 'No invoice numbers provided'}, status=400)
        invoice_list    = [n.strip() for n in numbers_str.replace(',', '\n').split('\n') if n.strip()]
        if not invoice_list:
            return JsonResponse({'error': 'No valid invoice numbers found'}, status=400)
        invoices = FeeInvoice.objects.filter(invoice_number__in=invoice_list, status__in=['PENDING', 'PARTIALLY_PAID', 'OVERDUE']).select_related('student')
        found_data       = [{
            'id': str(inv.id), 'number': inv.invoice_number,
            'student': inv.student.get_full_name(), 'student_id': str(inv.student.id),
            'balance': float(inv.balance), 'total': float(inv.total_amount),
            'paid': float(inv.paid_amount), 'date': inv.issue_date.strftime('%Y-%m-%d'),
            'session': inv.academic_session.name if inv.academic_session else 'N/A',
        } for inv in invoices]
        found_numbers    = {inv.invoice_number for inv in invoices}
        missing          = [n for n in invoice_list if n not in found_numbers]
        return JsonResponse({'success': True, 'invoices': found_data, 'missing': missing, 'total_balance': sum(d['balance'] for d in found_data), 'count': len(found_data)})
    except json.JSONDecodeError:
        return JsonResponse({'error': 'Invalid JSON data'}, status=400)
    except Exception as e:
        logger.error(f"Error validating invoice numbers: {e}", exc_info=True)
        return JsonResponse({'error': str(e)}, status=500)


# =============================================================================
# SCHOLARSHIP PROGRAMS
# =============================================================================

def get_filtered_scholarship_programs(request):
    programs = ScholarshipProgram.objects.prefetch_related(
        'applicable_fee_categories', 'applicable_levels', 'valid_sessions'
    ).annotate(
        application_count=Count('applications', distinct=True),
        active_scholarship_count=Count('student_scholarships', filter=Q(student_scholarships__status='ACTIVE'), distinct=True),
    ).order_by('name')

    query                     = request.GET.get('q', '').strip()
    scholarship_type          = request.GET.get('scholarship_type', '')
    program_type              = request.GET.get('program_type', '')
    discount_type             = request.GET.get('discount_type', '')
    is_active                 = request.GET.get('is_active', '')
    is_accepting_applications = request.GET.get('is_accepting_applications', '')
    academic_session          = request.GET.get('academic_session', '')

    if query:
        words = query.split()
        q = Q()
        for w in words:
            q &= Q(name__icontains=w) | Q(code__icontains=w) | Q(description__icontains=w) | Q(sponsor_name__icontains=w)
        programs = programs.filter(q)

    if scholarship_type:
        programs = programs.filter(scholarship_type=scholarship_type)
    if program_type:
        programs = programs.filter(program_type=program_type)
    if discount_type:
        programs = programs.filter(discount_type=discount_type)
    if academic_session:
        programs = programs.filter(valid_sessions__id=academic_session)
    if is_active:
        programs = programs.filter(is_active=(is_active.lower() == 'true'))
    if is_accepting_applications:
        programs = programs.filter(is_accepting_applications=(is_accepting_applications.lower() == 'true'))

    return programs


@login_required
def scholarship_program_list(request):
    filter_form = ScholarshipProgramFilterForm(request.GET or None)
    programs    = get_filtered_scholarship_programs(request)

    stats = {
        'total':                    programs.count(),
        'active':                   programs.filter(is_active=True).count(),
        'accepting_applications':   programs.filter(is_accepting_applications=True).count(),
        'total_budget':             programs.aggregate(Sum('total_budget_amount'))['total_budget_amount__sum'] or 0,
        'total_used':               programs.aggregate(Sum('current_budget_used'))['current_budget_used__sum'] or 0,
        'total_recipients':         programs.aggregate(Sum('current_recipient_count'))['current_recipient_count__sum'] or 0,
    }

    paginator     = Paginator(programs, 20)
    programs_page = paginator.get_page(request.GET.get('page', 1))
    is_htmx       = request.headers.get('HX-Request') == 'true'

    context = {'programs_page': programs_page, 'paginator': paginator, 'stats': stats, 'filter_form': filter_form, 'is_htmx': is_htmx}
    if is_htmx:
        return render(request, 'fees/scholarships/partials/_program_results.html', context)
    return render(request, 'fees/scholarships/program_list.html', context)


@login_required
def scholarship_program_create(request):
    if request.method == 'POST':
        form = ScholarshipProgramForm(request.POST)
        if form.is_valid():
            try:
                program = form.save()
                is_htmx = request.headers.get('HX-Request') == 'true'
                if is_htmx:
                    r = HttpResponse()
                    r['HX-Alert-Message'] = f"Scholarship program '{program.name}' created successfully!"
                    r['HX-Alert-Type']    = 'success'
                    r['HX-Alert-Title']   = 'Created!'
                    r['HX-Redirect']      = reverse('fees:scholarship_program_detail', kwargs={'pk': program.pk})
                    return r
                messages.success(request, f"Scholarship program '{program.name}' created successfully!", extra_tags='sweetalert')
                return redirect('fees:scholarship_program_detail', pk=program.pk)
            except Exception as e:
                logger.error(f"Error creating scholarship program: {e}")
                is_htmx = request.headers.get('HX-Request') == 'true'
                if is_htmx:
                    r = HttpResponse()
                    r['HX-Alert-Message'] = f'Error creating program: {str(e)}'
                    r['HX-Alert-Type']    = 'error'
                    r['HX-Alert-Title']   = 'Error!'
                    return r
                messages.error(request, f'Error creating program: {str(e)}')
    else:
        form = ScholarshipProgramForm()

    return render(request, 'fees/scholarships/program_form.html', {
        'form': form, 'title': 'Create Scholarship Program', 'submit_text': 'Create Program', 'submit_icon': 'fa-plus',
        'all_sessions':   AcademicSession.objects.filter(is_active=True).order_by('-start_date'),
        'all_categories': FeesCategory.objects.filter(is_active=True).order_by('display_group__display_order', 'display_order'),
    })


@login_required
def scholarship_program_detail(request, pk):
    program          = get_object_or_404(ScholarshipProgram, pk=pk)
    all_scholarships = program.student_scholarships.all()
    all_applications = program.applications.all()

    stats = {
        'active_recipients':  all_scholarships.filter(status='ACTIVE').count(),
        'suspended':          all_scholarships.filter(status='SUSPENDED').count(),
        'terminated':         all_scholarships.filter(status='TERMINATED').count(),
        'completed':          all_scholarships.filter(status='COMPLETED').count(),
        'total_applications': all_applications.count(),
        'pending_review':     all_applications.filter(status__in=['SUBMITTED', 'UNDER_REVIEW']).count(),
        'approved':           all_applications.filter(status='APPROVED').count(),
        'rejected':           all_applications.filter(status='REJECTED').count(),
        'total_disbursed':    all_scholarships.aggregate(total=Sum('total_amount_used'))['total'] or Decimal('0.00'),
        'budget_remaining':   (program.total_budget_amount - (program.current_budget_used or Decimal('0'))) if program.total_budget_amount is not None else None,
    }

    category_template = None
    all_categories    = None
    if program.is_category_specific_discount():
        category_template = program.get_category_discount_template()
        all_categories    = FeesCategory.objects.filter(is_active=True).order_by('display_group__display_order', 'display_order')

    return render(request, 'fees/scholarships/program_detail.html', {
        'program':           program,
        'stats':             stats,
        'discount_summary':  program.get_discount_summary(),
        'category_template': category_template,
        'all_categories':    all_categories,
        'today':             get_school_today(),
    })


@login_required
def scholarship_program_recipients_partial(request, pk):
    program = get_object_or_404(ScholarshipProgram, pk=pk)
    qs      = program.student_scholarships.select_related('student', 'scholarship_program')
    q             = request.GET.get('q', '').strip()
    status        = request.GET.get('status', '')
    discount_mode = request.GET.get('discount_mode', '')
    if q:
        qs = qs.filter(Q(student__first_name__icontains=q) | Q(student__last_name__icontains=q) | Q(student__admission_number__icontains=q))
    qs = qs.filter(status=status) if status else qs.filter(status='ACTIVE')
    if discount_mode == 'global':
        qs = qs.filter(use_category_specific_discounts=False)
    elif discount_mode == 'category':
        qs = qs.filter(use_category_specific_discounts=True)
    paginator = Paginator(qs.order_by('-start_date'), 20)
    page      = paginator.get_page(request.GET.get('page', 1))
    return render(request, 'fees/scholarships/partials/_recipient_results.html', {'recipients': page, 'program': program})


@login_required
def scholarship_program_applications_partial(request, pk):
    program = get_object_or_404(ScholarshipProgram, pk=pk)
    qs      = program.applications.select_related('student', 'academic_session')
    q       = request.GET.get('q', '').strip()
    status  = request.GET.get('status', '')
    if q:
        qs = qs.filter(Q(student__first_name__icontains=q) | Q(student__last_name__icontains=q) | Q(student__admission_number__icontains=q) | Q(application_number__icontains=q))
    if status:
        qs = qs.filter(status=status)
    paginator = Paginator(qs.order_by('-application_date'), 20)
    page      = paginator.get_page(request.GET.get('page', 1))
    return render(request, 'fees/scholarships/partials/_application_results.html', {'applications': page, 'program': program})


@login_required
def scholarship_program_edit(request, pk):
    program = get_object_or_404(ScholarshipProgram, pk=pk)
    if request.method == 'POST':
        form = ScholarshipProgramForm(request.POST, instance=program)
        if form.is_valid():
            try:
                program = form.save()
                is_htmx = request.headers.get('HX-Request') == 'true'
                if is_htmx:
                    r = HttpResponse()
                    r['HX-Alert-Message'] = f"Scholarship program '{program.name}' updated successfully!"
                    r['HX-Alert-Type']    = 'success'
                    r['HX-Alert-Title']   = 'Updated!'
                    r['HX-Redirect']      = reverse('fees:scholarship_program_detail', kwargs={'pk': program.pk})
                    return r
                messages.success(request, f"Scholarship program '{program.name}' updated successfully!", extra_tags='sweetalert')
                return redirect('fees:scholarship_program_detail', pk=program.pk)
            except Exception as e:
                logger.error(f"Error updating scholarship program: {e}")
                is_htmx = request.headers.get('HX-Request') == 'true'
                if is_htmx:
                    r = HttpResponse()
                    r['HX-Alert-Message'] = f'Error updating program: {str(e)}'
                    r['HX-Alert-Type']    = 'error'
                    r['HX-Alert-Title']   = 'Error!'
                    return r
                messages.error(request, f'Error updating program: {str(e)}')
    else:
        form = ScholarshipProgramForm(instance=program)

    existing_sessions = program.valid_sessions.all()
    return render(request, 'fees/scholarships/program_form.html', {
        'form': form, 'program': program,
        'title': f'Edit Scholarship Program — {program.name}', 'submit_text': 'Update Program', 'submit_icon': 'fa-save',
        'all_sessions':   AcademicSession.objects.filter(is_active=True).order_by('-start_date'),
        'all_categories': FeesCategory.objects.filter(is_active=True).order_by('display_group__display_order', 'display_order'),
        'initial_session_scope': 'specific' if existing_sessions.exists() else 'all',
    })


@login_required
@require_http_methods(["POST"])
def scholarship_program_delete(request, pk):
    program = get_object_or_404(ScholarshipProgram, pk=pk)
    if program.student_scholarships.filter(status='ACTIVE').exists():
        is_htmx = request.headers.get('HX-Request') == 'true'
        if is_htmx:
            r = HttpResponse()
            r['HX-Alert-Message'] = f"Cannot delete '{program.name}' — it has active scholarships"
            r['HX-Alert-Type']    = 'error'
            r['HX-Alert-Title']   = 'Cannot Delete'
            r['HX-Close-Modal']   = 'true'
            return r
        messages.error(request, f"Cannot delete '{program.name}' — it has active scholarships", extra_tags='sweetalert-error')
        return redirect('fees:scholarship_program_detail', pk=pk)
    name = program.name
    program.delete()
    is_htmx = request.headers.get('HX-Request') == 'true'
    if is_htmx:
        r = HttpResponse()
        r['HX-Alert-Message'] = f"Scholarship program '{name}' deleted successfully"
        r['HX-Alert-Type']    = 'success'
        r['HX-Alert-Title']   = 'Deleted!'
        r['HX-Close-Modal']   = 'true'
        r['HX-Redirect']      = reverse('fees:scholarship_program_list')
        return r
    messages.success(request, f"Scholarship program '{name}' deleted successfully", extra_tags='sweetalert')
    return redirect('fees:scholarship_program_list')


@login_required
@require_http_methods(["POST"])
def scholarship_program_activate(request, pk):
    program = get_object_or_404(ScholarshipProgram, pk=pk)
    try:
        program.is_active = True
        program.save()
        is_htmx = request.headers.get('HX-Request') == 'true'
        if is_htmx:
            r = HttpResponse()
            r['HX-Alert-Message'] = f"Scholarship program '{program.name}' activated!"
            r['HX-Alert-Type']    = 'success'
            r['HX-Alert-Title']   = 'Program Activated'
            r['HX-Close-Modal']   = 'true'
            r['HX-Redirect']      = reverse('fees:scholarship_program_detail', kwargs={'pk': program.pk})
            return r
        messages.success(request, f"Scholarship program '{program.name}' activated!")
        return redirect('fees:scholarship_program_detail', pk=program.pk)
    except Exception as e:
        logger.error(f"Error activating scholarship program: {e}")
        is_htmx = request.headers.get('HX-Request') == 'true'
        if is_htmx:
            r = HttpResponse()
            r['HX-Alert-Message'] = f'Error activating program: {str(e)}'
            r['HX-Alert-Type']    = 'error'
            r['HX-Alert-Title']   = 'Error'
            return r
        messages.error(request, f'Error activating program: {str(e)}')
        return redirect('fees:scholarship_program_detail', pk=pk)


@login_required
@require_http_methods(["POST"])
def scholarship_program_deactivate(request, pk):
    program = get_object_or_404(ScholarshipProgram, pk=pk)
    try:
        reason = request.POST.get('deactivation_reason', '')
        program.is_active = False
        program.is_accepting_applications = False
        if reason:
            program.description = f"{program.description}\n\nDeactivated: {reason}".strip()
        program.save()
        is_htmx = request.headers.get('HX-Request') == 'true'
        if is_htmx:
            r = HttpResponse()
            r['HX-Alert-Message'] = f"Scholarship program '{program.name}' deactivated!"
            r['HX-Alert-Type']    = 'success'
            r['HX-Alert-Title']   = 'Program Deactivated'
            r['HX-Close-Modal']   = 'true'
            r['HX-Redirect']      = reverse('fees:scholarship_program_detail', kwargs={'pk': program.pk})
            return r
        messages.success(request, f"Scholarship program '{program.name}' deactivated!")
        return redirect('fees:scholarship_program_detail', pk=program.pk)
    except Exception as e:
        logger.error(f"Error deactivating scholarship program: {e}")
        is_htmx = request.headers.get('HX-Request') == 'true'
        if is_htmx:
            r = HttpResponse()
            r['HX-Alert-Message'] = f'Error deactivating program: {str(e)}'
            r['HX-Alert-Type']    = 'error'
            r['HX-Alert-Title']   = 'Error'
            return r
        messages.error(request, f'Error deactivating program: {str(e)}')
        return redirect('fees:scholarship_program_detail', pk=pk)


@login_required
@require_http_methods(["POST"])
def scholarship_toggle_accepting(request, pk):
    program = get_object_or_404(ScholarshipProgram, pk=pk)
    try:
        program.is_accepting_applications = not program.is_accepting_applications
        program.save()
        status  = "accepting" if program.is_accepting_applications else "not accepting"
        is_htmx = request.headers.get('HX-Request') == 'true'
        if is_htmx:
            r = HttpResponse()
            r['HX-Alert-Message'] = f"Program is now {status} applications!"
            r['HX-Alert-Type']    = 'success'
            r['HX-Alert-Title']   = 'Status Updated'
            r['HX-Close-Modal']   = 'true'
            r['HX-Redirect']      = reverse('fees:scholarship_program_detail', kwargs={'pk': program.pk})
            return r
        messages.success(request, f"Program is now {status} applications!")
        return redirect('fees:scholarship_program_detail', pk=program.pk)
    except Exception as e:
        logger.error(f"Error toggling application status: {e}")
        is_htmx = request.headers.get('HX-Request') == 'true'
        if is_htmx:
            r = HttpResponse()
            r['HX-Alert-Message'] = f'Error updating status: {str(e)}'
            r['HX-Alert-Type']    = 'error'
            r['HX-Alert-Title']   = 'Error'
            return r
        messages.error(request, f'Error updating status: {str(e)}')
        return redirect('fees:scholarship_program_detail', pk=pk)


@login_required
def scholarship_program_list_print_view(request):
    FIELD_NAMES_FULL = {
        'code':               'Code',
        'name':               'Program Name',
        'scholarship_type':   'Scholarship Type',
        'program_type':       'Program Type',
        'discount_type':      'Discount Type',
        'discount_summary':   'Discount Summary',
        'total_budget':       'Total Budget',
        'budget_used':        'Budget Used',
        'budget_remaining':   'Remaining Budget',
        'max_recipients':     'Max Recipients',
        'current_recipients': 'Current Recipients',
        'sponsor_name':       'Sponsor',
        'is_active':          'Active',
        'accepting_apps':     'Accepting Applications',
    }
    FIELD_NAMES_SHORT = {
        'code':               'Code',
        'name':               'Program',
        'scholarship_type':   'Type',
        'program_type':       'Prog. Type',
        'discount_type':      'Disc. Type',
        'discount_summary':   'Discount',
        'total_budget':       'Budget',
        'budget_used':        'Used',
        'budget_remaining':   'Remaining',
        'max_recipients':     'Max',
        'current_recipients': 'Recipients',
        'sponsor_name':       'Sponsor',
        'is_active':          'Active',
        'accepting_apps':     'Accepting',
    }
    DEFAULT_FIELDS  = ['code', 'name', 'scholarship_type', 'program_type', 'discount_summary', 'total_budget', 'budget_used', 'current_recipients', 'is_active', 'accepting_apps']
    selected_fields = request.GET.getlist('fields') or DEFAULT_FIELDS
    short_headers   = request.GET.get('short_headers', 'false').lower() == 'true'
    landscape       = request.GET.get('landscape', 'false').lower() == 'true'
    include_stats   = request.GET.get('include_stats', 'true').lower() == 'true'
    field_names     = FIELD_NAMES_SHORT if short_headers else FIELD_NAMES_FULL
    programs        = get_filtered_scholarship_programs(request)

    stats = None
    if include_stats:
        stats = {
            'total':                  programs.count(),
            'active':                 programs.filter(is_active=True).count(),
            'accepting_applications': programs.filter(is_accepting_applications=True).count(),
            'total_budget':           programs.aggregate(Sum('total_budget_amount'))['total_budget_amount__sum'] or Decimal('0.00'),
            'total_used':             programs.aggregate(Sum('current_budget_used'))['current_budget_used__sum'] or Decimal('0.00'),
            'total_recipients':       programs.aggregate(Sum('current_recipient_count'))['current_recipient_count__sum'] or 0,
        }

    return render(request, 'fees/scholarships/print_program_list.html', {
        **get_print_school_context(request),
        'programs':             programs,
        'stats':                stats,
        'selected_fields':      selected_fields,
        'selected_field_names': [field_names.get(f, f.replace('_', ' ').title()) for f in selected_fields],
        'field_names':          field_names,
        'short_headers':        short_headers,
        'landscape':            landscape,
        'now':                  timezone.now(),
        'print_date':           get_school_today(),
        'printed_by':           request.user.get_full_name() or request.user.username,
        'title':                'Scholarship Programs',
    })


@login_required
def export_scholarship_programs_excel(request):
    ALL_COLUMNS = [
        ('code',               'Code',               lambda o: o.code),
        ('name',               'Name',               lambda o: o.name),
        ('scholarship_type',   'Scholarship Type',   lambda o: o.get_scholarship_type_display()),
        ('program_type',       'Program Type',       lambda o: o.get_program_type_display()),
        ('discount_type',      'Discount Type',      lambda o: o.get_discount_type_display()),
        ('discount_summary',   'Discount Summary',   lambda o: o.get_discount_summary()),
        ('total_budget',       'Total Budget',       lambda o: float(o.total_budget_amount) if o.total_budget_amount else ''),
        ('budget_used',        'Budget Used',        lambda o: float(o.current_budget_used or 0)),
        ('budget_remaining',   'Budget Remaining',   lambda o: float(o.get_remaining_budget()) if o.get_remaining_budget() is not None else 'Unlimited'),
        ('max_recipients',     'Max Recipients',     lambda o: o.maximum_recipients or 'Unlimited'),
        ('current_recipients', 'Current Recipients', lambda o: o.current_recipient_count or 0),
        ('renewal_policy',     'Renewal Policy',     lambda o: o.get_renewal_policy_display()),
        ('max_duration',       'Max Duration (Yrs)', lambda o: o.maximum_duration_years),
        ('sponsor_name',       'Sponsor',            lambda o: o.sponsor_name or ''),
        ('is_active',          'Active',             lambda o: 'Yes' if o.is_active else 'No'),
        ('accepting_apps',     'Accepting Apps',     lambda o: 'Yes' if o.is_accepting_applications else 'No'),
        ('app_start',          'App. Opens',         lambda o: o.application_start_date.strftime('%Y-%m-%d') if o.application_start_date else ''),
        ('app_end',            'App. Closes',        lambda o: o.application_end_date.strftime('%Y-%m-%d') if o.application_end_date else ''),
    ]
    DEFAULT_FIELDS = ['code', 'name', 'scholarship_type', 'program_type', 'discount_type', 'discount_summary', 'total_budget', 'budget_used', 'current_recipients', 'is_active', 'accepting_apps']
    programs = get_filtered_scholarship_programs(request)
    columns  = _resolve_columns(ALL_COLUMNS, request.GET.getlist('fields'), DEFAULT_FIELDS)
    return _xlsx_response(_make_workbook('Scholarship Programs', columns, programs), 'scholarship_programs')


# =============================================================================
# SCHOLARSHIP APPLICATIONS
# =============================================================================

def get_filtered_scholarship_applications(request):
    applications = StudentScholarshipApplication.objects.select_related(
        'student', 'scholarship_program', 'academic_session'
    ).order_by('-application_date')

    query            = request.GET.get('q', '').strip()
    status           = request.GET.get('status', '')
    program          = request.GET.get('program', '')
    academic_session = request.GET.get('academic_session', '')
    date_from        = request.GET.get('date_from', '')
    date_to          = request.GET.get('date_to', '')

    if query:
        words = query.split()
        q = Q()
        for w in words:
            q &= (Q(application_number__icontains=w) |
                  Q(student__first_name__icontains=w) |
                  Q(student__last_name__icontains=w) |
                  Q(student__admission_number__icontains=w) |
                  Q(scholarship_program__name__icontains=w))
        applications = applications.filter(q)

    if status:
        applications = applications.filter(status=status)
    if program:
        applications = applications.filter(scholarship_program_id=program)
    if academic_session:
        applications = applications.filter(academic_session_id=academic_session)
    if date_from:
        applications = applications.filter(application_date__gte=date_from)
    if date_to:
        applications = applications.filter(application_date__lte=date_to)

    return applications


@login_required
def scholarship_application_list(request):
    filter_form  = ScholarshipApplicationFilterForm(request.GET or None)
    applications = get_filtered_scholarship_applications(request)

    stats = {
        'total':        applications.count(),
        'submitted':    applications.filter(status='SUBMITTED').count(),
        'under_review': applications.filter(status='UNDER_REVIEW').count(),
        'approved':     applications.filter(status='APPROVED').count(),
        'rejected':     applications.filter(status='REJECTED').count(),
        'waitlisted':   applications.filter(status='WAITLISTED').count(),
        'withdrawn':    applications.filter(status='WITHDRAWN').count(),
    }

    paginator        = Paginator(applications, 20)
    applications_page = paginator.get_page(request.GET.get('page', 1))
    is_htmx          = request.headers.get('HX-Request') == 'true'

    context = {
        'applications_page': applications_page,
        'paginator':         paginator,
        'stats':             stats,
        'filter_form':       filter_form,
        'is_htmx':           is_htmx,
    }
    if is_htmx:
        return render(request, 'fees/scholarships/partials/_application_list_results.html', context)
    return render(request, 'fees/scholarships/application_list.html', context)


@login_required
def scholarship_application_create(request):
    program_id = request.GET.get('program')
    program    = get_object_or_404(ScholarshipProgram, pk=program_id) if program_id else None

    if request.method == 'POST':
        form = StudentScholarshipApplicationForm(request.POST, request.FILES)
        if form.is_valid():
            try:
                application = form.save(commit=False)
                if not application.application_number:
                    application.application_number = application.scholarship_program.code + '-APP-' + get_school_today().strftime('%Y%m%d-%H%M%S')
                application.save()
                form.save_m2m()
                is_htmx = request.headers.get('HX-Request') == 'true'
                if is_htmx:
                    r = HttpResponse()
                    r['HX-Alert-Message'] = f"Application '{application.application_number}' submitted successfully!"
                    r['HX-Alert-Type']    = 'success'
                    r['HX-Alert-Title']   = 'Application Submitted!'
                    r['HX-Redirect']      = reverse('fees:scholarship_application_detail', kwargs={'pk': application.pk})
                    return r
                messages.success(request, f"Application '{application.application_number}' submitted successfully!", extra_tags='sweetalert')
                return redirect('fees:scholarship_application_detail', pk=application.pk)
            except Exception as e:
                logger.error(f"Error creating scholarship application: {e}")
                is_htmx = request.headers.get('HX-Request') == 'true'
                if is_htmx:
                    r = HttpResponse()
                    r['HX-Alert-Message'] = f'Error submitting application: {str(e)}'
                    r['HX-Alert-Type']    = 'error'
                    r['HX-Alert-Title']   = 'Error!'
                    return r
                messages.error(request, f'Error submitting application: {str(e)}')
    else:
        initial = {'scholarship_program': program} if program else {}
        form    = StudentScholarshipApplicationForm(initial=initial)

    return render(request, 'fees/scholarships/application_form.html', {
        'form': form, 'program': program,
        'title': 'Submit Scholarship Application', 'submit_text': 'Submit Application',
    })


@login_required
def scholarship_application_detail(request, pk):
    application = get_object_or_404(
        StudentScholarshipApplication.objects.select_related(
            'student', 'scholarship_program', 'academic_session'
        ),
        pk=pk,
    )
    existing_scholarship = StudentScholarship.objects.filter(
        student=application.student,
        scholarship_program=application.scholarship_program,
        status='ACTIVE',
    ).first()
 
    return render(request, 'fees/scholarships/application_detail.html', {
        'application':          application,
        'existing_scholarship': existing_scholarship,
        'today':                get_school_today(),
    })


@login_required
def scholarship_application_edit(request, pk):
    application = get_object_or_404(StudentScholarshipApplication, pk=pk)
    if application.status not in ['DRAFT', 'SUBMITTED']:
        messages.error(request, 'Only draft or submitted applications can be edited.')
        return redirect('fees:scholarship_application_detail', pk=pk)

    if request.method == 'POST':
        form = StudentScholarshipApplicationForm(request.POST, request.FILES, instance=application)
        if form.is_valid():
            try:
                application = form.save()
                is_htmx     = request.headers.get('HX-Request') == 'true'
                if is_htmx:
                    r = HttpResponse()
                    r['HX-Alert-Message'] = f"Application '{application.application_number}' updated successfully!"
                    r['HX-Alert-Type']    = 'success'
                    r['HX-Alert-Title']   = 'Updated!'
                    r['HX-Redirect']      = reverse('fees:scholarship_application_detail', kwargs={'pk': application.pk})
                    return r
                messages.success(request, f"Application '{application.application_number}' updated successfully!", extra_tags='sweetalert')
                return redirect('fees:scholarship_application_detail', pk=application.pk)
            except Exception as e:
                logger.error(f"Error updating scholarship application: {e}")
                is_htmx = request.headers.get('HX-Request') == 'true'
                if is_htmx:
                    r = HttpResponse()
                    r['HX-Alert-Message'] = f'Error updating application: {str(e)}'
                    r['HX-Alert-Type']    = 'error'
                    r['HX-Alert-Title']   = 'Error!'
                    return r
                messages.error(request, f'Error updating application: {str(e)}')
    else:
        form = StudentScholarshipApplicationForm(instance=application)

    return render(request, 'fees/scholarships/application_form.html', {
        'form': form, 'application': application,
        'title': f'Edit Application — {application.application_number}', 'submit_text': 'Update Application',
    })


@login_required
@require_http_methods(["POST"])
def scholarship_application_review(request, pk):
    application = get_object_or_404(StudentScholarshipApplication, pk=pk)
    form        = ScholarshipApplicationApprovalForm(application, request.user, request.POST)
 
    if form.is_valid():
        try:
            decision       = form.cleaned_data['decision']
            reviewer_notes = form.cleaned_data.get('reviewer_notes', '')
            with transaction.atomic():
                application.status         = decision
                application.reviewed_by_id = str(request.user.id)   # CharField
                application.reviewed_at    = get_school_current_time()
                application.reviewer_notes = reviewer_notes
 
                if decision == 'APPROVED':
                    application.approved_amount = form.cleaned_data.get(
                        'approved_amount', application.requested_amount
                    )
                    application.effective_date = form.cleaned_data.get(
                        'effective_date', get_school_today()
                    )
                    application.save()
 
                    if form.cleaned_data.get('auto_create_scholarship'):
                        StudentScholarship.objects.create(
                            student=application.student,
                            scholarship_program=application.scholarship_program,
                            amount_awarded=application.approved_amount or Decimal('0.00'),
                            status='ACTIVE',
                            start_date=application.effective_date,
                            awarded_by_id=str(request.user.id),   # CharField
                            notes=f"Auto-created from application {application.application_number}",
                        )
                else:
                    application.save()
 
            is_htmx = request.headers.get('HX-Request') == 'true'
            msg     = f"Application {application.application_number}: {decision}"
            if is_htmx:
                r = HttpResponse()
                r['HX-Alert-Message'] = msg
                r['HX-Alert-Type']    = 'success'
                r['HX-Alert-Title']   = 'Decision Recorded'
                r['HX-Close-Modal']   = 'true'
                r['HX-Redirect']      = reverse(
                    'fees:scholarship_application_detail',
                    kwargs={'pk': application.pk},
                )
                return r
            messages.success(request, msg)
            return redirect('fees:scholarship_application_detail', pk=application.pk)
 
        except Exception as e:
            logger.error(f"Error reviewing scholarship application: {e}", exc_info=True)
            is_htmx = request.headers.get('HX-Request') == 'true'
            if is_htmx:
                r = HttpResponse()
                r['HX-Alert-Message'] = f'Error recording decision: {str(e)}'
                r['HX-Alert-Type']    = 'error'
                r['HX-Alert-Title']   = 'Error'
                return r
            messages.error(request, f'Error recording decision: {str(e)}')
            return redirect('fees:scholarship_application_detail', pk=pk)
 
    for field, errors in form.errors.items():
        for error in errors:
            messages.error(request, f'{field}: {error}')
    return redirect('fees:scholarship_application_detail', pk=pk)


@login_required
@require_http_methods(["POST"])
def scholarship_application_delete(request, pk):
    application = get_object_or_404(StudentScholarshipApplication, pk=pk)
    if application.status not in ['DRAFT', 'WITHDRAWN']:
        is_htmx = request.headers.get('HX-Request') == 'true'
        if is_htmx:
            r = HttpResponse()
            r['HX-Alert-Message'] = f"Cannot delete application with status '{application.status}'"
            r['HX-Alert-Type']    = 'error'
            r['HX-Alert-Title']   = 'Cannot Delete'
            r['HX-Close-Modal']   = 'true'
            return r
        messages.error(request, f"Cannot delete application with status '{application.status}'", extra_tags='sweetalert-error')
        return redirect('fees:scholarship_application_detail', pk=pk)
    number = application.application_number
    application.delete()
    is_htmx = request.headers.get('HX-Request') == 'true'
    if is_htmx:
        r = HttpResponse()
        r['HX-Alert-Message'] = f"Application '{number}' deleted successfully"
        r['HX-Alert-Type']    = 'success'
        r['HX-Alert-Title']   = 'Deleted!'
        r['HX-Close-Modal']   = 'true'
        r['HX-Redirect']      = reverse('fees:scholarship_application_list')
        return r
    messages.success(request, f"Application '{number}' deleted successfully", extra_tags='sweetalert')
    return redirect('fees:scholarship_application_list')


@login_required
def scholarship_application_list_print_view(request):
    FIELD_NAMES_FULL = {
        'application_number': 'Application Number',
        'student_id':         'Admission Number',
        'student_name':       'Student Name',
        'program':            'Scholarship Program',
        'scholarship_type':   'Scholarship Type',
        'academic_session':   'Academic Session',
        'application_date':   'Application Date',
        'requested_amount':   'Amount Requested',
        'approved_amount':    'Amount Approved',
        'status':             'Status',
        'reviewer':           'Reviewed By',
        'reviewed_at':        'Review Date',
        'reviewer_notes':     'Review Notes',
    }
    FIELD_NAMES_SHORT = {
        'application_number': 'App. No.',
        'student_id':         'Adm. No.',
        'student_name':       'Student',
        'program':            'Program',
        'scholarship_type':   'Type',
        'academic_session':   'Session',
        'application_date':   'Date',
        'requested_amount':   'Requested',
        'approved_amount':    'Approved',
        'status':             'Status',
        'reviewer':           'Reviewer',
        'reviewed_at':        'Review Date',
        'reviewer_notes':     'Notes',
    }
    DEFAULT_FIELDS  = ['application_number', 'student_id', 'student_name', 'program', 'scholarship_type', 'application_date', 'requested_amount', 'approved_amount', 'status']
    selected_fields = request.GET.getlist('fields') or DEFAULT_FIELDS
    short_headers   = request.GET.get('short_headers', 'false').lower() == 'true'
    landscape       = request.GET.get('landscape', 'true').lower() == 'true'
    include_stats   = request.GET.get('include_stats', 'true').lower() == 'true'
    field_names     = FIELD_NAMES_SHORT if short_headers else FIELD_NAMES_FULL
    applications    = get_filtered_scholarship_applications(request)

    stats = None
    if include_stats:
        stats = {
            'total':        applications.count(),
            'submitted':    applications.filter(status='SUBMITTED').count(),
            'under_review': applications.filter(status='UNDER_REVIEW').count(),
            'approved':     applications.filter(status='APPROVED').count(),
            'rejected':     applications.filter(status='REJECTED').count(),
        }

    if applications.count() > MAX_PRINT_RECORDS:
        applications = applications[:MAX_PRINT_RECORDS]

    return render(request, 'fees/scholarships/print_application_list.html', {
        **get_print_school_context(request),
        'applications':         applications,
        'stats':                stats,
        'selected_fields':      selected_fields,
        'selected_field_names': [field_names.get(f, f.replace('_', ' ').title()) for f in selected_fields],
        'field_names':          field_names,
        'short_headers':        short_headers,
        'landscape':            landscape,
        'now':                  timezone.now(),
        'print_date':           get_school_today(),
        'printed_by':           request.user.get_full_name() or request.user.username,
        'title':                'Scholarship Applications',
    })


@login_required
def export_scholarship_applications_excel(request):
    def _reviewed_by(o):
        user = o.get_reviewed_by_user()
        return user.get_full_name() if user else ''
 
    ALL_COLUMNS = [
        ('application_number', 'App. Number',      lambda o: o.application_number),
        ('student_id',         'Admission No.',    lambda o: o.student.admission_number),
        ('student_name',       'Student Name',     lambda o: o.student.get_full_name()),
        ('program',            'Program',          lambda o: o.scholarship_program.name),
        ('scholarship_type',   'Type',             lambda o: o.scholarship_program.get_scholarship_type_display()),
        ('academic_session',   'Session',          lambda o: o.academic_session.name if o.academic_session else ''),
        ('application_date',   'Application Date', lambda o: o.application_date.strftime('%Y-%m-%d') if o.application_date else ''),
        ('requested_amount',   'Amount Requested', lambda o: float(o.requested_amount) if o.requested_amount else ''),
        ('approved_amount',    'Amount Approved',  lambda o: float(o.approved_amount) if o.approved_amount else ''),
        ('status',             'Status',           lambda o: o.get_status_display()),
        ('reviewer',           'Reviewed By',      _reviewed_by),
        ('reviewed_at',        'Review Date',      lambda o: o.reviewed_at.strftime('%Y-%m-%d') if o.reviewed_at else ''),
        ('reviewer_notes',     'Review Notes',     lambda o: o.reviewer_notes or ''),
    ]
    DEFAULT_FIELDS = [
        'application_number', 'student_id', 'student_name', 'program',
        'scholarship_type', 'application_date', 'requested_amount',
        'approved_amount', 'status',
    ]
    applications = get_filtered_scholarship_applications(request)
    columns      = _resolve_columns(ALL_COLUMNS, request.GET.getlist('fields'), DEFAULT_FIELDS)
    return _xlsx_response(
        _make_workbook('Scholarship Applications', columns, applications),
        'scholarship_applications',
    )


# =============================================================================
# STUDENT SCHOLARSHIPS
# =============================================================================

def get_filtered_student_scholarships(request):
    scholarships = StudentScholarship.objects.select_related(
        'student', 'scholarship_program'
    ).annotate(
        status_order=Case(
            When(status='ACTIVE',    then=Value(0)),
            When(status='PENDING',   then=Value(1)),
            When(status='SUSPENDED', then=Value(2)),
            When(status='COMPLETED', then=Value(3)),
            When(status='TERMINATED',then=Value(4)),
            default=Value(5),
            output_field=IntegerField(),
        )
    ).order_by('status_order', '-start_date')
 
    query            = request.GET.get('q', '').strip()
    status           = request.GET.get('status', '')
    program          = request.GET.get('program', '')
    academic_session = request.GET.get('academic_session', '')
    discount_mode    = request.GET.get('discount_mode', '')
 
    if query:
        words = query.split()
        q = Q()
        for w in words:
            q &= (Q(student__first_name__icontains=w) |
                  Q(student__last_name__icontains=w) |
                  Q(student__admission_number__icontains=w) |
                  Q(scholarship_program__name__icontains=w) |
                  Q(scholarship_program__code__icontains=w))
        scholarships = scholarships.filter(q)
 
    if status:
        scholarships = scholarships.filter(status=status)
    if program:
        scholarships = scholarships.filter(scholarship_program_id=program)
    if academic_session:
        scholarships = scholarships.filter(academic_session_id=academic_session)
    if discount_mode == 'global':
        scholarships = scholarships.filter(use_category_specific_discounts=False)
    elif discount_mode == 'category':
        scholarships = scholarships.filter(use_category_specific_discounts=True)
 
    return scholarships


@login_required
def student_scholarship_list(request):
    filter_form  = StudentScholarshipFilterForm(request.GET or None)
    scholarships = get_filtered_student_scholarships(request)

    stats = {
        'total':           scholarships.count(),
        'active':          scholarships.filter(status='ACTIVE').count(),
        'suspended':       scholarships.filter(status='SUSPENDED').count(),
        'terminated':      scholarships.filter(status='TERMINATED').count(),
        'completed':       scholarships.filter(status='COMPLETED').count(),
        'global_discount': scholarships.filter(use_category_specific_discounts=False).count(),
        'category_disc':   scholarships.filter(use_category_specific_discounts=True).count(),
        'total_awarded':   scholarships.aggregate(Sum('amount_awarded'))['amount_awarded__sum'] or 0,
        'total_used':      scholarships.aggregate(Sum('total_amount_used'))['total_amount_used__sum'] or 0,
    }

    paginator        = Paginator(scholarships, 20)
    scholarships_page = paginator.get_page(request.GET.get('page', 1))
    is_htmx          = request.headers.get('HX-Request') == 'true'

    context = {
        'scholarships_page': scholarships_page,
        'paginator':         paginator,
        'stats':             stats,
        'filter_form':       filter_form,
        'is_htmx':           is_htmx,
    }
    if is_htmx:
        return render(request, 'fees/scholarships/partials/_scholarship_results.html', context)
    return render(request, 'fees/scholarships/scholarship_list.html', context)


@login_required
def student_scholarship_create(request):
    if request.method == 'POST':
        form = StudentScholarshipForm(request.POST)
        if form.is_valid():
            try:
                scholarship                = form.save(commit=False)
                scholarship.awarded_by_id  = str(request.user.id)   # CharField
                scholarship.awarded_date   = get_school_today()
                scholarship.save()
                form.save_m2m()
                is_htmx = request.headers.get('HX-Request') == 'true'
                if is_htmx:
                    r = HttpResponse()
                    r['HX-Alert-Message'] = f"Scholarship awarded to {scholarship.student.get_full_name()} successfully!"
                    r['HX-Alert-Type']    = 'success'
                    r['HX-Alert-Title']   = 'Scholarship Awarded!'
                    r['HX-Redirect']      = reverse(
                        'fees:student_scholarship_detail',
                        kwargs={'pk': scholarship.pk},
                    )
                    return r
                messages.success(
                    request,
                    f"Scholarship awarded to {scholarship.student.get_full_name()} successfully!",
                    extra_tags='sweetalert',
                )
                return redirect('fees:student_scholarship_detail', pk=scholarship.pk)
            except Exception as e:
                logger.error(f"Error creating scholarship: {e}")
                is_htmx = request.headers.get('HX-Request') == 'true'
                if is_htmx:
                    r = HttpResponse()
                    r['HX-Alert-Message'] = f'Error awarding scholarship: {str(e)}'
                    r['HX-Alert-Type']    = 'error'
                    r['HX-Alert-Title']   = 'Error!'
                    return r
                messages.error(request, f'Error awarding scholarship: {str(e)}')
    else:
        form = StudentScholarshipForm()
 
    return render(request, 'fees/scholarships/scholarship_form.html', {
        'form': form, 'title': 'Award Scholarship', 'submit_text': 'Award Scholarship',
    })


@login_required
def student_scholarship_detail(request, pk):
    scholarship = get_object_or_404(
        StudentScholarship.objects.select_related(
            'student', 'scholarship_program'
        ).prefetch_related('application_logs__invoice'),
        pk=pk,
    )
    application_logs = ScholarshipApplicationLog.objects.filter(
        scholarship=scholarship
    ).select_related('invoice').order_by('-application_date')[:20]
 
    return render(request, 'fees/scholarships/scholarship_detail.html', {
        'scholarship':      scholarship,
        'application_logs': application_logs,
        'discount_summary': scholarship.get_discount_display_summary(),
        'today':            get_school_today(),
    })


@login_required
def student_scholarship_edit(request, pk):
    scholarship = get_object_or_404(StudentScholarship, pk=pk)
    if request.method == 'POST':
        form = StudentScholarshipForm(request.POST, instance=scholarship)
        if form.is_valid():
            try:
                scholarship = form.save()
                is_htmx     = request.headers.get('HX-Request') == 'true'
                if is_htmx:
                    r = HttpResponse()
                    r['HX-Alert-Message'] = f"Scholarship for {scholarship.student.get_full_name()} updated successfully!"
                    r['HX-Alert-Type']    = 'success'
                    r['HX-Alert-Title']   = 'Updated!'
                    r['HX-Redirect']      = reverse('fees:student_scholarship_detail', kwargs={'pk': scholarship.pk})
                    return r
                messages.success(request, f"Scholarship for {scholarship.student.get_full_name()} updated successfully!", extra_tags='sweetalert')
                return redirect('fees:student_scholarship_detail', pk=scholarship.pk)
            except Exception as e:
                logger.error(f"Error updating scholarship: {e}")
                is_htmx = request.headers.get('HX-Request') == 'true'
                if is_htmx:
                    r = HttpResponse()
                    r['HX-Alert-Message'] = f'Error updating scholarship: {str(e)}'
                    r['HX-Alert-Type']    = 'error'
                    r['HX-Alert-Title']   = 'Error!'
                    return r
                messages.error(request, f'Error updating scholarship: {str(e)}')
    else:
        form = StudentScholarshipForm(instance=scholarship)

    return render(request, 'fees/scholarships/scholarship_form.html', {
        'form': form, 'scholarship': scholarship,
        'title': f'Edit Scholarship — {scholarship.student.get_full_name()}', 'submit_text': 'Update Scholarship',
    })


@login_required
@require_http_methods(["POST"])
def student_scholarship_suspend(request, pk):
    scholarship = get_object_or_404(StudentScholarship, pk=pk)
    if scholarship.status != 'ACTIVE':
        messages.error(request, 'Only active scholarships can be suspended.')
        return redirect('fees:student_scholarship_detail', pk=pk)
    try:
        reason                 = request.POST.get('suspension_reason', '')
        scholarship.status     = 'SUSPENDED'
        scholarship.award_notes = f"{scholarship.award_notes}\n\nSuspended on {get_school_today()}: {reason}".strip()
        scholarship.save()
        is_htmx = request.headers.get('HX-Request') == 'true'
        if is_htmx:
            r = HttpResponse()
            r['HX-Alert-Message'] = f"Scholarship for {scholarship.student.get_full_name()} suspended!"
            r['HX-Alert-Type']    = 'warning'
            r['HX-Alert-Title']   = 'Scholarship Suspended'
            r['HX-Close-Modal']   = 'true'
            r['HX-Redirect']      = reverse('fees:student_scholarship_detail', kwargs={'pk': scholarship.pk})
            return r
        messages.success(request, f"Scholarship for {scholarship.student.get_full_name()} suspended!")
        return redirect('fees:student_scholarship_detail', pk=scholarship.pk)
    except Exception as e:
        logger.error(f"Error suspending scholarship: {e}")
        is_htmx = request.headers.get('HX-Request') == 'true'
        if is_htmx:
            r = HttpResponse()
            r['HX-Alert-Message'] = f'Error suspending scholarship: {str(e)}'
            r['HX-Alert-Type']    = 'error'
            r['HX-Alert-Title']   = 'Error'
            return r
        messages.error(request, f'Error suspending scholarship: {str(e)}')
        return redirect('fees:student_scholarship_detail', pk=pk)


@login_required
@require_http_methods(["POST"])
def student_scholarship_terminate(request, pk):
    scholarship = get_object_or_404(StudentScholarship, pk=pk)
    if scholarship.status == 'TERMINATED':
        messages.error(request, 'Scholarship is already terminated.')
        return redirect('fees:student_scholarship_detail', pk=pk)
    try:
        reason                 = request.POST.get('termination_reason', '')
        scholarship.status     = 'TERMINATED'
        scholarship.end_date   = get_school_today()
        scholarship.award_notes = f"{scholarship.award_notes}\n\nTerminated on {get_school_today()}: {reason}".strip()
        scholarship.save()
        is_htmx = request.headers.get('HX-Request') == 'true'
        if is_htmx:
            r = HttpResponse()
            r['HX-Alert-Message'] = f"Scholarship for {scholarship.student.get_full_name()} terminated!"
            r['HX-Alert-Type']    = 'warning'
            r['HX-Alert-Title']   = 'Scholarship Terminated'
            r['HX-Close-Modal']   = 'true'
            r['HX-Redirect']      = reverse('fees:student_scholarship_detail', kwargs={'pk': scholarship.pk})
            return r
        messages.success(request, f"Scholarship for {scholarship.student.get_full_name()} terminated!")
        return redirect('fees:student_scholarship_detail', pk=scholarship.pk)
    except Exception as e:
        logger.error(f"Error terminating scholarship: {e}", exc_info=True)
        is_htmx = request.headers.get('HX-Request') == 'true'
        if is_htmx:
            r = HttpResponse()
            r['HX-Alert-Message'] = f'Error terminating scholarship: {str(e)}'
            r['HX-Alert-Type']    = 'error'
            r['HX-Alert-Title']   = 'Error'
            return r
        messages.error(request, f'Error terminating scholarship: {str(e)}')
        return redirect('fees:student_scholarship_detail', pk=pk)

@login_required
@require_http_methods(["POST"])
def student_scholarship_complete(request, pk):
    scholarship = get_object_or_404(StudentScholarship, pk=pk)

    if scholarship.status != 'ACTIVE':
        is_htmx = request.headers.get('HX-Request') == 'true'
        msg = f"Only active scholarships can be marked as completed. Current status: {scholarship.get_status_display()}"
        if is_htmx:
            r = HttpResponse()
            r['HX-Alert-Message'] = msg
            r['HX-Alert-Type']    = 'error'
            r['HX-Alert-Title']   = 'Cannot Complete'
            r['HX-Close-Modal']   = 'true'
            return r
        messages.error(request, msg)
        return redirect('fees:student_scholarship_detail', pk=pk)

    try:
        today                      = get_school_today()
        completion_notes           = request.POST.get('completion_notes', '').strip()
        scholarship.status         = 'COMPLETED'
        scholarship.end_date       = scholarship.end_date or today
        scholarship.notes          = (
            f"{scholarship.notes}\n\nCompleted on {today}: {completion_notes}".strip()
            if completion_notes
            else f"{scholarship.notes}\n\nMarked as completed on {today}.".strip()
        )
        scholarship.save(update_fields=['status', 'end_date', 'notes'])

        is_htmx = request.headers.get('HX-Request') == 'true'
        if is_htmx:
            r = HttpResponse()
            r['HX-Alert-Message'] = f"Scholarship for {scholarship.student.get_full_name()} marked as completed."
            r['HX-Alert-Type']    = 'success'
            r['HX-Alert-Title']   = 'Scholarship Completed'
            r['HX-Close-Modal']   = 'true'
            r['HX-Redirect']      = reverse('fees:student_scholarship_detail', kwargs={'pk': scholarship.pk})
            return r

        messages.success(
            request,
            f"Scholarship for {scholarship.student.get_full_name()} marked as completed.",
            extra_tags='sweetalert',
        )
        return redirect('fees:student_scholarship_detail', pk=pk)

    except Exception as e:
        logger.error(f"Error completing scholarship: {e}", exc_info=True)
        is_htmx = request.headers.get('HX-Request') == 'true'
        if is_htmx:
            r = HttpResponse()
            r['HX-Alert-Message'] = f'Error completing scholarship: {str(e)}'
            r['HX-Alert-Type']    = 'error'
            r['HX-Alert-Title']   = 'Error'
            return r
        messages.error(request, f'Error completing scholarship: {str(e)}')
        return redirect('fees:student_scholarship_detail', pk=pk)
    
@login_required
@require_http_methods(["POST"])
def student_scholarship_delete(request, pk):
    scholarship = get_object_or_404(StudentScholarship, pk=pk)
    if scholarship.status == 'ACTIVE' and scholarship.applications.exists():
        is_htmx = request.headers.get('HX-Request') == 'true'
        if is_htmx:
            r = HttpResponse()
            r['HX-Alert-Message'] = "Cannot delete an active scholarship that has been applied to invoices"
            r['HX-Alert-Type']    = 'error'
            r['HX-Alert-Title']   = 'Cannot Delete'
            r['HX-Close-Modal']   = 'true'
            return r
        messages.error(request, "Cannot delete an active scholarship that has been applied to invoices", extra_tags='sweetalert-error')
        return redirect('fees:student_scholarship_detail', pk=pk)
    student_name = scholarship.student.get_full_name()
    scholarship.delete()
    is_htmx = request.headers.get('HX-Request') == 'true'
    if is_htmx:
        r = HttpResponse()
        r['HX-Alert-Message'] = f"Scholarship for '{student_name}' deleted successfully"
        r['HX-Alert-Type']    = 'success'
        r['HX-Alert-Title']   = 'Deleted!'
        r['HX-Close-Modal']   = 'true'
        r['HX-Redirect']      = reverse('fees:student_scholarship_list')
        return r
    messages.success(request, f"Scholarship for '{student_name}' deleted successfully", extra_tags='sweetalert')
    return redirect('fees:student_scholarship_list')


@login_required
def student_scholarship_list_print_view(request):
    FIELD_NAMES_FULL = {
        'student_id':          'Admission Number',
        'student_name':        'Student Name',
        'program':             'Scholarship Program',
        'scholarship_type':    'Scholarship Type',
        'discount_mode':       'Discount Mode',
        'discount_summary':    'Discount Summary',
        'amount_awarded':      'Amount Awarded',
        'total_used':          'Total Used',
        'balance_remaining':   'Balance Remaining',
        'start_date':          'Start Date',
        'end_date':            'End Date',
        'status':              'Status',
        'awarded_by':          'Awarded By',
        'award_notes':         'Notes',
    }
    FIELD_NAMES_SHORT = {
        'student_id':          'Adm. No.',
        'student_name':        'Student',
        'program':             'Program',
        'scholarship_type':    'Type',
        'discount_mode':       'Disc. Mode',
        'discount_summary':    'Discount',
        'amount_awarded':      'Awarded',
        'total_used':          'Used',
        'balance_remaining':   'Remaining',
        'start_date':          'Start',
        'end_date':            'End',
        'status':              'Status',
        'awarded_by':          'By',
        'award_notes':         'Notes',
    }
    DEFAULT_FIELDS  = ['student_id', 'student_name', 'program', 'scholarship_type', 'discount_summary', 'amount_awarded', 'total_used', 'start_date', 'status']
    selected_fields = request.GET.getlist('fields') or DEFAULT_FIELDS
    short_headers   = request.GET.get('short_headers', 'false').lower() == 'true'
    landscape       = request.GET.get('landscape', 'true').lower() == 'true'
    include_stats   = request.GET.get('include_stats', 'true').lower() == 'true'
    field_names     = FIELD_NAMES_SHORT if short_headers else FIELD_NAMES_FULL
    scholarships    = get_filtered_student_scholarships(request)

    stats = None
    if include_stats:
        stats = {
            'total':         scholarships.count(),
            'active':        scholarships.filter(status='ACTIVE').count(),
            'suspended':     scholarships.filter(status='SUSPENDED').count(),
            'terminated':    scholarships.filter(status='TERMINATED').count(),
            'total_awarded': scholarships.aggregate(Sum('amount_awarded'))['amount_awarded__sum'] or Decimal('0.00'),
            'total_used':    scholarships.aggregate(Sum('total_amount_used'))['total_amount_used__sum'] or Decimal('0.00'),
        }

    if scholarships.count() > MAX_PRINT_RECORDS:
        scholarships = scholarships[:MAX_PRINT_RECORDS]

    return render(request, 'fees/scholarships/print_scholarship_list.html', {
        **get_print_school_context(request),
        'scholarships':         scholarships,
        'stats':                stats,
        'selected_fields':      selected_fields,
        'selected_field_names': [field_names.get(f, f.replace('_', ' ').title()) for f in selected_fields],
        'field_names':          field_names,
        'short_headers':        short_headers,
        'landscape':            landscape,
        'now':                  timezone.now(),
        'print_date':           get_school_today(),
        'printed_by':           request.user.get_full_name() or request.user.username,
        'title':                'Student Scholarships',
    })


@login_required
def export_student_scholarships_excel(request):
    def _awarded_by(o):
        if not o.awarded_by_id:
            return ''
        try:
            from django.contrib.auth import get_user_model
            user = get_user_model().objects.using('default').get(id=o.awarded_by_id)
            return user.get_full_name()
        except Exception:
            return ''
 
    ALL_COLUMNS = [
        ('student_id',         'Admission No.',    lambda o: o.student.admission_number),
        ('student_name',       'Student Name',     lambda o: o.student.get_full_name()),
        ('program',            'Program',          lambda o: o.scholarship_program.name),
        ('scholarship_type',   'Type',             lambda o: o.scholarship_program.get_scholarship_type_display()),
        ('discount_mode',      'Discount Mode',    lambda o: 'Category-Specific' if o.use_category_specific_discounts else 'Global'),
        ('discount_summary',   'Discount Summary', lambda o: str(o.get_discount_display_summary())),
        ('amount_awarded',     'Amount Awarded',   lambda o: float(o.amount_awarded) if o.amount_awarded else ''),
        ('total_used',         'Total Used',       lambda o: float(o.total_amount_used or 0)),
        ('start_date',         'Start Date',       lambda o: o.start_date.strftime('%Y-%m-%d') if o.start_date else ''),
        ('end_date',           'End Date',         lambda o: o.end_date.strftime('%Y-%m-%d') if o.end_date else ''),
        ('status',             'Status',           lambda o: o.get_status_display()),
        ('awarded_by',         'Awarded By',       _awarded_by),
        ('times_renewed',      'Renewal Count',    lambda o: o.times_renewed or 0),
        ('notes',              'Notes',            lambda o: o.notes or ''),
    ]
    DEFAULT_FIELDS = [
        'student_id', 'student_name', 'program', 'scholarship_type',
        'discount_summary', 'amount_awarded', 'total_used', 'start_date', 'status',
    ]
    scholarships = get_filtered_student_scholarships(request)
    columns      = _resolve_columns(ALL_COLUMNS, request.GET.getlist('fields'), DEFAULT_FIELDS)
    return _xlsx_response(
        _make_workbook('Student Scholarships', columns, scholarships),
        'student_scholarships',
    )


# =============================================================================
# DISCOUNTS  (DiscountPolicy + DiscountTier + StudentDiscount + DiscountApplication)
# =============================================================================

def get_filtered_discounts(request):
    """Filter DiscountPolicy objects."""
    policies = DiscountPolicy.objects.prefetch_related(
        'student_awards', 'valid_sessions'
    ).annotate(
        application_count=Count('student_awards__applications', distinct=True),
        active_award_count=Count(
            'student_awards',
            filter=Q(student_awards__status='ACTIVE'),
            distinct=True,
        ),
        status_order=Case(
            When(is_active=True,  then=Value(0)),
            When(is_active=False, then=Value(1)),
            default=Value(1),
            output_field=IntegerField(),
        )
    ).order_by('status_order', 'priority', 'name')

    query              = request.GET.get('q', '').strip()
    category           = request.GET.get('category', '')
    value_mode         = request.GET.get('value_mode', '')
    combination_mode   = request.GET.get('combination_mode', '')
    application_method = request.GET.get('application_method', '')
    is_active          = request.GET.get('is_active', '')
    auto_apply         = request.GET.get('auto_apply', '')
    academic_session   = request.GET.get('academic_session', '')

    if query:
        words = query.split()
        q = Q()
        for w in words:
            q &= Q(name__icontains=w) | Q(code__icontains=w) | Q(description__icontains=w)
        policies = policies.filter(q)

    if category:
        policies = policies.filter(category=category)
    if value_mode:
        policies = policies.filter(value_mode=value_mode)
    if combination_mode:
        policies = policies.filter(combination_mode=combination_mode)
    if application_method:
        policies = policies.filter(application_method=application_method)
    if is_active:
        policies = policies.filter(is_active=(is_active.lower() == 'true'))
    if auto_apply:
        policies = policies.filter(auto_apply=(auto_apply.lower() == 'true'))
    if academic_session:
        policies = policies.filter(valid_sessions__id=academic_session)

    return policies


def get_filtered_student_discounts(request):
    """Filter StudentDiscount (award) objects."""
    
    discounts = StudentDiscount.objects.select_related(
        'student', 'policy'
    ).annotate(
        status_order=Case(
            When(status='ACTIVE',    then=Value(0)),
            When(status='PENDING',   then=Value(1)),
            When(status='SUSPENDED', then=Value(2)),
            When(status='EXPIRED',   then=Value(3)),
            When(status='REVOKED',   then=Value(4)),
            default=Value(5),
            output_field=IntegerField(),
        )
    ).order_by(
        'status_order',
        '-updated_at',   
        '-start_date'
    )

    query          = request.GET.get('q', '').strip()
    policy         = request.GET.get('policy', '')
    status         = request.GET.get('status', '')
    active_on_date = request.GET.get('active_on_date', '')

    if query:
        words = query.split()
        q = Q()
        for w in words:
            q &= (
                Q(student__first_name__icontains=w) |
                Q(student__last_name__icontains=w)  |
                Q(student__admission_number__icontains=w) |
                Q(policy__name__icontains=w) |
                Q(policy__code__icontains=w)
            )
        discounts = discounts.filter(q)

    if policy:
        discounts = discounts.filter(policy_id=policy)
    if status:
        discounts = discounts.filter(status=status)
    if active_on_date:
        try:
            d = date.fromisoformat(active_on_date)
            discounts = discounts.filter(
                start_date__lte=d, status='ACTIVE'
            ).filter(
                Q(end_date__isnull=True) | Q(end_date__gte=d)
            )
        except ValueError:
            pass

    return discounts


def get_filtered_refunds(request):
    """Filter refunded Payments — no separate Refund model exists."""
    refunds = Payment.objects.filter(refunded=True).select_related(
        'student', 'invoice', 'payment_method', 'academic_session', 'fiscal_period'
    ).order_by('-refunded_on')

    query            = request.GET.get('q', '').strip()
    refund_method    = request.GET.get('refund_method', '')
    academic_session = request.GET.get('academic_session', '')
    fiscal_period    = request.GET.get('fiscal_period', '')
    student          = request.GET.get('student', '')
    date_from        = request.GET.get('date_from', '')
    date_to          = request.GET.get('date_to', '')
    min_amount       = request.GET.get('min_amount', '')
    max_amount       = request.GET.get('max_amount', '')

    if query:
        words = query.split()
        q = Q()
        for w in words:
            q &= (
                Q(payment_number__icontains=w)            |
                Q(refund_reference__icontains=w)          |
                Q(student__first_name__icontains=w)       |
                Q(student__last_name__icontains=w)        |
                Q(student__admission_number__icontains=w) |
                Q(refund_notes__icontains=w)
            )
        refunds = refunds.filter(q)

    if refund_method:
        refunds = refunds.filter(refund_method=refund_method)
    if academic_session:
        refunds = refunds.filter(academic_session_id=academic_session)
    if fiscal_period:
        refunds = refunds.filter(fiscal_period_id=fiscal_period)
    if student:
        refunds = refunds.filter(student_id=student)
    if date_from:
        refunds = refunds.filter(refunded_on__date__gte=date_from)
    if date_to:
        refunds = refunds.filter(refunded_on__date__lte=date_to)
    if min_amount:
        try:
            refunds = refunds.filter(amount__gte=Decimal(min_amount))
        except (ValueError, TypeError):
            pass
    if max_amount:
        try:
            refunds = refunds.filter(amount__lte=Decimal(max_amount))
        except (ValueError, TypeError):
            pass

    return refunds


# =============================================================================
# DISCOUNT POLICY VIEWS
# =============================================================================

@login_required
def discount_list(request):
    filter_form = DiscountPolicyFilterForm(request.GET or None)
    policies    = get_filtered_discounts(request)

    try:
        raw_stats = fees_stats.get_discount_statistics()
    except Exception as e:
        logger.error(f"Error getting discount statistics: {e}")
        raw_stats = {}

    budget    = raw_stats.get('budget', {})
    app_stats = raw_stats.get('applications', {})
    stats = {
        'total':                  policies.count(),
        'active':                 policies.filter(is_active=True).count(),
        'inactive':               policies.filter(is_active=False).count(),
        'auto_apply':             policies.filter(auto_apply=True).count(),
        'total_budget_allocated': budget.get('total_budget_allocated', 0),
        'total_budget_used':      budget.get('total_budget_used', 0),
        'total_discount_value':   app_stats.get('total_discount_value', 0),
        'students_benefited':     app_stats.get('students_benefited', 0),
    }

    paginator     = Paginator(policies, 20)
    policies_page = paginator.get_page(request.GET.get('page', 1))
    is_htmx       = request.headers.get('HX-Request') == 'true'

    context = {
        'policies_page': policies_page,
        'paginator':     paginator,
        'stats':         stats,
        'filter_form':   filter_form,
        'is_htmx':       is_htmx,
    }
    if is_htmx:
        return render(request, 'fees/discounts/partials/_discount_results.html', context)
    return render(request, 'fees/discounts/list.html', context)


@login_required
def discount_create(request):
    if request.method == 'POST':
        form          = DiscountPolicyForm(request.POST)
        tiers_formset = DiscountTierFormSet(request.POST, prefix='tiers')

        value_mode_is_tiered = request.POST.get('value_mode') == 'TIERED'
        tiers_valid          = tiers_formset.is_valid() if value_mode_is_tiered else True

        if form.is_valid() and tiers_valid:
            try:
                with transaction.atomic():
                    policy = form.save()
                    if value_mode_is_tiered:
                        tiers_formset.instance = policy
                        tiers_formset.save()

                is_htmx = request.headers.get('HX-Request') == 'true'
                if is_htmx:
                    r = HttpResponse()
                    r['HX-Alert-Message'] = f"Discount policy '{policy.name}' created successfully!"
                    r['HX-Alert-Type']    = 'success'
                    r['HX-Alert-Title']   = 'Created!'
                    r['HX-Redirect']      = reverse('fees:discount_detail', kwargs={'pk': policy.pk})
                    return r
                messages.success(request, f"Discount policy '{policy.name}' created successfully!", extra_tags='sweetalert')
                return redirect('fees:discount_detail', pk=policy.pk)
            except Exception as e:
                logger.error(f"Error creating discount policy: {e}", exc_info=True)
                is_htmx = request.headers.get('HX-Request') == 'true'
                if is_htmx:
                    r = HttpResponse()
                    r['HX-Alert-Message'] = f'Error creating policy: {str(e)}'
                    r['HX-Alert-Type']    = 'error'
                    r['HX-Alert-Title']   = 'Error!'
                    return r
                messages.error(request, f'Error creating policy: {str(e)}')
    else:
        form          = DiscountPolicyForm()
        tiers_formset = DiscountTierFormSet(prefix='tiers')

    matrix_category_types = (
        FeesCategory.objects
        .filter(is_active=True)
        .exclude(category_type='')
        .values_list('category_type', 'name')
        .distinct()
        .order_by('category_type')
    )

    return render(request, 'fees/discounts/form.html', {
        'form':                  form,
        'tiers_formset':         tiers_formset,
        'title':                 'Create Discount Policy',
        'submit_text':           'Create Policy',
        'submit_icon':           'fa-plus',
        'matrix_category_types': matrix_category_types,
    })


@login_required
def discount_detail(request, pk):
    policy = get_object_or_404(
        DiscountPolicy.objects.prefetch_related(
            'tiers', 'valid_sessions', 'applicable_categories'
        ),
        pk=pk,
    )

    # Direct queries — avoids FK traversal hitting wrong DB shard
    active_awards = StudentDiscount.objects.filter(
        policy=policy,
        status='ACTIVE',
    ).select_related('student').order_by('student__last_name')[:20]

    active_awards_total = StudentDiscount.objects.filter(
        policy=policy,
        status='ACTIVE',
    ).count()

    all_awards_total = StudentDiscount.objects.filter(
        policy=policy,
    ).count()

    recent_applications = DiscountApplication.objects.filter(
        student_discount__policy=policy,
        is_reversed=False,
    ).select_related(
        'student_discount__student', 'invoice'
    ).order_by('-created_at')[:20]

    application_count = DiscountApplication.objects.filter(
        student_discount__policy=policy, is_reversed=False
    ).count()

    total_discounted = DiscountApplication.objects.filter(
        student_discount__policy=policy, is_reversed=False
    ).aggregate(total=Sum('amount_discounted'))['total'] or Decimal('0.00')

    budget_remaining = None
    if policy.total_budget is not None:
        budget_remaining = policy.total_budget - (policy.budget_used or Decimal('0.00'))

    return render(request, 'fees/discounts/detail.html', {
        'policy':              policy,
        'recent_applications': recent_applications,
        'active_awards':       active_awards,
        'active_awards_total': active_awards_total,
        'all_awards_total':    all_awards_total,
        'application_count':   application_count,
        'total_discounted':    total_discounted,
        'budget_remaining':    budget_remaining,
        'today':               get_school_today(),
    })

@login_required
def discount_edit(request, pk):
    policy = get_object_or_404(
        DiscountPolicy.objects.annotate(
            active_award_count=Count(
                'student_awards',
                filter=Q(student_awards__status='ACTIVE'),
                distinct=True,
            ),
            application_count=Count(
                'student_awards__applications',
                distinct=True,
            ),
        ),
        pk=pk,
    )

    if request.method == 'POST':
        form          = DiscountPolicyForm(request.POST, instance=policy)
        tiers_formset = DiscountTierFormSet(request.POST, instance=policy, prefix='tiers')

        value_mode_is_tiered = request.POST.get('value_mode') == 'TIERED'
        tiers_valid          = tiers_formset.is_valid() if value_mode_is_tiered else True

        if form.is_valid() and tiers_valid:
            try:
                with transaction.atomic():
                    policy = form.save()
                    if value_mode_is_tiered:
                        tiers_formset.instance = policy
                        tiers_formset.save()
                    else:
                        # Value mode changed away from TIERED — remove orphaned tiers.
                        policy.tiers.all().delete()

                is_htmx = request.headers.get('HX-Request') == 'true'
                if is_htmx:
                    r = HttpResponse()
                    r['HX-Alert-Message'] = f"Discount policy '{policy.name}' updated successfully!"
                    r['HX-Alert-Type']    = 'success'
                    r['HX-Alert-Title']   = 'Updated!'
                    r['HX-Redirect']      = reverse('fees:discount_detail', kwargs={'pk': policy.pk})
                    return r
                messages.success(request, f"Discount policy '{policy.name}' updated successfully!", extra_tags='sweetalert')
                return redirect('fees:discount_detail', pk=policy.pk)
            except Exception as e:
                logger.error(f"Error updating discount policy: {e}", exc_info=True)
                is_htmx = request.headers.get('HX-Request') == 'true'
                if is_htmx:
                    r = HttpResponse()
                    r['HX-Alert-Message'] = f'Error updating policy: {str(e)}'
                    r['HX-Alert-Type']    = 'error'
                    r['HX-Alert-Title']   = 'Error!'
                    return r
                messages.error(request, f'Error updating policy: {str(e)}')
    else:
        form          = DiscountPolicyForm(instance=policy)
        tiers_formset = DiscountTierFormSet(instance=policy, prefix='tiers')

    matrix_category_types = (
        FeesCategory.objects
        .filter(is_active=True)
        .exclude(category_type='')
        .values_list('category_type', 'name')
        .distinct()
        .order_by('category_type')
    )

    return render(request, 'fees/discounts/form.html', {
        'form':                  form,
        'tiers_formset':         tiers_formset,
        'policy':                policy,
        'title':                 f'Edit Discount Policy — {policy.name}',
        'submit_text':           'Update Policy',
        'submit_icon':           'fa-save',
        'matrix_category_types': matrix_category_types,
    })


@login_required
@require_http_methods(["POST"])
def discount_delete(request, pk):
    policy            = get_object_or_404(DiscountPolicy, pk=pk)
    application_count = DiscountApplication.objects.filter(
        student_discount__policy=policy
    ).count()

    if application_count > 0:
        is_htmx = request.headers.get('HX-Request') == 'true'
        if is_htmx:
            r = HttpResponse()
            r['HX-Alert-Message'] = f"Cannot delete '{policy.name}' — it has {application_count} discount application(s)"
            r['HX-Alert-Type']    = 'error'
            r['HX-Alert-Title']   = 'Cannot Delete'
            r['HX-Close-Modal']   = 'true'
            return r
        messages.error(request, f"Cannot delete '{policy.name}' — it has {application_count} discount application(s)", extra_tags='sweetalert-error')
        return redirect('fees:discount_detail', pk=pk)

    name = policy.name
    policy.delete()
    is_htmx = request.headers.get('HX-Request') == 'true'
    if is_htmx:
        r = HttpResponse()
        r['HX-Alert-Message'] = f"Discount policy '{name}' deleted successfully"
        r['HX-Alert-Type']    = 'success'
        r['HX-Alert-Title']   = 'Deleted!'
        r['HX-Close-Modal']   = 'true'
        r['HX-Redirect']      = reverse('fees:discount_list')
        return r
    messages.success(request, f"Discount policy '{name}' deleted successfully", extra_tags='sweetalert')
    return redirect('fees:discount_list')


@login_required
@require_http_methods(["POST"])
def discount_toggle_active(request, pk):
    policy = get_object_or_404(DiscountPolicy, pk=pk)
    try:
        policy.is_active = not policy.is_active
        policy.save()
        status  = "activated" if policy.is_active else "deactivated"
        is_htmx = request.headers.get('HX-Request') == 'true'
        if is_htmx:
            r = HttpResponse()
            r['HX-Alert-Message'] = f"Discount policy '{policy.name}' {status}!"
            r['HX-Alert-Type']    = 'success'
            r['HX-Alert-Title']   = 'Status Updated'
            r['HX-Close-Modal']   = 'true'
            r['HX-Redirect']      = reverse('fees:discount_detail', kwargs={'pk': policy.pk})
            return r
        messages.success(request, f"Discount policy '{policy.name}' {status}!")
        return redirect('fees:discount_detail', pk=policy.pk)
    except Exception as e:
        logger.error(f"Error toggling discount policy status: {e}")
        is_htmx = request.headers.get('HX-Request') == 'true'
        if is_htmx:
            r = HttpResponse()
            r['HX-Alert-Message'] = f'Error updating status: {str(e)}'
            r['HX-Alert-Type']    = 'error'
            r['HX-Alert-Title']   = 'Error'
            return r
        messages.error(request, f'Error updating status: {str(e)}')
        return redirect('fees:discount_detail', pk=pk)


@login_required
def discount_list_print_view(request):
    FIELD_NAMES_FULL = {
        'code':               'Code',
        'name':               'Policy Name',
        'category':           'Category',
        'value_mode':         'Value Mode',
        'flat_percentage':    'Percentage',
        'flat_fixed_amount':  'Fixed Amount',
        'combination_mode':   'Combination Mode',
        'application_method': 'Application Method',
        'total_budget':       'Total Budget',
        'budget_used':        'Budget Used',
        'max_beneficiaries':  'Max Beneficiaries',
        'priority':           'Priority',
        'auto_apply':         'Auto Apply',
        'application_count':  'Applications',
        'active_award_count': 'Active Awards',
        'is_active':          'Active',
    }
    FIELD_NAMES_SHORT = {
        'code':               'Code',
        'name':               'Policy',
        'category':           'Category',
        'value_mode':         'Mode',
        'flat_percentage':    '%',
        'flat_fixed_amount':  'Fixed',
        'combination_mode':   'Combine',
        'application_method': 'Method',
        'total_budget':       'Budget',
        'budget_used':        'Used',
        'max_beneficiaries':  'Max',
        'priority':           'Pri.',
        'auto_apply':         'Auto',
        'application_count':  'Apps',
        'active_award_count': 'Awards',
        'is_active':          'Active',
    }
    DEFAULT_FIELDS  = ['code', 'name', 'category', 'value_mode', 'flat_percentage', 'flat_fixed_amount', 'combination_mode', 'total_budget', 'budget_used', 'application_count', 'is_active']
    selected_fields = request.GET.getlist('fields') or DEFAULT_FIELDS
    short_headers   = request.GET.get('short_headers', 'false').lower() == 'true'
    landscape       = request.GET.get('landscape', 'true').lower() == 'true'
    include_stats   = request.GET.get('include_stats', 'true').lower() == 'true'
    field_names     = FIELD_NAMES_SHORT if short_headers else FIELD_NAMES_FULL
    policies        = get_filtered_discounts(request)

    stats = None
    if include_stats:
        stats = {
            'total':    policies.count(),
            'active':   policies.filter(is_active=True).count(),
            'inactive': policies.filter(is_active=False).count(),
        }

    return render(request, 'fees/discounts/print_discount_list.html', {
        **get_print_school_context(request),
        'policies':             policies,
        'stats':                stats,
        'selected_fields':      selected_fields,
        'selected_field_names': [field_names.get(f, f.replace('_', ' ').title()) for f in selected_fields],
        'field_names':          field_names,
        'short_headers':        short_headers,
        'landscape':            landscape,
        'now':                  timezone.now(),
        'print_date':           get_school_today(),
        'printed_by':           request.user.get_full_name() or request.user.username,
        'title':                'Discount Policies',
    })


@login_required
def export_discounts_excel(request):
    ALL_COLUMNS = [
        ('code',               'Code',               lambda o: o.code),
        ('name',               'Name',               lambda o: o.name),
        ('category',           'Category',           lambda o: o.get_category_display()),
        ('value_mode',         'Value Mode',         lambda o: o.get_value_mode_display()),
        ('flat_percentage',    'Percentage',         lambda o: float(o.flat_percentage) if o.flat_percentage else ''),
        ('flat_fixed_amount',  'Fixed Amount',       lambda o: float(o.flat_fixed_amount) if o.flat_fixed_amount else ''),
        ('combination_mode',   'Combination Mode',   lambda o: o.get_combination_mode_display()),
        ('application_method', 'Application Method', lambda o: o.get_application_method_display()),
        ('total_budget',       'Total Budget',       lambda o: float(o.total_budget) if o.total_budget else 'Unlimited'),
        ('budget_used',        'Budget Used',        lambda o: float(o.budget_used or 0)),
        ('max_beneficiaries',  'Max Beneficiaries',  lambda o: o.max_beneficiaries or 'Unlimited'),
        ('priority',           'Priority',           lambda o: o.priority),
        ('auto_apply',         'Auto Apply',         lambda o: 'Yes' if o.auto_apply else 'No'),
        ('application_count',  'Applications',       lambda o: o.application_count or 0),
        ('active_award_count', 'Active Awards',      lambda o: o.active_award_count or 0),
        ('is_active',          'Active',             lambda o: 'Yes' if o.is_active else 'No'),
        ('description',        'Description',        lambda o: o.description or ''),
    ]
    DEFAULT_FIELDS = ['code', 'name', 'category', 'value_mode', 'flat_percentage', 'flat_fixed_amount', 'combination_mode', 'total_budget', 'budget_used', 'application_count', 'is_active']
    policies = get_filtered_discounts(request)
    columns  = _resolve_columns(ALL_COLUMNS, request.GET.getlist('fields'), DEFAULT_FIELDS)
    return _xlsx_response(_make_workbook('Discount Policies', columns, policies), 'discount_policies')


# =============================================================================
# STUDENT DISCOUNT AWARD VIEWS
# =============================================================================

@login_required
def student_discount_list(request):
    filter_form = StudentDiscountFilterForm(request.GET or None)
    discounts   = get_filtered_student_discounts(request)

    stats = {
        'total':           discounts.count(),
        'active':          discounts.filter(status='ACTIVE').count(),
        'suspended':       discounts.filter(status='SUSPENDED').count(),
        'expired':         discounts.filter(status='EXPIRED').count(),
        'revoked':         discounts.filter(status='REVOKED').count(),
        'pending':         discounts.filter(status='PENDING').count(),
        'total_discounted': discounts.aggregate(
            v=Sum('applications__amount_discounted')
        )['v'] or Decimal('0.00'),
    }

    paginator      = Paginator(discounts, 20)
    discounts_page = paginator.get_page(request.GET.get('page', 1))
    is_htmx        = request.headers.get('HX-Request') == 'true'

    context = {
        'discounts_page': discounts_page,
        'paginator':      paginator,
        'stats':          stats,
        'filter_form':    filter_form,
        'is_htmx':        is_htmx,
    }
    if is_htmx:
        return render(request, 'fees/discounts/partials/_student_discount_results.html', context)
    return render(request, 'fees/discounts/student_discount_list.html', context)

@login_required
def student_discount_create(request):
    student_id = request.GET.get('student')
    student    = get_object_or_404(Student, pk=student_id) if student_id else None

    if request.method == 'POST':
        form = StudentDiscountForm(request.POST, student=student)
        if form.is_valid():
            try:
                discount               = form.save(commit=False)
                discount.status = 'ACTIVE'
                discount.awarded_by_id = str(request.user.id)
                if not discount.awarded_date:
                    discount.awarded_date = get_school_today()
                discount.save()
                is_htmx = request.headers.get('HX-Request') == 'true'
                if is_htmx:
                    r = HttpResponse()
                    r['HX-Alert-Message'] = f"Discount awarded to {discount.student.get_full_name()} successfully!"
                    r['HX-Alert-Type']    = 'success'
                    r['HX-Alert-Title']   = 'Discount Awarded!'
                    r['HX-Redirect']      = reverse('fees:student_discount_detail', kwargs={'pk': discount.pk})
                    return r
                messages.success(request, f"Discount awarded to {discount.student.get_full_name()} successfully!", extra_tags='sweetalert')
                return redirect('fees:student_discount_detail', pk=discount.pk)
            except Exception as e:
                logger.error(f"Error creating student discount: {e}", exc_info=True)
                is_htmx = request.headers.get('HX-Request') == 'true'
                if is_htmx:
                    r = HttpResponse()
                    r['HX-Alert-Message'] = f'Error awarding discount: {str(e)}'
                    r['HX-Alert-Type']    = 'error'
                    r['HX-Alert-Title']   = 'Error!'
                    return r
                messages.error(request, f'Error awarding discount: {str(e)}')
    else:
        form = StudentDiscountForm(student=student)

    return render(request, 'fees/discounts/student_discount_form.html', {
        'form':        form,
        'student':     student,
        'title':       'Award Discount',
        'submit_text': 'Award Discount',
        'submit_icon': 'fa-plus',
    })


@login_required
def student_discount_detail(request, pk):
    discount = get_object_or_404(
        StudentDiscount.objects.select_related(
            'student', 'policy'
        ).prefetch_related('policy__tiers'),
        pk=pk,
    )
    applications = DiscountApplication.objects.filter(
        student_discount=discount
    ).select_related(
        'invoice', 'invoice_item__fee_category'
    ).order_by('-created_at')

    total_discounted = applications.filter(
        is_reversed=False
    ).aggregate(v=Sum('amount_discounted'))['v'] or Decimal('0.00')

    budget_remaining = None
    if discount.policy.max_discount_per_student:
        budget_remaining = discount.policy.max_discount_per_student - total_discounted

    active_application_count   = applications.filter(is_reversed=False).count()
    reversed_application_count = applications.filter(is_reversed=True).count()

    return render(request, 'fees/discounts/student_discount_detail.html', {
        'discount':                   discount,
        'applications':               applications,
        'total_discounted':           total_discounted,
        'budget_remaining':           budget_remaining,
        'active_application_count':   active_application_count,
        'reversed_application_count': reversed_application_count,
        'today':                      get_school_today(),
    })


@login_required
def student_discount_edit(request, pk):
    discount = get_object_or_404(
        StudentDiscount.objects.select_related(
            'student', 'policy'
        ).prefetch_related('policy__tiers').annotate(
            application_count=Count('applications', distinct=True),
            total_discounted=Sum('applications__amount_discounted'),
        ),
        pk=pk,
    )

    if request.method == 'POST':
        form = StudentDiscountForm(request.POST, instance=discount, student=discount.student)
        if form.is_valid():
            try:
                discount = form.save()
                is_htmx  = request.headers.get('HX-Request') == 'true'
                if is_htmx:
                    r = HttpResponse()
                    r['HX-Alert-Message'] = f"Discount for {discount.student.get_full_name()} updated successfully!"
                    r['HX-Alert-Type']    = 'success'
                    r['HX-Alert-Title']   = 'Updated!'
                    r['HX-Redirect']      = reverse('fees:student_discount_detail', kwargs={'pk': discount.pk})
                    return r
                messages.success(request, f"Discount for {discount.student.get_full_name()} updated successfully!", extra_tags='sweetalert')
                return redirect('fees:student_discount_detail', pk=discount.pk)
            except Exception as e:
                logger.error(f"Error updating student discount: {e}", exc_info=True)
                is_htmx = request.headers.get('HX-Request') == 'true'
                if is_htmx:
                    r = HttpResponse()
                    r['HX-Alert-Message'] = f'Error updating discount: {str(e)}'
                    r['HX-Alert-Type']    = 'error'
                    r['HX-Alert-Title']   = 'Error!'
                    return r
                messages.error(request, f'Error updating discount: {str(e)}')
    else:
        form = StudentDiscountForm(instance=discount, student=discount.student)

    return render(request, 'fees/discounts/student_discount_form.html', {
        'form':        form,
        'discount':    discount,
        'title':       f'Edit Discount — {discount.student.get_full_name()}',
        'submit_text': 'Update Discount',
        'submit_icon': 'fa-save',
    })


@login_required
@require_http_methods(["POST"])
def student_discount_delete(request, pk):
    discount = get_object_or_404(StudentDiscount, pk=pk)

    if discount.applications.filter(is_reversed=False).exists():
        is_htmx = request.headers.get('HX-Request') == 'true'
        if is_htmx:
            r = HttpResponse()
            r['HX-Alert-Message'] = "Cannot delete — this discount has been applied to invoices"
            r['HX-Alert-Type']    = 'error'
            r['HX-Alert-Title']   = 'Cannot Delete'
            r['HX-Close-Modal']   = 'true'
            return r
        messages.error(request, "Cannot delete — this discount has been applied to invoices", extra_tags='sweetalert-error')
        return redirect('fees:student_discount_detail', pk=pk)

    student_name = discount.student.get_full_name()
    discount.delete()
    is_htmx = request.headers.get('HX-Request') == 'true'
    if is_htmx:
        r = HttpResponse()
        r['HX-Alert-Message'] = f"Discount for '{student_name}' deleted successfully"
        r['HX-Alert-Type']    = 'success'
        r['HX-Alert-Title']   = 'Deleted!'
        r['HX-Close-Modal']   = 'true'
        r['HX-Redirect']      = reverse('fees:student_discount_list')
        return r
    messages.success(request, f"Discount for '{student_name}' deleted successfully", extra_tags='sweetalert')
    return redirect('fees:student_discount_list')


@login_required
@require_http_methods(["POST"])
def student_discount_suspend(request, pk):
    discount = get_object_or_404(StudentDiscount, pk=pk)
    if discount.status != 'ACTIVE':
        is_htmx = request.headers.get('HX-Request') == 'true'
        if is_htmx:
            r = HttpResponse()
            r['HX-Alert-Message'] = 'Only active discounts can be suspended.'
            r['HX-Alert-Type']    = 'warning'
            r['HX-Alert-Title']   = 'Cannot Suspend'
            return r
        messages.error(request, 'Only active discounts can be suspended.')
        return redirect('fees:student_discount_detail', pk=pk)
    try:
        reason                    = request.POST.get('suspension_reason', '')
        discount.status           = 'SUSPENDED'
        discount.suspension_reason = reason
        discount.save()
        is_htmx = request.headers.get('HX-Request') == 'true'
        if is_htmx:
            r = HttpResponse()
            r['HX-Alert-Message'] = f"Discount for {discount.student.get_full_name()} suspended!"
            r['HX-Alert-Type']    = 'warning'
            r['HX-Alert-Title']   = 'Discount Suspended'
            r['HX-Close-Modal']   = 'true'
            r['HX-Redirect']      = reverse('fees:student_discount_detail', kwargs={'pk': discount.pk})
            return r
        messages.success(request, f"Discount for {discount.student.get_full_name()} suspended!")
        return redirect('fees:student_discount_detail', pk=pk)
    except Exception as e:
        logger.error(f"Error suspending student discount: {e}", exc_info=True)
        is_htmx = request.headers.get('HX-Request') == 'true'
        if is_htmx:
            r = HttpResponse()
            r['HX-Alert-Message'] = f'Error suspending discount: {str(e)}'
            r['HX-Alert-Type']    = 'error'
            r['HX-Alert-Title']   = 'Error'
            return r
        messages.error(request, f'Error suspending discount: {str(e)}')
        return redirect('fees:student_discount_detail', pk=pk)


@login_required
@require_http_methods(["POST"])
def student_discount_revoke(request, pk):
    discount = get_object_or_404(StudentDiscount, pk=pk)
    if discount.status == 'REVOKED':
        is_htmx = request.headers.get('HX-Request') == 'true'
        if is_htmx:
            r = HttpResponse()
            r['HX-Alert-Message'] = 'Discount is already revoked.'
            r['HX-Alert-Type']    = 'warning'
            r['HX-Alert-Title']   = 'Already Revoked'
            return r
        messages.error(request, 'Discount is already revoked.')
        return redirect('fees:student_discount_detail', pk=pk)
    try:
        reason                    = request.POST.get('revocation_reason', '')
        discount.status           = 'REVOKED'
        discount.revocation_reason = reason
        discount.end_date         = get_school_today()
        discount.save()
        is_htmx = request.headers.get('HX-Request') == 'true'
        if is_htmx:
            r = HttpResponse()
            r['HX-Alert-Message'] = f"Discount for {discount.student.get_full_name()} revoked!"
            r['HX-Alert-Type']    = 'warning'
            r['HX-Alert-Title']   = 'Discount Revoked'
            r['HX-Close-Modal']   = 'true'
            r['HX-Redirect']      = reverse('fees:student_discount_detail', kwargs={'pk': discount.pk})
            return r
        messages.success(request, f"Discount for {discount.student.get_full_name()} revoked!")
        return redirect('fees:student_discount_detail', pk=pk)
    except Exception as e:
        logger.error(f"Error revoking student discount: {e}", exc_info=True)
        is_htmx = request.headers.get('HX-Request') == 'true'
        if is_htmx:
            r = HttpResponse()
            r['HX-Alert-Message'] = f'Error revoking discount: {str(e)}'
            r['HX-Alert-Type']    = 'error'
            r['HX-Alert-Title']   = 'Error'
            return r
        messages.error(request, f'Error revoking discount: {str(e)}')
        return redirect('fees:student_discount_detail', pk=pk)


@login_required
def student_discount_list_print_view(request):
    FIELD_NAMES_FULL = {
        'student_id':        'Admission Number',
        'student_name':      'Student Name',
        'policy_code':       'Policy Code',
        'policy_name':       'Policy Name',
        'category':          'Category',
        'value_mode':        'Value Mode',
        'override_pct':      'Override %',
        'override_fixed':    'Override Fixed Amount',
        'start_date':        'Start Date',
        'end_date':          'End Date',
        'status':            'Status',
        'application_count': 'Times Applied',
        'total_discounted':  'Total Discounted',
        'notes':             'Notes',
    }
    FIELD_NAMES_SHORT = {
        'student_id':        'Adm. No.',
        'student_name':      'Student',
        'policy_code':       'Code',
        'policy_name':       'Policy',
        'category':          'Category',
        'value_mode':        'Mode',
        'override_pct':      'Override %',
        'override_fixed':    'Override Fixed',
        'start_date':        'Start',
        'end_date':          'End',
        'status':            'Status',
        'application_count': 'Applied',
        'total_discounted':  'Total Disc.',
        'notes':             'Notes',
    }
    DEFAULT_FIELDS  = ['student_id', 'student_name', 'policy_code', 'policy_name', 'category', 'start_date', 'end_date', 'status', 'application_count', 'total_discounted']
    selected_fields = request.GET.getlist('fields') or DEFAULT_FIELDS
    short_headers   = request.GET.get('short_headers', 'false').lower() == 'true'
    landscape       = request.GET.get('landscape', 'true').lower() == 'true'
    include_stats   = request.GET.get('include_stats', 'true').lower() == 'true'
    field_names     = FIELD_NAMES_SHORT if short_headers else FIELD_NAMES_FULL
    discounts       = get_filtered_student_discounts(request)

    stats = None
    if include_stats:
        stats = {
            'total':           discounts.count(),
            'active':          discounts.filter(status='ACTIVE').count(),
            'suspended':       discounts.filter(status='SUSPENDED').count(),
            'revoked':         discounts.filter(status='REVOKED').count(),
            'total_discounted': discounts.aggregate(
                v=Sum('applications__amount_discounted')
            )['v'] or Decimal('0.00'),
        }

    if discounts.count() > MAX_PRINT_RECORDS:
        discounts = discounts[:MAX_PRINT_RECORDS]

    return render(request, 'fees/discounts/print_student_discount_list.html', {
        **get_print_school_context(request),
        'discounts':            discounts,
        'stats':                stats,
        'selected_fields':      selected_fields,
        'selected_field_names': [field_names.get(f, f.replace('_', ' ').title()) for f in selected_fields],
        'field_names':          field_names,
        'short_headers':        short_headers,
        'landscape':            landscape,
        'now':                  timezone.now(),
        'print_date':           get_school_today(),
        'printed_by':           request.user.get_full_name() or request.user.username,
        'title':                'Student Discount Awards',
    })


@login_required
def export_student_discounts_excel(request):
    ALL_COLUMNS = [
        ('student_id',        'Admission No.',    lambda o: o.student.admission_number),
        ('student_name',      'Student Name',     lambda o: o.student.get_full_name()),
        ('policy_code',       'Policy Code',      lambda o: o.policy.code),
        ('policy_name',       'Policy Name',      lambda o: o.policy.name),
        ('category',          'Category',         lambda o: o.policy.get_category_display()),
        ('value_mode',        'Value Mode',       lambda o: o.policy.get_value_mode_display()),
        ('override_pct',      'Override %',       lambda o: float(o.override_percentage) if o.override_percentage is not None else ''),
        ('override_fixed',    'Override Fixed',   lambda o: float(o.override_fixed_amount) if o.override_fixed_amount is not None else ''),
        ('start_date',        'Start Date',       lambda o: o.start_date.strftime('%Y-%m-%d')),
        ('end_date',          'End Date',         lambda o: o.end_date.strftime('%Y-%m-%d') if o.end_date else ''),
        ('status',            'Status',           lambda o: o.get_status_display()),
        ('application_count', 'Times Applied',    lambda o: o.application_count or 0),
        ('total_discounted',  'Total Discounted', lambda o: float(o.total_discounted or 0)),
        ('awarded_by_id',     'Awarded By ID',    lambda o: o.awarded_by_id or ''),
        ('awarded_date',      'Award Date',       lambda o: o.awarded_date.strftime('%Y-%m-%d') if o.awarded_date else ''),
        ('notes',             'Notes',            lambda o: o.notes or ''),
    ]
    DEFAULT_FIELDS = ['student_id', 'student_name', 'policy_code', 'policy_name', 'category', 'start_date', 'end_date', 'status', 'application_count', 'total_discounted']
    discounts = get_filtered_student_discounts(request)
    columns   = _resolve_columns(ALL_COLUMNS, request.GET.getlist('fields'), DEFAULT_FIELDS)
    return _xlsx_response(_make_workbook('Student Discounts', columns, discounts), 'student_discounts')


# =============================================================================
# DISCOUNT APPLICATION VIEWS
# =============================================================================

@login_required
@require_http_methods(["POST"])
def apply_discount_to_invoice(request, pk):
    """
    Manually apply a DiscountPolicy to an invoice.
    Finds or creates a StudentDiscount award, then records a DiscountApplication.
    """
    invoice = get_object_or_404(FeeInvoice, pk=pk)

    if invoice.status in ['PAID', 'CANCELLED', 'VOID']:
        is_htmx = request.headers.get('HX-Request') == 'true'
        if is_htmx:
            r = HttpResponse()
            r['HX-Alert-Message'] = f"Cannot apply discount to a {invoice.get_status_display()} invoice"
            r['HX-Alert-Type']    = 'error'
            r['HX-Alert-Title']   = 'Cannot Apply Discount'
            return r
        messages.error(request, f"Cannot apply discount to a {invoice.get_status_display()} invoice")
        return redirect('fees:invoice_detail', pk=pk)

    try:
        policy_id = request.POST.get('policy_id')
        policy    = get_object_or_404(DiscountPolicy, pk=policy_id, is_active=True)

        with transaction.atomic():
            # Find or create a StudentDiscount award using correct field names
            student_discount, _ = StudentDiscount.objects.get_or_create(
                student=invoice.student,
                policy=policy,
                status='ACTIVE',
                defaults={
                    'awarded_by_id': str(request.user.id),   # CharField
                    'awarded_date':  get_school_today(),       # correct field name
                    'notes':         f"Manual award via invoice {invoice.invoice_number}",
                    'start_date':    get_school_today(),       # required field
                },
            )

            # Calculate discount amount based on value_mode
            if policy.value_mode == 'FLAT_PERCENTAGE' and policy.flat_percentage:
                discount_amount = (
                    invoice.subtotal_amount * policy.flat_percentage / Decimal('100')
                ).quantize(Decimal('0.01'))
            elif policy.value_mode == 'FLAT_FIXED' and policy.flat_fixed_amount:
                discount_amount = policy.flat_fixed_amount
            elif policy.value_mode == 'FLAT_WAIVER':
                discount_amount = invoice.balance
            else:
                discount_amount = Decimal('0.00')

            # Cap at remaining balance and per-student budget if set
            discount_amount = min(discount_amount, invoice.balance)
            if policy.max_discount_per_student:
                already_given = DiscountApplication.objects.filter(
                    student_discount__student=invoice.student,
                    student_discount__policy=policy,
                    is_reversed=False,
                ).aggregate(total=Sum('amount_discounted'))['total'] or Decimal('0.00')
                remaining_cap   = policy.max_discount_per_student - already_given
                discount_amount = min(discount_amount, remaining_cap)

            if discount_amount <= Decimal('0.00'):
                is_htmx = request.headers.get('HX-Request') == 'true'
                if is_htmx:
                    r = HttpResponse()
                    r['HX-Alert-Message'] = "Discount amount is zero — nothing to apply"
                    r['HX-Alert-Type']    = 'warning'
                    r['HX-Alert-Title']   = 'No Discount Applied'
                    return r
                messages.warning(request, "Discount amount is zero — nothing to apply")
                return redirect('fees:invoice_detail', pk=pk)

            # Record the application against the invoice (no item — invoice level)
            DiscountApplication.objects.create(
                student_discount=student_discount,
                invoice=invoice,
                invoice_item=None,
                amount_discounted=discount_amount,
                applied_by_id=str(request.user.id),
                notes=request.POST.get('notes', ''),
            )

            # Update invoice totals
            invoice.discount_amount       += discount_amount
            invoice.has_discounts_applied  = True
            invoice.balance                = invoice.total_amount - invoice.paid_amount - invoice.discount_amount
            invoice.save(update_fields=['discount_amount', 'has_discounts_applied', 'balance'])

            # Update policy budget tracking if applicable
            if policy.total_budget:
                DiscountPolicy.objects.filter(pk=policy.pk).update(
                    budget_used=F('budget_used') + discount_amount
                )

        is_htmx = request.headers.get('HX-Request') == 'true'
        if is_htmx:
            r = HttpResponse()
            r['HX-Alert-Message'] = f"Discount of {discount_amount:,.2f} applied to invoice {invoice.invoice_number}!"
            r['HX-Alert-Type']    = 'success'
            r['HX-Alert-Title']   = 'Discount Applied'
            r['HX-Close-Modal']   = 'true'
            r['HX-Redirect']      = reverse('fees:invoice_detail', kwargs={'pk': invoice.pk})
            return r
        messages.success(request, f"Discount of {discount_amount:,.2f} applied to invoice {invoice.invoice_number}!")
        return redirect('fees:invoice_detail', pk=invoice.pk)

    except Exception as e:
        logger.error(f"Error applying discount to invoice: {e}", exc_info=True)
        is_htmx = request.headers.get('HX-Request') == 'true'
        if is_htmx:
            r = HttpResponse()
            r['HX-Alert-Message'] = f'Error applying discount: {str(e)}'
            r['HX-Alert-Type']    = 'error'
            r['HX-Alert-Title']   = 'Error'
            return r
        messages.error(request, f'Error applying discount: {str(e)}')
        return redirect('fees:invoice_detail', pk=pk)


@login_required
@require_http_methods(["POST"])
def discount_application_reverse(request, pk):
    """
    Reverse a DiscountApplication — marks it reversed and restores
    the discount amount on the parent invoice.
    """
    application = get_object_or_404(DiscountApplication, pk=pk)

    if application.is_reversed:
        is_htmx = request.headers.get('HX-Request') == 'true'
        if is_htmx:
            r = HttpResponse()
            r['HX-Alert-Message'] = 'This discount application is already reversed.'
            r['HX-Alert-Type']    = 'warning'
            r['HX-Alert-Title']   = 'Already Reversed'
            return r
        messages.error(request, 'This discount application is already reversed.')
        return redirect('fees:invoice_detail', pk=application.invoice.pk)

    reason = request.POST.get('reversal_reason', '').strip()
    if not reason:
        is_htmx = request.headers.get('HX-Request') == 'true'
        if is_htmx:
            r = HttpResponse()
            r['HX-Alert-Message'] = 'A reversal reason is required.'
            r['HX-Alert-Type']    = 'error'
            r['HX-Alert-Title']   = 'Validation Error'
            return r
        messages.error(request, 'A reversal reason is required.')
        return redirect('fees:invoice_detail', pk=application.invoice.pk)

    try:
        with transaction.atomic():
            application.is_reversed    = True
            application.reversed_date  = get_school_today()
            application.reversed_by_id = str(request.user.id)
            application.reversal_reason = reason
            application.save()

            invoice = application.invoice
            invoice.discount_amount = max(
                Decimal('0.00'),
                invoice.discount_amount - application.amount_discounted,
            )
            invoice.has_discounts_applied = invoice.discount_amount > Decimal('0.00')
            invoice.balance = invoice.total_amount - invoice.paid_amount - invoice.discount_amount
            invoice.save(update_fields=['discount_amount', 'has_discounts_applied', 'balance'])

            # Restore policy budget if tracked
            if application.student_discount.policy.total_budget:
                DiscountPolicy.objects.filter(
                    pk=application.student_discount.policy_id
                ).update(
                    budget_used=F('budget_used') - application.amount_discounted
                )

        is_htmx = request.headers.get('HX-Request') == 'true'
        if is_htmx:
            r = HttpResponse()
            r['HX-Alert-Message'] = f"Discount application of {application.amount_discounted:,.2f} reversed successfully!"
            r['HX-Alert-Type']    = 'success'
            r['HX-Alert-Title']   = 'Application Reversed'
            r['HX-Close-Modal']   = 'true'
            r['HX-Redirect']      = reverse('fees:invoice_detail', kwargs={'pk': invoice.pk})
            return r
        messages.success(request, "Discount application reversed successfully!")
        return redirect('fees:invoice_detail', pk=application.invoice.pk)

    except Exception as e:
        logger.error(f"Error reversing discount application: {e}", exc_info=True)
        is_htmx = request.headers.get('HX-Request') == 'true'
        if is_htmx:
            r = HttpResponse()
            r['HX-Alert-Message'] = f'Error reversing application: {str(e)}'
            r['HX-Alert-Type']    = 'error'
            r['HX-Alert-Title']   = 'Error'
            return r
        messages.error(request, f'Error reversing application: {str(e)}')
        return redirect('fees:invoice_detail', pk=application.invoice.pk)


# =============================================================================
# REFUNDS LIST / PRINT / EXPORT
# =============================================================================

@login_required
def refund_list(request):
    refunds = get_filtered_refunds(request)
    today   = get_school_today()

    stats = {
        'total':        refunds.count(),
        'total_amount': refunds.aggregate(Sum('amount'))['amount__sum'] or Decimal('0.00'),
        'this_month':   refunds.filter(
            refunded_on__month=today.month,
            refunded_on__year=today.year,
        ).count(),
        'by_method': refunds.values('refund_method').annotate(
            count=Count('id'), total=Sum('amount')
        ).order_by('-total')[:5],
    }

    paginator    = Paginator(refunds, 20)
    refunds_page = paginator.get_page(request.GET.get('page', 1))
    is_htmx      = request.headers.get('HX-Request') == 'true'

    context = {
        'refunds_page':       refunds_page,
        'paginator':          paginator,
        'stats':              stats,
        'is_htmx':            is_htmx,
        'all_sessions':       AcademicSession.objects.filter(is_active=True).order_by('-start_date'),
        'all_fiscal_periods': FiscalPeriod.objects.all().order_by('-start_date'),
    }
    if is_htmx:
        return render(request, 'fees/discounts/partials/_refund_results.html', context)
    return render(request, 'fees/discounts/refund_list.html', context)


@login_required
def refund_list_print_view(request):
    FIELD_NAMES_FULL = {
        'payment_number':   'Payment Number',
        'student_id':       'Admission Number',
        'student_name':     'Student Name',
        'invoice_number':   'Invoice Number',
        'amount':           'Amount',
        'refund_method':    'Refund Method',
        'refund_reference': 'Reference',
        'refunded_on':      'Refund Date',
        'refund_notes':     'Notes',
        'academic_session': 'Academic Session',
        'fiscal_period':    'Fiscal Period',
    }
    FIELD_NAMES_SHORT = {
        'payment_number':   'Pmt. No.',
        'student_id':       'Adm. No.',
        'student_name':     'Student',
        'invoice_number':   'Inv. No.',
        'amount':           'Amount',
        'refund_method':    'Method',
        'refund_reference': 'Ref.',
        'refunded_on':      'Date',
        'refund_notes':     'Notes',
        'academic_session': 'Session',
        'fiscal_period':    'Period',
    }
    DEFAULT_FIELDS  = ['payment_number', 'student_id', 'student_name', 'invoice_number', 'amount', 'refund_method', 'refund_reference', 'refunded_on', 'refund_notes']
    selected_fields = request.GET.getlist('fields') or DEFAULT_FIELDS
    short_headers   = request.GET.get('short_headers', 'false').lower() == 'true'
    landscape       = request.GET.get('landscape', 'true').lower() == 'true'
    include_stats   = request.GET.get('include_stats', 'true').lower() == 'true'
    field_names     = FIELD_NAMES_SHORT if short_headers else FIELD_NAMES_FULL
    refunds         = get_filtered_refunds(request)

    stats = None
    if include_stats:
        stats = {
            'total':        refunds.count(),
            'total_amount': refunds.aggregate(Sum('amount'))['amount__sum'] or Decimal('0.00'),
        }

    if refunds.count() > MAX_PRINT_RECORDS:
        refunds = refunds[:MAX_PRINT_RECORDS]

    return render(request, 'fees/discounts/print_refund_list.html', {
        **get_print_school_context(request),
        'refunds':              refunds,
        'stats':                stats,
        'selected_fields':      selected_fields,
        'selected_field_names': [field_names.get(f, f.replace('_', ' ').title()) for f in selected_fields],
        'field_names':          field_names,
        'short_headers':        short_headers,
        'landscape':            landscape,
        'now':                  timezone.now(),
        'print_date':           get_school_today(),
        'printed_by':           request.user.get_full_name() or request.user.username,
        'title':                'Refunds',
    })


@login_required
def export_refunds_excel(request):
    ALL_COLUMNS = [
        ('payment_number',    'Payment Number',         lambda o: o.payment_number),
        ('student_id',        'Admission No.',          lambda o: o.student.admission_number),
        ('student_name',      'Student Name',           lambda o: o.student.get_full_name()),
        ('invoice_number',    'Invoice Number',         lambda o: o.invoice.invoice_number if o.invoice else ''),
        ('amount',            'Amount',                 lambda o: float(o.amount)),
        ('refund_method',     'Refund Method',          lambda o: o.refund_method or ''),
        ('refund_reference',  'Refund Reference',       lambda o: o.refund_reference or ''),
        ('refunded_on',       'Refund Date',            lambda o: o.refunded_on.strftime('%Y-%m-%d') if o.refunded_on else ''),
        ('refund_notes',      'Refund Notes',           lambda o: o.refund_notes or ''),
        ('orig_payment_date', 'Original Payment Date',  lambda o: o.payment_date.strftime('%Y-%m-%d') if o.payment_date else ''),
        ('payment_method',    'Original Method',        lambda o: o.payment_method.name if o.payment_method else ''),
        ('academic_session',  'Academic Session',       lambda o: o.academic_session.name if o.academic_session else ''),
        ('fiscal_period',     'Fiscal Period',          lambda o: o.fiscal_period.name if o.fiscal_period else ''),
    ]
    DEFAULT_FIELDS = ['payment_number', 'student_id', 'student_name', 'invoice_number', 'amount', 'refund_method', 'refund_reference', 'refunded_on', 'refund_notes']
    refunds = get_filtered_refunds(request)
    columns = _resolve_columns(ALL_COLUMNS, request.GET.getlist('fields'), DEFAULT_FIELDS)
    return _xlsx_response(_make_workbook('Refunds', columns, refunds), 'refunds')


# =============================================================================
# REPORTS
# =============================================================================

@login_required
def reports_index(request):
    """Reports landing page."""
    return render(request, 'fees/reports/index.html', {
        'current_session': get_active_academic_session(),
    })


@login_required
def financial_summary_report(request):
    from django.db.models import Sum, Count, Q
    from django.db.models.functions import TruncDate, TruncMonth
    from finance.models import Expense, Budget, BudgetLine
    from uniforms.models import UniformSale, UniformSaleItem
    from fees.models import FeeInvoice, FeeInvoiceItem, Payment
    from academics.models import AcademicSession, StudentClassEnrollment
    from hr.models import Payroll
    import json

    # ── Session selection ──────────────────────────────────────────
    academic_sessions = AcademicSession.objects.all().order_by('-start_date')
    session_id = request.GET.get('academic_session')

    if session_id:
        try:
            this_session = AcademicSession.objects.get(pk=session_id)
        except AcademicSession.DoesNotExist:
            this_session = AcademicSession.get_current_session()
    else:
        this_session = AcademicSession.get_current_session()

    # ── Querysets scoped to session ────────────────────────────────
    if this_session:
        invoices     = FeeInvoice.objects.filter(academic_session=this_session)
        payments_qs  = Payment.objects.filter(
            status='COMPLETED', reversed=False, refunded=False,
            academic_session=this_session,
        )
        expenses_qs  = Expense.objects.filter(
            academic_session=this_session,
            status__in=['APPROVED', 'PAID'],
        )
        payroll_qs   = Payroll.objects.filter(
            pay_period_start__gte=this_session.start_date,
            pay_period_end__lte=this_session.end_date,
            reversed=False,
            status__in=['APPROVED', 'PARTIAL', 'PAID'],
        )
        budget_qs    = Budget.objects.filter(
            academic_session=this_session,
            status__in=['APPROVED', 'ACTIVE'],
        )
    else:
        invoices     = FeeInvoice.objects.none()
        payments_qs  = Payment.objects.none()
        expenses_qs  = Expense.objects.none()
        payroll_qs   = Payroll.objects.none()
        budget_qs    = Budget.objects.none()

    # ── Fee collection totals ──────────────────────────────────────
    inv_totals     = invoices.aggregate(billed=Sum('total_amount'), collected=Sum('paid_amount'))
    fees_billed    = inv_totals['billed']    or 0
    fees_collected = inv_totals['collected'] or 0
    total_outstanding  = fees_billed - fees_collected
    _fees_billed_f     = float(fees_billed)
    _fees_collected_f  = float(fees_collected)
    overall_rate       = round(_fees_collected_f / _fees_billed_f * 100, 1) if _fees_billed_f else 0
    fees_coverage_percentage = overall_rate

    total_pupils_with_invoices = invoices.values('student').distinct().count()
    fully_paid_pupils = invoices.filter(status='PAID').values('student').distinct().count()
    pupil_payment_completion_rate = round(
        fully_paid_pupils / total_pupils_with_invoices * 100, 1
    ) if total_pupils_with_invoices else 0

    # ── Revenue ────────────────────────────────────────────────────
    total_paid   = payments_qs.aggregate(t=Sum('amount_in_school_currency'))['t'] or 0
    _total_paid_f = float(total_paid)

    # ── Expenses ───────────────────────────────────────────────────
    total_non_salary_expenses = expenses_qs.aggregate(t=Sum('total_amount'))['t'] or 0
    total_employee_salaries   = float(
        payroll_qs.aggregate(t=Sum('employer_cost_in_school_currency'))['t'] or 0
    )
    from decimal import Decimal
    total_expenditure = float(total_non_salary_expenses) + total_employee_salaries
    total_profit      = _total_paid_f - total_expenditure

    revenue_change = 0
    expense_change = 0
    profit_change  = 0

    # ── Budget ────────────────────────────────────────────────────
    total_budget         = float(budget_qs.aggregate(t=Sum('total_expense_budget'))['t'] or 0)
    total_revenue_budget = float(budget_qs.aggregate(t=Sum('total_revenue_budget'))['t'] or 0)
    total_utilization    = round(float(total_expenditure) / float(total_budget) * 100, 1) if total_budget else 0

    # ── Budget helpers ────────────────────────────────────────────
    expense_lines = BudgetLine.objects.filter(budget__in=budget_qs, line_type='EXPENSE')

    def _budget_by_expense_type(expense_types):
        return float(expense_lines.filter(
            account__expense_type__in=expense_types
        ).aggregate(t=Sum('budgeted_amount'))['t'] or 0)

    def _exp_actual(cat_type):
        return float(expenses_qs.filter(
            category__category_type=cat_type
        ).aggregate(t=Sum('total_amount'))['t'] or 0)

    def _variance(budget_val, actual_val):
        if not budget_val:
            return '—'
        return round((float(actual_val) - float(budget_val)) / float(budget_val) * 100, 1)

    # Employee salaries — actual from Payroll (already computed above)
    eb = _budget_by_expense_type(['TEACHING_SALARIES', 'ADMIN_SALARIES', 'STAFF_BENEFITS'])
    employee_salaries_budget_total = eb
    # total_employee_salaries already set above from Payroll
    employee_salaries_variance = _variance(eb, total_employee_salaries)

    fb = _budget_by_expense_type(['FOOD_CATERING', 'BOARDING_SUPPLIES'])
    fa = _exp_actual('MEALS')
    food_purchases_budget_total = fb
    food_purchases_expenses_total = fa
    food_purchases_variance = _variance(fb, fa)

    xb = _budget_by_expense_type(['TEACHING_MATERIALS', 'LIBRARY_MATERIALS'])
    xa = _exp_actual('EXAMINATION')
    examination_budget_total = xb
    examination_expenses_total = xa
    examination_variance = _variance(xb, xa)

    ab = _budget_by_expense_type(['OFFICE_SUPPLIES', 'COMMUNICATION', 'PROFESSIONAL_FEES'])
    aa = _exp_actual('ADMINISTRATIVE')
    administration_budget_total = ab
    administration_expenses_total = aa
    administration_variance = _variance(ab, aa)

    tb = _budget_by_expense_type(['VEHICLE_EXPENSES', 'TRANSPORT_SERVICES'])
    ta = _exp_actual('TRANSPORT')
    transport_budget_total = tb
    transport_expenses_total = ta
    transport_variance = _variance(tb, ta)

    rb = _budget_by_expense_type(['MAINTENANCE_REPAIRS', 'SECURITY', 'CLEANING_SANITATION'])
    ra = _exp_actual('FACILITIES')
    structural_repairs_budget_total = rb
    structural_repairs_total = ra
    structural_repairs_variance = _variance(rb, ra)

    cb = _budget_by_expense_type(['DEPRECIATION'])
    ca = _exp_actual('CAPITAL')
    capital_budget_total = cb
    capital_expenditure_total = ca
    capital_variance = _variance(cb, ca)

    sb = _budget_by_expense_type(['TEACHING_MATERIALS'])
    sa = _exp_actual('SCHOLASTIC')
    scholastic_materials_budget_total = sb
    scholastic_materials_total = sa
    scholastic_materials_variance = _variance(sb, sa)

    pb = _budget_by_expense_type(['MISCELLANEOUS'])
    pa = _exp_actual('PTA')
    parent_teacher_association_budget_total = pb
    parent_teacher_association_total = pa
    parent_teacher_association_variance = _variance(pb, pa)

    mb = _budget_by_expense_type(['INSURANCE'])
    ma = _exp_actual('MEDICAL')
    medical_supplies_budget_total = mb
    medical_supplies_total = ma
    medical_supplies_variance = _variance(mb, ma)

    spb = _budget_by_expense_type(['IT_EXPENSES'])
    spa = _exp_actual('SPORTS')
    sports_facilitation_budget_total = spb
    sports_facilitation_total = spa
    sports_facilitation_variance = _variance(spb, spa)

    drb = 0
    dra = _exp_actual('DRAWINGS')
    drawings_budget_total = drb
    drawings_total = dra
    drawings_variance = _variance(drb, dra) if drb else None

    lb = _budget_by_expense_type(['PROFESSIONAL_FEES', 'BANK_CHARGES', 'TAXES'])
    la = _exp_actual('LEGAL')
    legal_compliance_budget_total = lb
    legal_compliance_expenses_total = la
    legal_compliance_variance = _variance(lb, la)

    total_budget_variance = _variance(total_budget, total_expenditure)

    # ── Revenue breakdown by payment method ───────────────────────
    rev_by_method = (
        payments_qs
        .values('payment_method__name')
        .annotate(total=Sum('amount_in_school_currency'))
        .order_by('-total')
    )
    total_revenue     = total_paid
    revenue_breakdown = []
    revenue_sources   = []
    revenue_amounts   = []
    for item in rev_by_method:
        src = item['payment_method__name'] or 'Unknown'
        amt = float(item['total'] or 0)
        pct = round(amt / float(total_revenue) * 100, 1) if total_revenue else 0
        revenue_breakdown.append({'source': src, 'amount': amt, 'percentage': pct})
        revenue_sources.append(src)
        revenue_amounts.append(amt)

    # ── Expense breakdown by category ─────────────────────────────
    # Start with payroll (salaries) as the first/largest line
    total_expenses     = total_expenditure
    expense_breakdown  = []
    expense_categories = []
    expense_amounts    = []

    if total_employee_salaries > 0:
        pct = round(total_employee_salaries / float(total_expenses) * 100, 1) if total_expenses else 0
        expense_breakdown.append({'category': 'Staff Salaries', 'amount': total_employee_salaries, 'percentage': pct})
        expense_categories.append('Staff Salaries')
        expense_amounts.append(total_employee_salaries)

    exp_by_cat = (
        expenses_qs
        .values('category__name', 'category__category_type')
        .annotate(total=Sum('total_amount'))
        .order_by('-total')
    )
    for item in exp_by_cat:
        cat = item['category__name'] or item['category__category_type'] or 'Other'
        amt = float(item['total'] or 0)
        pct = round(amt / float(total_expenses) * 100, 1) if total_expenses else 0
        expense_breakdown.append({'category': cat, 'amount': amt, 'percentage': pct})
        expense_categories.append(cat)
        expense_amounts.append(amt)

    # ── Fee collection by fee category ────────────────────────────
    # FeeInvoiceItem.fee_category → FeesCategory (direct FK)
    inv_by_cat = (
        invoices
        .values('items__fee_category__name', 'items__fee_category__category_type')
        .annotate(
            expected=Sum('items__amount'),
            collected=Sum('items__final_amount'),
        )
        .order_by('-expected')
    )
    fee_collection_data = []
    fee_types_list      = []
    exp_amounts_list    = []
    coll_amounts_list   = []
    for row in inv_by_cat:
        cat_name = row['items__fee_category__name'] or row['items__fee_category__category_type'] or 'Other'
        exp  = row['expected']  or 0
        coll = row['collected'] or 0
        rate = round(float(coll) / float(exp) * 100) if exp else 0
        fee_collection_data.append({
            'fee_type_name': cat_name,
            'expected':      exp,
            'collected':     coll,
            'rate':          rate,
        })
        fee_types_list.append(cat_name)
        exp_amounts_list.append(float(exp))
        coll_amounts_list.append(float(coll))

    fee_types_json         = json.dumps(fee_types_list)
    expected_amounts_json  = json.dumps(exp_amounts_list)
    collected_amounts_json = json.dumps(coll_amounts_list)

    # ── Class-wise collection ──────────────────────────────────────
    # Go through invoices directly to avoid multi-table fan-out from StudentClassEnrollment
    class_data = (
        invoices
        .filter(
            student__class_enrollments__academic_session=this_session,
            student__class_enrollments__is_active=True,
        )
        .values('student__class_enrollments__class_instance__academic_level__name')
        .annotate(
            student_count=Count('student', distinct=True),
            expected=Sum('total_amount'),
            collected=Sum('paid_amount'),
        )
        .order_by('student__class_enrollments__class_instance__academic_level__name')
    ) if this_session else []

    class_fee_collection = []
    class_levels_list    = []
    exp_cls_list         = []
    coll_cls_list        = []
    out_cls_list         = []
    total_students       = 0

    for row in class_data:
        level = row['student__class_enrollments__class_instance__academic_level__name'] or 'Unknown'
        exp   = row['expected']  or 0
        coll  = row['collected'] or 0
        out   = exp - coll
        rate  = round(float(coll) / float(exp) * 100, 1) if exp else 0
        cnt   = row['student_count'] or 0
        total_students += cnt
        class_fee_collection.append({
            'class_level':        level,
            'student_count':      cnt,
            'expected_amount':    exp,
            'collected_amount':   coll,
            'outstanding_amount': out,
            'collection_rate':    rate,
        })
        class_levels_list.append(level)
        exp_cls_list.append(float(exp))
        coll_cls_list.append(float(coll))
        out_cls_list.append(float(out))

    overall_collection_rate = overall_rate

    # ── Payment trend (ApexCharts area) ───────────────────────────
    daily = (
        payments_qs
        .annotate(day=TruncDate('payment_date'))
        .values('day')
        .annotate(total=Sum('amount_in_school_currency'))
        .order_by('day')
    )
    chart_data = json.dumps({
        'dates':   [str(r['day']) for r in daily],
        'amounts': [float(r['total'] or 0) for r in daily],
    })

    # ── Monthly trends (Chart.js tabs) ────────────────────────────
    monthly_rev = (
        payments_qs
        .annotate(month=TruncMonth('payment_date'))
        .values('month')
        .annotate(total=Sum('amount_in_school_currency'))
        .order_by('month')
    )
    # Non-salary expenses by month
    monthly_exp = (
        expenses_qs
        .annotate(month=TruncMonth('expense_date'))
        .values('month')
        .annotate(total=Sum('total_amount'))
        .order_by('month')
    )
    # Salary (payroll) costs by month — use payment_date as the month anchor
    monthly_payroll = (
        payroll_qs
        .annotate(month=TruncMonth('payment_date'))
        .values('month')
        .annotate(total=Sum('employer_cost_in_school_currency'))
        .order_by('month')
    )
    all_months = sorted(set(
        [r['month'] for r in monthly_rev     if r['month']] +
        [r['month'] for r in monthly_exp     if r['month']] +
        [r['month'] for r in monthly_payroll if r['month']]
    ))
    rev_map     = {r['month']: float(r['total'] or 0) for r in monthly_rev}
    exp_map     = {r['month']: float(r['total'] or 0) for r in monthly_exp}
    payroll_map = {r['month']: float(r['total'] or 0) for r in monthly_payroll}
    months_labels = [m.strftime('%b %Y') for m in all_months]
    revenue_data  = [rev_map.get(m, 0) for m in all_months]
    # Combined expense = non-salary + payroll per month
    expense_data  = [exp_map.get(m, 0) + payroll_map.get(m, 0) for m in all_months]
    profit_data   = [revenue_data[i] - expense_data[i] for i in range(len(all_months))]

    # ── Cross-session payments ─────────────────────────────────────
    if this_session:
        current_session_payments = float(
            Payment.objects.filter(
                status='COMPLETED', reversed=False, refunded=False,
                academic_session=this_session,
            ).aggregate(t=Sum('amount_in_school_currency'))['t'] or 0
        )
        past_session_recoveries = float(
            Payment.objects.filter(
                status='COMPLETED', reversed=False, refunded=False,
                invoice__academic_session__start_date__lt=this_session.start_date,
                payment_date__range=(this_session.start_date, this_session.end_date),
            ).aggregate(t=Sum('amount_in_school_currency'))['t'] or 0
        )
        future_session_advance_payments = float(
            Payment.objects.filter(
                status='COMPLETED', reversed=False, refunded=False,
                invoice__academic_session__start_date__gt=this_session.end_date,
                payment_date__range=(this_session.start_date, this_session.end_date),
            ).aggregate(t=Sum('amount_in_school_currency'))['t'] or 0
        )
        total_payments_in_session = (
            current_session_payments + past_session_recoveries + future_session_advance_payments
        )

        def _pct(val):
            return round(val / total_payments_in_session * 100, 1) if total_payments_in_session else 0

        current_session_percentage  = _pct(current_session_payments)
        past_session_percentage     = _pct(past_session_recoveries)
        future_session_percentage   = _pct(future_session_advance_payments)
    else:
        current_session_payments = past_session_recoveries = future_session_advance_payments = 0
        total_payments_in_session = 0
        current_session_percentage = past_session_percentage = future_session_percentage = 0

    # ── Uniform sales ──────────────────────────────────────────────
    # UniformSale has no academic_session field — derive via fiscal_period
    uniform_sales_qs = UniformSale.objects.filter(
        fiscal_period__related_academic_session=this_session,
        cancelled=False,
        returned=False,
        status__in=['PAID', 'ISSUED'],
    ) if this_session else UniformSale.objects.none()

    total_uniform_revenue = uniform_sales_qs.aggregate(t=Sum('total_amount'))['t'] or 0
    total_units = UniformSaleItem.objects.filter(
        sale__in=uniform_sales_qs
    ).aggregate(t=Sum('quantity'))['t'] or 0

    uni_by_type = (
        UniformSaleItem.objects.filter(sale__in=uniform_sales_qs)
        .values('uniform_item__name')
        .annotate(units=Sum('quantity'), revenue=Sum('total_price'))
        .order_by('-revenue')
    )
    uniform_sales_data = []
    uniform_types_list = []
    uniform_rev_list   = []
    for row in uni_by_type:
        name = row['uniform_item__name'] or 'Unknown'
        rev  = float(row['revenue'] or 0)
        uniform_sales_data.append({
            'uniform_type': name,
            'units_sold':   row['units'] or 0,
            'revenue':      rev,
        })
        uniform_types_list.append(name)
        uniform_rev_list.append(rev)

    return render(request, 'fees/reports/financial_summary.html', {
        # Session
        'academic_sessions':               academic_sessions,
        'this_session':                    this_session,
        # KPI
        'total_paid':                      total_paid,
        'total_expenditure':               Decimal(str(round(total_expenditure))),
        'total_profit':                    Decimal(str(round(total_profit))),
        'revenue_change':                  revenue_change,
        'expense_change':                  expense_change,
        'profit_change':                   profit_change,
        'fees_coverage_percentage':        fees_coverage_percentage,
        'fees_collected':                  fees_collected,
        'fees_billed':                     fees_billed,
        'total_utilization':               total_utilization,
        'fully_paid_pupils':               fully_paid_pupils,
        'total_pupils_with_invoices':      total_pupils_with_invoices,
        'pupil_payment_completion_rate':   pupil_payment_completion_rate,
        # Charts
        'chart_data':                      chart_data,
        'months':                          json.dumps(months_labels),
        'revenue_data':                    json.dumps(revenue_data),
        'expense_data':                    json.dumps(expense_data),
        'profit_data':                     json.dumps(profit_data),
        # Revenue breakdown
        'revenue_breakdown':               revenue_breakdown,
        'revenue_sources':                 json.dumps(revenue_sources),
        'revenue_amounts':                 json.dumps(revenue_amounts),
        'total_revenue':                   total_revenue,
        # Expense breakdown
        'expense_breakdown':               expense_breakdown,
        'expense_categories':              json.dumps(expense_categories),
        'expense_amounts':                 json.dumps(expense_amounts),
        'total_expenses':                  total_expenses,
        # Fee collection by category
        'fee_collection_data':             fee_collection_data,
        'fee_types_json':                  fee_types_json,
        'expected_amounts_json':           expected_amounts_json,
        'collected_amounts_json':          collected_amounts_json,
        'overall_rate':                    overall_rate,
        # Class-wise
        'class_fee_collection':            class_fee_collection,
        'class_levels':                    json.dumps(class_levels_list),
        'expected_amounts':                json.dumps(exp_cls_list),
        'collected_amounts':               json.dumps(coll_cls_list),
        'outstanding_amounts':             json.dumps(out_cls_list),
        'total_students':                  total_students,
        'total_outstanding':               total_outstanding,
        'overall_collection_rate':         overall_collection_rate,
        # Budget
        'total_budget':                    total_budget,
        'total_budget_variance':           total_budget_variance,
        'employee_salaries_budget_total':  employee_salaries_budget_total,
        'total_employee_salaries':         total_employee_salaries,
        'employee_salaries_variance':      employee_salaries_variance,
        'food_purchases_budget_total':     food_purchases_budget_total,
        'food_purchases_expenses_total':   food_purchases_expenses_total,
        'food_purchases_variance':         food_purchases_variance,
        'examination_budget_total':        examination_budget_total,
        'examination_expenses_total':      examination_expenses_total,
        'examination_variance':            examination_variance,
        'administration_budget_total':     administration_budget_total,
        'administration_expenses_total':   administration_expenses_total,
        'administration_variance':         administration_variance,
        'transport_budget_total':          transport_budget_total,
        'transport_expenses_total':        transport_expenses_total,
        'transport_variance':              transport_variance,
        'structural_repairs_budget_total': structural_repairs_budget_total,
        'structural_repairs_total':        structural_repairs_total,
        'structural_repairs_variance':     structural_repairs_variance,
        'capital_budget_total':            capital_budget_total,
        'capital_expenditure_total':       capital_expenditure_total,
        'capital_variance':                capital_variance,
        'scholastic_materials_budget_total': scholastic_materials_budget_total,
        'scholastic_materials_total':      scholastic_materials_total,
        'scholastic_materials_variance':   scholastic_materials_variance,
        'parent_teacher_association_budget_total': parent_teacher_association_budget_total,
        'parent_teacher_association_total': parent_teacher_association_total,
        'parent_teacher_association_variance': parent_teacher_association_variance,
        'medical_supplies_budget_total':   medical_supplies_budget_total,
        'medical_supplies_total':          medical_supplies_total,
        'medical_supplies_variance':       medical_supplies_variance,
        'sports_facilitation_budget_total': sports_facilitation_budget_total,
        'sports_facilitation_total':       sports_facilitation_total,
        'sports_facilitation_variance':    sports_facilitation_variance,
        'drawings_budget_total':           drawings_budget_total,
        'drawings_total':                  drawings_total,
        'drawings_variance':               drawings_variance,
        'legal_compliance_budget_total':   legal_compliance_budget_total,
        'legal_compliance_expenses_total': legal_compliance_expenses_total,
        'legal_compliance_variance':       legal_compliance_variance,
        # Cross-session
        'current_session_payments':        current_session_payments,
        'past_session_recoveries':         past_session_recoveries,
        'future_session_advance_payments': future_session_advance_payments,
        'total_payments_in_session':       total_payments_in_session,
        'current_session_percentage':      current_session_percentage,
        'past_session_percentage':         past_session_percentage,
        'future_session_percentage':       future_session_percentage,
        # Uniform sales
        'uniform_sales_data':              uniform_sales_data,
        'uniform_types':                   json.dumps(uniform_types_list),
        'uniform_revenue_amounts':         json.dumps(uniform_rev_list),
        'total_units':                     total_units,
        'total_uniform_revenue':           total_uniform_revenue,
        # Meta
        'generated_at':                    get_school_current_time(),
        'generated_by':                    request.user.get_full_name() or request.user.username,
    })


@login_required
def collection_report(request):
    """Fee collection report: payment totals by method, student, session."""
    academic_session_id = request.GET.get('academic_session', '')
    fiscal_period_id    = request.GET.get('fiscal_period', '')
    date_from           = request.GET.get('date_from', '')
    date_to             = request.GET.get('date_to', '')

    filters = {
        'academic_session': academic_session_id if academic_session_id else None,
        'fiscal_period':    fiscal_period_id    if fiscal_period_id    else None,
    }
    if date_from:
        filters['date_from'] = date_from
    if date_to:
        filters['date_to'] = date_to

    try:
        payment_stats = fees_stats.get_payment_statistics(filters=filters)
    except Exception as e:
        logger.error(f"Error generating collection report: {e}", exc_info=True)
        payment_stats = {}

    payments_qs = Payment.objects.filter(status='COMPLETED', reversed=False, refunded=False)
    if academic_session_id:
        payments_qs = payments_qs.filter(academic_session_id=academic_session_id)
    if fiscal_period_id:
        payments_qs = payments_qs.filter(fiscal_period_id=fiscal_period_id)
    if date_from:
        payments_qs = payments_qs.filter(payment_date__gte=date_from)
    if date_to:
        payments_qs = payments_qs.filter(payment_date__lte=date_to)

    by_method = payments_qs.values(
        'payment_method__name', 'payment_method__method_type'
    ).annotate(
        count=Count('id'),
        total=Sum('amount_in_school_currency'),
    ).order_by('-total')

    sessions       = AcademicSession.objects.all().order_by('-start_date')
    fiscal_periods = FiscalPeriod.objects.filter(is_active=True).order_by('-start_date')

    return render(request, 'fees/reports/collection_report.html', {
        'payment_stats':    payment_stats,
        'by_method':        by_method,
        'sessions':         sessions,
        'fiscal_periods':   fiscal_periods,
        'selected_session': academic_session_id,
        'selected_period':  fiscal_period_id,
        'date_from':        date_from,
        'date_to':          date_to,
        'generated_at':     get_school_current_time(),
        'generated_by':     request.user.get_full_name() or request.user.username,
    })


@login_required
def outstanding_report(request):
    """Outstanding fees report: aging analysis, top debtors."""
    academic_session_id = request.GET.get('academic_session', '')
    fiscal_period_id    = request.GET.get('fiscal_period', '')
    today               = get_school_today()

    invoices_qs = FeeInvoice.objects.filter(
        status__in=['PENDING', 'PARTIALLY_PAID', 'OVERDUE']
    ).select_related('student', 'academic_session', 'fiscal_period')

    if academic_session_id:
        invoices_qs = invoices_qs.filter(academic_session_id=academic_session_id)
    if fiscal_period_id:
        invoices_qs = invoices_qs.filter(fiscal_period_id=fiscal_period_id)

    aging = {
        'current':      invoices_qs.filter(due_date__gte=today).aggregate(count=Count('id'), total=Sum('balance')),
        '1_30':         invoices_qs.filter(due_date__lt=today, due_date__gte=today - timedelta(days=30)).aggregate(count=Count('id'), total=Sum('balance')),
        '31_60':        invoices_qs.filter(due_date__lt=today - timedelta(days=30), due_date__gte=today - timedelta(days=60)).aggregate(count=Count('id'), total=Sum('balance')),
        '61_90':        invoices_qs.filter(due_date__lt=today - timedelta(days=60), due_date__gte=today - timedelta(days=90)).aggregate(count=Count('id'), total=Sum('balance')),
        'over_90':      invoices_qs.filter(due_date__lt=today - timedelta(days=90)).aggregate(count=Count('id'), total=Sum('balance')),
    }

    top_debtors = invoices_qs.values(
        'student__first_name', 'student__last_name', 'student__admission_number'
    ).annotate(
        total_outstanding=Sum('balance'),
        invoice_count=Count('id'),
    ).order_by('-total_outstanding')[:20]

    try:
        account_stats = fees_stats.get_student_account_statistics()
    except Exception as e:
        logger.error(f"Error getting account stats: {e}")
        account_stats = {}

    sessions       = AcademicSession.objects.filter(is_active=True).order_by('-start_date')
    fiscal_periods = FiscalPeriod.objects.filter(is_active=True).order_by('-start_date')

    return render(request, 'fees/reports/outstanding_report.html', {
        'aging':            aging,
        'top_debtors':      top_debtors,
        'account_stats':    account_stats,
        'total_outstanding': invoices_qs.aggregate(Sum('balance'))['balance__sum'] or 0,
        'sessions':         sessions,
        'fiscal_periods':   fiscal_periods,
        'selected_session': academic_session_id,
        'selected_period':  fiscal_period_id,
        'today':            today,
        'generated_at':     get_school_current_time(),
        'generated_by':     request.user.get_full_name() or request.user.username,
    })


@login_required
def scholarship_report(request):
    """Scholarship utilisation report."""
    try:
        scholarship_stats = fees_stats.get_scholarship_statistics()
    except Exception as e:
        logger.error(f"Error getting scholarship stats: {e}")
        scholarship_stats = {}

    programs_qs = ScholarshipProgram.objects.filter(is_active=True).annotate(
        active_recipients=Count('student_scholarships', filter=Q(student_scholarships__status='ACTIVE')),
        total_disbursed=Sum('student_scholarships__total_amount_used'),
    ).order_by('-total_disbursed')[:20]

    return render(request, 'fees/reports/scholarship_report.html', {
        'scholarship_stats': scholarship_stats,
        'programs':          programs_qs,
        'generated_at':      get_school_current_time(),
        'generated_by':      request.user.get_full_name() or request.user.username,
    })


@login_required
def discount_report(request):
    """Discount utilisation report."""
    try:
        discount_stats = fees_stats.get_discount_statistics()
        refund_stats   = fees_stats.get_refund_statistics()
    except Exception as e:
        logger.error(f"Error getting discount/refund stats: {e}")
        discount_stats = {}
        refund_stats   = {}

    return render(request, 'fees/reports/discount_report.html', {
        'discount_stats': discount_stats,
        'refund_stats':   refund_stats,
        'generated_at':   get_school_current_time(),
        'generated_by':   request.user.get_full_name() or request.user.username,
    })