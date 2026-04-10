# boarding/views.py

"""
Boarding management views.

CHANGES FROM ORIGINAL
---------------------
IMPORTS
  - Removed DormitoryQuickAddForm and BoardingApprovalForm (deleted from forms.py)

execute_bulk_boarding_enrollment()
  - Removed entirely; BulkBoardingEnrollmentService replaces it

bulk_enrollment_step2()
  - POST handler now validates through BulkBoardingEnrollmentConfirmationForm
    and delegates execution to BulkBoardingEnrollmentService
  - Handles the structured result dict: enrolled_count, failed_count,
    warnings (non-blocking), errors (blocking per-student failures)

boarding_enrollment_terminate()
  - FIX: was calling terminate() then saving effective_end_date separately
    (two saves, two occupancy-count syncs).  Now passes effective_date to
    terminate() so the model method handles everything in one save.

boarding_enrollment_approve()
  - FIX: approved_by is FK to hr.Staff, but the view was passing request.user
    (the auth User model), which would raise an IntegrityError.
    Now looks up the Staff record for the user and falls back to None.

boarding_enrollment_delete()
  - FIX: enrollment.delete() had no try/except, so a ValidationError raised by
    the pre_delete signal (e.g. finalised invoice exists) caused a 500 error.
    Now catches ValidationError explicitly and returns a user-facing message.

get_school_today() used consistently — no bare timezone.now().date() remains.
"""

from django.shortcuts import render, redirect, get_object_or_404
from django.urls import reverse
from django.http import HttpResponseNotAllowed
from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator
from django.contrib import messages
from django.db.models import (
    Q, Count, Sum, Avg, F, Max, Min,
    Case, When, FloatField,
)
from django.utils import timezone
from django.http import JsonResponse, HttpResponse
from django.db import transaction
from django.core.exceptions import ValidationError
from datetime import timedelta, date, datetime
from decimal import Decimal
import logging

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from core.utils import get_school_today, get_school_current_time, format_money

from .models import Dormitory, BoardingEnrollment
from students.models import Student
from academics.models import AcademicSession, AcademicLevel, Class

from .forms import (
    DormitoryFilterForm,
    BoardingEnrollmentFilterForm,
    DormitoryForm,
    BoardingEnrollmentForm,
    BoardingTerminationForm,
    BulkBoardingEnrollmentStudentSelectionForm,
    BulkBoardingEnrollmentConfirmationForm,
)

logger = logging.getLogger(__name__)


@login_required
def boarding_dashboard(request):
    """
    Main boarding dashboard with overview statistics.

    CHANGES from original:
      - Removed Sum('current_occupancy') from dormitory_stats aggregate —
        current_occupancy is now a live @property, not a DB column.
        Total occupancy is counted directly from BoardingEnrollment instead.
      - Removed Sum(F('total_capacity') - F('current_occupancy')) for the
        same reason — available_beds is derived after the queries.
      - Removed Count('id', filter=Q(current_occupancy__gte=F('total_capacity')))
        from the aggregate — cannot filter on a property in ORM.
      - full_dormitories queryset replaced with a Python comprehension using
        the live current_occupancy property.
      - occupancy_percentage now calculated from the live BoardingEnrollment
        count so it always matches what the dormitory cards show.
    """

    today = get_school_today()

    try:
        # ── Dormitory aggregates (DB columns only) ────────────────────────────
        dormitory_stats = Dormitory.objects.aggregate(
            total_dormitories=Count('id'),
            active_dormitories=Count('id', filter=Q(is_active=True)),
            boys_dormitories=Count('id',   filter=Q(dormitory_type='BOYS')),
            girls_dormitories=Count('id',  filter=Q(dormitory_type='GIRLS')),
            mixed_dormitories=Count('id',  filter=Q(dormitory_type='MIXED')),
            total_capacity=Sum('total_capacity'),
            needs_maintenance=Count(
                'id',
                filter=Q(next_maintenance_due__lte=today),
            ),
        )

        # ── Live occupancy — count directly from BoardingEnrollment ──────────
        # current_occupancy is a @property so it cannot appear in aggregate().
        total_occupancy = BoardingEnrollment.objects.filter(
            status='ACTIVE'
        ).count()

        total_capacity = dormitory_stats.get('total_capacity') or 0
        dormitory_stats['total_occupancy'] = total_occupancy
        dormitory_stats['available_beds']  = max(total_capacity - total_occupancy, 0)

        # ── Enrollment aggregates ─────────────────────────────────────────────
        enrollment_stats = BoardingEnrollment.objects.aggregate(
            total_enrollments=Count('id'),
            pending_approval=Count('id',   filter=Q(status='PENDING')),
            active_enrollments=Count('id', filter=Q(status='ACTIVE')),
            suspended_enrollments=Count('id', filter=Q(status='SUSPENDED')),
            terminated_enrollments=Count('id', filter=Q(status='TERMINATED')),
            full_boarders=Count('id',   filter=Q(boarding_type='FULL_BOARDER',   status='ACTIVE')),
            weekly_boarders=Count('id', filter=Q(boarding_type='WEEKLY_BOARDER', status='ACTIVE')),
            flexi_boarders=Count('id',  filter=Q(boarding_type='FLEXI_BOARDER',  status='ACTIVE')),
            with_consent=Count('id',    filter=Q(guardian_consent=True)),
            without_consent=Count('id', filter=Q(guardian_consent=False)),
            with_invoice=Count('id',    filter=Q(boarding_invoice__isnull=False)),
        )

    except Exception as e:
        logger.error("Error getting boarding dashboard statistics: %s", e)
        dormitory_stats  = {}
        enrollment_stats = {}
        total_capacity   = 0
        total_occupancy  = 0

    occupancy_percentage = round(
        (total_occupancy / total_capacity * 100) if total_capacity else 0, 1
    )

    recent_enrollments = BoardingEnrollment.objects.select_related(
        'student', 'dormitory', 'academic_session',
    ).order_by('-created_at')[:10]

    pending_approvals = BoardingEnrollment.objects.select_related(
        'student', 'dormitory', 'academic_session',
    ).filter(status='PENDING').order_by('created_at')[:10]

    # full_dormitories — current_occupancy is a @property so we cannot use
    # ORM filter Q(current_occupancy__gte=F('total_capacity')).
    # Evaluate in Python; at most a handful of dormitories per school.
    full_dormitories = [
        d for d in Dormitory.objects.filter(
            is_active=True,
        ).order_by('dormitory_type', 'name')
        if d.current_occupancy >= d.total_capacity
    ][:10]

    maintenance_due = Dormitory.objects.filter(
        is_active=True,
        next_maintenance_due__lte=today,
    ).order_by('next_maintenance_due')[:10]

    missing_consent = BoardingEnrollment.objects.select_related(
        'student', 'dormitory',
    ).filter(
        status__in=('PENDING', 'ACTIVE'),
        guardian_consent=False,
    ).order_by('enrollment_date')[:10]

    return render(request, 'boarding/dashboard.html', {
        'dormitory_stats':      dormitory_stats,
        'enrollment_stats':     enrollment_stats,
        'occupancy_percentage': occupancy_percentage,
        'recent_enrollments':   recent_enrollments,
        'pending_approvals':    pending_approvals,
        'full_dormitories':     full_dormitories,
        'maintenance_due':      maintenance_due,
        'missing_consent':      missing_consent,
    })


# =============================================================================
# QUERYSET HELPERS
# =============================================================================

def get_filtered_dormitories(request):
    """
    Return a filtered, annotated queryset of Dormitory objects for the list view.
 
    Occupancy annotations (occupancy_ratio, available_beds) have been removed
    because current_occupancy is now a live @property on the model, not a
    database column.  The dormitory list template calls model methods directly:
        dormitory.get_occupancy_percentage()
        dormitory.get_available_capacity()
 
    Remaining annotation:
        active_enrollment_count — used in the Boarders column of the list table.
    """
    from django.db.models import Count, Q, Case, When, Value, BooleanField
    from .models import Dormitory
 
    qs = Dormitory.objects.annotate(
        active_enrollment_count=Count(
            'boarding_enrollments',
            filter=Q(boarding_enrollments__status='ACTIVE'),
        ),
    )
 
    # ── Text search ─────────────────────────────────────────────────────────
    query = request.GET.get('q', '').strip()
    if query:
        words = query.split()
        q_filter = Q()
        for word in words:
            q_filter &= (
                Q(name__icontains=word)        |
                Q(code__icontains=word)        |
                Q(building__icontains=word)    |
                Q(description__icontains=word)
            )
        qs = qs.filter(q_filter)
 
    # ── Choice filters ───────────────────────────────────────────────────────
    for param, field in (
        ('dormitory_type',    'dormitory_type'),
        ('maintenance_status', 'maintenance_status'),
    ):
        v = request.GET.get(param, '')
        if v:
            qs = qs.filter(**{field: v})
 
    for param in ('is_active', 'is_available_for_new_admissions'):
        v = request.GET.get(param, '')
        if v:
            qs = qs.filter(**{param: v.lower() == 'true'})
 
    master = request.GET.get('dormitory_master', '')
    if master:
        try:
            qs = qs.filter(dormitory_master_id=int(master))
        except (ValueError, TypeError):
            pass
 
    # ── Occupancy level filter ───────────────────────────────────────────────
    # occupancy_level can't be a DB annotation anymore, so filter in Python.
    # Since we have at most ~10 dormitories per school this is fine.
    occupancy_level = request.GET.get('occupancy_level', '')
    if occupancy_level:
        filtered_ids = []
        for d in qs:
            pct = d.get_occupancy_percentage()
            if occupancy_level == 'EMPTY'   and pct == 0:              filtered_ids.append(d.pk)
            elif occupancy_level == 'LOW'   and 0 < pct < 50:          filtered_ids.append(d.pk)
            elif occupancy_level == 'MED'   and 50 <= pct < 80:        filtered_ids.append(d.pk)
            elif occupancy_level == 'HIGH'  and 80 <= pct < 100:       filtered_ids.append(d.pk)
            elif occupancy_level == 'FULL'  and pct >= 100:            filtered_ids.append(d.pk)
        qs = qs.filter(pk__in=filtered_ids)
 
    return qs.order_by('dormitory_type', 'name')


def get_filtered_boarding_enrollments(request):
    """Build a filtered boarding enrollment queryset from request.GET."""
    enrollments = BoardingEnrollment.objects.select_related(
        'student__current_academic_level',
        'academic_session',
        'dormitory',
        'consenting_guardian',
        'approved_by',
        'boarding_invoice',
    ).order_by('-academic_session__start_date', 'dormitory', 'boarding_roll_number')

    query             = request.GET.get('q',                '').strip()
    status            = request.GET.get('status',           '')
    boarding_type     = request.GET.get('boarding_type',    '')
    dormitory         = request.GET.get('dormitory',        '')
    academic_session  = request.GET.get('academic_session', '')
    guardian_consent  = request.GET.get('guardian_consent', '')
    date_from         = request.GET.get('enrollment_date_from', '')
    date_to           = request.GET.get('enrollment_date_to',   '')
    student_gender    = request.GET.get('student_gender',   '')

    if query:
        words = query.split()
        q_filter = Q()
        for word in words:
            q_filter &= (
                Q(student__first_name__icontains=word)     |
                Q(student__last_name__icontains=word)      |
                Q(student__admission_number__icontains=word) |
                Q(boarding_roll_number__icontains=word)    |
                Q(room_number__icontains=word)             |
                Q(bed_number__icontains=word)
            )
        enrollments = enrollments.filter(q_filter)

    if status:
        enrollments = enrollments.filter(status=status)
    if boarding_type:
        enrollments = enrollments.filter(boarding_type=boarding_type)
    if dormitory:
        try:
            enrollments = enrollments.filter(dormitory_id=int(dormitory))
        except (ValueError, TypeError):
            pass
    if academic_session:
        try:
            enrollments = enrollments.filter(academic_session_id=int(academic_session))
        except (ValueError, TypeError):
            pass
    if guardian_consent:
        enrollments = enrollments.filter(
            guardian_consent=(guardian_consent.lower() == 'true')
        )
    if date_from:
        enrollments = enrollments.filter(enrollment_date__gte=date_from)
    if date_to:
        enrollments = enrollments.filter(enrollment_date__lte=date_to)
    if student_gender:
        enrollments = enrollments.filter(student__gender=student_gender)

    return enrollments


# =============================================================================
# DORMITORY VIEWS
# =============================================================================

@login_required
def dormitory_list(request):
    """
    Dormitory list — handles both full page loads and HTMX filter requests.

    Stats are computed on a separate unannotated queryset with the same
    filters applied so that the active_enrollment_count annotation on the
    main queryset does not interfere with aggregate calculations.

    CHANGE: current_occupancy is now a live @property on Dormitory, not a
    DB column, so it cannot be used in ORM aggregations or filters.

      - Sum('current_occupancy')             → separate BoardingEnrollment count
      - filter(current_occupancy__gte=...)   → Python comprehension over stats_qs
    """
    filter_form  = DormitoryFilterForm(request.GET or None)
    dormitories  = get_filtered_dormitories(request)

    # -- Stats on a clean (unannotated) queryset --------------------------------
    stats_qs = Dormitory.objects.all()

    query = request.GET.get('q', '').strip()
    if query:
        words    = query.split()
        q_filter = Q()
        for word in words:
            q_filter &= (
                Q(name__icontains=word)     |
                Q(code__icontains=word)     |
                Q(building__icontains=word) |
                Q(description__icontains=word)
            )
        stats_qs = stats_qs.filter(q_filter)

    for param, field in (
        ('dormitory_type',    'dormitory_type'),
        ('maintenance_status', 'maintenance_status'),
    ):
        v = request.GET.get(param, '')
        if v:
            stats_qs = stats_qs.filter(**{field: v})

    for param in ('is_active', 'is_available_for_new_admissions'):
        v = request.GET.get(param, '')
        if v:
            stats_qs = stats_qs.filter(**{param: v.lower() == 'true'})

    master = request.GET.get('dormitory_master', '')
    if master:
        try:
            stats_qs = stats_qs.filter(dormitory_master_id=int(master))
        except (ValueError, TypeError):
            pass

    # Aggregate the fields that are real DB columns.
    stats = stats_qs.aggregate(
        total=Count('id'),
        active=Count('id', filter=Q(is_active=True)),
        boys=Count('id',   filter=Q(dormitory_type='BOYS')),
        girls=Count('id',  filter=Q(dormitory_type='GIRLS')),
        mixed=Count('id',  filter=Q(dormitory_type='MIXED')),
        total_capacity=Sum('total_capacity'),
    )

    # Live occupancy — current_occupancy is a @property, not a DB column, so
    # Sum('current_occupancy') would raise FieldError.  Count directly from
    # BoardingEnrollment for the filtered dormitory set instead.
    total_occ = BoardingEnrollment.objects.filter(
        dormitory__in=stats_qs,
        status='ACTIVE',
    ).count()
    stats['total_occupancy'] = total_occ

    total_cap = stats.get('total_capacity') or 0
    stats['available_beds'] = total_cap - total_occ
    stats['avg_occupancy']  = round(total_occ / total_cap * 100, 1) if total_cap else 0

    # full_dormitories — current_occupancy__gte=F('total_capacity') is not a
    # valid ORM filter on a property.  Evaluate in Python; stats_qs is small.
    stats['full_dormitories'] = sum(
        1 for d in stats_qs if d.current_occupancy >= d.total_capacity
    )

    paginator        = Paginator(dormitories, 20)
    dormitories_page = paginator.get_page(request.GET.get('page', 1))
    is_htmx          = request.headers.get('HX-Request') == 'true'

    context = {
        'dormitories_page': dormitories_page,
        'paginator':        paginator,
        'stats':            stats,
        'filter_form':      filter_form,
        'is_htmx':          is_htmx,
    }

    if is_htmx:
        return render(
            request,
            'boarding/dormitories/partials/_dormitory_results.html',
            context,
        )
    return render(request, 'boarding/dormitories/list.html', context)


@login_required
def dormitory_create(request):
    """Create a new dormitory."""
    if request.method == 'POST':
        form = DormitoryForm(request.POST)
        if form.is_valid():
            dormitory = form.save()
            messages.success(
                request,
                f"Dormitory '{dormitory.name}' created successfully",
                extra_tags='sweetalert',
            )
            return redirect('boarding:dormitory_detail', pk=dormitory.pk)
    else:
        form = DormitoryForm()

    return render(request, 'boarding/dormitories/form.html', {
        'form':  form,
        'title': 'Create Dormitory',
    })


@login_required
def dormitory_edit(request, pk):
    """Edit an existing dormitory."""
    dormitory = get_object_or_404(Dormitory, pk=pk)

    if request.method == 'POST':
        form = DormitoryForm(request.POST, instance=dormitory)
        if form.is_valid():
            dormitory = form.save()
            messages.success(
                request,
                f"Dormitory '{dormitory.name}' updated successfully",
                extra_tags='sweetalert',
            )
            return redirect('boarding:dormitory_detail', pk=dormitory.pk)
    else:
        form = DormitoryForm(instance=dormitory)

    return render(request, 'boarding/dormitories/form.html', {
        'form':      form,
        'dormitory': dormitory,
        'title':     f'Edit {dormitory.name}',
    })


@login_required
def dormitory_detail(request, pk):
    """Dormitory detail page — facility info + HTMX-loaded session-scoped residents."""
    dormitory = get_object_or_404(
        Dormitory.objects.select_related(
            'dormitory_master',
            'assistant_dormitory_master',
        ),
        pk=pk,
    )
 
    all_sessions = AcademicSession.objects.filter(
        Q(boarding_enrollments__dormitory=dormitory) | Q(is_active=True)
    ).distinct().order_by('-start_date')
 
    default_session = AcademicSession.get_current_session()
    if not default_session:
        default_session = AcademicSession.objects.filter(
            is_active=True
        ).order_by('-start_date').first()
 
    return render(request, 'boarding/dormitories/detail.html', {
        'dormitory':       dormitory,
        'all_sessions':    all_sessions,
        'default_session': default_session,
    })
 
 
@login_required
def dormitory_residents_partial(request, pk):
    """
    HTMX endpoint: session-scoped resident table for a dormitory.
 
    Gender is enforced automatically:
        BOYS  dormitory → student__gender='M' (cannot be overridden by filters)
        GIRLS dormitory → student__gender='F' (cannot be overridden by filters)
        MIXED dormitory → no restriction
 
    Accepts GET params:
        session_id       — which academic session to show
        q                — search (name, admission no., roll no.)
        status           — enrollment status filter
        boarding_type    — FULL_BOARDER / WEEKLY_BOARDER / FLEXI_BOARDER
        guardian_consent — true / false
        has_invoice      — true / false
        page             — pagination
    """
    dormitory  = get_object_or_404(Dormitory, pk=pk)
    session_id = request.GET.get('session_id', '').strip()
 
    # Resolve session
    session = None
    if session_id:
        try:
            session = AcademicSession.objects.get(pk=session_id)
        except (AcademicSession.DoesNotExist, ValueError):
            pass
 
    if session is None:
        session = AcademicSession.get_current_session()
    if session is None:
        session = AcademicSession.objects.filter(
            is_active=True,
        ).order_by('-start_date').first()
 
    if session is None:
        return render(
            request,
            'boarding/dormitories/partials/_dormitory_residents.html',
            {
                'dormitory':             dormitory,
                'session':               None,
                'enrollments':           [],
                'stats':                 {},
                'missing_invoice_count': 0,
            },
        )
 
    # ── Base queryset ──────────────────────────────────────────────────────────
    qs = BoardingEnrollment.objects.filter(
        dormitory=dormitory,
        academic_session=session,
    ).select_related(
        'student',
        'boarding_invoice',
        'approved_by',
    )
 
    # ── Gender auto-filter — enforced before any user-supplied filters ─────────
    # This is a hard constraint, not a user-facing option.  It ensures that a
    # Boys dormitory can never surface female students regardless of what the
    # filter bar sends.
    if dormitory.dormitory_type == 'BOYS':
        qs = qs.filter(student__gender='M')
    elif dormitory.dormitory_type == 'GIRLS':
        qs = qs.filter(student__gender='F')
    # MIXED / STAFF → no restriction
 
    # Keep a copy of the gender-filtered queryset for the summary strip so that
    # the aggregate counts are also gender-correct.
    base = qs
 
    # ── User-supplied filters ──────────────────────────────────────────────────
    q             = request.GET.get('q',             '').strip()
    status        = request.GET.get('status',        '')
    boarding_type = request.GET.get('boarding_type', '')
    consent       = request.GET.get('guardian_consent', '')
    has_invoice   = request.GET.get('has_invoice',   '')
 
    if q:
        qs = qs.filter(
            Q(student__first_name__icontains=q)       |
            Q(student__last_name__icontains=q)        |
            Q(student__admission_number__icontains=q) |
            Q(boarding_roll_number__icontains=q)
        )
    if status:
        qs = qs.filter(status=status)
    if boarding_type:
        qs = qs.filter(boarding_type=boarding_type)
    if consent:
        qs = qs.filter(guardian_consent=(consent.lower() == 'true'))
    if has_invoice == 'true':
        qs = qs.filter(boarding_invoice__isnull=False)
    elif has_invoice == 'false':
        qs = qs.filter(boarding_invoice__isnull=True)
 
    qs = qs.order_by('boarding_roll_number', 'student__last_name')
 
    # ── Summary strip stats — always unfiltered (except gender) ───────────────
    stats = base.aggregate(
        total=Count('id'),
        active=Count('id',          filter=Q(status='ACTIVE')),
        pending=Count('id',         filter=Q(status='PENDING')),
        full_boarders=Count('id',   filter=Q(boarding_type='FULL_BOARDER',   status='ACTIVE')),
        weekly_boarders=Count('id', filter=Q(boarding_type='WEEKLY_BOARDER', status='ACTIVE')),
        flexi_boarders=Count('id',  filter=Q(boarding_type='FLEXI_BOARDER',  status='ACTIVE')),
        with_consent=Count('id',    filter=Q(guardian_consent=True,  status='ACTIVE')),
        without_consent=Count('id', filter=Q(guardian_consent=False, status='ACTIVE')),
    )
 
    missing_invoice_count = base.filter(
        boarding_invoice__isnull=True,
        status='ACTIVE',
    ).count()
 
    # ── Pagination ─────────────────────────────────────────────────────────────
    paginator = Paginator(qs, 10)
    page_obj  = paginator.get_page(request.GET.get('page', 1))
 
    return render(
        request,
        'boarding/dormitories/partials/_dormitory_residents.html',
        {
            'dormitory':             dormitory,
            'session':               session,
            'enrollments':           page_obj,
            'stats':                 stats,
            'missing_invoice_count': missing_invoice_count,
        },
    )


@login_required
def dormitory_delete(request, pk):
    """Delete a dormitory — blocked if it has any boarding enrollments."""
    dormitory = get_object_or_404(Dormitory, pk=pk)
    is_htmx   = request.headers.get('HX-Request') == 'true'

    if request.method == 'POST':
        if dormitory.boarding_enrollments.exists():
            msg = f"Cannot delete '{dormitory.name}' — it has boarding enrollments"
            if is_htmx:
                r = HttpResponse()
                r['HX-Alert-Message'] = msg
                r['HX-Alert-Type']    = 'error'
                r['HX-Close-Modal']   = 'true'
                return r
            messages.error(request, msg, extra_tags='sweetalert-error')
            return redirect('boarding:dormitory_list')

        name = dormitory.name
        dormitory.delete()

        msg = f"Dormitory '{name}' deleted successfully"
        if is_htmx:
            r = HttpResponse()
            r['HX-Alert-Message'] = msg
            r['HX-Alert-Type']    = 'success'
            r['HX-Close-Modal']   = 'true'
            r['HX-Redirect']      = reverse('boarding:dormitory_list')
            return r
        messages.success(request, msg, extra_tags='sweetalert')
        return redirect('boarding:dormitory_list')


@login_required
def dormitory_activate(request, pk):
    """Activate a dormitory."""
    dormitory = get_object_or_404(Dormitory, pk=pk)
    is_htmx   = request.headers.get('HX-Request') == 'true'

    if request.method == 'POST':
        if dormitory.is_active:
            msg = f"{dormitory.name} is already active"
            if is_htmx:
                r = HttpResponse()
                r['HX-Alert-Message'] = msg
                r['HX-Alert-Type']    = 'warning'
                r['HX-Close-Modal']   = 'true'
                return r
            messages.warning(request, msg, extra_tags='sweetalert')
            return redirect('boarding:dormitory_detail', pk=pk)

        dormitory.is_active = True
        dormitory.save(update_fields=['is_active', 'updated_at'])

        msg = f"{dormitory.name} activated successfully"
        if is_htmx:
            r = HttpResponse()
            r['HX-Alert-Message'] = msg
            r['HX-Alert-Type']    = 'success'
            r['HX-Close-Modal']   = 'true'
            r['HX-Redirect']      = reverse(
                'boarding:dormitory_detail', kwargs={'pk': pk}
            )
            return r
        messages.success(request, msg, extra_tags='sweetalert')
        return redirect('boarding:dormitory_detail', pk=pk)


@login_required
def dormitory_deactivate(request, pk):
    """Deactivate a dormitory."""
    dormitory = get_object_or_404(Dormitory, pk=pk)
    is_htmx   = request.headers.get('HX-Request') == 'true'

    if request.method == 'POST':
        if not dormitory.is_active:
            msg = f"{dormitory.name} is already inactive"
            if is_htmx:
                r = HttpResponse()
                r['HX-Alert-Message'] = msg
                r['HX-Alert-Type']    = 'warning'
                r['HX-Close-Modal']   = 'true'
                return r
            messages.warning(request, msg, extra_tags='sweetalert')
            return redirect('boarding:dormitory_detail', pk=pk)

        dormitory.is_active = False
        dormitory.save(update_fields=['is_active', 'updated_at'])

        msg = f"{dormitory.name} deactivated"
        if is_htmx:
            r = HttpResponse()
            r['HX-Alert-Message'] = msg
            r['HX-Alert-Type']    = 'warning'
            r['HX-Close-Modal']   = 'true'
            r['HX-Redirect']      = reverse(
                'boarding:dormitory_detail', kwargs={'pk': pk}
            )
            return r
        messages.warning(request, msg, extra_tags='sweetalert')
        return redirect('boarding:dormitory_detail', pk=pk)
    
@login_required
def dormitory_update_maintenance(request, pk):
    """POST — update maintenance status and dates for a dormitory."""
    dormitory = get_object_or_404(Dormitory, pk=pk)

    if request.method != 'POST':
        return HttpResponseNotAllowed(['POST'])

    maintenance_status    = request.POST.get('maintenance_status')
    last_maintenance_date = request.POST.get('last_maintenance_date') or None
    next_maintenance_due  = request.POST.get('next_maintenance_due') or None

    update_fields = []

    if maintenance_status in dict(Dormitory.MAINTENANCE_STATUS_CHOICES):
        dormitory.maintenance_status = maintenance_status
        update_fields.append('maintenance_status')

    if last_maintenance_date:
        from datetime import date
        try:
            dormitory.last_maintenance_date = date.fromisoformat(last_maintenance_date)
            update_fields.append('last_maintenance_date')
        except ValueError:
            pass
    else:
        dormitory.last_maintenance_date = None
        update_fields.append('last_maintenance_date')

    if next_maintenance_due:
        try:
            dormitory.next_maintenance_due = date.fromisoformat(next_maintenance_due)
            update_fields.append('next_maintenance_due')
        except ValueError:
            pass
    else:
        dormitory.next_maintenance_due = None
        update_fields.append('next_maintenance_due')

    if update_fields:
        dormitory.save(update_fields=update_fields)

    response = HttpResponse()
    response['HX-Close-Modal']    = 'true'
    response['HX-Alert-Message']  = f"Maintenance updated for {dormitory.name}."
    response['HX-Alert-Type']     = 'success'
    return response


@login_required
def dormitory_print_view(request):
    """Printable dormitory list."""
    selected_fields = request.GET.getlist('fields') or [
        'code', 'name', 'dormitory_type', 'total_capacity',
        'current_occupancy', 'is_active',
    ]
    include_stats = request.GET.get('include_stats') == 'true'
    landscape     = request.GET.get('landscape') == 'true'
    dormitories   = get_filtered_dormitories(request)

    stats = None
    if include_stats:
        stats = dormitories.aggregate(
            total=Count('id'),
            active=Count('id', filter=Q(is_active=True)),
            total_capacity=Sum('total_capacity'),
            total_occupancy=Sum('current_occupancy'),
        )

    field_names = {
        'code':              'Dormitory Code',
        'name':              'Dormitory Name',
        'dormitory_type':    'Type',
        'total_capacity':    'Capacity',
        'current_occupancy': 'Occupancy',
        'is_active':         'Active',
    }

    return render(request, 'boarding/dormitories/print.html', {
        'dormitories':          dormitories,
        'stats':                stats,
        'now':                  timezone.now(),
        'selected_fields':      selected_fields,
        'selected_field_names': [
            field_names.get(f, f.replace('_', ' ').title())
            for f in selected_fields
        ],
        'field_names': field_names,
        'landscape':   landscape,
    })


@login_required
def export_dormitories_excel(request):
    """Export filtered dormitories to Excel."""
    dormitories = get_filtered_dormitories(request)

    wb = Workbook()
    ws = wb.active
    ws.title = "Dormitories"

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)

    headers = [
        '#', 'Code', 'Name', 'Type', 'Building', 'Capacity',
        'Occupancy', 'Available', 'Active', 'Dormitory Master',
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.fill  = header_fill
        cell.font  = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')

    for idx, d in enumerate(dormitories, start=1):
        ws.append([
            idx,
            d.code,
            d.name,
            d.get_dormitory_type_display(),
            d.building or '',
            d.total_capacity,
            d.current_occupancy,
            d.get_available_capacity(),
            'Yes' if d.is_active else 'No',
            str(d.dormitory_master) if d.dormitory_master else '',
        ])

    for col in ws.columns:
        max_len = max(
            (len(str(c.value)) for c in col if c.value), default=0
        )
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 2, 50)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = (
        f'attachment; filename="dormitories_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
    )
    wb.save(response)
    return response


# =============================================================================
# BOARDING ENROLLMENT VIEWS
# =============================================================================

@login_required
def boarding_enrollment_create(request):
    """Create a new boarding enrollment."""
    if request.method == 'POST':
        form = BoardingEnrollmentForm(request.POST)
        if form.is_valid():
            enrollment = form.save()
            messages.success(
                request,
                f"Boarding enrollment for "
                f"{enrollment.student.get_full_name()} created successfully",
                extra_tags='sweetalert',
            )
            return redirect('boarding:enrollment_detail', pk=enrollment.pk)
    else:
        form = BoardingEnrollmentForm()

    return render(request, 'boarding/enrollments/form.html', {
        'form':       form,
        'enrollment': None,
        'title':      'Create Boarding Enrollment',
    })


@login_required
def boarding_enrollment_edit(request, pk):
    """Edit an existing boarding enrollment."""
    enrollment = get_object_or_404(BoardingEnrollment, pk=pk)

    if request.method == 'POST':
        form = BoardingEnrollmentForm(request.POST, instance=enrollment)
        if form.is_valid():
            enrollment = form.save()
            messages.success(
                request,
                "Boarding enrollment updated successfully",
                extra_tags='sweetalert',
            )
            return redirect('boarding:enrollment_detail', pk=enrollment.pk)
    else:
        form = BoardingEnrollmentForm(instance=enrollment)

    return render(request, 'boarding/enrollments/form.html', {
        'form':       form,
        'enrollment': enrollment,
        'title':      'Edit Boarding Enrollment',
    })


@login_required
def boarding_enrollment_detail(request, pk):
    """Boarding enrollment detail page."""
    enrollment = get_object_or_404(
        BoardingEnrollment.objects.select_related(
            'student', 'academic_session', 'dormitory',
            'consenting_guardian', 'approved_by', 'boarding_invoice',
        ),
        pk=pk,
    )
    return render(request, 'boarding/enrollments/detail.html', {
        'enrollment': enrollment,
    })


@login_required
def boarding_enrollment_approve(request, pk):
    """
    Approve a PENDING boarding enrollment.

    FIX: approved_by is FK to hr.Staff, not to the auth User model.
    Now looks up the Staff record for the current user and falls back to None
    rather than passing request.user directly (which would cause an IntegrityError).
    """
    enrollment = get_object_or_404(BoardingEnrollment, pk=pk)
    is_htmx    = request.headers.get('HX-Request') == 'true'

    if request.method == 'POST':
        if enrollment.status != 'PENDING':
            msg = 'Only pending enrollments can be approved'
            if is_htmx:
                r = HttpResponse()
                r['HX-Alert-Message'] = msg
                r['HX-Alert-Type']    = 'error'
                r['HX-Close-Modal']   = 'true'
                return r
            messages.error(request, msg, extra_tags='sweetalert-error')
            return redirect('boarding:enrollment_detail', pk=pk)

        # Resolve the Staff record for the approving user.
        approving_staff = None
        try:
            from hr.models import Staff
            approving_staff = Staff.objects.get(user=request.user)
        except Exception:
            logger.warning(
                "Could not resolve Staff for user %s — approved_by will be null",
                request.user,
            )

        enrollment.approve(approved_by=approving_staff)

        msg = 'Boarding enrollment approved successfully'
        if is_htmx:
            r = HttpResponse()
            r['HX-Alert-Message'] = msg
            r['HX-Alert-Type']    = 'success'
            r['HX-Close-Modal']   = 'true'
            r['HX-Redirect']      = reverse(
                'boarding:enrollment_detail', kwargs={'pk': pk}
            )
            return r
        messages.success(request, msg, extra_tags='sweetalert')
        return redirect('boarding:enrollment_detail', pk=pk)


@login_required
def boarding_enrollment_terminate(request, pk):
    """
    Terminate a boarding enrollment.

    FIX: original called enrollment.terminate() then set effective_end_date
    and saved again — two saves, two occupancy-count syncs.  Now passes
    effective_date to terminate() so everything happens in one save.
    """
    enrollment = get_object_or_404(BoardingEnrollment, pk=pk)
    is_htmx    = request.headers.get('HX-Request') == 'true'

    if request.method == 'POST':
        form = BoardingTerminationForm(request.POST)
        if form.is_valid():
            enrollment.terminate(
                reason=form.cleaned_data['termination_reason'],
                effective_date=form.cleaned_data['effective_termination_date'],
            )

            msg = 'Boarding enrollment terminated successfully'
            if is_htmx:
                r = HttpResponse()
                r['HX-Alert-Message'] = msg
                r['HX-Alert-Type']    = 'success'
                r['HX-Close-Modal']   = 'true'
                r['HX-Redirect']      = reverse(
                    'boarding:enrollment_detail', kwargs={'pk': pk}
                )
                return r
            messages.success(request, msg, extra_tags='sweetalert')
            return redirect('boarding:enrollment_detail', pk=pk)

        # Form has errors — re-render the modal with validation messages.
        return render(
            request,
            'boarding/enrollments/modals/terminate_enrollment.html',
            {'form': form, 'enrollment': enrollment},
        )


@login_required
def boarding_enrollment_suspend(request, pk):
    """Suspend an ACTIVE boarding enrollment."""
    enrollment = get_object_or_404(BoardingEnrollment, pk=pk)
    is_htmx    = request.headers.get('HX-Request') == 'true'

    if request.method == 'POST':
        reason = request.POST.get('reason', '').strip()
        if not reason:
            return render(
                request,
                'boarding/enrollments/modals/suspend_enrollment.html',
                {'enrollment': enrollment, 'error_message': 'Suspension reason is required'},
            )

        enrollment.suspend(reason=reason)

        msg = 'Boarding enrollment suspended successfully'
        if is_htmx:
            r = HttpResponse()
            r['HX-Alert-Message'] = msg
            r['HX-Alert-Type']    = 'success'
            r['HX-Close-Modal']   = 'true'
            r['HX-Redirect']      = reverse(
                'boarding:enrollment_detail', kwargs={'pk': pk}
            )
            return r
        messages.success(request, msg, extra_tags='sweetalert')
        return redirect('boarding:enrollment_detail', pk=pk)


@login_required
def boarding_enrollment_delete(request, pk):
    """
    Delete a boarding enrollment.

    FIX: enrollment.delete() previously had no error handling.  A ValidationError
    raised by the pre_delete signal (e.g. a finalised invoice is linked) would
    cause a 500.  Now catches ValidationError explicitly and returns a user-facing
    error message.
    """
    enrollment = get_object_or_404(BoardingEnrollment, pk=pk)
    is_htmx    = request.headers.get('HX-Request') == 'true'

    if request.method == 'POST':
        student_name = enrollment.student.get_full_name()

        try:
            enrollment.delete()
        except ValidationError as exc:
            # Pre-delete signal raised because invoice cannot be safely removed.
            err_msg = (
                exc.message
                if hasattr(exc, 'message')
                else '; '.join(exc.messages)
            )
            if is_htmx:
                r = HttpResponse()
                r['HX-Alert-Message'] = err_msg
                r['HX-Alert-Type']    = 'error'
                r['HX-Close-Modal']   = 'true'
                return r
            messages.error(request, err_msg, extra_tags='sweetalert-error')
            return redirect('boarding:enrollment_detail', pk=pk)
        except Exception as exc:
            logger.error("Error deleting boarding enrollment %s: %s", pk, exc, exc_info=True)
            err_msg = f"Could not delete enrollment: {exc}"
            if is_htmx:
                r = HttpResponse()
                r['HX-Alert-Message'] = err_msg
                r['HX-Alert-Type']    = 'error'
                r['HX-Close-Modal']   = 'true'
                return r
            messages.error(request, err_msg, extra_tags='sweetalert-error')
            return redirect('boarding:enrollment_detail', pk=pk)

        msg = f"Boarding enrollment for '{student_name}' deleted successfully"
        if is_htmx:
            r = HttpResponse()
            r['HX-Alert-Message'] = msg
            r['HX-Alert-Type']    = 'success'
            r['HX-Close-Modal']   = 'true'
            r['HX-Redirect']      = reverse('boarding:enrollment_list')
            return r
        messages.success(request, msg, extra_tags='sweetalert')
        return redirect('boarding:enrollment_list')


@login_required
def boarding_enrollment_print_view(request):
    """Printable enrollment list."""
    selected_fields = request.GET.getlist('fields') or [
        'student_name', 'dormitory', 'boarding_type',
        'status', 'enrollment_date', 'boarding_roll_number',
    ]
    include_stats = request.GET.get('include_stats') == 'true'
    landscape     = request.GET.get('landscape') == 'true'
    enrollments   = get_filtered_boarding_enrollments(request)

    stats = None
    if include_stats:
        stats = enrollments.aggregate(
            total=Count('id'),
            active=Count('id', filter=Q(status='ACTIVE')),
        )

    field_names = {
        'student_name':       'Student Name',
        'dormitory':          'Dormitory',
        'boarding_type':      'Boarding Type',
        'status':             'Status',
        'enrollment_date':    'Enrollment Date',
        'boarding_roll_number': 'Roll Number',
    }

    return render(request, 'boarding/enrollments/print.html', {
        'enrollments':          enrollments,
        'stats':                stats,
        'now':                  timezone.now(),
        'selected_fields':      selected_fields,
        'selected_field_names': [
            field_names.get(f, f.replace('_', ' ').title())
            for f in selected_fields
        ],
        'field_names': field_names,
        'landscape':   landscape,
    })


@login_required
def export_boarding_enrollments_excel(request):
    """Export filtered boarding enrollments to Excel."""
    enrollments = get_filtered_boarding_enrollments(request)

    wb = Workbook()
    ws = wb.active
    ws.title = "Boarding Enrollments"

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)

    headers = [
        '#', 'Student Name', 'Admission No.', 'Dormitory', 'Boarding Type',
        'Status', 'Enrollment Date', 'Room / Bed', 'Session',
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.fill  = header_fill
        cell.font  = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')

    for idx, e in enumerate(enrollments, start=1):
        room_bed = (
            f"{e.room_number}/{e.bed_number}" if e.room_number else ''
        )
        ws.append([
            idx,
            e.student.get_full_name(),
            e.student.admission_number,
            e.dormitory.name,
            e.get_boarding_type_display(),
            e.get_status_display(),
            e.enrollment_date.strftime('%Y-%m-%d'),
            room_bed,
            str(e.academic_session),
        ])

    for col in ws.columns:
        max_len = max(
            (len(str(c.value)) for c in col if c.value), default=0
        )
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 2, 50)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = (
        f'attachment; filename="boarding_enrollments_'
        f'{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
    )
    wb.save(response)
    return response


# =============================================================================
# BULK ENROLLMENT — STEP 1: STUDENT SELECTION
# =============================================================================

@login_required
def bulk_enrollment_step1(request):
    """
    Bulk enrollment Step 1 — filter and select students.

    Full page loads and HTMX filter requests are both handled here.
    The student list partial is returned for HTMX requests.

    CHANGE: Gender is now automatically restricted based on the target
    dormitory type BEFORE any form filters are applied:
        BOYS  dormitory → only male students shown (hard constraint)
        GIRLS dormitory → only female students shown (hard constraint)
        MIXED / STAFF   → no restriction

    This means the user-facing dormitory_type filter in the form bar is
    redundant when a target dormitory is pre-selected — the template hides
    it in that case.
    """
    dormitory_id = request.GET.get('dormitory_id')
    session_id   = request.GET.get('session_id')

    target_dormitory = get_object_or_404(Dormitory, pk=dormitory_id) if dormitory_id else None
    target_session   = get_object_or_404(AcademicSession, pk=session_id) if session_id else None

    form = BulkBoardingEnrollmentStudentSelectionForm(
        request.GET,
        academic_session=target_session,
        target_dormitory=target_dormitory,
    )

    students = Student.objects.filter(
        enrollment_status='ACTIVE',
    ).select_related('current_academic_level').order_by('first_name', 'last_name')

    # ── Auto-apply gender filter based on target dormitory type ──────────────
    # Hard constraint applied BEFORE form filters so it can never be bypassed.
    if target_dormitory:
        if target_dormitory.dormitory_type == 'BOYS':
            students = students.filter(gender='M')
        elif target_dormitory.dormitory_type == 'GIRLS':
            students = students.filter(gender='F')
    # MIXED / STAFF → no restriction

    if form.is_valid():
        search                = form.cleaned_data.get('search')
        current_level         = form.cleaned_data.get('current_level')
        current_class         = form.cleaned_data.get('current_class')
        enrollment_status     = form.cleaned_data.get('enrollment_status')
        gender                = form.cleaned_data.get('gender')
        exclude_enrolled      = form.cleaned_data.get('exclude_already_enrolled')
        dormitory_type_filter = form.cleaned_data.get('dormitory_type')
        sort_by               = form.cleaned_data.get('sort_by', 'name')

        if search:
            students = students.filter(
                Q(first_name__icontains=search)     |
                Q(last_name__icontains=search)      |
                Q(middle_name__icontains=search)    |
                Q(admission_number__icontains=search)
            )
        if current_level:
            students = students.filter(current_academic_level=current_level)
        if current_class:
            students = students.filter(current_class=current_class)
        if enrollment_status:
            students = students.filter(enrollment_status=enrollment_status)

        # Only apply the form gender filter when there is no target dormitory —
        # if a dormitory is set the hard constraint above already scopes gender
        # correctly and a conflicting form filter would simply return an empty
        # queryset (e.g. BOYS dormitory + user picks "Female").
        if gender and not target_dormitory:
            students = students.filter(gender=gender)

        # Only apply the form dormitory_type filter when there is no target
        # dormitory — same reason as gender above.
        if dormitory_type_filter and not target_dormitory:
            if dormitory_type_filter == 'BOYS':
                students = students.filter(gender='M')
            elif dormitory_type_filter == 'GIRLS':
                students = students.filter(gender='F')

        if exclude_enrolled and target_session:
            students = students.exclude(
                boarding_enrollments__academic_session=target_session,
                boarding_enrollments__status__in=('PENDING', 'ACTIVE'),
            )

        sort_map = {
            'name':             'first_name',
            '-name':            '-first_name',
            'admission_number': 'admission_number',
            '-admission_date':  '-admission_date',
            'admission_date':   'admission_date',
        }
        students = students.order_by(sort_map.get(sort_by, 'first_name'))

    paginator = Paginator(students, 20)
    page_obj  = paginator.get_page(request.GET.get('page', 1))
    is_htmx   = request.headers.get('HX-Request') == 'true'

    context = {
        'form':             form,
        'students':         page_obj,
        'page_obj':         page_obj,
        'target_dormitory': target_dormitory,
        'target_session':   target_session,
        'total_count':      students.count(),
        'is_htmx':          is_htmx,
        'title':            'Bulk Boarding Enrollment — Select Students',
    }

    if is_htmx:
        return render(
            request,
            'boarding/bulk_enrollment/partials/_student_results.html',
            context,
        )
    return render(request, 'boarding/bulk_enrollment/step1.html', context)


# =============================================================================
# BULK ENROLLMENT — STEP 2: CONFIRMATION AND EXECUTION
# =============================================================================

@login_required
def bulk_enrollment_step2(request):
    """
    Bulk enrollment Step 2 — review details and execute.

    CHANGED: POST now validates through BulkBoardingEnrollmentConfirmationForm
    and delegates execution to BulkBoardingEnrollmentService.  The service
    handles per-student atomic savepoints, invoice generation, and the
    structured result dict (enrolled_count, failed_count, warnings, errors).

    The removed execute_bulk_boarding_enrollment() helper is no longer needed.
    """
    # Student IDs arrive from step 1 via GET (initial render) or are embedded
    # in the hidden field on POST.
    student_ids_from_get = request.GET.get('student_ids', '')
    dormitory_id         = request.GET.get('dormitory_id')
    session_id           = request.GET.get('session_id')

    if request.method == 'POST':
        # Parse student count from the hidden field so the form label is accurate.
        ids_str     = request.POST.get('selected_student_ids', '')
        ids         = [i.strip() for i in ids_str.split(',') if i.strip()]
        form        = BulkBoardingEnrollmentConfirmationForm(
            request.POST, student_count=len(ids)
        )

        if form.is_valid():
            from boarding.services import BulkBoardingEnrollmentService

            service = BulkBoardingEnrollmentService()
            result  = service.enroll_students(
                student_ids=form.cleaned_data['selected_student_ids'],
                academic_session=form.cleaned_data['academic_session'],
                dormitory=form.cleaned_data['dormitory'],
                boarding_type=form.cleaned_data['boarding_type'],
                enrollment_date=form.cleaned_data['enrollment_date'],
                effective_start_date=form.cleaned_data['effective_start_date'],
                effective_end_date=form.cleaned_data.get('effective_end_date'),
                boarding_days=form.cleaned_data.get('boarding_days') or None,
                auto_create_invoice=form.cleaned_data.get('auto_create_invoice', True),
                require_guardian_consent=form.cleaned_data.get(
                    'require_guardian_consent', False
                ),
                reason_for_boarding=form.cleaned_data.get('reason_for_boarding', ''),
                created_by=request.user,
            )

            # Non-blocking warnings (skipped students, invoice failures)
            for warning in result.get('warnings', []):
                messages.warning(request, warning, extra_tags='sweetalert')

            if result['success']:
                enrolled = result['enrolled_count']
                failed   = result['failed_count']
                dormitory = form.cleaned_data['dormitory']

                if failed:
                    for err in result.get('errors', []):
                        messages.warning(request, err, extra_tags='sweetalert')
                    messages.warning(
                        request,
                        f"Enrolled {enrolled} student(s). "
                        f"{failed} could not be enrolled — see details above.",
                        extra_tags='sweetalert',
                    )
                else:
                    messages.success(
                        request,
                        f"Successfully enrolled {enrolled} student(s) into "
                        f"{dormitory.name}.",
                        extra_tags='sweetalert',
                    )

                return redirect(
                    reverse('boarding:enrollment_list') +
                    f'?dormitory={dormitory.pk}'
                )
            else:
                # All students failed pre-flight — no enrollments created.
                for err in result.get('errors', []):
                    messages.error(request, err, extra_tags='sweetalert-error')

        else:
            messages.error(
                request,
                'Please correct the errors below.',
                extra_tags='sweetalert-error',
            )

        # Re-render step 2 on failure.
        ids_str  = request.POST.get('selected_student_ids', '')
        ids      = [i.strip() for i in ids_str.split(',') if i.strip()]
        students = Student.objects.filter(
            pk__in=ids
        ).select_related('current_academic_level')

        return render(request, 'boarding/bulk_enrollment/step2.html', {
            'form':             form,
            'selected_students': students,
            'student_count':    len(ids),
            'title':            'Bulk Boarding Enrollment — Confirm',
        })

    # -------------------------------------------------------------------------
    # GET — initial render of step 2
    # -------------------------------------------------------------------------
    ids     = [i.strip() for i in student_ids_from_get.split(',') if i.strip()]
    if not ids:
        messages.error(
            request,
            'No students selected. Please go back and select students.',
            extra_tags='sweetalert-error',
        )
        return redirect('boarding:bulk_enrollment_step1')

    students = Student.objects.filter(
        pk__in=ids
    ).select_related('current_academic_level')

    initial = {
        'selected_student_ids': student_ids_from_get,
        'enrollment_date':      get_school_today(),
        'effective_start_date': get_school_today(),
        'auto_create_invoice':  True,
    }
    if dormitory_id:
        initial['dormitory'] = get_object_or_404(Dormitory, pk=dormitory_id)
    if session_id:
        initial['academic_session'] = get_object_or_404(AcademicSession, pk=session_id)

    form = BulkBoardingEnrollmentConfirmationForm(
        initial=initial, student_count=len(ids)
    )

    return render(request, 'boarding/bulk_enrollment/step2.html', {
        'form':              form,
        'selected_students': students,
        'student_count':     len(ids),
        'title':             'Bulk Boarding Enrollment — Confirm',
    })


# =============================================================================
# AJAX / JSON UTILITY ENDPOINTS
# =============================================================================

@login_required
def check_dormitory_capacity_ajax(request, pk):
    """JSON: Check whether a dormitory has capacity for N students."""
    try:
        dormitory     = get_object_or_404(Dormitory, pk=pk)
        student_count = int(request.GET.get('student_count', 0))
        available     = dormitory.get_available_capacity()
        can_accommodate = available >= student_count

        return JsonResponse({
            'success':             True,
            'can_accommodate':     can_accommodate,
            'dormitory_name':      dormitory.name,
            'total_capacity':      dormitory.total_capacity,
            'current_occupancy':   dormitory.current_occupancy,
            'available_capacity':  available,
            'requested_count':     student_count,
            'occupancy_percentage': dormitory.get_occupancy_percentage(),
            'message': (
                f"Dormitory can accommodate {student_count} student(s)."
                if can_accommodate
                else f"Insufficient capacity. Only {available} bed(s) available."
            ),
        })
    except (ValueError, TypeError):
        return JsonResponse(
            {'success': False, 'error': 'Invalid student count provided.'}, status=400
        )
    except Exception as exc:
        logger.error("Error checking dormitory capacity: %s", exc)
        return JsonResponse(
            {'success': False, 'error': 'An error occurred while checking capacity.'}, status=500
        )


@login_required
def check_student_boarding_eligibility_ajax(request, student_id):
    """JSON: Check whether a student can be assigned to a dormitory."""
    try:
        student      = get_object_or_404(Student, pk=student_id)
        dormitory_id = request.GET.get('dormitory_id')

        if dormitory_id:
            dormitory = get_object_or_404(Dormitory, pk=dormitory_id)
            can, msg  = dormitory.can_accommodate(student)
            return JsonResponse({
                'success':         True,
                'eligible':        can,
                'student_name':    student.get_full_name(),
                'dormitory_name':  dormitory.name,
                'message':         msg,
            })

        # No specific dormitory — list all compatible ones.
        compatible = []
        for dorm in Dormitory.objects.filter(
            is_active=True, is_available_for_new_admissions=True
        ):
            can, _ = dorm.can_accommodate(student)
            if can:
                compatible.append({
                    'id':                dorm.id,
                    'name':              dorm.name,
                    'type':              dorm.get_dormitory_type_display(),
                    'available_capacity': dorm.get_available_capacity(),
                })

        return JsonResponse({
            'success':                  True,
            'student_name':             student.get_full_name(),
            'compatible_dormitories':   compatible,
        })

    except Exception as exc:
        logger.error("Error checking student eligibility: %s", exc)
        return JsonResponse(
            {'success': False, 'error': 'An error occurred while checking eligibility.'},
            status=500,
        )


@login_required
def boarding_quick_stats_ajax(request):
    """JSON: Key boarding metrics for dashboard widgets."""
    try:
        today = get_school_today()

        dorm_agg = Dormitory.objects.filter(is_active=True).aggregate(
            total_capacity=Sum('total_capacity'),
            total_occupancy=Sum('current_occupancy'),
        )
        cap = dorm_agg['total_capacity']  or 0
        occ = dorm_agg['total_occupancy'] or 0

        return JsonResponse({
            'success': True,
            'dormitories': {
                'total':  Dormitory.objects.filter(is_active=True).count(),
                'boys':   Dormitory.objects.filter(is_active=True, dormitory_type='BOYS').count(),
                'girls':  Dormitory.objects.filter(is_active=True, dormitory_type='GIRLS').count(),
                'full':   Dormitory.objects.filter(
                    is_active=True,
                    current_occupancy__gte=F('total_capacity'),
                ).count(),
            },
            'capacity': {
                'total':               cap,
                'occupied':            occ,
                'available':           cap - occ,
                'occupancy_percentage': round(occ / cap * 100, 1) if cap else 0,
            },
            'enrollments': {
                'total':   BoardingEnrollment.objects.count(),
                'active':  BoardingEnrollment.objects.filter(status='ACTIVE').count(),
                'pending': BoardingEnrollment.objects.filter(status='PENDING').count(),
            },
        })
    except Exception as exc:
        logger.error("Error getting boarding quick stats: %s", exc)
        return JsonResponse(
            {'success': False, 'error': 'An error occurred while fetching statistics.'},
            status=500,
        )


@login_required
def get_student_guardians_api(request, student_id):
    """JSON: Return the active guardians for a student (for dynamic dropdown)."""
    try:
        from students.models import StudentGuardian

        student = get_object_or_404(Student, pk=student_id)

        student_guardians = StudentGuardian.objects.filter(
            student=student,
            is_active=True,
        ).select_related('guardian').order_by(
            '-is_primary', 'guardian__first_name', 'guardian__last_name'
        )

        guardians_data = [
            {
                'id':           str(sg.guardian.id),
                'full_name':    sg.guardian.get_full_name(),
                'first_name':   sg.guardian.first_name,
                'last_name':    sg.guardian.last_name,
                'relationship': sg.relationship or 'Guardian',
                'is_primary':   sg.is_primary,
            }
            for sg in student_guardians
        ]

        return JsonResponse({
            'success':      True,
            'student_id':   str(student.id),
            'student_name': student.get_full_name(),
            'guardians':    guardians_data,
            'count':        len(guardians_data),
        })

    except Student.DoesNotExist:
        return JsonResponse(
            {'success': False, 'error': 'Student not found'}, status=404
        )
    except Exception as exc:
        logger.error(
            "Error fetching guardians for student %s: %s", student_id, exc, exc_info=True
        )
        return JsonResponse(
            {'success': False, 'error': 'An error occurred while fetching guardians'},
            status=500,
        )