"""
exams/views.py

Examination Management Views.

Results flow (2 pages):
  1. results_by_class  — pick session + class
  2. class_marks       — category tabs + read-only grid
     (marks entered per-student via the modal defined in modal_views.py)

Print / Export / Report-card views are at the bottom of this file.
All modal views live in modal_views.py.
"""

from datetime import datetime
from decimal import Decimal

import logging

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.core.exceptions import PermissionDenied
from django.core.paginator import Paginator
from django.db import transaction
from django.db.models import Avg, Count, Max, Min, Q
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse
from django.utils import timezone

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from core.utils import (
    get_active_academic_session,
    get_school_current_time,
    get_school_today,
)
from core.view_helpers import get_print_school_context

from academics.models import AcademicSession, Class, Subject

from students.models import Student

from .forms import (
    ClassGradingSystemFilterForm,
    ClassGradingSystemForm,
    ExamCategoryFilterForm,
    ExamCategoryForm,
    ExaminationFilterForm,
    ExaminationForm,
    GradeLockForm,
    GradeUnlockForm,
    GradingRangeInlineFormSet,
    GradingSystemFilterForm,
    GradingSystemForm,
    ResultPublishForm,
    apply_examination_filters,
)
from .models import (
    ClassGradingSystem,
    ExamCategory,
    Examination,
    GradingSystem,
    StudentExamResult,
)

logger = logging.getLogger(__name__)


# =============================================================================
# DASHBOARD
# =============================================================================

@login_required
def exams_dashboard(request):
    """Main exams dashboard with overview statistics."""
    today           = get_school_today()
    current_session = get_active_academic_session()

    try:
        overview = {
            'total_categories':      ExamCategory.objects.filter(is_active=True).count(),
            'total_grading_systems': GradingSystem.objects.filter(is_active=True).count(),
            'total_examinations':    Examination.objects.count(),
            'upcoming_exams':        Examination.objects.filter(
                exam_date__gte=today, status__in=['PLANNED', 'SCHEDULED']
            ).count(),
            'ongoing_exams':         Examination.objects.filter(status='ONGOING').count(),
            'completed_exams':       Examination.objects.filter(status='COMPLETED').count(),
            'total_results':         StudentExamResult.objects.count(),
            'published_results':     StudentExamResult.objects.filter(is_published=True).count(),
            'locked_grades':         StudentExamResult.objects.filter(is_grade_locked=True).count(),
            'pending_results':       StudentExamResult.objects.filter(
                status='SUBMITTED', is_published=False
            ).count(),
        }

        session_stats = {}
        if current_session:
            session_exams = Examination.objects.filter(academic_session=current_session)
            session_stats = {
                'total_exams': session_exams.count(),
                'completed':   session_exams.filter(status='COMPLETED').count(),
                'ongoing':     session_exams.filter(status='ONGOING').count(),
                'upcoming':    session_exams.filter(
                    exam_date__gte=today, status__in=['PLANNED', 'SCHEDULED']
                ).count(),
            }
    except Exception as e:
        logger.error("Error building dashboard statistics: %s", e)
        overview, session_stats = {}, {}

    recent_examinations = Examination.objects.select_related(
        'subject', 'academic_session', 'exam_category'
    ).order_by('-created_at')[:10]

    upcoming_examinations = Examination.objects.filter(
        exam_date__gte=today, status__in=['PLANNED', 'SCHEDULED']
    ).select_related('subject', 'exam_category').order_by('exam_date', 'start_time')[:10]

    unpublished_results = (
        Examination.objects.filter(status='COMPLETED', results_published=False)
        .annotate(results_count=Count('student_results'))
        .filter(results_count__gt=0)
        .order_by('exam_date')[:10]
    )

    return render(request, 'exams/dashboard.html', {
        'overview':              overview,
        'current_session':       current_session,
        'session_stats':         session_stats,
        'recent_examinations':   recent_examinations,
        'upcoming_examinations': upcoming_examinations,
        'unpublished_results':   unpublished_results,
    })


# =============================================================================
# SHARED FILTER HELPERS
# =============================================================================

def _get_filtered_exam_categories(request):
    qs = ExamCategory.objects.prefetch_related(
        'applicable_levels', 'valid_sessions'
    ).order_by('category_type', 'name')

    q             = request.GET.get('q', '').strip()
    category_type = request.GET.get('category_type', '')
    frequency     = request.GET.get('frequency', '')
    is_active     = request.GET.get('is_active', '')
    curriculum    = request.GET.get('curriculum_compatibility', '')

    if q:
        qs = qs.filter(
            Q(name__icontains=q) | Q(abbreviation__icontains=q) |
            Q(code__icontains=q) | Q(description__icontains=q)
        )
    if category_type: qs = qs.filter(category_type=category_type)
    if frequency:     qs = qs.filter(frequency=frequency)
    if is_active:     qs = qs.filter(is_active=is_active.lower() == 'true')
    if curriculum:    qs = qs.filter(curriculum_compatibility=curriculum)
    return qs


def _get_filtered_grading_systems(request):
    qs = GradingSystem.objects.prefetch_related(
        'applicable_levels', 'applicable_subjects', 'ranges'
    ).annotate(ranges_count=Count('ranges')).order_by('grading_type', 'name')

    q            = request.GET.get('q', '').strip()
    grading_type = request.GET.get('grading_type', '')
    scale_type   = request.GET.get('scale_type', '')
    is_active    = request.GET.get('is_active', '')
    is_default   = request.GET.get('is_default', '')
    uses_gpa     = request.GET.get('uses_gpa', '')
    curriculum   = request.GET.get('curriculum_compatibility', '')

    if q:
        qs = qs.filter(
            Q(name__icontains=q) | Q(code__icontains=q) | Q(description__icontains=q)
        )
    if grading_type: qs = qs.filter(grading_type=grading_type)
    if scale_type:   qs = qs.filter(scale_type=scale_type)
    if is_active:    qs = qs.filter(is_active=is_active.lower() == 'true')
    if is_default:   qs = qs.filter(is_default=is_default.lower() == 'true')
    if uses_gpa:     qs = qs.filter(uses_gpa=uses_gpa.lower() == 'true')
    if curriculum:   qs = qs.filter(curriculum_compatibility=curriculum)
    return qs


def _get_filtered_class_grading_systems(request):
    qs = ClassGradingSystem.objects.select_related(
        'class_instance__academic_level', 'grading_system', 'academic_session', 'subject'
    ).order_by(
        '-academic_session__start_date',
        'class_instance__academic_level__order',
        'priority',
    )

    q                = request.GET.get('q', '').strip()
    class_id         = request.GET.get('class_id', '')
    academic_session = request.GET.get('academic_session', '')
    grading_system   = request.GET.get('grading_system', '')
    subject          = request.GET.get('subject', '')
    is_active        = request.GET.get('is_active', '')

    if q:
        qs = qs.filter(
            Q(class_instance__academic_level__name__icontains=q) |
            Q(grading_system__name__icontains=q) |
            Q(subject__name__icontains=q)
        )
    if class_id:         qs = qs.filter(class_instance_id=class_id)
    if academic_session: qs = qs.filter(academic_session_id=academic_session)
    if grading_system:   qs = qs.filter(grading_system_id=grading_system)
    if subject:          qs = qs.filter(subject_id=subject)
    if is_active:        qs = qs.filter(is_active=is_active.lower() == 'true')
    return qs


def _get_filtered_examinations(request):
    qs = Examination.objects.select_related(
        'subject', 'academic_session', 'exam_category', 'grading_system', 'classroom'
    ).prefetch_related('target_classes', 'invigilators').order_by('-exam_date', 'start_time')

    q                = request.GET.get('q', '').strip()
    academic_session = request.GET.get('academic_session', '')
    exam_category    = request.GET.get('exam_category', '')
    subject          = request.GET.get('subject', '')
    status           = request.GET.get('status', '')
    exam_mode        = request.GET.get('exam_mode', '')
    date_from        = request.GET.get('exam_date_from', '')
    date_to          = request.GET.get('exam_date_to', '')

    if q:
        qs = qs.filter(
            Q(name__icontains=q) | Q(code__icontains=q) |
            Q(description__icontains=q) | Q(subject__name__icontains=q)
        )
    if academic_session: qs = qs.filter(academic_session_id=academic_session)
    if exam_category:    qs = qs.filter(exam_category_id=exam_category)
    if subject:          qs = qs.filter(subject_id=subject)
    if status:           qs = qs.filter(status=status)
    if exam_mode:        qs = qs.filter(exam_mode=exam_mode)
    if date_from:
        try:
            qs = qs.filter(exam_date__gte=datetime.strptime(date_from, '%Y-%m-%d').date())
        except (ValueError, TypeError):
            pass
    if date_to:
        try:
            qs = qs.filter(exam_date__lte=datetime.strptime(date_to, '%Y-%m-%d').date())
        except (ValueError, TypeError):
            pass
    return qs


# =============================================================================
# EXAM CATEGORY VIEWS
# =============================================================================

@login_required
def exam_category_list(request):
    filter_form = ExamCategoryFilterForm(request.GET or None)
    categories  = _get_filtered_exam_categories(request)

    stats = {
        'total':     categories.count(),
        'active':    categories.filter(is_active=True).count(),
        'inactive':  categories.filter(is_active=False).count(),
        'formative': categories.filter(category_type='FORMATIVE').count(),
        'summative': categories.filter(category_type='SUMMATIVE').count(),
        'internal':  categories.filter(category_type='INTERNAL').count(),
        'external':  categories.filter(category_type='EXTERNAL').count(),
    }

    paginator       = Paginator(categories, 20)
    categories_page = paginator.get_page(request.GET.get('page', 1))
    is_htmx         = request.headers.get('HX-Request') == 'true'

    context = {
        'categories_page': categories_page,
        'paginator':       paginator,
        'stats':           stats,
        'filter_form':     filter_form,
        'is_htmx':         is_htmx,
    }
    template = (
        'exams/categories/partials/_category_results.html'
        if is_htmx else 'exams/categories/list.html'
    )
    return render(request, template, context)


@login_required
def exam_category_detail(request, pk):
    category = get_object_or_404(ExamCategory, pk=pk)
    today    = get_school_today()

    stats = {
        'total_exams':     category.examinations.count(),
        'active_exams':    category.examinations.filter(status='ONGOING').count(),
        'completed_exams': category.examinations.filter(status='COMPLETED').count(),
        'upcoming_exams':  category.examinations.filter(
            exam_date__gte=today, status__in=['PLANNED', 'SCHEDULED']
        ).count(),
    }
    return render(request, 'exams/categories/detail.html', {
        'category':         category,
        'stats':            stats,
        'exam_filter_form': ExaminationFilterForm(),
    })


@login_required
def category_examinations_partial(request, pk):
    """HTMX partial: examinations table inside a category detail page."""
    category    = get_object_or_404(ExamCategory, pk=pk)
    filter_form = ExaminationFilterForm(request.GET)

    qs = category.examinations.select_related(
        'subject', 'academic_session'
    ).order_by('-exam_date')
    qs = apply_examination_filters(qs, filter_form)

    paginator = Paginator(qs, 10)
    page      = paginator.get_page(request.GET.get('page', 1))

    return render(request, 'exams/categories/partials/_exam_results.html', {
        'examinations': page,
        'category':     category,
    })


@login_required
def exam_category_create(request):
    if request.method == 'POST':
        form = ExamCategoryForm(request.POST)
        if form.is_valid():
            try:
                category = form.save()
                messages.success(request, f'Exam category "{category.name}" created successfully.')
                return redirect('exams:category_detail', pk=category.pk)
            except Exception as e:
                logger.error("Error creating exam category: %s", e)
                messages.error(request, f'Error creating exam category: {e}')
        else:
            messages.error(request, 'Please correct the errors below.')
    else:
        form = ExamCategoryForm()

    return render(request, 'exams/categories/form.html', {
        'form':        form,
        'title':       'Create Exam Category',
        'submit_text': 'Create Category',
    })


@login_required
def exam_category_edit(request, pk):
    category = get_object_or_404(ExamCategory, pk=pk)
    today    = get_school_today()

    stats = {
        'total_exams':     category.examinations.count(),
        'active_exams':    category.examinations.filter(status='ONGOING').count(),
        'completed_exams': category.examinations.filter(status='COMPLETED').count(),
        'upcoming_exams':  category.examinations.filter(
            exam_date__gte=today, status__in=['PLANNED', 'SCHEDULED']
        ).count(),
        'planned_exams':   category.examinations.filter(status='PLANNED').count(),
        'scheduled_exams': category.examinations.filter(status='SCHEDULED').count(),
    }

    if request.method == 'POST':
        form = ExamCategoryForm(request.POST, instance=category)
        if form.is_valid():
            try:
                category = form.save()
                messages.success(request, f'Exam category "{category.name}" updated successfully.')
                return redirect('exams:category_detail', pk=category.pk)
            except Exception as e:
                logger.error("Error updating exam category: %s", e)
                messages.error(request, f'Error updating exam category: {e}')
        else:
            messages.error(request, 'Please correct the errors below.')
    else:
        form = ExamCategoryForm(instance=category)

    return render(request, 'exams/categories/form.html', {
        'form':        form,
        'category':    category,
        'title':       'Edit Exam Category',
        'submit_text': 'Update Category',
        'stats':       stats,
    })


@login_required
def exam_category_delete(request, pk):
    category = get_object_or_404(ExamCategory, pk=pk)
    is_htmx  = request.headers.get('HX-Request') == 'true'

    if request.method != 'POST':
        return redirect('exams:category_detail', pk=pk)

    if category.examinations.exists():
        msg = 'Cannot delete a category that has existing examinations.'
        if is_htmx:
            return _htmx_alert('error', msg)
        messages.error(request, msg)
        return redirect('exams:category_detail', pk=pk)

    try:
        name = category.name
        category.delete()
        if is_htmx:
            return _htmx_redirect(reverse('exams:category_list'), 'success',
                                   f'Category "{name}" deleted.')
        messages.success(request, f'Exam category "{name}" deleted successfully.')
        return redirect('exams:category_list')
    except Exception as e:
        logger.error("Error deleting exam category: %s", e)
        if is_htmx:
            return _htmx_alert('error', str(e))
        messages.error(request, str(e))
        return redirect('exams:category_detail', pk=pk)


@login_required
def exam_category_toggle_active(request, pk):
    if request.method != 'POST':
        return redirect('exams:category_detail', pk=pk)
    category = get_object_or_404(ExamCategory, pk=pk)
    try:
        category.is_active = not category.is_active
        category.save()
        status = 'activated' if category.is_active else 'deactivated'
        messages.success(request, f'Category "{category.name}" {status}.')
    except Exception as e:
        logger.error("Error toggling exam category: %s", e)
        messages.error(request, str(e))
    return redirect('exams:category_detail', pk=pk)


# =============================================================================
# GRADING SYSTEM VIEWS
# =============================================================================

@login_required
def grading_system_list(request):
    filter_form = GradingSystemFilterForm(request.GET or None)
    systems     = _get_filtered_grading_systems(request)

    stats = {
        'total':        systems.count(),
        'active':       systems.filter(is_active=True).count(),
        'default':      systems.filter(is_default=True).count(),
        'with_gpa':     systems.filter(uses_gpa=True).count(),
        'letter_grade': systems.filter(grading_type='LETTER').count(),
        'numerical':    systems.filter(grading_type='NUMERICAL').count(),
    }

    paginator    = Paginator(systems, 20)
    systems_page = paginator.get_page(request.GET.get('page', 1))
    is_htmx      = request.headers.get('HX-Request') == 'true'

    context = {
        'systems_page': systems_page,
        'paginator':    paginator,
        'stats':        stats,
        'filter_form':  filter_form,
        'is_htmx':      is_htmx,
    }
    template = (
        'exams/grading_systems/partials/_system_results.html'
        if is_htmx else 'exams/grading_systems/list.html'
    )
    return render(request, template, context)


@login_required
def grading_system_detail(request, pk):
    system = get_object_or_404(GradingSystem, pk=pk)
    ranges = system.ranges.all().order_by('-min_score')

    class_assignments = system.class_assignments.select_related(
        'class_instance__academic_level', 'academic_session'
    ).filter(is_active=True).order_by('-academic_session__start_date')[:20]

    examinations = system.examinations.select_related(
        'subject', 'academic_session'
    ).order_by('-exam_date')[:20]

    stats = {
        'total_ranges':      ranges.count(),
        'class_assignments': system.class_assignments.filter(is_active=True).count(),
        'examinations':      system.examinations.count(),
        'passing_ranges':    ranges.filter(is_passing_grade=True).count(),
        'failing_ranges':    ranges.filter(is_passing_grade=False).count(),
    }

    return render(request, 'exams/grading_systems/detail.html', {
        'system':           system,
        'ranges':           ranges,
        'class_assignments':class_assignments,
        'examinations':     examinations,
        'stats':            stats,
        'coverage_status':  _check_grading_system_coverage(system, ranges),
    })


@login_required
def grading_system_create(request):
    if request.method == 'POST':
        form    = GradingSystemForm(request.POST)
        formset = GradingRangeInlineFormSet(request.POST)
        if form.is_valid() and formset.is_valid():
            try:
                with transaction.atomic():
                    system           = form.save()
                    formset.instance = system
                    formset.save()
                messages.success(request, f'Grading system "{system.name}" created successfully.')
                return redirect('exams:grading_system_detail', pk=system.pk)
            except Exception as e:
                logger.error("Error creating grading system: %s", e, exc_info=True)
                messages.error(request, f'Error creating grading system: {e}')
        else:
            if form.errors:
                messages.error(request, 'Please correct the errors in the grading system details.')
            if formset.errors or formset.non_form_errors():
                messages.error(request, 'Please correct the errors in the grade ranges.')
    else:
        form    = GradingSystemForm()
        formset = GradingRangeInlineFormSet()

    return render(request, 'exams/grading_systems/form.html', {
        'form':        form,
        'formset':     formset,
        'title':       'Create Grading System',
        'submit_text': 'Create Grading System',
        'is_edit':     False,
    })


@login_required
def grading_system_edit(request, pk):
    system = get_object_or_404(GradingSystem, pk=pk)
    today  = get_school_today()

    locked_count = StudentExamResult.objects.filter(
        examination__grading_system=system, is_grade_locked=True
    ).count()

    stats = {
        'total_ranges':        system.ranges.count(),
        'class_assignments':   system.class_assignments.filter(is_active=True).count(),
        'examinations':        system.examinations.count(),
        'active_assignments':  system.class_assignments.filter(
            is_active=True, effective_date__lte=today
        ).filter(Q(end_date__isnull=True) | Q(end_date__gte=today)).count(),
        'locked_grades_count': locked_count,
    }

    if request.method == 'POST':
        form    = GradingSystemForm(request.POST, instance=system)
        formset = GradingRangeInlineFormSet(request.POST, instance=system)
        if form.is_valid() and formset.is_valid():
            try:
                with transaction.atomic():
                    system = form.save()
                    formset.save()
                if locked_count:
                    messages.warning(
                        request,
                        f'{locked_count} locked grade(s) were not recalculated — '
                        'unlock them individually if a regrade is needed.',
                    )
                messages.success(request, f'Grading system "{system.name}" updated successfully.')
                return redirect('exams:grading_system_detail', pk=system.pk)
            except Exception as e:
                logger.error("Error updating grading system: %s", e, exc_info=True)
                messages.error(request, f'Error updating grading system: {e}')
        else:
            if form.errors:
                messages.error(request, 'Please correct the errors in the grading system details.')
            if formset.errors or formset.non_form_errors():
                messages.error(request, 'Please correct the errors in the grade ranges.')
    else:
        form    = GradingSystemForm(instance=system)
        formset = GradingRangeInlineFormSet(instance=system)

    return render(request, 'exams/grading_systems/form.html', {
        'form':              form,
        'formset':           formset,
        'system':            system,
        'title':             f'Edit {system.name}',
        'submit_text':       'Update Grading System',
        'is_edit':           True,
        'stats':             stats,
        'has_locked_grades': locked_count > 0,
        'has_examinations':  stats['examinations'] > 0,
    })


@login_required
def grading_system_delete(request, pk):
    system  = get_object_or_404(GradingSystem, pk=pk)
    is_htmx = request.headers.get('HX-Request') == 'true'

    if request.method != 'POST':
        return redirect('exams:grading_system_detail', pk=pk)

    if system.is_default:
        msg = 'Cannot delete the default grading system.'
        if is_htmx:
            return _htmx_alert('error', msg)
        messages.error(request, msg)
        return redirect('exams:grading_system_detail', pk=pk)

    if system.class_assignments.exists() or system.examinations.exists():
        msg = 'Cannot delete a grading system that has class assignments or examinations.'
        if is_htmx:
            return _htmx_alert('error', msg)
        messages.error(request, msg)
        return redirect('exams:grading_system_detail', pk=pk)

    try:
        name = system.name
        system.delete()
        if is_htmx:
            return _htmx_redirect(
                reverse('exams:grading_system_list'), 'success', f'System "{name}" deleted.'
            )
        messages.success(request, f'Grading system "{name}" deleted successfully.')
        return redirect('exams:grading_system_list')
    except Exception as e:
        logger.error("Error deleting grading system: %s", e, exc_info=True)
        if is_htmx:
            return _htmx_alert('error', str(e))
        messages.error(request, str(e))
        return redirect('exams:grading_system_detail', pk=pk)


@login_required
def grading_system_toggle_active(request, pk):
    if request.method != 'POST':
        return redirect('exams:grading_system_detail', pk=pk)
    system = get_object_or_404(GradingSystem, pk=pk)
    try:
        system.is_active = not system.is_active
        system.save()
        status = 'activated' if system.is_active else 'deactivated'
        messages.success(request, f'Grading system "{system.name}" {status}.')
    except Exception as e:
        logger.error("Error toggling grading system: %s", e, exc_info=True)
        messages.error(request, str(e))
    return redirect('exams:grading_system_detail', pk=pk)


@login_required
def grading_system_set_default(request, pk):
    if request.method != 'POST':
        return redirect('exams:grading_system_detail', pk=pk)
    system = get_object_or_404(GradingSystem, pk=pk)
    try:
        with transaction.atomic():
            GradingSystem.objects.filter(is_default=True).update(is_default=False)
            system.is_default = True
            system.is_active  = True
            system.save()
        messages.success(request, f'"{system.name}" is now the default grading system.')
    except Exception as e:
        logger.error("Error setting default grading system: %s", e, exc_info=True)
        messages.error(request, str(e))
    return redirect('exams:grading_system_detail', pk=pk)


# =============================================================================
# GRADING SYSTEM PRIVATE HELPER
# =============================================================================

def _check_grading_system_coverage(system, ranges):
    """Return a coverage-status dict for a grading system's ranges."""
    if not ranges.exists():
        return {
            'has_coverage': False, 'has_gaps': True, 'gaps': [],
            'message': 'No grade ranges defined.',
        }

    sorted_ranges = list(ranges.order_by('min_score'))
    gaps = []

    if sorted_ranges[0].min_score > system.minimum_score:
        gaps.append({
            'start':   system.minimum_score,
            'end':     sorted_ranges[0].min_score,
            'message': f'Gap from {system.minimum_score} to {sorted_ranges[0].min_score}',
        })

    for i in range(len(sorted_ranges) - 1):
        cur = sorted_ranges[i]
        nxt = sorted_ranges[i + 1]
        if nxt.min_score - cur.max_score > Decimal('0.01'):
            gaps.append({
                'start':   cur.max_score,
                'end':     nxt.min_score,
                'message': f'Gap from {cur.max_score} to {nxt.min_score}',
            })

    if sorted_ranges[-1].max_score < system.maximum_score:
        gaps.append({
            'start':   sorted_ranges[-1].max_score,
            'end':     system.maximum_score,
            'message': f'Gap from {sorted_ranges[-1].max_score} to {system.maximum_score}',
        })

    return {
        'has_coverage': not gaps,
        'has_gaps':     bool(gaps),
        'gaps':         gaps,
        'message':      'Complete coverage' if not gaps else f'{len(gaps)} gap(s) found',
    }


# =============================================================================
# CLASS GRADING SYSTEM ASSIGNMENT VIEWS
# =============================================================================

@login_required
def class_grading_system_list(request):
    filter_form  = ClassGradingSystemFilterForm(request.GET or None)
    assignments  = _get_filtered_class_grading_systems(request)
    current_sess = get_active_academic_session()

    stats = {
        'total':           assignments.count(),
        'active':          assignments.filter(is_active=True).count(),
        'inactive':        assignments.filter(is_active=False).count(),
        'current_session': assignments.filter(
            academic_session=current_sess, is_active=True
        ).count() if current_sess else 0,
    }

    paginator        = Paginator(assignments, 20)
    assignments_page = paginator.get_page(request.GET.get('page', 1))
    is_htmx          = request.headers.get('HX-Request') == 'true'

    context = {
        'assignments_page': assignments_page,
        'paginator':        paginator,
        'stats':            stats,
        'filter_form':      filter_form,
        'is_htmx':          is_htmx,
    }
    template = (
        'exams/class_grading_systems/partials/_assignment_results.html'
        if is_htmx else 'exams/class_grading_systems/list.html'
    )
    return render(request, template, context)


@login_required
def class_grading_system_detail(request, pk):
    assignment = get_object_or_404(
        ClassGradingSystem.objects.select_related(
            'class_instance__academic_level', 'grading_system',
            'academic_session', 'subject', 'assigned_by',
        ),
        pk=pk,
    )
    return render(request, 'exams/class_grading_systems/detail.html', {
        'assignment':          assignment,
        'grading_ranges':      assignment.grading_system.ranges.all().order_by('-min_score'),
        'is_currently_active': assignment.is_currently_active(),
    })


@login_required
def class_grading_system_create(request, class_pk=None):
    initial = {}
    if class_pk:
        cls     = get_object_or_404(Class, pk=class_pk)
        initial = {'class_instance': cls, 'academic_session': cls.academic_session}

    if request.method == 'POST':
        form = ClassGradingSystemForm(request.POST)
        if form.is_valid():
            try:
                with transaction.atomic():
                    assignment             = form.save(commit=False)
                    assignment.assigned_by = request.user
                    assignment.save()
                    form.save_m2m()
                messages.success(request, 'Grading system assignment created successfully.')
                return redirect('exams:class_grading_system_detail', pk=assignment.pk)
            except Exception as e:
                logger.error("Error creating class grading system assignment: %s", e, exc_info=True)
                messages.error(request, f'Error creating assignment: {e}')
        else:
            messages.error(request, 'Please correct the errors below.')
    else:
        form = ClassGradingSystemForm(initial=initial)

    return render(request, 'exams/class_grading_systems/form.html', {
        'form':  form,
        'title': 'Assign Grading System to Class',
    })


@login_required
def class_grading_system_edit(request, pk):
    assignment = get_object_or_404(ClassGradingSystem, pk=pk)

    if request.method == 'POST':
        form = ClassGradingSystemForm(request.POST, instance=assignment)
        if form.is_valid():
            try:
                assignment = form.save()
                messages.success(request, 'Grading system assignment updated successfully.')
                return redirect('exams:class_grading_system_detail', pk=assignment.pk)
            except Exception as e:
                logger.error("Error updating class grading system assignment: %s", e, exc_info=True)
                messages.error(request, f'Error updating assignment: {e}')
        else:
            messages.error(request, 'Please correct the errors below.')
    else:
        form = ClassGradingSystemForm(instance=assignment)

    return render(request, 'exams/class_grading_systems/form.html', {
        'form':       form,
        'assignment': assignment,
        'title':      'Edit Grading System Assignment',
    })


@login_required
def class_grading_system_delete(request, pk):
    assignment = get_object_or_404(ClassGradingSystem, pk=pk)
    is_htmx    = request.headers.get('HX-Request') == 'true'

    if request.method != 'POST':
        return redirect('exams:class_grading_system_detail', pk=pk)

    try:
        assignment.delete()
        if is_htmx:
            return _htmx_redirect(
                reverse('exams:class_grading_system_list'),
                'success', 'Assignment deleted successfully.',
            )
        messages.success(request, 'Grading system assignment deleted successfully.')
        return redirect('exams:class_grading_system_list')
    except Exception as e:
        logger.error("Error deleting class grading system assignment: %s", e, exc_info=True)
        if is_htmx:
            return _htmx_alert('error', str(e))
        messages.error(request, str(e))
        return redirect('exams:class_grading_system_detail', pk=pk)


@login_required
def class_grading_system_toggle_active(request, pk):
    if request.method != 'POST':
        return redirect('exams:class_grading_system_detail', pk=pk)
    assignment = get_object_or_404(ClassGradingSystem, pk=pk)
    try:
        assignment.is_active = not assignment.is_active
        assignment.save()
        status = 'activated' if assignment.is_active else 'deactivated'
        messages.success(request, f'Assignment {status}.')
    except Exception as e:
        logger.error("Error toggling class grading system assignment: %s", e, exc_info=True)
        messages.error(request, str(e))
    return redirect('exams:class_grading_system_detail', pk=pk)


@login_required
def bulk_class_grading_system_assign(request):
    if request.method == 'POST':
        try:
            grading_system   = get_object_or_404(
                GradingSystem, pk=request.POST.get('grading_system')
            )
            academic_session = get_object_or_404(
                AcademicSession, pk=request.POST.get('academic_session')
            )
            class_ids  = request.POST.getlist('classes')
            subject_id = request.POST.get('subject')
            subject    = get_object_or_404(Subject, pk=subject_id) if subject_id else None

            created = skipped = 0
            with transaction.atomic():
                for class_id in class_ids:
                    cls = get_object_or_404(Class, pk=class_id)
                    _, was_created = ClassGradingSystem.objects.get_or_create(
                        class_instance   = cls,
                        grading_system   = grading_system,
                        academic_session = academic_session,
                        subject          = subject,
                        defaults={
                            'assigned_by':    request.user,
                            'effective_date': get_school_today(),
                        },
                    )
                    if was_created:
                        created += 1
                    else:
                        skipped += 1

            if created: messages.success(request, f'Assigned grading system to {created} class(es).')
            if skipped: messages.info(request,    f'Skipped {skipped} class(es) — assignment already exists.')
            return redirect('exams:class_grading_system_list')
        except Exception as e:
            logger.error("Error in bulk grading system assignment: %s", e, exc_info=True)
            messages.error(request, str(e))
            return redirect('exams:class_grading_system_list')

    return render(request, 'exams/class_grading_systems/bulk_assign.html', {
        'grading_systems': GradingSystem.objects.filter(is_active=True).order_by('name'),
        'sessions':        AcademicSession.objects.filter(is_active=True).order_by('-start_date'),
        'classes':         Class.objects.filter(is_active=True)
                               .select_related('academic_level')
                               .order_by('academic_level__order', 'section'),
        'subjects':        Subject.objects.filter(is_active=True).order_by('name'),
        'title':           'Bulk Assign Grading System',
    })


@login_required
def class_grading_system_print_list(request):
    return render(request, 'exams/class_grading_systems/print.html', {
        'assignments': _get_filtered_class_grading_systems(request),
        'now':         timezone.now(),
        **get_print_school_context(request),
    })


# =============================================================================
# EXAMINATION VIEWS
# =============================================================================

@login_required
def examination_list(request):
    filter_form  = ExaminationFilterForm(request.GET or None)
    examinations = _get_filtered_examinations(request)
    today        = get_school_today()

    stats = {
        'total':             examinations.count(),
        'planned':           examinations.filter(status='PLANNED').count(),
        'scheduled':         examinations.filter(status='SCHEDULED').count(),
        'ongoing':           examinations.filter(status='ONGOING').count(),
        'completed':         examinations.filter(status='COMPLETED').count(),
        'upcoming':          examinations.filter(
            exam_date__gte=today, status__in=['PLANNED', 'SCHEDULED']
        ).count(),
        'results_published': examinations.filter(results_published=True).count(),
    }

    paginator         = Paginator(examinations, 20)
    examinations_page = paginator.get_page(request.GET.get('page', 1))
    is_htmx           = request.headers.get('HX-Request') == 'true'

    context = {
        'examinations_page': examinations_page,
        'paginator':         paginator,
        'stats':             stats,
        'filter_form':       filter_form,
        'is_htmx':           is_htmx,
    }
    template = (
        'exams/examinations/partials/_examination_results.html'
        if is_htmx else 'exams/examinations/list.html'
    )
    return render(request, template, context)


@login_required
def examination_detail(request, pk):
    """Full detail page for a single examination."""
    examination = get_object_or_404(
        Examination.objects.select_related(
            'subject', 'academic_session', 'exam_category',
            'grading_system', 'classroom',
        ).prefetch_related('target_classes', 'invigilators'),
        pk=pk,
    )

    results = examination.student_results.select_related('student').order_by(
        'student__last_name', 'student__first_name'
    )

    agg = results.filter(
        status__in=['COMPLETED', 'SUBMITTED'], score__isnull=False
    ).aggregate(
        average=Avg('score'),
        highest=Max('score'),
        lowest=Min('score'),
        total=Count('id'),
        pass_count=Count('id', filter=Q(is_pass=True)),
    )
    total = agg['total'] or 0

    stats = {
        'total_results':    results.count(),
        'completed':        results.filter(status__in=['COMPLETED', 'SUBMITTED']).count(),
        'absent':           results.filter(status='ABSENT').count(),
        'published':        results.filter(is_published=True).count(),
        'locked':           results.filter(is_grade_locked=True).count(),
        'average_score':    round(float(agg['average']), 2) if agg['average'] else None,
        'highest_score':    agg['highest'],
        'lowest_score':     agg['lowest'],
        'pass_count':       agg['pass_count'] or 0,
        'pass_rate':        round((agg['pass_count'] or 0) / total * 100, 2) if total else 0,
    }

    grading_system = examination.get_effective_grading_system()

    return render(request, 'exams/examinations/detail.html', {
        'examination':    examination,
        'results':        results,
        'stats':          stats,
        'grading_system': grading_system,
        'publish_form':   ResultPublishForm(),
    })


@login_required
def examination_create(request):
    if request.method == 'POST':
        form = ExaminationForm(request.POST)
        if form.is_valid():
            try:
                examination = form.save()
                messages.success(request, f'Examination "{examination.name}" created successfully.')
                return redirect('exams:examination_detail', pk=examination.pk)
            except Exception as e:
                logger.error("Error creating examination: %s", e)
                messages.error(request, f'Error creating examination: {e}')
        else:
            messages.error(request, 'Please correct the errors below.')
    else:
        form = ExaminationForm()

    return render(request, 'exams/examinations/form.html', {
        'form':  form,
        'title': 'Create Examination',
    })


@login_required
def examination_edit(request, pk):
    examination = get_object_or_404(Examination, pk=pk)

    if request.method == 'POST':
        form = ExaminationForm(request.POST, instance=examination)
        if form.is_valid():
            try:
                examination = form.save()
                messages.success(request, f'Examination "{examination.name}" updated successfully.')
                return redirect('exams:examination_detail', pk=examination.pk)
            except Exception as e:
                logger.error("Error updating examination: %s", e)
                messages.error(request, f'Error updating examination: {e}')
        else:
            messages.error(request, 'Please correct the errors below.')
    else:
        form = ExaminationForm(instance=examination)

    return render(request, 'exams/examinations/form.html', {
        'form':        form,
        'examination': examination,
        'title':       'Edit Examination',
    })


@login_required
def examination_delete(request, pk):
    examination = get_object_or_404(Examination, pk=pk)
    is_htmx     = request.headers.get('HX-Request') == 'true'

    if request.method != 'POST':
        return redirect('exams:examination_detail', pk=pk)

    if examination.student_results.exists():
        msg = 'Cannot delete an examination that already has results.'
        if is_htmx:
            return _htmx_alert('error', msg)
        messages.error(request, msg)
        return redirect('exams:examination_detail', pk=pk)

    if examination.status in ('ONGOING', 'COMPLETED'):
        msg = 'Cannot delete an ongoing or completed examination.'
        if is_htmx:
            return _htmx_alert('error', msg)
        messages.error(request, msg)
        return redirect('exams:examination_detail', pk=pk)

    try:
        name = examination.name
        examination.delete()
        if is_htmx:
            return _htmx_redirect(
                reverse('exams:examination_list'), 'success',
                f'Examination "{name}" deleted.',
            )
        messages.success(request, f'Examination "{name}" deleted successfully.')
        return redirect('exams:examination_list')
    except Exception as e:
        logger.error("Error deleting examination: %s", e)
        if is_htmx:
            return _htmx_alert('error', str(e))
        messages.error(request, str(e))
        return redirect('exams:examination_detail', pk=pk)


@login_required
def examination_update_status(request, pk):
    if request.method != 'POST':
        return redirect('exams:examination_detail', pk=pk)
    examination = get_object_or_404(Examination, pk=pk)
    new_status  = request.POST.get('status')

    if new_status not in dict(Examination.EXAM_STATUS_CHOICES):
        messages.error(request, 'Invalid status.')
    else:
        try:
            examination.status = new_status
            examination.save()
            messages.success(request, f'Status updated to {examination.get_status_display()}.')
        except Exception as e:
            logger.error("Error updating examination status: %s", e)
            messages.error(request, str(e))
    return redirect('exams:examination_detail', pk=pk)


@login_required
def publish_results(request, pk):
    """Publish all completed results for an examination."""
    examination = get_object_or_404(Examination, pk=pk)
    is_htmx     = request.headers.get('HX-Request') == 'true'

    if request.method == 'POST':
        form = ResultPublishForm(request.POST)
        if form.is_valid():
            auto_lock = form.cleaned_data['auto_lock_grades']
            try:
                with transaction.atomic():
                    now = get_school_current_time()

                    results = examination.student_results.filter(
                        status__in=['COMPLETED', 'SUBMITTED']
                    )
                    for result in results:
                        result.is_published     = True
                        result.publication_date = now
                        result.save()
                        if auto_lock and not result.is_grade_locked:
                            result.lock_grade(
                                locked_by=request.user,
                                reason='Auto-locked during result publication',
                            )

                    examination.results_published        = True
                    examination.results_publication_date = now
                    examination.save()

                count  = results.count()
                locked = results.filter(is_grade_locked=True).count() if auto_lock else 0
                msg    = f'Published {count} result(s).'
                if auto_lock:
                    msg += f' {locked} grade(s) auto-locked.'

                if is_htmx:
                    return _htmx_redirect(
                        reverse('exams:examination_detail', kwargs={'pk': pk}),
                        'success', msg,
                    )
                messages.success(request, msg)
                return redirect('exams:examination_detail', pk=pk)
            except Exception as e:
                logger.error("Error publishing results: %s", e)
                if is_htmx:
                    return _htmx_alert('error', str(e))
                messages.error(request, str(e))
                return redirect('exams:examination_detail', pk=pk)

    return render(request, 'exams/examinations/publish_form.html', {
        'examination': examination,
        'form':        ResultPublishForm(),
    })


@login_required
def unpublish_results(request, pk):
    if request.method != 'POST':
        return redirect('exams:examination_detail', pk=pk)
    examination = get_object_or_404(Examination, pk=pk)
    try:
        with transaction.atomic():
            examination.results_published        = False
            examination.results_publication_date = None
            examination.save()
            examination.student_results.all().update(
                is_published=False,
                publication_date=None,
            )
        messages.success(request, 'Results unpublished successfully.')
    except Exception as e:
        logger.error("Error unpublishing results: %s", e)
        messages.error(request, str(e))
    return redirect('exams:examination_detail', pk=pk)


# =============================================================================
# RESULTS — PAGE 1: SELECT SESSION + CLASS
# =============================================================================

@login_required
def results_by_class(request):
    """
    Landing page for results.
    Shows a session dropdown and class cards for that session.

    URL:  /exams/results/
          /exams/results/?session=<pk>
    """
    session_id = request.GET.get('session')
    if session_id:
        session = get_object_or_404(AcademicSession, pk=session_id)
    else:
        session = AcademicSession.get_current_session()
        if not session:
            session = AcademicSession.objects.filter(is_active=True).order_by('-start_date').first()

    if not session:
        messages.warning(request, 'No academic session found. Please create one first.')
        return redirect('academics:session_create')

    classes = (
        Class.objects
        .filter(is_active=True, academic_session=session)
        .select_related('academic_level', 'class_teacher')
        .annotate(
            # Count ALL enrollments — no is_active/completion_status filter
            # because past-term enrollments are COMPLETED (is_active=False)
            session_student_count=Count('enrollments')
        )
        .order_by('academic_level__order', 'section')
    )

    all_sessions = AcademicSession.objects.order_by('-start_date')

    return render(request, 'exams/results/class_selector.html', {
        'classes':      classes,
        'session':      session,
        'all_sessions': all_sessions,
    })


# =============================================================================
# RESULTS — PAGE 2: CLASS MARKS (category tabs + read-only grid)
# =============================================================================

@login_required
def class_marks(request, class_pk):
    """
    Page 2 of the results flow.

    Category tabs across the top; clicking a tab swaps the student × subject
    grid via HTMX. Each cell opens the per-student score entry modal
    (modal_views.py).

    URL:  /exams/results/<class_pk>/
          /exams/results/<class_pk>/?session=<pk>&tab=<abbr>
    """
    class_instance = get_object_or_404(Class, pk=class_pk)

    session_id = request.GET.get('session')
    session    = (
        get_object_or_404(AcademicSession, pk=session_id)
        if session_id
        else class_instance.academic_session or AcademicSession.get_current_session()
    )
    if not session:
        messages.warning(request, 'No active session found.')
        return redirect('exams:results_by_class')

    # Cross-session class resolution: each term creates new Class rows.
    if class_instance.academic_session_id != session.pk:
        equivalent = Class.objects.filter(
            academic_level   = class_instance.academic_level,
            section          = class_instance.section,
            academic_session = session,
        ).first()
        if not equivalent:
            messages.warning(
                request,
                f'No class found for {class_instance.academic_level.name} '
                f'in {session}. Classes may not have been created for that term yet.',
            )
            return redirect(reverse('exams:results_by_class') + f'?session={session.pk}')
        class_instance = equivalent

    all_sessions = AcademicSession.objects.filter(is_active=True).order_by('-start_date')

    # Students currently enrolled in this class + session
    today             = get_school_today()
    enrollment_filter = dict(
        class_enrollments__class_instance=class_instance,
        class_enrollments__academic_session=session,
    )
    if session.end_date >= today:
        enrollment_filter['class_enrollments__is_active']        = True
        enrollment_filter['class_enrollments__completion_status'] = 'ONGOING'

    students = Student.objects.filter(
        **enrollment_filter
    ).distinct().order_by('first_name', 'last_name')

    # Examinations grouped by exam category
    examinations = Examination.objects.filter(
        target_classes=class_instance,
        academic_session=session,
    ).select_related('exam_category', 'subject').order_by('subject__name')

    exams_by_category: dict = {}
    for exam in examinations:
        abbr = exam.exam_category.abbreviation
        if abbr not in exams_by_category:
            exams_by_category[abbr] = {'category': exam.exam_category, 'exams': []}
        exams_by_category[abbr]['exams'].append(exam)

    selected_abbr = request.GET.get('tab')
    if not selected_abbr and exams_by_category:
        selected_abbr = next(iter(exams_by_category))

    # Build read-only grid for the selected tab
    grid_rows: list      = []
    category_subjects: list = []
    selected_category    = None

    if selected_abbr and selected_abbr in exams_by_category:
        cat_data          = exams_by_category[selected_abbr]
        selected_category = cat_data['category']
        category_exams    = cat_data['exams']
        category_subjects = [e.subject.name for e in category_exams]

        # Single query — avoids N+1
        result_map = {
            (r.student_id, r.examination_id): r
            for r in StudentExamResult.objects.filter(
                examination__in=category_exams,
                student__in=students,
            ).select_related('examination')
        }

        for student in students:
            cells  = []
            scores = []
            for exam in category_exams:
                result = result_map.get((student.pk, exam.pk))
                cells.append({
                    'exam':         exam,
                    'result':       result,
                    'score':        result.score if result else None,
                    'grade':        result.grade if result else '',
                    'is_locked':    result.is_grade_locked if result else False,
                    'is_published': result.is_published if result else False,
                })
                if result and result.score is not None:
                    scores.append(result.score)

            grid_rows.append({
                'student':        student,
                'cells':          cells,
                'total':          sum(scores),
                'average':        round(sum(scores) / len(scores), 2) if scores else 0,
                'subjects_taken': len(scores),
            })

        # Rank by total descending; ties share a position
        grid_rows.sort(key=lambda r: r['total'], reverse=True)
        position   = 1
        prev_total = None
        for i, row in enumerate(grid_rows):
            if row['total'] != prev_total:
                position = i + 1
            row['position'] = position
            prev_total      = row['total']

    is_htmx = request.headers.get('HX-Request') == 'true'
    context = {
        'class_instance':    class_instance,
        'session':           session,
        'all_sessions':      all_sessions,
        'exams_by_category': exams_by_category,
        'selected_abbr':     selected_abbr,
        'selected_category': selected_category,
        'category_subjects': category_subjects,
        'grid_rows':         grid_rows,
        'student_count':     students.count(),
    }
    template = (
        'exams/results/partials/_marks_grid.html'
        if is_htmx else 'exams/results/class_marks.html'
    )
    return render(request, template, context)


# =============================================================================
# INDIVIDUAL RESULT VIEWS
# =============================================================================

@login_required
def result_detail(request, pk):
    """Single result detail — used for the report-card link from the grid."""
    result = get_object_or_404(
        StudentExamResult.objects.select_related(
            'student',
            'examination__subject',
            'examination__academic_session',
            'examination__exam_category',
            'verified_by',
            'moderator',
            'grade_locked_by',
        ),
        pk=pk,
    )
    return render(request, 'exams/results/detail.html', {
        'result':              result,
        'grade_history':       result.get_grade_history() if result.is_grade_locked else None,
        'performance_summary': result.get_performance_summary(),
        'grading_system':      result.examination.get_effective_grading_system(),
    })


@login_required
def lock_grade(request, pk):
    """Lock a single result's grade (POST only)."""
    result  = get_object_or_404(StudentExamResult, pk=pk)
    is_htmx = request.headers.get('HX-Request') == 'true'

    if not request.user.has_perm('exams.lock_grades'):
        raise PermissionDenied("You don't have permission to lock grades.")

    if request.method == 'POST':
        form = GradeLockForm(request.POST)
        if form.is_valid():
            try:
                success = result.lock_grade(
                    locked_by=request.user,
                    reason=form.cleaned_data['lock_reason'],
                )
                if not success:
                    raise Exception('lock_grade returned False — grade may already be locked or missing.')
                if is_htmx:
                    return _htmx_redirect(
                        reverse('exams:result_detail', kwargs={'pk': pk}),
                        'success', 'Grade locked successfully.',
                    )
                messages.success(request, 'Grade locked successfully.')
                return redirect('exams:result_detail', pk=pk)
            except Exception as e:
                logger.error("Error locking grade (result pk=%s): %s", pk, e)
                if is_htmx:
                    return _htmx_alert('error', str(e))
                messages.error(request, str(e))

    return render(request, 'exams/results/lock_grade_form.html', {
        'result': result,
        'form':   GradeLockForm(),
    })


@login_required
def unlock_grade(request, pk):
    """Unlock a single result's grade (POST only)."""
    result  = get_object_or_404(StudentExamResult, pk=pk)
    is_htmx = request.headers.get('HX-Request') == 'true'

    if not result.can_unlock_grade(request.user):
        raise PermissionDenied(
            "You don't have permission to unlock this grade, or the unlock window has expired."
        )

    if request.method == 'POST':
        form = GradeUnlockForm(request.POST)
        if form.is_valid():
            try:
                success = result.unlock_grade(
                    unlocked_by=request.user,
                    reason=form.cleaned_data['unlock_reason'],
                )
                if not success:
                    raise Exception('unlock_grade returned False — grade may already be unlocked.')
                if is_htmx:
                    return _htmx_redirect(
                        reverse('exams:result_detail', kwargs={'pk': pk}),
                        'success', 'Grade unlocked successfully.',
                    )
                messages.success(request, 'Grade unlocked successfully.')
                return redirect('exams:result_detail', pk=pk)
            except Exception as e:
                logger.error("Error unlocking grade (result pk=%s): %s", pk, e)
                if is_htmx:
                    return _htmx_alert('error', str(e))
                messages.error(request, str(e))

    return render(request, 'exams/results/unlock_grade_form.html', {
        'result': result,
        'form':   GradeUnlockForm(),
    })


# =============================================================================
# REPORTS (grade sheet, mark sheet, rank list, merit list)
# =============================================================================

@login_required
def grade_sheet_report(request, examination_pk):
    examination = get_object_or_404(Examination, pk=examination_pk)
    return render(request, 'exams/reports/grade_sheet.html', {
        'examination': examination,
        'results':     examination.student_results
                           .select_related('student')
                           .order_by('student__last_name', 'student__first_name'),
    })


@login_required
def mark_sheet_report(request, examination_pk):
    examination = get_object_or_404(Examination, pk=examination_pk)
    return render(request, 'exams/reports/mark_sheet.html', {
        'examination': examination,
        'results':     examination.student_results
                           .select_related('student')
                           .order_by('student__last_name', 'student__first_name'),
    })


@login_required
def rank_list_report(request, examination_pk):
    examination = get_object_or_404(Examination, pk=examination_pk)
    return render(request, 'exams/reports/rank_list.html', {
        'examination': examination,
        'results':     examination.student_results
                           .select_related('student')
                           .order_by('-score'),
    })


@login_required
def merit_list_report(request, examination_pk):
    examination = get_object_or_404(Examination, pk=examination_pk)
    return render(request, 'exams/reports/merit_list.html', {
        'examination': examination,
        'results':     examination.student_results
                           .filter(is_pass=True)
                           .select_related('student')
                           .order_by('-score')[:50],
    })


# =============================================================================
# TIMETABLE
# =============================================================================

@login_required
def exam_timetable(request):
    return render(request, 'exams/timetable/index.html', {
        'sessions': AcademicSession.objects.filter(is_active=True).order_by('-start_date'),
    })


@login_required
def exam_timetable_session(request, session_pk):
    session = get_object_or_404(AcademicSession, pk=session_pk)
    return render(request, 'exams/timetable/session.html', {
        'session':      session,
        'examinations': Examination.objects.filter(academic_session=session)
                            .select_related('subject', 'exam_category')
                            .order_by('exam_date', 'start_time'),
    })


@login_required
def exam_timetable_print(request, session_pk):
    session = get_object_or_404(AcademicSession, pk=session_pk)
    return render(request, 'exams/timetable/print.html', {
        'session':      session,
        'examinations': Examination.objects.filter(academic_session=session)
                            .select_related('subject', 'exam_category')
                            .order_by('exam_date', 'start_time'),
        'print_date':   get_school_current_time(),
        **get_print_school_context(request),
    })


# =============================================================================
# IMPORT / EXPORT TEMPLATES
# =============================================================================

@login_required
def download_results_template(request):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Results Template'
    ws.append(['Admission Number', 'Examination Code', 'Score', 'Comments'])
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="results_template.xlsx"'
    wb.save(response)
    return response


@login_required
def download_examinations_template(request):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Examinations Template'
    ws.append(['Name', 'Code', 'Subject Code', 'Exam Date', 'Start Time', 'End Time', 'Total Marks'])
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="examinations_template.xlsx"'
    wb.save(response)
    return response


# =============================================================================
# SETTINGS
# =============================================================================

@login_required
def exam_settings(request):
    return render(request, 'exams/settings/exam_settings.html', {})


@login_required
def grade_locking_settings(request):
    return render(request, 'exams/settings/grade_locking.html', {})


# =============================================================================
# AJAX UTILITY ENDPOINTS
# =============================================================================

@login_required
def ajax_get_grading_system_ranges(request, system_pk):
    try:
        system = get_object_or_404(GradingSystem, pk=system_pk)
        return JsonResponse({'ranges': [
            {
                'id':         r.id,
                'grade':      r.grade,
                'min_score':  float(r.min_score),
                'max_score':  float(r.max_score),
                'gpa_points': float(r.gpa_points) if r.gpa_points else None,
            }
            for r in system.ranges.all().order_by('-min_score')
        ]})
    except Exception as e:
        logger.error("ajax_get_grading_system_ranges error: %s", e)
        return JsonResponse({'error': str(e)}, status=400)


@login_required
def ajax_get_examinations_for_session(request, session_pk):
    try:
        session = get_object_or_404(AcademicSession, pk=session_pk)
        return JsonResponse({'examinations': [
            {
                'id':        e.id,
                'name':      e.name,
                'code':      e.code,
                'subject':   e.subject.name,
                'exam_date': e.exam_date.strftime('%Y-%m-%d'),
            }
            for e in Examination.objects.filter(
                academic_session=session
            ).select_related('subject')
        ]})
    except Exception as e:
        logger.error("ajax_get_examinations_for_session error: %s", e)
        return JsonResponse({'error': str(e)}, status=400)


@login_required
def ajax_calculate_grade(request):
    try:
        score          = Decimal(request.GET.get('score', 0))
        grading_system = get_object_or_404(GradingSystem, pk=request.GET.get('grading_system'))
        grade_info     = grading_system.get_grade_for_score(float(score))
        if grade_info:
            return JsonResponse({
                'grade':      grade_info['grade'],
                'gpa_points': grade_info['gpa_points'],
                'is_passing': grade_info['is_passing'],
                'comments':   grade_info['comments'],
            })
        return JsonResponse({'error': 'No grade found for this score.'})
    except Exception as e:
        logger.error("ajax_calculate_grade error: %s", e)
        return JsonResponse({'error': str(e)}, status=400)


@login_required
def ajax_get_exam_statistics(request, examination_pk):
    try:
        examination = get_object_or_404(Examination, pk=examination_pk)
        results     = examination.student_results.filter(status='COMPLETED')
        agg         = results.aggregate(
            total      = Count('id'),
            highest    = Max('score'),
            lowest     = Min('score'),
            average    = Avg('score'),
            pass_count = Count('id', filter=Q(is_pass=True)),
        )
        total = agg['total'] or 0
        return JsonResponse({
            'total_results': total,
            'highest_score': float(agg['highest']) if agg['highest'] else 0,
            'lowest_score':  float(agg['lowest'])  if agg['lowest']  else 0,
            'average_score': round(float(agg['average']), 2) if agg['average'] else 0,
            'pass_count':    agg['pass_count'] or 0,
            'pass_rate':     round((agg['pass_count'] or 0) / total * 100, 2) if total else 0,
        })
    except Exception as e:
        logger.error("ajax_get_exam_statistics error: %s", e)
        return JsonResponse({'error': str(e)}, status=400)


@login_required
def ajax_validate_grade_unlock(request, result_pk):
    try:
        result = get_object_or_404(StudentExamResult, pk=result_pk)
        return JsonResponse({
            'can_unlock': result.can_unlock_grade(request.user),
            'is_locked':  result.is_grade_locked,
            'locked_by':  result.grade_locked_by.get_full_name() if result.grade_locked_by else None,
            'locked_at':  result.grade_locked_at.strftime('%Y-%m-%d %H:%M') if result.grade_locked_at else None,
        })
    except Exception as e:
        logger.error("ajax_validate_grade_unlock error: %s", e)
        return JsonResponse({'error': str(e)}, status=400)


# =============================================================================
# PRIVATE HELPERS — shared by print and export views
# =============================================================================

def _htmx_alert(alert_type: str, message: str) -> HttpResponse:
    """Return a bare HTMX response that triggers a front-end alert."""
    r = HttpResponse()
    r['HX-Trigger']      = 'showAlert'
    r['HX-Trigger-Data'] = f'{{"type":"{alert_type}","message":"{message}"}}'
    return r


def _htmx_redirect(url: str, alert_type: str, message: str) -> HttpResponse:
    """Return a bare HTMX response that redirects and triggers a front-end alert."""
    r = HttpResponse()
    r['HX-Redirect']     = url
    r['HX-Trigger']      = 'showAlert'
    r['HX-Trigger-Data'] = f'{{"type":"{alert_type}","message":"{message}"}}'
    return r


def _resolve_class_and_session(class_pk, request):
    """
    Resolve Class and AcademicSession from URL + query params.

    Performs cross-session class lookup when the session differs from the
    class's own session. Returns (class_instance, session).
    """
    class_instance = get_object_or_404(Class, pk=class_pk)

    session_id = request.GET.get('session')
    session    = (
        get_object_or_404(AcademicSession, pk=session_id)
        if session_id
        else class_instance.academic_session or AcademicSession.get_current_session()
    )

    if session and class_instance.academic_session_id != session.pk:
        equivalent = Class.objects.filter(
            academic_level   = class_instance.academic_level,
            section          = class_instance.section,
            academic_session = session,
        ).first()
        if equivalent:
            class_instance = equivalent

    return class_instance, session


def _build_results_grid(class_instance, session, category_abbr=None):
    """
    Build the student × exam grid used by print and export views.

    Returns (examinations, students, grid_rows, category).
    Rows are ranked by total score descending; ties share a position.
    """
    exam_qs = Examination.objects.filter(
        target_classes=class_instance,
        academic_session=session,
    ).select_related('exam_category', 'subject').order_by(
        'exam_category__name', 'subject__name'
    )
    if category_abbr:
        exam_qs = exam_qs.filter(exam_category__abbreviation=category_abbr)

    examinations = list(exam_qs)
    category     = (
        exam_qs.first().exam_category
        if category_abbr and exam_qs.exists()
        else None
    )

    today             = get_school_today()
    enrollment_filter = dict(
        class_enrollments__class_instance=class_instance,
        class_enrollments__academic_session=session,
    )
    if session.end_date >= today:
        enrollment_filter['class_enrollments__is_active']        = True
        enrollment_filter['class_enrollments__completion_status'] = 'ONGOING'

    students = Student.objects.filter(
        **enrollment_filter
    ).distinct().order_by('first_name', 'last_name')

    result_map = {
        (r.student_id, r.examination_id): r
        for r in StudentExamResult.objects.filter(
            examination__in=examinations,
            student__in=students,
        ).select_related('examination')
    }

    grid_rows = []
    for student in students:
        cells  = []
        scores = []
        for exam in examinations:
            result = result_map.get((student.pk, exam.pk))
            score  = result.score if result else None
            cells.append({
                'score':        score,
                'grade':        result.grade if result else '',
                'is_locked':    result.is_grade_locked if result else False,
                'is_published': result.is_published if result else False,
            })
            if score is not None:
                scores.append(score)

        total   = sum(scores)
        average = round(total / len(scores), 2) if scores else Decimal('0')
        grid_rows.append({
            'student':        student,
            'cells':          cells,
            'total':          total,
            'average':        average,
            'subjects_taken': len(scores),
        })

    grid_rows.sort(key=lambda r: r['total'], reverse=True)
    position   = 1
    prev_total = None
    for i, row in enumerate(grid_rows):
        if row['total'] != prev_total:
            position = i + 1
        row['position'] = position
        prev_total      = row['total']

    return examinations, students, grid_rows, category


def _build_student_card_data(results_qs, grading_system):
    """
    Enrich a queryset of StudentExamResult with grade_info from the grading
    system and compute PLE-style aggregate totals and division.

    PLE standard (Uganda Primary Leaving Examinations):
      - Each subject score maps to a grade (D1–F9) via GradingRange.aggregate
        e.g. D1=1, D2=2, C3=3, C4=4, P5=5, P6=6, P7=7, F8=8, F9=9
      - Best 4 subject aggregates are summed
      - Division I:   4–12   (best performance)
      - Division II:  13–23
      - Division III: 24–29
      - Division IV:  30–35
      - Ungraded (U): 36+

    Returns:
        enriched_rows  list of dicts — one per result, with grade_info attached
        aggregate_total  int — sum of best-N aggregate weights
        division         str — 'Division I' … 'Division U' or None
        gpa              float or None
    """
    enriched_rows    = []
    aggregate_values = []   # numeric weights only (the digit in D1, C3, P7 etc.)
    gpa_points_list  = []

    for result in results_qs:
        grade_info = None
        agg_numeric = None

        if grading_system and result.score is not None:
            grade_info = grading_system.get_grade_for_score(result.score)

            if grade_info:
                # Extract the numeric weight from the aggregate string
                # "D1" → 1, "C3" → 3, "P7" → 7, "F9" → 9
                agg_str    = str(grade_info.get('aggregate', '') or '')
                digits     = ''.join(filter(str.isdigit, agg_str))
                if digits:
                    agg_numeric = int(digits)
                    aggregate_values.append(agg_numeric)

                gpa = grade_info.get('gpa_points')
                if gpa is not None:
                    gpa_points_list.append(float(gpa))

        enriched_rows.append({
            'result':      result,
            'grade_info':  grade_info,   # full dict from get_grade_for_score()
            'agg_numeric': agg_numeric,  # int or None
        })

    # ── Determine how many subjects to use for division ───────────────────────
    # Respect GradingSystem.maximum_subjects_considered (PLE uses best 4)
    max_n = None
    if grading_system and grading_system.maximum_subjects_considered:
        max_n = grading_system.maximum_subjects_considered

    if max_n and len(aggregate_values) > max_n:
        # PLE: lowest aggregate = best result, so sort ascending and take first N
        agg_for_division = sorted(aggregate_values)[:max_n]
    else:
        agg_for_division = aggregate_values

    aggregate_total = sum(agg_for_division) if agg_for_division else None

    # ── PLE Division thresholds ───────────────────────────────────────────────
    division = None
    if aggregate_total is not None:
        if   aggregate_total <= 12: division = 'Division I'
        elif aggregate_total <= 23: division = 'Division II'
        elif aggregate_total <= 29: division = 'Division III'
        elif aggregate_total <= 34: division = 'Division IV'
        else:                       division = 'Ungraded (U)'

    # ── GPA (only if grading system uses it) ─────────────────────────────────
    gpa = None
    if grading_system and grading_system.uses_gpa and gpa_points_list:
        gpa = round(sum(gpa_points_list) / len(gpa_points_list), 2)

    return enriched_rows, aggregate_total, division, gpa


def _style_header_row(ws, header_fill_hex='1E3A5F'):
    """Apply standard header styling to row 1 of a worksheet."""
    fill  = PatternFill(start_color=header_fill_hex, end_color=header_fill_hex, fill_type='solid')
    font  = Font(bold=True, color='FFFFFF', size=11)
    align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    for cell in ws[1]:
        cell.fill      = fill
        cell.font      = font
        cell.alignment = align
    ws.row_dimensions[1].height = 28


def _auto_size_columns(ws, max_width=60):
    """Auto-size all columns, capped at max_width characters."""
    for col_cells in ws.columns:
        letter  = get_column_letter(col_cells[0].column)
        max_len = max((len(str(c.value)) if c.value else 0) for c in col_cells)
        ws.column_dimensions[letter].width = min(max_len + 4, max_width)


def _excel_response(wb, filename_prefix: str) -> HttpResponse:
    """Return an HttpResponse that streams an xlsx workbook as a download."""
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = (
        f'attachment; filename="{filename_prefix}_{datetime.now():%Y%m%d_%H%M}.xlsx"'
    )
    wb.save(response)
    return response


# =============================================================================
# PRINT VIEWS
# =============================================================================

@login_required
def examination_print_list(request):
    """
    Printable examination list with selectable columns.

    Query params:
      ?fields=code,name,...    columns to include (ordered)
      ?short_headers=true      abbreviated column headers
      ?include_stats=true      append a summary statistics block
      ?landscape=true          landscape page layout
    """
    selected_fields = request.GET.getlist('fields') or [
        'code', 'name', 'subject', 'exam_category',
        'exam_date', 'start_time', 'total_marks', 'status',
    ]

    include_stats  = request.GET.get('include_stats') == 'true'
    landscape_mode = request.GET.get('landscape') == 'true'
    short_headers  = request.GET.get('short_headers') == 'true'

    examinations = _get_filtered_examinations(request)

    stats = None
    if include_stats:
        today = get_school_today()
        stats = {
            'total':     examinations.count(),
            'planned':   examinations.filter(status='PLANNED').count(),
            'scheduled': examinations.filter(status='SCHEDULED').count(),
            'ongoing':   examinations.filter(status='ONGOING').count(),
            'completed': examinations.filter(status='COMPLETED').count(),
            'upcoming':  examinations.filter(
                exam_date__gte=today, status__in=['PLANNED', 'SCHEDULED']
            ).count(),
            'published': examinations.filter(results_published=True).count(),
        }

    _FIELD_NAMES_FULL = {
        'code':              'Examination Code',
        'name':              'Examination Name',
        'subject':           'Subject',
        'exam_category':     'Category',
        'academic_session':  'Academic Session',
        'exam_date':         'Exam Date',
        'start_time':        'Start Time',
        'end_time':          'End Time',
        'duration_minutes':  'Duration (min)',
        'total_marks':       'Total Marks',
        'pass_marks':        'Pass Marks',
        'exam_mode':         'Mode',
        'status':            'Status',
        'results_published': 'Results Published',
        'examination_venue': 'Venue',
        'target_classes':    'Target Classes',
    }
    _FIELD_NAMES_SHORT = {
        'code':              'Code',
        'name':              'Name',
        'subject':           'Subject',
        'exam_category':     'Category',
        'academic_session':  'Session',
        'exam_date':         'Date',
        'start_time':        'Start',
        'end_time':          'End',
        'duration_minutes':  'Dur.',
        'total_marks':       'Marks',
        'pass_marks':        'Pass',
        'exam_mode':         'Mode',
        'status':            'Status',
        'results_published': 'Published',
        'examination_venue': 'Venue',
        'target_classes':    'Classes',
    }

    field_names = _FIELD_NAMES_SHORT if short_headers else _FIELD_NAMES_FULL

    return render(request, 'exams/examinations/print.html', {
        'examinations':         examinations,
        'stats':                stats,
        'now':                  timezone.now(),
        'selected_fields':      selected_fields,
        'selected_field_names': [field_names.get(f, f.replace('_', ' ').title()) for f in selected_fields],
        'field_names':          field_names,
        'landscape':            landscape_mode,
        'short_headers':        short_headers,
        **get_print_school_context(request),
    })


@login_required
def exam_category_print_list(request):
    """Printable exam-category list with selectable columns."""
    selected_fields = request.GET.getlist('fields') or [
        'name', 'abbreviation', 'category_type', 'frequency',
        'weight_percentage', 'allows_retakes', 'is_active',
    ]

    include_stats  = request.GET.get('include_stats') == 'true'
    landscape_mode = request.GET.get('landscape') == 'true'
    short_headers  = request.GET.get('short_headers') == 'true'

    categories = _get_filtered_exam_categories(request)

    stats = None
    if include_stats:
        stats = {
            'total':     categories.count(),
            'active':    categories.filter(is_active=True).count(),
            'inactive':  categories.filter(is_active=False).count(),
            'formative': categories.filter(category_type='FORMATIVE').count(),
            'summative': categories.filter(category_type='SUMMATIVE').count(),
        }

    _FIELD_NAMES_FULL = {
        'name':                     'Category Name',
        'abbreviation':             'Abbreviation',
        'code':                     'Code',
        'category_type':            'Category Type',
        'frequency':                'Frequency',
        'weight_percentage':        'Weight (%)',
        'allows_retakes':           'Allows Retakes',
        'max_retakes':              'Max Retakes',
        'public_results':           'Public Results',
        'is_active':                'Active',
        'curriculum_compatibility': 'Curriculum',
    }
    _FIELD_NAMES_SHORT = {
        'name':                     'Name',
        'abbreviation':             'Abbr.',
        'code':                     'Code',
        'category_type':            'Type',
        'frequency':                'Freq.',
        'weight_percentage':        'Weight',
        'allows_retakes':           'Retakes',
        'max_retakes':              'Max Ret.',
        'public_results':           'Public',
        'is_active':                'Active',
        'curriculum_compatibility': 'Curriculum',
    }

    field_names = _FIELD_NAMES_SHORT if short_headers else _FIELD_NAMES_FULL

    return render(request, 'exams/categories/print.html', {
        'categories':           categories,
        'stats':                stats,
        'now':                  timezone.now(),
        'selected_fields':      selected_fields,
        'selected_field_names': [field_names.get(f, f.replace('_', ' ').title()) for f in selected_fields],
        'field_names':          field_names,
        'landscape':            landscape_mode,
        'short_headers':        short_headers,
        **get_print_school_context(request),
    })


@login_required
def grading_system_print_list(request):
    """
    Printable grading-system list with selectable columns.
    Pass ?include_ranges=true to append per-system grade-range tables.
    """
    selected_fields = request.GET.getlist('fields') or [
        'name', 'code', 'grading_type', 'scale_type',
        'minimum_score', 'maximum_score', 'pass_mark',
        'uses_gpa', 'is_active', 'is_default',
    ]

    include_stats  = request.GET.get('include_stats') == 'true'
    landscape_mode = request.GET.get('landscape') == 'true'
    short_headers  = request.GET.get('short_headers') == 'true'
    include_ranges = request.GET.get('include_ranges') == 'true'

    systems = _get_filtered_grading_systems(request)

    stats = None
    if include_stats:
        stats = {
            'total':    systems.count(),
            'active':   systems.filter(is_active=True).count(),
            'default':  systems.filter(is_default=True).count(),
            'with_gpa': systems.filter(uses_gpa=True).count(),
        }

    _FIELD_NAMES_FULL = {
        'name':          'System Name',
        'code':          'Code',
        'grading_type':  'Grading Type',
        'scale_type':    'Scale Type',
        'minimum_score': 'Min Score',
        'maximum_score': 'Max Score',
        'pass_mark':     'Pass Mark',
        'uses_gpa':      'Uses GPA',
        'is_active':     'Active',
        'is_default':    'Default',
        'ranges_count':  'Grade Ranges',
    }
    _FIELD_NAMES_SHORT = {
        'name':          'Name',
        'code':          'Code',
        'grading_type':  'Type',
        'scale_type':    'Scale',
        'minimum_score': 'Min',
        'maximum_score': 'Max',
        'pass_mark':     'Pass',
        'uses_gpa':      'GPA',
        'is_active':     'Active',
        'is_default':    'Default',
        'ranges_count':  'Ranges',
    }

    field_names = _FIELD_NAMES_SHORT if short_headers else _FIELD_NAMES_FULL

    return render(request, 'exams/grading_systems/print.html', {
        'systems':              systems,
        'stats':                stats,
        'now':                  timezone.now(),
        'selected_fields':      selected_fields,
        'selected_field_names': [field_names.get(f, f.replace('_', ' ').title()) for f in selected_fields],
        'field_names':          field_names,
        'landscape':            landscape_mode,
        'short_headers':        short_headers,
        'include_ranges':       include_ranges,
        **get_print_school_context(request),
    })


@login_required
def class_results_print_view(request, class_pk):
    """
    Printable marks sheet for one class + session.

    Query params:
      ?session=<pk>           override session
      ?category=<abbr>        limit to one exam-category tab
      ?show_grade=true        show letter grade column
      ?show_position=true     show class position column
      ?show_average=true      show average column
      ?include_stats=true     append class summary statistics
      ?landscape=true         landscape layout
    """
    class_instance, session = _resolve_class_and_session(class_pk, request)
    if not session:
        messages.warning(request, 'No active session found.')
        return redirect('exams:results_by_class')

    category_abbr = request.GET.get('category', '')
    include_stats = request.GET.get('include_stats') == 'true'
    landscape     = request.GET.get('landscape') == 'true'
    show_grade    = request.GET.get('show_grade', 'true') == 'true'
    show_position = request.GET.get('show_position', 'true') == 'true'
    show_average  = request.GET.get('show_average', 'true') == 'true'

    examinations, _students, grid_rows, category = _build_results_grid(
        class_instance, session, category_abbr or None
    )

    all_categories = (
        ExamCategory.objects.filter(
            examinations__target_classes=class_instance,
            examinations__academic_session=session,
        ).distinct().order_by('name')
    )

    stats = None
    if include_stats and grid_rows:
        totals   = [r['total']   for r in grid_rows if r['subjects_taken'] > 0]
        averages = [r['average'] for r in grid_rows if r['subjects_taken'] > 0]
        stats = {
            'student_count': len(grid_rows),
            'highest_total': max(totals)  if totals   else 0,
            'lowest_total':  min(totals)  if totals   else 0,
            'class_average': round(sum(averages) / len(averages), 2) if averages else 0,
        }

    return render(request, 'exams/results/print_class_results.html', {
        'class_instance': class_instance,
        'session':        session,
        'category':       category,
        'all_categories': all_categories,
        'examinations':   examinations,
        'grid_rows':      grid_rows,
        'stats':          stats,
        'now':            timezone.now(),
        'landscape':      landscape,
        'show_grade':     show_grade,
        'show_position':  show_position,
        'show_average':   show_average,
        **get_print_school_context(request),
    })


# =============================================================================
# REPORT CARD — SINGLE STUDENT
# =============================================================================

@login_required
def student_report_card(request, student_pk):
    """
    Full academic report card for a single student.

    URL:  /exams/report-card/<student_pk>/
          /exams/report-card/<student_pk>/?session=<pk>
          /exams/report-card/<student_pk>/?session=<pk>&class=<pk>
    """
    from collections import defaultdict
    from django.db.models import Sum as DSum
    from academics.models import StudentClassEnrollment

    student = get_object_or_404(
        Student.objects.select_related('current_academic_level'),
        pk=student_pk,
    )

    session_id = request.GET.get('session')
    session    = (
        get_object_or_404(AcademicSession, pk=session_id)
        if session_id
        else get_active_academic_session()
             or AcademicSession.objects.filter(is_active=True).order_by('-start_date').first()
    )
    if not session:
        messages.warning(request, 'No academic session found.')
        return redirect('students:student_profile', pk=student_pk)

    class_pk   = request.GET.get('class')
    enrollment = (
        StudentClassEnrollment.objects.filter(
            student=student,
            class_instance_id=class_pk,
            academic_session=session,
        ).select_related('class_instance__academic_level').first()
        if class_pk
        else StudentClassEnrollment.objects.filter(
            student=student,
            academic_session=session,
        ).select_related('class_instance__academic_level').order_by(
            '-is_active', '-class_instance__academic_level__order'
        ).first()
    )
    class_instance = enrollment.class_instance if enrollment else None

    results_qs = StudentExamResult.objects.filter(
        student=student,
        examination__academic_session=session,
        status__in=['COMPLETED', 'SUBMITTED'],
    ).select_related(
        'examination__subject',
        'examination__exam_category',
        'examination__grading_system',
    ).order_by(
        'examination__exam_category__name',
        'examination__subject__name',
    )

    # Filter to a single category if ?category=<pk> is present
    category_id = request.GET.get('category')
    selected_category = None
    if category_id:
        results_qs = results_qs.filter(examination__exam_category_id=category_id)
        from .models import ExamCategory
        selected_category = ExamCategory.objects.filter(pk=category_id).first()

    categories_data: dict = {}
    for result in results_qs:
        cat  = result.examination.exam_category
        abbr = cat.abbreviation
        if abbr not in categories_data:
            categories_data[abbr] = {'category': cat, 'rows': [], 'scores': []}
        categories_data[abbr]['rows'].append(result)
        if result.score is not None:
            categories_data[abbr]['scores'].append(result.score)

    for data in categories_data.values():
        scores         = data['scores']
        data['total']   = sum(scores)
        data['average'] = round(sum(scores) / len(scores), 2) if scores else Decimal('0')

    all_scores = [r.score for r in results_qs if r.score is not None]
    overall = {
        'total':    sum(all_scores),
        'average':  round(sum(all_scores) / len(all_scores), 2) if all_scores else Decimal('0'),
        'subjects': len(all_scores),
        'passed':   sum(1 for r in results_qs if r.is_pass),
        'failed':   sum(1 for r in results_qs if not r.is_pass and r.score is not None),
    }

    class_position = class_size = None
    if class_instance:
        peer_totals = list(
            StudentExamResult.objects.filter(
                examination__academic_session=session,
                examination__target_classes=class_instance,
                status__in=['COMPLETED', 'SUBMITTED'],
            )
            .values('student_id')
            .annotate(total=DSum('score'))
            .order_by('-total')
        )
        class_size = len(peer_totals)
        position   = 1
        prev_total = None
        for i, peer in enumerate(peer_totals):
            if peer['total'] != prev_total:
                position = i + 1
            if peer['student_id'] == student.pk:
                class_position = position
            prev_total = peer['total']

    grading_system = None
    if class_instance:
        grading_system = ClassGradingSystem.get_active_grading_system(class_instance, session)
    if not grading_system:
        grading_system = GradingSystem.objects.filter(is_default=True, is_active=True).first()

    grading_ranges = grading_system.ranges.all().order_by('-min_score') if grading_system else []

    # Build subject → teacher initials map from ClassSubject assignments
    teacher_initials_map = {}
    if class_instance:
        from academics.models import ClassSubject
        for cs in ClassSubject.objects.filter(
            class_instance=class_instance,
            is_active=True,
        ).select_related('subject', 'teacher'):
            if cs.teacher:
                parts    = [cs.teacher.first_name or '', cs.teacher.last_name or '']
                initials = '.'.join(p[0].upper() for p in parts if p) + '.'
                teacher_initials_map[cs.subject_id] = initials

    # Enrich results with grade_info (comments, aggregate, gpa_points) and
    # compute PLE division from the grading system's aggregate weights
    enriched_rows, aggregate_total, division, gpa = _build_student_card_data(
        list(results_qs), grading_system
    )

    # Attach teacher initials to each row so templates can just use row.teacher_initials
    for row in enriched_rows:
        subj_id = row['result'].examination.subject_id
        row['teacher_initials'] = teacher_initials_map.get(subj_id, '')

    # Rebuild categories_data using the enriched rows
    categories_data: dict = {}
    all_scores: list = []
    for row in enriched_rows:
        result = row['result']
        cat    = result.examination.exam_category
        abbr   = cat.abbreviation
        if abbr not in categories_data:
            categories_data[abbr] = {'category': cat, 'rows': [], 'scores': []}
        categories_data[abbr]['rows'].append(row)
        if result.score is not None:
            categories_data[abbr]['scores'].append(result.score)
            all_scores.append(result.score)

    for data in categories_data.values():
        scores           = data['scores']
        data['total']    = sum(scores)
        data['average']  = round(sum(scores) / len(scores), 2) if scores else Decimal('0')
        data['agg_total'] = sum(
            row['agg_numeric'] for row in data['rows']
            if row.get('agg_numeric') is not None
        )

    overall = {
        'total':    sum(all_scores),
        'average':  round(sum(all_scores) / len(all_scores), 2) if all_scores else Decimal('0'),
        'subjects': len(all_scores),
        'passed':   sum(1 for row in enriched_rows if row['result'].is_pass),
        'failed':   sum(
            1 for row in enriched_rows
            if not row['result'].is_pass and row['result'].score is not None
        ),
    }

    # Find the next academic session after this one
    next_session = (
        AcademicSession.objects
        .filter(start_date__gt=session.end_date)
        .order_by('start_date')
        .first()
    )

    return render(request, 'exams/results/report_card.html', {
        'student':              student,
        'session':              session,
        'all_sessions':         AcademicSession.objects.filter(is_active=True).order_by('-start_date'),
        'enrollment':           enrollment,
        'class_instance':       class_instance,
        'categories_data':      categories_data,
        'selected_category':    selected_category,
        'overall':              overall,
        'class_position':       class_position,
        'class_size':           class_size,
        'grading_system':       grading_system,
        'grading_ranges':       grading_ranges,
        'aggregate_total':      aggregate_total,
        'division':             division,
        'gpa':                  gpa,
        'teacher_initials_map': teacher_initials_map,
        'next_session':         next_session,
        'now':                  timezone.now(),
        **get_print_school_context(request),
    })


# =============================================================================
# REPORT CARD — BULK (all students in a class, separated by page-break)
# =============================================================================

@login_required
def class_report_cards(request, class_pk):
    """
    Bulk report-card print view — one card per student.

    URL:  /exams/report-cards/class/<class_pk>/
          /exams/report-cards/class/<class_pk>/?session=<pk>
          /exams/report-cards/class/<class_pk>/?session=<pk>&category=<pk>
    """
    from collections import defaultdict
    from django.db.models import Sum as DSum

    class_instance, session = _resolve_class_and_session(class_pk, request)
    if not session:
        messages.warning(request, 'No active session found.')
        return redirect('exams:results_by_class')

    # Optional category filter — passed from the marks grid Print button
    category_id       = request.GET.get('category')
    selected_category = None
    if category_id:
        selected_category = ExamCategory.objects.filter(pk=category_id).first()

    today             = get_school_today()
    enrollment_filter = dict(
        class_enrollments__class_instance=class_instance,
        class_enrollments__academic_session=session,
    )
    if session.end_date >= today:
        enrollment_filter['class_enrollments__is_active']        = True
        enrollment_filter['class_enrollments__completion_status'] = 'ONGOING'

    students = Student.objects.filter(
        **enrollment_filter
    ).distinct().select_related('current_academic_level').order_by('first_name', 'last_name')

    # ── Payment eligibility check ─────────────────────────────────────────────
    # Filter students based on ExamCategory.required_payment_percentage.
    # If no category is selected or the category has no payment requirement,
    # all enrolled students are included.
    eligible_student_ids = None   # None = no filter applied

    if selected_category and selected_category.required_payment_percentage > 0:
        from fees.models import FeeInvoice
        from django.db.models import Sum as FSum

        required_pct = selected_category.required_payment_percentage
        consider_all = selected_category.consider_all_outstanding_balances

        eligible_student_ids = set()
        ineligible_count = 0

        for student in students:
            if consider_all:
                # All sessions up to and including the current session
                invs = FeeInvoice.objects.filter(
                    student=student,
                    fiscal_period__related_academic_session__start_date__lte=session.start_date,
                ).exclude(total_amount=0)
            else:
                # Current session only
                invs = FeeInvoice.objects.filter(
                    student=student,
                    fiscal_period__related_academic_session=session,
                ).exclude(total_amount=0)

            totals = invs.aggregate(
                total=FSum('total_amount'),
                paid=FSum('paid_amount'),
            )
            total_amount = totals['total'] or 0
            paid_amount  = totals['paid']  or 0

            if total_amount == 0:
                # No invoice or all voided — include by default
                eligible_student_ids.add(student.pk)
            else:
                pct_paid = (paid_amount / total_amount) * 100
                if pct_paid >= required_pct:
                    eligible_student_ids.add(student.pk)
                else:
                    ineligible_count += 1

        students = students.filter(pk__in=eligible_student_ids)

    results_qs = StudentExamResult.objects.filter(
        student__in=students,
        examination__academic_session=session,
        examination__target_classes=class_instance,
        status__in=['COMPLETED', 'SUBMITTED'],
    ).select_related(
        'student',
        'examination__subject',
        'examination__exam_category',
    ).order_by(
        'student_id',
        'examination__exam_category__name',
        'examination__subject__name',
    )

    # Filter to a single category if requested
    if selected_category:
        results_qs = results_qs.filter(examination__exam_category=selected_category)

    results_by_student: dict = defaultdict(lambda: defaultdict(list))
    for result in results_qs:
        results_by_student[result.student_id][
            result.examination.exam_category.abbreviation
        ].append(result)

    # Class positions — scoped to the category if one is selected
    position_qs = StudentExamResult.objects.filter(
        examination__academic_session=session,
        examination__target_classes=class_instance,
        status__in=['COMPLETED', 'SUBMITTED'],
    )
    if selected_category:
        position_qs = position_qs.filter(examination__exam_category=selected_category)

    peer_totals = {
        p['student_id']: p['total'] or Decimal('0')
        for p in (
            position_qs
            .values('student_id')
            .annotate(total=DSum('score'))
        )
    }
    sorted_peers = sorted(peer_totals.items(), key=lambda x: x[1], reverse=True)
    rank_map: dict = {}
    position   = 1
    prev_total = None
    for i, (sid, total) in enumerate(sorted_peers):
        if total != prev_total:
            position = i + 1
        rank_map[sid] = position
        prev_total    = total

    class_size = len(sorted_peers)

    grading_system = ClassGradingSystem.get_active_grading_system(class_instance, session)
    if not grading_system:
        grading_system = GradingSystem.objects.filter(is_default=True, is_active=True).first()

    grading_ranges = grading_system.ranges.all().order_by('-min_score') if grading_system else []

    # Build subject → teacher initials map — same for all students in this class
    from academics.models import ClassSubject
    teacher_initials_map = {}
    for cs in ClassSubject.objects.filter(
        class_instance=class_instance,
        is_active=True,
    ).select_related('subject', 'teacher'):
        if cs.teacher:
            parts    = [cs.teacher.first_name or '', cs.teacher.last_name or '']
            initials = '.'.join(p[0].upper() for p in parts if p) + '.'
            teacher_initials_map[cs.subject_id] = initials

    cards = []
    for student in students:
        student_results = results_by_student.get(student.pk, {})

        # Flatten all results for this student into a single list for _build_student_card_data
        flat_results = [r for rows in student_results.values() for r in rows]

        if not flat_results:
            continue

        enriched_rows, aggregate_total, division, gpa = _build_student_card_data(
            flat_results, grading_system
        )

        # Attach teacher initials per row
        for row in enriched_rows:
            subj_id = row['result'].examination.subject_id
            row['teacher_initials'] = teacher_initials_map.get(subj_id, '')

        # Rebuild categories_data using enriched rows
        categories_data: dict = {}
        all_scores: list = []
        for row in enriched_rows:
            result = row['result']
            cat    = result.examination.exam_category
            abbr   = cat.abbreviation
            if abbr not in categories_data:
                categories_data[abbr] = {'category': cat, 'rows': [], 'scores': []}
            categories_data[abbr]['rows'].append(row)
            if result.score is not None:
                categories_data[abbr]['scores'].append(result.score)
                all_scores.append(result.score)

        for data in categories_data.values():
            scores            = data['scores']
            data['total']     = sum(scores)
            data['average']   = round(sum(scores) / len(scores), 2) if scores else Decimal('0')
            data['agg_total'] = sum(
                row['agg_numeric'] for row in data['rows']
                if row.get('agg_numeric') is not None
            )

        overall = {
            'total':    sum(all_scores),
            'average':  round(sum(all_scores) / len(all_scores), 2) if all_scores else Decimal('0'),
            'subjects': len(all_scores),
            'passed':   sum(1 for row in enriched_rows if row['result'].is_pass),
            'failed':   sum(
                1 for row in enriched_rows
                if not row['result'].is_pass and row['result'].score is not None
            ),
        }

        cards.append({
            'student':         student,
            'categories_data': categories_data,
            'overall':         overall,
            'class_position':  rank_map.get(student.pk),
            'class_size':      class_size,
            'aggregate_total': aggregate_total,
            'division':        division,
            'gpa':             gpa,
        })

    # Sort best to worst — students with no position go to the end
    cards.sort(key=lambda c: c['class_position'] if c['class_position'] else 9999)

    next_session = (
        AcademicSession.objects
        .filter(start_date__gt=session.end_date)
        .order_by('start_date')
        .first()
    )

    return render(request, 'exams/results/class_report_cards.html', {
        'class_instance':       class_instance,
        'session':              session,
        'selected_category':    selected_category,
        'all_sessions':         AcademicSession.objects.filter(is_active=True).order_by('-start_date'),
        'cards':                cards,
        'grading_system':       grading_system,
        'grading_ranges':       grading_ranges,
        'teacher_initials_map': teacher_initials_map,
        'next_session':         next_session,
        'ineligible_count':     ineligible_count if eligible_student_ids is not None else 0,
        'required_pct':         selected_category.required_payment_percentage if selected_category else None,
        'now':                  timezone.now(),
        **get_print_school_context(request),
    })


# =============================================================================
# REPORT CARD ELIGIBILITY — printable list of eligible/ineligible students
# =============================================================================

@login_required
def report_card_eligibility(request, class_pk):
    """
    Printable eligibility list showing which students are eligible for
    report cards based on fee payment status.

    URL:  /exams/report-cards/class/<class_pk>/eligibility/
          ?session=<pk>&category=<pk>
    """
    from fees.models import FeeInvoice
    from django.db.models import Sum as FSum

    class_instance, session = _resolve_class_and_session(class_pk, request)
    if not session:
        messages.warning(request, 'No active session found.')
        return redirect('exams:results_by_class')

    category_id       = request.GET.get('category')
    selected_category = None
    if category_id:
        selected_category = ExamCategory.objects.filter(pk=category_id).first()

    today             = get_school_today()
    enrollment_filter = dict(
        class_enrollments__class_instance=class_instance,
        class_enrollments__academic_session=session,
    )
    if session.end_date >= today:
        enrollment_filter['class_enrollments__is_active']        = True
        enrollment_filter['class_enrollments__completion_status'] = 'ONGOING'

    students = Student.objects.filter(
        **enrollment_filter
    ).distinct().select_related('current_academic_level').order_by('first_name', 'last_name')

    required_pct  = selected_category.required_payment_percentage if selected_category else 0
    consider_all  = selected_category.consider_all_outstanding_balances if selected_category else False

    eligible   = []
    ineligible = []

    for student in students:
        if consider_all:
            # All sessions up to and including the current session
            invs = FeeInvoice.objects.filter(
                student=student,
                fiscal_period__related_academic_session__start_date__lte=session.start_date,
            ).exclude(total_amount=0)
        else:
            invs = FeeInvoice.objects.filter(
                student=student,
                fiscal_period__related_academic_session=session,
            ).exclude(total_amount=0)

        totals       = invs.aggregate(total=FSum('total_amount'), paid=FSum('paid_amount'))
        total_amount = totals['total'] or 0
        paid_amount  = totals['paid']  or 0
        pct_paid     = round((paid_amount / total_amount) * 100, 1) if total_amount else 100
        balance      = total_amount - paid_amount

        entry = {
            'student':       student,
            'total_amount':  total_amount,
            'paid_amount':   paid_amount,
            'balance':       balance,
            'pct_paid':      pct_paid,
            'is_eligible':   pct_paid >= required_pct if required_pct else True,
        }

        if entry['is_eligible']:
            eligible.append(entry)
        else:
            ineligible.append(entry)

    # Sort each list by payment percentage descending
    eligible.sort(key=lambda x: x['pct_paid'], reverse=True)
    ineligible.sort(key=lambda x: x['pct_paid'], reverse=True)

    return render(request, 'exams/results/report_card_eligibility.html', {
        'class_instance':    class_instance,
        'session':           session,
        'selected_category': selected_category,
        'eligible':          eligible,
        'ineligible':        ineligible,
        'required_pct':      required_pct,
        'consider_all':      consider_all,
        'now':               timezone.now(),
        **get_print_school_context(request),
    })


# =============================================================================
# EXCEL EXPORTS
# =============================================================================

@login_required
def export_examinations_excel(request):
    """
    Export examinations to Excel with selectable columns.

    Query params:
      ?fields=code,name,...   ordered columns to include
      All filter params from _get_filtered_examinations are respected.
    """
    ALL_COLUMNS = [
        ('code',              'Code',              lambda e: e.code),
        ('name',              'Examination Name',  lambda e: e.name),
        ('subject',           'Subject',           lambda e: e.subject.name),
        ('exam_category',     'Category',          lambda e: e.exam_category.name),
        ('academic_session',  'Session',           lambda e: e.academic_session.name),
        ('exam_date',         'Exam Date',         lambda e: e.exam_date.strftime('%Y-%m-%d')),
        ('start_time',        'Start Time',        lambda e: e.start_time.strftime('%H:%M')),
        ('end_time',          'End Time',          lambda e: e.end_time.strftime('%H:%M')),
        ('duration_minutes',  'Duration (min)',    lambda e: e.duration_minutes),
        ('total_marks',       'Total Marks',       lambda e: float(e.total_marks)),
        ('pass_marks',        'Pass Marks',        lambda e: float(e.pass_marks)),
        ('exam_mode',         'Mode',              lambda e: e.get_exam_mode_display()),
        ('status',            'Status',            lambda e: e.get_status_display()),
        ('results_published', 'Results Published', lambda e: 'Yes' if e.results_published else 'No'),
        ('examination_venue', 'Venue',             lambda e: e.examination_venue or ''),
        ('target_classes',    'Target Classes',    lambda e: ', '.join(str(c) for c in e.target_classes.all())),
    ]

    COLUMN_MAP     = {col[0]: col for col in ALL_COLUMNS}
    DEFAULT_FIELDS = ['code', 'name', 'subject', 'exam_category', 'exam_date', 'start_time', 'total_marks', 'status']

    selected = request.GET.getlist('fields') or DEFAULT_FIELDS
    columns  = [COLUMN_MAP[f] for f in selected if f in COLUMN_MAP] or [COLUMN_MAP[f] for f in DEFAULT_FIELDS]

    examinations = _get_filtered_examinations(request)
    wb = Workbook()
    ws = wb.active
    ws.title = 'Examinations'

    ws.append([col[1] for col in columns])
    _style_header_row(ws)

    data_align = Alignment(vertical='center', wrap_text=False)
    for exam in examinations:
        ws.append([col[2](exam) for col in columns])
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = data_align

    _auto_size_columns(ws)
    return _excel_response(wb, 'examinations')


@login_required
def export_exam_categories_excel(request):
    """Export exam categories to Excel with selectable columns."""
    ALL_COLUMNS = [
        ('name',                     'Name',          lambda c: c.name),
        ('abbreviation',             'Abbreviation',  lambda c: c.abbreviation),
        ('code',                     'Code',          lambda c: c.code),
        ('category_type',            'Category Type', lambda c: c.get_category_type_display()),
        ('frequency',                'Frequency',     lambda c: c.get_frequency_display()),
        ('weight_percentage',        'Weight (%)',    lambda c: float(c.weight_percentage)),
        ('allows_retakes',           'Allows Retakes',lambda c: 'Yes' if c.allows_retakes else 'No'),
        ('max_retakes',              'Max Retakes',   lambda c: c.max_retakes),
        ('public_results',           'Public Results',lambda c: 'Yes' if c.public_results else 'No'),
        ('is_active',                'Active',        lambda c: 'Yes' if c.is_active else 'No'),
        ('curriculum_compatibility', 'Curriculum',    lambda c: c.get_curriculum_compatibility_display()),
    ]

    COLUMN_MAP     = {col[0]: col for col in ALL_COLUMNS}
    DEFAULT_FIELDS = ['name', 'abbreviation', 'category_type', 'frequency', 'weight_percentage', 'allows_retakes', 'is_active']

    selected = request.GET.getlist('fields') or DEFAULT_FIELDS
    columns  = [COLUMN_MAP[f] for f in selected if f in COLUMN_MAP] or [COLUMN_MAP[f] for f in DEFAULT_FIELDS]

    categories = _get_filtered_exam_categories(request)
    wb = Workbook()
    ws = wb.active
    ws.title = 'Exam Categories'

    ws.append([col[1] for col in columns])
    _style_header_row(ws)

    data_align = Alignment(vertical='center', wrap_text=False)
    for cat in categories:
        ws.append([col[2](cat) for col in columns])
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = data_align

    _auto_size_columns(ws)
    return _excel_response(wb, 'exam_categories')


@login_required
def export_grading_systems_excel(request):
    """
    Export grading systems to Excel.
    Sheet 1: Systems summary. Sheet 2: All grade ranges.
    """
    systems = _get_filtered_grading_systems(request)

    wb  = Workbook()
    ws1 = wb.active
    ws1.title = 'Grading Systems'
    ws1.append([
        'Name', 'Code', 'Type', 'Scale', 'Min Score', 'Max Score',
        'Pass Mark', 'Uses GPA', 'Active', 'Default', 'Grade Ranges',
    ])
    _style_header_row(ws1)

    data_align = Alignment(vertical='center', wrap_text=False)
    for sys in systems:
        ws1.append([
            sys.name,
            sys.code,
            sys.get_grading_type_display(),
            sys.get_scale_type_display(),
            float(sys.minimum_score),
            float(sys.maximum_score),
            float(sys.pass_mark),
            'Yes' if sys.uses_gpa   else 'No',
            'Yes' if sys.is_active  else 'No',
            'Yes' if sys.is_default else 'No',
            sys.ranges.count(),
        ])
    for row in ws1.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = data_align
    _auto_size_columns(ws1)

    ws2 = wb.create_sheet('Grade Ranges')
    ws2.append([
        'Grading System', 'Grade', 'Grade Name', 'Min Score', 'Max Score',
        'Aggregate', 'GPA Points', 'Passing', 'Color',
    ])
    _style_header_row(ws2, header_fill_hex='2E7D32')

    for sys in systems:
        for gr in sys.ranges.all().order_by('-min_score'):
            ws2.append([
                sys.name,
                gr.grade,
                gr.grade_name or '',
                float(gr.min_score),
                float(gr.max_score),
                gr.aggregate or '',
                float(gr.gpa_points) if gr.gpa_points else '',
                'Yes' if gr.is_passing_grade else 'No',
                gr.color_code or '',
            ])
    _auto_size_columns(ws2)

    return _excel_response(wb, 'grading_systems')


@login_required
def export_class_grading_systems_excel(request):
    """Export class grading system assignments to Excel."""
    assignments = _get_filtered_class_grading_systems(request)

    wb = Workbook()
    ws = wb.active
    ws.title = 'Class Grading Systems'
    ws.append([
        'Class', 'Academic Level', 'Grading System', 'Session', 'Subject',
        'Effective Date', 'End Date', 'Priority', 'Active', 'Default for Class',
    ])
    _style_header_row(ws)

    data_align = Alignment(vertical='center', wrap_text=False)
    for a in assignments:
        ws.append([
            str(a.class_instance),
            a.class_instance.academic_level.name,
            a.grading_system.name,
            a.academic_session.name,
            a.subject.name if a.subject else 'All Subjects',
            a.effective_date.strftime('%Y-%m-%d'),
            a.end_date.strftime('%Y-%m-%d') if a.end_date else '',
            a.priority,
            'Yes' if a.is_active            else 'No',
            'Yes' if a.is_default_for_class else 'No',
        ])
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = data_align

    _auto_size_columns(ws)
    return _excel_response(wb, 'class_grading_systems')


@login_required
def export_class_results_excel(request, class_pk):
    """
    Export the results grid for one class + session to Excel.

    Sheet 1: Student × subject score grid.
    Sheet 2: Flat result rows (one per result, for data analysis).

    Query params:
      ?session=<pk>     override session
      ?category=<abbr>  limit to one category tab
    """
    class_instance, session = _resolve_class_and_session(class_pk, request)
    if not session:
        return HttpResponse('No session found.', status=400)

    category_abbr = request.GET.get('category', '')
    examinations, _students, grid_rows, category = _build_results_grid(
        class_instance, session, category_abbr or None
    )

    wb = Workbook()

    # ── Sheet 1: Score grid ───────────────────────────────────────
    ws1       = wb.active
    ws1.title = 'Marks Grid'

    ws1.append(
        ['#', 'Student', 'Adm. No.'] +
        [f'{e.subject.name} /{float(e.total_marks):.0f}' for e in examinations] +
        ['Total', 'Average', 'Position']
    )
    _style_header_row(ws1)

    centre_align = Alignment(vertical='center', horizontal='center')
    left_align   = Alignment(vertical='center', horizontal='left')
    pass_fill    = PatternFill(start_color='E8F5E9', end_color='E8F5E9', fill_type='solid')
    fail_fill    = PatternFill(start_color='FFEBEE', end_color='FFEBEE', fill_type='solid')
    default_pass = float(examinations[0].pass_marks) if examinations else 50.0

    for idx, row in enumerate(grid_rows, 1):
        ws1.append(
            [idx, row['student'].get_full_name(), row['student'].admission_number] +
            [float(c['score']) if c['score'] is not None else '' for c in row['cells']] +
            [float(row['total']), float(row['average']), row.get('position', '')]
        )
        xl_row      = ws1.max_row
        name_cell   = ws1.cell(row=xl_row, column=2)
        name_cell.fill      = pass_fill if float(row['average']) >= default_pass else fail_fill
        name_cell.alignment = left_align
        for col_idx in range(4, len(row['cells']) + 7):
            ws1.cell(row=xl_row, column=col_idx).alignment = centre_align

    _auto_size_columns(ws1)

    # ── Sheet 2: Flat results ─────────────────────────────────────
    ws2       = wb.create_sheet('Flat Results')
    ws2.append([
        'Student', 'Adm. No.', 'Category', 'Subject',
        'Score', 'Total Marks', 'Percentage', 'Grade', 'Pass/Fail',
        'Published', 'Locked',
    ])
    _style_header_row(ws2, header_fill_hex='4A235A')

    flat_results = StudentExamResult.objects.filter(
        student__in=[r['student'] for r in grid_rows],
        examination__in=examinations,
    ).select_related(
        'student', 'examination__subject', 'examination__exam_category',
    ).order_by(
        'student__first_name',
        'examination__exam_category__name',
        'examination__subject__name',
    )

    data_align = Alignment(vertical='center', wrap_text=False)
    for result in flat_results:
        ws2.append([
            result.student.get_full_name(),
            result.student.admission_number,
            result.examination.exam_category.name,
            result.examination.subject.name,
            float(result.score)      if result.score      is not None else '',
            float(result.examination.total_marks),
            float(result.percentage) if result.percentage is not None else '',
            result.grade or '',
            'Pass' if result.is_pass        else 'Fail',
            'Yes'  if result.is_published   else 'No',
            'Yes'  if result.is_grade_locked else 'No',
        ])
    for row in ws2.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = data_align

    _auto_size_columns(ws2)

    filename = f'{class_instance}_results'
    if category:
        filename += f'_{category.abbreviation}'
    return _excel_response(wb, filename)